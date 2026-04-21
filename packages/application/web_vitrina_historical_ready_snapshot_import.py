"""One-time historical workbook -> ready snapshot importer for web-vitrina."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
import json
from pathlib import Path
import sqlite3
from typing import Any, Mapping

from openpyxl import load_workbook

from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime
from packages.application.sheet_vitrina_v1_live_plan import (
    HISTORICAL_CLOSED_DAY_SOURCE_KEYS,
)
from packages.contracts.sheet_vitrina_v1 import (
    SheetVitrinaV1Envelope,
    SheetVitrinaV1TemporalSlot,
    SheetVitrinaWriteTarget,
)

ARTIFACT_NAME = "web_vitrina_historical_ready_snapshot_artifact"
ARTIFACT_VERSION = "v1"
HISTORICAL_READY_SNAPSHOT_PLAN_VERSION = "sheet_vitrina_v1_historical_import_v1__sheet_scaffold_v1"
HISTORICAL_READY_SNAPSHOT_SLOT_KEY = "historical_import"
HISTORICAL_READY_SNAPSHOT_SLOT_LABEL = "Historical import"
DEFAULT_WORKBOOK_SHEET_NAME = "DATA_VITRINA"
DEFAULT_DATE_FROM = "2026-03-01"
DEFAULT_DATE_TO = "2026-04-19"


@dataclass(frozen=True)
class HistoricalArtifactRow:
    row_id: str
    label: str
    values_by_date: dict[str, Any]


@dataclass(frozen=True)
class HistoricalArtifact:
    artifact_name: str
    artifact_version: str
    source_file: str
    sheet_name: str
    date_from: str
    date_to: str
    dates: list[str]
    rows: list[HistoricalArtifactRow]


def extract_historical_artifact_from_workbook(
    *,
    workbook_path: Path,
    sheet_name: str = DEFAULT_WORKBOOK_SHEET_NAME,
    date_from: str = DEFAULT_DATE_FROM,
    date_to: str = DEFAULT_DATE_TO,
) -> HistoricalArtifact:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"workbook does not contain sheet {sheet_name!r}")
    worksheet = workbook[sheet_name]
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("historical workbook sheet is empty") from exc
    date_columns = _resolve_date_columns(header, date_from=date_from, date_to=date_to)
    if not date_columns:
        raise ValueError(f"historical workbook does not cover requested date window {date_from}..{date_to}")

    artifact_rows: list[HistoricalArtifactRow] = []
    current_scope: str | None = None
    seen_row_ids: set[str] = set()
    for row in rows_iter:
        key = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        label = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        if not key:
            continue
        if "|" in key:
            row_id = key
        elif key == "TOTAL" or key.startswith("GROUP:") or key.startswith("SKU:"):
            current_scope = key
            continue
        else:
            if current_scope is None:
                raise ValueError(f"metric row {key!r} appears before any scope marker")
            row_id = f"{current_scope}|{key}"
        if row_id in seen_row_ids:
            raise ValueError(f"historical workbook contains duplicate row_id: {row_id}")
        seen_row_ids.add(row_id)
        values_by_date: dict[str, Any] = {}
        for column_index, snapshot_date in date_columns:
            raw_value = row[column_index] if column_index < len(row) else None
            values_by_date[snapshot_date] = raw_value
        artifact_rows.append(
            HistoricalArtifactRow(
                row_id=row_id,
                label=label,
                values_by_date=values_by_date,
            )
        )

    return HistoricalArtifact(
        artifact_name=ARTIFACT_NAME,
        artifact_version=ARTIFACT_VERSION,
        source_file=str(workbook_path),
        sheet_name=sheet_name,
        date_from=date_columns[0][1],
        date_to=date_columns[-1][1],
        dates=[snapshot_date for _, snapshot_date in date_columns],
        rows=artifact_rows,
    )


def historical_artifact_to_payload(artifact: HistoricalArtifact) -> dict[str, Any]:
    return {
        "artifact_name": artifact.artifact_name,
        "artifact_version": artifact.artifact_version,
        "source_file": artifact.source_file,
        "sheet_name": artifact.sheet_name,
        "date_from": artifact.date_from,
        "date_to": artifact.date_to,
        "dates": list(artifact.dates),
        "rows": [
            {
                "row_id": row.row_id,
                "label": row.label,
                "values_by_date": dict(row.values_by_date),
            }
            for row in artifact.rows
        ],
    }


def historical_artifact_from_payload(payload: Mapping[str, Any]) -> HistoricalArtifact:
    artifact_name = str(payload.get("artifact_name") or "")
    artifact_version = str(payload.get("artifact_version") or "")
    if artifact_name != ARTIFACT_NAME or artifact_version != ARTIFACT_VERSION:
        raise ValueError(
            f"unsupported artifact identity: expected {ARTIFACT_NAME} {ARTIFACT_VERSION}, "
            f"got {artifact_name!r} {artifact_version!r}"
        )
    rows_payload = list(payload.get("rows") or [])
    return HistoricalArtifact(
        artifact_name=artifact_name,
        artifact_version=artifact_version,
        source_file=str(payload.get("source_file") or ""),
        sheet_name=str(payload.get("sheet_name") or DEFAULT_WORKBOOK_SHEET_NAME),
        date_from=str(payload.get("date_from") or ""),
        date_to=str(payload.get("date_to") or ""),
        dates=[str(item) for item in list(payload.get("dates") or [])],
        rows=[
            HistoricalArtifactRow(
                row_id=str(item["row_id"]),
                label=str(item.get("label") or ""),
                values_by_date={
                    str(snapshot_date): value
                    for snapshot_date, value in dict(item.get("values_by_date") or {}).items()
                },
            )
            for item in rows_payload
        ],
    )


def save_historical_artifact(*, artifact: HistoricalArtifact, output_path: Path) -> None:
    output_path.write_text(
        json.dumps(historical_artifact_to_payload(artifact), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_historical_artifact(*, artifact_path: Path) -> HistoricalArtifact:
    return historical_artifact_from_payload(json.loads(artifact_path.read_text(encoding="utf-8")))


def compare_historical_artifact_against_runtime(
    *,
    runtime: RegistryUploadDbBackedRuntime,
    artifact: HistoricalArtifact,
    date_from: str = DEFAULT_DATE_FROM,
    date_to: str = DEFAULT_DATE_TO,
) -> dict[str, Any]:
    selected_dates = _selected_dates(artifact, date_from=date_from, date_to=date_to)
    template = _load_ready_snapshot_template(runtime)
    artifact_row_ids = [row.row_id for row in artifact.rows]
    ready_snapshot_dates = _list_ready_snapshot_dates(runtime, date_from=date_from, date_to=date_to)
    available_dates = [snapshot_date for snapshot_date in selected_dates if snapshot_date in ready_snapshot_dates]
    missing_dates = [snapshot_date for snapshot_date in selected_dates if snapshot_date not in ready_snapshot_dates]
    return {
        "status": "success",
        "requested_window": {
            "date_from": date_from,
            "date_to": date_to,
        },
        "artifact_summary": {
            "source_file": artifact.source_file,
            "sheet_name": artifact.sheet_name,
            "date_count": len(selected_dates),
            "row_count": len(artifact.rows),
        },
        "template_summary": {
            "template_as_of_date": template["template_as_of_date"],
            "template_snapshot_id": template["template_snapshot_id"],
            "row_count": len(template["row_ids"]),
            "same_row_id_set": set(template["row_ids"]) == set(artifact_row_ids),
            "same_row_id_order": template["row_ids"] == artifact_row_ids,
            "template_only_row_ids": sorted(set(template["row_ids"]) - set(artifact_row_ids)),
            "artifact_only_row_ids": sorted(set(artifact_row_ids) - set(template["row_ids"])),
        },
        "ready_snapshot_coverage": {
            "available_count": len(available_dates),
            "missing_count": len(missing_dates),
            "available_dates": available_dates,
            "missing_dates": missing_dates,
        },
        "source_family_coverage": _source_family_coverage(runtime, date_from=date_from, date_to=date_to),
    }


def materialize_historical_ready_snapshots(
    *,
    runtime: RegistryUploadDbBackedRuntime,
    artifact: HistoricalArtifact,
    captured_at: str,
    date_from: str = DEFAULT_DATE_FROM,
    date_to: str = DEFAULT_DATE_TO,
    replace_existing: bool = False,
) -> dict[str, Any]:
    selected_dates = _selected_dates(artifact, date_from=date_from, date_to=date_to)
    if not selected_dates:
        raise ValueError(f"artifact does not contain requested window {date_from}..{date_to}")
    template = _load_ready_snapshot_template(runtime)
    current_state = runtime.load_current_state()
    artifact_values_by_row_id = {
        row.row_id: dict(row.values_by_date)
        for row in artifact.rows
    }
    template_row_ids = list(template["row_ids"])
    if set(template_row_ids) != set(artifact_values_by_row_id):
        raise ValueError(
            "historical artifact row universe does not match current DATA_VITRINA template; "
            "cannot truthfully materialize shared ready snapshots"
        )

    saved_count = 0
    overwritten_count = 0
    skipped_count = 0
    saved_dates: list[str] = []
    for snapshot_date in selected_dates:
        exists = _ready_snapshot_exists(runtime, as_of_date=snapshot_date)
        if exists and not replace_existing:
            skipped_count += 1
            continue
        if exists:
            overwritten_count += 1
        data_rows = [
            [
                template["label_by_row_id"][row_id],
                row_id,
                artifact_values_by_row_id[row_id].get(snapshot_date),
            ]
            for row_id in template_row_ids
        ]
        data_sheet = SheetVitrinaWriteTarget(
            sheet_name="DATA_VITRINA",
            write_start_cell=str(template["data_write_start_cell"]),
            write_rect=_build_write_rect(column_count=3, row_count=len(data_rows)),
            clear_range=str(template["data_clear_range"]),
            write_mode=str(template["data_write_mode"]),
            partial_update_allowed=bool(template["data_partial_update_allowed"]),
            header=["label", "key", snapshot_date],
            rows=data_rows,
            row_count=len(data_rows),
            column_count=3,
        )
        sheets = [data_sheet]
        if template["status_header"]:
            sheets.append(
                SheetVitrinaWriteTarget(
                    sheet_name="STATUS",
                    write_start_cell=str(template["status_write_start_cell"]),
                    write_rect=_build_write_rect(
                        column_count=len(template["status_header"]),
                        row_count=0,
                    ),
                    clear_range=str(template["status_clear_range"]),
                    write_mode=str(template["status_write_mode"]),
                    partial_update_allowed=bool(template["status_partial_update_allowed"]),
                    header=list(template["status_header"]),
                    rows=[],
                    row_count=0,
                    column_count=len(template["status_header"]),
                )
            )
        plan = SheetVitrinaV1Envelope(
            plan_version=HISTORICAL_READY_SNAPSHOT_PLAN_VERSION,
            snapshot_id=f"{snapshot_date}__web_vitrina_historical_import_v1__ready",
            as_of_date=snapshot_date,
            date_columns=[snapshot_date],
            temporal_slots=[
                SheetVitrinaV1TemporalSlot(
                    slot_key=HISTORICAL_READY_SNAPSHOT_SLOT_KEY,
                    slot_label=HISTORICAL_READY_SNAPSHOT_SLOT_LABEL,
                    column_date=snapshot_date,
                )
            ],
            source_temporal_policies={},
            sheets=sheets,
        )
        runtime.save_sheet_vitrina_ready_snapshot(
            current_state=current_state,
            refreshed_at=captured_at,
            plan=plan,
        )
        saved_count += 1
        saved_dates.append(snapshot_date)

    ready_snapshot_dates = _list_ready_snapshot_dates(runtime, date_from=date_from, date_to=date_to)
    available_dates = [snapshot_date for snapshot_date in selected_dates if snapshot_date in ready_snapshot_dates]
    missing_dates = [snapshot_date for snapshot_date in selected_dates if snapshot_date not in ready_snapshot_dates]
    return {
        "status": "success",
        "requested_window": {
            "date_from": date_from,
            "date_to": date_to,
        },
        "captured_at": captured_at,
        "replace_existing": replace_existing,
        "saved_snapshot_count": saved_count,
        "overwritten_snapshot_count": overwritten_count,
        "skipped_existing_snapshot_count": skipped_count,
        "saved_dates": saved_dates,
        "ready_snapshot_coverage_after": {
            "available_count": len(available_dates),
            "missing_count": len(missing_dates),
            "available_dates": available_dates,
            "missing_dates": missing_dates,
        },
    }


def _load_ready_snapshot_template(runtime: RegistryUploadDbBackedRuntime) -> dict[str, Any]:
    template_plan = runtime.load_sheet_vitrina_ready_snapshot()
    data_sheet = next((item for item in template_plan.sheets if item.sheet_name == "DATA_VITRINA"), None)
    if data_sheet is None:
        raise ValueError("latest ready snapshot does not contain DATA_VITRINA")
    status_sheet = next((item for item in template_plan.sheets if item.sheet_name == "STATUS"), None)
    row_ids: list[str] = []
    label_by_row_id: dict[str, str] = {}
    for row in data_sheet.rows:
        if len(row) < 2:
            continue
        row_id = str(row[1] or "").strip()
        if not row_id:
            continue
        row_ids.append(row_id)
        label_by_row_id[row_id] = str(row[0] or "")
    return {
        "template_as_of_date": template_plan.as_of_date,
        "template_snapshot_id": template_plan.snapshot_id,
        "row_ids": row_ids,
        "label_by_row_id": label_by_row_id,
        "data_write_start_cell": data_sheet.write_start_cell,
        "data_clear_range": data_sheet.clear_range,
        "data_write_mode": data_sheet.write_mode,
        "data_partial_update_allowed": data_sheet.partial_update_allowed,
        "status_header": list(status_sheet.header) if status_sheet is not None else [],
        "status_write_start_cell": status_sheet.write_start_cell if status_sheet is not None else "A1",
        "status_clear_range": status_sheet.clear_range if status_sheet is not None else "A:Z",
        "status_write_mode": status_sheet.write_mode if status_sheet is not None else "overwrite",
        "status_partial_update_allowed": (
            status_sheet.partial_update_allowed if status_sheet is not None else False
        ),
    }


def _list_ready_snapshot_dates(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    date_from: str,
    date_to: str,
) -> list[str]:
    current_state = runtime.load_current_state()
    with sqlite3.connect(runtime.db_path) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT as_of_date
            FROM sheet_vitrina_v1_ready_snapshots
            WHERE bundle_version = ?
              AND as_of_date >= ?
              AND as_of_date <= ?
            ORDER BY as_of_date
            """,
            (current_state.bundle_version, date_from, date_to),
        ).fetchall()
    return [str(row["as_of_date"]) for row in rows]


def _ready_snapshot_exists(runtime: RegistryUploadDbBackedRuntime, *, as_of_date: str) -> bool:
    try:
        runtime.load_sheet_vitrina_ready_snapshot(as_of_date=as_of_date)
    except ValueError:
        return False
    return True


def _source_family_coverage(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    date_from: str,
    date_to: str,
) -> dict[str, Any]:
    exact_date_sources = [
        "sales_funnel_history",
        "stocks",
        "promo_by_price",
        "seller_funnel_snapshot",
        "web_source_snapshot",
    ]
    exact_snapshot_coverage: dict[str, dict[str, Any]] = {}
    for source_key in exact_date_sources:
        dates = [
            snapshot_date
            for snapshot_date in runtime.list_temporal_source_snapshot_dates(source_key=source_key)
            if date_from <= snapshot_date <= date_to
        ]
        exact_snapshot_coverage[source_key] = {
            "count_in_window": len(dates),
            "first_date": dates[0] if dates else None,
            "last_date": dates[-1] if dates else None,
        }

    success_states = runtime.list_temporal_source_closure_states(
        source_keys=sorted(HISTORICAL_CLOSED_DAY_SOURCE_KEYS),
        slot_kind="yesterday_closed",
        states=["success"],
    )
    accepted_closed_dates_by_source: dict[str, list[str]] = {}
    for state in success_states:
        if date_from <= state.target_date <= date_to:
            accepted_closed_dates_by_source.setdefault(state.source_key, []).append(state.target_date)
    accepted_closed_coverage = {
        source_key: {
            "count_in_window": len(sorted(set(accepted_closed_dates_by_source.get(source_key) or []))),
            "first_date": (
                sorted(set(accepted_closed_dates_by_source.get(source_key) or []))[0]
                if accepted_closed_dates_by_source.get(source_key)
                else None
            ),
            "last_date": (
                sorted(set(accepted_closed_dates_by_source.get(source_key) or []))[-1]
                if accepted_closed_dates_by_source.get(source_key)
                else None
            ),
        }
        for source_key in sorted(HISTORICAL_CLOSED_DAY_SOURCE_KEYS)
    }
    return {
        "exact_snapshot_coverage": exact_snapshot_coverage,
        "accepted_closed_day_coverage": accepted_closed_coverage,
    }


def _selected_dates(
    artifact: HistoricalArtifact,
    *,
    date_from: str,
    date_to: str,
) -> list[str]:
    return [
        snapshot_date
        for snapshot_date in artifact.dates
        if date_from <= snapshot_date <= date_to
    ]


def _resolve_date_columns(header: tuple[Any, ...], *, date_from: str, date_to: str) -> list[tuple[int, str]]:
    selected: list[tuple[int, str]] = []
    for column_index, raw_value in enumerate(header[2:], start=2):
        snapshot_date = _coerce_header_date(raw_value)
        if snapshot_date is None:
            continue
        if date_from <= snapshot_date <= date_to:
            selected.append((column_index, snapshot_date))
    return selected


def _coerce_header_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        serial_days = int(value)
        return (_excel_epoch() + timedelta(days=serial_days)).isoformat()
    text = str(value or "").strip()
    if not text:
        return None
    for parser in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, parser).date().isoformat()
        except ValueError:
            continue
    return None


def _build_write_rect(*, column_count: int, row_count: int) -> str:
    last_column = _column_letters(column_count)
    last_row = row_count + 1
    return f"A1:{last_column}{last_row}"


def _column_letters(column_count: int) -> str:
    if column_count <= 0:
        raise ValueError("column_count must be positive")
    current = column_count
    letters: list[str] = []
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def _excel_epoch() -> date:
    return date(1899, 12, 30)
