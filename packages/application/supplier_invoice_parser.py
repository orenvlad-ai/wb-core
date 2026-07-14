"""Bounded XLSX parser for supplier invoice shipment registry."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import math
from pathlib import Path
import re
from typing import Any, Iterable, Mapping
import unicodedata
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from packages.contracts.supplier_shipments import (
    LINE_TYPE_EXTRA,
    LINE_TYPE_PRODUCT,
    MATCH_STATUS_EXTRA,
    MATCH_STATUS_UNMATCHED,
    PRODUCT_TYPE_ANTI_SPY,
    PRODUCT_TYPE_CLEAR,
    PRODUCT_TYPE_MATTE,
    SUPPLIER_INVOICE_PARSER_VERSION,
)

BARCODE_HEADER_ALIASES = (
    "barcode",
    "braocde",
    "条形码",
    "條形碼",
    "штрихкод",
    "шк",
)
BARCODE_PROFILE_MIN_DIGITS = 8
BARCODE_PROFILE_MAX_DIGITS = 32
MAX_EXACT_FLOAT_INTEGER = 2**53


def parse_supplier_invoice_xlsx(
    workbook_bytes: bytes,
    *,
    filename: str = "",
) -> dict[str, Any]:
    """Parse a supplier XLSX invoice into editable shipment payload."""

    parser = SupplierInvoiceParser()
    return parser.parse(workbook_bytes, filename=filename)


class SupplierInvoiceParser:
    def parse(self, workbook_bytes: bytes, *, filename: str = "") -> dict[str, Any]:
        if not workbook_bytes:
            raise ValueError("supplier invoice workbook is empty")
        try:
            workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl owns exact exception types
            raise ValueError(f"supplier invoice workbook must be a readable XLSX file: {exc}") from exc
        if not workbook.worksheets:
            raise ValueError("supplier invoice workbook does not contain worksheets")

        worksheet: Worksheet | None = None
        merged_values: dict[tuple[int, int], Any] = {}
        header_row = 0
        columns: dict[str, int] = {}
        barcode_diagnostics: dict[str, Any] = {}
        last_header_error: ValueError | None = None
        for candidate_sheet in workbook.worksheets:
            candidate_merged_values = _build_merged_value_index(candidate_sheet)
            try:
                candidate_header_row, candidate_columns, candidate_diagnostics = _find_header_row(
                    candidate_sheet,
                    candidate_merged_values,
                )
            except ValueError as exc:
                if "ambiguous barcode column" in str(exc).lower():
                    raise
                last_header_error = exc
                continue
            worksheet = candidate_sheet
            merged_values = candidate_merged_values
            header_row = candidate_header_row
            columns = candidate_columns
            barcode_diagnostics = candidate_diagnostics
            break
        if worksheet is None:
            raise last_header_error or ValueError("supplier invoice table headers not found")
        metadata = _extract_metadata(
            worksheet,
            merged_values,
            header_row=header_row,
            filename=filename,
            workbook_text_blocks=_extract_workbook_text_blocks(workbook_bytes),
        )
        lines, warnings = self._parse_lines(worksheet, merged_values, header_row=header_row, columns=columns)
        if metadata.get("declared_invoice_total") is None:
            metadata["declared_invoice_total"] = _extract_declared_invoice_total(
                worksheet,
                merged_values,
                header_row=header_row,
            )
        summary = _build_summary(lines, declared_total=_to_number(metadata.get("declared_invoice_total")))
        warnings.extend(_metadata_warnings(metadata))
        errors: list[str] = []
        if summary.get("checksum_error"):
            errors.append(
                "invoice total checksum mismatch: declared "
                f"{summary.get('declared_invoice_total')} vs parsed {summary.get('invoice_amount_total')}"
            )

        return {
            "parser_version": SUPPLIER_INVOICE_PARSER_VERSION,
            "diagnostics": {
                "worksheet": worksheet.title,
                "header_row": header_row,
                "barcode_column": barcode_diagnostics,
            },
            "metadata": metadata,
            "summary": summary,
            "lines": lines,
            "warnings": warnings,
            "errors": errors,
        }

    def _parse_lines(
        self,
        worksheet: Worksheet,
        merged_values: Mapping[tuple[int, int], Any],
        *,
        header_row: int,
        columns: Mapping[str, int],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        lines: list[dict[str, Any]] = []
        warnings: list[str] = []
        current_product_type = ""
        blank_run = 0
        sort_order = 0

        for row_index in range(header_row + 1, worksheet.max_row + 1):
            row_values = {
                key: _cell_value(worksheet, row_index, col_index, merged_values)
                for key, col_index in columns.items()
            }
            row_text = " ".join(_stringify(value) for value in row_values.values() if _stringify(value))
            detected_type = detect_product_type(row_text)

            model_raw = _stringify(row_values.get("models"))
            name_spec = _stringify(row_values.get("name_spec"))
            comment = _stringify(row_values.get("comment")) or name_spec
            qty = _to_number(row_values.get("qty"))
            unit_price = _to_number(row_values.get("unit_price"))
            amount = _to_number(row_values.get("amount"))
            source_no = _stringify(row_values.get("no"))
            barcode = ""
            barcode_error = ""
            try:
                barcode = normalize_barcode_value(row_values.get("barcode"))
            except ValueError as exc:
                barcode_error = str(exc)

            if not row_text:
                blank_run += 1
                if blank_run >= 12 and lines:
                    break
                continue
            blank_run = 0

            has_numeric_payload = qty is not None or unit_price is not None or amount is not None
            if lines and _is_total_row(row_text):
                break
            if not (model_raw or name_spec or source_no or has_numeric_payload):
                continue
            if has_numeric_payload and not (model_raw or name_spec or source_no):
                continue

            is_extra = _is_extra_row(source_no, model_raw, name_spec)
            if not is_extra and detected_type:
                current_product_type = detected_type

            if is_extra:
                sort_order += 1
                lines.append(
                    {
                        "line_type": LINE_TYPE_EXTRA,
                        "sort_order": sort_order,
                        "source_no": source_no,
                        "barcode": barcode,
                        "product_type": "",
                        "model_raw": model_raw or name_spec,
                        "model_normalized": normalize_invoice_model(model_raw or name_spec),
                        "match_key": "",
                        "internal_sku": "",
                        "internal_nm_id": None,
                        "internal_name": "",
                        "qty": qty,
                        "unit_price": unit_price,
                        "amount": amount,
                        "currency": "",
                        "comment": comment,
                        "match_status": MATCH_STATUS_EXTRA,
                        "manual_override": False,
                        "match_evidence": {"method": "barcode", "status": "extra"},
                        "raw": _raw_row_payload(row_index, row_values),
                    }
                )
                continue

            if not model_raw and not has_numeric_payload:
                continue
            product_type = detected_type or current_product_type
            normalized_model = normalize_invoice_model(model_raw)
            match_key = f"{product_type}|{normalized_model}" if product_type and normalized_model else ""
            if not product_type:
                warnings.append(f"row {row_index}: product type is not detected")
            if not normalized_model:
                warnings.append(f"row {row_index}: model is empty or unsupported")
            if barcode_error:
                warnings.append(f"row {row_index}: barcode cannot be normalized losslessly ({barcode_error})")
            elif not barcode:
                warnings.append(f"row {row_index}: product barcode is missing")

            sort_order += 1
            lines.append(
                {
                    "line_type": LINE_TYPE_PRODUCT,
                    "sort_order": sort_order,
                    "source_no": source_no,
                    "barcode": barcode,
                    "product_type": product_type,
                    "model_raw": model_raw,
                    "model_normalized": normalized_model,
                    "match_key": match_key,
                    "internal_sku": "",
                    "internal_nm_id": None,
                    "internal_name": "",
                    "qty": qty,
                    "unit_price": unit_price,
                    "amount": amount,
                    "currency": "",
                    "comment": comment,
                    "match_status": MATCH_STATUS_UNMATCHED,
                    "manual_override": False,
                    "match_evidence": {
                        "method": "barcode",
                        "status": "unmatched",
                        "reason": "barcode_invalid" if barcode_error else "barcode_not_resolved" if barcode else "barcode_missing",
                    },
                    "raw": _raw_row_payload(row_index, row_values),
                }
            )

        if not lines:
            warnings.append("invoice table was found but no editable invoice rows were parsed")
        return lines, warnings


def detect_product_type(value: Any) -> str:
    text = _stringify(value).lower()
    if not text:
        return ""
    if "防窥膜" in text or "anti-spy" in text or "anti spy" in text:
        return PRODUCT_TYPE_ANTI_SPY
    if "磨砂膜" in text or "matte" in text:
        return PRODUCT_TYPE_MATTE
    if "高清膜" in text or re.search(r"\bsmk\b", text):
        return PRODUCT_TYPE_CLEAR
    return ""


def normalize_invoice_model(value: Any) -> str:
    text = _stringify(value).lower()
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\((?:anti[-\s]?spy|matte|smk|clear)[^)]*\)", " ", text, flags=re.IGNORECASE)
    text = text.replace("iphone", "iphone ")
    text = re.sub(r"\bpromax\b", "pro max", text)
    text = re.sub(r"\bip\s*hone\b", "iphone", text)
    text = text.replace("&", " ")
    text = text.replace("+", " plus ")
    text = re.sub(r"[/\\]+", "_", text)
    text = re.sub(r"[-–—]+", "_", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    text = re.sub(r"^iphone_", "iphone_", text)
    return text


def extract_iphone_model_keys(value: Any) -> list[str]:
    """Return normalized iPhone model tokens for compatibility matching."""

    text = _stringify(value).lower()
    if not text:
        return []
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\((?:anti[-\s]?spy|matte|smk|clear)[^)]*\)", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bip\s*hone\b", "iphone", text)
    text = re.sub(r"\biphone\b", " iphone ", text)
    text = re.sub(r"\bpromax\b", "pro max", text)
    text = re.sub(r"(?<=\d)pro\s*max\b", " pro max", text)
    text = re.sub(r"(?<=\d)pro\b", " pro", text)
    text = re.sub(r"(?<=\d)plus\b", " plus", text)
    text = re.sub(r"(?<=\d)air\b", " air", text)
    text = text.replace("_", " ")
    text = re.sub(r"[/\\,;|]+", " ", text)
    text = re.sub(r"[-–—]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    pattern = re.compile(
        r"(?:\biphone\s*)?"
        r"\b(?P<number>1[0-9]|2[0-9])\s*"
        r"(?P<suffix>pro\s*max|pro|max|plus|mini|air|e)?\b",
        flags=re.IGNORECASE,
    )
    keys: list[str] = []
    seen: set[str] = set()
    for match in pattern.finditer(text):
        number = match.group("number")
        suffix = re.sub(r"\s+", "_", (match.group("suffix") or "").strip().lower())
        key = f"iphone_{number}" + (suffix if suffix == "e" else f"_{suffix}" if suffix else "")
        if key not in seen:
            seen.add(key)
            keys.append(key)
    return keys


def _build_merged_value_index(worksheet: Worksheet) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row_index, col_index)] = top_left
    return merged_values


def _cell_value(
    worksheet: Worksheet,
    row_index: int,
    col_index: int,
    merged_values: Mapping[tuple[int, int], Any],
) -> Any:
    if (row_index, col_index) in merged_values:
        return merged_values[(row_index, col_index)]
    return worksheet.cell(row_index, col_index).value


def _find_header_row(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
) -> tuple[int, dict[str, int], dict[str, Any]]:
    for row_index in range(1, min(worksheet.max_row, 80) + 1):
        columns: dict[str, int] = {}
        for col_index in range(1, worksheet.max_column + 1):
            role = _invoice_header_role(_cell_value(worksheet, row_index, col_index, merged_values))
            if role and role not in columns:
                columns[role] = col_index
        required = {"no", "models", "qty", "unit_price", "amount"}
        if required.issubset(columns):
            barcode_column, diagnostics = _detect_barcode_column(
                worksheet,
                merged_values,
                header_row=row_index,
                columns=columns,
            )
            return row_index, {**columns, "barcode": barcode_column}, diagnostics
    raise ValueError("supplier invoice table headers not found: expected NO., MODELS, QTY, U.PRICE, AMOUNT")


def _invoice_header_role(value: Any) -> str:
    compact = _normalize_header_compact(value)
    if not compact:
        return ""
    if compact in {"no", "number", "序号", "序號"}:
        return "no"
    if "model" in compact or "型号" in compact or "型號" in compact:
        return "models"
    if "qty" in compact or "quantity" in compact or "数量" in compact or "數量" in compact:
        return "qty"
    if "uprice" in compact or "unitprice" in compact or "单价" in compact or "單價" in compact:
        return "unit_price"
    if "amount" in compact or "总价" in compact or "總價" in compact or "金额" in compact or "金額" in compact:
        return "amount"
    if ("name" in compact and ("spec" in compact or "description" in compact)) or any(
        token in compact for token in ("品名规格", "品名規格", "产品名称", "產品名稱")
    ):
        return "name_spec"
    if compact in {"comment", "comments", "remark", "remarks", "备注", "備註"}:
        return "comment"
    return ""


def _detect_barcode_column(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    *,
    header_row: int,
    columns: Mapping[str, int],
) -> tuple[int, dict[str, Any]]:
    known_columns = set(columns.values())
    semantic_candidates = [
        col_index
        for col_index in range(1, worksheet.max_column + 1)
        if col_index not in known_columns
        and _is_barcode_header(_cell_value(worksheet, header_row, col_index, merged_values))
    ]
    semantic_confirmed = _confirmed_barcode_candidates(
        worksheet,
        merged_values,
        header_row=header_row,
        columns=columns,
        candidates=semantic_candidates,
    )
    if len(semantic_confirmed) > 1:
        raise _ambiguous_barcode_column_error(worksheet, merged_values, header_row, semantic_confirmed)
    if len(semantic_confirmed) == 1:
        return _barcode_column_result(
            worksheet,
            merged_values,
            header_row=header_row,
            column=semantic_confirmed[0],
            method="header_alias",
            columns=columns,
        )
    if semantic_candidates:
        candidate_text = _barcode_candidate_text(worksheet, merged_values, header_row, semantic_candidates)
        raise ValueError(
            "supplier invoice barcode column header was found but its values could not be confirmed: "
            + candidate_text
        )

    name_column = columns.get("name_spec")
    qty_column = columns.get("qty")
    relative_candidates: list[int] = []
    if name_column is not None and qty_column is not None and name_column < qty_column:
        relative_candidates = [
            col_index
            for col_index in range(name_column + 1, qty_column)
            if col_index not in known_columns
        ]
    relative_confirmed = _confirmed_barcode_candidates(
        worksheet,
        merged_values,
        header_row=header_row,
        columns=columns,
        candidates=relative_candidates,
    )
    if len(relative_confirmed) > 1:
        raise _ambiguous_barcode_column_error(worksheet, merged_values, header_row, relative_confirmed)
    if len(relative_confirmed) == 1:
        column = relative_confirmed[0]
        method = "positional_d" if column == 4 and _has_confirmed_current_template_structure(columns) else "relative_structure"
        return _barcode_column_result(
            worksheet,
            merged_values,
            header_row=header_row,
            column=column,
            method=method,
            columns=columns,
        )

    if _has_confirmed_current_template_structure(columns) and 4 not in known_columns:
        positional_confirmed = _confirmed_barcode_candidates(
            worksheet,
            merged_values,
            header_row=header_row,
            columns=columns,
            candidates=[4],
        )
        if positional_confirmed:
            return _barcode_column_result(
                worksheet,
                merged_values,
                header_row=header_row,
                column=4,
                method="positional_d",
                columns=columns,
            )

    raise ValueError(
        "supplier invoice barcode column not found: no semantic alias or structure-confirmed barcode-like column"
    )


def _confirmed_barcode_candidates(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    *,
    header_row: int,
    columns: Mapping[str, int],
    candidates: Iterable[int],
) -> list[int]:
    return [
        col_index
        for col_index in candidates
        if _barcode_value_profile(
            worksheet,
            merged_values,
            header_row=header_row,
            columns=columns,
            candidate_column=col_index,
        )["confirmed"]
    ]


def _barcode_value_profile(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    *,
    header_row: int,
    columns: Mapping[str, int],
    candidate_column: int,
) -> dict[str, Any]:
    product_rows = 0
    nonempty_values = 0
    valid_values = 0
    digit_values = 0
    barcode_length_values = 0
    invalid_values = 0
    blank_run = 0
    for row_index in range(header_row + 1, min(worksheet.max_row, header_row + 160) + 1):
        role_values = {
            key: _cell_value(worksheet, row_index, col_index, merged_values)
            for key, col_index in columns.items()
        }
        row_text = " ".join(_stringify(value) for value in role_values.values() if _stringify(value))
        if not row_text:
            blank_run += 1
            if blank_run >= 12 and product_rows:
                break
            continue
        blank_run = 0
        if _is_total_row(row_text):
            break
        model_raw = _stringify(role_values.get("models"))
        name_spec = _stringify(role_values.get("name_spec"))
        source_no = _stringify(role_values.get("no"))
        has_numeric_payload = any(
            _to_number(role_values.get(key)) is not None for key in ("qty", "unit_price", "amount")
        )
        if not has_numeric_payload or not (model_raw or name_spec or source_no):
            continue
        if _is_extra_row(source_no, model_raw, name_spec):
            continue
        product_rows += 1
        raw_value = _cell_value(worksheet, row_index, candidate_column, merged_values)
        if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
            continue
        nonempty_values += 1
        try:
            barcode = normalize_barcode_value(raw_value)
        except ValueError:
            invalid_values += 1
            continue
        if not barcode:
            continue
        valid_values += 1
        if barcode.isdigit():
            digit_values += 1
            if BARCODE_PROFILE_MIN_DIGITS <= len(barcode) <= BARCODE_PROFILE_MAX_DIGITS:
                barcode_length_values += 1
    coverage = nonempty_values / product_rows if product_rows else 0.0
    digit_ratio = digit_values / valid_values if valid_values else 0.0
    length_ratio = barcode_length_values / digit_values if digit_values else 0.0
    confirmed = bool(
        product_rows
        and valid_values
        and invalid_values == 0
        and coverage >= 0.6
        and digit_ratio >= 0.8
        and length_ratio >= 0.6
    )
    return {
        "confirmed": confirmed,
        "product_row_count": product_rows,
        "nonempty_count": nonempty_values,
        "valid_count": valid_values,
        "digit_count": digit_values,
    }


def _barcode_column_result(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    *,
    header_row: int,
    column: int,
    method: str,
    columns: Mapping[str, int],
) -> tuple[int, dict[str, Any]]:
    header_value = _cell_value(worksheet, header_row, column, merged_values)
    profile = _barcode_value_profile(
        worksheet,
        merged_values,
        header_row=header_row,
        columns=columns,
        candidate_column=column,
    )
    return column, {
        "method": method,
        "column_index": column,
        "column_letter": get_column_letter(column),
        "header_normalized": _normalize_header_compact(header_value),
        "value_profile": {
            "product_row_count": profile["product_row_count"],
            "nonempty_count": profile["nonempty_count"],
            "valid_count": profile["valid_count"],
            "digit_count": profile["digit_count"],
        },
    }


def _has_confirmed_current_template_structure(columns: Mapping[str, int]) -> bool:
    return all(
        columns.get(role) == expected
        for role, expected in {
            "no": 1,
            "models": 2,
            "name_spec": 3,
            "qty": 5,
            "unit_price": 6,
            "amount": 7,
        }.items()
    )


def _ambiguous_barcode_column_error(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    header_row: int,
    candidates: Iterable[int],
) -> ValueError:
    return ValueError(
        "ambiguous barcode column: multiple equally confirmed candidates: "
        + _barcode_candidate_text(worksheet, merged_values, header_row, candidates)
    )


def _barcode_candidate_text(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    header_row: int,
    candidates: Iterable[int],
) -> str:
    return ", ".join(
        f"{get_column_letter(col_index)} ({_normalize_header_compact(_cell_value(worksheet, header_row, col_index, merged_values)) or 'empty'})"
        for col_index in candidates
    )


def _is_barcode_header(value: Any) -> bool:
    compact = _normalize_header_compact(value)
    if not compact:
        return False
    for alias in BARCODE_HEADER_ALIASES:
        if alias == "шк":
            if compact == alias:
                return True
            continue
        if alias in compact:
            return True
    return False


def _extract_metadata(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    *,
    header_row: int,
    filename: str,
    workbook_text_blocks: Iterable[str] | None = None,
) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "invoice_no": "",
        "invoice_date": "",
        "contract_no": "",
        "contract_date": "",
        "supplier_name": "",
        "customer_name": "",
        "currency": "",
        "declared_invoice_total": None,
    }
    scan_rows = range(1, min(max(header_row, 1), 40) + 1)
    for row_index in scan_rows:
        row = [
            _stringify(_cell_value(worksheet, row_index, col_index, merged_values))
            for col_index in range(1, worksheet.max_column + 1)
        ]
        for col_index, text in enumerate(row, start=1):
            if not text:
                continue
            next_value = _stringify(
                _cell_value(worksheet, row_index, min(col_index + 1, worksheet.max_column), merged_values)
            )
            _maybe_assign_metadata(metadata, text, next_value)
            if not metadata["currency"]:
                currency = _detect_currency(text)
                if currency:
                    metadata["currency"] = currency
            if metadata["declared_invoice_total"] is None and re.search(r"\btotal\b|合计|总计", text, re.IGNORECASE):
                number = _to_number(next_value) or _last_number(text)
                if number is not None:
                    metadata["declared_invoice_total"] = number

    text_blocks = [item for item in (workbook_text_blocks or []) if _stringify(item)]
    for index, text in enumerate(text_blocks):
        next_value = text_blocks[index + 1] if index + 1 < len(text_blocks) else ""
        _maybe_assign_metadata(metadata, text, next_value)
        if not metadata["currency"]:
            currency = _detect_currency(text)
            if currency:
                metadata["currency"] = currency
        if metadata["declared_invoice_total"] is None and re.search(r"\btotal\b|合计|总计", text, re.IGNORECASE):
            number = _to_number(next_value) or _last_number(text)
            if number is not None:
                metadata["declared_invoice_total"] = number

    filename_meta = _metadata_from_filename(filename)
    for key, value in filename_meta.items():
        if not metadata.get(key) and value:
            metadata[key] = value
    return metadata


def _extract_declared_invoice_total(
    worksheet: Worksheet,
    merged_values: Mapping[tuple[int, int], Any],
    *,
    header_row: int,
) -> float | None:
    for row_index in range(header_row + 1, min(worksheet.max_row, header_row + 120) + 1):
        row = [
            _cell_value(worksheet, row_index, col_index, merged_values)
            for col_index in range(1, worksheet.max_column + 1)
        ]
        row_text = " ".join(_stringify(value) for value in row if _stringify(value))
        if not _is_total_row(row_text):
            continue
        numeric_values = [
            float(value)
            for value in row
            if isinstance(value, (int, float)) and not isinstance(value, bool)
        ]
        if not numeric_values:
            numeric_values = [number for number in (_to_number(value) for value in row) if number is not None]
        if numeric_values:
            return numeric_values[-1]
    return None


def _maybe_assign_metadata(metadata: dict[str, Any], text: str, next_value: str) -> None:
    normalized = re.sub(r"\s+", " ", text).strip().replace("：", ":")
    lower = normalized.lower()
    value_after_colon = ""
    if ":" in normalized:
        value_after_colon = normalized.split(":", 1)[1].strip()
    candidate = value_after_colon or next_value
    if not candidate:
        return
    has_contract_no_label = bool(
        (
            "contract" in lower
            and ("no" in lower or "number" in lower or "№" in lower)
        )
        or re.search(r"合同\s*(号|编号|号码|no|number)", normalized, flags=re.IGNORECASE)
        or re.search(r"合约\s*(号|编号|号码|no|number)", normalized, flags=re.IGNORECASE)
    )
    has_contract_date_label = bool(
        ("contract" in lower and "date" in lower)
        or "date of contract" in lower
        or re.search(r"(合同|合约)\s*(日期|时间)", normalized)
        or "下单日期" in normalized
        or "订单日期" in normalized
    )
    if has_contract_no_label and not metadata["contract_no"]:
        metadata["contract_no"] = candidate
    elif has_contract_date_label and not metadata["contract_date"]:
        metadata["contract_date"] = _normalize_date(candidate) or candidate
    elif "invoice" in lower and ("no" in lower or "number" in lower) and not metadata["invoice_no"]:
        metadata["invoice_no"] = candidate
    elif "invoice" in lower and "date" in lower and not metadata["invoice_date"]:
        metadata["invoice_date"] = _normalize_date(candidate) or candidate
    elif "supplier" in lower and not metadata["supplier_name"]:
        metadata["supplier_name"] = candidate
    elif "customer" in lower and not metadata["customer_name"]:
        metadata["customer_name"] = candidate
    elif "currency" in lower and not metadata["currency"]:
        metadata["currency"] = _detect_currency(candidate) or candidate


def _metadata_from_filename(filename: str) -> dict[str, str]:
    stem = Path(str(filename or "")).stem
    result = {"invoice_no": "", "invoice_date": ""}
    if not stem:
        return result
    date_match = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", stem)
    if date_match:
        day, month, year = date_match.groups()
        if len(year) == 2:
            year = "20" + year
        result["invoice_date"] = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    invoice_match = re.search(r"(\d{2}[A-Z]{1,5}\d{2,})", stem, flags=re.IGNORECASE)
    if invoice_match:
        result["invoice_no"] = invoice_match.group(1).upper()
    return result


def _normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _stringify(value)
    if not text:
        return ""
    year_first = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", text)
    if year_first:
        year, month, day = year_first.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    date_match = re.search(r"(\d{1,2})([.\-/])(\d{1,2})\2(\d{2,4})", text)
    if not date_match:
        return ""
    first, separator, second, year = date_match.groups()
    if len(year) == 2:
        year = "20" + year
    if separator == "/" and int(first) <= 12 < int(second):
        month, day = first, second
    else:
        day, month = first, second
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def _extract_workbook_text_blocks(workbook_bytes: bytes) -> list[str]:
    blocks: list[str] = []
    try:
        with ZipFile(BytesIO(workbook_bytes)) as workbook_zip:
            for name in workbook_zip.namelist():
                lower_name = name.lower()
                if not (
                    lower_name.startswith("xl/drawings/")
                    and (lower_name.endswith(".xml") or lower_name.endswith(".vml"))
                ):
                    continue
                try:
                    root = ElementTree.fromstring(workbook_zip.read(name))
                except ElementTree.ParseError:
                    continue
                blocks.extend(_drawing_text_blocks(root))
    except BadZipFile:
        return []
    return [block for block in blocks if block]


def _drawing_text_blocks(root: ElementTree.Element) -> list[str]:
    blocks: list[str] = []
    for paragraph in root.iter():
        if _xml_local_name(paragraph.tag) != "p":
            continue
        parts = [
            str(node.text or "")
            for node in paragraph.iter()
            if _xml_local_name(node.tag) == "t" and str(node.text or "").strip()
        ]
        text = re.sub(r"\s+", " ", "".join(parts)).strip()
        if text:
            blocks.append(text)
    if blocks:
        return blocks
    parts = [
        str(node.text or "")
        for node in root.iter()
        if _xml_local_name(node.tag) == "t" and str(node.text or "").strip()
    ]
    text = re.sub(r"\s+", " ", " ".join(parts)).strip()
    return [text] if text else []


def _xml_local_name(tag: str) -> str:
    return str(tag).rsplit("}", 1)[-1]


def _detect_currency(value: Any) -> str:
    text = _stringify(value).upper()
    if "USD" in text or "$" in text:
        return "USD"
    if "CNY" in text or "RMB" in text or "¥" in text:
        return "RMB"
    if "EUR" in text:
        return "EUR"
    return ""


def _metadata_warnings(metadata: Mapping[str, Any]) -> list[str]:
    warnings: list[str] = []
    if not metadata.get("invoice_no"):
        warnings.append("invoice no was not detected; operator must verify metadata")
    if not metadata.get("invoice_date"):
        warnings.append("invoice date was not detected; operator must verify metadata")
    if not metadata.get("supplier_name"):
        warnings.append("supplier name was not detected; operator must verify metadata")
    return warnings


def _build_summary(lines: list[Mapping[str, Any]], *, declared_total: float | None) -> dict[str, Any]:
    product_qty_total = _sum_numeric(item.get("qty") for item in lines if item.get("line_type") == LINE_TYPE_PRODUCT)
    product_amount_total = _sum_numeric(item.get("amount") for item in lines if item.get("line_type") == LINE_TYPE_PRODUCT)
    extras_amount_total = _sum_numeric(item.get("amount") for item in lines if item.get("line_type") == LINE_TYPE_EXTRA)
    invoice_amount_total = round(product_amount_total + extras_amount_total, 2)
    checksum_error = bool(declared_total is not None and abs(round(declared_total - invoice_amount_total, 2)) > 0.02)
    return {
        "product_qty_total": product_qty_total,
        "product_amount_total": product_amount_total,
        "extras_amount_total": extras_amount_total,
        "invoice_amount_total": invoice_amount_total,
        "declared_invoice_total": declared_total,
        "checksum_error": checksum_error,
    }


def _is_extra_row(*values: Any) -> bool:
    lower = " ".join(_stringify(value) for value in values if _stringify(value)).lower()
    if not lower:
        return False
    strong_markers = (
        "opp",
        "label",
        "labels",
        "card",
        "cards",
        "shipping",
        "freight",
        "售后卡",
        "定制卡",
        "卡片",
        "标签",
        "标贴",
        "运费",
    )
    if any(marker in lower for marker in strong_markers):
        return True
    packaging_markers = ("bag", "packet", "package", "packaging", "袋", "包装")
    return any(marker in lower for marker in packaging_markers) and not detect_product_type(lower) and "iphone" not in lower


def _is_total_row(row_text: str) -> bool:
    return bool(re.search(r"\btotal\b|合计|总计|总值", _stringify(row_text), re.IGNORECASE))


def _normalize_header_compact(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _stringify(value)).casefold()
    return "".join(character for character in text if character.isalnum())


def normalize_barcode_value(value: Any) -> str:
    """Return an exact barcode identity or reject lossy Excel representations."""

    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        raise ValueError("boolean barcode value is invalid")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, Decimal):
        if not value.is_finite() or value != value.to_integral_value():
            raise ValueError("numeric barcode is not an exact integer")
        return format(value, "f").split(".", 1)[0]
    if isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer() or abs(value) > MAX_EXACT_FLOAT_INTEGER:
            raise ValueError("numeric barcode cannot be restored losslessly")
        return str(int(value))
    if not isinstance(value, str):
        raise ValueError(f"unsupported barcode cell type: {type(value).__name__}")
    text = value.replace("\u00a0", "").replace("\u202f", "")
    text = re.sub(r"\s+", "", text)
    if text.startswith("'"):
        text = text[1:]
    if not text:
        return ""
    if re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)[eE][+-]?\d+", text):
        raise ValueError("scientific-notation barcode text is not accepted")
    return text


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _stringify(value)
    if not text:
        return None
    text = text.replace("\u00a0", " ")
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _last_number(value: Any) -> float | None:
    matches = re.findall(r"[-+]?\d+(?:[,.]\d+)?", _stringify(value))
    if not matches:
        return None
    return _to_number(matches[-1])


def _sum_numeric(values: Iterable[Any]) -> float:
    total = 0.0
    for value in values:
        number = _to_number(value)
        if number is not None:
            total += number
    return round(total, 2)


def _raw_row_payload(row_index: int, row_values: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "worksheet_row": row_index,
        "values": {key: _stringify(value) for key, value in row_values.items()},
    }
