"""Atomic XLSX/ZIP export helpers for WB regional supply recommendations."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timezone
from hashlib import sha256
from io import BytesIO
import math
from pathlib import Path
import re
from typing import Any, Iterable, Mapping, Sequence
import unicodedata
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from packages.application.supplier_invoice_parser import normalize_barcode_value
from packages.contracts.wb_regional_supply import WbRegionalSupplyDistrictResult


WB_UPLOAD_TEMPLATE_PATH = Path(__file__).with_name("resources") / "wb_supply_upload_template.xlsx"
WB_UPLOAD_TEMPLATE_SHA256 = "db02c1d7f01759b24fca1b513d32600a61da2473db871b64d427d1cdbf21df79"
WB_UPLOAD_SHEET_NAME = "Sheet1"
WB_UPLOAD_HEADERS = ("Баркод", "Количество")

_VALID_BARCODE_STATUSES = {"ready", "manual"}
_WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}
_INVALID_PATH_CHARACTERS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_MULTIPLE_UNDERSCORES = re.compile(r"_+")


@dataclass(frozen=True)
class WbUploadRow:
    barcode: str
    quantity: int


class WbSupplyExportValidationError(ValueError):
    def __init__(self, issues: Sequence[str]) -> None:
        normalized = tuple(str(issue).strip() for issue in issues if str(issue).strip())
        self.issues = normalized
        super().__init__("; ".join(normalized))


def validate_raw_district_quantities(district_payload: Mapping[str, Any]) -> tuple[str, ...]:
    """Reject lossy/negative quantities before dataclass loading can coerce them."""

    issues: list[str] = []
    total = 0
    for row in district_payload.get("rows") or []:
        if not isinstance(row, Mapping):
            issues.append("обнаружена некорректная строка рекомендации")
            continue
        raw_quantity = row.get("allocated_qty")
        sku = _sku_reference(row)
        if raw_quantity is None or raw_quantity == "":
            continue
        quantity = _exact_integer(raw_quantity)
        if quantity is None:
            issues.append(f"{sku}: количество {raw_quantity!r} не является целым числом")
            continue
        if quantity < 0:
            issues.append(f"{sku}: количество {quantity} не может быть отрицательным")
            continue
        total += quantity

    expected_total = _exact_integer(district_payload.get("total_qty"))
    if expected_total is None or expected_total < 0:
        issues.append(
            "итог рекомендации должен быть целым неотрицательным числом, "
            f"получено {district_payload.get('total_qty')!r}"
        )
    elif total != expected_total:
        issues.append(f"итог строк {total} не совпадает с итогом рекомендации {expected_total}")
    return tuple(issues)


def build_wb_upload_rows(
    *,
    district: WbRegionalSupplyDistrictResult,
    nomenclature_items: Iterable[Mapping[str, Any]],
) -> tuple[WbUploadRow, ...]:
    """Resolve exact canonical barcodes and merge duplicate barcode quantities."""

    nomenclature_by_nm_id: dict[int, list[Mapping[str, Any]]] = {}
    for item in nomenclature_items:
        if not isinstance(item, Mapping) or not bool(item.get("is_active", True)):
            continue
        nm_id = _positive_integer(item.get("nm_id"))
        if nm_id > 0:
            nomenclature_by_nm_id.setdefault(nm_id, []).append(item)

    issues: list[str] = []
    quantity_by_barcode: dict[str, int] = {}
    barcode_order: list[str] = []
    source_total = 0

    for row in district.rows:
        quantity = _exact_integer(row.allocated_qty)
        sku = _sku_reference({"nm_id": row.nm_id, "sku_comment": row.sku_comment})
        if quantity is None:
            issues.append(f"{sku}: количество {row.allocated_qty!r} не является целым числом")
            continue
        if quantity < 0:
            issues.append(f"{sku}: количество {quantity} не может быть отрицательным")
            continue
        if quantity == 0:
            continue
        source_total += quantity

        candidates = nomenclature_by_nm_id.get(int(row.nm_id), [])
        if not candidates:
            issues.append(f"{sku}: отсутствует баркод в каноническом справочнике товаров")
            continue
        if len(candidates) > 1:
            issues.append(f"{sku}: найдено несколько активных записей номенклатуры")
            continue

        item = candidates[0]
        barcode_status = str(item.get("barcode_status") or "").strip().lower()
        raw_barcodes: list[Any] = []
        if item.get("barcode") not in (None, ""):
            raw_barcodes.append(item.get("barcode"))
        values = item.get("barcodes")
        if isinstance(values, (list, tuple)):
            raw_barcodes.extend(values)

        barcodes: list[str] = []
        invalid_barcode_error = ""
        for raw_barcode in raw_barcodes:
            try:
                barcode = normalize_barcode_value(raw_barcode)
            except ValueError as exc:
                invalid_barcode_error = str(exc)
                break
            if barcode and barcode not in barcodes:
                barcodes.append(barcode)
        if invalid_barcode_error:
            issues.append(f"{sku}: некорректный баркод ({invalid_barcode_error})")
            continue
        if not barcodes:
            issues.append(f"{sku}: отсутствует баркод в каноническом справочнике товаров")
            continue
        if len(barcodes) != 1 or barcode_status == "multiple":
            issues.append(f"{sku}: найдено несколько неоднозначных баркодов")
            continue
        if barcode_status and barcode_status not in _VALID_BARCODE_STATUSES:
            issues.append(f"{sku}: баркод имеет неподтверждённый статус {barcode_status!r}")
            continue

        barcode = barcodes[0]
        if barcode not in quantity_by_barcode:
            barcode_order.append(barcode)
            quantity_by_barcode[barcode] = 0
        quantity_by_barcode[barcode] += quantity

    expected_total = _exact_integer(district.total_qty)
    if expected_total is None or expected_total < 0:
        issues.append(f"итог рекомендации {district.total_qty!r} некорректен")
    elif source_total != expected_total:
        issues.append(f"итог положительных строк {source_total} не совпадает с итогом рекомендации {expected_total}")

    export_total = sum(quantity_by_barcode.values())
    if not issues and expected_total is not None and export_total != expected_total:
        issues.append(f"итог WB-файла {export_total} не совпадает с итогом рекомендации {expected_total}")
    if issues:
        raise WbSupplyExportValidationError(issues)

    return tuple(WbUploadRow(barcode=barcode, quantity=quantity_by_barcode[barcode]) for barcode in barcode_order)


def build_wb_upload_workbook_bytes(
    rows: Sequence[WbUploadRow],
    *,
    expected_total: int,
) -> bytes:
    """Populate a copy of the checked-in canonical WB workbook template."""

    template_bytes = _load_canonical_template_bytes()
    workbook = load_workbook(BytesIO(template_bytes))
    _validate_template_workbook(workbook)
    sheet = workbook[WB_UPLOAD_SHEET_NAME]

    prototype_styles = {
        column: copy(sheet.cell(row=2, column=column)._style)
        for column in (1, 2)
        if sheet.max_row >= 2
    }
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_index, item in enumerate(rows, start=2):
        barcode_cell = sheet.cell(row=row_index, column=1, value=str(item.barcode))
        quantity_cell = sheet.cell(row=row_index, column=2, value=int(item.quantity))
        if 1 in prototype_styles:
            barcode_cell._style = copy(prototype_styles[1])
        if 2 in prototype_styles:
            quantity_cell._style = copy(prototype_styles[2])
        barcode_cell.number_format = "@"
        quantity_cell.number_format = "0"

    last_row = 1 + len(rows)
    for table in sheet.tables.values():
        table.ref = f"A1:B{last_row}"
    if sheet.auto_filter.ref:
        sheet.auto_filter.ref = f"A1:B{last_row}"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    workbook_bytes = output.getvalue()
    _validate_generated_wb_workbook(
        workbook_bytes,
        expected_rows=rows,
        expected_total=expected_total,
    )
    return workbook_bytes


def recommendation_identity(*, report_date: str, calculation_id: str, ordinal: int) -> str:
    report_token = _report_date_token(report_date)
    return f"R{report_token}-{stable_run_token(calculation_id)}-{int(ordinal):03d}"


def recommendation_prefix(
    *,
    ordinal: int,
    recommendation_id: str,
    destination_name: str,
) -> str:
    destination = safe_path_component(destination_name, fallback="Направление", max_length=72)
    return safe_path_component(
        f"{int(ordinal):02d}_{recommendation_id}_{destination}",
        fallback=f"{int(ordinal):02d}_{recommendation_id}",
        max_length=96,
    )


def archive_filename(*, report_date: str, calculated_at: str, calculation_id: str) -> str:
    parsed_date = _parse_iso_date(report_date)
    timestamp = _parse_timestamp(calculated_at)
    if timestamp is not None:
        try:
            timestamp = timestamp.astimezone(ZoneInfo("Asia/Yekaterinburg"))
        except Exception:
            timestamp = timestamp.astimezone(timezone.utc)
        stamp_date = timestamp.date().isoformat()
        stamp_time = timestamp.strftime("%H-%M")
    else:
        stamp_date = (parsed_date or date(1970, 1, 1)).isoformat()
        stamp_time = "00-00"
    stem = safe_path_component(
        f"Рекомендации_поставок_{stamp_date}_{stamp_time}_{stable_run_token(calculation_id)}",
        fallback=f"supply_recommendations_{stamp_date}_{stamp_time}",
        max_length=190,
    )
    return f"{stem}.zip"


def safe_path_component(value: Any, *, fallback: str, max_length: int) -> str:
    normalized = unicodedata.normalize("NFC", str(value or "")).strip()
    normalized = _INVALID_PATH_CHARACTERS.sub("_", normalized)
    normalized = re.sub(r"\s+", "_", normalized)
    normalized = _MULTIPLE_UNDERSCORES.sub("_", normalized).strip(" ._")
    if not normalized:
        normalized = fallback
    if normalized.upper() in _WINDOWS_RESERVED_NAMES:
        normalized = f"_{normalized}"
    if len(normalized) > max_length:
        digest = sha256(normalized.encode("utf-8")).hexdigest()[:10]
        normalized = normalized[: max(1, max_length - len(digest) - 1)].rstrip(" ._") + "_" + digest
    return normalized.rstrip(" .") or fallback


def stable_run_token(calculation_id: str) -> str:
    raw = str(calculation_id or "").strip()
    if not raw:
        raise ValueError("ID запуска расчёта отсутствует")
    token = re.sub(r"[^A-Za-z0-9_-]+", "-", raw).strip("-_")
    if not token:
        token = "run-" + sha256(raw.encode("utf-8")).hexdigest()[:16]
    if len(token) > 40:
        token = token[:27].rstrip("-_") + "-" + sha256(raw.encode("utf-8")).hexdigest()[:12]
    return token


def _load_canonical_template_bytes() -> bytes:
    try:
        template_bytes = WB_UPLOAD_TEMPLATE_PATH.read_bytes()
    except OSError as exc:
        raise ValueError(f"Канонический WB-шаблон недоступен: {WB_UPLOAD_TEMPLATE_PATH}") from exc
    actual_sha256 = sha256(template_bytes).hexdigest()
    if actual_sha256 != WB_UPLOAD_TEMPLATE_SHA256:
        raise ValueError(
            "Канонический WB-шаблон изменён без обновления контракта: "
            f"expected={WB_UPLOAD_TEMPLATE_SHA256}, actual={actual_sha256}"
        )
    return template_bytes


def _validate_template_workbook(workbook: Any) -> None:
    if workbook.sheetnames != [WB_UPLOAD_SHEET_NAME]:
        raise ValueError(
            "Канонический WB-шаблон должен содержать только лист Sheet1: "
            f"получено {workbook.sheetnames}"
        )
    sheet = workbook[WB_UPLOAD_SHEET_NAME]
    headers = (sheet["A1"].value, sheet["B1"].value)
    if headers != WB_UPLOAD_HEADERS:
        raise ValueError(f"Заголовки WB-шаблона изменены: ожидалось {WB_UPLOAD_HEADERS}, получено {headers}")
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column > 2 and cell.value not in (None, ""):
                raise ValueError(f"WB-шаблон содержит дополнительный столбец: {cell.coordinate}")
    for table in sheet.tables.values():
        min_col, min_row, max_col, _ = _range_boundaries(table.ref)
        if (min_col, min_row, max_col) != (1, 1, 2):
            raise ValueError(f"Excel-таблица WB-шаблона должна начинаться в A1:B*: {table.ref}")


def _validate_generated_wb_workbook(
    workbook_bytes: bytes,
    *,
    expected_rows: Sequence[WbUploadRow],
    expected_total: int,
) -> None:
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    except Exception as exc:
        raise ValueError("Сформированный WB-файл повреждён") from exc
    try:
        _validate_template_workbook(workbook)
        sheet = workbook[WB_UPLOAD_SHEET_NAME]
        actual_rows: list[tuple[str, int]] = []
        for row_index in range(2, sheet.max_row + 1):
            barcode_cell = sheet.cell(row=row_index, column=1)
            quantity_cell = sheet.cell(row=row_index, column=2)
            if barcode_cell.value in (None, "") and quantity_cell.value in (None, ""):
                continue
            barcode = str(barcode_cell.value or "")
            quantity = _exact_integer(quantity_cell.value)
            if barcode_cell.data_type != "s" or barcode_cell.number_format != "@":
                raise ValueError(f"Баркод в строке {row_index} должен храниться как текст")
            if not barcode or quantity is None or quantity <= 0:
                raise ValueError(f"Некорректная строка WB-файла: {row_index}")
            actual_rows.append((barcode, quantity))
        expected_pairs = [(item.barcode, item.quantity) for item in expected_rows]
        if actual_rows != expected_pairs:
            raise ValueError(f"Строки WB-файла не совпадают с рекомендацией: {actual_rows!r}")
        if sum(quantity for _, quantity in actual_rows) != int(expected_total):
            raise ValueError("Итог WB-файла не совпадает с итогом рекомендации")
        expected_ref = f"A1:B{1 + len(expected_rows)}"
        for table in sheet.tables.values():
            if table.ref != expected_ref:
                raise ValueError(
                    f"Диапазон Excel-таблицы WB-файла не расширен: {table.ref}, ожидалось {expected_ref}"
                )
    finally:
        workbook.close()


def _range_boundaries(ref: str) -> tuple[int, int, int, int]:
    from openpyxl.utils.cell import range_boundaries

    return range_boundaries(ref)


def _exact_integer(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return int(value)
    return None


def _positive_integer(value: Any) -> int:
    numeric = _exact_integer(value)
    return numeric if numeric is not None and numeric > 0 else 0


def _sku_reference(row: Mapping[str, Any]) -> str:
    nm_id = row.get("nm_id")
    label = str(row.get("sku_comment") or row.get("sku_label") or "").strip()
    if label:
        return f"SKU nmId={nm_id} ({label})"
    return f"SKU nmId={nm_id}"


def _report_date_token(value: str) -> str:
    parsed = _parse_iso_date(value)
    if parsed is None:
        raise ValueError(f"Дата расчёта некорректна: {value!r}")
    return parsed.strftime("%y%m%d")


def _parse_iso_date(value: str) -> date | None:
    try:
        return date.fromisoformat(str(value or "").strip())
    except ValueError:
        return None


def _parse_timestamp(value: str) -> datetime | None:
    normalized = str(value or "").strip()
    if not normalized:
        return None
    try:
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed
