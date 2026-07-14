"""Server-owned supplier invoice shipment registry block."""

from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
from io import BytesIO
import json
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Any, Iterable, Mapping, Protocol
from uuid import uuid4
import zlib

from openpyxl import Workbook, load_workbook

from packages.adapters.official_api_runtime import OfficialApiRuntimeError
from packages.adapters.wb_content import (
    HttpBackedWbContentSource,
    WbContentCard,
    WbContentHttpStatusError,
    WbContentTransportError,
)
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime
from packages.application.ff_stock_ledger import FfStockLedgerBlock
from packages.application.supplier_invoice_parser import (
    extract_iphone_model_keys,
    normalize_invoice_model,
    parse_supplier_invoice_xlsx,
)
from packages.application.supplier_financial_documents import build_financial_summary
from packages.application.supplier_shipment_status import (
    supplier_business_today,
    validate_supplier_factual_dates,
)
from packages.contracts.supplier_shipments import (
    DEFAULT_SUPPLIER_NAME,
    LINE_TYPE_EXTRA,
    LINE_TYPE_PRODUCT,
    MATCH_STATUS_AMBIGUOUS,
    MATCH_STATUS_MATCHED,
    MATCH_STATUS_MATCHED_BY_COMPATIBILITY,
    MATCH_STATUS_UNMATCHED,
    ORDER_STATUS_DEFAULT,
    ORDER_STATUS_ACCEPTED_FF,
    ORDER_STATUS_IN_TRANSIT,
    ORDER_STATUS_PRODUCTION,
    ORDER_STATUSES,
    PRICE_CONFORMITY_CHECK_MODE_INITIAL_PARSE,
    PRICE_CONFORMITY_CHECK_MODE_MANUAL_RECHECK,
    PRICE_CONFORMITY_CHECK_MODE_MIGRATION_BACKFILL,
    PRICE_CONFORMITY_CHECK_MODE_NOT_CHECKED,
    PRICE_CONFORMITY_CHECK_MODES,
    PRICE_CONFORMITY_STATUS_INVOICE_PRICE_MISSING,
    PRICE_CONFORMITY_STATUS_MATCHED,
    PRICE_CONFORMITY_STATUS_MISMATCHED,
    PRICE_CONFORMITY_STATUS_NOT_CHECKED,
    PRICE_CONFORMITY_STATUS_REFERENCE_PRICE_MISSING,
    PRICE_CONFORMITY_STATUS_SKU_NOT_FOUND,
    PRICE_CONFORMITY_STATUSES,
    SHIPMENT_STATUS_ALL_MATCHED,
    SHIPMENT_STATUS_CHECKSUM_ERROR,
    SHIPMENT_STATUS_HAS_UNMATCHED,
    SHIPMENT_STATUS_MANUAL_OVERRIDE,
    NOMENCLATURE_BARCODE_SOURCE_ERROR,
    NOMENCLATURE_BARCODE_SOURCE_MANUAL,
    NOMENCLATURE_BARCODE_SOURCE_MISSING,
    NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT,
    NOMENCLATURE_BARCODE_SOURCES,
    NOMENCLATURE_BARCODE_STATUS_MANUAL,
    NOMENCLATURE_BARCODE_STATUS_MISSING,
    NOMENCLATURE_BARCODE_STATUS_MULTIPLE,
    NOMENCLATURE_BARCODE_STATUS_READY,
    NOMENCLATURE_BARCODE_STATUS_SYNC_ERROR,
    NOMENCLATURE_BARCODE_STATUS_TOKEN_MISSING,
    NOMENCLATURE_BARCODE_STATUSES,
    SUPPLIER_INVOICE_CONTENT_TYPE,
    SUPPLIER_INVOICE_PARSER_VERSION,
    TRADE_DOCUMENT_CONTRACT_PARSER_VERSION,
    TRADE_DOCUMENT_ALLOWED_EXTENSIONS,
    TRADE_DOCUMENT_CONTENT_TYPES_BY_EXTENSION,
    TRADE_DOCUMENT_LINK_SOURCE_MIGRATION,
    TRADE_DOCUMENT_LINK_SOURCE_OPERATOR,
    TRADE_DOCUMENT_LINK_SOURCE_SUPPLIER_SHIPMENT_AUTO,
    TRADE_DOCUMENT_SOURCE_MIGRATION_EXISTING_SUPPLIER_INVOICE,
    TRADE_DOCUMENT_SOURCE_SETTINGS_UPLOAD,
    TRADE_DOCUMENT_SOURCE_SUPPLIER_SHIPMENT_PARSE,
    TRADE_DOCUMENT_STATUS_ACTIVE,
    TRADE_DOCUMENT_TYPE_CONTRACT,
    TRADE_DOCUMENT_TYPE_INVOICE,
    TRADE_DOCUMENT_TYPES,
)


NOMENCLATURE_XLSX_FILENAME = "nomenclature.xlsx"
NOMENCLATURE_XLSX_CONTENT_TYPE = SUPPLIER_INVOICE_CONTENT_TYPE
NOMENCLATURE_XLSX_HEADERS = [
    "ID строки",
    "Включено",
    "Скрыто",
    "nmId",
    "ШК / barcode",
    "Все ШК",
    "Источник ШК",
    "Статус ШК",
    "Артикул продавца WB / vendorCode",
    "Название WB",
    "WB subject",
    "WB updatedAt",
    "Статус WB sync",
    "Номенклатура",
    "Группа",
    "Match key",
    "Цена закупки, ¥",
    "Совместимые модели",
    "Ключи совместимости",
    "Обновлено",
]
NOMENCLATURE_PRODUCT_TYPE_LABELS = {
    "clear": "Clean",
    "clean": "Clean",
    "anti_spy": "Anti-spy",
    "matte": "Matte",
    "no_frame_clean": "No Frame Clean",
    "no_frame_anti_spy": "No Frame Anti-spy",
    "no_frame_matte": "No Frame Matte",
    "extra": "Доп. строка",
    "other": "Другое",
}
NOMENCLATURE_LEGACY_PRODUCT_TYPE_LABELS = {
    "прозрачное": "clear",
    "антишпион": "anti_spy",
    "матовое": "matte",
}
NOMENCLATURE_PRODUCT_TYPE_BY_LABEL = {
    **{key: key for key in NOMENCLATURE_PRODUCT_TYPE_LABELS},
    **{value.casefold(): key for key, value in NOMENCLATURE_PRODUCT_TYPE_LABELS.items()},
    **NOMENCLATURE_LEGACY_PRODUCT_TYPE_LABELS,
}
DEFAULT_SKU_GROUPS: tuple[dict[str, Any], ...] = (
    {"group_key": "clean", "label": "Clean", "aliases": ["clean", "clear", "transparent", "прозрач"]},
    {"group_key": "anti_spy", "label": "Anti-spy", "aliases": ["anti-spy", "anti spy", "antispy", "privacy", "антишпион"]},
    {"group_key": "matte", "label": "Matte", "aliases": ["matte", "matt", "матов"]},
    {"group_key": "no_frame_clean", "label": "No Frame Clean", "aliases": ["no frame clean", "noframe clean", "no frame clear"]},
    {
        "group_key": "no_frame_anti_spy",
        "label": "No Frame Anti-spy",
        "aliases": ["no frame anti-spy", "no frame anti spy", "noframe anti-spy", "noframe privacy"],
    },
    {"group_key": "no_frame_matte", "label": "No Frame Matte", "aliases": ["no frame matte", "noframe matte", "no frame matt"]},
    {"group_key": "extra", "label": "Доп. строка", "aliases": []},
    {"group_key": "other", "label": "Другое", "aliases": []},
)
WB_CARD_SYNC_SOURCE = "wb_content_cards"
PRICE_CONFORMITY_MONEY_QUANT = Decimal("0.01")
APPROX_YUAN_RATE_QUANT = Decimal("0.0001")
PRICE_CONFORMITY_YUAN_CURRENCIES = {"CNY", "CNH", "RMB", "YUAN", "YUANS", "¥", "￥", "元"}
CONTRACT_OCR_PREFERRED_LANGUAGES = ("eng", "chi_sim", "rus")
CONTRACT_OCR_PSMS = ("6", "11", "4", "3")
CONTRACT_PDF_OCR_STRATEGIES: tuple[dict[str, Any], ...] = (
    {"id": "pdf_120dpi_full", "dpi": 120, "gray": False, "crop": None, "render_timeout": 60, "ocr_timeout": 60},
    {"id": "pdf_120dpi_gray", "dpi": 120, "gray": True, "crop": None, "render_timeout": 60, "ocr_timeout": 60},
    {"id": "pdf_150dpi_full", "dpi": 150, "gray": False, "crop": None, "render_timeout": 75, "ocr_timeout": 60},
    {"id": "pdf_150dpi_gray", "dpi": 150, "gray": True, "crop": None, "render_timeout": 75, "ocr_timeout": 60},
    {"id": "pdf_200dpi_full", "dpi": 200, "gray": False, "crop": None, "render_timeout": 90, "ocr_timeout": 75},
    {"id": "pdf_300dpi_full", "dpi": 300, "gray": False, "crop": None, "render_timeout": 120, "ocr_timeout": 90},
    {"id": "pdf_300dpi_gray", "dpi": 300, "gray": True, "crop": None, "render_timeout": 120, "ocr_timeout": 90},
    {"id": "pdf_300dpi_top", "dpi": 300, "gray": False, "crop": "top", "render_timeout": 120, "ocr_timeout": 90},
    {"id": "pdf_400dpi_top", "dpi": 400, "gray": False, "crop": "top", "render_timeout": 150, "ocr_timeout": 90},
    {"id": "pdf_600dpi_top", "dpi": 600, "gray": True, "crop": "top", "render_timeout": 180, "ocr_timeout": 120},
)
_TESSERACT_LANGUAGES_CACHE: list[str] | None = None


class NomenclatureBarcodeSource(Protocol):
    def fetch_barcodes_by_nm_ids(self, nm_ids: list[int]) -> Mapping[int, Any]:
        raise NotImplementedError

    def fetch_cards(self, *, limit: int | None = None, max_pages: int | None = None) -> list[Any]:
        raise NotImplementedError


class SupplierShipmentsBlock:
    def __init__(
        self,
        *,
        runtime: RegistryUploadDbBackedRuntime,
        barcode_source: NomenclatureBarcodeSource | None = None,
        timestamp_factory: callable | None = None,
    ) -> None:
        self.runtime = runtime
        self.barcode_source = barcode_source or HttpBackedWbContentSource()
        self.timestamp_factory = timestamp_factory or _default_timestamp_factory

    def list_shipments(self) -> dict[str, Any]:
        self.migrate_existing_supplier_shipments_into_trade_documents()
        rows = self.runtime.list_supplier_shipments()
        return {
            "contract_name": "sheet_vitrina_v1_supplier_shipments",
            "status": "ok",
            "shipments": [self._with_approx_cost_fields(self._with_document_fields(_with_invoice_download_path(row))) for row in rows],
        }

    def parse_upload(
        self,
        workbook_bytes: bytes,
        *,
        uploaded_filename: str | None = None,
        uploaded_content_type: str | None = None,
    ) -> dict[str, Any]:
        filename = _safe_filename(uploaded_filename or "supplier-invoice.xlsx")
        if not filename.lower().endswith(".xlsx"):
            raise ValueError("supplier invoice upload must be an .xlsx file")
        created_at = self.timestamp_factory()
        parsed_payload = parse_supplier_invoice_xlsx(
            workbook_bytes,
            filename=filename,
            aliases=self._active_nomenclature_aliases(),
        )
        parsed_payload["metadata"] = _supplier_order_metadata(parsed_payload.get("metadata"))
        nomenclature_items = self._active_nomenclature_items()
        parsed_payload["lines"] = _apply_nomenclature_matches(
            [dict(item) for item in parsed_payload.get("lines") or []],
            nomenclature_items,
        )
        parsed_payload["lines"] = _apply_price_conformity_checks(
            parsed_payload.get("lines") or [],
            nomenclature_items,
            checked_at=created_at,
            mode=PRICE_CONFORMITY_CHECK_MODE_INITIAL_PARSE,
            default_currency=str(parsed_payload.get("metadata", {}).get("currency") or ""),
        )
        upload_id = "upl_" + uuid4().hex
        sha256 = hashlib.sha256(workbook_bytes).hexdigest()
        relative_path = self._write_runtime_file(
            root_kind="uploads",
            entity_id=upload_id,
            filename=filename,
            body=workbook_bytes,
        )
        content_type = str(uploaded_content_type or "").strip() or SUPPLIER_INVOICE_CONTENT_TYPE
        self.runtime.save_supplier_shipment_upload(
            upload_id=upload_id,
            created_at=created_at,
            source_filename=filename,
            content_type=content_type,
            source_file_sha256=sha256,
            source_file_path=relative_path,
            parser_version=SUPPLIER_INVOICE_PARSER_VERSION,
            parsed_payload=parsed_payload,
        )
        payload = deepcopy(parsed_payload)
        payload.update(
            {
                "upload_id": upload_id,
                "created_at": created_at,
                "source_filename": filename,
                "source_file_sha256": sha256,
                "content_type": content_type,
                "contract_candidates": self.find_contract_candidates(
                    str(payload.get("metadata", {}).get("contract_no") or ""),
                    str(payload.get("metadata", {}).get("contract_date") or ""),
                ),
            }
        )
        return payload

    def create_shipment(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        upload_id = str(payload.get("upload_id") or "").strip()
        if not upload_id:
            raise ValueError("upload_id is required")
        upload = self.runtime.load_supplier_shipment_upload(upload_id)
        if upload is None:
            raise ValueError(f"supplier shipment upload not found: {upload_id}")
        now = self.timestamp_factory()
        business_today = supplier_business_today(timestamp=now)
        edited_payload = _resolve_edited_payload(payload, fallback=upload["parsed_payload"])
        shipment_date = _validate_iso_date(str(payload.get("shipment_date") or edited_payload.get("shipment_date") or ""))
        actual_shipment_date = _resolve_optional_date_field(payload, edited_payload, None, "actual_shipment_date")
        actual_ff_acceptance_date = _resolve_optional_date_field(payload, edited_payload, None, "actual_ff_acceptance_date")
        approx_yuan_rate = _resolve_optional_positive_decimal_field(payload, edited_payload, None, "approx_yuan_rate")
        status_resolution = validate_supplier_factual_dates(
            actual_shipment_date=actual_shipment_date,
            actual_ff_acceptance_date=actual_ff_acceptance_date,
            business_today=business_today,
        )
        order_status = status_resolution.order_status
        if "order_status" in payload:
            requested_status = _normalize_order_status(payload.get("order_status"))
            if requested_status != order_status:
                raise ValueError(
                    "Статус поставки вычисляется из фактических дат и не может быть задан вручную."
                )
        metadata, lines, warnings, errors, summary, match_status = _normalize_edit_payload(
            edited_payload,
            shipment_date=shipment_date,
            force_manual_override=False,
        )
        nomenclature_items = self._active_nomenclature_items()
        lines = _apply_nomenclature_matches(lines, nomenclature_items)
        _assert_atomic_supplier_product_matching(lines)
        lines = _apply_price_conformity_checks(
            lines,
            nomenclature_items,
            checked_at=now,
            mode=PRICE_CONFORMITY_CHECK_MODE_INITIAL_PARSE,
            default_currency=str(metadata.get("currency") or ""),
        )
        summary = _recalculate_summary(lines, declared_total=_optional_number(metadata.get("declared_invoice_total")))
        match_status = _shipment_match_status(lines, checksum_error=summary["checksum_error"])
        shipment_id = "sup_" + uuid4().hex
        source_filename = str(upload.get("source_filename") or "supplier-invoice.xlsx")
        source_path = self._copy_upload_to_shipment_file(
            upload_path=str(upload.get("source_file_path") or ""),
            shipment_id=shipment_id,
            filename=source_filename,
        )
        invoice_document = self._create_or_load_supplier_invoice_document(
            shipment_id=shipment_id,
            upload_id=upload_id,
            source_filename=source_filename,
            content_type=str(upload.get("content_type") or SUPPLIER_INVOICE_CONTENT_TYPE),
            source_file_sha256=str(upload.get("source_file_sha256") or ""),
            source_file_path=source_path,
            parser_version=str(upload.get("parser_version") or SUPPLIER_INVOICE_PARSER_VERSION),
            metadata=metadata,
            warnings=warnings,
            errors=errors,
            parsed_payload=edited_payload,
            created_at=now,
        )
        header = {
            "shipment_id": shipment_id,
            "created_at": now,
            "updated_at": now,
            "shipment_date": shipment_date,
            "actual_shipment_date": actual_shipment_date,
            "actual_ff_acceptance_date": actual_ff_acceptance_date,
            "order_status": order_status,
            "invoice_no": metadata.get("invoice_no") or "",
            "invoice_date": metadata.get("invoice_date") or "",
            "contract_no": metadata.get("contract_no") or "",
            "contract_date": metadata.get("contract_date") or "",
            "supplier_name": metadata.get("supplier_name") or "",
            "customer_name": metadata.get("customer_name") or "",
            "currency": metadata.get("currency") or "",
            "approx_yuan_rate": approx_yuan_rate,
            "product_qty_total": summary["product_qty_total"],
            "product_amount_total": summary["product_amount_total"],
            "extras_amount_total": summary["extras_amount_total"],
            "invoice_amount_total": summary["invoice_amount_total"],
            "declared_invoice_total": summary.get("declared_invoice_total"),
            "match_status": match_status,
            "source_filename": source_filename,
            "source_file_sha256": upload.get("source_file_sha256") or "",
            "source_file_path": source_path,
            "invoice_document_id": invoice_document.get("document_id") or "",
            "parser_version": upload.get("parser_version") or SUPPLIER_INVOICE_PARSER_VERSION,
            "warnings": warnings,
            "errors": errors,
        }
        self.runtime.save_supplier_shipment(header=header, lines=lines)
        if actual_ff_acceptance_date:
            self._record_ff_stock_receipt({"header": header, "lines": lines})
            self._materialize_ff_cost_layer(shipment_id)
        self._autolink_invoice_contract_from_metadata(
            invoice_document_id=str(invoice_document.get("document_id") or ""),
            contract_no=str(metadata.get("contract_no") or ""),
            contract_date=str(metadata.get("contract_date") or ""),
            linked_by="system",
            source=TRADE_DOCUMENT_LINK_SOURCE_SUPPLIER_SHIPMENT_AUTO,
        )
        return self.get_shipment(shipment_id)

    def get_shipment(self, shipment_id: str) -> dict[str, Any]:
        self.migrate_existing_supplier_shipments_into_trade_documents()
        detail = self.runtime.load_supplier_shipment(shipment_id)
        if detail is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        return self._with_approx_cost_fields(self._with_document_fields(_detail_payload(detail)))

    def update_shipment(
        self,
        shipment_id: str,
        payload: Mapping[str, Any],
    ) -> dict[str, Any]:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        now = self.timestamp_factory()
        business_today = supplier_business_today(timestamp=now)
        edited_payload = _resolve_edited_payload(payload, fallback=_detail_payload(existing))
        shipment_date = _validate_iso_date(
            str(payload.get("shipment_date") or edited_payload.get("shipment_date") or existing["header"].get("shipment_date") or "")
        )
        existing_header = dict(existing["header"])
        actual_shipment_date = _resolve_optional_date_field(payload, edited_payload, existing_header, "actual_shipment_date")
        actual_ff_acceptance_date = _resolve_optional_date_field(payload, edited_payload, existing_header, "actual_ff_acceptance_date")
        status_resolution = validate_supplier_factual_dates(
            actual_shipment_date=actual_shipment_date,
            actual_ff_acceptance_date=actual_ff_acceptance_date,
            business_today=business_today,
        )
        factual_date_changed = (
            str(existing_header.get("actual_shipment_date") or "").strip()
            != actual_shipment_date
        )
        if factual_date_changed:
            raise ValueError(
                "Изменение фактической даты отгрузки требует audited correction flow."
            )
        existing_ff_acceptance_date = str(existing_header.get("actual_ff_acceptance_date") or "").strip()
        if existing_ff_acceptance_date and actual_ff_acceptance_date != existing_ff_acceptance_date:
            if self._has_current_ff_cost_layer(shipment_id):
                raise ValueError(
                    "actual_ff_acceptance_date cannot be cleared or changed after ФФ cost layer materialization"
                )
        approx_yuan_rate = _resolve_optional_positive_decimal_field(payload, edited_payload, existing_header, "approx_yuan_rate")
        metadata, lines, warnings, errors, summary, match_status = _normalize_edit_payload(
            edited_payload,
            shipment_date=shipment_date,
            force_manual_override=False,
        )
        _assert_atomic_supplier_product_matching(lines)
        order_status = status_resolution.order_status
        if "order_status" in payload:
            requested_status = _normalize_order_status(payload.get("order_status"))
            if requested_status != order_status:
                raise ValueError(
                    "Статус поставки вычисляется из фактических дат и не может быть задан вручную."
                )
        cost_affecting_changed = (
            [dict(item) for item in existing.get("lines") or []] != [dict(item) for item in lines]
            or str(existing_header.get("currency") or "") != str(metadata.get("currency") or "")
            or _optional_number(existing_header.get("approx_yuan_rate")) != _optional_number(approx_yuan_rate)
        )
        header = {
            **existing_header,
            "updated_at": now,
            "shipment_date": shipment_date,
            "actual_shipment_date": actual_shipment_date,
            "actual_ff_acceptance_date": actual_ff_acceptance_date,
            "order_status": order_status,
            "invoice_no": metadata.get("invoice_no") or "",
            "invoice_date": metadata.get("invoice_date") or "",
            "contract_no": metadata.get("contract_no") or "",
            "contract_date": metadata.get("contract_date") or "",
            "supplier_name": metadata.get("supplier_name") or "",
            "customer_name": metadata.get("customer_name") or "",
            "currency": metadata.get("currency") or "",
            "approx_yuan_rate": approx_yuan_rate,
            "product_qty_total": summary["product_qty_total"],
            "product_amount_total": summary["product_amount_total"],
            "extras_amount_total": summary["extras_amount_total"],
            "invoice_amount_total": summary["invoice_amount_total"],
            "declared_invoice_total": summary.get("declared_invoice_total"),
            "match_status": match_status,
            "warnings": warnings,
            "errors": errors,
        }
        if cost_affecting_changed:
            header["expenses_complete"] = False
        self.runtime.save_supplier_shipment(header=header, lines=lines)
        if cost_affecting_changed:
            from packages.application.own_product_capital import OwnProductCapitalBlock

            OwnProductCapitalBlock(
                runtime=self.runtime,
                timestamp_factory=self.timestamp_factory,
            ).set_expenses_certification(
                shipment_id=shipment_id,
                expenses_complete=False,
            )
        if actual_ff_acceptance_date:
            self._record_ff_stock_receipt({"header": header, "lines": lines})
            self._materialize_ff_cost_layer(shipment_id)
        if "contract_document_id" in payload:
            contract_document_id = str(payload.get("contract_document_id") or "").strip()
            if contract_document_id:
                self.link_shipment_contract(
                    shipment_id,
                    contract_document_id=contract_document_id,
                    linked_by="operator",
                    source=TRADE_DOCUMENT_LINK_SOURCE_OPERATOR,
                )
            else:
                self.unlink_shipment_contract(shipment_id)
        return self.get_shipment(shipment_id)

    def update_order_status(self, shipment_id: str, order_status: Any) -> dict[str, Any]:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        raise ValueError(
            "order_status вычисляется из фактических дат; ручной status-only PATCH не поддерживается."
        )

    def factual_date_change_required(self, shipment_id: str, payload: Mapping[str, Any]) -> bool:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        edited_payload = _resolve_edited_payload(payload, fallback=_detail_payload(existing))
        desired = _resolve_optional_date_field(
            payload,
            edited_payload,
            dict(existing["header"]),
            "actual_shipment_date",
        )
        return desired != str(existing["header"].get("actual_shipment_date") or "").strip()

    def desired_actual_shipment_date(
        self,
        shipment_id: str,
        payload: Mapping[str, Any],
    ) -> str:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        edited_payload = _resolve_edited_payload(payload, fallback=_detail_payload(existing))
        return _resolve_optional_date_field(
            payload,
            edited_payload,
            dict(existing["header"]),
            "actual_shipment_date",
        )

    def factual_date_correction_has_other_changes(
        self,
        shipment_id: str,
        payload: Mapping[str, Any],
    ) -> bool:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        header = dict(existing["header"])
        edited_payload = _resolve_edited_payload(payload, fallback=_detail_payload(existing))
        shipment_date = _validate_iso_date(
            str(
                payload.get("shipment_date")
                or edited_payload.get("shipment_date")
                or header.get("shipment_date")
                or ""
            )
        )
        acceptance = _resolve_optional_date_field(
            payload,
            edited_payload,
            header,
            "actual_ff_acceptance_date",
        )
        approx_rate = _resolve_optional_positive_decimal_field(
            payload,
            edited_payload,
            header,
            "approx_yuan_rate",
        )
        metadata, lines, _, _, _, _ = _normalize_edit_payload(
            edited_payload,
            shipment_date=shipment_date,
            force_manual_override=False,
        )
        header_fields = (
            "invoice_no",
            "invoice_date",
            "contract_no",
            "contract_date",
            "supplier_name",
            "customer_name",
            "currency",
        )
        return bool(
            shipment_date != str(header.get("shipment_date") or "")
            or acceptance != str(header.get("actual_ff_acceptance_date") or "").strip()
            or _optional_number(approx_rate) != _optional_number(header.get("approx_yuan_rate"))
            or any(
                str(metadata.get(field) or "") != str(header.get(field) or "")
                for field in header_fields
            )
            or [dict(item) for item in lines]
            != [dict(item) for item in existing.get("lines") or []]
            or "contract_document_id" in payload
        )

    def update_expenses_complete(self, shipment_id: str, expenses_complete: Any) -> dict[str, Any]:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        normalized = _normalize_bool_field(expenses_complete, field_name="expenses_complete")
        updated = self.runtime.update_supplier_shipment_expenses_complete(
            shipment_id=shipment_id,
            expenses_complete=normalized,
            updated_at=self.timestamp_factory(),
        )
        if not updated:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        from packages.application.own_product_capital import OwnProductCapitalBlock

        OwnProductCapitalBlock(
            runtime=self.runtime,
            timestamp_factory=self.timestamp_factory,
        ).set_expenses_certification(
            shipment_id=shipment_id,
            expenses_complete=normalized,
        )
        return self.get_shipment(shipment_id)

    def _materialize_ff_cost_layer(self, shipment_id: str) -> None:
        from packages.application.our_wb_costs import OurWbCostBlock

        cost_block = OurWbCostBlock(
            runtime=self.runtime,
            timestamp_factory=self.timestamp_factory,
        )
        cost_block.materialize_supplier_ff_cost_layer(shipment_id)
        cost_block.materialize_wb_supply_cost_layers()
        cost_block.materialize_opening_baseline()
        cost_block.materialize_daily_state()

    def _record_ff_stock_receipt(self, shipment_detail: Mapping[str, Any]) -> dict[str, Any] | None:
        return FfStockLedgerBlock(
            runtime=self.runtime,
            timestamp_factory=self.timestamp_factory,
        ).record_supplier_acceptance(shipment_detail)

    def _has_current_ff_cost_layer(self, shipment_id: str) -> bool:
        from packages.application.our_wb_costs import OurWbCostBlock

        return OurWbCostBlock(
            runtime=self.runtime,
            timestamp_factory=self.timestamp_factory,
        ).has_current_supplier_ff_cost_layer(shipment_id)

    def recheck_shipment_prices(self, shipment_id: str, *, actor: str = "", context: Mapping[str, Any] | None = None) -> dict[str, Any]:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        now = self.timestamp_factory()
        header = dict(existing["header"])
        lines = _apply_price_conformity_checks(
            [dict(item) for item in existing.get("lines") or []],
            self._active_nomenclature_items(),
            checked_at=now,
            mode=PRICE_CONFORMITY_CHECK_MODE_MANUAL_RECHECK,
            actor=actor,
            context=context or {},
            default_currency=str(header.get("currency") or ""),
        )
        header["updated_at"] = now
        self.runtime.save_supplier_shipment(header=header, lines=lines)
        return self.get_shipment(shipment_id)

    def backfill_price_conformity_checks(self) -> dict[str, Any]:
        shipments = self.runtime.list_supplier_shipments()
        now = self.timestamp_factory()
        nomenclature_items = self._active_nomenclature_items()
        processed_shipments = 0
        skipped_shipments = 0
        matched = 0
        mismatched = 0
        missing = 0
        updated_line_count = 0
        for shipment in shipments:
            shipment_id = str(shipment.get("shipment_id") or "")
            if not shipment_id:
                continue
            detail = self.runtime.load_supplier_shipment(shipment_id)
            if detail is None:
                continue
            original_lines = [dict(item) for item in detail.get("lines") or []]
            next_lines = _apply_price_conformity_checks(
                original_lines,
                nomenclature_items,
                checked_at=now,
                mode=PRICE_CONFORMITY_CHECK_MODE_MIGRATION_BACKFILL,
                only_missing=True,
                default_currency=str((detail.get("header") or {}).get("currency") or ""),
            )
            changed = next_lines != original_lines
            if not changed:
                skipped_shipments += 1
                continue
            header = dict(detail["header"])
            self.runtime.save_supplier_shipment(header=header, lines=next_lines)
            processed_shipments += 1
            for line in next_lines:
                if not _line_was_backfilled(original_lines, line):
                    continue
                updated_line_count += 1
                status = str(line.get("price_conformity_status") or "")
                if status == PRICE_CONFORMITY_STATUS_MATCHED:
                    matched += 1
                elif status == PRICE_CONFORMITY_STATUS_MISMATCHED:
                    mismatched += 1
                elif status in {
                    PRICE_CONFORMITY_STATUS_SKU_NOT_FOUND,
                    PRICE_CONFORMITY_STATUS_REFERENCE_PRICE_MISSING,
                    PRICE_CONFORMITY_STATUS_INVOICE_PRICE_MISSING,
                }:
                    missing += 1
        return {
            "contract_name": "sheet_vitrina_v1_supplier_price_conformity_backfill",
            "status": "ok",
            "checked_at": now,
            "found_shipments": len(shipments),
            "processed_shipments": processed_shipments,
            "skipped_shipments": skipped_shipments,
            "updated_line_count": updated_line_count,
            "matched_count": matched,
            "mismatched_count": mismatched,
            "missing_count": missing,
        }

    def delete_shipment(self, shipment_id: str) -> dict[str, Any]:
        detail = self.runtime.load_supplier_shipment(shipment_id)
        if detail is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        header = dict(detail.get("header") or {})
        invoice_document_id = str(header.get("invoice_document_id") or "")
        deleted = self.runtime.delete_supplier_shipment(shipment_id)
        if not deleted:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        if invoice_document_id:
            self.runtime.delete_invoice_contract_link(invoice_document_id)
            try:
                self.runtime.archive_trade_document(invoice_document_id, updated_at=self.timestamp_factory())
            except ValueError:
                pass
        return {
            "contract_name": "sheet_vitrina_v1_supplier_shipments",
            "status": "ok",
            "deleted": True,
            "shipment_id": shipment_id,
        }

    def rematch_shipment(self, shipment_id: str, payload: Mapping[str, Any] | None = None) -> dict[str, Any]:
        existing = self.runtime.load_supplier_shipment(shipment_id)
        if existing is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        detail_payload = _detail_payload(existing)
        overwrite_manual = bool((payload or {}).get("overwrite_manual"))
        detail_payload["lines"] = _apply_nomenclature_matches(
            detail_payload.get("lines") or [],
            self._active_nomenclature_items(),
            overwrite_manual=overwrite_manual,
        )
        shipment_date = _validate_iso_date(str(detail_payload.get("shipment_date") or ""))
        metadata, lines, warnings, errors, summary, match_status = _normalize_edit_payload(
            detail_payload,
            shipment_date=shipment_date,
            force_manual_override=False,
        )
        existing_header = dict(existing["header"])
        now = self.timestamp_factory()
        header = {
            **existing_header,
            "updated_at": now,
            "invoice_no": metadata.get("invoice_no") or "",
            "invoice_date": metadata.get("invoice_date") or "",
            "contract_no": metadata.get("contract_no") or "",
            "contract_date": metadata.get("contract_date") or "",
            "supplier_name": metadata.get("supplier_name") or "",
            "customer_name": metadata.get("customer_name") or "",
            "currency": metadata.get("currency") or "",
            "product_qty_total": summary["product_qty_total"],
            "product_amount_total": summary["product_amount_total"],
            "extras_amount_total": summary["extras_amount_total"],
            "invoice_amount_total": summary["invoice_amount_total"],
            "declared_invoice_total": summary.get("declared_invoice_total"),
            "match_status": match_status,
            "warnings": warnings,
            "errors": errors,
        }
        self.runtime.save_supplier_shipment(header=header, lines=lines)
        return self.get_shipment(shipment_id)

    def download_invoice(self, shipment_id: str) -> tuple[bytes, str, str]:
        detail = self.runtime.load_supplier_shipment(shipment_id)
        if detail is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        header = detail["header"]
        file_path: Path | None = None
        source_file_path = str(header.get("source_file_path") or "")
        if source_file_path:
            file_path = self._resolve_runtime_file(source_file_path)
        if (file_path is None or not file_path.exists() or not file_path.is_file()) and str(header.get("invoice_document_id") or ""):
            document = self.runtime.load_trade_document(str(header.get("invoice_document_id") or ""))
            if document is not None:
                file_path = self._resolve_runtime_file(str(document.get("file_path") or ""))
        if file_path is None or not file_path.exists() or not file_path.is_file():
            raise ValueError(f"supplier invoice file is missing for shipment: {shipment_id}")
        content_type = SUPPLIER_INVOICE_CONTENT_TYPE
        return file_path.read_bytes(), str(header.get("source_filename") or "supplier-invoice.xlsx"), content_type

    def download_shipment_contract(self, shipment_id: str) -> tuple[bytes, str, str]:
        shipment = self.get_shipment(shipment_id)
        contract_document_id = str(shipment.get("contract_document_id") or "")
        if not contract_document_id:
            raise ValueError(f"supplier shipment contract is not linked: {shipment_id}")
        return self.download_trade_document_file(contract_document_id)

    def list_trade_documents(self, *, include_archived: bool = False) -> dict[str, Any]:
        documents = [self._with_document_download_path(item) for item in self.runtime.list_trade_documents(include_archived=include_archived)]
        return {
            "contract_name": "sheet_vitrina_v1_trade_documents",
            "status": "ok",
            "documents": documents,
        }

    def create_trade_document_from_upload(
        self,
        *,
        document_type: str,
        file_bytes: bytes,
        uploaded_filename: str | None = None,
        uploaded_content_type: str | None = None,
        number: str | None = None,
        document_date: str | None = None,
        supplier_name: str | None = None,
        currency: str | None = None,
        amount_total: Any = None,
    ) -> dict[str, Any]:
        normalized_type = _normalize_trade_document_type(document_type)
        if not file_bytes:
            raise ValueError("trade document upload file is empty")
        filename = _safe_document_filename(uploaded_filename or f"{normalized_type}.xlsx", document_type=normalized_type)
        extension = Path(filename).suffix.lower()
        if extension not in TRADE_DOCUMENT_ALLOWED_EXTENSIONS:
            raise ValueError("trade document upload must be one of: .pdf, .jpg, .jpeg, .png, .xlsx")
        content_type = _document_content_type(filename, uploaded_content_type)
        sha256 = hashlib.sha256(file_bytes).hexdigest()
        duplicate = self.runtime.find_settings_trade_document_duplicate(
            document_type=normalized_type,
            file_sha256=sha256,
        )
        if duplicate is not None:
            duplicate = self._backfill_trade_document_record(
                duplicate,
                file_bytes=file_bytes,
                filename=filename,
                manual_number=number,
                manual_document_date=document_date,
                manual_supplier_name=supplier_name,
            )
            duplicate.pop("_backfill_stats", None)
            payload = self._with_document_download_path(duplicate)
            payload["deduplicated"] = True
            return {
                "contract_name": "sheet_vitrina_v1_trade_documents",
                "status": "duplicate_existing",
                "document": payload,
            }

        now = self.timestamp_factory()
        parsed_metadata: dict[str, Any] = {}
        warnings: list[str] = []
        errors: list[str] = []
        parser_version = ""
        if normalized_type == TRADE_DOCUMENT_TYPE_INVOICE and extension == ".xlsx":
            parsed_metadata, warnings, errors, parser_version = self._parse_invoice_document_metadata(
                file_bytes,
                filename=filename,
            )
        elif normalized_type == TRADE_DOCUMENT_TYPE_CONTRACT:
            parsed_metadata, warnings, errors, parser_version = self._parse_contract_document_metadata(
                file_bytes,
                filename=filename,
            )
        manual_number = str(number or "").strip()
        manual_date = _optional_iso_date(document_date)
        if normalized_type == TRADE_DOCUMENT_TYPE_CONTRACT:
            parsed_number = str(parsed_metadata.get("parsed_number") or "").strip()
            parsed_date = _optional_iso_date(parsed_metadata.get("parsed_document_date"))
            if manual_number and parsed_number and _compact_compare(manual_number) != _compact_compare(parsed_number):
                warnings.append(
                    f"contract parser found number {parsed_number!r}, manual number {manual_number!r} kept"
                )
            if manual_date and parsed_date and manual_date != parsed_date:
                warnings.append(
                    f"contract parser found date {parsed_date!r}, manual date {manual_date!r} kept"
                )
            metadata_number = manual_number or parsed_number
            metadata_date = manual_date or parsed_date
        else:
            metadata_number = manual_number or str(parsed_metadata.get("invoice_no") or "").strip()
            metadata_date = manual_date or _optional_iso_date(parsed_metadata.get("invoice_date"))
        metadata_supplier = _document_supplier_name(supplier_name, parsed_metadata.get("supplier_name"))
        metadata_currency = str(currency or "").strip().upper() or str(parsed_metadata.get("currency") or "").strip().upper()
        metadata_amount = _optional_number(amount_total)
        if metadata_amount is None:
            metadata_amount = _optional_number(parsed_metadata.get("declared_invoice_total"))
        if metadata_amount is None:
            metadata_amount = _optional_number(parsed_metadata.get("invoice_amount_total"))

        document_id = "tdoc_" + uuid4().hex
        file_path = self._write_trade_document_file(
            document_type=normalized_type,
            document_id=document_id,
            filename=filename,
            body=file_bytes,
        )
        document = self.runtime.save_trade_document(
            {
                "document_id": document_id,
                "document_type": normalized_type,
                "number": metadata_number,
                "document_date": metadata_date,
                "supplier_name": metadata_supplier,
                "currency": metadata_currency,
                "amount_total": metadata_amount,
                "source": TRADE_DOCUMENT_SOURCE_SETTINGS_UPLOAD,
                "source_shipment_id": "",
                "source_upload_id": "",
                "file_original_name": filename,
                "file_content_type": content_type,
                "file_sha256": sha256,
                "file_path": file_path,
                "parser_version": parser_version,
                "parsed_metadata": parsed_metadata,
                "warnings": warnings,
                "errors": errors,
                "status": TRADE_DOCUMENT_STATUS_ACTIVE,
                "created_at": now,
                "updated_at": now,
            }
        )
        return {
            "contract_name": "sheet_vitrina_v1_trade_documents",
            "status": "ok",
            "document": self._with_document_download_path(document),
        }

    def update_trade_document(self, document_id: str, payload: Mapping[str, Any]) -> dict[str, Any]:
        allowed_fields = {"number", "document_date", "supplier_name"}
        unsupported_fields = sorted(set(payload.keys()) - allowed_fields)
        if unsupported_fields:
            raise ValueError("unsupported trade document metadata fields: " + ", ".join(unsupported_fields))
        updates: dict[str, Any] = {}
        if "number" in payload:
            updates["number"] = str(payload.get("number") or "").strip()
        if "document_date" in payload:
            updates["document_date"] = _optional_trade_document_date(payload.get("document_date"))
        if "supplier_name" in payload:
            updates["supplier_name"] = _document_supplier_name(payload.get("supplier_name"))
        document = self.runtime.update_trade_document(document_id, updates, updated_at=self.timestamp_factory())
        return {
            "contract_name": "sheet_vitrina_v1_trade_documents",
            "status": "ok",
            "document": self._with_document_download_path(document),
        }

    def archive_trade_document(self, document_id: str) -> dict[str, Any]:
        existing = self.runtime.load_trade_document(document_id)
        if existing is None:
            raise ValueError(f"trade document not found: {document_id}")
        if (
            str(existing.get("document_type") or "") == TRADE_DOCUMENT_TYPE_CONTRACT
            and self.runtime.count_contract_document_links(document_id) > 0
        ):
            raise ValueError("contract document has linked invoice documents and cannot be archived")
        if str(existing.get("document_type") or "") == TRADE_DOCUMENT_TYPE_INVOICE:
            self.runtime.delete_invoice_contract_link(document_id)
        document = self.runtime.archive_trade_document(document_id, updated_at=self.timestamp_factory())
        return {
            "contract_name": "sheet_vitrina_v1_trade_documents",
            "status": "ok",
            "document": self._with_document_download_path(document),
        }

    def download_trade_document_file(self, document_id: str) -> tuple[bytes, str, str]:
        document = self.runtime.load_trade_document(document_id)
        if document is None or str(document.get("status") or "") != TRADE_DOCUMENT_STATUS_ACTIVE:
            raise ValueError(f"trade document not found: {document_id}")
        file_path = self._resolve_runtime_file(str(document.get("file_path") or ""))
        if not file_path.exists() or not file_path.is_file():
            raise ValueError(f"trade document file is missing: {document_id}")
        return (
            file_path.read_bytes(),
            str(document.get("file_original_name") or "document"),
            str(document.get("file_content_type") or "application/octet-stream"),
        )

    def link_invoice_to_contract(
        self,
        invoice_document_id: str,
        *,
        contract_document_id: str,
        linked_by: str = "",
        source: str = TRADE_DOCUMENT_LINK_SOURCE_OPERATOR,
    ) -> dict[str, Any]:
        invoice = self.runtime.load_trade_document(invoice_document_id)
        if invoice is None or str(invoice.get("document_type") or "") != TRADE_DOCUMENT_TYPE_INVOICE:
            raise ValueError(f"invoice document not found: {invoice_document_id}")
        if str(invoice.get("status") or "") != TRADE_DOCUMENT_STATUS_ACTIVE:
            raise ValueError(f"invoice document is not active: {invoice_document_id}")
        contract = self.runtime.load_trade_document(contract_document_id)
        if contract is None or str(contract.get("document_type") or "") != TRADE_DOCUMENT_TYPE_CONTRACT:
            raise ValueError(f"contract document not found: {contract_document_id}")
        if str(contract.get("status") or "") != TRADE_DOCUMENT_STATUS_ACTIVE:
            raise ValueError(f"contract document is not active: {contract_document_id}")
        now = self.timestamp_factory()
        existing = self.runtime.load_invoice_contract_link(invoice_document_id)
        link = self.runtime.save_invoice_contract_link(
            invoice_document_id=invoice_document_id,
            contract_document_id=contract_document_id,
            created_at=str((existing or {}).get("created_at") or now),
            updated_at=now,
            linked_by=linked_by,
            source=source,
        )
        return {
            "contract_name": "sheet_vitrina_v1_invoice_contract_links",
            "status": "ok",
            "link": link,
            "invoice": self._with_document_download_path(self.runtime.load_trade_document(invoice_document_id) or invoice),
            "contract": self._with_document_download_path(contract),
        }

    def unlink_invoice_contract(self, invoice_document_id: str) -> dict[str, Any]:
        invoice = self.runtime.load_trade_document(invoice_document_id)
        if invoice is None or str(invoice.get("document_type") or "") != TRADE_DOCUMENT_TYPE_INVOICE:
            raise ValueError(f"invoice document not found: {invoice_document_id}")
        deleted = self.runtime.delete_invoice_contract_link(invoice_document_id)
        return {
            "contract_name": "sheet_vitrina_v1_invoice_contract_links",
            "status": "ok",
            "deleted": deleted,
            "invoice_document_id": invoice_document_id,
            "invoice": self._with_document_download_path(self.runtime.load_trade_document(invoice_document_id) or invoice),
        }

    def link_shipment_contract(
        self,
        shipment_id: str,
        *,
        contract_document_id: str,
        linked_by: str = "",
        source: str = TRADE_DOCUMENT_LINK_SOURCE_OPERATOR,
    ) -> dict[str, Any]:
        shipment = self._ensure_shipment_invoice_document(shipment_id)
        invoice_document_id = str(shipment.get("invoice_document_id") or "")
        if not invoice_document_id:
            raise ValueError(f"supplier shipment invoice document is missing: {shipment_id}")
        result = self.link_invoice_to_contract(
            invoice_document_id,
            contract_document_id=contract_document_id,
            linked_by=linked_by,
            source=source,
        )
        result["shipment"] = self.get_shipment(shipment_id)
        return result

    def unlink_shipment_contract(self, shipment_id: str) -> dict[str, Any]:
        shipment = self._ensure_shipment_invoice_document(shipment_id)
        invoice_document_id = str(shipment.get("invoice_document_id") or "")
        if not invoice_document_id:
            raise ValueError(f"supplier shipment invoice document is missing: {shipment_id}")
        result = self.unlink_invoice_contract(invoice_document_id)
        result["shipment"] = self.get_shipment(shipment_id)
        return result

    def upload_shipment_contract(
        self,
        shipment_id: str,
        *,
        file_bytes: bytes,
        uploaded_filename: str | None = None,
        uploaded_content_type: str | None = None,
        number: str | None = None,
        document_date: str | None = None,
        supplier_name: str | None = None,
    ) -> dict[str, Any]:
        shipment = self._ensure_shipment_invoice_document(shipment_id)
        header_supplier = str(shipment.get("supplier_name") or DEFAULT_SUPPLIER_NAME)
        created = self.create_trade_document_from_upload(
            document_type=TRADE_DOCUMENT_TYPE_CONTRACT,
            file_bytes=file_bytes,
            uploaded_filename=uploaded_filename,
            uploaded_content_type=uploaded_content_type,
            number=number or shipment.get("contract_no") or "",
            document_date=document_date or shipment.get("contract_date") or "",
            supplier_name=supplier_name or header_supplier,
        )
        contract_document_id = str((created.get("document") or {}).get("document_id") or "")
        if not contract_document_id:
            raise ValueError("uploaded contract document was not saved")
        linked = self.link_shipment_contract(
            shipment_id,
            contract_document_id=contract_document_id,
            linked_by="operator",
            source=TRADE_DOCUMENT_LINK_SOURCE_OPERATOR,
        )
        linked["document_upload"] = created
        return linked

    def backfill_trade_document_metadata(self, *, include_archived: bool = False) -> dict[str, Any]:
        rows = self.runtime.list_trade_documents(include_archived=include_archived)
        updated_count = 0
        supplier_backfilled_count = 0
        contract_parse_attempt_count = 0
        contract_metadata_backfilled_count = 0
        parser_warning_count = 0
        missing_file_count = 0
        unchanged_count = 0
        for document in rows:
            before = dict(document)
            backfilled = self._backfill_trade_document_record(before)
            stats = dict(backfilled.pop("_backfill_stats", {}) or {})
            if backfilled != before:
                updated_count += 1
                if not str(before.get("supplier_name") or "").strip() and str(backfilled.get("supplier_name") or "").strip():
                    supplier_backfilled_count += 1
                if (
                    str(before.get("document_type") or "") == TRADE_DOCUMENT_TYPE_CONTRACT
                    and (
                        (not str(before.get("number") or "").strip() and str(backfilled.get("number") or "").strip())
                        or (not str(before.get("document_date") or "").strip() and str(backfilled.get("document_date") or "").strip())
                    )
                ):
                    contract_metadata_backfilled_count += 1
            else:
                unchanged_count += 1
            if stats.get("contract_parse_attempted"):
                contract_parse_attempt_count += 1
            if stats.get("parser_warning"):
                parser_warning_count += 1
            if stats.get("missing_file"):
                missing_file_count += 1
        return {
            "contract_name": "sheet_vitrina_v1_trade_documents_metadata_backfill",
            "status": "ok",
            "scanned_documents": len(rows),
            "updated_documents": updated_count,
            "supplier_backfilled_documents": supplier_backfilled_count,
            "contract_parse_attempted_documents": contract_parse_attempt_count,
            "contract_metadata_backfilled_documents": contract_metadata_backfilled_count,
            "parser_warning_documents": parser_warning_count,
            "missing_file_documents": missing_file_count,
            "unchanged_documents": unchanged_count,
            "default_supplier_name": DEFAULT_SUPPLIER_NAME,
        }

    def find_contract_candidates(self, number: str, document_date: str = "") -> list[dict[str, Any]]:
        return [
            self._with_document_download_path(item)
            for item in self.runtime.find_contract_document_candidates(
                number=str(number or "").strip(),
                document_date=_optional_iso_date(document_date),
            )
        ]

    def migrate_existing_supplier_shipments_into_trade_documents(self) -> dict[str, Any]:
        rows = self.runtime.list_supplier_shipments()
        created_count = 0
        linked_count = 0
        skipped_count = 0
        for row in rows:
            shipment_id = str(row.get("shipment_id") or "")
            if not shipment_id:
                skipped_count += 1
                continue
            detail = self.runtime.load_supplier_shipment(shipment_id)
            if detail is None:
                skipped_count += 1
                continue
            header = dict(detail.get("header") or {})
            if str(header.get("invoice_document_id") or "").strip():
                skipped_count += 1
                continue
            source_file_path = str(header.get("source_file_path") or "").strip()
            if not source_file_path:
                skipped_count += 1
                continue
            file_sha256 = str(header.get("source_file_sha256") or "").strip() or self._sha256_for_existing_runtime_file(
                source_file_path,
                fallback_seed=f"{shipment_id}:{source_file_path}",
            )
            existing = self.runtime.find_trade_document_by_source_file(
                document_type=TRADE_DOCUMENT_TYPE_INVOICE,
                file_sha256=file_sha256,
                source_shipment_id=shipment_id,
            )
            now = self.timestamp_factory()
            if existing is None:
                existing = self.runtime.save_trade_document(
                    {
                        "document_id": "tdoc_" + uuid4().hex,
                        "document_type": TRADE_DOCUMENT_TYPE_INVOICE,
                        "number": header.get("invoice_no") or "",
                        "document_date": header.get("invoice_date") or "",
                        "supplier_name": header.get("supplier_name") or DEFAULT_SUPPLIER_NAME,
                        "currency": header.get("currency") or "",
                        "amount_total": header.get("invoice_amount_total"),
                        "source": TRADE_DOCUMENT_SOURCE_MIGRATION_EXISTING_SUPPLIER_INVOICE,
                        "source_shipment_id": shipment_id,
                        "source_upload_id": "",
                        "file_original_name": header.get("source_filename") or "supplier-invoice.xlsx",
                        "file_content_type": SUPPLIER_INVOICE_CONTENT_TYPE,
                        "file_sha256": file_sha256,
                        "file_path": source_file_path,
                        "parser_version": header.get("parser_version") or "",
                        "parsed_metadata": {
                            "invoice_no": header.get("invoice_no") or "",
                            "invoice_date": header.get("invoice_date") or "",
                            "contract_no": header.get("contract_no") or "",
                            "contract_date": header.get("contract_date") or "",
                            "supplier_name": header.get("supplier_name") or DEFAULT_SUPPLIER_NAME,
                            "currency": header.get("currency") or "",
                            "invoice_amount_total": header.get("invoice_amount_total"),
                            "declared_invoice_total": header.get("declared_invoice_total"),
                        },
                        "warnings": header.get("warnings") or [],
                        "errors": header.get("errors") or [],
                        "status": TRADE_DOCUMENT_STATUS_ACTIVE,
                        "created_at": now,
                        "updated_at": now,
                    }
                )
                created_count += 1
            updated = self.runtime.set_supplier_shipment_invoice_document_id(
                shipment_id=shipment_id,
                invoice_document_id=str(existing.get("document_id") or ""),
                updated_at=now,
            )
            if updated:
                contract_linked = self._autolink_invoice_contract_from_metadata(
                    invoice_document_id=str(existing.get("document_id") or ""),
                    contract_no=str(header.get("contract_no") or ""),
                    contract_date=str(header.get("contract_date") or ""),
                    linked_by="system",
                    source=TRADE_DOCUMENT_LINK_SOURCE_MIGRATION,
                )
                if contract_linked:
                    linked_count += 1
            else:
                skipped_count += 1
        return {
            "contract_name": "sheet_vitrina_v1_trade_documents_backfill",
            "status": "ok",
            "found_shipments": len(rows),
            "created_documents": created_count,
            "linked_contracts": linked_count,
            "skipped_shipments": skipped_count,
        }

    def _backfill_trade_document_record(
        self,
        document: Mapping[str, Any],
        *,
        file_bytes: bytes | None = None,
        filename: str | None = None,
        manual_number: Any = None,
        manual_document_date: Any = None,
        manual_supplier_name: Any = None,
    ) -> dict[str, Any]:
        document_id = str(document.get("document_id") or "").strip()
        if not document_id:
            return dict(document)
        existing = dict(document)
        updates: dict[str, Any] = {}
        stats: dict[str, Any] = {}
        existing_supplier = str(existing.get("supplier_name") or "").strip()
        if not existing_supplier:
            updates["supplier_name"] = _document_supplier_name(manual_supplier_name)

        document_type = str(existing.get("document_type") or "").strip()
        if (
            document_type == TRADE_DOCUMENT_TYPE_CONTRACT
            and str(existing.get("number") or "").strip()
            and str(existing.get("document_date") or "").strip()
        ):
            existing_warnings = _string_list(existing.get("warnings"))
            cleaned_warnings = _remove_stale_contract_parser_warnings(existing_warnings)
            if cleaned_warnings != existing_warnings:
                updates["warnings"] = cleaned_warnings
        needs_contract_parse = (
            document_type == TRADE_DOCUMENT_TYPE_CONTRACT
            and (not str(existing.get("number") or "").strip() or not str(existing.get("document_date") or "").strip())
        )
        if needs_contract_parse:
            payload_bytes = file_bytes
            payload_filename = filename or str(existing.get("file_original_name") or "")
            if payload_bytes is None:
                try:
                    payload_bytes = self._resolve_runtime_file(str(existing.get("file_path") or "")).read_bytes()
                except (OSError, ValueError):
                    payload_bytes = None
                    stats["missing_file"] = True
            if payload_bytes is not None:
                stats["contract_parse_attempted"] = True
                parsed_metadata, warnings, errors, parser_version = self._parse_contract_document_metadata(
                    payload_bytes,
                    filename=payload_filename,
                )
                manual_number_value = str(manual_number or "").strip()
                manual_date_value = _optional_iso_date(manual_document_date)
                parsed_number = str(parsed_metadata.get("parsed_number") or "").strip()
                parsed_date = _optional_iso_date(parsed_metadata.get("parsed_document_date"))
                if not str(existing.get("number") or "").strip() and (manual_number_value or parsed_number):
                    updates["number"] = manual_number_value or parsed_number
                if not str(existing.get("document_date") or "").strip() and (manual_date_value or parsed_date):
                    updates["document_date"] = manual_date_value or parsed_date
                if not existing_supplier:
                    updates["supplier_name"] = _document_supplier_name(
                        manual_supplier_name,
                        parsed_metadata.get("supplier_name"),
                    )
                existing_metadata = existing.get("parsed_metadata")
                existing_metadata = dict(existing_metadata) if isinstance(existing_metadata, Mapping) else {}
                merged_metadata = {**existing_metadata, **dict(parsed_metadata)}
                if merged_metadata != existing_metadata:
                    updates["parsed_metadata"] = merged_metadata
                existing_warnings = _string_list(existing.get("warnings"))
                merged_warnings = _merge_contract_parser_warnings(
                    existing_warnings,
                    warnings,
                    parsed_number=parsed_number,
                    parsed_date=parsed_date,
                )
                if merged_warnings != existing_warnings:
                    updates["warnings"] = merged_warnings
                existing_errors = _string_list(existing.get("errors"))
                merged_errors = _merge_string_lists(existing_errors, errors)
                if merged_errors != existing_errors:
                    updates["errors"] = merged_errors
                if parser_version and str(existing.get("parser_version") or "") != parser_version:
                    updates["parser_version"] = parser_version
                if warnings:
                    stats["parser_warning"] = True

        if not updates:
            result = dict(existing)
            result["_backfill_stats"] = stats
            return result
        updated = self.runtime.update_trade_document(document_id, updates, updated_at=self.timestamp_factory())
        result = dict(updated)
        result["_backfill_stats"] = stats
        return result

    def list_nomenclature(self, *, visibility: str = "visible") -> dict[str, Any]:
        self._ensure_nomenclature_ready()
        visibility_mode = _normalize_visibility_mode(visibility)
        all_items = self.runtime.list_nomenclature_items()
        if visibility_mode == "visible":
            items = [item for item in all_items if not bool(item.get("is_hidden"))]
        elif visibility_mode == "hidden":
            items = [item for item in all_items if bool(item.get("is_hidden"))]
        else:
            items = all_items
        return {
            "contract_name": "sheet_vitrina_v1_nomenclature",
            "status": "ok",
            "visibility": visibility_mode,
            "summary": _nomenclature_barcode_summary(all_items),
            "sku_groups": self.list_sku_groups(include_inactive=True)["groups"],
            "items": items,
        }

    def export_nomenclature_xlsx(self) -> tuple[bytes, str, str]:
        self._ensure_nomenclature_ready()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Номенклатура"
        worksheet.append(NOMENCLATURE_XLSX_HEADERS)
        for item in self.runtime.list_nomenclature_items():
            worksheet.append(
                [
                    str(item.get("item_id") or ""),
                    "да" if bool(item.get("is_active")) else "нет",
                    "да" if bool(item.get("is_hidden")) else "нет",
                    item.get("nm_id") if item.get("nm_id") is not None else "",
                    str(item.get("barcode") or ""),
                    ", ".join(str(barcode) for barcode in item.get("barcodes") or [] if str(barcode or "").strip()),
                    str(item.get("barcode_source") or ""),
                    str(item.get("barcode_status") or ""),
                    str(item.get("vendor_code") or ""),
                    str(item.get("wb_title") or ""),
                    str(item.get("wb_subject_name") or ""),
                    str(item.get("wb_updated_at") or ""),
                    str(item.get("wb_sync_status") or ""),
                    str(item.get("nomenclature_name") or ""),
                    NOMENCLATURE_PRODUCT_TYPE_LABELS.get(str(item.get("product_type") or ""), str(item.get("product_type") or "")),
                    str(item.get("match_key") or ""),
                    item.get("purchase_price_yuan") if item.get("purchase_price_yuan") is not None else "",
                    str(item.get("compatible_models_text") or ""),
                    ", ".join(str(key) for key in item.get("compatible_model_keys") or [] if str(key or "").strip()),
                    str(item.get("updated_at") or ""),
                ]
            )
        worksheet.freeze_panes = "A2"
        for index, width in enumerate([24, 12, 12, 14, 22, 34, 18, 18, 34, 36, 22, 22, 20, 34, 22, 28, 18, 34, 34, 24], start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width
        output = BytesIO()
        workbook.save(output)
        return output.getvalue(), NOMENCLATURE_XLSX_FILENAME, NOMENCLATURE_XLSX_CONTENT_TYPE

    def import_nomenclature_xlsx(
        self,
        workbook_bytes: bytes,
        *,
        uploaded_filename: str | None = None,
        uploaded_content_type: str | None = None,
        dry_run: bool = False,
    ) -> dict[str, Any]:
        del uploaded_content_type
        filename = _safe_filename(uploaded_filename or NOMENCLATURE_XLSX_FILENAME)
        if not filename.lower().endswith(".xlsx"):
            raise ValueError("nomenclature import upload must be an .xlsx file")
        try:
            workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl owns exact exception types
            raise ValueError("nomenclature import upload must be a readable .xlsx file") from exc

        self._ensure_nomenclature_ready()
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_keys = _nomenclature_import_header_keys(header_row or [])
        if not any(header_keys):
            raise ValueError("nomenclature import workbook must contain a header row")

        existing_items = self.runtime.list_nomenclature_items()
        existing_by_id = {str(item.get("item_id") or ""): dict(item) for item in existing_items}
        active_by_match_key: dict[str, list[dict[str, Any]]] = {}
        for item in existing_items:
            if bool(item.get("is_active")) and str(item.get("match_key") or "").strip():
                active_by_match_key.setdefault(str(item.get("match_key") or "").strip(), []).append(dict(item))

        now = self.timestamp_factory()
        operations: list[dict[str, Any]] = []
        skipped_count = 0
        errors: list[dict[str, Any]] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = {
                key: row[index]
                for index, key in enumerate(header_keys)
                if key and index < len(row)
            }
            if _nomenclature_import_row_empty(row_values):
                skipped_count += 1
                continue
            try:
                operation = _normalize_nomenclature_import_row(
                    row_values,
                    row_number=row_number,
                    existing_by_id=existing_by_id,
                    active_by_match_key=active_by_match_key,
                    now=now,
                )
            except ValueError as exc:
                errors.append({"row": row_number, "message": str(exc)})
                continue
            if operation is None:
                skipped_count += 1
                continue
            operations.append(operation)

        duplicate_errors = _nomenclature_import_duplicate_errors(existing_items, operations)
        errors.extend(duplicate_errors)
        if errors:
            return _nomenclature_import_result(
                status="error",
                dry_run=dry_run,
                operations=operations,
                skipped_count=skipped_count,
                errors=errors,
                items=[],
            )

        if dry_run:
            return _nomenclature_import_result(
                status="ok",
                dry_run=True,
                operations=operations,
                skipped_count=skipped_count,
                errors=[],
                items=[],
            )

        saved_items = self.runtime.save_nomenclature_items_atomic([operation["item"] for operation in operations])
        return _nomenclature_import_result(
            status="ok",
            dry_run=False,
            operations=operations,
            skipped_count=skipped_count,
            errors=[],
            items=saved_items,
        )

    def list_sku_groups(self, *, include_inactive: bool = True) -> dict[str, Any]:
        self._ensure_sku_groups_ready()
        groups = self.runtime.list_sku_groups(include_inactive=include_inactive)
        return {
            "contract_name": "sheet_vitrina_v1_sku_groups",
            "status": "ok",
            "groups": groups,
        }

    def create_sku_group(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        self._ensure_sku_groups_ready()
        now = self.timestamp_factory()
        group = _normalize_sku_group_payload(payload, created_at=now, updated_at=now)
        return {
            "contract_name": "sheet_vitrina_v1_sku_groups",
            "status": "ok",
            "group": self.runtime.save_sku_group(group),
            "groups": self.runtime.list_sku_groups(include_inactive=True),
        }

    def update_sku_group(self, group_key: str, payload: Mapping[str, Any]) -> dict[str, Any]:
        self._ensure_sku_groups_ready()
        existing = self.runtime.load_sku_group(group_key)
        if existing is None:
            raise ValueError(f"sku group not found: {group_key}")
        now = self.timestamp_factory()
        normalized = _normalize_sku_group_payload(
            {**existing, **dict(payload), "group_key": str(existing.get("group_key") or group_key)},
            created_at=str(existing.get("created_at") or now),
            updated_at=now,
        )
        if not bool(normalized.get("is_active")) and bool(existing.get("is_active")):
            usage_count = self.runtime.sku_group_active_item_count(str(existing.get("group_key") or group_key))
            if usage_count:
                raise ValueError(f"sku group is used by {usage_count} active nomenclature rows")
        return {
            "contract_name": "sheet_vitrina_v1_sku_groups",
            "status": "ok",
            "group": self.runtime.save_sku_group(normalized),
            "groups": self.runtime.list_sku_groups(include_inactive=True),
        }

    def deactivate_sku_group(self, group_key: str) -> dict[str, Any]:
        return self.update_sku_group(group_key, {"is_active": False})

    def create_nomenclature_item(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        now = self.timestamp_factory()
        prepared_payload = _prepare_nomenclature_barcode_payload(
            existing=None,
            payload=payload,
            updated_at=now,
        )
        if bool(prepared_payload.get("is_hidden")) and not str(prepared_payload.get("hidden_at") or "").strip():
            prepared_payload["hidden_at"] = now
            prepared_payload["hidden_reason"] = str(prepared_payload.get("hidden_reason") or "operator_hidden")
        item = _normalize_nomenclature_payload(
            prepared_payload,
            item_id="nom_" + uuid4().hex,
            created_at=now,
            updated_at=now,
        )
        item, barcode_sync = self._sync_nomenclature_barcode_item(
            item,
            reason="auto_save",
            allow_existing_non_manual=False,
        )
        self._validate_nomenclature_unique(item)
        return {
            "contract_name": "sheet_vitrina_v1_nomenclature",
            "status": "ok",
            "item": self.runtime.save_nomenclature_item(item),
            "barcode_sync": barcode_sync,
        }

    def update_nomenclature_item(self, item_id: str, payload: Mapping[str, Any]) -> dict[str, Any]:
        existing = self.runtime.load_nomenclature_item(item_id)
        if existing is None:
            raise ValueError(f"nomenclature item not found: {item_id}")
        now = self.timestamp_factory()
        prepared_payload = _prepare_nomenclature_barcode_payload(
            existing=existing,
            payload=payload,
            updated_at=now,
        )
        if "is_hidden" in payload:
            requested_hidden = bool(payload.get("is_hidden"))
            if requested_hidden and not bool(existing.get("is_hidden")):
                prepared_payload["hidden_at"] = now
                prepared_payload["hidden_reason"] = str(payload.get("hidden_reason") or "operator_hidden")
            elif not requested_hidden:
                prepared_payload["hidden_at"] = ""
                prepared_payload["hidden_reason"] = ""
        item = _normalize_nomenclature_payload(
            {**existing, **prepared_payload},
            item_id=item_id,
            created_at=str(existing.get("created_at") or now),
            updated_at=now,
        )
        item, barcode_sync = self._sync_nomenclature_barcode_item(
            item,
            reason="auto_save",
            allow_existing_non_manual=False,
        )
        self._validate_nomenclature_unique(item)
        return {
            "contract_name": "sheet_vitrina_v1_nomenclature",
            "status": "ok",
            "item": self.runtime.save_nomenclature_item(item),
            "barcode_sync": barcode_sync,
        }

    def deactivate_nomenclature_item(self, item_id: str) -> dict[str, Any]:
        item = self.runtime.delete_nomenclature_item(item_id, updated_at=self.timestamp_factory())
        return {
            "contract_name": "sheet_vitrina_v1_nomenclature",
            "status": "ok",
            "item": item,
        }

    def sync_nomenclature_item_barcode(self, item_id: str) -> dict[str, Any]:
        existing = self.runtime.load_nomenclature_item(item_id)
        if existing is None:
            raise ValueError(f"nomenclature item not found: {item_id}")
        item, barcode_sync = self._sync_nomenclature_barcode_item(
            existing,
            reason="manual_row_sync",
            allow_existing_non_manual=True,
        )
        saved = self.runtime.save_nomenclature_item(item) if barcode_sync.get("save_item", False) else existing
        return {
            "contract_name": "sheet_vitrina_v1_nomenclature_barcode_sync",
            "status": "ok",
            "item": saved,
            "barcode_sync": {key: value for key, value in barcode_sync.items() if key != "save_item"},
        }

    def sync_nomenclature_barcodes(self, payload: Mapping[str, Any] | None = None) -> dict[str, Any]:
        return self.sync_nomenclature_with_wb(payload)

    def sync_nomenclature_with_wb(self, payload: Mapping[str, Any] | None = None) -> dict[str, Any]:
        payload = payload or {}
        self._ensure_nomenclature_ready()
        limit = _bounded_int(payload.get("limit"), default=100, minimum=1, maximum=100)
        max_pages = _bounded_int(payload.get("max_pages"), default=500, minimum=1, maximum=5000)
        now = self.timestamp_factory()
        try:
            cards = self.barcode_source.fetch_cards(limit=limit, max_pages=max_pages)
        except OfficialApiRuntimeError as exc:
            return _nomenclature_wb_sync_error_result(
                status="token_missing",
                error=exc,
                summary=_nomenclature_barcode_summary(self.runtime.list_nomenclature_items()),
                limit=limit,
                max_pages=max_pages,
            )
        except (WbContentHttpStatusError, WbContentTransportError, RuntimeError) as exc:
            return _nomenclature_wb_sync_error_result(
                status="sync_error",
                error=exc,
                summary=_nomenclature_barcode_summary(self.runtime.list_nomenclature_items()),
                limit=limit,
                max_pages=max_pages,
            )

        groups = self.runtime.list_sku_groups(include_inactive=False)
        rows = self.runtime.list_nomenclature_items()
        matcher = _NomenclatureWbCardMatcher(rows)
        counts = {
            "cards_processed": 0,
            "matched_nm_id": 0,
            "matched_barcode": 0,
            "matched_vendor_code": 0,
            "created": 0,
            "created_needs_review": 0,
            "updated": 0,
            "manual_barcode_preserved": 0,
            "hidden_matched": 0,
            "skipped_invalid": 0,
        }
        results: list[dict[str, Any]] = []
        for raw_card in cards:
            card = _normalize_wb_card_for_sync(raw_card)
            if not _wb_card_has_identity(card):
                counts["skipped_invalid"] += 1
                continue
            counts["cards_processed"] += 1
            matched, match_type = matcher.match(card)
            if matched is not None:
                if match_type == "nm_id":
                    counts["matched_nm_id"] += 1
                elif match_type == "barcode":
                    counts["matched_barcode"] += 1
                elif match_type == "vendor_code":
                    counts["matched_vendor_code"] += 1
                if bool(matched.get("is_hidden")):
                    counts["hidden_matched"] += 1
                updated = _apply_wb_card_to_existing_nomenclature(
                    matched,
                    card=card,
                    match_type=match_type,
                    synced_at=now,
                )
                if updated.get("_manual_barcode_preserved"):
                    counts["manual_barcode_preserved"] += 1
                updated.pop("_manual_barcode_preserved", None)
                saved = self.runtime.save_nomenclature_item(updated)
                matcher.replace(saved)
                results.append(saved)
                counts["updated"] += 1
                continue
            created = _new_nomenclature_item_from_wb_card(
                card,
                groups=groups,
                created_at=now,
                updated_at=now,
            )
            saved = self.runtime.save_nomenclature_item(created)
            matcher.replace(saved)
            results.append(saved)
            counts["created"] += 1
            if str(saved.get("wb_sync_status") or "") == "needs_review":
                counts["created_needs_review"] += 1
        items = self.runtime.list_nomenclature_items()
        return {
            "contract_name": "sheet_vitrina_v1_nomenclature_wb_sync",
            "status": "ok",
            "limit": limit,
            "max_pages": max_pages,
            **counts,
            "items": results,
            "summary": _nomenclature_barcode_summary(items),
        }

    def _sync_nomenclature_barcode_item(
        self,
        item: Mapping[str, Any],
        *,
        reason: str,
        allow_existing_non_manual: bool,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        normalized_item = dict(item)
        if (
            str(normalized_item.get("barcode_source") or "") == NOMENCLATURE_BARCODE_SOURCE_MANUAL
            and str(normalized_item.get("barcode") or "").strip()
        ):
            return normalized_item, {"status": "skipped_manual", "reason": reason, "save_item": False}
        if str(normalized_item.get("barcode") or "").strip() and not allow_existing_non_manual:
            return normalized_item, {"status": "skipped_existing", "reason": reason, "save_item": False}
        nm_id = _optional_int(normalized_item.get("nm_id"))
        now = self.timestamp_factory()
        if nm_id is None:
            normalized_item.update(
                {
                    "barcode": "",
                    "barcodes": [],
                    "barcode_source": NOMENCLATURE_BARCODE_SOURCE_MISSING,
                    "barcode_status": NOMENCLATURE_BARCODE_STATUS_MISSING,
                    "barcode_evidence": {"reason": "missing_nm_id", "sync_reason": reason},
                    "barcode_updated_at": now,
                }
            )
            return normalized_item, {"status": "missing_nm_id", "reason": reason, "save_item": True}
        try:
            resolutions = self.barcode_source.fetch_barcodes_by_nm_ids([nm_id])
            resolution = resolutions.get(nm_id) if isinstance(resolutions, Mapping) else None
            barcodes = _barcode_resolution_barcodes(resolution)
            evidence = _barcode_resolution_evidence(resolution, nm_id=nm_id, sync_reason=reason)
        except OfficialApiRuntimeError as exc:
            return _nomenclature_barcode_sync_error_item(
                normalized_item,
                status=NOMENCLATURE_BARCODE_STATUS_TOKEN_MISSING,
                error=exc,
                updated_at=now,
                sync_reason=reason,
            )
        except (WbContentHttpStatusError, WbContentTransportError, RuntimeError) as exc:
            return _nomenclature_barcode_sync_error_item(
                normalized_item,
                status=NOMENCLATURE_BARCODE_STATUS_SYNC_ERROR,
                error=exc,
                updated_at=now,
                sync_reason=reason,
            )

        if not barcodes:
            normalized_item.update(
                {
                    "barcode": "",
                    "barcodes": [],
                    "barcode_source": NOMENCLATURE_BARCODE_SOURCE_MISSING,
                    "barcode_status": NOMENCLATURE_BARCODE_STATUS_MISSING,
                    "barcode_synced_at": now,
                    "barcode_updated_at": now,
                    "barcode_evidence": {**evidence, "result": "not_found"},
                }
            )
            return normalized_item, {"status": "not_found", "reason": reason, "save_item": True}

        deterministic_barcodes = sorted(_normalize_barcode_list(barcodes))
        primary_barcode = deterministic_barcodes[0]
        status = (
            NOMENCLATURE_BARCODE_STATUS_MULTIPLE
            if len(deterministic_barcodes) > 1
            else NOMENCLATURE_BARCODE_STATUS_READY
        )
        normalized_item.update(
            {
                "barcode": primary_barcode,
                "barcodes": deterministic_barcodes,
                "barcode_source": NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT,
                "barcode_status": status,
                "barcode_synced_at": now,
                "barcode_updated_at": now,
                "barcode_evidence": {
                    **evidence,
                    "result": "resolved",
                    "selected_primary": primary_barcode,
                    "barcode_count": len(deterministic_barcodes),
                },
            }
        )
        return normalized_item, {"status": status, "reason": reason, "save_item": True}

    def _copy_upload_to_shipment_file(self, *, upload_path: str, shipment_id: str, filename: str) -> str:
        source_path = self._resolve_runtime_file(upload_path)
        if not source_path.exists() or not source_path.is_file():
            raise ValueError("staged supplier invoice upload file is missing")
        safe_filename = _safe_filename(filename)
        target_dir = self.runtime.runtime_dir / "supplier_invoices" / "files" / shipment_id
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / safe_filename
        shutil.copy2(source_path, target_path)
        return _relative_to_runtime(self.runtime.runtime_dir, target_path)

    def _write_runtime_file(self, *, root_kind: str, entity_id: str, filename: str, body: bytes) -> str:
        safe_filename = _safe_filename(filename)
        target_dir = self.runtime.runtime_dir / "supplier_invoices" / root_kind / entity_id
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / safe_filename
        target_path.write_bytes(body)
        return _relative_to_runtime(self.runtime.runtime_dir, target_path)

    def _resolve_runtime_file(self, relative_path: str) -> Path:
        normalized = str(relative_path or "").strip()
        if not normalized:
            raise ValueError("runtime file path is empty")
        root = self.runtime.runtime_dir.resolve()
        path = (root / normalized).resolve()
        if root != path and root not in path.parents:
            raise ValueError("runtime file path escapes runtime dir")
        return path

    def _write_trade_document_file(self, *, document_type: str, document_id: str, filename: str, body: bytes) -> str:
        safe_filename = _safe_document_filename(filename, document_type=document_type)
        target_dir = self.runtime.runtime_dir / "trade_documents" / "files" / document_type / document_id
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / safe_filename
        target_path.write_bytes(body)
        return _relative_to_runtime(self.runtime.runtime_dir, target_path)

    def _parse_invoice_document_metadata(self, workbook_bytes: bytes, *, filename: str) -> tuple[dict[str, Any], list[str], list[str], str]:
        try:
            parsed_payload = parse_supplier_invoice_xlsx(
                workbook_bytes,
                filename=filename,
                aliases=self._active_nomenclature_aliases(),
            )
        except Exception as exc:
            return {}, [], [f"supplier invoice parser skipped: {exc}"], ""
        metadata = dict(parsed_payload.get("metadata") or {})
        summary = dict(parsed_payload.get("summary") or {})
        if "invoice_amount_total" not in metadata and summary.get("invoice_amount_total") is not None:
            metadata["invoice_amount_total"] = summary.get("invoice_amount_total")
        if "declared_invoice_total" not in metadata and summary.get("declared_invoice_total") is not None:
            metadata["declared_invoice_total"] = summary.get("declared_invoice_total")
        return (
            metadata,
            _string_list(parsed_payload.get("warnings")),
            _string_list(parsed_payload.get("errors")),
            str(parsed_payload.get("parser_version") or SUPPLIER_INVOICE_PARSER_VERSION),
        )

    def _parse_contract_document_metadata(self, file_bytes: bytes, *, filename: str) -> tuple[dict[str, Any], list[str], list[str], str]:
        extension = Path(str(filename or "")).suffix.lower()
        warnings: list[str] = []
        errors: list[str] = []
        lines: list[str] = []
        extraction_method = ""
        extraction_diagnostics: dict[str, Any] = {}
        try:
            if extension == ".xlsx":
                lines = _extract_contract_xlsx_lines(file_bytes)
                extraction_method = "xlsx_first_rows"
            elif extension == ".pdf":
                lines, extract_warnings, extraction_method, extraction_diagnostics = _extract_contract_pdf_lines(file_bytes)
                warnings.extend(extract_warnings)
            elif extension in {".jpg", ".jpeg", ".png"}:
                lines, extract_warnings, extraction_method, extraction_diagnostics = _extract_contract_image_lines(
                    file_bytes,
                    suffix=extension,
                )
                warnings.extend(extract_warnings)
            else:
                warnings.append(f"contract parser skipped unsupported extension: {extension or 'unknown'}")
        except Exception as exc:
            return (
                {"extraction_method": extraction_method or "contract_parser_error"},
                warnings,
                [f"contract parser failed: {exc}"],
                TRADE_DOCUMENT_CONTRACT_PARSER_VERSION,
            )

        lines = [_normalize_text_line(line) for line in lines]
        lines = [line for line in lines if line]
        top_text = "\n".join(lines[:40])[:4000]
        first_line = lines[0] if lines else ""
        parsed_number = _extract_contract_number(first_line, lines[:8])
        parsed_document_date = _extract_contract_document_date(top_text)
        if not lines and extension in {".pdf", ".jpg", ".jpeg", ".png"} and not warnings:
            warnings.append("contract parser found no readable text")
        extraction_diagnostics["number_found"] = bool(parsed_number)
        extraction_diagnostics["date_found"] = bool(parsed_document_date)
        metadata = {
            "parsed_number": parsed_number,
            "parsed_document_date": parsed_document_date,
            "first_non_empty_line": first_line,
            "extraction_method": extraction_method or "unknown",
            "text_line_count": len(lines),
            "diagnostics": extraction_diagnostics,
        }
        if not parsed_number and first_line:
            warnings.append("contract parser OCR/text exists but number pattern was not found")
        if not parsed_document_date and top_text:
            warnings.append("contract parser OCR/text exists but document date pattern was not found")
        return metadata, warnings, errors, TRADE_DOCUMENT_CONTRACT_PARSER_VERSION

    def _create_or_load_supplier_invoice_document(
        self,
        *,
        shipment_id: str,
        upload_id: str,
        source_filename: str,
        content_type: str,
        source_file_sha256: str,
        source_file_path: str,
        parser_version: str,
        metadata: Mapping[str, Any],
        warnings: list[str],
        errors: list[str],
        parsed_payload: Mapping[str, Any],
        created_at: str,
    ) -> dict[str, Any]:
        file_sha256 = str(source_file_sha256 or "").strip() or self._sha256_for_existing_runtime_file(
            source_file_path,
            fallback_seed=f"{shipment_id}:{source_file_path}",
        )
        existing = self.runtime.find_trade_document_by_source_file(
            document_type=TRADE_DOCUMENT_TYPE_INVOICE,
            file_sha256=file_sha256,
            source_shipment_id=shipment_id,
        )
        if existing is not None:
            return existing
        parsed_metadata = {
            "invoice_no": metadata.get("invoice_no") or "",
            "invoice_date": metadata.get("invoice_date") or "",
            "contract_no": metadata.get("contract_no") or "",
            "contract_date": metadata.get("contract_date") or "",
            "supplier_name": metadata.get("supplier_name") or DEFAULT_SUPPLIER_NAME,
            "currency": metadata.get("currency") or "",
            "invoice_amount_total": parsed_payload.get("summary", {}).get("invoice_amount_total")
            if isinstance(parsed_payload.get("summary"), Mapping)
            else None,
            "declared_invoice_total": metadata.get("declared_invoice_total"),
        }
        return self.runtime.save_trade_document(
            {
                "document_id": "tdoc_" + uuid4().hex,
                "document_type": TRADE_DOCUMENT_TYPE_INVOICE,
                "number": metadata.get("invoice_no") or "",
                "document_date": metadata.get("invoice_date") or "",
                "supplier_name": metadata.get("supplier_name") or DEFAULT_SUPPLIER_NAME,
                "currency": metadata.get("currency") or "",
                "amount_total": parsed_metadata.get("invoice_amount_total") or metadata.get("declared_invoice_total"),
                "source": TRADE_DOCUMENT_SOURCE_SUPPLIER_SHIPMENT_PARSE,
                "source_shipment_id": shipment_id,
                "source_upload_id": upload_id,
                "file_original_name": source_filename,
                "file_content_type": content_type or SUPPLIER_INVOICE_CONTENT_TYPE,
                "file_sha256": file_sha256,
                "file_path": source_file_path,
                "parser_version": parser_version or SUPPLIER_INVOICE_PARSER_VERSION,
                "parsed_metadata": parsed_metadata,
                "warnings": warnings,
                "errors": errors,
                "status": TRADE_DOCUMENT_STATUS_ACTIVE,
                "created_at": created_at,
                "updated_at": created_at,
            }
        )

    def _autolink_invoice_contract_from_metadata(
        self,
        *,
        invoice_document_id: str,
        contract_no: str,
        contract_date: str,
        linked_by: str,
        source: str,
    ) -> bool:
        if not invoice_document_id:
            return False
        if self.runtime.load_invoice_contract_link(invoice_document_id) is not None:
            return False
        candidates = self.find_contract_candidates(contract_no, contract_date)
        if len(candidates) != 1:
            return False
        self.link_invoice_to_contract(
            invoice_document_id,
            contract_document_id=str(candidates[0].get("document_id") or ""),
            linked_by=linked_by,
            source=source,
        )
        return True

    def _ensure_shipment_invoice_document(self, shipment_id: str) -> dict[str, Any]:
        detail = self.runtime.load_supplier_shipment(shipment_id)
        if detail is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        header = dict(detail.get("header") or {})
        if str(header.get("invoice_document_id") or "").strip():
            return self._with_document_fields(_detail_payload(detail))
        source_file_path = str(header.get("source_file_path") or "").strip()
        if not source_file_path:
            raise ValueError(f"supplier shipment invoice file is missing: {shipment_id}")
        self.migrate_existing_supplier_shipments_into_trade_documents()
        refreshed = self.runtime.load_supplier_shipment(shipment_id)
        if refreshed is None:
            raise ValueError(f"supplier shipment not found: {shipment_id}")
        return self._with_document_fields(_detail_payload(refreshed))

    def _with_document_fields(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        enriched = dict(payload)
        shipment_id = str(enriched.get("shipment_id") or "")
        invoice_document_id = str(enriched.get("invoice_document_id") or "")
        invoice_document = self.runtime.load_trade_document(invoice_document_id) if invoice_document_id else None
        if invoice_document is not None and str(invoice_document.get("status") or "") == TRADE_DOCUMENT_STATUS_ACTIVE:
            enriched["invoice_document_id"] = str(invoice_document.get("document_id") or "")
            enriched["invoice_download_path"] = _invoice_download_path(shipment_id)
        else:
            enriched["invoice_document_id"] = invoice_document_id
            enriched["invoice_download_path"] = _invoice_download_path(shipment_id)

        link = self.runtime.load_invoice_contract_link(invoice_document_id) if invoice_document_id else None
        contract_document = None
        if link is not None:
            contract_document = self.runtime.load_trade_document(str(link.get("contract_document_id") or ""))
        if contract_document is not None and str(contract_document.get("status") or "") == TRADE_DOCUMENT_STATUS_ACTIVE:
            contract_document_id = str(contract_document.get("document_id") or "")
            enriched["contract_document_id"] = contract_document_id
            enriched["contract_no"] = contract_document.get("number") or enriched.get("contract_no") or ""
            enriched["contract_date"] = contract_document.get("document_date") or enriched.get("contract_date") or ""
            enriched["contract_download_path"] = _contract_download_path(shipment_id)
            enriched["contract_link_status"] = "linked"
            enriched["contract_candidates"] = []
        else:
            contract_no = str(enriched.get("contract_no") or "")
            contract_date = str(enriched.get("contract_date") or "")
            candidates = self.find_contract_candidates(contract_no, contract_date)
            enriched["contract_document_id"] = ""
            enriched["contract_download_path"] = ""
            enriched["contract_link_status"] = (
                "missing" if not candidates else "single_candidate" if len(candidates) == 1 else "multiple_candidates"
            )
            enriched["contract_candidates"] = candidates
        return enriched

    def _with_approx_cost_fields(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        enriched = dict(payload)
        shipment_id = str(enriched.get("shipment_id") or "").strip()
        documents = self.runtime.list_supplier_financial_documents(shipment_id) if shipment_id else []
        expense_lines = self.runtime.list_supplier_financial_expense_lines(shipment_id) if shipment_id else []
        summary = build_financial_summary(documents, expense_lines, shipment={"header": enriched, "lines": []})
        per_unit = summary.get("per_unit") if isinstance(summary.get("per_unit"), Mapping) else {}
        enriched["approx_yuan_rate"] = _read_optional_positive_decimal(enriched.get("approx_yuan_rate"))
        enriched["approx_invoice_cost_rub"] = per_unit.get("approx_invoice_cost_rub")
        enriched["approx_landed_cost_per_unit_rub"] = per_unit.get("approx_landed_cost_per_unit_rub")
        enriched["exact_bank_fees_rub"] = per_unit.get("exact_bank_fees_rub")
        enriched["exact_currency_payment_cost_rub"] = per_unit.get("exact_currency_payment_cost_rub")
        enriched["exact_landed_cost_total_rub"] = per_unit.get("exact_landed_cost_total_rub")
        enriched["exact_landed_cost_per_unit_rub"] = per_unit.get("exact_landed_cost_per_unit_rub")
        enriched["exact_cost_status"] = per_unit.get("exact_cost_status") or "unavailable"
        enriched["exact_cost_blockers"] = list(per_unit.get("exact_cost_blockers") or [])
        enriched["exact_cost_warnings"] = list(per_unit.get("exact_cost_warnings") or [])
        return enriched

    def _with_document_download_path(self, document: Mapping[str, Any] | None) -> dict[str, Any]:
        if document is None:
            return {}
        payload = dict(document)
        payload["download_path"] = _trade_document_download_path(str(payload.get("document_id") or ""))
        parsed_metadata = payload.get("parsed_metadata")
        parsed_metadata = parsed_metadata if isinstance(parsed_metadata, Mapping) else {}
        payload["parsed_number"] = str(parsed_metadata.get("parsed_number") or "")
        payload["parsed_document_date"] = str(parsed_metadata.get("parsed_document_date") or "")
        payload["parser_warnings"] = _string_list(payload.get("warnings") if isinstance(payload.get("warnings"), list) else [])
        payload["parser_errors"] = _string_list(payload.get("errors") if isinstance(payload.get("errors"), list) else [])
        return payload

    def _sha256_for_existing_runtime_file(self, relative_path: str, *, fallback_seed: str) -> str:
        try:
            file_path = self._resolve_runtime_file(relative_path)
        except ValueError:
            file_path = None
        if file_path is not None and file_path.exists() and file_path.is_file():
            return hashlib.sha256(file_path.read_bytes()).hexdigest()
        return hashlib.sha256(str(fallback_seed or relative_path or "missing").encode("utf-8")).hexdigest()

    def _active_nomenclature_aliases(self) -> list[dict[str, Any]]:
        self._ensure_nomenclature_ready()
        aliases: list[dict[str, Any]] = []
        for item in self.runtime.list_nomenclature_items(active_only=True):
            aliases.extend(_nomenclature_item_aliases(item))
        return aliases

    def _active_nomenclature_items(self) -> list[dict[str, Any]]:
        self._ensure_nomenclature_ready()
        return self.runtime.list_nomenclature_items(active_only=True)

    def _ensure_nomenclature_ready(self) -> None:
        self._ensure_sku_groups_ready()
        self._seed_nomenclature_from_current_config_if_empty()
        self._backfill_nomenclature_compatible_models()

    def _ensure_sku_groups_ready(self) -> None:
        now = self.timestamp_factory()
        existing_keys = {str(group.get("group_key") or "") for group in self.runtime.list_sku_groups(include_inactive=True)}
        for group in DEFAULT_SKU_GROUPS:
            group_key = str(group.get("group_key") or "")
            if not group_key or group_key in existing_keys:
                continue
            self.runtime.save_sku_group(
                {
                    **group,
                    "is_active": True,
                    "is_system": True,
                    "created_at": now,
                    "updated_at": now,
                }
            )
            existing_keys.add(group_key)

    def _validate_nomenclature_unique(self, item: Mapping[str, Any]) -> None:
        if (
            bool(item.get("is_active"))
            and str(item.get("match_key") or "").strip()
            and self.runtime.active_nomenclature_match_key_exists(
                match_key=str(item.get("match_key") or "").strip(),
                exclude_item_id=str(item.get("item_id") or ""),
            )
        ):
            raise ValueError(f"duplicate active nomenclature match_key: {item.get('match_key')}")

    def _seed_nomenclature_from_current_config_if_empty(self) -> None:
        if self.runtime.list_nomenclature_items():
            return
        try:
            current_state = self.runtime.load_current_state()
        except Exception:
            return
        now = self.timestamp_factory()
        seen: set[str] = set()
        for config_item in getattr(current_state, "config_v2", []) or []:
            if not bool(getattr(config_item, "enabled", False)):
                continue
            display_name = str(getattr(config_item, "display_name", "") or "").strip()
            product_type = _product_type_from_config_item(display_name, str(getattr(config_item, "group", "") or ""))
            model_text = _model_text_from_nomenclature_name(display_name)
            normalized_model = normalize_invoice_model(model_text)
            compatible_model_keys = extract_iphone_model_keys(model_text)
            if not product_type or not normalized_model:
                continue
            match_key = f"{product_type}|{normalized_model}"
            if match_key in seen:
                continue
            seen.add(match_key)
            self.runtime.save_nomenclature_item(
                {
                    "item_id": f"nom_seed_{int(getattr(config_item, 'nm_id'))}",
                    "is_active": True,
                    "our_sku": "",
                    "nm_id": int(getattr(config_item, "nm_id")),
                    "nomenclature_name": display_name,
                    "product_type": product_type,
                    "match_key": match_key,
                    "aliases": [],
                    "compatible_models_text": model_text,
                    "compatible_model_keys": compatible_model_keys,
                    "purchase_price_yuan": None,
                    "comment": "seeded from current registry config_v2",
                    "created_at": now,
                    "updated_at": now,
                }
            )

    def _backfill_nomenclature_compatible_models(self) -> None:
        items = self.runtime.list_nomenclature_items()
        now = self.timestamp_factory()
        for item in items:
            if item.get("compatible_model_keys"):
                continue
            keys = _infer_compatible_model_keys(item)
            if not keys:
                continue
            text = str(item.get("compatible_models_text") or "").strip() or _compatible_models_text_from_keys(keys)
            updated = dict(item)
            updated["compatible_models_text"] = text
            updated["compatible_model_keys"] = keys
            updated["updated_at"] = now
            self.runtime.save_nomenclature_item(updated)


def _resolve_edited_payload(payload: Mapping[str, Any], *, fallback: Mapping[str, Any]) -> dict[str, Any]:
    raw = payload.get("payload")
    if raw is None:
        raw = payload.get("edited_payload")
    if isinstance(raw, Mapping):
        resolved = deepcopy(dict(raw))
    else:
        resolved = deepcopy(dict(fallback))
    for key in ("metadata", "lines", "summary", "warnings", "errors"):
        if key in payload and key not in resolved:
            resolved[key] = payload[key]
    for key in ("shipment_date", "actual_shipment_date", "actual_ff_acceptance_date"):
        if key in payload:
            resolved[key] = payload[key]
    return resolved


def _resolve_optional_date_field(
    payload: Mapping[str, Any],
    edited_payload: Mapping[str, Any],
    existing_header: Mapping[str, Any] | None,
    field_name: str,
) -> str:
    if field_name in payload:
        raw = payload.get(field_name)
    elif field_name in edited_payload:
        raw = edited_payload.get(field_name)
    elif existing_header is not None:
        raw = existing_header.get(field_name)
    else:
        raw = ""
    return _validate_optional_iso_date(raw, field_name=field_name)


def _normalize_edit_payload(
    payload: Mapping[str, Any],
    *,
    shipment_date: str,
    force_manual_override: bool,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[str], list[str], dict[str, Any], str]:
    metadata = _normalize_metadata(payload.get("metadata"))
    if not metadata.get("currency") and payload.get("currency"):
        metadata["currency"] = str(payload.get("currency") or "").strip()
    raw_lines = payload.get("lines") or []
    if not isinstance(raw_lines, list):
        raise ValueError("supplier shipment lines must be a list")
    lines = [
        _normalize_line(item, index=index, currency=str(metadata.get("currency") or ""), force_manual_override=force_manual_override)
        for index, item in enumerate(raw_lines, start=1)
        if isinstance(item, Mapping)
    ]
    if len(lines) != len(raw_lines):
        raise ValueError("supplier shipment lines must be JSON objects")
    summary = _recalculate_summary(lines, declared_total=_optional_number(metadata.get("declared_invoice_total")))
    warnings = _string_list(payload.get("warnings"))
    errors = _string_list(payload.get("errors"))
    if summary["checksum_error"] and not any("checksum" in item.lower() for item in errors):
        errors.append(
            "invoice total checksum mismatch: declared "
            f"{summary.get('declared_invoice_total')} vs parsed {summary.get('invoice_amount_total')}"
        )
    match_status = _shipment_match_status(lines, checksum_error=summary["checksum_error"])
    metadata["declared_invoice_total"] = summary.get("declared_invoice_total")
    return metadata, lines, warnings, errors, summary, match_status


def _normalize_metadata(raw: Any) -> dict[str, Any]:
    metadata = dict(raw) if isinstance(raw, Mapping) else {}
    return {
        "invoice_no": str(metadata.get("invoice_no") or "").strip(),
        "invoice_date": _optional_iso_date(metadata.get("invoice_date")),
        "contract_no": str(metadata.get("contract_no") or "").strip(),
        "contract_date": _optional_iso_date(metadata.get("contract_date")),
        "supplier_name": DEFAULT_SUPPLIER_NAME,
        "customer_name": "",
        "currency": str(metadata.get("currency") or "").strip().upper(),
        "declared_invoice_total": _optional_number(metadata.get("declared_invoice_total")),
    }


def _supplier_order_metadata(raw: Any) -> dict[str, Any]:
    metadata = _normalize_metadata(raw)
    metadata["supplier_name"] = DEFAULT_SUPPLIER_NAME
    metadata["customer_name"] = ""
    return metadata


def _normalize_line(
    raw: Mapping[str, Any],
    *,
    index: int,
    currency: str,
    force_manual_override: bool,
) -> dict[str, Any]:
    line_type = str(raw.get("line_type") or LINE_TYPE_PRODUCT).strip()
    if line_type not in {LINE_TYPE_PRODUCT, LINE_TYPE_EXTRA}:
        raise ValueError(f"line #{index}: line_type must be product or extra")
    qty = _optional_number(raw.get("qty"))
    unit_price = _optional_number(raw.get("unit_price"))
    amount = _optional_number(raw.get("amount"))
    if "qty" in raw and raw.get("qty") not in {None, ""} and qty is None:
        raise ValueError(f"line #{index}: qty must be numeric")
    if "unit_price" in raw and raw.get("unit_price") not in {None, ""} and unit_price is None:
        raise ValueError(f"line #{index}: unit_price must be numeric")
    if "amount" in raw and raw.get("amount") not in {None, ""} and amount is None:
        raise ValueError(f"line #{index}: amount must be numeric")
    internal_nm_id = _optional_int(raw.get("internal_nm_id"))
    product_type = str(raw.get("product_type") or "").strip()
    model_normalized = str(raw.get("model_normalized") or "").strip()
    match_key = str(raw.get("match_key") or "").strip()
    if line_type == LINE_TYPE_PRODUCT and product_type and model_normalized and not match_key:
        match_key = f"{product_type}|{model_normalized}"
    has_internal_match = bool(str(raw.get("internal_sku") or "").strip() or internal_nm_id or str(raw.get("internal_name") or "").strip())
    match_status = str(raw.get("match_status") or "").strip()
    if line_type == LINE_TYPE_EXTRA:
        match_status = "extra"
    elif match_status not in {
        MATCH_STATUS_MATCHED,
        MATCH_STATUS_MATCHED_BY_COMPATIBILITY,
        MATCH_STATUS_UNMATCHED,
        MATCH_STATUS_AMBIGUOUS,
    }:
        match_status = MATCH_STATUS_MATCHED if has_internal_match else MATCH_STATUS_UNMATCHED
    raw_payload = raw.get("raw") if isinstance(raw.get("raw"), Mapping) else {}
    return {
        "line_id": str(raw.get("line_id") or ("ln_" + uuid4().hex)).strip(),
        "line_type": line_type,
        "sort_order": _optional_int(raw.get("sort_order")) or index,
        "source_no": str(raw.get("source_no") or "").strip(),
        "product_type": product_type,
        "model_raw": str(raw.get("model_raw") or "").strip(),
        "model_normalized": model_normalized,
        "match_key": match_key,
        "internal_sku": str(raw.get("internal_sku") or "").strip(),
        "internal_nm_id": internal_nm_id,
        "internal_name": str(raw.get("internal_name") or "").strip(),
        "qty": qty,
        "unit_price": unit_price,
        "amount": amount,
        "currency": str(raw.get("currency") or currency or "").strip().upper(),
        "comment": str(raw.get("comment") or "").strip(),
        "match_status": match_status,
        "manual_override": bool(raw.get("manual_override")) or force_manual_override,
        "invoice_price_yuan_snapshot": _optional_number(raw.get("invoice_price_yuan_snapshot")),
        "reference_purchase_price_yuan_snapshot": _optional_number(raw.get("reference_purchase_price_yuan_snapshot")),
        "price_conformity_status": _normalize_price_conformity_status(raw.get("price_conformity_status")),
        "price_conformity_checked_at": _optional_timestamp(raw.get("price_conformity_checked_at")),
        "price_conformity_check_mode": _normalize_price_conformity_check_mode(raw.get("price_conformity_check_mode")),
        "price_conformity_reason": str(raw.get("price_conformity_reason") or "not_checked").strip() or "not_checked",
        "price_conformity_actor": str(raw.get("price_conformity_actor") or "").strip(),
        "price_conformity_context": _normalize_json_object(raw.get("price_conformity_context")),
        "raw": dict(raw_payload),
    }


def _recalculate_summary(lines: list[Mapping[str, Any]], *, declared_total: float | None) -> dict[str, Any]:
    product_qty_total = _sum_numeric(item.get("qty") for item in lines if item.get("line_type") == LINE_TYPE_PRODUCT)
    product_amount_total = _sum_numeric(item.get("amount") for item in lines if item.get("line_type") == LINE_TYPE_PRODUCT)
    extras_amount_total = _sum_numeric(item.get("amount") for item in lines if item.get("line_type") == LINE_TYPE_EXTRA)
    invoice_amount_total = round(product_amount_total + extras_amount_total, 2)
    checksum_error = bool(declared_total is not None and abs(round(declared_total - invoice_amount_total, 2)) > 0.02)
    return {
        "product_qty_total": product_qty_total,
        "product_amount_total": product_amount_total,
        "extras_amount_total": extras_amount_total,
        "invoice_amount_total": invoice_amount_total,
        "declared_invoice_total": declared_total,
        "checksum_error": checksum_error,
    }


def _assert_atomic_supplier_product_matching(lines: Iterable[Mapping[str, Any]]) -> None:
    problems: list[str] = []
    product_count = 0
    for index, line in enumerate(lines, start=1):
        if str(line.get("line_type") or "") != LINE_TYPE_PRODUCT:
            continue
        product_count += 1
        line_id = str(line.get("line_id") or line.get("source_no") or index)
        status = str(line.get("match_status") or "")
        if status not in {MATCH_STATUS_MATCHED, MATCH_STATUS_MATCHED_BY_COMPATIBILITY}:
            problems.append(f"строка {line_id}: SKU {status or 'не сопоставлен'}")
        if _optional_int(line.get("internal_nm_id")) is None:
            problems.append(f"строка {line_id}: отсутствует nmID")
        if (_optional_number(line.get("qty")) or 0) <= 0:
            problems.append(f"строка {line_id}: отсутствует положительное количество")
        if (_optional_number(line.get("unit_price")) or 0) <= 0:
            problems.append(f"строка {line_id}: отсутствует положительная цена")
        if (_optional_number(line.get("amount")) or 0) <= 0:
            problems.append(f"строка {line_id}: отсутствует положительная сумма")
    if product_count <= 0:
        problems.append("товарные строки отсутствуют")
    if problems:
        raise ValueError(
            "supplier document rejected atomically; correct authoritative nomenclature/aliases and reparse: "
            + "; ".join(problems)
        )


def _shipment_match_status(lines: list[Mapping[str, Any]], *, checksum_error: bool) -> str:
    if checksum_error:
        return SHIPMENT_STATUS_CHECKSUM_ERROR
    if any(bool(item.get("manual_override")) for item in lines):
        return SHIPMENT_STATUS_MANUAL_OVERRIDE
    if any(
        item.get("line_type") == LINE_TYPE_PRODUCT
        and item.get("match_status") not in {MATCH_STATUS_MATCHED, MATCH_STATUS_MATCHED_BY_COMPATIBILITY}
        for item in lines
    ):
        return SHIPMENT_STATUS_HAS_UNMATCHED
    return SHIPMENT_STATUS_ALL_MATCHED


def _detail_payload(detail: Mapping[str, Any]) -> dict[str, Any]:
    header = dict(detail.get("header") or {})
    header["supplier_name"] = DEFAULT_SUPPLIER_NAME
    header["customer_name"] = ""
    header["order_status"] = _normalize_order_status(header.get("order_status"))
    header["planned_shipment_date"] = header.get("shipment_date") or ""
    header["actual_shipment_date"] = header.get("actual_shipment_date") or ""
    header["actual_ff_acceptance_date"] = header.get("actual_ff_acceptance_date") or ""
    lines = [dict(item) for item in detail.get("lines") or []]
    summary = {
        "product_qty_total": header.get("product_qty_total"),
        "product_amount_total": header.get("product_amount_total"),
        "extras_amount_total": header.get("extras_amount_total"),
        "invoice_amount_total": header.get("invoice_amount_total"),
        "declared_invoice_total": header.get("declared_invoice_total"),
        "checksum_error": header.get("match_status") == SHIPMENT_STATUS_CHECKSUM_ERROR,
    }
    payload = {
        **header,
        "metadata": {
            "invoice_no": header.get("invoice_no") or "",
            "invoice_date": header.get("invoice_date") or "",
            "contract_no": header.get("contract_no") or "",
            "contract_date": header.get("contract_date") or "",
            "supplier_name": header.get("supplier_name") or "",
            "customer_name": header.get("customer_name") or "",
            "currency": header.get("currency") or "",
            "declared_invoice_total": header.get("declared_invoice_total"),
        },
        "summary": summary,
        "lines": lines,
        "product_lines": [item for item in lines if item.get("line_type") == LINE_TYPE_PRODUCT],
        "extra_lines": [item for item in lines if item.get("line_type") == LINE_TYPE_EXTRA],
        "invoice_download_path": _invoice_download_path(str(header.get("shipment_id") or "")),
    }
    return payload


def _with_invoice_download_path(row: Mapping[str, Any]) -> dict[str, Any]:
    payload = dict(row)
    payload["supplier_name"] = DEFAULT_SUPPLIER_NAME
    payload["customer_name"] = ""
    payload["order_status"] = _normalize_order_status(payload.get("order_status"))
    payload["planned_shipment_date"] = payload.get("shipment_date") or ""
    payload["actual_shipment_date"] = payload.get("actual_shipment_date") or ""
    payload["actual_ff_acceptance_date"] = payload.get("actual_ff_acceptance_date") or ""
    payload["invoice_download_path"] = _invoice_download_path(str(payload.get("shipment_id") or ""))
    return payload


def _apply_nomenclature_matches(
    lines: list[Mapping[str, Any]],
    nomenclature_items: list[Mapping[str, Any]],
    *,
    overwrite_manual: bool = False,
) -> list[dict[str, Any]]:
    index = _build_nomenclature_match_index(nomenclature_items)
    matched_lines: list[dict[str, Any]] = []
    for raw_line in lines:
        line = dict(raw_line)
        if line.get("line_type") != LINE_TYPE_PRODUCT:
            matched_lines.append(line)
            continue
        product_type = str(line.get("product_type") or "").strip()
        normalized_model = str(line.get("model_normalized") or "").strip()
        match_key = str(line.get("match_key") or "").strip()
        if not match_key and product_type and normalized_model:
            match_key = f"{product_type}|{normalized_model}"
            line["match_key"] = match_key
        if bool(line.get("manual_override")) and not overwrite_manual:
            matched_lines.append(line)
            continue
        resolution = _resolve_nomenclature_match(line, index)
        _apply_match_resolution(line, resolution)
        matched_lines.append(line)
    return matched_lines


def _build_nomenclature_match_index(items: list[Mapping[str, Any]]) -> dict[str, Any]:
    exact_by_key: dict[str, list[dict[str, Any]]] = {}
    alias_by_key: dict[str, list[dict[str, Any]]] = {}
    compatible: list[dict[str, Any]] = []
    for item in items:
        if not bool(item.get("is_active")):
            continue
        item_payload = _nomenclature_item_match_payload(item)
        base_match_key = str(item.get("match_key") or "").strip()
        if base_match_key:
            exact_by_key.setdefault(base_match_key, []).append(item_payload)
        for alias in _nomenclature_item_aliases(item):
            alias_key = str(alias.get("match_key") or "").strip()
            if alias_key and alias_key != base_match_key:
                alias_by_key.setdefault(alias_key, []).append(_nomenclature_item_match_payload({**item, **alias}))
        compatible_keys = _infer_compatible_model_keys(item)
        if compatible_keys and str(item.get("product_type") or "") in {"clear", "anti_spy", "matte"}:
            compatible.append({**item_payload, "compatible_model_keys": compatible_keys})
    return {
        "exact_by_key": exact_by_key,
        "alias_by_key": alias_by_key,
        "compatible": compatible,
    }


def _resolve_nomenclature_match(line: Mapping[str, Any], index: Mapping[str, Any]) -> dict[str, Any] | None:
    product_type = str(line.get("product_type") or "").strip()
    match_key = str(line.get("match_key") or "").strip()
    exact_candidates = list((index.get("exact_by_key") or {}).get(match_key) or [])
    if len(exact_candidates) == 1:
        return {**exact_candidates[0], "match_status": MATCH_STATUS_MATCHED}
    if len(exact_candidates) > 1:
        return {"match_status": MATCH_STATUS_AMBIGUOUS}

    alias_candidates = list((index.get("alias_by_key") or {}).get(match_key) or [])
    if len(alias_candidates) == 1:
        return {**alias_candidates[0], "match_status": MATCH_STATUS_MATCHED}
    if len(alias_candidates) > 1:
        return {"match_status": MATCH_STATUS_AMBIGUOUS}

    invoice_keys = _line_compatible_model_keys(line)
    if product_type not in {"clear", "anti_spy", "matte"} or not invoice_keys:
        return None
    invoice_key_set = set(invoice_keys)
    scored: list[tuple[tuple[int, int, int], dict[str, Any]]] = []
    for candidate in index.get("compatible") or []:
        if str(candidate.get("product_type") or "") != product_type:
            continue
        candidate_keys = [str(item) for item in candidate.get("compatible_model_keys") or [] if str(item or "").strip()]
        if not candidate_keys:
            continue
        candidate_key_set = set(candidate_keys)
        intersection = sorted(invoice_key_set & candidate_key_set)
        if not intersection:
            continue
        subset_bonus = 1 if candidate_key_set.issubset(invoice_key_set) or invoice_key_set.issubset(candidate_key_set) else 0
        exact_size_bonus = 1 if candidate_key_set == invoice_key_set else 0
        score = (subset_bonus, len(intersection), exact_size_bonus)
        scored.append((score, {**candidate, "matched_model_keys": intersection}))
    if not scored:
        return None
    scored.sort(key=lambda item: item[0], reverse=True)
    top_score, top_candidate = scored[0]
    if len(scored) > 1 and scored[1][0] == top_score:
        return {"match_status": MATCH_STATUS_AMBIGUOUS}
    return {**top_candidate, "match_status": MATCH_STATUS_MATCHED_BY_COMPATIBILITY}


def _apply_match_resolution(line: dict[str, Any], resolution: Mapping[str, Any] | None) -> None:
    if not resolution:
        line["internal_sku"] = ""
        line["internal_nm_id"] = None
        line["internal_name"] = ""
        line["match_status"] = MATCH_STATUS_UNMATCHED
        line["manual_override"] = False
        return
    if str(resolution.get("match_status") or "") == MATCH_STATUS_AMBIGUOUS:
        line["internal_sku"] = ""
        line["internal_nm_id"] = None
        line["internal_name"] = ""
        line["match_status"] = MATCH_STATUS_AMBIGUOUS
        line["manual_override"] = False
        return
    line["internal_sku"] = str(resolution.get("internal_sku") or resolution.get("our_sku") or "")
    line["internal_nm_id"] = _optional_int(resolution.get("internal_nm_id") or resolution.get("nm_id"))
    line["internal_name"] = str(resolution.get("internal_name") or resolution.get("nomenclature_name") or "")
    line["match_status"] = str(resolution.get("match_status") or MATCH_STATUS_MATCHED)
    line["manual_override"] = False


def _apply_price_conformity_checks(
    lines: list[Mapping[str, Any]],
    nomenclature_items: list[Mapping[str, Any]],
    *,
    checked_at: str,
    mode: str,
    actor: str = "",
    context: Mapping[str, Any] | None = None,
    only_missing: bool = False,
    default_currency: str = "",
) -> list[dict[str, Any]]:
    normalized_mode = _normalize_price_conformity_check_mode(mode)
    reference_index = _build_price_reference_index(nomenclature_items)
    checked_lines: list[dict[str, Any]] = []
    for raw_line in lines:
        line = dict(raw_line)
        if only_missing and not _price_conformity_missing(line):
            checked_lines.append(line)
            continue
        line.update(
            _price_conformity_check_payload(
                line,
                reference_index=reference_index,
                checked_at=checked_at,
                mode=normalized_mode,
                actor=actor,
                context=context or {},
                default_currency=default_currency,
            )
        )
        checked_lines.append(line)
    return checked_lines


def _price_conformity_check_payload(
    line: Mapping[str, Any],
    *,
    reference_index: Mapping[str, Any],
    checked_at: str,
    mode: str,
    actor: str,
    context: Mapping[str, Any],
    default_currency: str,
) -> dict[str, Any]:
    invoice_price = _parse_money_decimal(line.get("unit_price"))
    reference_item, reference_reason = _resolve_price_reference_item(line, reference_index)
    reference_price = _parse_money_decimal(reference_item.get("purchase_price_yuan")) if reference_item else None
    payload = {
        "invoice_price_yuan_snapshot": _decimal_snapshot(invoice_price),
        "reference_purchase_price_yuan_snapshot": _decimal_snapshot(reference_price),
        "price_conformity_checked_at": checked_at,
        "price_conformity_check_mode": mode,
        "price_conformity_actor": str(actor or "").strip() if mode == PRICE_CONFORMITY_CHECK_MODE_MANUAL_RECHECK else "",
        "price_conformity_context": dict(context or {}) if mode == PRICE_CONFORMITY_CHECK_MODE_MANUAL_RECHECK else {},
    }
    if line.get("line_type") != LINE_TYPE_PRODUCT:
        return {
            **payload,
            "price_conformity_status": PRICE_CONFORMITY_STATUS_NOT_CHECKED,
            "price_conformity_reason": "not_product_line",
        }
    if reference_item is None:
        return {
            **payload,
            "price_conformity_status": PRICE_CONFORMITY_STATUS_SKU_NOT_FOUND,
            "price_conformity_reason": reference_reason or "sku_not_found",
        }
    if reference_price is None:
        return {
            **payload,
            "price_conformity_status": PRICE_CONFORMITY_STATUS_REFERENCE_PRICE_MISSING,
            "price_conformity_reason": "reference_price_missing",
        }
    if invoice_price is None:
        return {
            **payload,
            "price_conformity_status": PRICE_CONFORMITY_STATUS_INVOICE_PRICE_MISSING,
            "price_conformity_reason": "invoice_price_missing",
        }
    currency_reason = _price_currency_blocker_reason(line.get("currency") or default_currency)
    if currency_reason:
        return {
            **payload,
            "price_conformity_status": PRICE_CONFORMITY_STATUS_NOT_CHECKED,
            "price_conformity_reason": currency_reason,
        }
    if invoice_price == reference_price:
        return {
            **payload,
            "price_conformity_status": PRICE_CONFORMITY_STATUS_MATCHED,
            "price_conformity_reason": "price_matched",
        }
    return {
        **payload,
        "price_conformity_status": PRICE_CONFORMITY_STATUS_MISMATCHED,
        "price_conformity_reason": "price_mismatch",
    }


def _build_price_reference_index(nomenclature_items: list[Mapping[str, Any]]) -> dict[str, Any]:
    active_items = [dict(item) for item in nomenclature_items if bool(item.get("is_active"))]
    by_item_id = {str(item.get("item_id") or ""): item for item in active_items if str(item.get("item_id") or "")}
    by_nm_id: dict[int, list[dict[str, Any]]] = {}
    for item in active_items:
        nm_id = _optional_int(item.get("nm_id"))
        if nm_id is not None:
            by_nm_id.setdefault(nm_id, []).append(item)
    return {
        "by_item_id": by_item_id,
        "by_nm_id": by_nm_id,
        "match_index": _build_nomenclature_match_index(active_items),
    }


def _resolve_price_reference_item(
    line: Mapping[str, Any],
    reference_index: Mapping[str, Any],
) -> tuple[dict[str, Any] | None, str]:
    internal_nm_id = _optional_int(line.get("internal_nm_id"))
    if internal_nm_id is not None:
        candidates = list((reference_index.get("by_nm_id") or {}).get(internal_nm_id) or [])
        if len(candidates) == 1:
            return candidates[0], ""
        if len(candidates) > 1:
            return None, "reference_sku_ambiguous"
    resolution = _resolve_nomenclature_match(line, reference_index.get("match_index") or {})
    if not resolution:
        return None, "sku_not_found"
    if str(resolution.get("match_status") or "") == MATCH_STATUS_AMBIGUOUS:
        return None, "reference_match_ambiguous"
    item_id = str(resolution.get("item_id") or "")
    item = (reference_index.get("by_item_id") or {}).get(item_id)
    if item is not None:
        return dict(item), ""
    return dict(resolution), ""


def _parse_money_decimal(value: Any) -> Decimal | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        numeric = value
    elif isinstance(value, (int, float)):
        numeric = Decimal(str(value))
    else:
        text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace("−", "-")
        if not text:
            return None
        text = re.sub(r"[^0-9,.\-]", "", text)
        if not text or text in {"-", ".", ","}:
            return None
        if text.count("-") > 1 or ("-" in text and not text.startswith("-")):
            return None
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            numeric = Decimal(text)
        except (InvalidOperation, ValueError):
            return None
    try:
        return numeric.quantize(PRICE_CONFORMITY_MONEY_QUANT, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def _decimal_snapshot(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _price_currency_blocker_reason(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "currency_missing"
    normalized = text.upper().replace("￥", "¥")
    if any(token in normalized for token in ("USD", "EUR", "RUB", "₽", "$")):
        return "currency_not_yuan"
    if normalized in PRICE_CONFORMITY_YUAN_CURRENCIES:
        return ""
    if any(alias in normalized for alias in PRICE_CONFORMITY_YUAN_CURRENCIES):
        return ""
    return "currency_not_yuan"


def _price_conformity_missing(line: Mapping[str, Any]) -> bool:
    return not str(line.get("price_conformity_checked_at") or "").strip()


def _line_was_backfilled(original_lines: list[Mapping[str, Any]], line: Mapping[str, Any]) -> bool:
    line_id = str(line.get("line_id") or "")
    original = next((item for item in original_lines if str(item.get("line_id") or "") == line_id), None)
    return original is None or _price_conformity_missing(original)


def _normalize_nomenclature_payload(
    payload: Mapping[str, Any],
    *,
    item_id: str,
    created_at: str,
    updated_at: str,
) -> dict[str, Any]:
    product_type = str(payload.get("product_type") or "").strip()
    if not _valid_sku_group_key(product_type):
        raise ValueError("nomenclature group must be a stable key like clean, anti_spy, matte, extra or other")
    is_active = bool(payload.get("is_active", True))
    is_hidden = bool(payload.get("is_hidden", False))
    nomenclature_name = str(payload.get("nomenclature_name") or "").strip()
    match_key = _normalize_match_key(payload.get("match_key"))
    compatible_models_text = str(payload.get("compatible_models_text") or "").strip()
    compatible_model_keys = _normalize_compatible_model_keys(
        payload.get("compatible_model_keys"),
        fallback_text=compatible_models_text,
        item_hint={**dict(payload), "match_key": match_key, "nomenclature_name": nomenclature_name},
    )
    if not compatible_models_text and compatible_model_keys:
        compatible_models_text = _compatible_models_text_from_keys(compatible_model_keys)
    if is_active and product_type not in {"extra", "other"}:
        if not nomenclature_name:
            raise ValueError("active product nomenclature item requires nomenclature_name")
    purchase_price_yuan = _optional_nonnegative_number(
        payload.get("purchase_price_yuan"),
        field_name="nomenclature purchase_price_yuan",
    )
    barcode = _normalize_barcode(payload.get("barcode") or payload.get("primary_barcode"))
    barcodes = _normalize_barcode_list([barcode, *_raw_barcode_list(payload.get("barcodes"))])
    if not barcode and barcodes:
        barcode = barcodes[0]
    barcode_source = _normalize_barcode_source(payload.get("barcode_source"), has_barcode=bool(barcode))
    barcode_status = _normalize_barcode_status(
        payload.get("barcode_status"),
        barcode_source=barcode_source,
        barcode_count=len(barcodes),
        has_barcode=bool(barcode),
    )
    return {
        "item_id": item_id,
        "is_active": is_active,
        "is_hidden": is_hidden,
        "hidden_at": str(payload.get("hidden_at") or "").strip(),
        "hidden_reason": str(payload.get("hidden_reason") or "").strip(),
        "our_sku": str(payload.get("our_sku") or "").strip(),
        "nm_id": _optional_int(payload.get("nm_id")),
        "barcode": barcode,
        "primary_barcode": barcode,
        "barcodes": barcodes,
        "barcode_source": barcode_source,
        "barcode_status": barcode_status,
        "barcode_synced_at": str(payload.get("barcode_synced_at") or "").strip(),
        "barcode_updated_at": str(payload.get("barcode_updated_at") or "").strip(),
        "barcode_evidence": payload.get("barcode_evidence") if isinstance(payload.get("barcode_evidence"), Mapping) else {},
        "vendor_code": str(payload.get("vendor_code") or payload.get("seller_article") or "").strip(),
        "seller_article": str(payload.get("vendor_code") or payload.get("seller_article") or "").strip(),
        "wb_title": str(payload.get("wb_title") or "").strip(),
        "wb_subject_name": str(payload.get("wb_subject_name") or "").strip(),
        "wb_updated_at": str(payload.get("wb_updated_at") or "").strip(),
        "wb_synced_at": str(payload.get("wb_synced_at") or "").strip(),
        "wb_sync_status": str(payload.get("wb_sync_status") or "").strip(),
        "wb_sync_evidence": payload.get("wb_sync_evidence") if isinstance(payload.get("wb_sync_evidence"), Mapping) else {},
        "nomenclature_name": nomenclature_name,
        "product_type": product_type,
        "match_key": match_key,
        "purchase_price_yuan": purchase_price_yuan,
        "aliases": _normalize_alias_list(payload.get("aliases")),
        "compatible_models_text": compatible_models_text,
        "compatible_model_keys": compatible_model_keys,
        "comment": str(payload.get("comment") or "").strip(),
        "created_at": created_at,
        "updated_at": updated_at,
    }


def _prepare_nomenclature_barcode_payload(
    *,
    existing: Mapping[str, Any] | None,
    payload: Mapping[str, Any],
    updated_at: str,
) -> dict[str, Any]:
    prepared = dict(payload)
    barcode_present = "barcode" in payload or "primary_barcode" in payload
    if not barcode_present:
        return prepared
    incoming = _normalize_barcode(payload.get("barcode") if "barcode" in payload else payload.get("primary_barcode"))
    existing_barcode = _normalize_barcode((existing or {}).get("barcode") or (existing or {}).get("primary_barcode"))
    if incoming == existing_barcode and existing is not None:
        prepared.setdefault("barcodes", existing.get("barcodes") or ([incoming] if incoming else []))
        prepared.setdefault("barcode_source", existing.get("barcode_source") or "")
        prepared.setdefault("barcode_status", existing.get("barcode_status") or "")
        prepared.setdefault("barcode_synced_at", existing.get("barcode_synced_at") or "")
        prepared.setdefault("barcode_updated_at", existing.get("barcode_updated_at") or "")
        prepared.setdefault("barcode_evidence", existing.get("barcode_evidence") or {})
        return prepared
    if incoming:
        prepared.update(
            {
                "barcode": incoming,
                "primary_barcode": incoming,
                "barcodes": _normalize_barcode_list([incoming, *_raw_barcode_list(payload.get("barcodes"))]),
                "barcode_source": NOMENCLATURE_BARCODE_SOURCE_MANUAL,
                "barcode_status": NOMENCLATURE_BARCODE_STATUS_MANUAL,
                "barcode_updated_at": updated_at,
                "barcode_evidence": {"source": "manual_override"},
            }
        )
        return prepared
    prepared.update(
        {
            "barcode": "",
            "primary_barcode": "",
            "barcodes": [],
            "barcode_source": NOMENCLATURE_BARCODE_SOURCE_MISSING,
            "barcode_status": NOMENCLATURE_BARCODE_STATUS_MISSING,
            "barcode_updated_at": updated_at,
            "barcode_evidence": {"source": "manual_clear"},
        }
    )
    return prepared


def _normalize_barcode(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def _raw_barcode_list(value: Any) -> list[str]:
    if isinstance(value, str):
        return [item for item in re.split(r"[\n,;]+", value)]
    if isinstance(value, list):
        return [str(item) for item in value]
    return []


def _normalize_barcode_list(value: Any) -> list[str]:
    raw_items = _raw_barcode_list(value) if not isinstance(value, list) else value
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in raw_items:
        barcode = _normalize_barcode(raw)
        if not barcode or barcode in seen:
            continue
        seen.add(barcode)
        normalized.append(barcode)
    return normalized


def _normalize_barcode_source(value: Any, *, has_barcode: bool) -> str:
    normalized = str(value or "").strip()
    if normalized in NOMENCLATURE_BARCODE_SOURCES:
        return normalized
    return NOMENCLATURE_BARCODE_SOURCE_MANUAL if has_barcode else NOMENCLATURE_BARCODE_SOURCE_MISSING


def _normalize_barcode_status(value: Any, *, barcode_source: str, barcode_count: int, has_barcode: bool) -> str:
    normalized = str(value or "").strip()
    if normalized in NOMENCLATURE_BARCODE_STATUSES:
        return normalized
    if barcode_source == NOMENCLATURE_BARCODE_SOURCE_MANUAL and has_barcode:
        return NOMENCLATURE_BARCODE_STATUS_MANUAL
    if barcode_count > 1:
        return NOMENCLATURE_BARCODE_STATUS_MULTIPLE
    return NOMENCLATURE_BARCODE_STATUS_READY if has_barcode else NOMENCLATURE_BARCODE_STATUS_MISSING


def _barcode_resolution_barcodes(resolution: Any) -> list[str]:
    if resolution is None:
        return []
    if isinstance(resolution, Mapping):
        return _normalize_barcode_list(resolution.get("barcodes") or resolution.get("skus") or [])
    return _normalize_barcode_list(getattr(resolution, "barcodes", []))


def _barcode_resolution_evidence(resolution: Any, *, nm_id: int, sync_reason: str) -> dict[str, Any]:
    if isinstance(resolution, Mapping):
        cards_found = _optional_int(resolution.get("cards_found"))
        pages_fetched = _optional_int(resolution.get("pages_fetched"))
        endpoint = str(resolution.get("endpoint") or "/content/v2/get/cards/list")
    else:
        cards_found = _optional_int(getattr(resolution, "cards_found", None))
        pages_fetched = _optional_int(getattr(resolution, "pages_fetched", None))
        endpoint = str(getattr(resolution, "endpoint", "/content/v2/get/cards/list"))
    return {
        "source": NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT,
        "endpoint": endpoint,
        "nm_id": nm_id,
        "cards_found": cards_found or 0,
        "pages_fetched": pages_fetched or 0,
        "sync_reason": sync_reason,
    }


def _nomenclature_barcode_sync_error_item(
    item: Mapping[str, Any],
    *,
    status: str,
    error: Exception,
    updated_at: str,
    sync_reason: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    normalized = dict(item)
    existing_barcode = _normalize_barcode(normalized.get("barcode"))
    normalized.update(
        {
            "barcode_source": normalized.get("barcode_source") or NOMENCLATURE_BARCODE_SOURCE_ERROR,
            "barcode_status": status,
            "barcode_updated_at": updated_at,
            "barcode_evidence": {
                "source": NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT,
                "result": status,
                "sync_reason": sync_reason,
                "error": _safe_barcode_error(error),
            },
        }
    )
    if not existing_barcode:
        normalized["barcode_source"] = NOMENCLATURE_BARCODE_SOURCE_ERROR
    return normalized, {"status": status, "reason": sync_reason, "save_item": True}


def _safe_barcode_error(error: Exception) -> str:
    text = str(error or "")
    text = re.sub(r"(?i)(authorization|token|cookie|password|secret)([\"'=:\s]+)([^\\s\"'<>;,]+)", r"\1\2<redacted>", text)
    return re.sub(r"\s+", " ", text).strip()[:420]


def _nomenclature_barcode_summary(items: list[Mapping[str, Any]]) -> dict[str, Any]:
    active_rows = [item for item in items if bool(item.get("is_active"))]
    active_with_barcode = [item for item in active_rows if str(item.get("barcode") or "").strip()]
    hidden_rows = [item for item in items if bool(item.get("is_hidden"))]
    return {
        "total_rows": len(items),
        "active_rows": len(active_rows),
        "hidden_rows": len(hidden_rows),
        "visible_rows": len(items) - len(hidden_rows),
        "active_rows_with_barcode": len(active_with_barcode),
        "active_rows_missing_barcode": len(active_rows) - len(active_with_barcode),
        "manual_barcode_count": sum(
            1
            for item in items
            if str(item.get("barcode_source") or "") == NOMENCLATURE_BARCODE_SOURCE_MANUAL
            and str(item.get("barcode") or "").strip()
        ),
        "wb_content_barcode_count": sum(
            1
            for item in items
            if str(item.get("barcode_source") or "") == NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT
            and str(item.get("barcode") or "").strip()
        ),
        "sync_error_count": sum(
            1
            for item in items
            if str(item.get("barcode_status") or "") in {NOMENCLATURE_BARCODE_STATUS_SYNC_ERROR, NOMENCLATURE_BARCODE_STATUS_TOKEN_MISSING}
        ),
        "multiple_barcode_count": sum(
            1 for item in items if str(item.get("barcode_status") or "") == NOMENCLATURE_BARCODE_STATUS_MULTIPLE
        ),
    }


def _normalize_visibility_mode(value: Any) -> str:
    normalized = str(value or "visible").strip().casefold()
    if normalized in {"visible", "hidden", "all"}:
        return normalized
    return "visible"


def _normalize_sku_group_payload(payload: Mapping[str, Any], *, created_at: str, updated_at: str) -> dict[str, Any]:
    group_key = _normalize_sku_group_key(payload.get("group_key"))
    if not _valid_sku_group_key(group_key):
        raise ValueError("sku group_key must contain lowercase latin letters, digits or underscores")
    label = str(payload.get("label") or group_key).strip()
    aliases = _normalize_alias_list(payload.get("aliases"))
    return {
        "group_key": group_key,
        "label": label,
        "aliases": aliases,
        "is_active": bool(payload.get("is_active", True)),
        "is_system": bool(payload.get("is_system", False)),
        "created_at": created_at,
        "updated_at": updated_at,
    }


def _normalize_sku_group_key(value: Any) -> str:
    normalized = str(value or "").strip().casefold()
    normalized = normalized.replace("-", "_").replace(" ", "_")
    normalized = re.sub(r"[^a-z0-9_]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized


def _valid_sku_group_key(value: Any) -> bool:
    normalized = str(value or "").strip()
    return bool(re.fullmatch(r"[a-z][a-z0-9_]{0,63}", normalized))


def _normalize_wb_card_for_sync(raw_card: Any) -> dict[str, Any]:
    if isinstance(raw_card, WbContentCard):
        payload = raw_card.to_dict()
    elif isinstance(raw_card, Mapping):
        payload = dict(raw_card)
    else:
        payload = {}
    nm_id = _optional_int(payload.get("nm_id") or payload.get("nmID") or payload.get("nmId"))
    vendor_code = str(payload.get("vendor_code") or payload.get("vendorCode") or payload.get("seller_article") or "").strip()
    title = str(payload.get("title") or payload.get("wb_title") or payload.get("name") or "").strip()
    subject_name = str(payload.get("subject_name") or payload.get("subjectName") or payload.get("wb_subject_name") or "").strip()
    updated_at = str(payload.get("updated_at") or payload.get("updatedAt") or payload.get("wb_updated_at") or "").strip()
    return {
        "nm_id": nm_id,
        "vendor_code": vendor_code,
        "title": title,
        "subject_name": subject_name,
        "updated_at": updated_at,
        "barcodes": _normalize_barcode_list(payload.get("barcodes") or payload.get("skus") or []),
        "endpoint": str(payload.get("endpoint") or "/content/v2/get/cards/list"),
    }


def _wb_card_has_identity(card: Mapping[str, Any]) -> bool:
    return (
        _optional_int(card.get("nm_id")) is not None
        or bool(_normalize_barcode_list(card.get("barcodes") or []))
        or bool(_normalize_vendor_code(card.get("vendor_code")))
    )


class _NomenclatureWbCardMatcher:
    def __init__(self, rows: list[Mapping[str, Any]]) -> None:
        self.rows = [dict(row) for row in rows]
        self._rebuild()

    def _rebuild(self) -> None:
        self.by_nm_id: dict[int, dict[str, Any]] = {}
        self.by_barcode: dict[str, dict[str, Any]] = {}
        self.by_vendor_code: dict[str, dict[str, Any]] = {}
        for row in self.rows:
            nm_id = _optional_int(row.get("nm_id"))
            if nm_id is not None and nm_id > 0:
                self.by_nm_id.setdefault(nm_id, row)
            for barcode in _normalize_barcode_list([row.get("barcode"), *(row.get("barcodes") or [])]):
                self.by_barcode.setdefault(barcode, row)
            vendor_code = _normalize_vendor_code(row.get("vendor_code") or row.get("seller_article"))
            if vendor_code:
                self.by_vendor_code.setdefault(vendor_code, row)

    def match(self, card: Mapping[str, Any]) -> tuple[dict[str, Any] | None, str]:
        nm_id = _optional_int(card.get("nm_id"))
        if nm_id is not None and nm_id in self.by_nm_id:
            return dict(self.by_nm_id[nm_id]), "nm_id"
        for barcode in _normalize_barcode_list(card.get("barcodes") or []):
            if barcode in self.by_barcode:
                return dict(self.by_barcode[barcode]), "barcode"
        vendor_code = _normalize_vendor_code(card.get("vendor_code"))
        if vendor_code and vendor_code in self.by_vendor_code:
            return dict(self.by_vendor_code[vendor_code]), "vendor_code"
        return None, ""

    def replace(self, row: Mapping[str, Any]) -> None:
        item_id = str(row.get("item_id") or "")
        replaced = False
        for index, existing in enumerate(self.rows):
            if str(existing.get("item_id") or "") == item_id:
                self.rows[index] = dict(row)
                replaced = True
                break
        if not replaced:
            self.rows.append(dict(row))
        self._rebuild()


def _apply_wb_card_to_existing_nomenclature(
    existing: Mapping[str, Any],
    *,
    card: Mapping[str, Any],
    match_type: str,
    synced_at: str,
) -> dict[str, Any]:
    updated = dict(existing)
    card_nm_id = _optional_int(card.get("nm_id"))
    if _optional_int(updated.get("nm_id")) is None and card_nm_id is not None:
        updated["nm_id"] = card_nm_id
    barcodes = sorted(_normalize_barcode_list(card.get("barcodes") or []))
    manual_barcode = (
        str(updated.get("barcode_source") or "") == NOMENCLATURE_BARCODE_SOURCE_MANUAL
        and bool(str(updated.get("barcode") or "").strip())
    )
    if barcodes and not manual_barcode:
        updated["barcode"] = barcodes[0]
        updated["primary_barcode"] = barcodes[0]
        updated["barcodes"] = barcodes
        updated["barcode_source"] = NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT
        updated["barcode_status"] = (
            NOMENCLATURE_BARCODE_STATUS_MULTIPLE if len(barcodes) > 1 else NOMENCLATURE_BARCODE_STATUS_READY
        )
        updated["barcode_synced_at"] = synced_at
        updated["barcode_updated_at"] = synced_at
        updated["barcode_evidence"] = {
            "source": NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT,
            "endpoint": str(card.get("endpoint") or "/content/v2/get/cards/list"),
            "nm_id": card_nm_id,
            "result": "resolved",
            "selected_primary": barcodes[0],
            "barcode_count": len(barcodes),
            "sync_reason": "wb_card_sync",
        }
    elif manual_barcode:
        updated["_manual_barcode_preserved"] = True
    updated["vendor_code"] = str(card.get("vendor_code") or "")
    updated["seller_article"] = str(card.get("vendor_code") or "")
    updated["wb_title"] = str(card.get("title") or "")
    updated["wb_subject_name"] = str(card.get("subject_name") or "")
    updated["wb_updated_at"] = str(card.get("updated_at") or "")
    updated["wb_synced_at"] = synced_at
    updated["wb_sync_status"] = (
        "needs_review"
        if str(existing.get("product_type") or "") == "other" and str(existing.get("wb_sync_status") or "") == "needs_review"
        else f"matched_{match_type or 'unknown'}"
    )
    updated["wb_sync_evidence"] = _wb_card_sync_evidence(card, result="matched", match_type=match_type)
    updated["updated_at"] = synced_at
    return updated


def _new_nomenclature_item_from_wb_card(
    card: Mapping[str, Any],
    *,
    groups: list[Mapping[str, Any]],
    created_at: str,
    updated_at: str,
) -> dict[str, Any]:
    detection = _detect_sku_group_from_vendor_code(str(card.get("vendor_code") or ""), groups)
    group_key = str(detection.get("group_key") or "other")
    needs_review = not bool(detection.get("group_key"))
    name = str(card.get("vendor_code") or card.get("title") or card.get("nm_id") or "").strip()
    barcodes = sorted(_normalize_barcode_list(card.get("barcodes") or []))
    barcode = barcodes[0] if barcodes else ""
    barcode_status = NOMENCLATURE_BARCODE_STATUS_MULTIPLE if len(barcodes) > 1 else (
        NOMENCLATURE_BARCODE_STATUS_READY if barcode else NOMENCLATURE_BARCODE_STATUS_MISSING
    )
    match_suffix = _vendor_code_match_suffix(str(card.get("vendor_code") or name))
    evidence = _wb_card_sync_evidence(
        card,
        result="created",
        match_type="none",
        detected_group=detection,
        needs_review=needs_review,
    )
    return {
        "item_id": "nom_wb_" + uuid4().hex,
        "is_active": not needs_review and group_key not in {"extra", "other"},
        "is_hidden": False,
        "hidden_at": "",
        "hidden_reason": "",
        "our_sku": "",
        "nm_id": _optional_int(card.get("nm_id")),
        "barcode": barcode,
        "primary_barcode": barcode,
        "barcodes": barcodes,
        "barcode_source": NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT if barcode else NOMENCLATURE_BARCODE_SOURCE_MISSING,
        "barcode_status": barcode_status,
        "barcode_synced_at": updated_at if barcode else "",
        "barcode_updated_at": updated_at if barcode else "",
        "barcode_evidence": {
            "source": NOMENCLATURE_BARCODE_SOURCE_WB_CONTENT,
            "endpoint": str(card.get("endpoint") or "/content/v2/get/cards/list"),
            "nm_id": _optional_int(card.get("nm_id")),
            "result": "resolved" if barcode else "missing",
            "selected_primary": barcode,
            "barcode_count": len(barcodes),
            "sync_reason": "wb_card_sync",
        },
        "vendor_code": str(card.get("vendor_code") or ""),
        "seller_article": str(card.get("vendor_code") or ""),
        "wb_title": str(card.get("title") or ""),
        "wb_subject_name": str(card.get("subject_name") or ""),
        "wb_updated_at": str(card.get("updated_at") or ""),
        "wb_synced_at": updated_at,
        "wb_sync_status": "needs_review" if needs_review else "created",
        "wb_sync_evidence": evidence,
        "nomenclature_name": name,
        "product_type": group_key,
        "match_key": f"{group_key}|{match_suffix}" if match_suffix and group_key not in {"extra", "other"} else "",
        "purchase_price_yuan": None,
        "aliases": [],
        "compatible_models_text": "",
        "compatible_model_keys": [],
        "comment": "needs_review: WB vendorCode group is not recognized" if needs_review else "created from WB Content sync",
        "created_at": created_at,
        "updated_at": updated_at,
    }


def _detect_sku_group_from_vendor_code(vendor_code: str, groups: list[Mapping[str, Any]]) -> dict[str, Any]:
    normalized_vendor = _normalize_vendor_code(vendor_code)
    if not normalized_vendor:
        return {}
    candidates: list[tuple[int, int, str, str, str]] = []
    for group in groups:
        if not bool(group.get("is_active", True)):
            continue
        group_key = str(group.get("group_key") or "").strip()
        for alias in group.get("aliases") or []:
            normalized_alias = _normalize_vendor_code(alias)
            if not normalized_alias:
                continue
            if _normalized_alias_matches(normalized_vendor, normalized_alias):
                candidates.append((len(normalized_alias), normalized_alias.count(" "), group_key, str(group.get("label") or group_key), str(alias)))
    if not candidates:
        return {}
    candidates.sort(reverse=True)
    _, _, group_key, label, alias = candidates[0]
    return {"group_key": group_key, "label": label, "matched_alias": alias, "confidence": "high"}


def _normalized_alias_matches(normalized_vendor: str, normalized_alias: str) -> bool:
    if not normalized_alias:
        return False
    return f" {normalized_alias} " in f" {normalized_vendor} " or normalized_alias in normalized_vendor


def _normalize_vendor_code(value: Any) -> str:
    normalized = str(value or "").strip().casefold().replace("ё", "е")
    normalized = normalized.replace("–", "-").replace("—", "-").replace("−", "-")
    normalized = re.sub(r"[-_/]+", " ", normalized)
    normalized = re.sub(r"\bnoframe\b", "no frame", normalized)
    normalized = re.sub(r"\bno\s*frame\b", "no frame", normalized)
    normalized = re.sub(r"\bantispy\b", "anti spy", normalized)
    normalized = re.sub(r"\banti\s*spy\b", "anti spy", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _vendor_code_match_suffix(value: str) -> str:
    normalized = _normalize_vendor_code(value)
    normalized = re.sub(r"[^a-z0-9а-я]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized[:96]


def _wb_card_sync_evidence(
    card: Mapping[str, Any],
    *,
    result: str,
    match_type: str,
    detected_group: Mapping[str, Any] | None = None,
    needs_review: bool = False,
) -> dict[str, Any]:
    return {
        "source": WB_CARD_SYNC_SOURCE,
        "endpoint": str(card.get("endpoint") or "/content/v2/get/cards/list"),
        "result": result,
        "match_type": match_type,
        "nm_id": _optional_int(card.get("nm_id")),
        "vendor_code": str(card.get("vendor_code") or ""),
        "barcode_count": len(_normalize_barcode_list(card.get("barcodes") or [])),
        "detected_group": dict(detected_group or {}),
        "needs_review": bool(needs_review),
    }


def _nomenclature_wb_sync_error_result(
    *,
    status: str,
    error: Exception,
    summary: Mapping[str, Any],
    limit: int,
    max_pages: int,
) -> dict[str, Any]:
    return {
        "contract_name": "sheet_vitrina_v1_nomenclature_wb_sync",
        "status": status,
        "limit": limit,
        "max_pages": max_pages,
        "cards_processed": 0,
        "created": 0,
        "updated": 0,
        "error": _safe_barcode_error(error),
        "summary": dict(summary),
        "items": [],
    }


def _normalize_match_key(value: Any) -> str:
    normalized = str(value or "").strip().lower()
    normalized = normalized.replace(" ", "_")
    normalized = re.sub(r"_+", "_", normalized)
    normalized = normalized.strip("_")
    if normalized and "|" not in normalized:
        raise ValueError("nomenclature match_key must use product_type|normalized_model")
    return normalized


def _normalize_alias_list(value: Any) -> list[str]:
    if isinstance(value, str):
        raw_items = re.split(r"[\n,;]+", value)
    elif isinstance(value, list):
        raw_items = [str(item) for item in value]
    else:
        raw_items = []
    aliases: list[str] = []
    seen: set[str] = set()
    for item in raw_items:
        alias = str(item or "").strip()
        if not alias or alias in seen:
            continue
        seen.add(alias)
        aliases.append(alias)
    return aliases


def _normalize_compatible_model_keys(
    value: Any,
    *,
    fallback_text: str = "",
    item_hint: Mapping[str, Any] | None = None,
) -> list[str]:
    raw_items: list[str]
    if isinstance(value, str):
        raw_items = re.split(r"[\n,;]+", value)
    elif isinstance(value, list):
        raw_items = [str(item) for item in value]
    else:
        raw_items = []
    keys: list[str] = []
    seen: set[str] = set()
    for raw_item in raw_items:
        for key in extract_iphone_model_keys(raw_item):
            if key not in seen:
                seen.add(key)
                keys.append(key)
    for source in [fallback_text]:
        for key in extract_iphone_model_keys(source):
            if key not in seen:
                seen.add(key)
                keys.append(key)
    if not keys and item_hint:
        for key in _infer_compatible_model_keys(item_hint):
            if key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def _nomenclature_import_header_keys(header_row: tuple[Any, ...]) -> list[str]:
    header_aliases = {
        "id строки": "item_id",
        "id": "item_id",
        "item_id": "item_id",
        "включено": "is_active",
        "active": "is_active",
        "is_active": "is_active",
        "скрыто": "is_hidden",
        "hidden": "is_hidden",
        "is_hidden": "is_hidden",
        "nmid": "nm_id",
        "nm id": "nm_id",
        "nm_id": "nm_id",
        "шк": "barcode",
        "штрихкод": "barcode",
        "штрихкод wb": "barcode",
        "шк / barcode": "barcode",
        "barcode": "barcode",
        "primary_barcode": "barcode",
        "все шк": "barcodes",
        "barcodes": "barcodes",
        "barcode_list": "barcodes",
        "источник шк": "barcode_source",
        "barcode_source": "barcode_source",
        "статус шк": "barcode_status",
        "barcode_status": "barcode_status",
        "артикул продавца wb / vendorcode": "vendor_code",
        "артикул продавца wb": "vendor_code",
        "vendorcode": "vendor_code",
        "vendor_code": "vendor_code",
        "seller_article": "vendor_code",
        "название wb": "wb_title",
        "wb_title": "wb_title",
        "wb subject": "wb_subject_name",
        "wb_subject_name": "wb_subject_name",
        "wb updatedat": "wb_updated_at",
        "wb_updated_at": "wb_updated_at",
        "статус wb sync": "wb_sync_status",
        "wb_sync_status": "wb_sync_status",
        "номенклатура": "nomenclature_name",
        "nomenclature": "nomenclature_name",
        "nomenclature_name": "nomenclature_name",
        "тип": "product_type",
        "группа": "product_type",
        "product_type": "product_type",
        "match key": "match_key",
        "match_key": "match_key",
        "цена закупки, ¥": "purchase_price_yuan",
        "цена закупки": "purchase_price_yuan",
        "purchase_price_yuan": "purchase_price_yuan",
        "совместимые модели": "compatible_models_text",
        "compatible_models_text": "compatible_models_text",
        "ключи совместимости": "compatible_model_keys",
        "compatible_model_keys": "compatible_model_keys",
        "обновлено": "updated_at",
        "updated_at": "updated_at",
    }
    return [header_aliases.get(_normalize_nomenclature_header(value), "") for value in header_row]


def _normalize_nomenclature_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold().replace("ё", "е"))


def _nomenclature_import_row_empty(row_values: Mapping[str, Any]) -> bool:
    for key, value in row_values.items():
        if key == "updated_at":
            continue
        if _cell_text(value):
            return False
    return True


def _normalize_nomenclature_import_row(
    row_values: Mapping[str, Any],
    *,
    row_number: int,
    existing_by_id: Mapping[str, dict[str, Any]],
    active_by_match_key: Mapping[str, list[dict[str, Any]]],
    now: str,
) -> dict[str, Any] | None:
    raw_item_id = _cell_text(row_values.get("item_id"))
    existing: dict[str, Any] | None = None
    if raw_item_id:
        existing = existing_by_id.get(raw_item_id)
        if existing is None:
            raise ValueError(f"Строка {row_number}: ID строки не найден: {raw_item_id}")
    raw_match_key = _normalize_match_key(row_values.get("match_key")) if "match_key" in row_values else ""
    if existing is None and raw_match_key:
        candidates = active_by_match_key.get(raw_match_key, [])
        if len(candidates) == 1:
            existing = candidates[0]
        elif len(candidates) > 1:
            raise ValueError(f"Строка {row_number}: match key неоднозначен: {raw_match_key}")

    base = dict(existing) if existing is not None else {}
    is_active = (
        _parse_nomenclature_bool(row_values.get("is_active"), default=bool(base.get("is_active", True)))
        if "is_active" in row_values
        else bool(base.get("is_active", True))
    )
    is_hidden = (
        _parse_nomenclature_bool(row_values.get("is_hidden"), default=bool(base.get("is_hidden", False)))
        if "is_hidden" in row_values
        else bool(base.get("is_hidden", False))
    )
    product_type = (
        _parse_nomenclature_product_type(row_values.get("product_type"))
        if "product_type" in row_values and _cell_text(row_values.get("product_type"))
        else str(base.get("product_type") or "")
    )
    if not product_type:
        raise ValueError(f"Строка {row_number}: Тип обязателен")
    nomenclature_name = (
        _cell_text(row_values.get("nomenclature_name"))
        if "nomenclature_name" in row_values
        else str(base.get("nomenclature_name") or "")
    )
    match_key = raw_match_key if "match_key" in row_values else str(base.get("match_key") or "")
    nm_id = _parse_nomenclature_nm_id(row_values.get("nm_id"), row_number=row_number) if "nm_id" in row_values else base.get("nm_id")
    if "barcode" in row_values:
        barcode = _normalize_barcode(row_values.get("barcode"))
        barcode_payload = _prepare_nomenclature_barcode_payload(
            existing=base,
            payload={
                "barcode": barcode,
                "barcodes": _cell_text(row_values.get("barcodes")) if "barcodes" in row_values else base.get("barcodes") or [],
            },
            updated_at=now,
        )
    else:
        barcode_payload = {
            "barcode": base.get("barcode") or "",
            "barcodes": base.get("barcodes") or [],
            "barcode_source": base.get("barcode_source") or "",
            "barcode_status": base.get("barcode_status") or "",
            "barcode_synced_at": base.get("barcode_synced_at") or "",
            "barcode_updated_at": base.get("barcode_updated_at") or "",
            "barcode_evidence": base.get("barcode_evidence") or {},
        }
    compatible_models_text = (
        _cell_text(row_values.get("compatible_models_text"))
        if "compatible_models_text" in row_values
        else str(base.get("compatible_models_text") or "")
    )
    if "compatible_model_keys" in row_values and _cell_text(row_values.get("compatible_model_keys")):
        compatible_model_keys_source: Any = _cell_text(row_values.get("compatible_model_keys"))
    elif "compatible_model_keys" in row_values:
        compatible_model_keys_source = []
    else:
        compatible_model_keys_source = base.get("compatible_model_keys") or []
    if "purchase_price_yuan" in row_values:
        try:
            purchase_price_yuan = _optional_nonnegative_number(
                row_values.get("purchase_price_yuan"),
                field_name="Цена закупки, ¥",
            )
        except ValueError as exc:
            raise ValueError(f"Строка {row_number}: {exc}") from exc
    else:
        purchase_price_yuan = base.get("purchase_price_yuan")

    item_id = str(base.get("item_id") or raw_item_id or ("nom_" + uuid4().hex))
    created_at = str(base.get("created_at") or now)
    item = _normalize_nomenclature_payload(
        {
            **base,
            "is_active": is_active,
            "is_hidden": is_hidden,
            "hidden_at": base.get("hidden_at") or "",
            "hidden_reason": base.get("hidden_reason") or "",
            "nm_id": nm_id,
            **barcode_payload,
            "vendor_code": _cell_text(row_values.get("vendor_code")) if "vendor_code" in row_values else base.get("vendor_code") or "",
            "wb_title": _cell_text(row_values.get("wb_title")) if "wb_title" in row_values else base.get("wb_title") or "",
            "wb_subject_name": (
                _cell_text(row_values.get("wb_subject_name"))
                if "wb_subject_name" in row_values
                else base.get("wb_subject_name") or ""
            ),
            "wb_updated_at": _cell_text(row_values.get("wb_updated_at")) if "wb_updated_at" in row_values else base.get("wb_updated_at") or "",
            "wb_synced_at": base.get("wb_synced_at") or "",
            "wb_sync_status": _cell_text(row_values.get("wb_sync_status")) if "wb_sync_status" in row_values else base.get("wb_sync_status") or "",
            "wb_sync_evidence": base.get("wb_sync_evidence") or {},
            "nomenclature_name": nomenclature_name,
            "product_type": product_type,
            "match_key": match_key,
            "purchase_price_yuan": purchase_price_yuan,
            "compatible_models_text": compatible_models_text,
            "compatible_model_keys": compatible_model_keys_source,
        },
        item_id=item_id,
        created_at=created_at,
        updated_at=now,
    )
    if existing is not None and not _nomenclature_item_changed(existing, item):
        return None
    action = "created"
    if existing is not None:
        action = "deactivated" if bool(existing.get("is_active")) and not bool(item.get("is_active")) else "updated"
    return {"row": row_number, "action": action, "item": item}


def _parse_nomenclature_bool(value: Any, *, default: bool) -> bool:
    if value is None or _cell_text(value) == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value) == 1.0:
            return True
        if float(value) == 0.0:
            return False
    normalized = _cell_text(value).casefold()
    if normalized in {"да", "true", "1", "yes", "y", "on"}:
        return True
    if normalized in {"нет", "false", "0", "no", "n", "off"}:
        return False
    raise ValueError("Включено должно быть да/нет, true/false или 1/0")


def _parse_nomenclature_product_type(value: Any) -> str:
    normalized = _cell_text(value).casefold().replace("ё", "е")
    if normalized in NOMENCLATURE_PRODUCT_TYPE_BY_LABEL:
        return NOMENCLATURE_PRODUCT_TYPE_BY_LABEL[normalized]
    normalized_key = _normalize_sku_group_key(normalized)
    if _valid_sku_group_key(normalized_key):
        return normalized_key
    normalized_no_dot = normalized.replace(".", "")
    if normalized_no_dot in {"доп строка", "дополнительная строка"}:
        return "extra"
    raise ValueError("Группа должна быть stable key или label из справочника групп SKU")


def _parse_nomenclature_nm_id(value: Any, *, row_number: int) -> int | None:
    if value is None or _cell_text(value) == "":
        return None
    parsed = _optional_int(value)
    if parsed is None:
        raise ValueError(f"Строка {row_number}: nmId должен быть целым числом")
    return parsed


def _optional_nonnegative_number(value: Any, *, field_name: str) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and not value.strip():
        return None
    parsed = _optional_number(value)
    if parsed is None or parsed < 0:
        raise ValueError(f"{field_name} должна быть числом >= 0")
    return parsed


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _nomenclature_item_changed(existing: Mapping[str, Any], item: Mapping[str, Any]) -> bool:
    keys = [
        "is_active",
        "is_hidden",
        "hidden_at",
        "hidden_reason",
        "our_sku",
        "nm_id",
        "barcode",
        "barcodes",
        "barcode_source",
        "barcode_status",
        "barcode_synced_at",
        "barcode_updated_at",
        "vendor_code",
        "wb_title",
        "wb_subject_name",
        "wb_updated_at",
        "wb_synced_at",
        "wb_sync_status",
        "nomenclature_name",
        "product_type",
        "match_key",
        "purchase_price_yuan",
        "aliases",
        "compatible_models_text",
        "compatible_model_keys",
        "comment",
    ]
    for key in keys:
        if key == "purchase_price_yuan":
            left = _optional_number(existing.get(key))
            right = _optional_number(item.get(key))
        else:
            left = existing.get(key)
            right = item.get(key)
        if left != right:
            return True
    return False


def _nomenclature_import_duplicate_errors(
    existing_items: list[dict[str, Any]],
    operations: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    projected: dict[str, set[str]] = {}
    for item in existing_items:
        if bool(item.get("is_active")) and str(item.get("match_key") or "").strip():
            projected.setdefault(str(item.get("match_key") or "").strip(), set()).add(str(item.get("item_id") or ""))
    for operation in operations:
        item = operation["item"]
        item_id = str(item.get("item_id") or "")
        for match_key in list(projected.keys()):
            projected[match_key].discard(item_id)
            if not projected[match_key]:
                del projected[match_key]
        if bool(item.get("is_active")) and str(item.get("match_key") or "").strip():
            projected.setdefault(str(item.get("match_key") or "").strip(), set()).add(item_id)
    duplicate_match_keys = {match_key for match_key, item_ids in projected.items() if len(item_ids) > 1}
    errors: list[dict[str, Any]] = []
    for operation in operations:
        item = operation["item"]
        match_key = str(item.get("match_key") or "").strip()
        if bool(item.get("is_active")) and match_key in duplicate_match_keys:
            errors.append(
                {
                    "row": operation["row"],
                    "message": f"Строка {operation['row']}: duplicate active match key: {match_key}",
                }
            )
    return errors


def _nomenclature_import_result(
    *,
    status: str,
    dry_run: bool,
    operations: list[dict[str, Any]],
    skipped_count: int,
    errors: list[dict[str, Any]],
    items: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "contract_name": "sheet_vitrina_v1_nomenclature_import",
        "status": status,
        "dry_run": dry_run,
        "created_count": sum(1 for operation in operations if operation["action"] == "created"),
        "updated_count": sum(1 for operation in operations if operation["action"] == "updated"),
        "deactivated_count": sum(1 for operation in operations if operation["action"] == "deactivated"),
        "skipped_count": skipped_count,
        "error_count": len(errors),
        "errors": errors,
        "items": items,
    }


def _line_compatible_model_keys(line: Mapping[str, Any]) -> list[str]:
    parts = [
        line.get("model_raw"),
        line.get("model_normalized"),
        str(line.get("match_key") or "").split("|", 1)[1] if "|" in str(line.get("match_key") or "") else "",
    ]
    keys: list[str] = []
    seen: set[str] = set()
    for part in parts:
        for key in extract_iphone_model_keys(part):
            if key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def _infer_compatible_model_keys(item: Mapping[str, Any]) -> list[str]:
    raw_keys = item.get("compatible_model_keys")
    if isinstance(raw_keys, list):
        keys = [str(key).strip() for key in raw_keys if str(key or "").strip()]
        if keys:
            return _dedupe(keys)
    keys: list[str] = []
    seen: set[str] = set()
    sources: list[Any] = [
        item.get("compatible_models_text"),
        str(item.get("match_key") or "").split("|", 1)[1] if "|" in str(item.get("match_key") or "") else "",
        item.get("nomenclature_name"),
    ]
    sources.extend(item.get("aliases") or [])
    for source in sources:
        for key in extract_iphone_model_keys(source):
            if key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def _nomenclature_item_match_payload(item: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "active": True,
        "item_id": str(item.get("item_id") or ""),
        "product_type": str(item.get("product_type") or ""),
        "factory_type": str(item.get("product_type") or ""),
        "internal_sku": str(item.get("our_sku") or item.get("internal_sku") or ""),
        "internal_nm_id": _optional_int(item.get("nm_id") or item.get("internal_nm_id")),
        "internal_name": str(item.get("nomenclature_name") or item.get("internal_name") or ""),
        "nomenclature_name": str(item.get("nomenclature_name") or item.get("internal_name") or ""),
        "match_key": str(item.get("match_key") or ""),
        "purchase_price_yuan": item.get("purchase_price_yuan"),
        "compatible_model_keys": _infer_compatible_model_keys(item),
        "group": "nomenclature",
    }


def _compatible_models_text_from_keys(keys: list[str]) -> str:
    return ", ".join(_model_key_to_label(key) for key in keys)


def _model_key_to_label(key: str) -> str:
    normalized = str(key or "").strip()
    if not normalized.startswith("iphone_"):
        return normalized
    parts = normalized.removeprefix("iphone_").split("_")
    if not parts:
        return normalized
    number = parts[0]
    suffix = " ".join(part.capitalize() if part != "e" else "e" for part in parts[1:])
    return ("iPhone " + number + (" " + suffix if suffix else "")).strip()


def _dedupe(items: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for item in items:
        normalized = str(item or "").strip()
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return result


def _nomenclature_item_aliases(item: Mapping[str, Any]) -> list[dict[str, Any]]:
    if not bool(item.get("is_active")):
        return []
    base_match_key = str(item.get("match_key") or "").strip()
    payload_base = {
        "active": True,
        "product_type": str(item.get("product_type") or ""),
        "factory_type": str(item.get("product_type") or ""),
        "internal_sku": str(item.get("our_sku") or ""),
        "internal_nm_id": _optional_int(item.get("nm_id")),
        "internal_name": str(item.get("nomenclature_name") or ""),
        "nomenclature_name": str(item.get("nomenclature_name") or ""),
        "item_id": str(item.get("item_id") or ""),
        "purchase_price_yuan": item.get("purchase_price_yuan"),
        "group": "nomenclature",
    }
    aliases: list[dict[str, Any]] = []
    if base_match_key:
        aliases.append({**payload_base, "match_key": base_match_key})
    product_type = str(item.get("product_type") or "").strip()
    for raw_alias in item.get("aliases") or []:
        alias_text = str(raw_alias or "").strip()
        if not alias_text:
            continue
        if "|" in alias_text:
            aliases.append({**payload_base, "match_key": _normalize_match_key(alias_text)})
            continue
        normalized_model = normalize_invoice_model(alias_text)
        if product_type and normalized_model:
            aliases.append(
                {
                    **payload_base,
                    "normalized_model": normalized_model,
                    "match_key": f"{product_type}|{normalized_model}",
                }
            )
    return aliases


def _product_type_from_config_item(display_name: str, group: str) -> str:
    text = f"{group} {display_name}".lower()
    if "anti" in text and "spy" in text:
        return "anti_spy"
    if "matte" in text:
        return "matte"
    if "clean" in text or "clear" in text:
        return "clear"
    return ""


def _model_text_from_nomenclature_name(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^\s*(no\s*frame\s*)?(clean|clear|matte|anti[-\s]?spy)\s+", "", text, flags=re.IGNORECASE)
    return text.strip()


def _extract_contract_xlsx_lines(file_bytes: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        worksheet = workbook.worksheets[0]
        lines: list[str] = []
        for row in worksheet.iter_rows(min_row=1, max_row=16, max_col=8, values_only=True):
            parts = [_normalize_text_line(cell) for cell in row if _normalize_text_line(cell)]
            if parts:
                lines.append(" ".join(parts))
        return lines
    finally:
        workbook.close()


def _extract_contract_pdf_lines(file_bytes: bytes) -> tuple[list[str], list[str], str, dict[str, Any]]:
    warnings: list[str] = []
    diagnostics: dict[str, Any] = {}
    if shutil.which("pdftotext"):
        text = _extract_pdf_text_with_pdftotext(file_bytes)
        lines = _text_to_lines(text)
        if lines:
            diagnostics["text_layer_extraction"] = "pdftotext"
            return lines, warnings, "pdf_pdftotext", diagnostics
        diagnostics["pdftotext_text_nonempty"] = False

    text = _extract_pdf_text_layer(file_bytes)
    lines = _text_to_lines(text)
    if lines:
        diagnostics["text_layer_extraction"] = "embedded_pdf_streams"
        return lines, warnings, "pdf_text_layer", diagnostics

    ocr_lines, ocr_warning, ocr_diagnostics = _extract_contract_pdf_ocr_lines(file_bytes)
    diagnostics.update(ocr_diagnostics)
    if ocr_lines:
        return ocr_lines, warnings, "pdf_ocr_first_page", diagnostics
    if ocr_warning:
        warnings.append(ocr_warning)
    return [], warnings, "pdf_no_readable_text", diagnostics


def _extract_pdf_text_with_pdftotext(file_bytes: bytes) -> str:
    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as handle:
            handle.write(file_bytes)
            temp_path = handle.name
        result = subprocess.run(
            ["pdftotext", "-f", "1", "-l", "1", "-layout", temp_path, "-"],
            capture_output=True,
            timeout=8,
            check=False,
        )
        if result.returncode != 0:
            return ""
        return result.stdout.decode("utf-8", "replace")
    except (OSError, subprocess.SubprocessError):
        return ""
    finally:
        if temp_path:
            try:
                Path(temp_path).unlink()
            except OSError:
                pass


def _extract_pdf_text_layer(file_bytes: bytes) -> str:
    fragments: list[str] = []
    for stream_dict, stream_bytes in _iter_pdf_streams(file_bytes):
        filters = _pdf_stream_filters(stream_dict)
        if "DCTDecode" in filters or b"Image" in stream_dict:
            continue
        payload = stream_bytes
        if "FlateDecode" in filters:
            try:
                payload = zlib.decompress(stream_bytes)
            except zlib.error:
                continue
        if not any(marker in payload for marker in (b"Tj", b"TJ", b"'", b'"')):
            continue
        fragments.extend(_extract_pdf_text_fragments(payload))
        if len(fragments) >= 80:
            break
    return "\n".join(fragment for fragment in fragments if _normalize_text_line(fragment))


def _iter_pdf_streams(file_bytes: bytes) -> list[tuple[bytes, bytes]]:
    raw = file_bytes[:8_000_000]
    streams: list[tuple[bytes, bytes]] = []
    for match in re.finditer(rb"<<(?P<dict>.*?)>>\s*stream\r?\n(?P<body>.*?)\r?\nendstream", raw, flags=re.S):
        streams.append((match.group("dict"), match.group("body")))
    return streams


def _pdf_stream_filters(stream_dict: bytes) -> set[str]:
    filters: set[str] = set()
    single = re.search(rb"/Filter\s*/([A-Za-z0-9]+)", stream_dict)
    if single:
        filters.add(single.group(1).decode("ascii", "ignore"))
    for match in re.finditer(rb"/([A-Za-z0-9]+)", stream_dict):
        name = match.group(1).decode("ascii", "ignore")
        if name.endswith("Decode"):
            filters.add(name)
    return filters


def _extract_pdf_text_fragments(content: bytes) -> list[str]:
    fragments: list[str] = []
    for array_match in re.finditer(rb"\[(?P<body>.*?)\]\s*TJ", content, flags=re.S):
        fragments.extend(_decode_pdf_string_tokens(array_match.group("body")))
    for literal_match in re.finditer(rb"(?P<token>\((?:\\.|[^\\()])*\))\s*(?:Tj|'|\")", content, flags=re.S):
        decoded = _decode_pdf_literal_string(literal_match.group("token"))
        if decoded:
            fragments.append(decoded)
    return [_normalize_text_line(fragment) for fragment in fragments if _normalize_text_line(fragment)]


def _decode_pdf_string_tokens(value: bytes) -> list[str]:
    fragments: list[str] = []
    for literal_match in re.finditer(rb"\((?:\\.|[^\\()])*\)", value, flags=re.S):
        decoded = _decode_pdf_literal_string(literal_match.group(0))
        if decoded:
            fragments.append(decoded)
    for hex_match in re.finditer(rb"<([0-9A-Fa-f\s]+)>", value):
        raw_hex = re.sub(rb"\s+", b"", hex_match.group(1))
        if len(raw_hex) % 2:
            raw_hex += b"0"
        try:
            decoded = _decode_pdf_bytes(bytes.fromhex(raw_hex.decode("ascii")))
        except ValueError:
            decoded = ""
        if decoded:
            fragments.append(decoded)
    return fragments


def _decode_pdf_literal_string(token: bytes) -> str:
    body = token[1:-1]
    result = bytearray()
    index = 0
    while index < len(body):
        char = body[index]
        if char != 0x5C:
            result.append(char)
            index += 1
            continue
        index += 1
        if index >= len(body):
            break
        escaped = body[index]
        if escaped in b"nrtbf":
            result.extend({ord("n"): b"\n", ord("r"): b"\r", ord("t"): b"\t", ord("b"): b"\b", ord("f"): b"\f"}[escaped])
            index += 1
            continue
        if escaped in b"()\\":
            result.append(escaped)
            index += 1
            continue
        if 48 <= escaped <= 55:
            octal = bytes([escaped])
            index += 1
            for _ in range(2):
                if index < len(body) and 48 <= body[index] <= 55:
                    octal += bytes([body[index]])
                    index += 1
            result.append(int(octal, 8))
            continue
        result.append(escaped)
        index += 1
    return _decode_pdf_bytes(bytes(result))


def _decode_pdf_bytes(value: bytes) -> str:
    if not value:
        return ""
    if value.startswith(b"\xfe\xff"):
        return value[2:].decode("utf-16-be", "ignore")
    if value.startswith(b"\xff\xfe"):
        return value[2:].decode("utf-16-le", "ignore")
    if value.count(b"\x00") > max(2, len(value) // 4):
        decoded = value.decode("utf-16-be", "ignore")
        if _printable_text_score(decoded) > 0.5:
            return decoded
    best = ""
    best_score = -1.0
    for encoding in ("utf-8", "cp1251", "latin-1"):
        decoded = value.decode(encoding, "ignore")
        score = _printable_text_score(decoded)
        if score > best_score:
            best = decoded
            best_score = score
    return best


def _printable_text_score(value: str) -> float:
    if not value:
        return 0.0
    useful = sum(1 for char in value if char.isalnum() or char.isspace() or char in "№#./:-_«»(),")
    return useful / max(1, len(value))


def _extract_contract_pdf_ocr_lines(file_bytes: bytes) -> tuple[list[str], str, dict[str, Any]]:
    diagnostics: dict[str, Any] = {
        "ocr_attempted": True,
        "ocr_available": bool(shutil.which("pdftoppm") and shutil.which("tesseract")),
        "ocr_engine": "tesseract",
        "ocr_text_nonempty": False,
        "ocr_attempt_count": 0,
    }
    if not shutil.which("pdftoppm") or not shutil.which("tesseract"):
        return [], "contract parser skipped OCR: OCR tools missing (pdftoppm/tesseract)", diagnostics
    languages = _preferred_tesseract_languages()
    diagnostics["ocr_languages"] = languages.split("+") if languages else []
    best_lines: list[str] = []
    best_score = -1
    best_strategy = ""
    failure_kinds: set[str] = set()
    try:
        with tempfile.TemporaryDirectory(prefix="wb-contract-ocr-") as tmp:
            pdf_path = Path(tmp) / "contract.pdf"
            pdf_path.write_bytes(file_bytes)
            for strategy in CONTRACT_PDF_OCR_STRATEGIES:
                image_prefix = Path(tmp) / str(strategy["id"])
                render_result = _render_pdf_first_page_for_ocr(pdf_path, image_prefix=image_prefix, strategy=strategy)
                if not render_result.get("ok"):
                    diagnostics["ocr_attempt_count"] = int(diagnostics["ocr_attempt_count"]) + 1
                    failure_kinds.add(str(render_result.get("reason") or "render_failed"))
                    continue
                image_path = Path(str(image_prefix) + ".png")
                for psm in CONTRACT_OCR_PSMS:
                    diagnostics["ocr_attempt_count"] = int(diagnostics["ocr_attempt_count"]) + 1
                    strategy_id = f"{strategy['id']}_psm{psm}"
                    lines, warning, attempt_diagnostics = _run_tesseract_image(
                        image_path,
                        psm=psm,
                        languages=languages,
                        strategy_id=strategy_id,
                        timeout_seconds=float(strategy.get("ocr_timeout") or 60),
                    )
                    if warning:
                        failure_kinds.add(warning)
                    score_payload = _score_contract_metadata_lines(lines)
                    score = int(score_payload["score"])
                    if score > best_score:
                        best_score = score
                        best_lines = lines
                        best_strategy = strategy_id
                        diagnostics.update(attempt_diagnostics)
                        diagnostics.update(
                            {
                                "ocr_strategy_used": strategy_id,
                                "ocr_text_nonempty": bool(lines),
                                "ocr_best_number_found": bool(score_payload["number_found"]),
                                "ocr_best_date_found": bool(score_payload["date_found"]),
                            }
                        )
                    if lines and score_payload["number_found"] and score_payload["date_found"]:
                        return lines, "", diagnostics
            if best_lines:
                diagnostics["ocr_strategy_used"] = best_strategy
                diagnostics["ocr_text_nonempty"] = True
                return (
                    best_lines,
                    "contract parser OCR text exists but number/date pattern was not fully found",
                    diagnostics,
                )
            diagnostics["ocr_failure_kinds"] = sorted(failure_kinds)[:8]
            return [], "contract parser OCR text empty after bounded first-page strategies", diagnostics
    except (OSError, subprocess.SubprocessError):
        return [], "contract parser skipped OCR: OCR command failed", diagnostics


def _render_pdf_first_page_for_ocr(pdf_path: Path, *, image_prefix: Path, strategy: Mapping[str, Any]) -> dict[str, Any]:
    command = [
        "pdftoppm",
        "-f",
        "1",
        "-singlefile",
        "-r",
        str(int(strategy.get("dpi") or 150)),
        "-png",
    ]
    if strategy.get("gray"):
        command.append("-gray")
    crop = str(strategy.get("crop") or "")
    if crop == "top":
        crop_pixels = _pdf_first_page_top_crop_pixels(pdf_path, dpi=int(strategy.get("dpi") or 150))
        if crop_pixels:
            command.extend(["-x", "0", "-y", "0", "-W", str(crop_pixels[0]), "-H", str(crop_pixels[1])])
    command.extend([str(pdf_path), str(image_prefix)])
    try:
        result = subprocess.run(
            command,
            capture_output=True,
            timeout=float(strategy.get("render_timeout") or 60),
            check=False,
        )
    except subprocess.TimeoutExpired:
        return {"ok": False, "reason": "pdf_render_timeout"}
    except OSError:
        return {"ok": False, "reason": "pdf_render_os_error"}
    image_path = Path(str(image_prefix) + ".png")
    if result.returncode != 0 or not image_path.exists():
        return {"ok": False, "reason": "pdf_render_failed"}
    return {"ok": True}


def _pdf_first_page_top_crop_pixels(pdf_path: Path, *, dpi: int) -> tuple[int, int] | None:
    if not shutil.which("pdfinfo"):
        return None
    try:
        result = subprocess.run(
            ["pdfinfo", str(pdf_path)],
            capture_output=True,
            text=True,
            timeout=8,
            check=False,
        )
    except (OSError, subprocess.SubprocessError):
        return None
    if result.returncode != 0:
        return None
    match = re.search(r"Page size:\s*([0-9.]+)\s+x\s+([0-9.]+)\s+pts", result.stdout)
    if not match:
        return None
    width = max(1, int(float(match.group(1)) * dpi / 72))
    height = max(1, int(float(match.group(2)) * dpi / 72 * 0.45))
    return width, height


def _extract_contract_image_lines(file_bytes: bytes, *, suffix: str) -> tuple[list[str], list[str], str, dict[str, Any]]:
    diagnostics: dict[str, Any] = {
        "ocr_attempted": True,
        "ocr_available": bool(shutil.which("tesseract")),
        "ocr_engine": "tesseract",
        "ocr_text_nonempty": False,
        "ocr_attempt_count": 0,
    }
    if not shutil.which("tesseract"):
        return [], ["contract parser skipped OCR: OCR tool missing (tesseract)"], "image_no_ocr", diagnostics
    languages = _preferred_tesseract_languages()
    diagnostics["ocr_languages"] = languages.split("+") if languages else []
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as handle:
            handle.write(file_bytes)
            image_path = Path(handle.name)
        try:
            best_lines: list[str] = []
            best_score = -1
            best_warning = ""
            for psm in CONTRACT_OCR_PSMS:
                diagnostics["ocr_attempt_count"] = int(diagnostics["ocr_attempt_count"]) + 1
                strategy_id = f"image_original_psm{psm}"
                lines, warning, attempt_diagnostics = _run_tesseract_image(
                    image_path,
                    psm=psm,
                    languages=languages,
                    strategy_id=strategy_id,
                    timeout_seconds=60,
                )
                score_payload = _score_contract_metadata_lines(lines)
                score = int(score_payload["score"])
                if score > best_score:
                    best_score = score
                    best_lines = lines
                    best_warning = warning
                    diagnostics.update(attempt_diagnostics)
                    diagnostics.update(
                        {
                            "ocr_strategy_used": strategy_id,
                            "ocr_text_nonempty": bool(lines),
                            "ocr_best_number_found": bool(score_payload["number_found"]),
                            "ocr_best_date_found": bool(score_payload["date_found"]),
                        }
                    )
                if lines and score_payload["number_found"] and score_payload["date_found"]:
                    return lines, [], "image_ocr", diagnostics
            warning = ""
            if best_lines:
                warning = "contract parser OCR text exists but number/date pattern was not fully found"
            elif best_warning:
                warning = best_warning
            else:
                warning = "contract parser OCR text empty after bounded image strategies"
            return best_lines, ([warning] if warning else []), "image_ocr", diagnostics
        finally:
            try:
                image_path.unlink()
            except OSError:
                pass
    except OSError:
        return [], ["contract parser skipped OCR: image temp file could not be created"], "image_no_ocr", diagnostics


def _run_tesseract_image(
    image_path: Path,
    *,
    psm: str,
    languages: str,
    strategy_id: str,
    timeout_seconds: float,
) -> tuple[list[str], str, dict[str, Any]]:
    diagnostics: dict[str, Any] = {
        "ocr_engine": "tesseract",
        "ocr_strategy_used": strategy_id,
        "ocr_languages": languages.split("+") if languages else [],
    }
    command = ["tesseract", str(image_path), "stdout"]
    if languages:
        command.extend(["-l", languages])
    command.extend(["--psm", str(psm)])
    try:
        result = subprocess.run(
            command,
            capture_output=True,
            timeout=timeout_seconds,
            check=False,
        )
    except subprocess.TimeoutExpired:
        return [], "contract parser OCR command timed out", diagnostics
    except (OSError, subprocess.SubprocessError):
        return [], "contract parser skipped OCR: tesseract command failed", diagnostics
    if result.returncode != 0:
        return [], "contract parser skipped OCR: tesseract returned no readable text", diagnostics
    lines = _text_to_lines(result.stdout.decode("utf-8", "replace"))
    diagnostics["ocr_text_nonempty"] = bool(lines)
    diagnostics["ocr_line_count"] = len(lines)
    if not lines:
        return [], "contract parser skipped OCR: tesseract returned no readable text", diagnostics
    return lines, "", diagnostics


def _preferred_tesseract_languages() -> str:
    available = _available_tesseract_languages()
    selected = [language for language in CONTRACT_OCR_PREFERRED_LANGUAGES if language in available]
    return "+".join(selected)


def _available_tesseract_languages() -> list[str]:
    global _TESSERACT_LANGUAGES_CACHE
    if _TESSERACT_LANGUAGES_CACHE is not None:
        return list(_TESSERACT_LANGUAGES_CACHE)
    if not shutil.which("tesseract"):
        _TESSERACT_LANGUAGES_CACHE = []
        return []
    try:
        result = subprocess.run(
            ["tesseract", "--list-langs"],
            capture_output=True,
            text=True,
            timeout=8,
            check=False,
        )
    except (OSError, subprocess.SubprocessError):
        _TESSERACT_LANGUAGES_CACHE = []
        return []
    languages: list[str] = []
    if result.returncode == 0:
        for line in result.stdout.splitlines():
            language = line.strip()
            if language and not language.lower().startswith("list of available"):
                languages.append(language)
    _TESSERACT_LANGUAGES_CACHE = languages
    return list(languages)


def _score_contract_metadata_lines(lines: list[str]) -> dict[str, Any]:
    normalized_lines = [_normalize_text_line(line) for line in lines]
    normalized_lines = [line for line in normalized_lines if line]
    first_line = normalized_lines[0] if normalized_lines else ""
    top_text = "\n".join(normalized_lines[:40])[:4000]
    number_found = bool(_extract_contract_number(first_line, normalized_lines[:8]))
    date_found = bool(_extract_contract_document_date(top_text))
    return {
        "number_found": number_found,
        "date_found": date_found,
        "score": (2 if number_found else 0) + (2 if date_found else 0) + (1 if normalized_lines else 0),
    }


def _text_to_lines(value: str) -> list[str]:
    return [line for line in (_normalize_text_line(line) for line in str(value or "").splitlines()) if line]


def _extract_contract_number(first_line: str, header_lines: list[str] | tuple[str, ...] | None = None) -> str:
    line = _normalize_text_line(first_line)
    if not line:
        return ""
    header_lines = [_normalize_text_line(item) for item in (header_lines or []) if _normalize_text_line(item)]
    header_text = " ".join(header_lines[:8])
    repeated_number = _extract_repeated_contract_number_from_ocr_line(line)
    if repeated_number:
        return repeated_number
    patterns = (
        r"(?iu)(?:合同(?:编号|号))\s*[:：#№\-]*\s*([A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9._/\-]+)",
        r"(?iu)\b(?:contract|контракт|договор|kontrakt|koнtrakt|kohtpakt)\s*(?:no\.?|number|№|#|n[o0b]?\.?|nb|j)?\s*[:：#№\-]*\s*([A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9._/\-]+)",
        r"(?iu)(?:^|\s)(?:no\.?|n[o0b]?\.?|nb|№|#|j)\s*([A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9._/\-]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, line)
        if match and _looks_like_contract_number(match.group(1)):
            return _clean_contract_number(match.group(1))
    if _is_generic_contract_heading(line):
        for pattern in patterns:
            match = re.search(pattern, header_text)
            if match and _looks_like_contract_number(match.group(1)):
                return _clean_contract_number(match.group(1))
        for candidate_line in header_lines[1:8]:
            candidate = _clean_contract_number(candidate_line)
            if _looks_like_contract_number(candidate):
                return candidate
        return ""
    if _has_contract_number_context(line):
        candidates = _contract_number_candidates(line)
        if candidates:
            return candidates[0]
    cleaned = _clean_contract_number(line)
    return cleaned if _looks_like_contract_number(cleaned) else ""


def _extract_repeated_contract_number_from_ocr_line(value: str) -> str:
    line = _normalize_text_line(value)
    if not _has_contract_number_context(line):
        return ""
    candidates = _contract_number_candidates(line)
    if not candidates:
        return ""
    normalized = [_normalize_contract_number_candidate(candidate) for candidate in candidates]
    counts: dict[str, int] = {}
    for candidate in normalized:
        counts[candidate] = counts.get(candidate, 0) + 1
    for candidate in normalized:
        if counts.get(candidate, 0) > 1:
            return candidate
    return ""


def _contract_number_candidates(value: str) -> list[str]:
    line = _normalize_text_line(value)
    candidates: list[str] = []
    patterns = (
        r"\b\d{1,8}\s*/\s*\d{1,8}\b",
        r"\b[A-Za-zА-Яа-я]{1,12}[-_/]\d[A-Za-zА-Яа-я0-9._/\-]{1,40}\b",
        r"\b\d{1,8}[-_][A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9._/\-]{1,40}\b",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, line):
            candidate = _normalize_contract_number_candidate(match.group(0))
            if _looks_like_contract_number(candidate) and candidate not in candidates:
                candidates.append(candidate)
    return candidates


def _normalize_contract_number_candidate(value: str) -> str:
    cleaned = _clean_contract_number(value)
    cleaned = re.sub(r"\s*/\s*", "/", cleaned)
    cleaned = re.sub(r"\s*[-_]\s*", lambda match: match.group(0).strip(), cleaned)
    return cleaned


def _has_contract_number_context(value: str) -> bool:
    line = _normalize_text_line(value).casefold()
    return bool(
        re.search(
            r"(?iu)(contract|контракт|договор|kontrakt|koнtrakt|kohtpakt|合同|no\.?|nb|n[o0b]?\.?|№|#)",
            line,
        )
    )


def _clean_contract_number(value: str) -> str:
    cleaned = _normalize_text_line(value)
    cleaned = re.sub(r"(?iu)\b(?:dated|date|от|дата)\b.*$", "", cleaned).strip()
    cleaned = re.sub(r"(?iu)\b\d{4}[-./]\d{1,2}[-./]\d{1,2}\b.*$", "", cleaned).strip()
    cleaned = cleaned.strip(" :：#№-—\"'«»")
    return cleaned[:120]


def _is_generic_contract_heading(value: str) -> bool:
    normalized = re.sub(r"[\s:：#№.\-—\"'«»]+", "", str(value or "").casefold())
    return normalized in {
        "contract",
        "salescontract",
        "purchasecontract",
        "kontrakt",
        "koнtrakt",
        "kohtpakt",
        "контракт",
        "договор",
        "договорпоставки",
        "合同",
        "销售合同",
    }


def _looks_like_contract_number(value: str) -> bool:
    cleaned = _clean_contract_number(value)
    if not cleaned:
        return False
    if _is_generic_contract_heading(cleaned):
        return False
    if len(cleaned) > 80 and not re.search(r"\d", cleaned):
        return False
    return bool(re.search(r"\d", cleaned) or re.search(r"[A-Za-zА-Яа-я]{2,}[-_/]\w", cleaned))


def _extract_contract_document_date(text: str) -> str:
    bounded = str(text or "")[:4000]
    if not bounded.strip():
        return ""
    label_pattern = re.compile(
        r"(?iu)(contract\s+date|date\s+of\s+contract|合同日期|签订日期|date|дата)\s*[:：\-]?\s*([^\n]{0,120})"
    )
    for match in label_pattern.finditer(bounded):
        parsed = _parse_contract_date_fragment(match.group(0))
        if parsed:
            return parsed
    return _parse_contract_date_fragment(bounded)


def _parse_contract_date_fragment(value: str) -> str:
    text = _normalize_text_line(value)
    if not text:
        return ""
    for pattern in (
        r"\b(\d{4})[-./](\d{1,2})[-./](\d{1,2})\b",
        r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日",
    ):
        match = re.search(pattern, text)
        if match:
            return _format_contract_date(match.group(1), match.group(2), match.group(3))
    match = re.search(r"\b(\d{1,2})\.(\d{1,2})\.(\d{4})\b", text)
    if match:
        return _format_contract_date(match.group(3), match.group(2), match.group(1))
    match = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", text)
    if match:
        return _format_contract_date(match.group(3), match.group(1), match.group(2))
    month_names = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    match = re.search(
        r"(?iu)\b("
        + "|".join(re.escape(name) for name in sorted(month_names, key=len, reverse=True))
        + r")\.?\s+(\d{1,2}),?\s+(\d{4})\b",
        text,
    )
    if match:
        return _format_contract_date(match.group(3), str(month_names[match.group(1).lower()]), match.group(2))
    match = re.search(
        r"(?iu)\b(\d{1,2})\s+("
        + "|".join(re.escape(name) for name in sorted(month_names, key=len, reverse=True))
        + r")\.?,?\s+(\d{4})\b",
        text,
    )
    if match:
        return _format_contract_date(match.group(3), str(month_names[match.group(2).lower()]), match.group(1))
    ru_months = {
        "января": 1,
        "февраля": 2,
        "марта": 3,
        "апреля": 4,
        "мая": 5,
        "июня": 6,
        "июля": 7,
        "августа": 8,
        "сентября": 9,
        "октября": 10,
        "ноября": 11,
        "декабря": 12,
    }
    match = re.search(
        r"(?iu)[«\"]?(\d{1,2})[»\"]?\s+("
        + "|".join(ru_months)
        + r")\s+(\d{4})",
        text,
    )
    if match:
        return _format_contract_date(match.group(3), str(ru_months[match.group(2).lower()]), match.group(1))
    return ""


def _format_contract_date(year: Any, month: Any, day: Any) -> str:
    try:
        return date(int(year), int(month), int(day)).isoformat()
    except (TypeError, ValueError):
        return ""


def _normalize_text_line(value: Any) -> str:
    text = str(value or "").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _compact_compare(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def _document_supplier_name(*values: Any) -> str:
    for value in values:
        supplier = str(value or "").strip()
        if supplier:
            return supplier
    return DEFAULT_SUPPLIER_NAME


def _merge_string_lists(existing: Any, incoming: Any) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for value in [*_string_list(existing), *_string_list(incoming)]:
        item = str(value or "").strip()
        if not item or item in seen:
            continue
        seen.add(item)
        merged.append(item)
    return merged


def _merge_contract_parser_warnings(
    existing: Any,
    incoming: Any,
    *,
    parsed_number: str,
    parsed_date: str,
) -> list[str]:
    base = _string_list(existing)
    if str(parsed_number or "").strip() and str(parsed_date or "").strip():
        base = _remove_stale_contract_parser_warnings(base)
    return _merge_string_lists(base, incoming)


def _remove_stale_contract_parser_warnings(warnings: Any) -> list[str]:
    return [warning for warning in _string_list(warnings) if not _is_stale_contract_parser_warning(warning)]


def _is_stale_contract_parser_warning(warning: str) -> bool:
    normalized = str(warning or "").strip().lower()
    return normalized.startswith(
        (
            "contract parser skipped ocr",
            "contract parser ocr text",
            "contract parser found no readable text",
            "contract parser could not extract",
            "contract parser found no text through pdftotext",
        )
    )


def _invoice_download_path(shipment_id: str) -> str:
    if not shipment_id:
        return ""
    return f"/v1/sheet-vitrina-v1/supply/supplier-shipments/{shipment_id}/invoice"


def _contract_download_path(shipment_id: str) -> str:
    if not shipment_id:
        return ""
    return f"/v1/sheet-vitrina-v1/supply/supplier-shipments/{shipment_id}/contract"


def _trade_document_download_path(document_id: str) -> str:
    if not document_id:
        return ""
    return f"/v1/sheet-vitrina-v1/settings/documents/{document_id}/file"


def _normalize_trade_document_type(value: Any) -> str:
    document_type = str(value or "").strip().lower()
    if document_type not in TRADE_DOCUMENT_TYPES:
        raise ValueError("document_type must be contract or invoice")
    return document_type


def _document_content_type(filename: str, uploaded_content_type: str | None = None) -> str:
    extension = Path(str(filename or "")).suffix.lower()
    fallback = TRADE_DOCUMENT_CONTENT_TYPES_BY_EXTENSION.get(extension, "application/octet-stream")
    content_type = str(uploaded_content_type or "").strip().split(";", 1)[0].strip().lower()
    return content_type or fallback


def _safe_document_filename(value: str, *, document_type: str) -> str:
    fallback = f"{document_type or 'document'}.xlsx"
    name = _safe_filename(value or fallback)
    if not Path(name).suffix:
        name = f"{name}.xlsx"
    return name


def _safe_filename(value: str) -> str:
    name = Path(str(value or "")).name.strip()
    name = name.replace("\x00", "").replace("/", "_").replace("\\", "_")
    name = re.sub(r"[\r\n\t]+", " ", name).strip()
    if not name:
        name = "supplier-invoice.xlsx"
    if len(name) > 180:
        stem = Path(name).stem[:150] or "supplier-invoice"
        suffix = Path(name).suffix[:16] or ".xlsx"
        name = stem + suffix
    return name


def _relative_to_runtime(runtime_dir: Path, path: Path) -> str:
    root = runtime_dir.resolve()
    resolved = path.resolve()
    if root != resolved and root not in resolved.parents:
        raise ValueError("runtime file path escapes runtime dir")
    return resolved.relative_to(root).as_posix()


def _validate_iso_date(value: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        raise ValueError("shipment_date is required")
    try:
        date.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError("shipment_date must be an ISO date YYYY-MM-DD") from exc
    return normalized


def _validate_optional_iso_date(value: Any, *, field_name: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
        raise ValueError(f"{field_name} must be an ISO date YYYY-MM-DD or blank")
    try:
        date.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be an ISO date YYYY-MM-DD or blank") from exc
    return normalized


def _normalize_order_status(value: Any) -> str:
    normalized = str(value or ORDER_STATUS_DEFAULT).strip()
    if not normalized:
        return ORDER_STATUS_DEFAULT
    if normalized not in ORDER_STATUSES:
        raise ValueError(f"unsupported supplier order_status: {normalized}")
    return normalized


def _normalize_bool_field(value: Any, *, field_name: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "1", "yes", "y", "да"}:
            return True
        if normalized in {"false", "0", "no", "n", "нет", ""}:
            return False
    raise ValueError(f"{field_name} must be boolean")


def _resolve_optional_positive_decimal_field(
    payload: Mapping[str, Any],
    edited_payload: Mapping[str, Any],
    existing_header: Mapping[str, Any] | None,
    field_name: str,
) -> float | None:
    if field_name in payload:
        return _validate_optional_positive_decimal(payload.get(field_name), field_name=field_name)
    if isinstance(payload.get("payload"), Mapping) and field_name in payload.get("payload", {}):
        return _validate_optional_positive_decimal(payload.get("payload", {}).get(field_name), field_name=field_name)
    if field_name in edited_payload:
        return _validate_optional_positive_decimal(edited_payload.get(field_name), field_name=field_name)
    if existing_header is not None:
        return _validate_optional_positive_decimal(existing_header.get(field_name), field_name=field_name)
    return None


def _validate_optional_positive_decimal(value: Any, *, field_name: str) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError(f"{field_name} must be a number greater than 0 or blank")
    if isinstance(value, Decimal):
        parsed = value
    elif isinstance(value, (int, float)):
        parsed = Decimal(str(value))
    else:
        text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace("−", "-")
        if not text:
            return None
        text = re.sub(r"[^0-9,.\-]", "", text)
        if not text or text in {"-", ".", ","}:
            raise ValueError(f"{field_name} must be a number greater than 0 or blank")
        if text.count("-") > 1 or ("-" in text and not text.startswith("-")):
            raise ValueError(f"{field_name} must be a number greater than 0 or blank")
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            parsed = Decimal(text)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError(f"{field_name} must be a number greater than 0 or blank") from exc
    if parsed <= 0:
        raise ValueError(f"{field_name} must be greater than 0")
    return float(parsed.quantize(APPROX_YUAN_RATE_QUANT, rounding=ROUND_HALF_UP))


def _read_optional_positive_decimal(value: Any) -> float | None:
    try:
        return _validate_optional_positive_decimal(value, field_name="approx_yuan_rate")
    except ValueError:
        return None


def _normalize_price_conformity_status(value: Any) -> str:
    normalized = str(value or PRICE_CONFORMITY_STATUS_NOT_CHECKED).strip()
    return normalized if normalized in PRICE_CONFORMITY_STATUSES else PRICE_CONFORMITY_STATUS_NOT_CHECKED


def _normalize_price_conformity_check_mode(value: Any) -> str:
    normalized = str(value or PRICE_CONFORMITY_CHECK_MODE_NOT_CHECKED).strip()
    return normalized if normalized in PRICE_CONFORMITY_CHECK_MODES else PRICE_CONFORMITY_CHECK_MODE_NOT_CHECKED


def _optional_timestamp(value: Any) -> str:
    return str(value or "").strip()


def _normalize_json_object(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return dict(parsed) if isinstance(parsed, Mapping) else {}
    return {}


def _optional_iso_date(value: Any) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    try:
        date.fromisoformat(normalized)
    except ValueError:
        return normalized
    return normalized


def _optional_trade_document_date(value: Any) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
        raise ValueError("document_date must be an ISO date YYYY-MM-DD or blank")
    try:
        date.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError("document_date must be an ISO date YYYY-MM-DD or blank") from exc
    return normalized


def _optional_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", " ")
    if not text:
        return None
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _bounded_int(value: Any, *, default: int, minimum: int, maximum: int) -> int:
    try:
        normalized = int(value)
    except (TypeError, ValueError):
        normalized = int(default)
    return min(max(normalized, minimum), maximum)


def _sum_numeric(values: Any) -> float:
    total = 0.0
    for value in values:
        number = _optional_number(value)
        if number is not None:
            total += number
    return round(total, 2)


def _string_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item) for item in value if str(item or "").strip()]


def _default_timestamp_factory() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
