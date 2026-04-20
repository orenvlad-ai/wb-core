"""Server-owned live source seam for promo-backed daily metrics."""

from __future__ import annotations

from dataclasses import asdict
from datetime import date, datetime
import json
import os
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from packages.adapters.promo_xlsx_collector_block import (
    PlaywrightPromoCollectorDriver,
)
from packages.application.promo_campaign_archive import (
    materialize_promo_result_from_archive,
    promo_campaign_archive_root,
    sync_promo_campaign_archive,
)
from packages.application.promo_xlsx_collector_block import PromoXlsxCollectorBlock
from packages.contracts.promo_live_source import (
    PromoLiveSourceEnvelope,
    PromoLiveSourceIncomplete,
    PromoLiveSourceItem,
    PromoLiveSourceRequest,
    PromoLiveSourceSuccess,
)
from packages.contracts.promo_xlsx_collector_block import PromoMetadata, PromoOutcome, PromoXlsxCollectorRequest


BUSINESS_TIMEZONE = ZoneInfo("Asia/Yekaterinburg")
PRICE_EPSILON = 0.5
PROMO_RUNTIME_DIRNAME = "promo_xlsx_collector_runs"
HEADER_NM_ID = "Артикул WB"
HEADER_PLAN_PRICE = "Плановая цена для акции"
HEADER_CURRENT_PRICE = "Текущая розничная цена"
HEADER_CURRENT_DISCOUNT = "Текущая скидка на сайте, %"
HEADER_UPLOADABLE_DISCOUNT = "Загружаемая скидка для участия в акции"


class PromoLiveSourceBlock:
    def __init__(
        self,
        *,
        runtime_dir: Path,
        collector_block: PromoXlsxCollectorBlock | None = None,
        now_factory=None,
        storage_state_path: str | None = None,
        headless: bool = True,
        max_candidates: int | None = None,
        max_downloads: int | None = None,
    ) -> None:
        self.runtime_dir = runtime_dir
        self.collector_block = collector_block
        self.now_factory = now_factory or _default_now_factory
        self.storage_state_path = (
            storage_state_path
            or str(os.environ.get("PROMO_XLSX_COLLECTOR_STORAGE_STATE_PATH", "")).strip()
        )
        self.headless = headless
        self.max_candidates = max_candidates
        self.max_downloads = max_downloads

    def execute(self, request: PromoLiveSourceRequest) -> PromoLiveSourceEnvelope:
        run_dir = (
            self.runtime_dir
            / PROMO_RUNTIME_DIRNAME
            / f"{request.snapshot_date}__{self.now_factory().strftime('%Y%m%d_%H%M%S')}"
        )
        run_dir.mkdir(parents=True, exist_ok=True)
        archive_root = promo_campaign_archive_root(self.runtime_dir)
        collector_summary = None
        if request.snapshot_date == self.now_factory().date().isoformat():
            collector_request = PromoXlsxCollectorRequest(
                output_root=str(run_dir),
                storage_state_path=request.storage_state_path or self.storage_state_path,
                archive_root=str(archive_root),
                source_tab=request.source_tab,
                source_filter_code=request.source_filter_code,
                headless=request.headless if request.headless is not None else self.headless,
                hydration_attempt_budget=request.hydration_attempt_budget,
                hydration_wait_sec=request.hydration_wait_sec,
                max_candidates=request.max_candidates if request.max_candidates is not None else self.max_candidates,
                max_downloads=request.max_downloads if request.max_downloads is not None else self.max_downloads,
            )
            collector_block = self.collector_block or PromoXlsxCollectorBlock(PlaywrightPromoCollectorDriver(run_dir))
            collector_summary = collector_block.execute(collector_request)
        sync_summary = sync_promo_campaign_archive(self.runtime_dir)
        if collector_summary is None:
            detail_prefix = "collector_mode=archive_only"
            trace_run_dir = str(archive_root)
        else:
            detail_prefix = (
                f"collector_mode=live_refresh; "
                f"trace_run_dir={collector_summary.run_dir}; "
                f"collector_status={collector_summary.status}; "
                f"hydration_attempts={len(collector_summary.hydration_attempts)}; "
                f"archive_reuse_enabled=true"
            )
            trace_run_dir = collector_summary.run_dir
        result = materialize_promo_result_from_archive(
            runtime_dir=self.runtime_dir,
            snapshot_date=request.snapshot_date,
            requested_nm_ids=request.nm_ids,
            sync_summary=sync_summary,
            trace_run_dir=trace_run_dir,
            detail_prefix=detail_prefix,
        )
        (run_dir / "derived_promo_live_source.json").write_text(
            json.dumps(asdict(result), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return PromoLiveSourceEnvelope(result=result)


def _derive_live_source_result(
    *,
    summary,
    requested_nm_ids: list[int],
    snapshot_date: str,
) -> PromoLiveSourceSuccess | PromoLiveSourceIncomplete:
    slot_date = date.fromisoformat(snapshot_date)
    collected_date = date.fromisoformat(summary.started_at[:10])
    item_accumulator = {
        nm_id: {"promo_count_by_price": 0.0, "promo_entry_price_best": 0.0}
        for nm_id in requested_nm_ids
    }
    current_promos = 0
    current_promos_downloaded = 0
    current_promos_blocked = 0
    future_promos = 0
    skipped_past_promos = 0
    ambiguous_promos = 0
    export_kinds: set[str] = set()
    blockers: list[str] = []

    for outcome in summary.promos:
        relation = _classify_outcome_for_slot(
            metadata=outcome.metadata,
            slot_date=slot_date,
            collected_date=collected_date,
        )
        if relation == "past":
            skipped_past_promos += 1
            continue
        if relation == "future":
            future_promos += 1
            continue
        if relation == "ambiguous":
            ambiguous_promos += 1
            blockers.append(
                f"ambiguous_promo={outcome.promo_title}; status={outcome.status}; period={outcome.metadata.promo_period_text}"
            )
            continue

        current_promos += 1
        if outcome.status != "downloaded" or not outcome.saved_path:
            current_promos_blocked += 1
            blockers.append(
                f"current_promo_not_downloaded={outcome.promo_title}; status={outcome.status}; blocker={outcome.blocker or ''}"
            )
            continue

        current_promos_downloaded += 1
        if outcome.export_kind:
            export_kinds.add(outcome.export_kind)
        _merge_workbook_rows(
            accumulator=item_accumulator,
            workbook_path=Path(outcome.saved_path),
            requested_nm_ids=requested_nm_ids,
        )

    items = [
        PromoLiveSourceItem(
            snapshot_date=snapshot_date,
            nm_id=nm_id,
            promo_count_by_price=round(values["promo_count_by_price"], 6),
            promo_entry_price_best=round(values["promo_entry_price_best"], 6),
            promo_participation=1.0 if values["promo_count_by_price"] > 0 else 0.0,
        )
        for nm_id, values in sorted(item_accumulator.items())
    ]
    detail = _build_detail(
        trace_run_dir=summary.run_dir,
        collector_status=summary.status,
        current_promos=current_promos,
        current_promos_downloaded=current_promos_downloaded,
        current_promos_blocked=current_promos_blocked,
        future_promos=future_promos,
        skipped_past_promos=skipped_past_promos,
        ambiguous_promos=ambiguous_promos,
        hydration_attempts=len(summary.hydration_attempts),
        blockers=blockers,
    )
    common = dict(
        snapshot_date=snapshot_date,
        date_from=snapshot_date,
        date_to=snapshot_date,
        requested_count=len(requested_nm_ids),
        items=items,
        detail=detail,
        trace_run_dir=summary.run_dir,
        current_promos=current_promos,
        current_promos_downloaded=current_promos_downloaded,
        current_promos_blocked=current_promos_blocked,
        future_promos=future_promos,
        skipped_past_promos=skipped_past_promos,
        ambiguous_promos=ambiguous_promos,
        current_download_export_kinds=sorted(export_kinds),
    )
    if summary.status == "blocked" or current_promos_blocked > 0 or ambiguous_promos > 0:
        return PromoLiveSourceIncomplete(
            kind="incomplete",
            covered_count=0 if current_promos_downloaded <= 0 else len(requested_nm_ids),
            missing_nm_ids=sorted(requested_nm_ids),
            **common,
        )
    return PromoLiveSourceSuccess(
        kind="success",
        covered_count=len(requested_nm_ids),
        **common,
    )


def _merge_workbook_rows(
    *,
    accumulator: dict[int, dict[str, float]],
    workbook_path: Path,
    requested_nm_ids: Iterable[int],
) -> None:
    requested = set(requested_nm_ids)
    workbook = load_workbook(filename=str(workbook_path), read_only=False, data_only=True)
    try:
        sheet, header_row_index = _find_workbook_data_sheet(workbook)
        header = list(next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True)))
        header_index = {str(name).strip(): idx for idx, name in enumerate(header) if name not in (None, "")}
        for required in (
            HEADER_NM_ID,
            HEADER_PLAN_PRICE,
            HEADER_CURRENT_PRICE,
            HEADER_CURRENT_DISCOUNT,
            HEADER_UPLOADABLE_DISCOUNT,
        ):
            if required not in header_index:
                raise ValueError(f"promo workbook missing required header: {required}")
        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            nm_id = _parse_int(row[header_index[HEADER_NM_ID]])
            if nm_id is None or nm_id not in requested:
                continue
            plan_price = _parse_float(row[header_index[HEADER_PLAN_PRICE]])
            current_price = _parse_float(row[header_index[HEADER_CURRENT_PRICE]])
            current_discount = _parse_float(row[header_index[HEADER_CURRENT_DISCOUNT]])
            uploadable_discount = _parse_float(row[header_index[HEADER_UPLOADABLE_DISCOUNT]])
            if plan_price is None:
                continue
            accumulator[nm_id]["promo_entry_price_best"] = max(
                accumulator[nm_id]["promo_entry_price_best"],
                plan_price,
            )
            effective_discount = uploadable_discount if uploadable_discount is not None else current_discount
            if current_price is None or effective_discount is None:
                continue
            effective_price = current_price * (100.0 - effective_discount) / 100.0
            if effective_price <= plan_price + PRICE_EPSILON:
                accumulator[nm_id]["promo_count_by_price"] += 1.0
    finally:
        workbook.close()


def _classify_outcome_for_slot(
    *,
    metadata: PromoMetadata,
    slot_date: date,
    collected_date: date,
) -> str:
    if metadata.promo_start_at and metadata.promo_end_at:
        start_date = date.fromisoformat(metadata.promo_start_at[:10])
        end_date = date.fromisoformat(metadata.promo_end_at[:10])
        if slot_date < start_date:
            return "future"
        if slot_date > end_date:
            return "past"
        return "current"
    if slot_date == collected_date:
        return metadata.temporal_classification
    return "ambiguous"


def _build_detail(
    *,
    trace_run_dir: str,
    collector_status: str,
    current_promos: int,
    current_promos_downloaded: int,
    current_promos_blocked: int,
    future_promos: int,
    skipped_past_promos: int,
    ambiguous_promos: int,
    hydration_attempts: int,
    blockers: list[str],
) -> str:
    parts = [
        f"trace_run_dir={trace_run_dir}",
        f"collector_status={collector_status}",
        f"hydration_attempts={hydration_attempts}",
        f"current_promos={current_promos}",
        f"current_promos_downloaded={current_promos_downloaded}",
        f"current_promos_blocked={current_promos_blocked}",
        f"future_promos={future_promos}",
        f"skipped_past_promos={skipped_past_promos}",
        f"ambiguous_promos={ambiguous_promos}",
    ]
    if blockers:
        parts.append("blockers=" + " | ".join(blockers))
    return "; ".join(parts)


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace("%", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_int(value: object) -> int | None:
    numeric = _parse_float(value)
    if numeric is None:
        return None
    return int(numeric)


def _find_workbook_data_sheet(workbook) -> tuple[object, int]:
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_index = _find_workbook_header_row_index(sheet)
        if row_index is not None:
            return sheet, row_index
    first_sheet = workbook[workbook.sheetnames[0]]
    row_index = _find_workbook_header_row_index(first_sheet)
    return first_sheet, row_index or 1


def _find_workbook_header_row_index(sheet) -> int | None:
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=True), start=1):
        header = [value for value in row if value not in (None, "")]
        normalized = {str(value).strip() for value in header}
        if {HEADER_NM_ID, HEADER_PLAN_PRICE, HEADER_CURRENT_PRICE}.issubset(normalized):
            return row_index
    return None


def _default_now_factory() -> datetime:
    return datetime.now(BUSINESS_TIMEZONE)
