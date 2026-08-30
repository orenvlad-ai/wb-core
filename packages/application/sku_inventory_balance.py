"""Immutable SKU inventory-balance calculations and fail-closed apply jobs."""

from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from hashlib import sha256
from io import BytesIO
import json
import logging
from pathlib import Path
import re
import sqlite3
import threading
import time
from typing import Any, Callable, Mapping, Protocol, Sequence
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from packages.application.sku_management import SkuManagementError
from packages.application.sheet_vitrina_v1_ads import (
    AdsBidSafetyThresholdPolicy,
    SheetVitrinaV1AdsError,
)
from packages.application.change_registry import ChangeRegistryRepository
from packages.application.change_registry_writer import (
    InternalWriterRegistry,
    InternalWriterRegistryError,
    PreparedWriterOperation,
)
from packages.adapters.wb_promotion import WbPromotionApiError
from packages.business_time import current_business_date_iso


CALCULATION_CONTRACT = "sheet_vitrina_v1_sku_inventory_balance/v2"
FORMULA_VERSION = "sku_inventory_balance_conservative_pace_v2"
CALCULATION_OPERATION_CONTRACT = "sheet_vitrina_v1_sku_inventory_balance_operation/v1"
CALCULATION_OPERATION_ACCEPTANCE_CONTRACT = (
    "sheet_vitrina_v1_sku_inventory_balance_operation_acceptance/v1"
)
CALCULATION_OPERATIONS_PATH = (
    "/v1/sheet-vitrina-v1/sku-management/inventory-balance/operations"
)
CONFIG_KEY = "sku_inventory_balance"
CONFIG_SCHEMA_VERSION = 1
DRY_RUN_MODE = "dry_run"
LIVE_MODE = "live_wb"
MANUAL_PENDING_CONTRACT = "change_registry_manual_pending/v1"
BALANCE_OWNER_CONFIRMATION_POLICY_CONTRACT = (
    "inventory_balance_owner_confirmed_bid_thresholds/v1"
)
TERMINAL_ITEM_STATES = {"succeeded", "failed", "skipped", "ambiguous"}
ACTIVE_JOB_STATES = {"pending", "running", "delayed", "stalled"}
ACTIVE_ITEM_STATES = {
    "pending",
    "preflighting",
    "ready",
    "submitting",
    "submitted",
    "verifying",
    "delayed",
}
SUPPORTED_CAMPAIGN_STATUSES = {4, 9, 11}
CAMPAIGN_STATE_BY_STATUS = {4: "ready", 9: "active", 11: "paused"}
CAMPAIGN_STATE_ACTION_BY_STATUS = {4: "start", 9: "pause", 11: "start"}
EXCLUSION_POLICY_VERSION = "sku_inventory_balance_exclusions_v1"
EXCLUDED_NM_IDS = {
    497413772: "iPhone Air glass is outside inventory-balance scope",
    497415593: "iPhone Air glass is outside inventory-balance scope",
    497416931: "iPhone Air glass is outside inventory-balance scope",
}

BALANCE_COLUMNS = (
    "select",
    "product",
    "status",
    "known_stock_units",
    "current_daily_sales",
    "target_daily_sales",
    "pace_change_pct",
    "days_cover",
    "bottleneck_date",
    "next_inbound",
    "subsequent_inbound",
    "new_cpc_campaigns",
    "old_cpm_campaigns",
    "quality",
)
MANDATORY_COLUMNS = ("select", "product")
DEFAULT_VISIBLE_COLUMNS = BALANCE_COLUMNS
_OPERATION_ID_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._:-]{7,127}$")
_LOGGER = logging.getLogger(__name__)


class SkuInventoryBalanceError(SkuManagementError):
    """Controlled inventory-balance contract error."""


class InventoryBalanceApplyAdapter(Protocol):
    mode: str
    external_writes_enabled: bool

    def apply(self, target: Mapping[str, Any], *, actor: str) -> dict[str, Any]:
        """Apply or simulate one exact campaign target."""


class DryRunInventoryBalanceApplyAdapter:
    """Default adapter: durable simulation with zero WB calls."""

    mode = DRY_RUN_MODE
    external_writes_enabled = False

    def apply(self, target: Mapping[str, Any], *, actor: str) -> dict[str, Any]:
        return {
            "status": "succeeded",
            "adapter": "dry_run_noop",
            "actor": actor,
            "target_key": str(target.get("target_key") or ""),
            "requested_bid_rub": target.get("final_target_bid_rub"),
            "readback_status": "simulated_matching",
            "wb_patch_called": False,
        }


class LiveWbInventoryBalanceApplyAdapter:
    """Balance-owned bulk transport over the incumbent guarded Ads source."""

    mode = LIVE_MODE

    def __init__(
        self,
        *,
        sku_management_block: Any,
    ) -> None:
        self.sku_management_block = sku_management_block

    @property
    def external_writes_enabled(self) -> bool:
        ads = getattr(self.sku_management_block, "ads_block", None)
        safety = getattr(ads, "safety", None)
        return bool(safety is not None and getattr(safety, "write_enabled", False))

    def apply(self, target: Mapping[str, Any], *, actor: str) -> dict[str, Any]:
        del target, actor
        raise SkuInventoryBalanceError(
            "live Balance jobs use batch preflight/submit/readback, not single apply",
            http_status=409,
        )

    def preflight(
        self,
        targets: Sequence[Mapping[str, Any]],
        *,
        min_bid_interval_seconds: float,
        sleep: Callable[[float], None],
        safety_threshold_policy: AdsBidSafetyThresholdPolicy,
    ) -> list[dict[str, Any]]:
        if not self.external_writes_enabled:
            raise SkuInventoryBalanceError(
                "live WB inventory-balance capability is disabled",
                http_status=403,
            )
        return self.sku_management_block.ads_block.preflight_bid_targets(
            targets,
            min_bid_interval_seconds=min_bid_interval_seconds,
            sleep=sleep,
            safety_threshold_policy=safety_threshold_policy,
        )

    def owner_confirmation_policy(self) -> dict[str, Any]:
        ads = self.sku_management_block.ads_block
        safety = ads.safety
        return {
            "contract_name": BALANCE_OWNER_CONFIRMATION_POLICY_CONTRACT,
            "safety_threshold_policy": (
                AdsBidSafetyThresholdPolicy.OWNER_CONFIRMED_BALANCE.value
            ),
            "warnings_only": [
                "absolute_max_bid",
                "max_absolute_increase",
                "max_percent_increase",
            ],
            "thresholds": {
                "absolute_max_bid_rub": safety.absolute_max_bid_kopecks / 100,
                "max_absolute_increase_rub": (
                    safety.max_absolute_increase_kopecks / 100
                ),
                "max_percent_increase": float(safety.max_percent_increase),
            },
            "direct_submit": True,
            "staircase_submit": False,
        }

    def safety_threshold_warnings(
        self, target: Mapping[str, Any]
    ) -> list[dict[str, Any]]:
        return self.sku_management_block.ads_block.bid_safety_threshold_warnings(
            old_bid_kopecks=int(target["current_bid_minor"]),
            new_bid_kopecks=int(target["final_target_bid_minor"]),
        )

    def submit_batch(
        self, targets: Sequence[Mapping[str, Any]]
    ) -> Mapping[str, Any]:
        if not self.external_writes_enabled:
            raise SkuInventoryBalanceError(
                "live WB inventory-balance capability is disabled",
                http_status=403,
            )
        return self.sku_management_block.ads_block.submit_bid_targets(targets)

    def preflight_state(
        self, targets: Sequence[Mapping[str, Any]]
    ) -> list[dict[str, Any]]:
        if not self.external_writes_enabled:
            raise SkuInventoryBalanceError(
                "live WB inventory-balance capability is disabled",
                http_status=403,
            )
        return self.sku_management_block.ads_block.preflight_campaign_state_targets(
            targets
        )

    def submit_state(self, target: Mapping[str, Any]) -> Mapping[str, Any]:
        if not self.external_writes_enabled:
            raise SkuInventoryBalanceError(
                "live WB inventory-balance capability is disabled",
                http_status=403,
            )
        return self.sku_management_block.ads_block.submit_campaign_state(target)

    def readback_state(
        self, targets: Sequence[Mapping[str, Any]]
    ) -> list[dict[str, Any]]:
        return self.sku_management_block.ads_block.read_campaign_state_targets(
            targets
        )

    def readback(
        self, targets: Sequence[Mapping[str, Any]]
    ) -> list[dict[str, Any]]:
        return self.sku_management_block.ads_block.read_bid_targets(targets)


class SkuInventoryBalanceBlock:
    """Server-owned calculations, overrides, workbook export and apply jobs."""

    def __init__(
        self,
        *,
        runtime: Any,
        sku_management_block: Any,
        now_factory: Callable[[], datetime] | None = None,
        timestamp_factory: Callable[[], str] | None = None,
        apply_adapter: InventoryBalanceApplyAdapter | None = None,
        manual_pending_registry: ChangeRegistryRepository | None = None,
        writer_registry: InternalWriterRegistry | None = None,
        seller_id: str = "",
        account_scope: str = "seller-portal-primary",
        live_batch_size: int = 10,
        min_bid_interval_seconds: float = 3.0,
        patch_interval_seconds: float = 0.2,
        readback_initial_delay_seconds: float = 30.0,
        readback_poll_seconds: float = 5.0,
        readback_deadline_seconds: float = 90.0,
        apply_lease_seconds: int = 180,
        sleep: Callable[[float], None] = time.sleep,
        monotonic_factory: Callable[[], float] = time.monotonic,
    ) -> None:
        self.runtime = runtime
        self.sku_management_block = sku_management_block
        self.now_factory = now_factory or (lambda: datetime.now(timezone.utc))
        self.timestamp_factory = timestamp_factory or (
            lambda: datetime.now(timezone.utc).isoformat()
        )
        self.apply_adapter = apply_adapter or DryRunInventoryBalanceApplyAdapter()
        self.manual_pending_registry = manual_pending_registry
        self.writer_registry = writer_registry
        self.seller_id = str(seller_id or "").strip()
        self.account_scope = str(account_scope or "").strip()
        self.live_batch_size = min(max(int(live_batch_size), 1), 50)
        self.min_bid_interval_seconds = max(float(min_bid_interval_seconds), 0.0)
        self.patch_interval_seconds = max(float(patch_interval_seconds), 0.0)
        self.readback_initial_delay_seconds = max(
            float(readback_initial_delay_seconds), 0.0
        )
        self.readback_poll_seconds = max(float(readback_poll_seconds), 0.0)
        self.readback_deadline_seconds = max(float(readback_deadline_seconds), 0.0)
        self.apply_lease_seconds = max(int(apply_lease_seconds), 30)
        self.sleep = sleep
        self.monotonic_factory = monotonic_factory
        if self.apply_adapter.mode not in {DRY_RUN_MODE, LIVE_MODE}:
            raise SkuInventoryBalanceError(
                "inventory-balance apply adapter mode is unsupported",
                http_status=500,
            )
        if self.apply_adapter.mode == LIVE_MODE and (
            not self.apply_adapter.external_writes_enabled
            or self.writer_registry is None
            or not self.seller_id
        ):
            raise SkuInventoryBalanceError(
                "live inventory-balance runtime requires enabled Ads writes and registry binding",
                http_status=500,
            )
        self._calculation_worker_lock = threading.Lock()
        self._calculation_worker_thread: threading.Thread | None = None
        self._calculation_worker_operation_id = ""
        self._apply_worker_lock = threading.Lock()
        self._apply_worker_thread: threading.Thread | None = None
        self._apply_worker_wakeup = threading.Event()
        self._apply_worker_stop = threading.Event()
        self.ensure_schema()
        self._terminalize_interrupted_calculation_operations()
        self._start_apply_worker_if_needed()

    def ensure_schema(self) -> None:
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS sheet_vitrina_v1_inventory_balance_calculations (
                    calculation_id TEXT PRIMARY KEY,
                    operation_id TEXT,
                    previous_calculation_id TEXT,
                    contract_name TEXT NOT NULL,
                    formula_version TEXT NOT NULL,
                    source_digest TEXT NOT NULL,
                    settings_json TEXT NOT NULL,
                    payload_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    created_by TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS inventory_balance_calculations_created
                ON sheet_vitrina_v1_inventory_balance_calculations(created_at DESC, calculation_id DESC);
                CREATE TABLE IF NOT EXISTS sheet_vitrina_v1_inventory_balance_operations (
                    operation_id TEXT PRIMARY KEY,
                    user_key TEXT NOT NULL,
                    idempotency_key TEXT NOT NULL,
                    contract_name TEXT NOT NULL,
                    request_digest TEXT NOT NULL,
                    request_json TEXT NOT NULL,
                    acceptance_receipt_json TEXT NOT NULL,
                    state TEXT NOT NULL,
                    phase TEXT NOT NULL,
                    progress_percent INTEGER NOT NULL,
                    calculation_id TEXT,
                    error_code TEXT NOT NULL DEFAULT '',
                    error_message TEXT NOT NULL DEFAULT '',
                    outcome_json TEXT NOT NULL DEFAULT '{}',
                    active_slot INTEGER,
                    created_at TEXT NOT NULL,
                    started_at TEXT,
                    finished_at TEXT,
                    updated_at TEXT NOT NULL,
                    created_by TEXT NOT NULL,
                    UNIQUE(user_key, idempotency_key),
                    UNIQUE(calculation_id)
                );
                CREATE UNIQUE INDEX IF NOT EXISTS inventory_balance_operations_active_slot
                ON sheet_vitrina_v1_inventory_balance_operations(active_slot)
                WHERE active_slot IS NOT NULL;
                CREATE INDEX IF NOT EXISTS inventory_balance_operations_user_created
                ON sheet_vitrina_v1_inventory_balance_operations(user_key, created_at DESC, operation_id DESC);
                CREATE TABLE IF NOT EXISTS sheet_vitrina_v1_inventory_balance_overrides (
                    calculation_id TEXT NOT NULL,
                    target_key TEXT NOT NULL,
                    nm_id INTEGER NOT NULL,
                    advert_id INTEGER NOT NULL,
                    placement TEXT NOT NULL,
                    calculated_target_bid_rub TEXT NOT NULL,
                    manual_target_bid_rub TEXT,
                    updated_at TEXT NOT NULL,
                    updated_by TEXT NOT NULL,
                    PRIMARY KEY(calculation_id, target_key),
                    FOREIGN KEY(calculation_id) REFERENCES sheet_vitrina_v1_inventory_balance_calculations(calculation_id)
                );
                CREATE TABLE IF NOT EXISTS sheet_vitrina_v1_inventory_balance_apply_jobs (
                    job_id TEXT PRIMARY KEY,
                    calculation_id TEXT NOT NULL,
                    mode TEXT NOT NULL,
                    state TEXT NOT NULL,
                    idempotency_key TEXT NOT NULL UNIQUE,
                    apply_manifest_digest TEXT NOT NULL,
                    apply_manifest_json TEXT NOT NULL,
                    selection_json TEXT NOT NULL,
                    summary_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    created_by TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(calculation_id) REFERENCES sheet_vitrina_v1_inventory_balance_calculations(calculation_id)
                );
                CREATE TABLE IF NOT EXISTS sheet_vitrina_v1_inventory_balance_apply_items (
                    job_id TEXT NOT NULL,
                    target_key TEXT NOT NULL,
                    nm_id INTEGER NOT NULL,
                    target_json TEXT NOT NULL,
                    state TEXT NOT NULL,
                    attempt_count INTEGER NOT NULL DEFAULT 0,
                    result_json TEXT NOT NULL DEFAULT '{}',
                    error TEXT NOT NULL DEFAULT '',
                    updated_at TEXT NOT NULL,
                    PRIMARY KEY(job_id, target_key),
                    FOREIGN KEY(job_id) REFERENCES sheet_vitrina_v1_inventory_balance_apply_jobs(job_id)
                );
                CREATE INDEX IF NOT EXISTS inventory_balance_apply_items_state
                ON sheet_vitrina_v1_inventory_balance_apply_items(job_id, state, target_key);
                CREATE TABLE IF NOT EXISTS sheet_vitrina_v1_inventory_balance_outcomes (
                    observation_id TEXT PRIMARY KEY,
                    calculation_id TEXT NOT NULL,
                    target_key TEXT NOT NULL,
                    observed_from TEXT,
                    observed_to TEXT,
                    observed_orders REAL,
                    observed_spend_rub REAL,
                    observed_cpo_rub REAL,
                    observed_drr REAL,
                    stockout_observed INTEGER,
                    outcome_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    source TEXT NOT NULL,
                    FOREIGN KEY(calculation_id) REFERENCES sheet_vitrina_v1_inventory_balance_calculations(calculation_id)
                );
                CREATE TRIGGER IF NOT EXISTS inventory_balance_calculations_no_update
                BEFORE UPDATE ON sheet_vitrina_v1_inventory_balance_calculations
                BEGIN SELECT RAISE(ABORT, 'inventory balance calculations are immutable'); END;
                CREATE TRIGGER IF NOT EXISTS inventory_balance_calculations_no_delete
                BEFORE DELETE ON sheet_vitrina_v1_inventory_balance_calculations
                BEGIN SELECT RAISE(ABORT, 'inventory balance calculations are immutable'); END;
                CREATE TRIGGER IF NOT EXISTS inventory_balance_outcomes_no_update
                BEFORE UPDATE ON sheet_vitrina_v1_inventory_balance_outcomes
                BEGIN SELECT RAISE(ABORT, 'inventory balance outcomes are append-only'); END;
                CREATE TRIGGER IF NOT EXISTS inventory_balance_outcomes_no_delete
                BEFORE DELETE ON sheet_vitrina_v1_inventory_balance_outcomes
                BEGIN SELECT RAISE(ABORT, 'inventory balance outcomes are append-only'); END;
                """
            )
            calculation_columns = {
                str(row[1])
                for row in conn.execute(
                    "PRAGMA table_info(sheet_vitrina_v1_inventory_balance_calculations)"
                ).fetchall()
            }
            if "operation_id" not in calculation_columns:
                conn.execute(
                    "ALTER TABLE sheet_vitrina_v1_inventory_balance_calculations "
                    "ADD COLUMN operation_id TEXT"
                )
            conn.execute(
                """CREATE UNIQUE INDEX IF NOT EXISTS inventory_balance_calculations_operation
                   ON sheet_vitrina_v1_inventory_balance_calculations(operation_id)
                   WHERE operation_id IS NOT NULL"""
            )
            job_columns = {
                str(row[1])
                for row in conn.execute(
                    "PRAGMA table_info(sheet_vitrina_v1_inventory_balance_apply_jobs)"
                ).fetchall()
            }
            for name, declaration in (
                ("phase", "TEXT NOT NULL DEFAULT 'queued'"),
                ("worker_token", "TEXT NOT NULL DEFAULT ''"),
                ("lease_expires_at", "TEXT NOT NULL DEFAULT ''"),
                ("started_at", "TEXT NOT NULL DEFAULT ''"),
                ("finished_at", "TEXT NOT NULL DEFAULT ''"),
                ("error_code", "TEXT NOT NULL DEFAULT ''"),
                ("error_message", "TEXT NOT NULL DEFAULT ''"),
            ):
                if name not in job_columns:
                    conn.execute(
                        "ALTER TABLE sheet_vitrina_v1_inventory_balance_apply_jobs "
                        f"ADD COLUMN {name} {declaration}"
                    )
            item_columns = {
                str(row[1])
                for row in conn.execute(
                    "PRAGMA table_info(sheet_vitrina_v1_inventory_balance_apply_items)"
                ).fetchall()
            }
            for name, declaration in (
                ("phase", "TEXT NOT NULL DEFAULT 'queued'"),
                ("submit_group", "INTEGER NOT NULL DEFAULT 0"),
                ("submitted_at", "TEXT NOT NULL DEFAULT ''"),
                ("readback_deadline_at", "TEXT NOT NULL DEFAULT ''"),
                ("registry_operation_id", "TEXT NOT NULL DEFAULT ''"),
                ("registry_receipt_reference", "TEXT NOT NULL DEFAULT ''"),
                ("error_code", "TEXT NOT NULL DEFAULT ''"),
                ("last_observed_bid_minor", "INTEGER"),
            ):
                if name not in item_columns:
                    conn.execute(
                        "ALTER TABLE sheet_vitrina_v1_inventory_balance_apply_items "
                        f"ADD COLUMN {name} {declaration}"
                    )
            conn.commit()

    def get_settings(self, *, user_key: str) -> dict[str, Any]:
        record = self.runtime.load_sheet_vitrina_user_config(
            user_key=user_key,
            config_key=CONFIG_KEY,
        )
        config = dict(record.get("config") or {}) if record.get("status") == "ok" else {}
        return {
            "status": "ok",
            "revision": int(record.get("revision") or 0),
            "updated_at": str(record.get("updated_at") or ""),
            "calculation": _sanitize_calculation_settings(config.get("calculation")),
            "table": _sanitize_table_preferences(config.get("table")),
            "canonical_store": "server_runtime_user_config",
        }

    def save_settings(self, payload: Mapping[str, Any], *, user_key: str) -> dict[str, Any]:
        calculation = _sanitize_calculation_settings(payload.get("calculation"))
        table = _sanitize_table_preferences(payload.get("table"))
        saved = self.runtime.save_sheet_vitrina_user_config(
            user_key=user_key,
            config_key=CONFIG_KEY,
            schema_version=CONFIG_SCHEMA_VERSION,
            payload={"calculation": calculation, "table": table},
            updated_at=self.timestamp_factory(),
            expected_revision=_optional_int(payload.get("base_revision")),
        )
        if saved.get("status") == "conflict":
            raise SkuInventoryBalanceError(
                "inventory balance settings revision conflict",
                http_status=409,
                payload=saved,
            )
        return self.get_settings(user_key=user_key)

    def start_calculation_operation(
        self,
        payload: Mapping[str, Any],
        *,
        user_key: str,
        actor: str,
    ) -> dict[str, Any]:
        operation_id = _operation_token(payload.get("operation_id"), "operation_id")
        idempotency_key = _operation_token(
            payload.get("idempotency_key"),
            "idempotency_key",
        )
        settings_payload = self.get_settings(user_key=user_key)
        settings = _sanitize_calculation_settings(
            payload.get("calculation")
            if isinstance(payload.get("calculation"), Mapping)
            else settings_payload.get("calculation")
        )
        request_payload = {"calculation": settings}
        request_json = _json(request_payload)
        request_digest = _digest(request_payload)
        created_at = self.timestamp_factory()
        acceptance_receipt = {
            "contract_name": CALCULATION_OPERATION_ACCEPTANCE_CONTRACT,
            "operation_contract": CALCULATION_OPERATION_CONTRACT,
            "operation_id": operation_id,
            "state": "accepted",
            "status_path": f"{CALCULATION_OPERATIONS_PATH}/{operation_id}",
            "accepted": True,
            "idempotent": True,
        }
        acceptance_json = _json(acceptance_receipt)
        acceptance_receipt = json.loads(acceptance_json)
        created = False
        with self._connect() as conn:
            conn.execute("BEGIN IMMEDIATE")
            existing = conn.execute(
                """SELECT operation_id,request_digest,acceptance_receipt_json
                   FROM sheet_vitrina_v1_inventory_balance_operations
                   WHERE user_key=? AND idempotency_key=?""",
                (user_key, idempotency_key),
            ).fetchone()
            if existing is not None:
                if str(existing["request_digest"]) != request_digest:
                    raise SkuInventoryBalanceError(
                        "Ключ операции уже использован с другими параметрами расчёта.",
                        http_status=409,
                        payload={"code": "idempotency_key_payload_mismatch"},
                    )
                conn.rollback()
                return json.loads(str(existing["acceptance_receipt_json"]))
            identity_collision = conn.execute(
                """SELECT user_key,idempotency_key FROM sheet_vitrina_v1_inventory_balance_operations
                   WHERE operation_id=?""",
                (operation_id,),
            ).fetchone()
            if identity_collision is not None:
                raise SkuInventoryBalanceError(
                    "Идентификатор операции уже занят другой операцией.",
                    http_status=409,
                    payload={"code": "operation_id_conflict"},
                )
            active = conn.execute(
                """SELECT operation_id FROM sheet_vitrina_v1_inventory_balance_operations
                   WHERE active_slot=1 LIMIT 1"""
            ).fetchone()
            if active is not None:
                raise SkuInventoryBalanceError(
                    "Другой расчёт баланса запасов уже выполняется. Дождитесь его завершения.",
                    http_status=409,
                    payload={"code": "calculation_operation_busy"},
                )
            conn.execute(
                """INSERT INTO sheet_vitrina_v1_inventory_balance_operations(
                       operation_id,user_key,idempotency_key,contract_name,request_digest,
                       request_json,acceptance_receipt_json,state,phase,progress_percent,
                       active_slot,created_at,updated_at,created_by
                   ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    operation_id,
                    user_key,
                    idempotency_key,
                    CALCULATION_OPERATION_CONTRACT,
                    request_digest,
                    request_json,
                    acceptance_json,
                    "accepted",
                    "accepted",
                    0,
                    1,
                    created_at,
                    created_at,
                    actor,
                ),
            )
            conn.commit()
            created = True
        self._log_calculation_operation(
            operation_id,
            phase="accepted",
            duration_ms=0,
            outcome="accepted",
        )
        if created and not self._start_calculation_worker(operation_id):
            self._fail_calculation_operation(
                operation_id,
                error_code="worker_capacity_unavailable",
                error_message="Не удалось запустить bounded worker. Создайте новую операцию.",
                release_slot=True,
            )
        return acceptance_receipt

    def get_calculation_operation(
        self,
        operation_id: str,
        *,
        user_key: str,
    ) -> dict[str, Any]:
        normalized_id = _operation_token(operation_id, "operation_id")
        with self._connect() as conn:
            row = conn.execute(
                """SELECT * FROM sheet_vitrina_v1_inventory_balance_operations
                   WHERE operation_id=? AND user_key=?""",
                (normalized_id, user_key),
            ).fetchone()
        if row is None:
            raise SkuInventoryBalanceError(
                "Операция расчёта не найдена.",
                http_status=404,
                payload={"code": "calculation_operation_not_found"},
            )
        return self._calculation_operation_payload(row)

    def latest_calculation_operation(self, *, user_key: str) -> dict[str, Any] | None:
        with self._connect() as conn:
            row = conn.execute(
                """SELECT * FROM sheet_vitrina_v1_inventory_balance_operations
                   WHERE user_key=? ORDER BY created_at DESC,operation_id DESC LIMIT 1""",
                (user_key,),
            ).fetchone()
        return self._calculation_operation_payload(row) if row is not None else None

    def _start_calculation_worker(self, operation_id: str) -> bool:
        with self._calculation_worker_lock:
            active = self._calculation_worker_thread
            if active is not None and active.is_alive():
                return self._calculation_worker_operation_id == operation_id
            worker = threading.Thread(
                target=self._execute_calculation_operation,
                args=(operation_id,),
                name="sku-inventory-balance-operation",
                daemon=True,
            )
            self._calculation_worker_operation_id = operation_id
            self._calculation_worker_thread = worker
            worker.start()
            return True

    def _execute_calculation_operation(self, operation_id: str) -> None:
        started_monotonic = time.monotonic()
        try:
            started_at = self.timestamp_factory()
            with self._connect() as conn:
                claimed = conn.execute(
                    """UPDATE sheet_vitrina_v1_inventory_balance_operations
                       SET state='running',phase='building_evidence',progress_percent=10,
                           started_at=?,updated_at=?
                       WHERE operation_id=? AND state='accepted' AND active_slot=1""",
                    (started_at, started_at, operation_id),
                ).rowcount
                conn.commit()
                if claimed != 1:
                    return
                row = conn.execute(
                    """SELECT request_json,user_key,created_by
                       FROM sheet_vitrina_v1_inventory_balance_operations
                       WHERE operation_id=?""",
                    (operation_id,),
                ).fetchone()
            if row is None:
                return
            self._log_calculation_operation(
                operation_id,
                phase="building_evidence",
                duration_ms=0,
                outcome="running",
            )
            request_payload = json.loads(str(row["request_json"]))
            self.calculate(
                request_payload,
                user_key=str(row["user_key"]),
                actor=str(row["created_by"]),
                operation_id=operation_id,
            )
            duration_ms = int((time.monotonic() - started_monotonic) * 1000)
            self._log_calculation_operation(
                operation_id,
                phase="succeeded",
                duration_ms=duration_ms,
                outcome="calculation_created",
            )
        except Exception:
            self._fail_calculation_operation(
                operation_id,
                error_code="calculation_failed",
                error_message=(
                    "Расчёт не завершён. Результат не создан; повторите действие новой операцией."
                ),
            )
            duration_ms = int((time.monotonic() - started_monotonic) * 1000)
            self._log_calculation_operation(
                operation_id,
                phase="failed",
                duration_ms=duration_ms,
                outcome="failed_before_calculation",
            )
        finally:
            with self._calculation_worker_lock:
                if self._calculation_worker_operation_id == operation_id:
                    self._calculation_worker_operation_id = ""
                    self._calculation_worker_thread = None
            self._release_calculation_operation_slot(operation_id)

    def _fail_calculation_operation(
        self,
        operation_id: str,
        *,
        error_code: str,
        error_message: str,
        release_slot: bool = False,
    ) -> None:
        finished_at = self.timestamp_factory()
        with self._connect() as conn:
            conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_operations
                   SET state='failed',phase='failed',progress_percent=100,
                       error_code=?,error_message=?,outcome_json=?,active_slot=?,
                       finished_at=?,updated_at=?
                   WHERE operation_id=? AND state IN ('accepted','running')
                         AND calculation_id IS NULL""",
                (
                    error_code,
                    error_message,
                    _json({"durable_outcome": "no_calculation_created", "retryable": True}),
                    None if release_slot else 1,
                    finished_at,
                    finished_at,
                    operation_id,
                ),
            )
            conn.commit()

    def _release_calculation_operation_slot(self, operation_id: str) -> None:
        with self._connect() as conn:
            conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_operations
                   SET active_slot=NULL
                   WHERE operation_id=? AND state IN ('succeeded','failed')""",
                (operation_id,),
            )
            conn.commit()

    def _terminalize_interrupted_calculation_operations(self) -> None:
        with self._connect() as conn:
            conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_operations
                   SET active_slot=NULL WHERE state IN ('succeeded','failed')"""
            )
            conn.commit()
            active_rows = conn.execute(
                """SELECT operation_id FROM sheet_vitrina_v1_inventory_balance_operations
                   WHERE state IN ('accepted','running') AND active_slot=1"""
            ).fetchall()
        if not active_rows:
            return
        for row in active_rows:
            operation_id = str(row["operation_id"])
            self._fail_calculation_operation(
                operation_id,
                error_code="runtime_interrupted",
                error_message=(
                    "Процесс расчёта был прерван до создания результата. "
                    "Запустите новый расчёт новой операцией."
                ),
                release_slot=True,
            )
            self._log_calculation_operation(
                operation_id,
                phase="failed",
                duration_ms=0,
                outcome="runtime_interrupted",
            )

    def _calculation_operation_payload(self, row: sqlite3.Row) -> dict[str, Any]:
        state = str(row["state"])
        calculation_id = str(row["calculation_id"] or "")
        started_at = str(row["started_at"] or "")
        finished_at = str(row["finished_at"] or "")
        result = self.get_calculation(calculation_id) if state == "succeeded" and calculation_id else None
        return {
            "contract_name": CALCULATION_OPERATION_CONTRACT,
            "operation_id": str(row["operation_id"]),
            "state": state,
            "phase": str(row["phase"]),
            "progress": {
                "percent": int(row["progress_percent"] or 0),
                "terminal": state in {"succeeded", "failed"},
            },
            "calculation_id": calculation_id or None,
            "result": result,
            "error": (
                {
                    "code": str(row["error_code"]),
                    "message": str(row["error_message"]),
                }
                if row["error_code"]
                else None
            ),
            "outcome": json.loads(str(row["outcome_json"] or "{}")),
            "created_at": str(row["created_at"]),
            "started_at": started_at,
            "finished_at": finished_at,
            "duration_ms": _iso_duration_ms(started_at, finished_at),
            "updated_at": str(row["updated_at"]),
            "retryable_by_new_operation": state == "failed" and not calculation_id,
            "blind_resubmit_allowed": False,
        }

    @staticmethod
    def _log_calculation_operation(
        operation_id: str,
        *,
        phase: str,
        duration_ms: int,
        outcome: str,
    ) -> None:
        _LOGGER.info(
            "inventory_balance_operation id=%s phase=%s duration_ms=%d outcome=%s",
            operation_id[:48],
            phase,
            max(int(duration_ms), 0),
            outcome,
        )

    def latest(self, *, user_key: str) -> dict[str, Any]:
        settings = self.get_settings(user_key=user_key)
        with self._connect() as conn:
            row = conn.execute(
                """SELECT * FROM sheet_vitrina_v1_inventory_balance_calculations
                   ORDER BY created_at DESC, calculation_id DESC LIMIT 1"""
            ).fetchone()
            job_row = conn.execute(
                """SELECT job_id FROM sheet_vitrina_v1_inventory_balance_apply_jobs
                   WHERE created_by=? ORDER BY created_at DESC, job_id DESC LIMIT 1""",
                (user_key,),
            ).fetchone()
        return {
            "status": "ok",
            "settings": settings,
            "calculation": self._calculation_payload(row) if row is not None else None,
            "apply_job": (
                self.get_apply_job(str(job_row["job_id"]))
                if job_row is not None
                else None
            ),
            "apply_capability": self._apply_capability(),
            "calculation_operation": self.latest_calculation_operation(user_key=user_key),
        }

    def list_registry(self, *, limit: int = 20) -> dict[str, Any]:
        normalized_limit = min(max(int(limit), 1), 100)
        with self._connect() as conn:
            calculations = conn.execute(
                """SELECT * FROM sheet_vitrina_v1_inventory_balance_calculations
                   ORDER BY created_at DESC,calculation_id DESC LIMIT ?""",
                (normalized_limit,),
            ).fetchall()
            items = []
            for calculation in calculations:
                jobs = conn.execute(
                    """SELECT j.job_id,j.mode,j.state,j.apply_manifest_digest,
                              j.created_at,j.created_by,j.updated_at,
                              COUNT(i.target_key) AS target_count,
                              SUM(CASE WHEN i.state IN ('succeeded','failed','skipped','ambiguous') THEN 1 ELSE 0 END) AS terminal_count
                       FROM sheet_vitrina_v1_inventory_balance_apply_jobs j
                       LEFT JOIN sheet_vitrina_v1_inventory_balance_apply_items i ON i.job_id=j.job_id
                       WHERE j.calculation_id=?
                       GROUP BY j.job_id
                       ORDER BY j.created_at DESC,j.job_id DESC""",
                    (str(calculation["calculation_id"]),),
                ).fetchall()
                immutable = json.loads(str(calculation["payload_json"]))
                items.append(
                    {
                        "calculation_id": str(calculation["calculation_id"]),
                        "created_at": str(calculation["created_at"]),
                        "created_by": str(calculation["created_by"]),
                        "previous_calculation_id": calculation["previous_calculation_id"],
                        "formula_version": str(calculation["formula_version"]),
                        "source_digest": str(calculation["source_digest"]),
                        "row_count": len(immutable.get("rows") or []),
                        "apply_protocols": list(immutable.get("apply_protocols") or []),
                        "apply_jobs": [
                            {
                                "job_id": str(job["job_id"]),
                                "mode": str(job["mode"]),
                                "state": str(job["state"]),
                                "apply_manifest_digest": str(job["apply_manifest_digest"]),
                                "target_count": int(job["target_count"] or 0),
                                "terminal_count": int(job["terminal_count"] or 0),
                                "created_at": str(job["created_at"]),
                                "created_by": str(job["created_by"]),
                                "updated_at": str(job["updated_at"]),
                            }
                            for job in jobs
                        ],
                    }
                )
        return {
            "contract_name": "sheet_vitrina_v1_inventory_balance_registry/v1",
            "limit": normalized_limit,
            "items": items,
            "immutable_calculations": True,
            "apply_jobs_linked": True,
        }

    def calculate(
        self,
        payload: Mapping[str, Any],
        *,
        user_key: str,
        actor: str,
        operation_id: str | None = None,
    ) -> dict[str, Any]:
        normalized_operation_id = (
            _operation_token(operation_id, "operation_id") if operation_id else None
        )
        settings_payload = self.get_settings(user_key=user_key)
        settings = _sanitize_calculation_settings(
            payload.get("calculation")
            if isinstance(payload.get("calculation"), Mapping)
            else settings_payload.get("calculation")
        )
        source = self.sku_management_block.build_inventory_balance_evidence(
            user_key=user_key,
            sales_period_days=int(settings["sales_period_days"]),
        )
        as_of_date = str(
            ((source.get("meta") or {}).get("metric_policy") or {}).get("business_date")
            or current_business_date_iso(self.now_factory())
        )
        date_to = (date.fromisoformat(as_of_date) - timedelta(days=1)).isoformat()
        date_from = (
            date.fromisoformat(date_to) - timedelta(days=settings["sales_period_days"] - 1)
        ).isoformat()
        rows: list[dict[str, Any]] = []
        source_errors: list[dict[str, Any]] = []
        excluded_rows: list[dict[str, Any]] = []
        for raw in source.get("rows") or []:
            source_nm_id = int(raw.get("nm_id") or 0)
            if source_nm_id in EXCLUDED_NM_IDS:
                excluded_rows.append(
                    {
                        "nm_id": source_nm_id,
                        "reason": EXCLUDED_NM_IDS[source_nm_id],
                        "policy_version": EXCLUSION_POLICY_VERSION,
                    }
                )
                continue
            row = calculate_inventory_balance_row(raw, settings=settings)
            try:
                ads_detail = self.sku_management_block.ads_block.build_sku_detail(
                    int(row["nm_id"]),
                    params={"date_from": date_from, "date_to": date_to},
                )
                campaign_rows = list(ads_detail.get("rows") or [])
                campaign_meta = {
                    **dict(ads_detail.get("meta") or {}),
                    "period": dict(ads_detail.get("period") or {}),
                }
            except Exception as exc:
                campaign_rows = list(raw.get("ad_options") or [])
                campaign_meta = {"stats_status": "error", "stats_error": str(exc)}
                source_errors.append({"nm_id": row["nm_id"], "source": "ads_detail", "error": str(exc)})
            recommendations = [
                _campaign_recommendation(
                    item,
                    nm_id=int(row["nm_id"]),
                    pace_ratio=row.get("pace_ratio"),
                )
                for item in campaign_rows
            ]
            recommendations = _allocate_campaign_targets(
                recommendations,
                pace_ratio=row.get("pace_ratio"),
            )
            row["campaign_recommendations"] = recommendations
            row["new_cpc_campaigns"] = [
                item for item in recommendations if item["campaign_group"] == "new_cpc"
            ]
            row["old_cpm_campaigns"] = [
                item for item in recommendations if item["campaign_group"] == "old_cpm"
            ]
            row["select_available"] = any(
                item["can_apply"] or item.get("state_action_available")
                for item in recommendations
            )
            row["ads_evidence"] = campaign_meta
            row["outcome_observation"] = {
                "status": "not_observed",
                "observed_from": None,
                "observed_to": None,
                "orders": None,
                "spend_rub": None,
                "cpo_rub": None,
                "drr": None,
                "stockout_observed": None,
                "automatic_training": False,
            }
            rows.append(row)
        rows.sort(key=_row_sort_key)
        now = self.timestamp_factory()
        with self._connect() as conn:
            previous = conn.execute(
                """SELECT calculation_id FROM sheet_vitrina_v1_inventory_balance_calculations
                   ORDER BY created_at DESC, calculation_id DESC LIMIT 1"""
            ).fetchone()
            previous_id = str(previous["calculation_id"]) if previous else ""
            calculation_id = f"ibc_{uuid4().hex}"
            immutable = {
                "contract_name": CALCULATION_CONTRACT,
                "formula_version": FORMULA_VERSION,
                "calculation_id": calculation_id,
                "operation_id": normalized_operation_id,
                "previous_calculation_id": previous_id or None,
                "created_at": now,
                "created_by": actor,
                "as_of_date": as_of_date,
                "ads_period": {"date_from": date_from, "date_to": date_to},
                "settings": settings,
                "rows": rows,
                "source_errors": source_errors,
                "source_contract": str(source.get("contract_name") or ""),
                "source_generated_at": str(source.get("generated_at") or ""),
                "sales_evidence_window": dict(
                    ((source.get("meta") or {}).get("inventory_balance_evidence") or {})
                ),
                "lineage": {
                    "previous_calculation_id": previous_id or None,
                    "comparison_status": "available" if previous_id else "first_calculation",
                    "sales_evidence_window": {
                        key: value
                        for key, value in dict(
                            ((source.get("meta") or {}).get("inventory_balance_evidence") or {})
                        ).items()
                        if key != "supplier_eta"
                    },
                    "supplier_eta_evidence": dict(
                        (((source.get("meta") or {}).get("inventory_balance_evidence") or {}).get("supplier_eta") or {})
                    ),
                    "exclusion_policy": {
                        "version": EXCLUSION_POLICY_VERSION,
                        "excluded_nm_ids": sorted(EXCLUDED_NM_IDS),
                        "matched_rows": excluded_rows,
                        "identity_rule": "exact_nm_id_only",
                    },
                    "wb_stock_evidence": [
                        {
                            "nm_id": int(item.get("nm_id") or 0),
                            **dict(item.get("wb_stock_evidence") or {}),
                        }
                        for item in rows
                    ],
                },
                "excluded_rows": excluded_rows,
                "apply_protocols": self._apply_protocols(),
                "automatic_ml_or_training": False,
            }
            digest = _digest(
                {
                    "settings": settings,
                    "source_generated_at": immutable["source_generated_at"],
                    "rows": rows,
                }
            )
            immutable["source_digest"] = digest
            if normalized_operation_id is not None:
                operation = conn.execute(
                    """SELECT state,calculation_id FROM sheet_vitrina_v1_inventory_balance_operations
                       WHERE operation_id=? AND user_key=?""",
                    (normalized_operation_id, user_key),
                ).fetchone()
                if (
                    operation is None
                    or str(operation["state"]) != "running"
                    or operation["calculation_id"] is not None
                ):
                    raise SkuInventoryBalanceError(
                        "Операция расчёта потеряла exact running identity.",
                        http_status=409,
                        payload={"code": "calculation_operation_state_conflict"},
                    )
            conn.execute(
                """INSERT INTO sheet_vitrina_v1_inventory_balance_calculations(
                       calculation_id,operation_id,previous_calculation_id,contract_name,formula_version,
                       source_digest,settings_json,payload_json,created_at,created_by
                   ) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (
                    calculation_id,
                    normalized_operation_id,
                    previous_id or None,
                    CALCULATION_CONTRACT,
                    FORMULA_VERSION,
                    digest,
                    _json(settings),
                    _json(immutable),
                    now,
                    actor,
                ),
            )
            if normalized_operation_id is not None:
                updated = conn.execute(
                    """UPDATE sheet_vitrina_v1_inventory_balance_operations
                       SET state='succeeded',phase='succeeded',progress_percent=100,
                           calculation_id=?,error_code='',error_message='',outcome_json=?,
                           finished_at=?,updated_at=?
                       WHERE operation_id=? AND user_key=? AND state='running'
                             AND calculation_id IS NULL""",
                    (
                        calculation_id,
                        _json(
                            {
                                "durable_outcome": "calculation_created",
                                "calculation_id": calculation_id,
                            }
                        ),
                        now,
                        now,
                        normalized_operation_id,
                        user_key,
                    ),
                ).rowcount
                if updated != 1:
                    raise SkuInventoryBalanceError(
                        "Операция расчёта не приняла exact calculation result.",
                        http_status=409,
                        payload={"code": "calculation_operation_result_conflict"},
                    )
            conn.commit()
        return self.get_calculation(calculation_id)

    def get_calculation(self, calculation_id: str) -> dict[str, Any]:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM sheet_vitrina_v1_inventory_balance_calculations WHERE calculation_id=?",
                (str(calculation_id),),
            ).fetchone()
        if row is None:
            raise SkuInventoryBalanceError("inventory balance calculation not found", http_status=404)
        return self._calculation_payload(row)

    def save_override(
        self,
        calculation_id: str,
        payload: Mapping[str, Any],
        *,
        actor: str,
    ) -> dict[str, Any]:
        target_key = str(payload.get("target_key") or "").strip()
        calculation = self.get_calculation(calculation_id)
        target = _find_target(calculation, target_key)
        if target is None:
            raise SkuInventoryBalanceError("calculation target not found", http_status=404)
        if not target.get("manual_override_allowed"):
            raise SkuInventoryBalanceError(
                "manual target override is unavailable without complete inventory pacing evidence",
                http_status=422,
            )
        raw_manual = payload.get("manual_target_bid_rub")
        manual = None if raw_manual in {None, ""} else _positive_money(raw_manual, "manual_target_bid_rub")
        now = self.timestamp_factory()
        with self._connect() as conn:
            conn.execute(
                """INSERT INTO sheet_vitrina_v1_inventory_balance_overrides(
                       calculation_id,target_key,nm_id,advert_id,placement,
                       calculated_target_bid_rub,manual_target_bid_rub,updated_at,updated_by
                   ) VALUES(?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(calculation_id,target_key) DO UPDATE SET
                       manual_target_bid_rub=excluded.manual_target_bid_rub,
                       updated_at=excluded.updated_at,updated_by=excluded.updated_by""",
                (
                    calculation_id,
                    target_key,
                    int(target["nm_id"]),
                    int(target["advert_id"]),
                    str(target["placement"]),
                    str(target["calculated_target_bid_rub"]),
                    None if manual is None else str(manual),
                    now,
                    actor,
                ),
            )
            conn.commit()
        return self.get_calculation(calculation_id)

    def start_apply(self, payload: Mapping[str, Any], *, actor: str) -> dict[str, Any]:
        calculation_id = str(payload.get("calculation_id") or "").strip()
        if not calculation_id:
            raise SkuInventoryBalanceError("calculation_id is required")
        if payload.get("confirmed") is not True:
            raise SkuInventoryBalanceError("explicit confirmation is required", http_status=409)
        requested_mode = str(payload.get("mode") or DRY_RUN_MODE)
        if requested_mode not in {DRY_RUN_MODE, LIVE_MODE}:
            raise SkuInventoryBalanceError(
                "unsupported inventory-balance apply mode", http_status=422
            )
        if requested_mode == LIVE_MODE and (
            self.apply_adapter.mode != LIVE_MODE
            or not self.apply_adapter.external_writes_enabled
            or self.writer_registry is None
        ):
            raise SkuInventoryBalanceError(
                "live WB inventory-balance apply is unavailable", http_status=503
            )
        calculation = self.get_calculation(calculation_id)
        selected_nm_ids = {
            int(item) for item in (payload.get("nm_ids") or []) if _optional_int(item)
        }
        selected_target_keys = {
            str(item) for item in (payload.get("target_keys") or []) if str(item).strip()
        }
        raw_state_actions = payload.get("state_actions") or []
        if not isinstance(raw_state_actions, Sequence) or isinstance(
            raw_state_actions, (str, bytes)
        ):
            raise SkuInventoryBalanceError(
                "state_actions must be an array", http_status=422
            )
        if not selected_nm_ids and not selected_target_keys and not raw_state_actions:
            raise SkuInventoryBalanceError(
                "at least one selected bid or campaign state action is required",
                http_status=422,
            )
        bid_targets = []
        if selected_nm_ids or selected_target_keys:
            for row in calculation.get("rows") or []:
                if selected_nm_ids and int(row["nm_id"]) not in selected_nm_ids:
                    continue
                for target in row.get("campaign_recommendations") or []:
                    if selected_target_keys and str(target["target_key"]) not in selected_target_keys:
                        continue
                    if target.get("can_apply"):
                        bid_targets.append(dict(target))
        campaign_index: dict[tuple[int, int], dict[str, Any]] = {}
        campaign_placements: dict[tuple[int, int], set[str]] = {}
        for row in calculation.get("rows") or []:
            for target in row.get("campaign_recommendations") or []:
                identity = (int(target["nm_id"]), int(target["advert_id"]))
                existing = campaign_index.get(identity)
                if existing is not None and (
                    int(existing.get("campaign_status") or 0)
                    != int(target.get("campaign_status") or 0)
                    or str(existing.get("payment_type") or "")
                    != str(target.get("payment_type") or "")
                ):
                    raise SkuInventoryBalanceError(
                        "calculation contains conflicting campaign identity evidence",
                        http_status=409,
                    )
                campaign_index[identity] = dict(target)
                placement_evidence = str(target.get("placement") or "")
                if placement_evidence:
                    campaign_placements.setdefault(identity, set()).add(
                        placement_evidence
                    )
        state_targets: list[dict[str, Any]] = []
        seen_state_adverts: set[int] = set()
        for raw_action in raw_state_actions:
            if not isinstance(raw_action, Mapping):
                raise SkuInventoryBalanceError(
                    "campaign state action must be an object", http_status=422
                )
            nm_id = _optional_int(raw_action.get("nm_id")) or 0
            advert_id = _optional_int(raw_action.get("advert_id")) or 0
            action = str(raw_action.get("action") or "").strip().lower()
            if nm_id <= 0 or advert_id <= 0 or action not in {"start", "pause"}:
                raise SkuInventoryBalanceError(
                    "campaign state action identity is invalid", http_status=422
                )
            if advert_id in seen_state_adverts:
                raise SkuInventoryBalanceError(
                    "campaign state action is duplicated", http_status=422
                )
            seen_state_adverts.add(advert_id)
            source_target = campaign_index.get((nm_id, advert_id))
            if source_target is None or not source_target.get("identity_valid"):
                raise SkuInventoryBalanceError(
                    "campaign state action has no exact calculation identity",
                    http_status=409,
                )
            current_status = int(source_target.get("campaign_status") or 0)
            expected_action = CAMPAIGN_STATE_ACTION_BY_STATUS.get(current_status, "")
            if action != expected_action:
                raise SkuInventoryBalanceError(
                    "campaign state action is unavailable from current state",
                    http_status=409,
                )
            current_state = CAMPAIGN_STATE_BY_STATUS[current_status]
            requested_state = "active" if action == "start" else "paused"
            state_target_key = f"state:{nm_id}:{advert_id}"
            recommendation_basis = {
                "contract": "sku_inventory_balance_campaign_state_recommendation/v1",
                "calculation_id": calculation_id,
                "target": {
                    "seller_id": self.seller_id,
                    "account_scope": self.account_scope,
                    "target_kind": "campaign",
                    "nm_id": nm_id,
                    "advert_id": advert_id,
                    "placement": "",
                    "parameter_field": "campaign_state",
                },
                "before_value": current_state,
                "requested_value": requested_state,
            }
            state_targets.append(
                {
                    "action_type": "campaign_state",
                    "target_key": state_target_key,
                    "nm_id": nm_id,
                    "advert_id": advert_id,
                    "campaign_name": str(source_target.get("campaign_name") or ""),
                    "payment_type": str(source_target.get("payment_type") or ""),
                    "placement": "",
                    "placement_evidence": sorted(campaign_placements[(nm_id, advert_id)])[0],
                    "current_campaign_status": current_status,
                    "current_campaign_state": current_state,
                    "requested_campaign_state": requested_state,
                    "state_action": action,
                    "state_action_label": str(source_target.get("state_action_label") or action),
                    "recommendation_item_id": "ibsr_"
                    + sha256(
                        json.dumps(
                            recommendation_basis,
                            ensure_ascii=False,
                            sort_keys=True,
                            separators=(",", ":"),
                        ).encode("utf-8")
                    ).hexdigest(),
                }
            )
        targets = [
            {**item, "action_type": "bid_change"} for item in bid_targets
        ] + state_targets
        if not targets:
            raise SkuInventoryBalanceError(
                "selection has no valid bid or campaign state changes",
                http_status=422,
            )
        selection = {
            "nm_ids": sorted({int(item["nm_id"]) for item in targets}),
            "target_keys": sorted(str(item["target_key"]) for item in targets),
            "bid_target_keys": sorted(
                str(item["target_key"])
                for item in targets
                if item["action_type"] == "bid_change"
            ),
            "campaign_state_target_keys": sorted(
                str(item["target_key"])
                for item in targets
                if item["action_type"] == "campaign_state"
            ),
        }
        confirmation_policy = self._balance_owner_confirmation_policy()
        job_targets = []
        manifest_targets = []
        for item in targets:
            safety_warnings = (
                self._balance_safety_threshold_warnings(item)
                if item["action_type"] == "bid_change"
                else []
            )
            job_targets.append(
                {
                    **item,
                    "owner_confirmation_policy": confirmation_policy[
                        "contract_name"
                    ],
                    "safety_warnings": safety_warnings,
                }
            )
            if item["action_type"] == "bid_change":
                manifest_target = {
                    "action_type": "bid_change",
                    "target_key": str(item["target_key"]),
                    "nm_id": int(item["nm_id"]),
                    "advert_id": int(item["advert_id"]),
                    "placement": str(item["placement"]),
                    "payment_type": str(item.get("payment_type") or ""),
                    "current_bid_rub": item.get("current_bid_rub"),
                    "current_bid_minor": int(item["current_bid_minor"]),
                    "calculated_target_bid_rub": item.get("calculated_target_bid_rub"),
                    "manual_target_bid_rub": item.get("manual_target_bid_rub"),
                    "final_target_bid_rub": item.get("final_target_bid_rub"),
                    "final_target_bid_minor": int(item["final_target_bid_minor"]),
                    "recommendation_item_id": str(item["recommendation_item_id"]),
                    "override_updated_at": item.get("override_updated_at") or "",
                    "safety_warnings": safety_warnings,
                }
            else:
                manifest_target = {
                    "action_type": "campaign_state",
                    "target_key": str(item["target_key"]),
                    "nm_id": int(item["nm_id"]),
                    "advert_id": int(item["advert_id"]),
                    "placement": "",
                    "placement_evidence": str(item["placement_evidence"]),
                    "payment_type": str(item["payment_type"]),
                    "current_campaign_status": int(item["current_campaign_status"]),
                    "current_campaign_state": str(item["current_campaign_state"]),
                    "requested_campaign_state": str(item["requested_campaign_state"]),
                    "state_action": str(item["state_action"]),
                    "recommendation_item_id": str(item["recommendation_item_id"]),
                    "safety_warnings": [],
                }
            manifest_targets.append(manifest_target)
        manifest_targets.sort(key=lambda item: item["target_key"])
        apply_manifest = {
            "contract_name": "sheet_vitrina_v1_inventory_balance_apply_manifest/v1",
            "calculation_id": calculation_id,
            "mode": requested_mode,
            "targets": manifest_targets,
            "external_writes": requested_mode == LIVE_MODE,
            "batch_size": self.live_batch_size if requested_mode == LIVE_MODE else 0,
            "canary_required": requested_mode == LIVE_MODE,
            "owner_confirmation": {
                **confirmation_policy,
                "confirmed": True,
                "scope": "exact_immutable_manifest_targets",
            },
        }
        manifest_digest = _digest(apply_manifest)
        idempotency_key = manifest_digest
        now = self.timestamp_factory()
        job_id = f"ibj_{uuid4().hex}"
        summary = {
            "sku_count": len(selection["nm_ids"]),
            "target_count": len(targets),
            "bid_count": len(bid_targets),
            "campaign_state_count": len(state_targets),
            "increase_count": sum(
                1 for item in bid_targets if float(item["final_target_bid_rub"]) > float(item["current_bid_rub"])
            ),
            "decrease_count": sum(
                1 for item in bid_targets if float(item["final_target_bid_rub"]) < float(item["current_bid_rub"])
            ),
            "unchanged_count": sum(
                1 for item in bid_targets if float(item["final_target_bid_rub"]) == float(item["current_bid_rub"])
            ),
            "external_writes": requested_mode == LIVE_MODE,
        }
        with self._connect() as conn:
            existing = conn.execute(
                "SELECT job_id FROM sheet_vitrina_v1_inventory_balance_apply_jobs WHERE idempotency_key=?",
                (idempotency_key,),
            ).fetchone()
            if existing is not None:
                existing_job_id = str(existing["job_id"])
                if requested_mode == LIVE_MODE:
                    self._start_apply_worker_if_needed()
                return self.get_apply_job(existing_job_id)
            conn.execute(
                """INSERT INTO sheet_vitrina_v1_inventory_balance_apply_jobs(
                       job_id,calculation_id,mode,state,idempotency_key,
                       apply_manifest_digest,apply_manifest_json,selection_json,
                       summary_json,created_at,created_by,updated_at
                   ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    job_id,
                    calculation_id,
                    requested_mode,
                    "pending",
                    idempotency_key,
                    manifest_digest,
                    _json(apply_manifest),
                    _json(selection),
                    _json(summary),
                    now,
                    actor,
                    now,
                ),
            )
            for target in job_targets:
                conn.execute(
                    """INSERT INTO sheet_vitrina_v1_inventory_balance_apply_items(
                           job_id,target_key,nm_id,target_json,state,updated_at
                       ) VALUES(?,?,?,?,?,?)""",
                    (
                        job_id,
                        str(target["target_key"]),
                        int(target["nm_id"]),
                        _json(target),
                        "pending",
                        now,
                    ),
                )
            conn.commit()
        if requested_mode == LIVE_MODE:
            self._start_apply_worker_if_needed()
        return self.get_apply_job(job_id)

    def start_manual_pending(
        self, payload: Mapping[str, Any], *, actor: str
    ) -> dict[str, Any]:
        if self.manual_pending_registry is None or not self.seller_id:
            raise SkuInventoryBalanceError(
                "manual portal fallback is unavailable", http_status=503
            )
        if payload.get("confirmed") is not True:
            raise SkuInventoryBalanceError(
                "explicit confirmation is required", http_status=409
            )
        calculation_id = str(payload.get("calculation_id") or "").strip()
        if not calculation_id:
            raise SkuInventoryBalanceError("calculation_id is required")
        calculation = self.get_calculation(calculation_id)
        selected_ids = {
            str(value).strip()
            for value in (payload.get("recommendation_item_ids") or [])
            if str(value).strip()
        }
        if not selected_ids:
            raise SkuInventoryBalanceError(
                "at least one recommendation_item_id is required", http_status=422
            )
        recommendations: list[dict[str, Any]] = []
        for row in calculation.get("rows") or []:
            for target in row.get("campaign_recommendations") or []:
                if str(target.get("recommendation_item_id") or "") not in selected_ids:
                    continue
                if not target.get("can_apply"):
                    raise SkuInventoryBalanceError(
                        "selected recommendation is not actionable", http_status=409
                    )
                recommendations.append(
                    {
                        "recommendation_item_id": target["recommendation_item_id"],
                        "action_type": target["action_type"],
                        "target": target["exact_target"],
                        "before_value": target["current_bid_minor"],
                        "requested_value": target["final_target_bid_minor"],
                    }
                )
        if {item["recommendation_item_id"] for item in recommendations} != selected_ids:
            raise SkuInventoryBalanceError(
                "selection contains unknown recommendation_item_id", http_status=404
            )
        receipt = self.manual_pending_registry.register_manual_pending(
            seller_id=self.seller_id,
            account_scope=self.account_scope,
            calculation_id=calculation_id,
            recommendations=recommendations,
            actor_principal=actor,
            requested_at=self.timestamp_factory(),
        )
        receipt["boundary"] = {
            "wb_upload_task_calls": 0,
            "wb_patch_bids_calls": 0,
            "balance_live_apply": False,
        }
        return receipt

    def resume_apply(self, job_id: str, *, actor: str, limit: int = 10) -> dict[str, Any]:
        existing = self.get_apply_job(job_id)
        if existing["mode"] == LIVE_MODE:
            if existing["state"] == "stalled":
                now = self.timestamp_factory()
                with self._connect() as conn:
                    resumed = conn.execute(
                        """UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs
                           SET state='running',phase='resume_requested',error_code='',
                               error_message='',worker_token='',lease_expires_at='',updated_at=?
                           WHERE job_id=? AND (
                               state='stalled'
                               OR (state IN ('pending','running','delayed')
                                   AND (lease_expires_at='' OR lease_expires_at<?))
                           )""",
                        (now, job_id, now),
                    ).rowcount
                    conn.commit()
                if resumed != 1:
                    raise SkuInventoryBalanceError(
                        "live apply job is active under a valid worker lease",
                        http_status=409,
                    )
            self._start_apply_worker_if_needed()
            return self.get_apply_job(job_id)
        limit = min(max(int(limit), 1), 50)
        self._terminalize_stale_running_items(job_id)
        for _ in range(limit):
            with self._connect() as conn:
                conn.execute("BEGIN IMMEDIATE")
                job = conn.execute(
                    "SELECT * FROM sheet_vitrina_v1_inventory_balance_apply_jobs WHERE job_id=?",
                    (job_id,),
                ).fetchone()
                if job is None:
                    conn.rollback()
                    raise SkuInventoryBalanceError("inventory balance apply job not found", http_status=404)
                if str(job["mode"]) != DRY_RUN_MODE:
                    conn.rollback()
                    raise SkuInventoryBalanceError("non-dry-run job is unreachable", http_status=403)
                item = conn.execute(
                    """SELECT * FROM sheet_vitrina_v1_inventory_balance_apply_items
                       WHERE job_id=? AND state='pending' ORDER BY target_key LIMIT 1""",
                    (job_id,),
                ).fetchone()
                if item is None:
                    conn.commit()
                    break
                now = self.timestamp_factory()
                conn.execute(
                    """UPDATE sheet_vitrina_v1_inventory_balance_apply_items
                       SET state='running',attempt_count=attempt_count+1,updated_at=?
                       WHERE job_id=? AND target_key=? AND state='pending'""",
                    (now, job_id, str(item["target_key"])),
                )
                conn.execute(
                    "UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs SET state='running',updated_at=? WHERE job_id=?",
                    (now, job_id),
                )
                conn.commit()
            target = json.loads(str(item["target_json"]))
            try:
                result = DryRunInventoryBalanceApplyAdapter().apply(target, actor=actor)
                if result.get("wb_patch_called") is not False:
                    raise SkuInventoryBalanceError("dry-run adapter did not prove no WB PATCH")
                state, error = "succeeded", ""
            except Exception as exc:
                result = {"status": "failed", "wb_patch_called": False}
                state, error = "failed", str(exc)
            with self._connect() as conn:
                conn.execute(
                    """UPDATE sheet_vitrina_v1_inventory_balance_apply_items
                       SET state=?,result_json=?,error=?,updated_at=?
                       WHERE job_id=? AND target_key=? AND state='running'""",
                    (state, _json(result), error, self.timestamp_factory(), job_id, str(item["target_key"])),
                )
                conn.commit()
        self._refresh_job_state(job_id)
        return self.get_apply_job(job_id)

    def _start_apply_worker_if_needed(self) -> bool:
        if self.apply_adapter.mode != LIVE_MODE:
            return False
        with self._apply_worker_lock:
            active = self._apply_worker_thread
            if active is not None and active.is_alive():
                self._apply_worker_wakeup.set()
                return True
            worker = threading.Thread(
                target=self._apply_worker_loop,
                name="sku-inventory-balance-live-apply",
                daemon=True,
            )
            self._apply_worker_thread = worker
            worker.start()
            return True

    def _apply_worker_loop(self) -> None:
        try:
            while not self._apply_worker_stop.is_set():
                claimed = self._claim_next_live_job()
                if claimed:
                    job_id, worker_token = claimed
                    try:
                        self._run_live_job(job_id, worker_token)
                    except Exception as exc:  # final worker containment
                        self._mark_job_worker_error(job_id, worker_token, exc)
                    continue
                if not self._has_active_live_jobs():
                    break
                self._apply_worker_wakeup.wait(timeout=1.0)
                self._apply_worker_wakeup.clear()
        finally:
            with self._apply_worker_lock:
                self._apply_worker_thread = None

    def _claim_next_live_job(self) -> tuple[str, str] | None:
        now = self.timestamp_factory()
        lease_until = (
            self.now_factory().astimezone(timezone.utc)
            + timedelta(seconds=self.apply_lease_seconds)
        ).isoformat()
        token = f"ibw_{uuid4().hex}"
        with self._connect() as conn:
            conn.execute("BEGIN IMMEDIATE")
            row = conn.execute(
                """SELECT job_id FROM sheet_vitrina_v1_inventory_balance_apply_jobs
                   WHERE mode=? AND state IN ('pending','running','delayed')
                     AND (worker_token='' OR lease_expires_at='' OR lease_expires_at<?)
                   ORDER BY created_at,job_id LIMIT 1""",
                (LIVE_MODE, now),
            ).fetchone()
            if row is None:
                conn.commit()
                return None
            job_id = str(row["job_id"])
            updated = conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs
                   SET worker_token=?,lease_expires_at=?,state='running',
                       phase=CASE WHEN phase='' OR phase='queued' THEN 'preflight' ELSE phase END,
                       started_at=CASE WHEN started_at='' THEN ? ELSE started_at END,
                       updated_at=?
                   WHERE job_id=? AND (worker_token='' OR lease_expires_at='' OR lease_expires_at<?)""",
                (token, lease_until, now, now, job_id, now),
            ).rowcount
            conn.commit()
        return (job_id, token) if updated == 1 else None

    def _has_active_live_jobs(self) -> bool:
        with self._connect() as conn:
            row = conn.execute(
                """SELECT 1 FROM sheet_vitrina_v1_inventory_balance_apply_jobs
                   WHERE mode=? AND state IN ('pending','running','delayed') LIMIT 1""",
                (LIVE_MODE,),
            ).fetchone()
        return row is not None

    def _renew_job_lease(
        self, job_id: str, worker_token: str, *, phase: str
    ) -> None:
        now = self.timestamp_factory()
        lease_until = (
            self.now_factory().astimezone(timezone.utc)
            + timedelta(seconds=self.apply_lease_seconds)
        ).isoformat()
        with self._connect() as conn:
            updated = conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs
                   SET lease_expires_at=?,phase=?,updated_at=?
                   WHERE job_id=? AND worker_token=?""",
                (lease_until, phase, now, job_id, worker_token),
            ).rowcount
            conn.commit()
        if updated != 1:
            raise SkuInventoryBalanceError("live apply worker lease was lost")

    def _release_job_lease(self, job_id: str, worker_token: str) -> None:
        with self._connect() as conn:
            conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs
                   SET worker_token='',lease_expires_at='',updated_at=?
                   WHERE job_id=? AND worker_token=?""",
                (self.timestamp_factory(), job_id, worker_token),
            )
            conn.commit()

    def _run_live_job(self, job_id: str, worker_token: str) -> None:
        self._recover_uncertain_submit_items(job_id)
        active_uncertain = self._load_job_items(
            job_id, states={"submitting", "submitted", "verifying", "delayed"}
        )
        if active_uncertain:
            self._renew_job_lease(job_id, worker_token, phase="verifying")
            self._verify_live_targets(
                job_id, worker_token, active_uncertain, wait_initial=False
            )

        pending = self._load_job_items(
            job_id, states={"pending", "preflighting", "ready"}
        )
        if pending:
            self._renew_job_lease(job_id, worker_token, phase="preflight")
            self._set_item_states(
                job_id,
                [item["target_key"] for item in pending],
                state="preflighting",
                phase="Проверяем текущие ставки и состояния кампаний",
            )
            pending_bid = [
                item
                for item in pending
                if self._transport_target(item)["action_type"] == "bid_change"
            ]
            pending_state = [
                item
                for item in pending
                if self._transport_target(item)["action_type"] == "campaign_state"
            ]
            preflight: list[dict[str, Any]] = []
            if pending_bid:
                preflight.extend(
                    self.apply_adapter.preflight(
                        [self._transport_target(item) for item in pending_bid],
                        min_bid_interval_seconds=self.min_bid_interval_seconds,
                        sleep=lambda seconds: self._sleep_with_lease(
                            job_id, worker_token, seconds, phase="preflight"
                        ),
                        safety_threshold_policy=self._job_safety_threshold_policy(job_id),
                    )
                )
            if pending_state:
                state_transports = [
                    self._transport_target(item) for item in pending_state
                ]
                try:
                    preflight.extend(
                        self.apply_adapter.preflight_state(state_transports)
                    )
                except WbPromotionApiError as exc:
                    code = (
                        "wb_rate_limited"
                        if exc.http_status == 429
                        else "campaign_state_preflight_unavailable"
                    )
                    message = (
                        "WB ограничил частоту проверки состояния кампании."
                        if exc.http_status == 429
                        else "Не удалось подтвердить текущее состояние кампании до отправки."
                    )
                    preflight.extend(
                        {
                            **target,
                            "ok": False,
                            "error_code": code,
                            "message": message,
                        }
                        for target in state_transports
                    )
            self._renew_job_lease(job_id, worker_token, phase="preflight_result")
            for result in preflight:
                if result.get("ok"):
                    self._update_item(
                        job_id,
                        str(result["target_key"]),
                        state="ready",
                        phase="Готово к применению",
                        result={"preflight": _public_preflight_result(result)},
                    )
                else:
                    self._update_item(
                        job_id,
                        str(result["target_key"]),
                        state="failed",
                        phase="Не применено",
                        error_code=str(result.get("error_code") or "preflight_failed"),
                        error=str(result.get("message") or "Проверка перед применением не пройдена."),
                        result={"preflight": _public_preflight_result(result)},
                    )

        succeeded = self._load_job_items(job_id, states={"succeeded"})
        ready = self._load_job_items(job_id, states={"ready"})
        if ready and not succeeded:
            canary = [ready[0]]
            self._renew_job_lease(job_id, worker_token, phase="canary_submit")
            submitted = self._submit_live_group(
                job_id, worker_token, canary, group_no=1
            )
            if submitted:
                self._renew_job_lease(
                    job_id, worker_token, phase="canary_verifying"
                )
                self._verify_live_targets(
                    job_id, worker_token, submitted, wait_initial=True
                )
            canary_result = self._load_job_items(
                job_id, target_keys=[canary[0]["target_key"]]
            )[0]
            if canary_result["state"] != "succeeded":
                remaining = self._load_job_items(job_id, states={"ready"})
                self._set_item_states(
                    job_id,
                    [item["target_key"] for item in remaining],
                    state="skipped",
                    phase="Не отправлено: контрольное изменение не подтверждено",
                    error_code="canary_not_confirmed",
                    error="Остальные изменения не отправлены: контрольное изменение не подтверждено.",
                )
                self._refresh_job_state(job_id)
                self._release_job_lease(job_id, worker_token)
                return

        ready = self._load_job_items(job_id, states={"ready"})
        ready_bid = [
            item
            for item in ready
            if self._transport_target(item)["action_type"] == "bid_change"
        ]
        ready_state = [
            item
            for item in ready
            if self._transport_target(item)["action_type"] == "campaign_state"
        ]
        group_no = 2
        submitted_all: list[dict[str, Any]] = []
        for offset in range(0, len(ready_bid), self.live_batch_size):
            batch = ready_bid[offset : offset + self.live_batch_size]
            self._renew_job_lease(job_id, worker_token, phase="batch_submit")
            submitted_all.extend(
                self._submit_live_group(
                    job_id, worker_token, batch, group_no=group_no
                )
            )
            group_no += 1
            if offset + self.live_batch_size < len(ready_bid) and self.patch_interval_seconds:
                self._sleep_with_lease(
                    job_id,
                    worker_token,
                    self.patch_interval_seconds,
                    phase="batch_submit",
                )
        for state_index, item in enumerate(ready_state):
            self._renew_job_lease(job_id, worker_token, phase="campaign_state_submit")
            submitted_all.extend(
                self._submit_live_group(
                    job_id, worker_token, [item], group_no=group_no
                )
            )
            group_no += 1
            if (
                state_index + 1 < len(ready_state)
                and self.patch_interval_seconds
            ):
                self._sleep_with_lease(
                    job_id,
                    worker_token,
                    self.patch_interval_seconds,
                    phase="campaign_state_submit",
                )
        if submitted_all:
            self._renew_job_lease(job_id, worker_token, phase="batch_verifying")
            self._verify_live_targets(
                job_id, worker_token, submitted_all, wait_initial=True
            )
        self._refresh_job_state(job_id)
        self._release_job_lease(job_id, worker_token)

    def _recover_uncertain_submit_items(self, job_id: str) -> None:
        items = self._load_job_items(job_id, states={"submitting"})
        for item in items:
            prepared = self._prepared_from_item(item)
            if prepared is not None:
                try:
                    self.writer_registry.ambiguous(
                        prepared,
                        error_code="worker_interrupted_at_submit",
                        error_message="Worker был прерван на границе отправки; повтор запрещён.",
                        receipt_reference=item["registry_receipt_reference"],
                    )
                except InternalWriterRegistryError:
                    pass
            self._update_item(
                job_id,
                item["target_key"],
                state="verifying",
                phase="Проверяем результат после перезапуска",
                error_code="worker_interrupted_at_submit",
                error="Отправка могла состояться; выполняется только проверка.",
            )

    def _submit_live_group(
        self,
        job_id: str,
        worker_token: str,
        items: Sequence[Mapping[str, Any]],
        *,
        group_no: int,
    ) -> list[dict[str, Any]]:
        if not items:
            return []
        job = self._job_row(job_id)
        action_types = {
            self._transport_target(item)["action_type"] for item in items
        }
        if len(action_types) != 1:
            raise SkuInventoryBalanceError(
                "one submit group cannot mix bid and campaign state transports"
            )
        action_type = next(iter(action_types))
        if action_type == "campaign_state" and len(items) != 1:
            raise SkuInventoryBalanceError(
                "campaign state submit group must contain exactly one campaign"
            )
        prepared_by_key: dict[str, PreparedWriterOperation] = {}
        try:
            for item in items:
                target = self._transport_target(item)
                receipt = f"inventory-balance:{job_id}:{target['target_key']}"
                prepared = self._prepared_from_item(item)
                if prepared is None:
                    common = {
                        "source_surface": "sku_inventory_balance",
                        "actor": str(job["created_by"]),
                        "native_operation_id": f"{job_id}:{target['target_key']}",
                        "nm_id": int(target["nm_id"]),
                        "advert_id": int(target["advert_id"]),
                        "requested_at": str(job["created_at"]),
                        "correlation_id": job_id,
                        "calculation_id": str(job["calculation_id"]),
                        "apply_operation_id": job_id,
                        "recommendation_item_id": str(target["recommendation_item_id"]),
                        "native_audit_reference": f"inventory-balance/apply-job/{job_id}",
                    }
                    if action_type == "campaign_state":
                        prepared = self.writer_registry.prepare_campaign_state(
                            **common,
                            before_state=str(target["current_campaign_state"]),
                            requested_state=str(target["requested_campaign_state"]),
                        )
                    else:
                        prepared = self.writer_registry.prepare_bid(
                            **common,
                            placement=str(target["placement"]),
                            before_bid_minor=int(target["current_bid_minor"]),
                            requested_bid_minor=int(target["requested_bid_minor"]),
                        )
                prepared_by_key[target["target_key"]] = prepared
                self._update_item(
                    job_id,
                    target["target_key"],
                    state="ready",
                    phase="Зарегистрировано перед отправкой",
                    registry_operation_id=prepared.operation_id,
                    registry_receipt_reference=receipt,
                )
        except Exception as exc:
            for key, prepared in prepared_by_key.items():
                try:
                    self.writer_registry.fail_before_submit(
                        prepared,
                        rejected=False,
                        error_code="batch_registry_prepare_failed",
                        error_message=str(exc),
                    )
                except Exception:
                    pass
                self._update_item(
                    job_id,
                    key,
                    state="failed",
                    phase="Не применено",
                    error_code="registry_prepare_failed",
                    error="Не удалось надёжно зарегистрировать изменение до отправки.",
                )
            for item in items:
                key = str(item["target_key"])
                if key not in prepared_by_key:
                    self._update_item(
                        job_id,
                        key,
                        state="failed",
                        phase="Не применено",
                        error_code="registry_prepare_failed",
                        error="Пакет остановлен до отправки из-за ошибки реестра.",
                    )
            return []

        keys = [str(item["target_key"]) for item in items]
        self._set_item_states(
            job_id,
            keys,
            state="submitting",
            phase="Отправляем в WB",
            submit_group=group_no,
        )
        transport_targets = [self._transport_target(item) for item in items]
        response: Mapping[str, Any] = {}
        for attempt in range(2):
            try:
                response = (
                    self.apply_adapter.submit_state(transport_targets[0])
                    if action_type == "campaign_state"
                    else self.apply_adapter.submit_batch(transport_targets)
                )
            except WbPromotionApiError as exc:
                self._renew_job_lease(
                    job_id, worker_token, phase="submit_error_received"
                )
                if (
                    action_type == "bid_change"
                    and exc.http_status == 429
                    and attempt == 0
                ):
                    delay = max(float(exc.retry_after_seconds or 1.0), 0.0)
                    self._set_item_states(
                        job_id,
                        keys,
                        state="delayed",
                        phase="WB ограничил частоту, ждём безопасное повторение",
                        error_code="wb_rate_limited",
                        error="WB временно ограничил частоту запросов.",
                    )
                    if delay:
                        self._sleep_with_lease(
                            job_id, worker_token, delay, phase="wb_rate_limited"
                        )
                    self._set_item_states(
                        job_id, keys, state="submitting", phase="Повторяем после лимита"
                    )
                    continue
                if exc.http_status is not None and exc.http_status < 500:
                    self._reject_pre_submit_group(
                        job_id, items, prepared_by_key, exc
                    )
                    return []
                self._mark_submit_ambiguous(
                    job_id, items, prepared_by_key, exc, group_no=group_no
                )
                return self._load_job_items(job_id, target_keys=keys)
            except (SheetVitrinaV1AdsError, SkuManagementError) as exc:
                self._renew_job_lease(
                    job_id, worker_token, phase="submit_rejected_locally"
                )
                self._fail_local_pre_submit_group(
                    job_id, items, prepared_by_key, exc
                )
                return []
            except Exception as exc:
                self._renew_job_lease(
                    job_id, worker_token, phase="submit_error_received"
                )
                self._mark_submit_ambiguous(
                    job_id, items, prepared_by_key, exc, group_no=group_no
                )
                return self._load_job_items(job_id, target_keys=keys)
            else:
                self._renew_job_lease(
                    job_id, worker_token, phase="submit_response_received"
                )
                break
        else:
            return []

        submitted_at = self.timestamp_factory()
        deadline_at = (
            self.now_factory().astimezone(timezone.utc)
            + timedelta(seconds=self.readback_deadline_seconds)
        ).isoformat()
        receipt_digest = _digest(
            {
                "job_id": job_id,
                "submit_group": group_no,
                "action_type": action_type,
                "target_keys": keys,
                "response": dict(response),
            }
        )
        for item in items:
            key = str(item["target_key"])
            receipt = f"inventory-balance:{job_id}:{key}"
            try:
                self.writer_registry.submitted(
                    prepared_by_key[key],
                    receipt_reference=receipt,
                    receipt_basis={
                        "job_id": job_id,
                        "submit_group": group_no,
                        "target_key": key,
                        "transport_receipt_digest": receipt_digest,
                    },
                )
            except InternalWriterRegistryError as exc:
                try:
                    self.writer_registry.ambiguous(
                        prepared_by_key[key],
                        error_code="registry_post_submit_failure",
                        error_message=str(exc),
                        receipt_reference=receipt,
                    )
                except InternalWriterRegistryError:
                    pass
            self._update_item(
                job_id,
                key,
                state="submitted",
                phase="Отправлено, ожидаем точное подтверждение WB",
                submitted_at=submitted_at,
                readback_deadline_at=deadline_at,
                result={"transport_receipt_digest": receipt_digest},
                error_code="",
                error="",
            )
        return self._load_job_items(job_id, target_keys=keys)

    def _reject_pre_submit_group(
        self,
        job_id: str,
        items: Sequence[Mapping[str, Any]],
        prepared_by_key: Mapping[str, PreparedWriterOperation],
        exc: WbPromotionApiError,
    ) -> None:
        code = f"wb_http_{exc.http_status}"
        for item in items:
            key = str(item["target_key"])
            try:
                self.writer_registry.fail_before_submit(
                    prepared_by_key[key],
                    rejected=True,
                    error_code=code,
                    error_message=str(exc),
                )
            except InternalWriterRegistryError:
                pass
            self._update_item(
                job_id,
                key,
                state="failed",
                phase="WB отклонил изменение",
                error_code=code,
                error="WB отклонил пакет до применения.",
            )

    def _fail_local_pre_submit_group(
        self,
        job_id: str,
        items: Sequence[Mapping[str, Any]],
        prepared_by_key: Mapping[str, PreparedWriterOperation],
        exc: Exception,
    ) -> None:
        for item in items:
            key = str(item["target_key"])
            try:
                self.writer_registry.fail_before_submit(
                    prepared_by_key[key],
                    rejected=False,
                    error_code="local_submit_guard",
                    error_message=str(exc),
                )
            except InternalWriterRegistryError:
                pass
            self._update_item(
                job_id,
                key,
                state="failed",
                phase="Не применено",
                error_code="local_submit_guard",
                error="Локальная проверка остановила отправку до обращения к WB.",
            )

    def _mark_submit_ambiguous(
        self,
        job_id: str,
        items: Sequence[Mapping[str, Any]],
        prepared_by_key: Mapping[str, PreparedWriterOperation],
        exc: Exception,
        *,
        group_no: int,
    ) -> None:
        submitted_at = self.timestamp_factory()
        deadline_at = (
            self.now_factory().astimezone(timezone.utc)
            + timedelta(seconds=self.readback_deadline_seconds)
        ).isoformat()
        for item in items:
            key = str(item["target_key"])
            receipt = f"inventory-balance:{job_id}:{key}"
            try:
                self.writer_registry.ambiguous(
                    prepared_by_key[key],
                    error_code="wb_submit_transport_unknown",
                    error_message=str(exc),
                    receipt_reference=receipt,
                )
            except InternalWriterRegistryError:
                pass
            self._update_item(
                job_id,
                key,
                state="verifying",
                phase="Ответ WB неясен, проверяем фактическое состояние",
                submit_group=group_no,
                submitted_at=submitted_at,
                readback_deadline_at=deadline_at,
                error_code="wb_submit_transport_unknown",
                error="Ответ WB не подтверждает результат; повторная отправка запрещена.",
            )

    def _verify_live_targets(
        self,
        job_id: str,
        worker_token: str,
        items: Sequence[Mapping[str, Any]],
        *,
        wait_initial: bool,
    ) -> None:
        keys = [str(item["target_key"]) for item in items]
        if wait_initial and self.readback_initial_delay_seconds:
            self._set_item_states(
                job_id,
                keys,
                state="delayed",
                phase="WB синхронизирует изменения, ожидаем проверку",
            )
            self._sleep_with_lease(
                job_id,
                worker_token,
                self.readback_initial_delay_seconds,
                phase="waiting_for_wb_sync",
            )
        while True:
            current = self._load_job_items(
                job_id,
                target_keys=keys,
                states={"submitting", "submitted", "verifying", "delayed"},
            )
            if not current:
                return
            self._renew_job_lease(job_id, worker_token, phase="readback")
            self._set_item_states(
                job_id,
                [item["target_key"] for item in current],
                state="verifying",
                phase="Проверяем фактические ставки и состояния в WB",
            )
            try:
                current_bid = [
                    item
                    for item in current
                    if self._transport_target(item)["action_type"] == "bid_change"
                ]
                current_state = [
                    item
                    for item in current
                    if self._transport_target(item)["action_type"] == "campaign_state"
                ]
                readbacks: list[dict[str, Any]] = []
                if current_bid:
                    readbacks.extend(
                        self.apply_adapter.readback(
                            [self._transport_target(item) for item in current_bid]
                        )
                    )
                if current_state:
                    readbacks.extend(
                        self.apply_adapter.readback_state(
                            [self._transport_target(item) for item in current_state]
                        )
                    )
            except Exception as exc:
                readbacks = [
                    {
                        **self._transport_target(item),
                        "ok": False,
                        "error_code": "readback_unavailable",
                        "message": str(exc),
                        "observed_bid_minor": None,
                        "observed_campaign_state": "",
                    }
                    for item in current
                ]
            self._renew_job_lease(
                job_id, worker_token, phase="readback_received"
            )
            by_key = {str(item["target_key"]): item for item in readbacks}
            now = self.now_factory().astimezone(timezone.utc)
            pending: list[dict[str, Any]] = []
            for item in current:
                key = str(item["target_key"])
                result = by_key.get(key) or {}
                transport = self._transport_target(item)
                action_type = transport["action_type"]
                observed_bid = _optional_int(result.get("observed_bid_minor"))
                observed_state = str(result.get("observed_campaign_state") or "")
                matched = (
                    result.get("ok")
                    and (
                        observed_state == str(transport["requested_campaign_state"])
                        if action_type == "campaign_state"
                        else observed_bid == int(transport["requested_bid_minor"])
                    )
                )
                if matched:
                    prepared = self._prepared_from_item(item)
                    if prepared is None:
                        self._update_item(
                            job_id,
                            key,
                            state="ambiguous",
                            phase="Требуется проверка",
                            error_code="registry_operation_missing",
                            error="Результат совпал, но связь с реестром не восстановлена.",
                            last_observed_bid_minor=observed_bid,
                        )
                        continue
                    try:
                        readback_basis = {
                            "job_id": job_id,
                            "calculation_id": str(
                                self._job_row(job_id)["calculation_id"]
                            ),
                            "recommendation_item_id": str(
                                transport["recommendation_item_id"]
                            ),
                            "nm_id": int(item["nm_id"]),
                            "advert_id": int(transport["advert_id"]),
                            "action_type": action_type,
                        }
                        if action_type == "campaign_state":
                            readback_basis["campaign_state"] = observed_state
                            self.writer_registry.confirm_campaign_state(
                                prepared,
                                confirmed_state=observed_state,
                                readback_basis=readback_basis,
                                receipt_reference=str(item["registry_receipt_reference"]),
                                native_audit_references=(
                                    f"inventory-balance/apply-job/{job_id}",
                                ),
                            )
                        else:
                            readback_basis["placement"] = str(transport["placement"])
                            readback_basis["bid_minor"] = observed_bid
                            self.writer_registry.confirm_bid(
                                prepared,
                                confirmed_bid_minor=int(observed_bid),
                                readback_basis=readback_basis,
                                receipt_reference=str(item["registry_receipt_reference"]),
                                native_audit_references=(
                                    f"inventory-balance/apply-job/{job_id}",
                                ),
                            )
                    except InternalWriterRegistryError as exc:
                        self._update_item(
                            job_id,
                            key,
                            state="ambiguous",
                            phase="Требуется проверка",
                            error_code="registry_confirmation_failed",
                            error="Изменение применено, но подтверждение реестра не завершено.",
                            last_observed_bid_minor=observed_bid,
                            result={"registry_error": str(exc)},
                        )
                        continue
                    if action_type == "bid_change":
                        self._persist_confirmed_bid_event(
                            job_id, item, int(observed_bid)
                        )
                    self._update_item(
                        job_id,
                        key,
                        state="succeeded",
                        phase="Применено",
                        error_code="",
                        error="",
                        last_observed_bid_minor=observed_bid,
                        result={
                            "readback_status": "matching",
                            "confirmed_bid_minor": observed_bid,
                            "confirmed_campaign_state": observed_state,
                        },
                    )
                    continue
                deadline = _parse_timestamp(str(item.get("readback_deadline_at") or ""))
                if deadline is None:
                    deadline = now + timedelta(seconds=self.readback_deadline_seconds)
                if now >= deadline:
                    prepared = self._prepared_from_item(item)
                    if prepared is not None:
                        try:
                            self.writer_registry.ambiguous(
                                prepared,
                                error_code="wb_readback_unconfirmed",
                                error_message="Exact readback did not confirm the requested change before deadline.",
                                receipt_reference=str(item["registry_receipt_reference"]),
                            )
                        except InternalWriterRegistryError:
                            pass
                    self._update_item(
                        job_id,
                        key,
                        state="ambiguous",
                        phase="Требуется проверка",
                        error_code=str(result.get("error_code") or "wb_readback_unconfirmed"),
                        error="WB не подтвердил запрошенное изменение в отведённое время.",
                        last_observed_bid_minor=observed_bid,
                    )
                else:
                    self._update_item(
                        job_id,
                        key,
                        state="delayed",
                        phase="WB задерживает подтверждение, продолжаем проверку",
                        error_code=str(result.get("error_code") or ""),
                        error="" if result.get("ok") else "Проверка временно недоступна.",
                        last_observed_bid_minor=observed_bid,
                    )
                    pending.append(item)
            if not pending:
                return
            if self.readback_poll_seconds:
                self._sleep_with_lease(
                    job_id,
                    worker_token,
                    self.readback_poll_seconds,
                    phase="readback",
                )

    def _sleep_with_lease(
        self,
        job_id: str,
        worker_token: str,
        seconds: float,
        *,
        phase: str,
    ) -> None:
        remaining = max(float(seconds), 0.0)
        slice_seconds = max(min(self.apply_lease_seconds / 3.0, 30.0), 1.0)
        while remaining > 0:
            delay = min(remaining, slice_seconds)
            self.sleep(delay)
            remaining -= delay
            self._renew_job_lease(job_id, worker_token, phase=phase)

    def _persist_confirmed_bid_event(
        self, job_id: str, item: Mapping[str, Any], confirmed_minor: int
    ) -> None:
        try:
            target = self._transport_target(item)
            job = self._job_row(job_id)
            self.sku_management_block.persist_balance_bid_result(
                job_id=job_id,
                calculation_id=str(job["calculation_id"]),
                actor=str(job["created_by"]),
                target=target,
                confirmed_bid_minor=confirmed_minor,
                requested_at=str(job["created_at"]),
                confirmed_at=self.timestamp_factory(),
            )
        except Exception as exc:
            _LOGGER.warning("balance apply native event persistence failed: %s", exc)

    def _prepared_from_item(
        self, item: Mapping[str, Any]
    ) -> PreparedWriterOperation | None:
        operation_id = str(item.get("registry_operation_id") or "")
        if not operation_id or self.writer_registry is None:
            return None
        try:
            stored = self.writer_registry.repository.read_operation(operation_id)
        except Exception:
            return None
        operation = stored["operation"]
        change_item_ids = {
            ":".join(
                (
                    str(value["target_kind"]),
                    str(value["nm_id"]),
                    str(value["advert_id"]),
                    str(value["placement"]),
                    str(value["parameter_field"]),
                )
            ): str(value["change_item_id"])
            for value in stored["items"]
        }
        return PreparedWriterOperation(
            operation_id=operation_id,
            change_item_ids=change_item_ids,
            source_surface=str(operation["source_surface"]),
            native_operation_id=str(operation["native_idempotency_key"]),
        )

    def _transport_target(self, item: Mapping[str, Any]) -> dict[str, Any]:
        target = json.loads(str(item["target_json"])) if "target_json" in item else dict(item)
        base = {
            **target,
            "target_key": str(target["target_key"]),
            "nm_id": int(target["nm_id"]),
            "advert_id": int(target["advert_id"]),
            "action_type": str(target.get("action_type") or "bid_change"),
            "recommendation_item_id": str(target["recommendation_item_id"]),
        }
        if base["action_type"] == "campaign_state":
            return {
                **base,
                "placement": "",
                "placement_evidence": str(target["placement_evidence"]),
                "payment_type": str(target["payment_type"]),
                "state_action": str(target["state_action"]),
                "current_campaign_status": int(target["current_campaign_status"]),
                "current_campaign_state": str(target["current_campaign_state"]),
                "requested_campaign_state": str(target["requested_campaign_state"]),
            }
        return {
            **base,
            "placement": str(target["placement"]),
            "current_bid_minor": int(target["current_bid_minor"]),
            "requested_bid_minor": int(target["final_target_bid_minor"]),
        }

    def _job_row(self, job_id: str) -> dict[str, Any]:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM sheet_vitrina_v1_inventory_balance_apply_jobs WHERE job_id=?",
                (job_id,),
            ).fetchone()
        if row is None:
            raise SkuInventoryBalanceError("inventory balance apply job not found", http_status=404)
        return dict(row)

    def _load_job_items(
        self,
        job_id: str,
        *,
        states: set[str] | None = None,
        target_keys: Sequence[str] | None = None,
    ) -> list[dict[str, Any]]:
        conditions = ["job_id=?"]
        params: list[Any] = [job_id]
        if states:
            conditions.append("state IN (%s)" % ",".join("?" for _ in states))
            params.extend(sorted(states))
        if target_keys:
            conditions.append("target_key IN (%s)" % ",".join("?" for _ in target_keys))
            params.extend(str(value) for value in target_keys)
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM sheet_vitrina_v1_inventory_balance_apply_items WHERE "
                + " AND ".join(conditions)
                + " ORDER BY target_key",
                params,
            ).fetchall()
        return [dict(row) for row in rows]

    def _set_item_states(
        self,
        job_id: str,
        target_keys: Sequence[str],
        *,
        state: str,
        phase: str,
        error_code: str | None = None,
        error: str | None = None,
        submit_group: int | None = None,
    ) -> None:
        for key in target_keys:
            self._update_item(
                job_id,
                str(key),
                state=state,
                phase=phase,
                error_code=error_code,
                error=error,
                submit_group=submit_group,
            )

    def _update_item(
        self,
        job_id: str,
        target_key: str,
        *,
        state: str | None = None,
        phase: str | None = None,
        result: Mapping[str, Any] | None = None,
        error_code: str | None = None,
        error: str | None = None,
        submit_group: int | None = None,
        submitted_at: str | None = None,
        readback_deadline_at: str | None = None,
        registry_operation_id: str | None = None,
        registry_receipt_reference: str | None = None,
        last_observed_bid_minor: int | None = None,
    ) -> None:
        assignments = ["updated_at=?"]
        params: list[Any] = [self.timestamp_factory()]
        values = {
            "state": state,
            "phase": phase,
            "result_json": _json(result) if result is not None else None,
            "error_code": error_code,
            "error": error,
            "submit_group": submit_group,
            "submitted_at": submitted_at,
            "readback_deadline_at": readback_deadline_at,
            "registry_operation_id": registry_operation_id,
            "registry_receipt_reference": registry_receipt_reference,
            "last_observed_bid_minor": last_observed_bid_minor,
        }
        for column, value in values.items():
            if value is None:
                continue
            assignments.append(f"{column}=?")
            params.append(value)
        params.extend((job_id, target_key))
        with self._connect() as conn:
            updated = conn.execute(
                "UPDATE sheet_vitrina_v1_inventory_balance_apply_items SET "
                + ",".join(assignments)
                + " WHERE job_id=? AND target_key=?",
                params,
            ).rowcount
            conn.execute(
                "UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs SET updated_at=? WHERE job_id=?",
                (self.timestamp_factory(), job_id),
            )
            conn.commit()
        if updated != 1:
            raise SkuInventoryBalanceError("inventory balance apply item not found")

    def _mark_job_worker_error(
        self, job_id: str, worker_token: str, exc: Exception
    ) -> None:
        now = self.timestamp_factory()
        with self._connect() as conn:
            conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs
                   SET state='stalled',phase='worker_error',error_code='worker_error',
                       error_message=?,worker_token='',lease_expires_at='',updated_at=?
                   WHERE job_id=? AND worker_token=?""",
                (_bounded_error(exc), now, job_id, worker_token),
            )
            conn.commit()

    def get_apply_job(self, job_id: str) -> dict[str, Any]:
        with self._connect() as conn:
            job = conn.execute(
                "SELECT * FROM sheet_vitrina_v1_inventory_balance_apply_jobs WHERE job_id=?",
                (job_id,),
            ).fetchone()
            if job is None:
                raise SkuInventoryBalanceError("inventory balance apply job not found", http_status=404)
            items = conn.execute(
                """SELECT * FROM sheet_vitrina_v1_inventory_balance_apply_items
                   WHERE job_id=? ORDER BY nm_id,target_key""",
                (job_id,),
            ).fetchall()
        states = {
            state: 0
            for state in (
                "pending",
                "preflighting",
                "ready",
                "submitting",
                "submitted",
                "verifying",
                "delayed",
                "running",
                "succeeded",
                "failed",
                "skipped",
                "ambiguous",
            )
        }
        payload_items = []
        by_nm: dict[int, list[str]] = {}
        for item in items:
            state = str(item["state"])
            states[state] = states.get(state, 0) + 1
            target = json.loads(str(item["target_json"]))
            by_nm.setdefault(int(item["nm_id"]), []).append(state)
            payload_items.append(
                {
                    **target,
                    "state": state,
                    "attempt_count": int(item["attempt_count"]),
                    "result": json.loads(str(item["result_json"] or "{}")),
                    "phase": str(item["phase"] or ""),
                    "submit_group": int(item["submit_group"] or 0),
                    "submitted_at": str(item["submitted_at"] or ""),
                    "readback_deadline_at": str(item["readback_deadline_at"] or ""),
                    "registry_operation_id": str(item["registry_operation_id"] or ""),
                    "registry_receipt_reference": str(item["registry_receipt_reference"] or ""),
                    "error_code": str(item["error_code"] or ""),
                    "last_observed_bid_minor": item["last_observed_bid_minor"],
                    "error": str(item["error"] or ""),
                    "updated_at": str(item["updated_at"]),
                }
            )
        sku_states = []
        for nm_id, item_states in sorted(by_nm.items()):
            if any(state in {"failed", "ambiguous"} for state in item_states):
                row_state = "failed"
            elif all(state in TERMINAL_ITEM_STATES for state in item_states):
                row_state = "succeeded"
            elif any(state in ACTIVE_ITEM_STATES or state == "running" for state in item_states):
                row_state = "running"
            else:
                row_state = "pending"
            sku_states.append({"nm_id": nm_id, "state": row_state, "target_count": len(item_states)})
        terminal = sum(states.get(state, 0) for state in TERMINAL_ITEM_STATES)
        total = len(items)
        verifying = sum(
            states.get(state, 0)
            for state in ("submitting", "submitted", "verifying", "delayed")
        )
        waiting = sum(
            states.get(state, 0)
            for state in ("pending", "preflighting", "ready")
        )
        failed = states.get("failed", 0) + states.get("skipped", 0)
        needs_check = states.get("ambiguous", 0)
        updated_at = str(job["updated_at"])
        stored_state = str(job["state"])
        stalled = stored_state == "stalled" or (
            stored_state in ACTIVE_JOB_STATES
            and _timestamp_age_seconds(updated_at, self.now_factory()) > self.apply_lease_seconds
        )
        return {
            "contract_name": "sheet_vitrina_v1_inventory_balance_apply_job/v1",
            "job_id": str(job["job_id"]),
            "calculation_id": str(job["calculation_id"]),
            "mode": str(job["mode"]),
            "state": "stalled" if stalled else stored_state,
            "stored_state": stored_state,
            "idempotency_key": str(job["idempotency_key"]),
            "apply_manifest_digest": str(job["apply_manifest_digest"]),
            "apply_manifest": json.loads(str(job["apply_manifest_json"])),
            "selection": json.loads(str(job["selection_json"])),
            "summary": json.loads(str(job["summary_json"])),
            "progress": {
                "total": total,
                "terminal": terminal,
                "percent": round(terminal / total * 100, 1) if total else 100.0,
                "states": states,
                "applied": states.get("succeeded", 0),
                "verifying": verifying,
                "waiting": waiting,
                "failed": failed,
                "needs_check": needs_check,
            },
            "sku_states": sku_states,
            "items": payload_items,
            "created_at": str(job["created_at"]),
            "created_by": str(job["created_by"]),
            "updated_at": updated_at,
            "phase": str(job["phase"] or ""),
            "error_code": str(job["error_code"] or ""),
            "error_message": str(job["error_message"] or ""),
            "stalled": stalled,
            "external_writes": str(job["mode"]) == LIVE_MODE,
            "wb_patch_called": str(job["mode"]) == LIVE_MODE and any(
                int(item["submit_group"] or 0) > 0
                and str(json.loads(str(item["target_json"])).get("action_type") or "bid_change")
                == "bid_change"
                for item in items
            ),
            "wb_campaign_action_called": str(job["mode"]) == LIVE_MODE and any(
                int(item["submit_group"] or 0) > 0
                and str(json.loads(str(item["target_json"])).get("action_type") or "bid_change")
                == "campaign_state"
                for item in items
            ),
        }

    def build_workbook(self, calculation_id: str) -> tuple[bytes, str]:
        calculation = self.get_calculation(calculation_id)
        wb = Workbook()
        decisions = wb.active
        decisions.title = "Решения"
        decision_headers = [
            "Товар", "nmID", "Статус", "Известный запас", "Продажи/день",
            "Целевой темп", "Изменение темпа", "Дней запаса", "Узкая дата",
            "Следующая поставка", "Последующая поставка", "Новая CPC", "Старая CPM",
            "WB источник", "Качество",
        ]
        decisions.append(decision_headers)
        for row in calculation.get("rows") or []:
            decisions.append(
                [
                    row.get("name"), row.get("nm_id"), row.get("status"),
                    row.get("known_stock_units"), row.get("current_daily_sales"),
                    row.get("target_daily_sales"), row.get("pace_change_pct"),
                    row.get("days_cover"), row.get("bottleneck_date"),
                    _inbound_label(row.get("next_inbound")),
                    _inbound_label(row.get("subsequent_inbound")),
                    _campaigns_label(row.get("new_cpc_campaigns")),
                    _campaigns_label(row.get("old_cpm_campaigns")),
                    _wb_stock_evidence_label(row), row.get("quality"),
                ]
            )
        calculation_sheet = wb.create_sheet("Расчёт")
        calculation_sheet.append(
            [
                "Товар", "nmID", "Начальный доступный остаток", "Текущий темп",
                "Жёсткий темп", "Темп с резервом", "Целевой темп", "Статус",
                "Узкая дата", "Следующая дата", "Следующая, шт",
                "Последующая дата", "Последующая, шт", "Остаток WB",
                "Коэффициент WB", "Учтено WB", "WB evidence mode", "Формула",
            ]
        )
        for row in calculation.get("rows") or []:
            next_inbound = row.get("next_inbound") or {}
            subsequent = row.get("subsequent_inbound") or {}
            calculation_sheet.append(
                [
                    row.get("name"), row.get("nm_id"), row.get("known_stock_units"),
                    row.get("current_daily_sales"), row.get("hard_daily_sales"),
                    row.get("reserve_daily_sales"), row.get("target_daily_sales"),
                    row.get("status"), row.get("bottleneck_date"),
                    next_inbound.get("date"), next_inbound.get("quantity"),
                    subsequent.get("date"), subsequent.get("quantity"),
                    row.get("stock_wb_units"), row.get("wb_confidence_coefficient"),
                    row.get("confidence_adjusted_wb_units"),
                    (row.get("wb_stock_evidence") or {}).get("mode"),
                    calculation.get("formula_version"),
                ]
            )
        campaigns = wb.create_sheet("Кампании")
        campaigns.append(
            [
                "Товар", "nmID", "Группа", "advertId", "Кампания", "Тип оплаты",
                "Размещение", "CPO", "Текущая ставка", "Расчётная ставка",
                "Ручная ставка", "Финальная ставка", "Можно применить", "target_key",
                "Outcome status", "Observed from", "Observed to", "Observed CPO",
            ]
        )
        for row in calculation.get("rows") or []:
            outcome = row.get("outcome_observation") or {}
            for target in row.get("campaign_recommendations") or []:
                campaigns.append(
                    [
                        row.get("name"), row.get("nm_id"), target.get("campaign_group"),
                        target.get("advert_id"), target.get("campaign_name"),
                        target.get("payment_type"), target.get("placement"), target.get("cpo_rub"),
                        target.get("current_bid_rub"), target.get("calculated_target_bid_rub"),
                        target.get("manual_target_bid_rub"), target.get("final_target_bid_rub"),
                        bool(target.get("can_apply")), target.get("target_key"),
                        outcome.get("status"), outcome.get("observed_from"),
                        outcome.get("observed_to"), outcome.get("cpo_rub"),
                    ]
                )
        inbounds = wb.create_sheet("Поставки")
        inbounds.append(
            [
                "Товар", "nmID", "Дата", "Количество", "Доступно до прибытия",
                "Накопительно после прибытия", "Source", "Source ID", "Роль",
                "ETA method", "ETA samples", "ETA mean days", "ETA applied days",
                "ETA quality", "Row quality",
            ]
        )
        eta = dict((calculation.get("lineage") or {}).get("supplier_eta_evidence") or {})
        for row in calculation.get("rows") or []:
            next_date = str((row.get("next_inbound") or {}).get("date") or "")
            subsequent_date = str((row.get("subsequent_inbound") or {}).get("date") or "")
            for milestone in row.get("milestones") or []:
                milestone_date = str(milestone.get("date") or "")
                role = "next" if milestone_date == next_date else "subsequent" if milestone_date == subsequent_date else "later"
                inbounds.append(
                    [
                        row.get("name"), row.get("nm_id"), milestone_date,
                        milestone.get("quantity"), milestone.get("available_before_arrival"),
                        milestone.get("cumulative_available"),
                        ", ".join(str(item) for item in milestone.get("sources") or []),
                        ", ".join(str(item) for item in milestone.get("source_ids") or []),
                        role, eta.get("method"), eta.get("sample_count"),
                        eta.get("mean_days_exact"), eta.get("applied_days"),
                        eta.get("quality"), row.get("quality"),
                    ]
                )
        sources = wb.create_sheet("Источники")
        sources.append(["Поле", "Значение"])
        for key, value in (
            ("calculation_id", calculation.get("calculation_id")),
            ("operation_id", calculation.get("operation_id")),
            ("previous_calculation_id", calculation.get("previous_calculation_id")),
            ("created_at", calculation.get("created_at")),
            ("source_digest", calculation.get("source_digest")),
            ("formula_version", calculation.get("formula_version")),
            ("source_contract", calculation.get("source_contract")),
            ("source_generated_at", calculation.get("source_generated_at")),
            ("sales_evidence_window", _json((calculation.get("lineage") or {}).get("sales_evidence_window") or {})),
            ("supplier_eta_evidence", _json((calculation.get("lineage") or {}).get("supplier_eta_evidence") or {})),
            ("wb_stock_evidence", _json((calculation.get("lineage") or {}).get("wb_stock_evidence") or [])),
            ("exclusion_policy", _json((calculation.get("lineage") or {}).get("exclusion_policy") or {})),
            ("excluded_rows", _json(calculation.get("excluded_rows") or [])),
            ("settings", _json(calculation.get("settings") or {})),
            ("automatic_ml_or_training", False),
        ):
            sources.append([key, value])
        history = wb.create_sheet("История расчётов")
        history.append(["calculation_id", "created_at", "created_by", "source_digest", "previous"])
        with self._connect() as conn:
            rows = conn.execute(
                """SELECT calculation_id,created_at,created_by,source_digest,previous_calculation_id
                   FROM sheet_vitrina_v1_inventory_balance_calculations
                   ORDER BY created_at DESC, calculation_id DESC"""
            ).fetchall()
        for row in rows:
            history.append(
                [row["calculation_id"], row["created_at"], row["created_by"], row["source_digest"], row["previous_calculation_id"]]
            )
        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="253449")
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            for column in range(1, sheet.max_column + 1):
                values = [str(sheet.cell(row=row, column=column).value or "") for row in range(1, min(sheet.max_row, 100) + 1)]
                sheet.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, values), default=8) + 2, 10), 42)
        stream = BytesIO()
        wb.save(stream)
        body = stream.getvalue()
        readback = load_workbook(BytesIO(body), read_only=True, data_only=False)
        if readback.sheetnames[:1] != ["Решения"] or not {"Кампании", "Поставки"}.issubset(readback.sheetnames):
            raise SkuInventoryBalanceError("generated workbook failed primary-sheet readback", http_status=500)
        filename = f"Баланс_запасов_{calculation_id}.xlsx"
        return body, filename

    def _calculation_payload(self, row: sqlite3.Row) -> dict[str, Any]:
        payload = json.loads(str(row["payload_json"]))
        with self._connect() as conn:
            overrides = conn.execute(
                """SELECT * FROM sheet_vitrina_v1_inventory_balance_overrides
                   WHERE calculation_id=?""",
                (str(row["calculation_id"]),),
            ).fetchall()
        by_key = {str(item["target_key"]): item for item in overrides}
        payload = deepcopy(payload)
        recommendation_ids: list[str] = []
        for balance_row in payload.get("rows") or []:
            for target in balance_row.get("campaign_recommendations") or []:
                override = by_key.get(str(target["target_key"]))
                manual = (
                    float(override["manual_target_bid_rub"])
                    if override is not None and override["manual_target_bid_rub"] is not None
                    else None
                )
                target["manual_target_bid_rub"] = manual
                target["final_target_bid_rub"] = (
                    manual if manual is not None else target.get("calculated_target_bid_rub")
                )
                target["override_updated_at"] = str(override["updated_at"]) if override else ""
                target["override_updated_by"] = str(override["updated_by"]) if override else ""
                target["can_apply"] = bool(
                    target.get("identity_valid")
                    and target.get("manual_override_allowed")
                    and target.get("current_bid_rub") is not None
                    and target.get("final_target_bid_rub") is not None
                    and float(target["current_bid_rub"]) != float(target["final_target_bid_rub"])
                )
                target["action_type"] = "bid_change"
                target["current_bid_minor"] = _rub_to_minor(
                    target.get("current_bid_rub")
                )
                target["final_target_bid_minor"] = _rub_to_minor(
                    target.get("final_target_bid_rub")
                )
                target["exact_target"] = {
                    "seller_id": self.seller_id,
                    "account_scope": self.account_scope,
                    "target_kind": "bid",
                    "nm_id": int(target["nm_id"]),
                    "advert_id": int(target["advert_id"]),
                    "placement": str(target["placement"]),
                    "parameter_field": "bid_minor",
                }
                recommendation_basis = {
                    "contract": "sku_inventory_balance_bid_recommendation/v1",
                    "calculation_id": str(row["calculation_id"]),
                    "target": target["exact_target"],
                    "before_value": target["current_bid_minor"],
                    "requested_value": target["final_target_bid_minor"],
                    "override_updated_at": target.get("override_updated_at") or "",
                }
                target["recommendation_item_id"] = (
                    "ibr_" + sha256(
                        json.dumps(
                            recommendation_basis,
                            ensure_ascii=False,
                            sort_keys=True,
                            separators=(",", ":"),
                        ).encode("utf-8")
                    ).hexdigest()
                )
                state_action = str(target.get("state_action") or "")
                requested_campaign_state = (
                    "active" if state_action == "start" else "paused"
                    if state_action == "pause"
                    else ""
                )
                target["state_target_key"] = (
                    f"state:{int(target['nm_id'])}:{int(target['advert_id'])}"
                )
                target["requested_campaign_state"] = requested_campaign_state
                state_basis = {
                    "contract": "sku_inventory_balance_campaign_state_recommendation/v1",
                    "calculation_id": str(row["calculation_id"]),
                    "target": {
                        "seller_id": self.seller_id,
                        "account_scope": self.account_scope,
                        "target_kind": "campaign",
                        "nm_id": int(target["nm_id"]),
                        "advert_id": int(target["advert_id"]),
                        "placement": "",
                        "parameter_field": "campaign_state",
                    },
                    "before_value": str(target.get("campaign_state") or ""),
                    "requested_value": requested_campaign_state,
                }
                target["campaign_state_recommendation_item_id"] = (
                    "ibsr_" + sha256(
                        json.dumps(
                            state_basis,
                            ensure_ascii=False,
                            sort_keys=True,
                            separators=(",", ":"),
                        ).encode("utf-8")
                    ).hexdigest()
                )
                target["manual_pending_available"] = bool(
                    target["can_apply"]
                    and self.manual_pending_registry is not None
                    and self.seller_id
                )
                recommendation_ids.append(target["recommendation_item_id"])
            balance_row["new_cpc_campaigns"] = [
                item for item in balance_row.get("campaign_recommendations") or []
                if item.get("campaign_group") == "new_cpc"
            ]
            balance_row["old_cpm_campaigns"] = [
                item for item in balance_row.get("campaign_recommendations") or []
                if item.get("campaign_group") == "old_cpm"
            ]
            balance_row["select_available"] = any(
                item.get("can_apply") or item.get("state_action_available")
                for item in balance_row.get("campaign_recommendations") or []
            )
        payload["registry_immutable"] = True
        payload["overrides_are_separate"] = True
        payload["apply_capability"] = self._apply_capability()
        statuses = (
            self.manual_pending_registry.manual_pending_statuses(recommendation_ids)
            if self.manual_pending_registry is not None and recommendation_ids
            else {}
        )
        for balance_row in payload.get("rows") or []:
            for target in balance_row.get("campaign_recommendations") or []:
                target["manual_pending"] = statuses.get(
                    str(target.get("recommendation_item_id") or "")
                )
        payload["manual_pending_capability"] = {
            "available": bool(self.manual_pending_registry is not None and self.seller_id),
            "contract_name": MANUAL_PENDING_CONTRACT,
            "expiration_hours": 24,
            "external_writes": False,
        }
        return payload

    def _terminalize_stale_running_items(self, job_id: str) -> None:
        cutoff = (self.now_factory().astimezone(timezone.utc) - timedelta(minutes=5)).isoformat()
        with self._connect() as conn:
            conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_apply_items
                   SET state='ambiguous',error='stale running target; no blind retry',updated_at=?
                   WHERE job_id=? AND state='running' AND updated_at<?""",
                (self.timestamp_factory(), job_id, cutoff),
            )
            conn.commit()

    def _refresh_job_state(self, job_id: str) -> None:
        with self._connect() as conn:
            counts = {
                str(row["state"]): int(row["count"])
                for row in conn.execute(
                    """SELECT state,COUNT(*) AS count
                       FROM sheet_vitrina_v1_inventory_balance_apply_items
                       WHERE job_id=? GROUP BY state""",
                    (job_id,),
                ).fetchall()
            }
            if any(counts.get(state) for state in ACTIVE_ITEM_STATES | {"running"}):
                state = "running"
            elif counts.get("failed") or counts.get("ambiguous"):
                state = "completed_with_errors"
            else:
                state = "completed"
            conn.execute(
                """UPDATE sheet_vitrina_v1_inventory_balance_apply_jobs
                   SET state=?,phase=?,finished_at=CASE WHEN ? IN ('completed','completed_with_errors')
                       THEN ? ELSE finished_at END,updated_at=? WHERE job_id=?""",
                (
                    state,
                    "complete" if state == "completed" else "complete_with_issues"
                    if state == "completed_with_errors"
                    else "running",
                    state,
                    self.timestamp_factory(),
                    self.timestamp_factory(),
                    job_id,
                ),
            )
            conn.commit()

    def _apply_capability(self) -> dict[str, Any]:
        live = bool(
            self.apply_adapter.mode == LIVE_MODE
            and self.apply_adapter.external_writes_enabled
            and self.writer_registry is not None
            and self.seller_id
        )
        return {
            "default_mode": LIVE_MODE if live else DRY_RUN_MODE,
            "accepted_modes": [DRY_RUN_MODE, LIVE_MODE] if live else [DRY_RUN_MODE],
            "live_wb_available": live,
            "external_writes_enabled": live,
            "guard_contract": "fresh batch current/min preflight -> canary -> micro-batches -> exact readback",
            "wb_patch_reachable": live,
            "wb_campaign_state_reachable": live,
            "campaign_state_actions": {
                "9": "pause",
                "4": "start",
                "11": "start",
            },
            "batch_size": self.live_batch_size,
            "canary_required": True,
            "reload_safe": True,
            "owner_confirmation_policy": self._balance_owner_confirmation_policy(),
        }

    def _balance_owner_confirmation_policy(self) -> dict[str, Any]:
        policy_factory = getattr(
            self.apply_adapter, "owner_confirmation_policy", None
        )
        if callable(policy_factory):
            policy = dict(policy_factory())
        else:
            policy = {
                "contract_name": BALANCE_OWNER_CONFIRMATION_POLICY_CONTRACT,
                "safety_threshold_policy": (
                    AdsBidSafetyThresholdPolicy.OWNER_CONFIRMED_BALANCE.value
                ),
                "warnings_only": [],
                "thresholds": {},
                "direct_submit": True,
                "staircase_submit": False,
            }
        return policy

    def _balance_safety_threshold_warnings(
        self, target: Mapping[str, Any]
    ) -> list[dict[str, Any]]:
        warning_factory = getattr(
            self.apply_adapter, "safety_threshold_warnings", None
        )
        if not callable(warning_factory):
            return []
        return [dict(item) for item in warning_factory(target)]

    def _job_safety_threshold_policy(
        self, job_id: str
    ) -> AdsBidSafetyThresholdPolicy:
        job = self._job_row(job_id)
        manifest = json.loads(str(job.get("apply_manifest_json") or "{}"))
        confirmation = manifest.get("owner_confirmation")
        if not isinstance(confirmation, Mapping):
            return AdsBidSafetyThresholdPolicy.STRICT
        if (
            confirmation.get("confirmed") is True
            and str(confirmation.get("contract_name") or "")
            == BALANCE_OWNER_CONFIRMATION_POLICY_CONTRACT
            and str(confirmation.get("safety_threshold_policy") or "")
            == AdsBidSafetyThresholdPolicy.OWNER_CONFIRMED_BALANCE.value
        ):
            return AdsBidSafetyThresholdPolicy.OWNER_CONFIRMED_BALANCE
        return AdsBidSafetyThresholdPolicy.STRICT

    def _apply_protocols(self) -> list[dict[str, Any]]:
        return [
            {
                "protocol": "inventory_balance_apply_job/v1",
                "mode": DRY_RUN_MODE,
                "durable": True,
                "resumable": True,
                "single_target_terminal_states": sorted(TERMINAL_ITEM_STATES),
                "external_writes": False,
            },
            {
                "protocol": "inventory_balance_live_wb_boundary/v1",
                "mode": LIVE_MODE,
                "available": self._apply_capability()["live_wb_available"],
                "fail_closed": True,
                "batch_size": self.live_batch_size,
                "canary_required": True,
                "server_owned_worker": True,
                "exact_readback": True,
                "blind_retry": False,
            },
        ]

    def _connect(self) -> sqlite3.Connection:
        registry = getattr(self.runtime, "store_registry", None)
        if registry is not None:
            manifest = registry.load()
            conn = registry.connect(
                "operational",
                mode="rw",
                operation="sku_inventory_balance",
                manifest=manifest,
            )
        else:
            db_path = Path(getattr(self.runtime, "db_path"))
            db_path.parent.mkdir(parents=True, exist_ok=True)
            conn = sqlite3.connect(db_path, timeout=30.0)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        return conn


def calculate_inventory_balance_row(
    raw: Mapping[str, Any],
    *,
    settings: Mapping[str, Any],
) -> dict[str, Any]:
    demand = _optional_float(raw.get("daily_demand"))
    stock_wb = _optional_float(raw.get("stock_wb"))
    stock_ff = _optional_float(raw.get("stock_ff"))
    wb_confidence_coefficient = float(settings["wb_confidence_coefficient"])
    wb_stock_evidence = _calculation_wb_stock_evidence(
        raw,
        stock_wb=stock_wb,
        wb_confidence_coefficient=wb_confidence_coefficient,
    )
    inbound_evidence = raw.get("inventory_balance_inbounds")
    as_of_date = _iso_date(str(raw.get("inventory_balance_as_of_date") or ""))
    identity = {
        "nm_id": int(raw.get("nm_id") or 0),
        "name": str(raw.get("name") or raw.get("display_name") or ""),
        "our_sku": str(raw.get("our_sku") or ""),
    }
    if (
        stock_wb is None
        or stock_ff is None
        or demand is None
        or demand < 0
        or not isinstance(inbound_evidence, list)
        or as_of_date is None
    ):
        missing = []
        if stock_wb is None:
            missing.append("WB")
        if stock_ff is None:
            missing.append("FF")
        if not isinstance(inbound_evidence, list):
            missing.append("dated registry inbounds")
        if as_of_date is None:
            missing.append("balance as_of_date")
        return {
            **identity,
            "status": "Недостаточно данных",
            "quality": "unknown",
            "quality_warnings": list(raw.get("quality_warnings") or [])
            + [
                "Нет exact stock evidence: " + ", ".join(missing)
                if missing
                else "Нет usable timeline/темпа продаж"
            ],
            "stock_wb_units": stock_wb,
            "stock_ff_units": stock_ff,
            "wb_confidence_coefficient": wb_confidence_coefficient,
            "wb_stock_evidence": wb_stock_evidence,
            "known_stock_units": None,
            "current_daily_sales": demand,
            "hard_daily_sales": None,
            "reserve_daily_sales": None,
            "target_daily_sales": None,
            "pace_ratio": None,
            "pace_change_pct": None,
            "days_cover": None,
            "bottleneck_date": None,
            "next_inbound": None,
            "subsequent_inbound": None,
            "milestones": [],
        }
    raw_opening = float(stock_wb)
    saleable_wb = raw_opening * wb_confidence_coefficient
    wb_stock_evidence["confidence_adjusted_wb_units"] = round(saleable_wb, 2)
    known_stock = saleable_wb + float(stock_ff)
    cumulative = known_stock
    milestones: list[dict[str, Any]] = []
    start_date = as_of_date
    dated_inbounds: dict[date, dict[str, Any]] = {}
    inbound_warnings: list[str] = []
    seen_inbound_identities: set[tuple[str, str, str, str]] = set()
    for raw_inbound in inbound_evidence:
        if not isinstance(raw_inbound, Mapping):
            inbound_warnings.append("Invalid registry inbound evidence excluded")
            continue
        inbound_date = _iso_date(str(raw_inbound.get("date") or ""))
        quantity = _optional_float(raw_inbound.get("quantity"))
        inbound_identity = (
            str(raw_inbound.get("source") or ""),
            str(raw_inbound.get("source_id") or ""),
            str(raw_inbound.get("date") or ""),
            str(raw_inbound.get("district_key") or ""),
        )
        if inbound_identity in seen_inbound_identities:
            inbound_warnings.append(
                f"Duplicate registry inbound identity excluded: {':'.join(inbound_identity)}"
            )
            continue
        seen_inbound_identities.add(inbound_identity)
        if (
            inbound_date is None
            or quantity is None
            or quantity <= 0
            or inbound_date < start_date
            or bool(raw_inbound.get("consumes_current_ff"))
        ):
            inbound_warnings.append(
                f"Registry inbound excluded: {raw_inbound.get('source')}:{raw_inbound.get('source_id')}"
            )
            continue
        bucket = dated_inbounds.setdefault(
            inbound_date,
            {"quantity": 0.0, "sources": [], "source_ids": []},
        )
        bucket["quantity"] += quantity
        bucket["sources"].append(str(raw_inbound.get("source") or "registry_inbound"))
        bucket["source_ids"].append(str(raw_inbound.get("source_id") or ""))
    if not dated_inbounds:
        return {
            **identity,
            "status": "Недостаточно данных",
            "quality": "unknown",
            "quality_warnings": list(raw.get("quality_warnings") or [])
            + inbound_warnings
            + ["Нет eligible exact production/in_transit поставок; target не рассчитывается"],
            "stock_wb_units": round(raw_opening, 2),
            "stock_ff_units": round(float(stock_ff), 2),
            "wb_confidence_coefficient": wb_confidence_coefficient,
            "wb_stock_evidence": wb_stock_evidence,
            "known_stock_units": round(known_stock, 2),
            "saleable_opening_units": round(known_stock, 2),
            "confidence_adjusted_wb_units": round(saleable_wb, 2),
            "current_daily_sales": round(demand, 4),
            "hard_daily_sales": None,
            "reserve_daily_sales": None,
            "target_daily_sales": None,
            "pace_ratio": None,
            "pace_ratio_basis": "no_supply_unknown",
            "pace_change_pct": None,
            "days_cover": round(known_stock / demand, 2) if demand > 0 else None,
            "bottleneck_date": None,
            "next_inbound": None,
            "subsequent_inbound": None,
            "milestones": [],
        }
    for inbound_date, evidence in sorted(dated_inbounds.items()):
        available_before = cumulative
        cumulative += float(evidence["quantity"])
        milestones.append(
            {
                "date": inbound_date.isoformat(),
                "quantity": round(float(evidence["quantity"]), 2),
                "available_before_arrival": round(available_before, 2),
                "cumulative_available": round(cumulative, 2),
                "days_from_calculation": (inbound_date - start_date).days + 1,
                "sources": list(evidence["sources"]),
                "source_ids": list(evidence["source_ids"]),
            }
        )
    horizon_date = max(dated_inbounds)
    safety_days = int(settings["safety_stock_days"])
    hard_candidates = [
        (float(item["available_before_arrival"]) / max(int(item["days_from_calculation"]), 1), item)
        for item in milestones
    ]
    reserve_candidates = [
        (
            float(item["available_before_arrival"])
            / max(int(item["days_from_calculation"]) + safety_days, 1),
            item,
        )
        for item in milestones
    ]
    hard, hard_milestone = min(hard_candidates, key=lambda pair: pair[0])
    reserve, reserve_milestone = min(reserve_candidates, key=lambda pair: pair[0])
    if demand > hard + 1e-9:
        status = "Дефицит"
        target = hard
        bottleneck = hard_milestone
    elif demand + 1e-9 < reserve:
        status = "Переизбыток"
        target = reserve
        bottleneck = reserve_milestone
    else:
        status = "Баланс"
        target = demand
        bottleneck = None
    zero_sales_launch = demand == 0 and target > 0
    pace_ratio = (
        target / demand
        if demand > 0
        else float(settings["bid_scale_max"])
        if zero_sales_launch
        else 1.0
    )
    quality_warnings = list(raw.get("quality_warnings") or []) + inbound_warnings
    if zero_sales_launch:
        quality_warnings.append(
            "Нет наблюдаемой эластичности при нулевых продажах; используется bounded launch ratio"
        )
    inbound_only = [item for item in milestones if float(item["quantity"]) > 0]
    return {
        **identity,
        "status": status,
        "quality": "partial" if quality_warnings else str(raw.get("quality") or "complete"),
        "quality_warnings": quality_warnings,
        "known_stock_units": round(known_stock, 2),
        "saleable_opening_units": round(known_stock, 2),
        "confidence_adjusted_wb_units": round(saleable_wb, 2),
        "stock_wb_units": round(raw_opening, 2),
        "stock_ff_units": round(float(stock_ff), 2),
        "raw_opening_stock_units": round(raw_opening, 2),
        "wb_confidence_coefficient": wb_confidence_coefficient,
        "wb_stock_evidence": wb_stock_evidence,
        "current_daily_sales": round(demand, 4),
        "hard_daily_sales": round(hard, 4),
        "reserve_daily_sales": round(reserve, 4),
        "target_daily_sales": round(target, 4),
        "pace_ratio": round(pace_ratio, 6),
        "pace_ratio_basis": "zero_sales_bounded_launch" if zero_sales_launch else "target_to_observed_sales",
        "pace_change_pct": round((pace_ratio - 1.0) * 100.0, 2),
        "days_cover": round(known_stock / demand, 2) if demand > 0 else None,
        "bottleneck_date": bottleneck.get("date") if bottleneck else None,
        "next_inbound": inbound_only[0] if inbound_only else None,
        "subsequent_inbound": inbound_only[1] if len(inbound_only) > 1 else None,
        "milestones": milestones,
        "horizon_date": horizon_date.isoformat(),
        "horizon_policy": "last_eligible_exact_registry_inbound",
    }


def _campaign_recommendation(
    raw: Mapping[str, Any],
    *,
    nm_id: int,
    pace_ratio: float | None,
) -> dict[str, Any]:
    advert_id = _optional_int(raw.get("advert_id")) or 0
    placement = str(raw.get("placement") or "")
    payment_type = str(raw.get("payment_type") or "").strip().lower()
    group = "new_cpc" if payment_type == "cpc" else "old_cpm" if payment_type == "cpm" else "other"
    current = _optional_float(raw.get("current_bid_rub"))
    min_bid = _optional_float(raw.get("min_bid_rub"))
    ratio = min(max(float(pace_ratio if pace_ratio is not None else 1.0), 0.25), 2.0)
    calculated = None
    if current is not None and current > 0:
        calculated = current * ratio
        if min_bid is not None:
            calculated = max(calculated, min_bid)
        calculated = float(Decimal(str(calculated)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    spend = _optional_float(raw.get("spend_rub") if "spend_rub" in raw else raw.get("sum"))
    orders = _optional_float(raw.get("orders"))
    cpo = spend / orders if spend is not None and orders and orders > 0 else None
    status = _optional_int(raw.get("status"))
    campaign_state = CAMPAIGN_STATE_BY_STATUS.get(status or 0, "")
    state_action = CAMPAIGN_STATE_ACTION_BY_STATUS.get(status or 0, "")
    identity_valid = bool(
        nm_id > 0
        and advert_id > 0
        and placement
        and payment_type in {"cpm", "cpc"}
        and status in SUPPORTED_CAMPAIGN_STATUSES
    )
    target_key = f"{nm_id}:{advert_id}:{placement}"
    return {
        "target_key": target_key,
        "nm_id": nm_id,
        "advert_id": advert_id,
        "campaign_name": str(raw.get("campaign_name") or raw.get("name") or ""),
        "campaign_group": group,
        "payment_type": payment_type,
        "placement": placement,
        "campaign_status": status,
        "campaign_state": campaign_state,
        "state_action_available": bool(identity_valid and state_action),
        "state_action": state_action,
        "state_action_label": (
            "остановить"
            if state_action == "pause"
            else "возобновить"
            if status == 11
            else "запустить"
            if state_action == "start"
            else ""
        ),
        "cpo_rub": round(cpo, 2) if cpo is not None else None,
        "orders": orders,
        "spend_rub": spend,
        "current_bid_rub": current,
        "calculated_target_bid_rub": calculated,
        "manual_target_bid_rub": None,
        "final_target_bid_rub": calculated,
        "min_bid_rub": min_bid,
        "identity_valid": identity_valid,
        "inventory_evidence_actionable": pace_ratio is not None,
        "manual_override_allowed": bool(identity_valid and pace_ratio is not None),
        "can_apply": bool(
            identity_valid
            and current is not None
            and calculated is not None
            and float(current) != float(calculated)
        ),
        "calculation_reason": "ставка масштабирована по отношению целевого темпа к текущему; WB minimum является нижней границей",
    }


def _rub_to_minor(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(
            (Decimal(str(value)) * Decimal("100")).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
        )
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_timestamp(value: str) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        moment = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if moment.tzinfo is None or moment.utcoffset() is None:
        return None
    return moment.astimezone(timezone.utc)


def _timestamp_age_seconds(value: str, now: datetime) -> float:
    moment = _parse_timestamp(value)
    if moment is None:
        return float("inf")
    return max(
        (now.astimezone(timezone.utc) - moment).total_seconds(),
        0.0,
    )


def _bounded_error(exc: Exception) -> str:
    text = " ".join(str(exc).split())
    return text[:800]


def _public_preflight_result(value: Mapping[str, Any]) -> dict[str, Any]:
    allowed = {
        "ok",
        "error_code",
        "message",
        "observed_bid_minor",
        "observed_campaign_status",
        "observed_campaign_state",
        "minimum_bid_minor",
        "payment_type",
        "candidate_nm_ids",
        "safety_threshold_policy",
        "safety_warnings",
    }
    return {key: value.get(key) for key in sorted(allowed) if key in value}


def _allocate_campaign_targets(
    recommendations: Sequence[Mapping[str, Any]],
    *,
    pace_ratio: float | None,
) -> list[dict[str, Any]]:
    """Route growth/cuts using relative group CPO; fail closed on weak evidence."""

    result = [dict(item) for item in recommendations]
    eligible = [
        item
        for item in result
        if item.get("identity_valid")
        and item.get("current_bid_rub") is not None
        and item.get("calculated_target_bid_rub") is not None
        and item.get("campaign_group") in {"new_cpc", "old_cpm"}
    ]
    group_rows: dict[str, list[dict[str, Any]]] = {}
    for item in eligible:
        group_rows.setdefault(str(item["campaign_group"]), []).append(item)
    group_evidence: dict[str, dict[str, Any]] = {}
    for group, items in group_rows.items():
        complete = all(
            item.get("spend_rub") is not None
            and item.get("orders") is not None
            and float(item.get("orders") or 0) > 0
            for item in items
        )
        spend = sum(float(item.get("spend_rub") or 0) for item in items)
        orders = sum(float(item.get("orders") or 0) for item in items)
        group_evidence[group] = {
            "campaign_count": len(items),
            "spend_rub": round(spend, 2),
            "orders": round(orders, 2),
            "cpo_rub": round(spend / orders, 2) if complete and orders > 0 else None,
            "quality": "complete" if complete and orders > 0 else "insufficient_stats",
        }
    selected_group: str | None = None
    action = "hold"
    quality = "complete"
    ratio = float(pace_ratio) if pace_ratio is not None else None
    if ratio is None:
        quality = "insufficient_inventory_evidence"
        action = "hold_conservative"
    elif abs(ratio - 1.0) <= 1e-9:
        action = "hold_balanced"
    elif len(group_evidence) == 1:
        selected_group = next(iter(group_evidence))
        action = "increase_single_group" if ratio > 1 else "decrease_single_group"
        quality = "single_group_no_relative_comparison"
    elif len(group_evidence) >= 2 and all(
        item["quality"] == "complete" for item in group_evidence.values()
    ):
        ranked = sorted(
            group_evidence,
            key=lambda group: (float(group_evidence[group]["cpo_rub"]), group),
        )
        if ratio > 1:
            selected_group = ranked[0]
            action = "increase_more_efficient_group"
        else:
            selected_group = ranked[-1]
            action = "decrease_less_efficient_group"
    else:
        quality = "insufficient_stats"
        action = "hold_conservative"
    for item in result:
        selected = selected_group is not None and item.get("campaign_group") == selected_group
        if not selected:
            item["calculated_target_bid_rub"] = item.get("current_bid_rub")
            item["final_target_bid_rub"] = item.get("current_bid_rub")
            item["can_apply"] = False
        item["relative_efficiency"] = {
            "basis": "group_cpo",
            "groups": group_evidence,
            "selected_group": selected_group,
            "quality": quality,
        }
        item["recommendation_quality"] = quality
        item["allocation_action"] = action if selected or selected_group is None else "hold_other_group"
        item["calculation_reason"] = (
            "рост направлен в группу с меньшим CPO"
            if action == "increase_more_efficient_group" and selected
            else "снижение начато с группы с большим CPO"
            if action == "decrease_less_efficient_group" and selected
            else "консервативно без изменения: недостаточно сопоставимой CPO evidence"
            if action == "hold_conservative"
            else "группа удерживается, пока изменение направлено в выбранную по CPO группу"
            if not selected
            else "единственная eligible группа масштабирована по целевому темпу"
        )
    return result


def _calculation_wb_stock_evidence(
    raw: Mapping[str, Any],
    *,
    stock_wb: float | None,
    wb_confidence_coefficient: float,
) -> dict[str, Any]:
    source = raw.get("wb_stock_evidence")
    evidence = dict(source) if isinstance(source, Mapping) else {}
    if not evidence:
        evidence = {
            "source": "sku_inventory_balance_input",
            "source_contract": "sku_management/current_incident_projection",
            "mode": (
                "warehouse_granular_incident_projection"
                if stock_wb is not None
                else "missing"
            ),
            "quality": "exact_warehouse_projection" if stock_wb is not None else "unknown",
            "warehouse_granularity_complete": stock_wb is not None,
            "incident_projection_applied": stock_wb is not None,
            "warning": "",
        }
    evidence["stock_wb_units"] = stock_wb
    evidence["wb_confidence_coefficient"] = wb_confidence_coefficient
    evidence["confidence_adjusted_wb_units"] = (
        round(float(stock_wb) * wb_confidence_coefficient, 2)
        if stock_wb is not None
        else None
    )
    return evidence


def _sanitize_calculation_settings(raw: Any) -> dict[str, Any]:
    source = dict(raw) if isinstance(raw, Mapping) else {}
    return {
        "wb_confidence_coefficient": _bounded_float(source.get("wb_confidence_coefficient"), 0.5, 0.0, 1.0),
        "safety_stock_days": _bounded_int(source.get("safety_stock_days"), 10, 0, 120),
        "sales_period_days": _choice_int(source.get("sales_period_days"), 7, {7, 14, 30, 60}),
        "bid_scale_min": 0.25,
        "bid_scale_max": 2.0,
        "automatic_training": False,
    }


def _sanitize_table_preferences(raw: Any) -> dict[str, Any]:
    source = dict(raw) if isinstance(raw, Mapping) else {}
    allowed = set(BALANCE_COLUMNS)
    order = list(MANDATORY_COLUMNS)
    for item in source.get("column_order") or []:
        key = str(item)
        if key in allowed and key not in order:
            order.append(key)
    for key in BALANCE_COLUMNS:
        if key not in order:
            order.append(key)
    visible = []
    requested_visible = source.get("visible_columns")
    for item in (
        requested_visible
        if isinstance(requested_visible, Sequence)
        and not isinstance(requested_visible, str)
        and len(requested_visible) > 0
        else DEFAULT_VISIBLE_COLUMNS
    ):
        key = str(item)
        if key in allowed and key not in visible:
            visible.append(key)
    for key in reversed(MANDATORY_COLUMNS):
        if key not in visible:
            visible.insert(0, key)
    widths = {}
    for key, value in (source.get("column_widths") or {}).items():
        if str(key) in allowed and _optional_int(value) is not None:
            widths[str(key)] = min(max(int(value), 60), 600)
    filters = dict(source.get("filters") or {})
    preset = str(source.get("preset") or "all")
    if preset not in {"all", "deficit", "excess", "actionable"}:
        preset = "all"
    return {
        "visible_columns": visible,
        "column_order": order,
        "column_widths": widths,
        "filters": {
            "search": str(filters.get("search") or "")[:200],
            "status": str(filters.get("status") or "")[:40],
        },
        "preset": preset,
    }


def _find_target(calculation: Mapping[str, Any], target_key: str) -> dict[str, Any] | None:
    for row in calculation.get("rows") or []:
        for target in row.get("campaign_recommendations") or []:
            if str(target.get("target_key") or "") == target_key:
                return dict(target)
    return None


def _row_sort_key(row: Mapping[str, Any]) -> tuple[int, str, int]:
    rank = {"Дефицит": 0, "Переизбыток": 1, "Баланс": 2, "Недостаточно данных": 3}
    return rank.get(str(row.get("status")), 4), str(row.get("bottleneck_date") or "9999-12-31"), int(row.get("nm_id") or 0)


def _campaigns_label(items: Any) -> str:
    values = []
    for item in items or []:
        values.append(
            f"{item.get('advert_id')} {item.get('campaign_name') or ''}: "
            f"CPO {item.get('cpo_rub') if item.get('cpo_rub') is not None else '—'}; "
            f"{item.get('current_bid_rub')} → {item.get('final_target_bid_rub')}"
        )
    return " | ".join(values)


def _inbound_label(item: Any) -> str:
    if not isinstance(item, Mapping):
        return ""
    return f"{item.get('date') or ''}: {item.get('quantity') or 0}"


def _wb_stock_evidence_label(row: Mapping[str, Any]) -> str:
    evidence = row.get("wb_stock_evidence")
    if not isinstance(evidence, Mapping):
        return ""
    coefficient = row.get("wb_confidence_coefficient")
    if str(evidence.get("mode") or "") == "aggregate_per_sku_total":
        return (
            f"Официальный агрегат WB × {coefficient}; без раскладки по складам; "
            f"{evidence.get('raw_rows_digest') or 'digest unavailable'}"
        )
    return f"Складская incident-проекция WB × {coefficient}"


def _operation_token(value: Any, field: str) -> str:
    normalized = str(value or "").strip()
    if not _OPERATION_ID_RE.fullmatch(normalized):
        raise SkuInventoryBalanceError(
            f"{field} должен содержать 8..128 безопасных символов.",
            http_status=422,
            payload={"code": f"invalid_{field}"},
        )
    return normalized


def _iso_duration_ms(started_at: str, finished_at: str) -> int:
    if not started_at or not finished_at:
        return 0
    try:
        started = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
        finished = datetime.fromisoformat(finished_at.replace("Z", "+00:00"))
    except ValueError:
        return 0
    return max(0, int((finished - started).total_seconds() * 1000))


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)


def _digest(value: Any) -> str:
    return "sha256:" + sha256(_json(value).encode("utf-8")).hexdigest()


def _positive_money(value: Any, field: str) -> float:
    try:
        normalized = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise SkuInventoryBalanceError(f"{field} must be a positive money value") from exc
    if normalized <= 0 or normalized > Decimal("100000"):
        raise SkuInventoryBalanceError(f"{field} must be in (0, 100000]")
    return float(normalized)


def _optional_int(value: Any) -> int | None:
    if isinstance(value, bool) or value in {None, ""}:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _optional_float(value: Any) -> float | None:
    if isinstance(value, bool) or value in {None, ""}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _iso_date(value: str) -> date | None:
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        return None


def _bounded_int(value: Any, default: int, minimum: int, maximum: int) -> int:
    normalized = _optional_int(value)
    if normalized is None:
        return default
    return min(max(normalized, minimum), maximum)


def _choice_int(value: Any, default: int, choices: set[int]) -> int:
    normalized = _optional_int(value)
    return normalized if normalized in choices else default


def _bounded_float(value: Any, default: float, minimum: float, maximum: float) -> float:
    normalized = _optional_float(value)
    if normalized is None:
        return default
    return min(max(normalized, minimum), maximum)
