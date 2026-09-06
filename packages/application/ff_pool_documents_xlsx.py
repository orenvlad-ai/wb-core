"""Bounded XLSX templates and parsers for Stage 2 FF pool documents.

The HTTP layer is intentionally absent.  A future route must call
``validate_xlsx_request_seam`` before buffering and then use these helpers.
"""

from __future__ import annotations

from dataclasses import asdict
from decimal import Decimal, InvalidOperation
import hashlib
import io
import json
from pathlib import PurePosixPath
import re
from typing import Any, Iterable, Mapping
import zipfile
import xml.etree.ElementTree as ET

from packages.contracts.ff_pool_documents import XlsxParserLimits


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ALLOWED_CONTENT_TYPES = frozenset({XLSX_CONTENT_TYPE, "application/octet-stream"})
CHINA_SHEET = "China acceptance"
INVENTORY_SHEET = "Inventory"
LISTS_SHEET = "__lists"
CONTRACT_SHEET = "__contract"
CHINA_HEADERS = (
    "nmId",
    "Штрихкод",
    "SKU",
    "Отправлено, шт",
    "Принято, шт",
    "Капитал принятия, RUB",
    "FBO, шт",
    "FBS, шт",
    "Расхождение",
    "Комментарий",
)
INVENTORY_HEADERS = (
    "nmId",
    "Штрихкод",
    "SKU",
    "FBO actual, шт",
    "FBS actual, шт",
)
CONTRACT_NAME = "ff_pool_documents_xlsx_v1"
DEFAULT_LIMITS = XlsxParserLimits()
MAX_SQLITE_INTEGER = 9_223_372_036_854_775_807


class FfPoolXlsxError(ValueError):
    """Machine-readable fail-closed workbook error."""

    def __init__(self, code: str, message: str, *, details: Any = None) -> None:
        super().__init__(message)
        self.code = code
        self.details = details


def validate_xlsx_request_seam(
    *,
    content_length: int | None,
    filename: str,
    content_type: str,
    limits: XlsxParserLimits = DEFAULT_LIMITS,
) -> None:
    """Validate the future HTTP seam before request bytes are buffered."""

    _validate_envelope(filename=filename, content_type=content_type)
    if content_length is None:
        return
    if isinstance(content_length, bool) or content_length < 0:
        raise FfPoolXlsxError(
            "invalid_content_length",
            "XLSX Content-Length must be a non-negative integer",
        )
    if content_length > limits.max_request_bytes:
        raise FfPoolXlsxError(
            "request_too_large",
            "XLSX request exceeds the pre-buffering limit",
            details={"limit_bytes": limits.max_request_bytes, "actual_bytes": content_length},
        )


def generate_china_acceptance_workbook(
    *,
    facilities: Iterable[Mapping[str, Any]],
    shipment_lines: Iterable[Mapping[str, Any]],
    source_revision: str,
    selected_facility_id: str = "",
) -> bytes:
    active = _active_facilities(facilities)
    if not active:
        raise FfPoolXlsxError(
            "no_active_facilities",
            "No active FF facilities are available for the template",
        )
    normalized_lines = _normalize_source_lines(shipment_lines)
    selected = str(selected_facility_id or "").strip()
    if selected and selected not in {item["facility_id"] for item in active}:
        raise FfPoolXlsxError(
            "unknown_facility",
            "Selected facility is not active",
            details={"facility_id": selected},
        )
    fingerprint = _template_fingerprint(
        profile="china_acceptance_v1",
        facilities=active,
        source_revision=source_revision,
        rows=normalized_lines,
    )
    workbook = _new_workbook()
    sheet = workbook.active
    sheet.title = CHINA_SHEET
    sheet.sheet_view.showGridLines = False
    _set_text(sheet["A1"], "Распределение приёмки Китай → FF")
    _set_text(sheet["A2"], "Facility")
    _set_text(sheet["B2"], _facility_label(active, selected) if selected else "")
    _set_text(sheet["A3"], "Source revision")
    _set_text(sheet["B3"], source_revision)
    sheet.append([])
    sheet.append(list(CHINA_HEADERS))
    for item in normalized_lines:
        sheet.append(
            [
                item["nm_id"],
                item["barcode"],
                item["sku"],
                item["quantity"],
                item["quantity"],
                _decimal_text(item["capital_rub"]),
                0,
                0,
                "",
                "",
            ]
        )
        _set_text(sheet.cell(sheet.max_row, 2), item["barcode"])
        _set_text(sheet.cell(sheet.max_row, 3), item["sku"])
    _format_sheet(sheet, header_row=5, max_row=max(5, sheet.max_row), max_column=10)
    for row in range(6, sheet.max_row + 1):
        sheet.cell(row, 2).number_format = "@"
        sheet.cell(row, 4).number_format = "0"
        sheet.cell(row, 5).number_format = "0"
        sheet.cell(row, 6).number_format = "#,##0.00"
        sheet.cell(row, 7).number_format = "0"
        sheet.cell(row, 8).number_format = "0"
    _add_table(sheet, name="ChinaAcceptanceTable", ref=f"A5:J{max(6, sheet.max_row)}")
    _add_lists_and_validation(
        workbook,
        sheet=sheet,
        active=active,
        facility_cell="B2",
    )
    _add_contract_sheet(
        workbook,
        profile="china_acceptance_v1",
        fingerprint=fingerprint,
        source_revision=str(source_revision),
        selected_facility_id=selected,
        scope="both",
    )
    return _save_workbook(workbook)


def generate_inventory_workbook(
    *,
    facilities: Iterable[Mapping[str, Any]],
    facility_id: str,
    scope: str,
    catalog: Iterable[Mapping[str, Any]],
    source_revision: str,
    targets: Mapping[tuple[int, str], Any] | None = None,
) -> bytes:
    active = _active_facilities(facilities)
    if not active:
        raise FfPoolXlsxError(
            "no_active_facilities",
            "No active FF facilities are available for the template",
        )
    selected = str(facility_id or "").strip()
    if selected not in {item["facility_id"] for item in active}:
        raise FfPoolXlsxError(
            "unknown_facility",
            "Selected facility is not active",
            details={"facility_id": selected},
        )
    normalized_scope = _scope(scope)
    normalized_catalog = _normalize_catalog(catalog)
    target_values = dict(targets or {})
    fingerprint = _template_fingerprint(
        profile="inventory_v1",
        facilities=active,
        source_revision=source_revision,
        rows=normalized_catalog,
        facility_id=selected,
        scope=normalized_scope,
    )
    workbook = _new_workbook()
    sheet = workbook.active
    sheet.title = INVENTORY_SHEET
    sheet.sheet_view.showGridLines = False
    _set_text(sheet["A1"], "Инвентаризация facility × pool")
    _set_text(sheet["A2"], "Facility")
    _set_text(sheet["B2"], _facility_label(active, selected))
    _set_text(sheet["A3"], "Scope")
    _set_text(sheet["B3"], normalized_scope)
    _set_text(sheet["A4"], "Source revision")
    _set_text(sheet["B4"], source_revision)
    sheet.append(list(INVENTORY_HEADERS))
    for item in normalized_catalog:
        nm_id = int(item["nm_id"])
        sheet.append(
            [
                nm_id,
                item["barcode"],
                item["sku"],
                _exact_nonnegative_int(target_values.get((nm_id, "FBO"), 0), field="FBO target"),
                _exact_nonnegative_int(target_values.get((nm_id, "FBS"), 0), field="FBS target"),
            ]
        )
        _set_text(sheet.cell(sheet.max_row, 2), item["barcode"])
        _set_text(sheet.cell(sheet.max_row, 3), item["sku"])
    _format_sheet(sheet, header_row=5, max_row=max(5, sheet.max_row), max_column=5)
    for row in range(6, sheet.max_row + 1):
        sheet.cell(row, 2).number_format = "@"
        sheet.cell(row, 4).number_format = "0"
        sheet.cell(row, 5).number_format = "0"
    _add_table(sheet, name="FacilityPoolInventoryTable", ref=f"A5:E{max(6, sheet.max_row)}")
    _add_lists_and_validation(
        workbook,
        sheet=sheet,
        active=active,
        facility_cell="B2",
        scope_cell="B3",
    )
    _add_contract_sheet(
        workbook,
        profile="inventory_v1",
        fingerprint=fingerprint,
        source_revision=str(source_revision),
        selected_facility_id=selected,
        scope=normalized_scope,
    )
    return _save_workbook(workbook)


def parse_china_acceptance_workbook(
    source_bytes: bytes,
    *,
    filename: str,
    content_type: str,
    facilities: Iterable[Mapping[str, Any]],
    shipment_lines: Iterable[Mapping[str, Any]],
    source_revision: str,
    limits: XlsxParserLimits = DEFAULT_LIMITS,
) -> dict[str, Any]:
    _preflight_xlsx(
        source_bytes,
        filename=filename,
        content_type=content_type,
        limits=limits,
    )
    workbook = _load_workbook(source_bytes)
    _assert_allowed_sheets(workbook, {CHINA_SHEET, LISTS_SHEET, CONTRACT_SHEET})
    active = _active_facilities(facilities)
    if not active:
        raise FfPoolXlsxError("no_active_facilities", "No active FF facilities exist")
    normalized_lines = _normalize_source_lines(shipment_lines)
    expected_fingerprint = _template_fingerprint(
        profile="china_acceptance_v1",
        facilities=active,
        source_revision=source_revision,
        rows=normalized_lines,
    )
    metadata = _read_contract(workbook)
    _assert_contract(
        metadata,
        profile="china_acceptance_v1",
        expected_fingerprint=expected_fingerprint,
        source_revision=source_revision,
    )
    sheet = workbook[CHINA_SHEET]
    _assert_headers(sheet, row=5, expected=CHINA_HEADERS)
    selected_facility = str(sheet["B2"].value or "").strip()
    facility_id = _facility_id_from_label(active, selected_facility)
    if not facility_id:
        raise FfPoolXlsxError(
            "unknown_or_inactive_facility",
            "Workbook must select one active facility",
            details={"facility": selected_facility},
        )
    resolved_source = {int(item["nm_id"]): item for item in normalized_lines}
    barcode_index = _barcode_index(normalized_lines)
    seen: set[int] = set()
    allocations: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for row_no in range(6, sheet.max_row + 1):
        values = [sheet.cell(row_no, column) for column in range(1, 11)]
        if all(cell.value in (None, "") for cell in values):
            continue
        try:
            resolved = _resolve_identity_cells(
                values[0],
                values[1],
                by_nm=resolved_source,
                by_barcode=barcode_index,
            )
            nm_id = int(resolved["nm_id"])
            if nm_id in seen:
                raise FfPoolXlsxError(
                    "duplicate_resolved_sku",
                    "Workbook resolves more than one row to the same SKU",
                    details={"nm_id": nm_id},
                )
            seen.add(nm_id)
            expected = _whole_cell(values[3], field="expected quantity", positive=True)
            if expected != int(resolved["quantity"]):
                raise FfPoolXlsxError(
                    "source_quantity_tampered",
                    "Expected shipment quantity differs from the immutable source",
                    details={"nm_id": nm_id, "expected": int(resolved["quantity"]), "actual": expected},
                )
            accepted = _whole_cell(values[4], field="accepted quantity")
            fbo = _whole_cell(values[6], field="FBO quantity")
            fbs = _whole_cell(values[7], field="FBS quantity")
            if fbo + fbs != accepted:
                raise FfPoolXlsxError(
                    "allocation_quantity_mismatch",
                    "FBO + FBS must exactly equal the accepted quantity",
                    details={"nm_id": nm_id, "accepted": accepted, "fbo": fbo, "fbs": fbs},
                )
            discrepancy = str(values[8].value or "").strip().casefold()
            derived = "" if accepted == expected else ("shortage" if accepted < expected else "surplus")
            aliases = {"": "", "недостача": "shortage", "shortage": "shortage", "излишек": "surplus", "surplus": "surplus", "пересорт": "mis_sort", "mis-sort": "mis_sort", "mis_sort": "mis_sort"}
            if discrepancy not in aliases:
                raise FfPoolXlsxError("invalid_discrepancy_type", "Unsupported discrepancy type", details={"value": discrepancy})
            declared = aliases[discrepancy]
            if declared and declared != derived and declared != "mis_sort":
                raise FfPoolXlsxError("discrepancy_type_mismatch", "Declared discrepancy does not match quantities")
            base_capital = Decimal(str(resolved["capital_rub"]))
            accepted_capital = base_capital * Decimal(accepted) / Decimal(expected)
            row = {
                    "nm_id": nm_id,
                    "barcode": str(resolved["barcode"]),
                    "sku": str(resolved["sku"]),
                    "expected_quantity": expected,
                    "accepted_quantity": accepted,
                    "accepted_capital_rub": _decimal_text(accepted_capital),
                    "quantity_fbo": fbo,
                    "quantity_fbs": fbs,
                    "discrepancy_type": declared or derived,
                    "discrepancy_quantity": abs(accepted - expected),
                    "comment": str(values[9].value or "").strip()[:500],
                    "row_no": row_no,
                    "identity_evidence_digest": str(resolved["identity_evidence_digest"]),
                }
            allocations.append(row)
        except FfPoolXlsxError as exc:
            errors.append({"row": row_no, "code": exc.code, "details": exc.details})
    missing = sorted(set(resolved_source) - seen)
    if missing:
        errors.append({"code": "incomplete_accepted_composition", "nm_ids": missing[:100]})
    if errors:
        raise FfPoolXlsxError(
            "invalid_china_acceptance_rows",
            "China acceptance workbook contains blocked rows",
            details=errors[:100],
        )
    if not any(int(item["accepted_quantity"]) > 0 for item in allocations):
        raise FfPoolXlsxError("empty_actual_acceptance", "At least one actual accepted quantity must be positive")
    return {
        "contract_name": CONTRACT_NAME,
        "profile": "china_acceptance_v1",
        "facility_id": facility_id,
        "source_revision": str(source_revision),
        "source_filename": str(filename),
        "source_sha256": _sha256(source_bytes),
        "template_fingerprint": expected_fingerprint,
        "allocations": sorted(allocations, key=lambda item: int(item["nm_id"])),
    }


def build_china_acceptance_form_manifest(
    *,
    shipment_lines: Iterable[Mapping[str, Any]],
    source_revision: str,
    facilities: Iterable[Mapping[str, Any]],
    facility_id: str,
    mode: str,
    rows: Any = None,
) -> dict[str, Any]:
    """Build the same receipt from operator quantities and trusted source only."""

    source = _normalize_source_lines(shipment_lines)
    if facility_id not in {str(item["facility_id"]) for item in _active_facilities(facilities)}:
        raise FfPoolXlsxError("unknown_or_inactive_facility", "Выберите действующий склад ФФ.")
    if mode not in {"FBS", "FBO", "split"}:
        raise FfPoolXlsxError("invalid_acceptance_mode", "Выберите способ приёмки.")

    def whole(value: Any) -> int:
        if isinstance(value, bool) or not re.fullmatch(r"[0-9]{1,16}", str(value)) or int(value) > 9_007_199_254_740_991:
            raise FfPoolXlsxError("invalid_acceptance_quantity", "Количество должно быть целым и неотрицательным.")
        return int(value)

    supplied: dict[int, Mapping[str, Any]] = {}
    if rows is not None:
        if not isinstance(rows, list):
            raise FfPoolXlsxError("invalid_acceptance_rows", "Некорректный состав приёмки.")
        for row in rows:
            if not isinstance(row, Mapping) or set(row) - {
                "nm_id", "accepted_quantity", "quantity_fbs", "quantity_fbo", "discrepancy_type", "comment"
            }:
                raise FfPoolXlsxError("acceptance_fields_not_allowed", "Состав и стоимость поставки определяются сервером.")
            nm_id = whole(row.get("nm_id"))
            if nm_id in supplied:
                raise FfPoolXlsxError("duplicate_resolved_sku", "SKU повторяется в приёмке.")
            supplied[nm_id] = row
        if set(supplied) != {int(item["nm_id"]) for item in source}:
            raise FfPoolXlsxError("incomplete_accepted_composition", "Состав приёмки должен совпадать с поставкой.")
    elif mode == "split":
        raise FfPoolXlsxError("incomplete_accepted_composition", "Заполните распределение по каждому SKU.")
    allocations = []
    for item in source:
        nm_id = int(item["nm_id"])
        row = supplied.get(nm_id, {})
        expected = int(item["quantity"])
        accepted = whole(row.get("accepted_quantity", expected))
        fbs = whole(row.get("quantity_fbs")) if mode == "split" else (accepted if mode == "FBS" else 0)
        fbo = whole(row.get("quantity_fbo")) if mode == "split" else (accepted if mode == "FBO" else 0)
        if fbs + fbo != accepted:
            raise FfPoolXlsxError("allocation_quantity_mismatch", "Сумма FBS и FBO должна равняться фактически принятому количеству.", details={"nm_id": nm_id})
        derived = "" if accepted == expected else ("shortage" if accepted < expected else "surplus")
        declared = str(row.get("discrepancy_type") or "")
        if declared not in {"", "shortage", "surplus", "mis_sort"} or (declared and declared not in {derived, "mis_sort"}):
            raise FfPoolXlsxError("discrepancy_type_mismatch", "Вид расхождения не соответствует количеству.")
        allocations.append({
            "nm_id": nm_id,
            "barcode": item["barcode"],
            "sku": item["sku"],
            "expected_quantity": expected,
            "accepted_quantity": accepted,
            "accepted_capital_rub": _decimal_text(Decimal(item["capital_rub"]) * Decimal(accepted) / Decimal(expected)),
            "quantity_fbs": fbs,
            "quantity_fbo": fbo,
            "discrepancy_type": declared or derived,
            "discrepancy_quantity": abs(accepted - expected),
            "comment": str(row.get("comment") or "")[:500],
            "identity_evidence_digest": item["identity_evidence_digest"],
        })
    if not any(item["accepted_quantity"] for item in allocations):
        raise FfPoolXlsxError("empty_actual_acceptance", "Укажите хотя бы одну принятую единицу.")
    return {
        "contract_name": CONTRACT_NAME,
        "profile": "china_acceptance_v1",
        "facility_id": facility_id,
        "source_revision": source_revision,
        "allocations": sorted(allocations, key=lambda item: item["nm_id"]),
    }


def parse_inventory_workbook(
    source_bytes: bytes,
    *,
    filename: str,
    content_type: str,
    facilities: Iterable[Mapping[str, Any]],
    catalog: Iterable[Mapping[str, Any]],
    source_revision: str,
    limits: XlsxParserLimits = DEFAULT_LIMITS,
) -> dict[str, Any]:
    _preflight_xlsx(
        source_bytes,
        filename=filename,
        content_type=content_type,
        limits=limits,
    )
    workbook = _load_workbook(source_bytes)
    _assert_allowed_sheets(workbook, {INVENTORY_SHEET, LISTS_SHEET, CONTRACT_SHEET})
    active = _active_facilities(facilities)
    if not active:
        raise FfPoolXlsxError("no_active_facilities", "No active FF facilities exist")
    normalized_catalog = _normalize_catalog(catalog)
    metadata = _read_contract(workbook)
    facility_id = str(metadata.get("selected_facility_id") or "").strip()
    normalized_scope = _scope(str(metadata.get("scope") or ""))
    expected_fingerprint = _template_fingerprint(
        profile="inventory_v1",
        facilities=active,
        source_revision=source_revision,
        rows=normalized_catalog,
        facility_id=facility_id,
        scope=normalized_scope,
    )
    _assert_contract(
        metadata,
        profile="inventory_v1",
        expected_fingerprint=expected_fingerprint,
        source_revision=source_revision,
    )
    sheet = workbook[INVENTORY_SHEET]
    _assert_headers(sheet, row=5, expected=INVENTORY_HEADERS)
    selected_in_sheet = _facility_id_from_label(
        active,
        str(sheet["B2"].value or "").strip(),
    )
    scope_in_sheet = _scope(str(sheet["B3"].value or ""))
    if selected_in_sheet != facility_id or scope_in_sheet != normalized_scope:
        raise FfPoolXlsxError(
            "template_scope_tampered",
            "Facility or pool scope differs from the signed template contract",
        )
    if facility_id not in {item["facility_id"] for item in active}:
        raise FfPoolXlsxError(
            "unknown_or_inactive_facility",
            "Workbook facility is not active",
            details={"facility_id": facility_id},
        )
    by_nm = {int(item["nm_id"]): item for item in normalized_catalog}
    by_barcode = _barcode_index(normalized_catalog)
    seen: set[int] = set()
    targets: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for row_no in range(6, sheet.max_row + 1):
        values = [sheet.cell(row_no, column) for column in range(1, 6)]
        if all(cell.value in (None, "") for cell in values):
            continue
        try:
            resolved = _resolve_identity_cells(
                values[0], values[1], by_nm=by_nm, by_barcode=by_barcode
            )
            nm_id = int(resolved["nm_id"])
            if nm_id in seen:
                raise FfPoolXlsxError(
                    "duplicate_resolved_sku",
                    "Workbook resolves more than one row to the same SKU",
                    details={"nm_id": nm_id},
                )
            seen.add(nm_id)
            target_fbo = _whole_cell(values[3], field="FBO target")
            target_fbs = _whole_cell(values[4], field="FBS target")
            targets.append(
                {
                    "nm_id": nm_id,
                    "barcode": str(resolved["barcode"]),
                    "target_fbo": target_fbo,
                    "target_fbs": target_fbs,
                    "row_no": row_no,
                    "identity_evidence_digest": str(resolved["identity_evidence_digest"]),
                }
            )
        except FfPoolXlsxError as exc:
            errors.append({"row": row_no, "code": exc.code, "details": exc.details})
    missing = sorted(set(by_nm) - seen)
    if missing:
        errors.append({"code": "incomplete_selected_scope_coverage", "nm_ids": missing[:100]})
    if errors:
        raise FfPoolXlsxError(
            "invalid_inventory_rows",
            "Inventory workbook contains blocked rows",
            details=errors[:100],
        )
    return {
        "contract_name": CONTRACT_NAME,
        "profile": "inventory_v1",
        "facility_id": facility_id,
        "scope": normalized_scope,
        "selected_pools": list(_selected_pools(normalized_scope)),
        "source_revision": str(source_revision),
        "source_filename": str(filename),
        "source_sha256": _sha256(source_bytes),
        "template_fingerprint": expected_fingerprint,
        "targets": sorted(targets, key=lambda item: int(item["nm_id"])),
    }


def _preflight_xlsx(
    source_bytes: bytes,
    *,
    filename: str,
    content_type: str,
    limits: XlsxParserLimits,
) -> None:
    _validate_envelope(filename=filename, content_type=content_type)
    if not source_bytes:
        raise FfPoolXlsxError("empty_file", "XLSX file is empty")
    if len(source_bytes) > limits.max_file_bytes:
        raise FfPoolXlsxError(
            "file_too_large",
            "XLSX file exceeds the parser limit",
            details={"limit_bytes": limits.max_file_bytes, "actual_bytes": len(source_bytes)},
        )
    try:
        archive = zipfile.ZipFile(io.BytesIO(source_bytes), "r")
    except zipfile.BadZipFile as exc:
        raise FfPoolXlsxError("malformed_ooxml", "File is not a valid OOXML ZIP") from exc
    with archive:
        infos = archive.infolist()
        if len(infos) > limits.max_zip_entries:
            raise FfPoolXlsxError(
                "zip_entry_limit_exceeded",
                "XLSX contains too many ZIP entries",
                details={"limit": limits.max_zip_entries, "actual": len(infos)},
            )
        names = [item.filename for item in infos]
        if len(names) != len(set(names)):
            raise FfPoolXlsxError("duplicate_zip_entry", "XLSX contains duplicate ZIP paths")
        required = {"[Content_Types].xml", "xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
        if not required.issubset(names):
            raise FfPoolXlsxError("malformed_ooxml", "XLSX is missing required OOXML parts")
        total_uncompressed = 0
        worksheet_cells = 0
        worksheet_rows = 0
        for info in infos:
            path = PurePosixPath(info.filename)
            if path.is_absolute() or ".." in path.parts or "\\" in info.filename:
                raise FfPoolXlsxError("unsafe_zip_path", "XLSX contains an unsafe ZIP path")
            lowered = info.filename.lower()
            if lowered.endswith(".bin") or "vbaproject" in lowered or "macrosheet" in lowered:
                raise FfPoolXlsxError("macros_forbidden", "Macros are forbidden in Stage 2 XLSX")
            if "externallinks" in lowered or "/embeddings/" in lowered or "/activex/" in lowered:
                raise FfPoolXlsxError(
                    "external_links_forbidden",
                    "External links and embedded active content are forbidden",
                )
            total_uncompressed += int(info.file_size)
            if total_uncompressed > limits.max_uncompressed_bytes:
                raise FfPoolXlsxError(
                    "zip_uncompressed_limit_exceeded",
                    "XLSX uncompressed size exceeds the parser limit",
                    details={
                        "limit": limits.max_uncompressed_bytes,
                        "actual": total_uncompressed,
                    },
                )
            if info.file_size > limits.max_entry_uncompressed_bytes:
                raise FfPoolXlsxError(
                    "zip_entry_too_large",
                    "One XLSX ZIP entry exceeds the uncompressed limit",
                    details={"entry": info.filename, "actual": info.file_size},
                )
            if info.file_size and info.compress_size == 0:
                raise FfPoolXlsxError("zip_bomb_ratio", "XLSX has an invalid compression ratio")
            if info.file_size / max(info.compress_size, 1) > limits.max_compression_ratio:
                raise FfPoolXlsxError(
                    "zip_bomb_ratio",
                    "XLSX compression ratio exceeds the parser limit",
                    details={"entry": info.filename, "limit": limits.max_compression_ratio},
                )
            if lowered.endswith(".xml") or lowered.endswith(".rels"):
                payload = archive.read(info)
                if b"\x00" in payload:
                    raise FfPoolXlsxError(
                        "unsupported_xml_encoding",
                        "OOXML XML parts must use UTF-8 encoding",
                    )
                try:
                    payload.decode("utf-8-sig")
                except UnicodeDecodeError as exc:
                    raise FfPoolXlsxError(
                        "unsupported_xml_encoding",
                        "OOXML XML parts must use UTF-8 encoding",
                    ) from exc
                declaration = re.match(
                    br"^(?:\xef\xbb\xbf)?\s*<\?xml[^>]*\bencoding=[\"']([^\"']+)[\"']",
                    payload,
                    re.I,
                )
                if declaration is not None and declaration.group(1).lower() not in {b"utf-8", b"utf8"}:
                    raise FfPoolXlsxError(
                        "unsupported_xml_encoding",
                        "OOXML XML parts must declare UTF-8 encoding",
                    )
                lowered_payload = payload.lower()
                if b"<!doctype" in lowered_payload or b"<!entity" in lowered_payload:
                    raise FfPoolXlsxError(
                        "unsafe_xml_forbidden",
                        "OOXML document type and entity declarations are forbidden",
                    )
                part_rows, part_cells = _inspect_xml_part(
                    payload,
                    worksheet=lowered.startswith("xl/worksheets/") and lowered.endswith(".xml"),
                    limits=limits,
                )
                worksheet_rows += part_rows
                worksheet_cells += part_cells
        if total_uncompressed > limits.max_uncompressed_bytes:
            raise FfPoolXlsxError(
                "zip_uncompressed_limit_exceeded",
                "XLSX uncompressed size exceeds the parser limit",
                details={"limit": limits.max_uncompressed_bytes, "actual": total_uncompressed},
            )
        if worksheet_rows > limits.max_rows or worksheet_cells > limits.max_cells:
            raise FfPoolXlsxError(
                "worksheet_size_limit_exceeded",
                "XLSX worksheet dimensions exceed parser limits",
                details={"rows": worksheet_rows, "cells": worksheet_cells},
            )
        if "xl/sharedStrings.xml" in names:
            payload = archive.read("xl/sharedStrings.xml")
            if len(payload) > limits.max_shared_strings_bytes:
                raise FfPoolXlsxError(
                    "shared_strings_limit_exceeded",
                    "XLSX shared strings exceed the parser limit",
                )


def _inspect_xml_part(
    payload: bytes,
    *,
    worksheet: bool,
    limits: XlsxParserLimits,
) -> tuple[int, int]:
    rows = 0
    cells = 0
    row_cells: int | None = None
    try:
        iterator = ET.iterparse(io.BytesIO(payload), events=("start", "end"))
        for event, element in iterator:
            local_name = str(element.tag).rsplit("}", 1)[-1]
            if event == "start":
                if local_name == "Relationship" and any(
                    str(key).rsplit("}", 1)[-1] == "TargetMode"
                    and str(value).strip().lower() == "external"
                    for key, value in element.attrib.items()
                ):
                    raise FfPoolXlsxError(
                        "external_links_forbidden",
                        "OOXML external relationships are forbidden",
                    )
                if worksheet:
                    if local_name == "f":
                        raise FfPoolXlsxError("formulas_forbidden", "Formulas are forbidden")
                    if local_name == "row":
                        if row_cells is not None:
                            raise FfPoolXlsxError(
                                "malformed_ooxml",
                                "OOXML worksheet contains nested rows",
                            )
                        rows += 1
                        row_cells = 0
                    elif local_name == "c":
                        if row_cells is None:
                            raise FfPoolXlsxError(
                                "malformed_ooxml",
                                "OOXML worksheet cell is outside a row",
                            )
                        cells += 1
                        row_cells += 1
                        if row_cells > limits.max_columns:
                            raise FfPoolXlsxError(
                                "worksheet_dimension_limit_exceeded",
                                "XLSX row exceeds the allowed bounded column profile",
                                details={"max_columns": limits.max_columns},
                            )
                        _assert_cell_reference(str(element.attrib.get("r") or ""), limits=limits)
                    elif local_name == "dimension":
                        for token in str(element.attrib.get("ref") or "").split(":"):
                            _assert_cell_reference(token, limits=limits)
            elif local_name == "t":
                text_bytes = str(element.text or "").encode("utf-8")
                if len(text_bytes) > limits.max_cell_text_bytes:
                    raise FfPoolXlsxError(
                        "cell_text_limit_exceeded",
                        "One XLSX text value exceeds the parser limit",
                    )
            if event == "end" and worksheet and local_name == "row":
                row_cells = None
            if event == "end":
                element.clear()
    except FfPoolXlsxError:
        raise
    except ET.ParseError as exc:
        raise FfPoolXlsxError("malformed_ooxml", "OOXML XML part is malformed") from exc
    return rows, cells


def _assert_cell_reference(token: str, *, limits: XlsxParserLimits) -> None:
    match = re.fullmatch(r"([A-Z]+)([0-9]+)", str(token or ""))
    if match is None:
        return
    column = 0
    for char in match.group(1):
        column = column * 26 + (ord(char) - 64)
    row = int(match.group(2))
    if row > limits.max_rows or column > limits.max_columns:
        raise FfPoolXlsxError(
            "worksheet_dimension_limit_exceeded",
            "XLSX cell lies outside the allowed bounded profile",
            details={"cell": token, "max_rows": limits.max_rows, "max_columns": limits.max_columns},
        )


def _validate_envelope(*, filename: str, content_type: str) -> None:
    normalized_filename = str(filename or "").strip()
    if not normalized_filename.lower().endswith(".xlsx"):
        raise FfPoolXlsxError("invalid_extension", "Stage 2 accepts only .xlsx files")
    if str(content_type or "").split(";", 1)[0].strip().lower() not in ALLOWED_CONTENT_TYPES:
        raise FfPoolXlsxError(
            "invalid_mime_type",
            "Stage 2 XLSX MIME type is not allowed",
            details={"content_type": str(content_type or "")},
        )


def _new_workbook() -> Any:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # controlled runtime dependency boundary
        raise FfPoolXlsxError(
            "xlsx_dependency_unavailable",
            "openpyxl is required for Stage 2 XLSX generation",
        ) from exc
    return Workbook()


def _load_workbook(source_bytes: bytes) -> Any:
    try:
        from openpyxl import load_workbook
        return load_workbook(
            io.BytesIO(source_bytes),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except FfPoolXlsxError:
        raise
    except Exception as exc:
        raise FfPoolXlsxError("malformed_ooxml", "openpyxl could not parse the workbook") from exc


def _save_workbook(workbook: Any) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _format_sheet(sheet: Any, *, header_row: int, max_row: int, max_column: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet["A1"].font = Font(name="Aptos Display", size=15, bold=True, color="1F2937")
    sheet["A1"].alignment = Alignment(vertical="center")
    for cell in sheet[header_row]:
        if cell.column > max_column:
            break
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [16, 22, 36, 18, 18, 24, 16, 16, 18, 34]
    for column in range(1, max_column + 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = widths[column - 1]
    for row in range(header_row + 1, max_row + 1):
        sheet.cell(row, 3).alignment = Alignment(horizontal="left")
    sheet.auto_filter.ref = f"A{header_row}:{sheet.cell(max_row, max_column).coordinate}"


def _add_table(sheet: Any, *, name: str, ref: str) -> None:
    from openpyxl.worksheet.table import Table, TableStyleInfo

    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _add_lists_and_validation(
    workbook: Any,
    *,
    sheet: Any,
    active: list[dict[str, str]],
    facility_cell: str,
    scope_cell: str = "",
) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    lists = workbook.create_sheet(LISTS_SHEET)
    lists.append(["facility_id", "facility_label", "scope"])
    for index, item in enumerate(active, start=2):
        _set_text(lists.cell(index, 1), item["facility_id"])
        _set_text(lists.cell(index, 2), item["label"])
    for index, value in enumerate(("FBS", "FBO", "both"), start=2):
        _set_text(lists.cell(index, 3), value)
    lists.sheet_state = "veryHidden"
    facility_validation = DataValidation(
        type="list",
        formula1=f"'{LISTS_SHEET}'!$B$2:$B${len(active) + 1}",
        allow_blank=False,
    )
    facility_validation.error = "Выберите active facility из списка"
    facility_validation.errorTitle = "Неверная facility"
    facility_validation.showErrorMessage = True
    sheet.add_data_validation(facility_validation)
    facility_validation.add(sheet[facility_cell])
    if scope_cell:
        scope_validation = DataValidation(
            type="list",
            formula1=f"'{LISTS_SHEET}'!$C$2:$C$4",
            allow_blank=False,
        )
        scope_validation.showErrorMessage = True
        sheet.add_data_validation(scope_validation)
        scope_validation.add(sheet[scope_cell])


def _add_contract_sheet(
    workbook: Any,
    *,
    profile: str,
    fingerprint: str,
    source_revision: str,
    selected_facility_id: str,
    scope: str,
) -> None:
    sheet = workbook.create_sheet(CONTRACT_SHEET)
    rows = [
        ("contract_name", CONTRACT_NAME),
        ("profile", profile),
        ("template_fingerprint", fingerprint),
        ("source_revision", source_revision),
        ("selected_facility_id", selected_facility_id),
        ("scope", scope),
    ]
    for row_index, (key, value) in enumerate(rows, start=1):
        _set_text(sheet.cell(row_index, 1), key)
        _set_text(sheet.cell(row_index, 2), value)
    sheet.sheet_state = "veryHidden"


def _read_contract(workbook: Any) -> dict[str, str]:
    sheet = workbook[CONTRACT_SHEET]
    result: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = str(row[0] or "").strip()
        if key:
            result[key] = str(row[1] or "").strip()
    return result


def _assert_contract(
    metadata: Mapping[str, str],
    *,
    profile: str,
    expected_fingerprint: str,
    source_revision: str,
) -> None:
    expected = {
        "contract_name": CONTRACT_NAME,
        "profile": profile,
        "template_fingerprint": expected_fingerprint,
        "source_revision": str(source_revision),
    }
    mismatches = {
        key: {"expected": value, "actual": str(metadata.get(key) or "")}
        for key, value in expected.items()
        if str(metadata.get(key) or "") != value
    }
    if mismatches:
        raise FfPoolXlsxError(
            "template_fingerprint_mismatch",
            "Workbook template/source fingerprint is stale or modified",
            details=mismatches,
        )


def _assert_allowed_sheets(workbook: Any, allowed: set[str]) -> None:
    actual = set(workbook.sheetnames)
    if actual != allowed:
        raise FfPoolXlsxError(
            "invalid_sheet_profile",
            "Workbook sheets do not match the exact Stage 2 profile",
            details={"expected": sorted(allowed), "actual": sorted(actual)},
        )


def _assert_headers(sheet: Any, *, row: int, expected: tuple[str, ...]) -> None:
    actual = tuple(str(sheet.cell(row, index).value or "").strip() for index in range(1, len(expected) + 1))
    if actual != expected:
        raise FfPoolXlsxError(
            "invalid_headers",
            "Workbook headers do not match the exact Stage 2 profile",
            details={"expected": expected, "actual": actual},
        )


def _active_facilities(facilities: Iterable[Mapping[str, Any]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in facilities:
        if not bool(item.get("active")):
            continue
        facility_id = str(item.get("facility_id") or "").strip()
        if not facility_id or facility_id in seen:
            continue
        seen.add(facility_id)
        result.append(
            {
                "facility_id": facility_id,
                "code": str(item.get("code") or facility_id).strip(),
                "name": str(item.get("name") or facility_id).strip(),
            }
        )
    result = sorted(result, key=lambda item: (item["code"], item["facility_id"]))
    return [{**item, "label": f"{item['code']} — {item['name']}"} for item in result]


def _facility_label(active: Iterable[Mapping[str, str]], facility_id: str) -> str:
    for item in active:
        if str(item.get("facility_id") or "") == facility_id:
            return str(item.get("label") or "")
    return ""


def _facility_id_from_label(active: Iterable[Mapping[str, str]], label: str) -> str:
    matches = [
        str(item.get("facility_id") or "")
        for item in active
        if str(item.get("label") or "") == label
    ]
    return matches[0] if len(matches) == 1 else ""


def _normalize_source_lines(lines: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[int] = set()
    for position, item in enumerate(lines, start=1):
        nm_id = _positive_int(item.get("nm_id"), field="nm_id")
        if nm_id in seen:
            raise FfPoolXlsxError(
                "duplicate_source_nm_id",
                "Source composition contains duplicate nmId",
                details={"nm_id": nm_id},
            )
        seen.add(nm_id)
        quantity = _positive_int(
            item.get("accepted_quantity", item.get("quantity")),
            field="accepted_quantity",
        )
        capital = _positive_decimal(
            item.get("accepted_capital_rub", item.get("capital_rub")),
            field="accepted_capital_rub",
        )
        barcodes = _barcodes(item)
        if not barcodes:
            raise FfPoolXlsxError(
                "exact_identity_evidence_missing",
                "Source SKU requires a canonical server-owned barcode",
                details={"nm_id": nm_id},
            )
        identity = {
            "nm_id": nm_id,
            "barcodes": barcodes,
            "source_identity_revision": str(item.get("identity_revision") or ""),
        }
        result.append(
            {
                "line_no": position,
                "nm_id": nm_id,
                "barcode": barcodes[0],
                "barcodes": barcodes,
                "sku": str(item.get("sku") or item.get("our_sku") or nm_id),
                "quantity": quantity,
                "capital_rub": _decimal_text(capital),
                "identity_evidence_digest": _fingerprint(identity),
            }
        )
    if not result:
        raise FfPoolXlsxError("empty_source_composition", "Source composition is empty")
    return result


def _normalize_catalog(catalog: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[int] = set()
    for item in catalog:
        if item.get("active") is False or bool(item.get("hidden")):
            continue
        nm_id = _positive_int(item.get("nm_id"), field="nm_id")
        if nm_id in seen:
            raise FfPoolXlsxError(
                "duplicate_catalog_nm_id",
                "Catalog contains duplicate active nmId",
                details={"nm_id": nm_id},
            )
        seen.add(nm_id)
        barcodes = _barcodes(item)
        if not barcodes:
            raise FfPoolXlsxError(
                "exact_identity_evidence_missing",
                "Catalog SKU requires a canonical server-owned barcode",
                details={"nm_id": nm_id},
            )
        rows.append(
            {
                "nm_id": nm_id,
                "barcode": barcodes[0],
                "barcodes": barcodes,
                "sku": str(item.get("sku") or item.get("our_sku") or nm_id),
                "identity_evidence_digest": _fingerprint(
                    {
                        "nm_id": nm_id,
                        "barcodes": barcodes,
                        "source_identity_revision": str(item.get("identity_revision") or ""),
                    }
                ),
            }
        )
    if not rows:
        raise FfPoolXlsxError("empty_catalog", "Selected inventory catalog is empty")
    return sorted(rows, key=lambda item: int(item["nm_id"]))


def _resolve_identity_cells(
    nm_cell: Any,
    barcode_cell: Any,
    *,
    by_nm: Mapping[int, Mapping[str, Any]],
    by_barcode: Mapping[str, list[Mapping[str, Any]]],
) -> Mapping[str, Any]:
    nm_id: int | None = None
    if nm_cell.value not in (None, ""):
        nm_id = _whole_cell(nm_cell, field="nmId", positive=True)
    barcode = ""
    if barcode_cell.value not in (None, ""):
        if barcode_cell.data_type != "s" or not isinstance(barcode_cell.value, str):
            raise FfPoolXlsxError(
                "barcode_must_be_text",
                "Barcode must be stored as an exact text cell",
            )
        barcode = str(barcode_cell.value).strip()
        if not barcode or re.search(r"[eE][+-]?[0-9]+$", barcode) or re.fullmatch(r"[0-9]+\.[0-9]+", barcode):
            raise FfPoolXlsxError(
                "unsafe_barcode_representation",
                "Scientific or fractional barcode representation is forbidden",
            )
    if nm_id is None and not barcode:
        raise FfPoolXlsxError("identity_missing", "nmId or exact barcode is required")
    nm_item = by_nm.get(nm_id) if nm_id is not None else None
    barcode_candidates = list(by_barcode.get(barcode, [])) if barcode else []
    if nm_id is not None and nm_item is None:
        raise FfPoolXlsxError("unknown_nm_id", "nmId is not in the exact source scope", details={"nm_id": nm_id})
    if barcode and not barcode_candidates:
        raise FfPoolXlsxError("unknown_barcode", "Barcode is not in the exact source scope", details={"barcode": barcode})
    if len(barcode_candidates) > 1:
        raise FfPoolXlsxError(
            "ambiguous_barcode",
            "Barcode resolves to more than one SKU",
            details={"barcode": barcode, "nm_ids": sorted(int(item["nm_id"]) for item in barcode_candidates)},
        )
    barcode_item = barcode_candidates[0] if barcode_candidates else None
    if nm_item is not None and barcode_item is not None and int(nm_item["nm_id"]) != int(barcode_item["nm_id"]):
        raise FfPoolXlsxError(
            "conflicting_identity",
            "nmId and barcode resolve to different SKUs",
            details={"nm_id": nm_id, "barcode": barcode},
        )
    return nm_item or barcode_item  # type: ignore[return-value]


def _barcode_index(rows: Iterable[Mapping[str, Any]]) -> dict[str, list[Mapping[str, Any]]]:
    result: dict[str, list[Mapping[str, Any]]] = {}
    for item in rows:
        for barcode in item.get("barcodes") or [item.get("barcode")]:
            token = str(barcode or "").strip()
            if token:
                result.setdefault(token, []).append(item)
    return result


def _barcodes(item: Mapping[str, Any]) -> list[str]:
    candidates: list[Any] = [item.get("barcode")]
    extra = item.get("barcodes") or item.get("barcodes_json") or []
    if isinstance(extra, str):
        try:
            decoded = json.loads(extra)
            extra = decoded if isinstance(decoded, list) else [extra]
        except json.JSONDecodeError:
            extra = [extra]
    if isinstance(extra, Iterable) and not isinstance(extra, (str, bytes, Mapping)):
        candidates.extend(extra)
    result: list[str] = []
    for value in candidates:
        token = str(value or "").strip()
        if not token:
            continue
        if re.search(r"[eE][+-]?[0-9]+$", token) or re.fullmatch(r"[0-9]+\.[0-9]+", token):
            raise FfPoolXlsxError(
                "unsafe_barcode_evidence",
                "Canonical barcode evidence must be exact text",
                details={"barcode": token},
            )
        if token not in result:
            result.append(token)
    return result


def _whole_cell(cell: Any, *, field: str, positive: bool = False) -> int:
    if cell.data_type == "f":
        raise FfPoolXlsxError("formulas_forbidden", f"{field} cannot be a formula")
    value = cell.value
    if isinstance(value, bool) or value in (None, ""):
        raise FfPoolXlsxError("invalid_integer", f"{field} must be an exact integer")
    if isinstance(value, int):
        result = value
    elif isinstance(value, Decimal) and value == value.to_integral_value():
        result = int(value)
    elif isinstance(value, str) and re.fullmatch(r"(?:0|[1-9][0-9]*)", value.strip()):
        token = value.strip()
        if len(token) > 19:
            raise FfPoolXlsxError("invalid_integer", f"{field} exceeds the exact integer range")
        result = int(token)
    else:
        raise FfPoolXlsxError(
            "invalid_integer",
            f"{field} must be a non-scientific whole number",
            details={"value": str(value)},
        )
    if result > MAX_SQLITE_INTEGER or result < 0 or positive and result <= 0:
        raise FfPoolXlsxError("invalid_integer", f"{field} is outside the allowed range")
    return result


def _positive_int(value: Any, *, field: str) -> int:
    result = _exact_nonnegative_int(value, field=field)
    if result <= 0:
        raise FfPoolXlsxError("invalid_integer", f"{field} must be positive")
    return result


def _exact_nonnegative_int(value: Any, *, field: str) -> int:
    if isinstance(value, bool):
        raise FfPoolXlsxError("invalid_integer", f"{field} must be an exact integer")
    if isinstance(value, int):
        result = value
    else:
        token = str(value if value is not None else "").strip()
        if len(token) > 19 or not re.fullmatch(r"(?:0|[1-9][0-9]*)", token):
            raise FfPoolXlsxError("invalid_integer", f"{field} must be an exact integer")
        result = int(token)
    if result < 0 or result > MAX_SQLITE_INTEGER:
        raise FfPoolXlsxError("invalid_integer", f"{field} must be non-negative")
    return result


def _positive_decimal(value: Any, *, field: str) -> Decimal:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise FfPoolXlsxError("invalid_decimal", f"{field} must be Decimal-safe") from exc
    if not amount.is_finite() or amount <= 0:
        raise FfPoolXlsxError("invalid_decimal", f"{field} must be positive")
    return amount


def _decimal_text(value: Any) -> str:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise FfPoolXlsxError("invalid_decimal", "Value must be Decimal-safe") from exc
    if not amount.is_finite():
        raise FfPoolXlsxError("invalid_decimal", "Value must be finite")
    return "0" if amount == 0 else format(amount, "f")


def _scope(value: str) -> str:
    normalized = str(value or "").strip()
    if normalized not in {"FBS", "FBO", "both"}:
        raise FfPoolXlsxError("invalid_pool_scope", "Pool scope must be FBS, FBO or both")
    return normalized


def _set_text(cell: Any, value: Any) -> None:
    cell.value = str(value if value is not None else "")
    cell.data_type = "s"


def _selected_pools(scope: str) -> tuple[str, ...]:
    return ("FBS", "FBO") if scope == "both" else (scope,)


def _template_fingerprint(
    *,
    profile: str,
    facilities: Iterable[Mapping[str, Any]],
    source_revision: str,
    rows: Iterable[Mapping[str, Any]],
    facility_id: str = "",
    scope: str = "",
) -> str:
    payload = {
        "contract_name": CONTRACT_NAME,
        "profile": profile,
        "facilities": [
            {key: str(item.get(key) or "") for key in ("facility_id", "code", "name")}
            for item in facilities
        ],
        "source_revision": str(source_revision),
        "facility_id": str(facility_id),
        "scope": str(scope),
        "rows": [
            {
                "nm_id": int(item["nm_id"]),
                "barcode": str(item.get("barcode") or ""),
                "quantity": int(item.get("quantity") or 0),
                "capital_rub": str(item.get("capital_rub") or "0"),
                "identity_evidence_digest": str(item.get("identity_evidence_digest") or ""),
            }
            for item in rows
        ],
        "limits": asdict(DEFAULT_LIMITS),
    }
    return _fingerprint(payload)


def _fingerprint(value: Any) -> str:
    payload = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return "sha256:" + hashlib.sha256(payload).hexdigest()


def _sha256(value: bytes) -> str:
    return "sha256:" + hashlib.sha256(value).hexdigest()
