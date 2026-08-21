"""Server-owned, UI-first single-SKU partner profitability report."""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_DOWN, ROUND_HALF_UP
import hashlib
from io import BytesIO
import json
from pathlib import Path
import re
import sqlite3
import time
from typing import Any, Callable, Iterable, Mapping
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from packages.application.wb_finance_weekly import (
    COST_METHOD_VERSION,
    PROFIT_METHOD_VERSION,
    SKU_AGGREGATE_FORMULA_VERSION,
    WbFinanceWeeklyBlock,
    _decimal,
    _money_text,
    _nomenclature_identity_index,
    _operation_date,
    _resolve_finance_nm_id,
    classify_deduction,
)
from packages.application.ads_snapshot_payload import resolve_ads_snapshot_payload
from packages.application.canonical_wb_cost_resolver import (
    resolve_channel_location_cost,
)
from packages.application.storage_registry import StoreRegistry


PARTNER_REPORT_FORMULA_VERSION = "partner_report_profitability_ui_first_v4"
PARTNER_REPORT_SCHEMA_VERSION = "partner_report_v4"
COMMON_EXPENSE_RULE = "net_revenue_share"
ADS_SOURCE_ROLE = "accepted_closed_day_snapshot"
ADS_SOURCE_KEY = "ads_compact"
ZERO = Decimal("0")
HUNDRED = Decimal("100")
MONEY_QUANT = Decimal("0.0001")
PERCENT_QUANT = Decimal("0.0001")
DISPLAY_MONEY_QUANT = Decimal("0.01")
OTHER_DIRECT_ALLOCATED_KEY = "other_direct_and_allocated_expenses"
OTHER_DIRECT_ALLOCATED_LABEL = "Прочие прямые и распределённые расходы"
OTHER_DIRECT_ALLOCATED_TOOLTIP = (
    "Итог включает только прямые и распределённые расходы, для которых нет "
    "отдельной основной строки. Расходы кабинета распределяются по доле чистой "
    "выручки SKU; маркетинг Finance исключён, потому что строка «Маркетинг WB» "
    "использует accepted ads_compact."
)
OTHER_EXPENSE_CATEGORIES: tuple[tuple[str, str], ...] = (
    (
        "uncapitalized_transit_logistics",
        "Транзитная логистика, не подтверждённая как капитализированная",
    ),
    ("wb_jam_subscription", "Подписка WB Jam"),
    ("wb_paid_services", "Платные сервисы WB"),
    ("review_points", "Баллы за отзывы"),
    ("other_withholdings", "Прочие удержания"),
)
ALLOCATED_MAIN_EXPENSE_KEYS: tuple[str, ...] = (
    "agent_remuneration",
    "acquiring",
    "logistics",
    "storage",
    "acceptance",
    "penalties_and_adjustments",
)
FINANCE_EXPORT_COLUMNS = (
    ("reportId", "report_id"),
    ("reportType", "report_type"),
    ("rrdId", "rrd_id"),
    ("rrDate", "operation_date"),
    ("docTypeName", "operation_type"),
    ("sellerOperName", "seller_operation"),
    ("nmId", "nm_id"),
    ("vendorCode", "vendor_code"),
    ("sku", "barcode"),
    ("quantity", "quantity"),
    ("retailPriceWithDisc", "revenue"),
    ("forPay", "for_pay"),
    ("deliveryService", "logistics"),
    ("paidStorage", "storage"),
    ("paidAcceptance", "paid_acceptance"),
    ("penalty", "penalty"),
    ("deduction", "deduction"),
    ("additionalPayment", "additional_payment"),
    ("acquiringFee", "acquiring_disclosure"),
    ("bonusTypeName", "deduction_name"),
)
REPORT_ROWS = (
    ("net_revenue", "Чистая выручка"),
    ("sales_without_cost_rub", "Продажи без себестоимости, ₽"),
    ("orders_without_cost", "Заказы без себестоимости, шт."),
    ("units_without_cost", "Единицы без себестоимости, шт."),
    ("cogs", "Себестоимость наша"),
    ("agent_remuneration", "Агентское вознаграждение WB"),
    ("acquiring", "Эквайринг"),
    ("logistics", "Логистика WB"),
    ("storage", "Хранение WB"),
    ("acceptance", "Платная приёмка"),
    ("ads", "Маркетинг WB"),
    ("penalties_and_adjustments", "Штрафы/корректировки"),
    (OTHER_DIRECT_ALLOCATED_KEY, OTHER_DIRECT_ALLOCATED_LABEL),
    ("finance_margin", "Маржа по финотчёту"),
    ("office", "Офисные расходы"),
    ("estimated_tax", "Налог"),
    ("replenishment_reserve", "На пополнение товарных остатков"),
    ("net_profit", "Чистая прибыль"),
    ("dividends", "Дивиденды"),
    ("annualized_return_pct", "Расчётная годовая доходность инвестора, %"),
)


class PartnerReportError(ValueError):
    def __init__(
        self,
        message: str,
        *,
        code: str = "partner_report_invalid",
        blockers: list[dict[str, Any]] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.blockers = blockers or []


def _strict_decimal(value: Any, *, field: str) -> Decimal:
    if value in (None, ""):
        raise PartnerReportError(f"{field} is required", code="settings_required")
    try:
        result = Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise PartnerReportError(f"{field} must be decimal", code="settings_invalid") from exc
    if not result.is_finite():
        raise PartnerReportError(f"{field} must be finite", code="settings_invalid")
    return result


def _decimal_text(value: Decimal | None, quant: Decimal = MONEY_QUANT) -> str | None:
    return None if value is None else format(value.quantize(quant), "f")


def _display_money(value: Decimal) -> Decimal:
    return value.quantize(DISPLAY_MONEY_QUANT, rounding=ROUND_HALF_UP)


def _display_breakdown(
    exact: Mapping[str, Decimal],
    *,
    target_total: Decimal | None = None,
) -> list[dict[str, str]]:
    """Round visible rows so their displayed sum equals the main row.

    Exact values remain in calculation/provenance.  Display kopecks use a
    deterministic largest-remainder apportionment; ``Прочие удержания`` wins
    exact ties, without changing the unrounded profit calculation.
    """

    labels = dict(OTHER_EXPENSE_CATEGORIES)
    keys = [key for key, _label in OTHER_EXPENSE_CATEGORIES]
    scaled = {key: exact.get(key, ZERO) * Decimal("100") for key in keys}
    cents = {
        key: int(value.to_integral_value(rounding=ROUND_DOWN))
        for key, value in scaled.items()
    }
    # The public main row is first serialized at MONEY_QUANT and only then
    # formatted to kopecks by UI/XLSX.  Reconcile against that same two-stage
    # representation rather than rounding the higher-precision exact sum via a
    # different path.
    serialized_total = (
        target_total
        if target_total is not None
        else sum((exact.get(key, ZERO) for key in keys), ZERO)
    ).quantize(MONEY_QUANT)
    target_cents = int(
        (_display_money(serialized_total) * Decimal("100")).to_integral_value()
    )
    residual = target_cents - sum(cents.values())
    tie_order = {key: index + 1 for index, key in enumerate(keys)}
    tie_order["other_withholdings"] = 0
    while residual:
        direction = 1 if residual > 0 else -1
        candidates = sorted(
            keys,
            key=lambda key: (
                -(scaled[key] - Decimal(cents[key])) * direction,
                tie_order[key],
            ),
        )
        cents[candidates[0]] += direction
        residual -= direction
    shown = {key: Decimal(cents[key]) / Decimal("100") for key in keys}
    return [
        {
            "key": key,
            "label": labels[key],
            "amount_rub": format(shown[key], "f"),
        }
        for key in keys
    ]


def _breakdown_map(rows: Iterable[Mapping[str, Any]]) -> dict[str, Decimal]:
    return {
        str(item.get("key") or ""): _decimal(item.get("amount_rub"))
        for item in rows
        if str(item.get("key") or "")
    }


def _sha256_json(value: Any) -> str:
    return "sha256:" + hashlib.sha256(
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()


def _now_iso(factory: Callable[[], datetime]) -> str:
    return factory().astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


class PartnerReportBlock:
    def __init__(
        self,
        runtime_dir: Path,
        *,
        seller_id: str = "canonical",
        now_factory: Callable[[], datetime] | None = None,
    ) -> None:
        self.runtime_dir = Path(runtime_dir)
        self.store_registry = StoreRegistry(self.runtime_dir)
        self.db_path = self.store_registry.resolve("operational")
        self.seller_id = seller_id or "canonical"
        self.now_factory = now_factory or (lambda: datetime.now(timezone.utc))
        self.finance = WbFinanceWeeklyBlock(
            self.runtime_dir,
            seller_id=self.seller_id,
            now_factory=self.now_factory,
        )

    def ensure_schema(self) -> None:
        self.finance.ensure_schema()
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS partner_report_settings_versions (
                    settings_version_id TEXT PRIMARY KEY,
                    seller_id TEXT NOT NULL,
                    nm_id TEXT NOT NULL,
                    product_name TEXT NOT NULL,
                    parameters_json TEXT NOT NULL,
                    fingerprint TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    created_by TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS partner_report_settings_versions_by_sku
                ON partner_report_settings_versions(seller_id,nm_id,created_at DESC);
                CREATE TABLE IF NOT EXISTS partner_report_settings_current (
                    seller_id TEXT NOT NULL,
                    nm_id TEXT NOT NULL,
                    settings_version_id TEXT NOT NULL REFERENCES partner_report_settings_versions(settings_version_id),
                    PRIMARY KEY(seller_id,nm_id)
                );
                CREATE TABLE IF NOT EXISTS partner_report_finalized_reports (
                    report_id TEXT PRIMARY KEY,
                    seller_id TEXT NOT NULL,
                    nm_id TEXT NOT NULL,
                    product_name TEXT NOT NULL,
                    settings_version_id TEXT NOT NULL,
                    formula_version TEXT NOT NULL,
                    selected_weeks_json TEXT NOT NULL,
                    report_json TEXT NOT NULL,
                    provenance_json TEXT NOT NULL,
                    source_digest TEXT NOT NULL,
                    loss_carry_in_rub TEXT NOT NULL,
                    loss_carry_out_rub TEXT NOT NULL,
                    total_partner_payout_rub TEXT NOT NULL,
                    period_roi_pct TEXT NOT NULL,
                    annualized_return_pct TEXT NOT NULL,
                    finalized_at TEXT NOT NULL,
                    finalized_by TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS partner_report_finalized_by_sku
                ON partner_report_finalized_reports(seller_id,nm_id,finalized_at DESC);
                CREATE TABLE IF NOT EXISTS partner_report_audit (
                    audit_id TEXT PRIMARY KEY,
                    seller_id TEXT NOT NULL,
                    action TEXT NOT NULL,
                    object_id TEXT NOT NULL,
                    actor TEXT NOT NULL,
                    payload_digest TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );
                """
            )
            conn.commit()

    def options(self) -> dict[str, Any]:
        self.ensure_schema()
        with self._connect() as conn:
            items = self._nomenclature_items(conn)
            settings_rows = conn.execute(
                """SELECT current.nm_id,versions.settings_version_id,
                          versions.parameters_json,versions.fingerprint,versions.created_at
                   FROM partner_report_settings_current AS current
                   JOIN partner_report_settings_versions AS versions
                     ON versions.settings_version_id=current.settings_version_id
                   WHERE current.seller_id=? ORDER BY current.nm_id""",
                (self.seller_id,),
            ).fetchall()
            week_rows = conn.execute(
                """SELECT week_start,week_end,status,report_count,raw_row_count
                   FROM wb_finance_weekly_sync WHERE seller_id=? ORDER BY week_start""",
                (self.seller_id,),
            ).fetchall()
        settings = {
            str(row["nm_id"]): {
                "settings_version_id": row["settings_version_id"],
                "parameters": json.loads(row["parameters_json"]),
                "fingerprint": row["fingerprint"],
                "created_at": row["created_at"],
            }
            for row in settings_rows
        }
        return {
            "status": "ok",
            "contract_version": PARTNER_REPORT_SCHEMA_VERSION,
            "formula_version": PARTNER_REPORT_FORMULA_VERSION,
            "common_expense_rules": [
                {
                    "value": COMMON_EXPENSE_RULE,
                    "label": "По доле чистой выручки карточки в чистой выручке недели",
                }
            ],
            "cards": [
                {
                    "nm_id": item["nm_id"],
                    "product_name": item["product_name"],
                    "vendor_code": item["vendor_code"],
                    "barcode": item["barcode"],
                    "is_active": item["is_active"],
                    "is_hidden": item["is_hidden"],
                    "settings": settings.get(item["nm_id"]),
                }
                for item in items
            ],
            "weeks": [dict(row) for row in week_rows],
            "export_contract": "UI preview first; source-digest-bound XLSX only",
        }

    def save_settings(self, payload: Mapping[str, Any], *, actor: str) -> dict[str, Any]:
        self.ensure_schema()
        parameters = self._validate_parameters(payload)
        nm_id = parameters["nm_id"]
        with self._connect() as conn:
            product = self._nomenclature_product(conn, nm_id)
            if product is None:
                raise PartnerReportError(
                    f"nmId {nm_id} is not present in canonical nomenclature",
                    code="nomenclature_missing",
                )
            now = _now_iso(self.now_factory)
            version_payload = {
                "seller_id": self.seller_id,
                "nm_id": nm_id,
                "product_name": product["product_name"],
                "parameters": parameters,
                "formula_version": PARTNER_REPORT_FORMULA_VERSION,
            }
            fingerprint = _sha256_json(version_payload)
            current = conn.execute(
                """SELECT versions.* FROM partner_report_settings_current AS current
                   JOIN partner_report_settings_versions AS versions
                     ON versions.settings_version_id=current.settings_version_id
                   WHERE current.seller_id=? AND current.nm_id=?""",
                (self.seller_id, nm_id),
            ).fetchone()
            if current is not None and str(current["fingerprint"]) == fingerprint:
                return self._settings_payload(current)
            version_id = "prs_" + hashlib.sha256(
                (
                    f"{fingerprint}|{now}|{actor}|"
                    f"{str(current['settings_version_id']) if current is not None else ''}"
                ).encode("utf-8")
            ).hexdigest()[:24]
            conn.execute(
                """INSERT INTO partner_report_settings_versions(
                   settings_version_id,seller_id,nm_id,product_name,parameters_json,
                   fingerprint,created_at,created_by
                   ) VALUES(?,?,?,?,?,?,?,?)""",
                (
                    version_id,
                    self.seller_id,
                    nm_id,
                    product["product_name"],
                    json.dumps(parameters, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
                    fingerprint,
                    now,
                    actor,
                ),
            )
            conn.execute(
                """INSERT INTO partner_report_settings_current(seller_id,nm_id,settings_version_id)
                   VALUES(?,?,?) ON CONFLICT(seller_id,nm_id) DO UPDATE SET
                   settings_version_id=excluded.settings_version_id""",
                (self.seller_id, nm_id, version_id),
            )
            self._audit(
                conn,
                action="settings_saved",
                object_id=version_id,
                actor=actor,
                payload_digest=fingerprint,
                created_at=now,
            )
            conn.commit()
            row = conn.execute(
                "SELECT * FROM partner_report_settings_versions WHERE settings_version_id=?",
                (version_id,),
            ).fetchone()
            return self._settings_payload(row)

    def preview(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        started = time.perf_counter()
        self.ensure_schema()
        nm_id = str(payload.get("nm_id") or "").strip()
        weeks = self._validate_selected_weeks(
            payload.get("selected_weeks"), require_continuous=False
        )
        with self._connect() as conn:
            settings = self._load_settings(conn, nm_id=nm_id)
            report = self._calculate_report(
                conn,
                settings=settings,
                selected_weeks=weeks,
                finalization=False,
            )[0]
        return {
            **report,
            "performance": {
                "duration_ms": round((time.perf_counter() - started) * 1000, 3),
                "raw_finance_full_scan": False,
                "source": "wb_finance_weekly_sku_aggregates indexed lookup",
            },
        }

    def finalize(self, payload: Mapping[str, Any], *, actor: str) -> dict[str, Any]:
        raise PartnerReportError(
            "Partner Report finalization is outside the current UI-first Excel scope",
            code="partner_finalization_removed",
        )
        # Historical finalized rows remain readable, but new writes are not
        # exposed until a separate payout-finalization contract is approved.
        self.ensure_schema()
        nm_id = str(payload.get("nm_id") or "").strip()
        weeks = self._validate_selected_weeks(
            payload.get("selected_weeks"), require_continuous=True
        )
        with self._connect() as conn:
            conn.execute("BEGIN IMMEDIATE")
            try:
                settings = self._load_settings(conn, nm_id=nm_id)
                report, provenance = self._calculate_report(
                    conn,
                    settings=settings,
                    selected_weeks=weeks,
                    finalization=True,
                )
                if report["status"] != "ready":
                    raise PartnerReportError(
                        "Partner Report finalization is blocked by source coverage",
                        code="source_coverage_incomplete",
                        blockers=report["blockers"],
                    )
                selected_weeks_json = json.dumps(
                    weeks, ensure_ascii=False, separators=(",", ":")
                )
                existing = conn.execute(
                    """SELECT report_json FROM partner_report_finalized_reports
                       WHERE seller_id=? AND nm_id=? AND settings_version_id=?
                         AND formula_version=? AND selected_weeks_json=? AND source_digest=?
                       ORDER BY finalized_at,report_id LIMIT 1""",
                    (
                        self.seller_id,
                        nm_id,
                        settings["settings_version_id"],
                        PARTNER_REPORT_FORMULA_VERSION,
                        selected_weeks_json,
                        report["source_digest"],
                    ),
                ).fetchone()
                if existing is not None:
                    conn.rollback()
                    return json.loads(existing["report_json"])
                period_conflict = conn.execute(
                    """SELECT report_id FROM partner_report_finalized_reports
                       WHERE seller_id=? AND nm_id=? AND selected_weeks_json=?
                       ORDER BY finalized_at,report_id LIMIT 1""",
                    (self.seller_id, nm_id, selected_weeks_json),
                ).fetchone()
                if period_conflict is not None:
                    raise PartnerReportError(
                        "the selected payout period is already finalized with different immutable inputs",
                        code="finalized_period_already_exists",
                        blockers=[
                            {
                                "code": "finalized_period_already_exists",
                                "report_id": str(period_conflict["report_id"]),
                            }
                        ],
                    )
                now = _now_iso(self.now_factory)
                source_digest = str(report["source_digest"])
                report_id = "prf_" + hashlib.sha256(
                    f"{self.seller_id}|{nm_id}|{settings['settings_version_id']}|{source_digest}|{now}".encode(
                        "utf-8"
                    )
                ).hexdigest()[:28]
                report = {
                    **report,
                    "report_id": report_id,
                    "finalized": True,
                    "finalized_at": now,
                    "finalized_by": actor,
                }
                totals = report["totals"]
                conn.execute(
                    """INSERT INTO partner_report_finalized_reports(
                       report_id,seller_id,nm_id,product_name,settings_version_id,
                       formula_version,selected_weeks_json,report_json,provenance_json,
                       source_digest,loss_carry_in_rub,loss_carry_out_rub,
                       total_partner_payout_rub,period_roi_pct,annualized_return_pct,
                       finalized_at,finalized_by
                       ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        report_id,
                        self.seller_id,
                        nm_id,
                        report["product_name"],
                        settings["settings_version_id"],
                        PARTNER_REPORT_FORMULA_VERSION,
                        selected_weeks_json,
                        json.dumps(report, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
                        json.dumps(provenance, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
                        source_digest,
                        totals["loss_carry_in"],
                        totals["loss_carry_out"],
                        totals["partner_payout"],
                        totals["period_roi_pct"],
                        totals["annualized_return_pct"],
                        now,
                        actor,
                    ),
                )
                self._audit(
                    conn,
                    action="report_finalized",
                    object_id=report_id,
                    actor=actor,
                    payload_digest=source_digest,
                    created_at=now,
                )
                conn.commit()
                return report
            except Exception:
                conn.rollback()
                raise

    def list_finalized(self, *, nm_id: str = "") -> dict[str, Any]:
        self.ensure_schema()
        with self._connect() as conn:
            rows = conn.execute(
                """SELECT report_id,nm_id,product_name,settings_version_id,formula_version,
                          selected_weeks_json,source_digest,total_partner_payout_rub,
                          period_roi_pct,annualized_return_pct,finalized_at,finalized_by
                   FROM partner_report_finalized_reports
                   WHERE seller_id=? AND (?='' OR nm_id=?)
                   ORDER BY finalized_at DESC,report_id DESC""",
                (self.seller_id, nm_id, nm_id),
            ).fetchall()
        return {
            "status": "ok",
            "reports": [
                {
                    **dict(row),
                    "selected_weeks": json.loads(row["selected_weeks_json"]),
                }
                for row in rows
            ],
        }

    def finalized_report(self, report_id: str) -> dict[str, Any]:
        self.ensure_schema()
        with self._connect() as conn:
            row = conn.execute(
                """SELECT report_json FROM partner_report_finalized_reports
                   WHERE seller_id=? AND report_id=?""",
                (self.seller_id, report_id),
            ).fetchone()
        if row is None:
            raise PartnerReportError("finalized report not found", code="report_not_found")
        return json.loads(row["report_json"])

    def build_finalized_package(self, report_id: str) -> tuple[bytes, str, dict[str, Any]]:
        raise PartnerReportError(
            "Partner ZIP packages and raw Finance exports were removed from this report scope",
            code="partner_package_removed",
        )

    def build_preview_package(
        self, payload: Mapping[str, Any]
    ) -> tuple[bytes, str, dict[str, Any]]:
        raise PartnerReportError(
            "Partner ZIP packages and raw Finance exports were removed from this report scope",
            code="partner_package_removed",
        )

    def build_preview_workbook(
        self,
        payload: Mapping[str, Any],
        *,
        expected_source_digest: str = "",
    ) -> tuple[bytes, str, dict[str, Any]]:
        self.ensure_schema()
        nm_id = str(payload.get("nm_id") or "").strip()
        weeks = self._validate_selected_weeks(
            payload.get("selected_weeks"), require_continuous=False
        )
        with self._connect() as conn:
            settings = self._load_settings(conn, nm_id=nm_id)
            report, provenance = self._calculate_report(
                conn,
                settings=settings,
                selected_weeks=weeks,
                finalization=False,
            )
            if report["status"] != "ready":
                raise PartnerReportError(
                    "preview Excel is blocked by source coverage",
                    code="source_coverage_incomplete",
                    blockers=report["blockers"],
                )
            if expected_source_digest and expected_source_digest != str(report["source_digest"]):
                raise PartnerReportError(
                    "preview inputs changed before Excel export; rebuild the on-screen report",
                    code="preview_source_digest_changed",
                )
            workbook = self._build_main_workbook(report)
            first_week = str(report["selected_weeks"][0])
            last_week = str(report["selected_weeks"][-1])
            filename = (
                f"Партнёрский_отчёт_{self._safe_filename(report['product_name']) or nm_id}_"
                f"{nm_id}_{first_week}_{last_week}.xlsx"
            )
            return workbook, filename, {
                "source_digest": report["source_digest"],
                "formula_version": report["formula_version"],
                "nm_id": nm_id,
                "selected_weeks": report["selected_weeks"],
            }

    def _calculate_report_legacy_deprecated(
        self,
        conn: sqlite3.Connection,
        *,
        settings: Mapping[str, Any],
        selected_weeks: list[str],
        finalization: bool,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        nm_id = str(settings["nm_id"])
        params = settings["parameters"]
        week_records: list[dict[str, Any]] = []
        blockers: list[dict[str, Any]] = []
        provenance_weeks: list[dict[str, Any]] = []
        alias_to_nm, ambiguous_aliases, _groups, _items = _nomenclature_identity_index(conn)
        for week_start_text in selected_weeks:
            week_row = conn.execute(
                """SELECT week_start,week_end,status FROM wb_finance_weekly_sync
                   WHERE seller_id=? AND week_start=?""",
                (self.seller_id, week_start_text),
            ).fetchone()
            if week_row is None:
                blockers.append({"code": "finance_week_missing", "week_start": week_start_text})
                continue
            week_start = date.fromisoformat(week_start_text)
            week_end = date.fromisoformat(str(week_row["week_end"]))
            if finalization and str(week_row["status"]) != "completed":
                blockers.append(
                    {
                        "code": "finance_week_not_final",
                        "week_start": week_start_text,
                        "status": str(week_row["status"]),
                    }
                )
            raw_rows = conn.execute(
                """SELECT report_id,rrd_id,row_hash,raw_json
                   FROM wb_finance_weekly_raw_rows
                   WHERE seller_id=? AND week_start=? AND week_end=?
                   ORDER BY report_id,rrd_id""",
                (self.seller_id, week_start.isoformat(), week_end.isoformat()),
            ).fetchall()
            parsed = [json.loads(row["raw_json"]) for row in raw_rows]
            selected_rows: list[dict[str, Any]] = []
            account_rows: list[dict[str, Any]] = []
            total_net_revenue = ZERO
            selected_net_revenue = ZERO
            selected_components = self._empty_components()
            account_breakdown_internal: list[dict[str, Any]] = []
            identity_blockers: list[dict[str, Any]] = []
            for raw in parsed:
                resolved_nm, identity_method, identity_problem = _resolve_finance_nm_id(
                    raw,
                    alias_to_nm=alias_to_nm,
                    ambiguous_aliases=ambiguous_aliases,
                )
                doc = str(raw.get("docTypeName") or "").casefold()
                row_revenue = self._row_net_revenue(raw)
                if doc in {"продажа", "возврат"}:
                    total_net_revenue += row_revenue
                if resolved_nm == nm_id:
                    selected_rows.append(raw)
                    selected_net_revenue += row_revenue
                    self._accumulate_components(selected_components, raw, week_start)
                elif not resolved_nm and identity_method == "unresolved" and str(raw.get("nmId") or "").strip() in {"", "0"}:
                    account_rows.append(raw)
                elif not resolved_nm and identity_method == "ambiguous_alias":
                    identity_blockers.append(
                        {
                            "code": "finance_identity_ambiguous",
                            "week_start": week_start_text,
                            "report_id": str(raw.get("reportId") or ""),
                            "rrd_id": str(raw.get("rrdId") or ""),
                            "reason": identity_problem,
                        }
                    )
            blockers.extend(identity_blockers)
            allocated_common = ZERO
            allocation_ratio: Decimal | None = None
            account_expense_total = sum(
                (self._row_account_expense(row, week_start) for row in account_rows),
                ZERO,
            )
            if account_expense_total:
                if total_net_revenue <= ZERO:
                    blockers.append(
                        {
                            "code": "common_expense_zero_revenue_base",
                            "week_start": week_start_text,
                        }
                    )
                else:
                    allocation_ratio = selected_net_revenue / total_net_revenue
                    for raw in account_rows:
                        source_amount = self._row_account_expense(raw, week_start)
                        allocated = source_amount * allocation_ratio
                        if not allocated:
                            continue
                        category = classify_deduction(raw) if _decimal(raw.get("deduction")) else "account_level_expense"
                        account_breakdown_internal.append(
                            {
                                "category": category,
                                "source_amount_rub": _decimal_text(source_amount),
                                "allocation_coefficient": _decimal_text(allocation_ratio, Decimal("0.00000001")),
                                "allocated_amount_rub": _decimal_text(allocated),
                                "source_row_hash": _sha256_json(raw),
                            }
                        )
                        allocated_common += allocated
            cogs_coverage = self.finance._calculate_cogs(
                conn,
                selected_rows,
                week_start,
                include_details=True,
            )
            if int(cogs_coverage["unmatched_units"]):
                blockers.append(
                    {
                        "code": "partner_cost_coverage_incomplete",
                        "week_start": week_start_text,
                        "problem_skus": cogs_coverage["problem_skus"],
                    }
                )
            ads_value, ads_rows, ads_blockers = self._ads_for_week(
                conn,
                nm_id=nm_id,
                week_start=week_start,
                week_end=week_end,
            )
            blockers.extend(ads_blockers)
            cogs = (
                _decimal(cogs_coverage["cogs_rub"])
                if cogs_coverage["cogs_rub"] is not None
                else None
            )
            week_values = self._week_formulas(
                components=selected_components,
                cogs=cogs,
                ads=ads_value,
                other_expense_categories=self._other_expense_categories(
                    selected_components,
                    {"other_withholdings": allocated_common},
                ),
                params=params,
            )
            week_record = {
                "week_start": week_start.isoformat(),
                "week_end": week_end.isoformat(),
                "label": f"{week_start.strftime('%d.%m')}–{week_end.strftime('%d.%m')}",
                "values": week_values,
                "coverage": {
                    "finance_status": str(week_row["status"]),
                    "finance_raw_row_count": len(raw_rows),
                    "selected_finance_row_count": len(selected_rows),
                    "account_level_row_count": len(account_rows),
                    "cost": {
                        key: cogs_coverage[key]
                        for key in (
                            "matched_units",
                            "unmatched_units",
                            "coverage_pct",
                            "problem_skus",
                            "cost_state_hash",
                        )
                    },
                    "ads_date_count": 7,
                    "ads_covered_date_count": 7 - len(ads_blockers),
                },
            }
            week_records.append(week_record)
            safe_finance_rows = [
                self._safe_finance_row(row, selected_nm_id=nm_id) for row in selected_rows
            ]
            provenance_weeks.append(
                {
                    "week_start": week_start.isoformat(),
                    "week_end": week_end.isoformat(),
                    "finance_rows": safe_finance_rows,
                    "finance_source_digest": _sha256_json(
                        [
                            [raw["report_id"], raw["rrd_id"], raw["row_hash"]]
                            for raw in raw_rows
                        ]
                    ),
                    "ads_rows": ads_rows,
                    "ads_source_digest": _sha256_json(ads_rows),
                    "cost_rows": cogs_coverage["detail_rows"],
                    "cost_source_digest": cogs_coverage["cost_state_hash"],
                    "common_expense_internal": account_breakdown_internal,
                    "common_expense_safe": [
                        {
                            "category": item["category"],
                            "allocated_amount_rub": item["allocated_amount_rub"],
                            "rule": COMMON_EXPENSE_RULE,
                            "formula_version": PARTNER_REPORT_FORMULA_VERSION,
                            "source_digest": item["source_row_hash"],
                        }
                        for item in account_breakdown_internal
                    ],
                }
            )

        loss_carry_in, loss_carry_source, loss_blockers = self._loss_carry_context(
            conn,
            nm_id=nm_id,
            selected_weeks=selected_weeks,
            finalization=finalization,
        )
        blockers.extend(loss_blockers)
        totals = self._period_totals(
            week_records,
            params=params,
            loss_carry_in=loss_carry_in,
        )
        source_manifest = {
            "settings_fingerprint": settings["fingerprint"],
            "formula_version": PARTNER_REPORT_FORMULA_VERSION,
            "finance": [
                {
                    "week_start": item["week_start"],
                    "digest": item["finance_source_digest"],
                }
                for item in provenance_weeks
            ],
            "ads": [
                {"week_start": item["week_start"], "digest": item["ads_source_digest"]}
                for item in provenance_weeks
            ],
            "cost": [
                {"week_start": item["week_start"], "digest": item["cost_source_digest"]}
                for item in provenance_weeks
            ],
            "loss_carry": loss_carry_source,
        }
        source_digest = _sha256_json(source_manifest)
        report = {
            "status": "incomplete" if blockers else "ready",
            "contract_version": PARTNER_REPORT_SCHEMA_VERSION,
            "formula_version": PARTNER_REPORT_FORMULA_VERSION,
            "finance_cost_formula_version": COST_METHOD_VERSION,
            "finance_profit_formula_version": PROFIT_METHOD_VERSION,
            "retro_cost_formula_version": RETRO_COST_FORMULA_VERSION,
            "settings_version_id": settings["settings_version_id"],
            "settings_fingerprint": settings["fingerprint"],
            "nm_id": nm_id,
            "product_name": settings["product_name"],
            "parameters": params,
            "selected_weeks": selected_weeks,
            "weeks": week_records,
            "totals": totals,
            "source_coverage": {
                "complete": not blockers,
                "blocker_count": len(blockers),
            },
            "source_manifest": source_manifest,
            "source_digest": source_digest,
            "blockers": blockers,
            "finalized": False,
            "generated_at": _now_iso(self.now_factory),
            "loss_policy": (
                "continuous_finalized_period; negative selected weeks offset period payout; "
                "persisted prior loss carries into the immediately following period"
            ),
            "package_retention": "ephemeral_response_only",
        }
        provenance = {
            "schema_version": "partner_report_provenance_v1",
            "seller_id": self.seller_id,
            "nm_id": nm_id,
            "weeks": provenance_weeks,
            "source_manifest": source_manifest,
            "source_digest": source_digest,
            "loss_carry": loss_carry_source,
        }
        return report, provenance

    def _calculate_report(
        self,
        conn: sqlite3.Connection,
        *,
        settings: Mapping[str, Any],
        selected_weeks: list[str],
        finalization: bool,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        """Build a preview from indexed Finance projections, never a raw scan."""

        nm_id = str(settings["nm_id"])
        params = settings["parameters"]
        week_records: list[dict[str, Any]] = []
        blockers: list[dict[str, Any]] = []
        provenance_weeks: list[dict[str, Any]] = []
        period_other_expenses = {
            key: ZERO for key, _label in OTHER_EXPENSE_CATEGORIES
        }
        visible_other_expense_keys: set[str] = set()
        for week_start_text in selected_weeks:
            sync = conn.execute(
                """SELECT week_start,week_end,status,content_hash,raw_row_count
                   FROM wb_finance_weekly_sync WHERE seller_id=? AND week_start=?""",
                (self.seller_id, week_start_text),
            ).fetchone()
            if sync is None:
                blockers.append({"code": "finance_week_missing", "week_start": week_start_text})
                continue
            week_start = date.fromisoformat(week_start_text)
            week_end = date.fromisoformat(str(sync["week_end"]))
            if finalization and str(sync["status"]) != "completed":
                blockers.append(
                    {
                        "code": "finance_week_not_final",
                        "week_start": week_start_text,
                        "status": str(sync["status"]),
                    }
                )
            sku_row = conn.execute(
                """SELECT * FROM wb_finance_weekly_sku_aggregates
                   WHERE seller_id=? AND week_start=? AND week_end=? AND nm_id=?""",
                (self.seller_id, week_start_text, week_end.isoformat(), nm_id),
            ).fetchone()
            account_row = conn.execute(
                """SELECT * FROM wb_finance_weekly_sku_aggregates
                   WHERE seller_id=? AND week_start=? AND week_end=? AND nm_id='__account__'""",
                (self.seller_id, week_start_text, week_end.isoformat()),
            ).fetchone()
            if sku_row is None or account_row is None:
                blockers.append(
                    {
                        "code": "finance_sku_aggregate_missing",
                        "week_start": week_start_text,
                        "nm_id": nm_id,
                    }
                )
                continue
            if (
                str(sku_row["formula_version"]) != SKU_AGGREGATE_FORMULA_VERSION
                or str(account_row["formula_version"]) != SKU_AGGREGATE_FORMULA_VERSION
            ):
                blockers.append(
                    {
                        "code": "finance_sku_aggregate_formula_stale",
                        "week_start": week_start_text,
                        "nm_id": nm_id,
                    }
                )
            sync_hash = str(sync["content_hash"] or "")
            if (
                not sync_hash
                or str(sku_row["week_content_hash"]) != sync_hash
                or str(account_row["week_content_hash"]) != sync_hash
            ):
                blockers.append(
                    {
                        "code": "finance_sku_aggregate_raw_stale",
                        "week_start": week_start_text,
                        "nm_id": nm_id,
                    }
                )
            metrics = json.loads(str(sku_row["metrics_json"] or "{}"))
            coverage = json.loads(str(sku_row["coverage_json"] or "{}"))
            account_metrics = json.loads(str(account_row["metrics_json"] or "{}"))
            account_coverage = json.loads(str(account_row["coverage_json"] or "{}"))
            stale_cost_rows: list[dict[str, Any]] = []
            checked_cost_dependencies: set[tuple[str, str]] = set()
            for detail in coverage.get("detail_rows") or []:
                dependency_key = (
                    str(detail.get("operation_date") or ""),
                    str(detail.get("source_digest") or ""),
                )
                if dependency_key in checked_cost_dependencies:
                    continue
                checked_cost_dependencies.add(dependency_key)
                try:
                    operation_day = date.fromisoformat(str(detail.get("operation_date") or ""))
                except ValueError:
                    stale_cost_rows.append({"reason": "operation_date_invalid", **dict(detail)})
                    continue
                current = resolve_channel_location_cost(
                    conn,
                    nm_id=nm_id,
                    operation_date=operation_day,
                    fbs_order_id=(
                        int(detail.get("fbs_order_id") or 0) or None
                    ),
                )
                if (
                    current.get("status") != "resolved"
                    or str(current.get("source_digest") or "")
                    != str(detail.get("source_digest") or "")
                    or str(detail.get("formula_version") or "")
                    != COST_METHOD_VERSION
                ):
                    stale_cost_rows.append(
                        {
                            "operation_date": operation_day.isoformat(),
                            "expected_source_digest": str(detail.get("source_digest") or ""),
                            "current_source_digest": str(current.get("source_digest") or ""),
                            "expected_formula_version": str(
                                detail.get("formula_version") or ""
                            ),
                            "current_formula_version": COST_METHOD_VERSION,
                            "current_reason": str(current.get("reason") or ""),
                        }
                    )
            if stale_cost_rows:
                blockers.append(
                    {
                        "code": "finance_sku_aggregate_cost_stale",
                        "week_start": week_start_text,
                        "nm_id": nm_id,
                        "rows": stale_cost_rows,
                    }
                )
            if account_coverage.get("identity_blockers"):
                blockers.append(
                    {
                        "code": "finance_identity_ambiguous",
                        "week_start": week_start_text,
                        "rows": account_coverage["identity_blockers"],
                    }
                )
            selected_revenue = _decimal(metrics.get("net_revenue"))
            total_revenue = _decimal(account_coverage.get("global_net_revenue"))
            account_expense = self._account_allocatable_expense_total(account_metrics)
            allocated_common = ZERO
            allocation_ratio: Decimal | None = None
            allocated_categories = {
                key: ZERO for key, _label in OTHER_EXPENSE_CATEGORIES
            }
            allocated_main_expenses = {
                key: ZERO for key in ALLOCATED_MAIN_EXPENSE_KEYS
            }
            account_sources = {
                **self._account_main_expense_source(account_metrics),
                **self._account_expense_category_source(account_metrics),
            }
            if any(value != ZERO for value in account_sources.values()):
                if total_revenue <= ZERO:
                    blockers.append(
                        {
                            "code": "common_expense_zero_revenue_base",
                            "week_start": week_start_text,
                        }
                    )
                else:
                    allocation_ratio = selected_revenue / total_revenue
                    allocated_common = account_expense * allocation_ratio
                    allocated_categories = self._allocated_account_expense_categories(
                        account_metrics,
                        allocation_ratio=allocation_ratio,
                    )
                    allocated_main_expenses = self._allocated_account_main_expenses(
                        account_metrics,
                        allocation_ratio=allocation_ratio,
                    )
                    if (
                        sum(allocated_categories.values(), ZERO)
                        + sum(allocated_main_expenses.values(), ZERO)
                    ).quantize(MONEY_QUANT) != allocated_common.quantize(MONEY_QUANT):
                        raise PartnerReportError(
                            "combined allocated Partner categories do not conserve",
                            code="common_expense_conservation_failed",
                        )
            ads_value, ads_rows, ads_blockers = self._ads_for_week(
                conn,
                nm_id=nm_id,
                week_start=week_start,
                week_end=week_end,
            )
            blockers.extend(ads_blockers)
            other_expense_categories = self._other_expense_categories(
                metrics,
                allocated_categories,
            )
            for category_key, amount in other_expense_categories.items():
                period_other_expenses[category_key] += amount
                if amount != ZERO:
                    visible_other_expense_keys.add(category_key)
            week_values = self._week_formulas(
                components=metrics,
                cogs=(
                    _decimal(metrics.get("cogs"))
                    if metrics.get("cogs") not in (None, "")
                    and metrics.get("profit_after_cogs") not in (None, "")
                    else None
                ),
                ads=ads_value,
                other_expense_categories=other_expense_categories,
                allocated_main_expenses=allocated_main_expenses,
                other_direct_and_allocated_total=sum(
                    other_expense_categories.values(), ZERO
                ),
                params=params,
            )
            week_record = {
                "week_start": week_start_text,
                "week_end": week_end.isoformat(),
                "label": f"{week_start.strftime('%d.%m')}–{week_end.strftime('%d.%m')}",
                "values": week_values,
                "other_expense_breakdown": _display_breakdown(
                    other_expense_categories,
                    target_total=_decimal(
                        week_values[OTHER_DIRECT_ALLOCATED_KEY]
                    ),
                ),
                "coverage": {
                    "finance_status": str(sync["status"]),
                    "finance_raw_row_count": int(sync["raw_row_count"] or 0),
                    "selected_finance_row_count": int(sku_row["raw_row_count"] or 0),
                    "cost": {
                        key: coverage.get(key)
                        for key in (
                            "matched_units",
                            "unmatched_units",
                            "coverage_pct",
                            "covered_sales_revenue_rub",
                            "uncovered_sales_revenue_rub",
                            "covered_sales_order_count",
                            "uncovered_sales_order_count",
                            "covered_sales_units",
                            "uncovered_sales_units",
                            "profit_coverage_status",
                            "daily_rows",
                            "problem_skus",
                            "cost_state_hash",
                        )
                    },
                    "ads_date_count": 7,
                    "ads_covered_date_count": 7 - len(ads_blockers),
                },
            }
            week_records.append(week_record)
            provenance_weeks.append(
                {
                    "week_start": week_start_text,
                    "week_end": week_end.isoformat(),
                    "finance_source_digest": str(sku_row["raw_source_digest"]),
                    "finance_week_content_hash": sync_hash,
                    "finance_formula_version": str(sku_row["formula_version"]),
                    "ads_rows": ads_rows,
                    "ads_source_digest": _sha256_json(ads_rows),
                    "cost_rows": coverage.get("detail_rows") or [],
                    "cost_source_digest": str(coverage.get("cost_state_hash") or ""),
                    "common_expense_safe": {
                        "allocated_amount_rub": _decimal_text(allocated_common),
                        "finance_marketing_excluded_rub": _decimal_text(
                            _decimal(account_metrics.get("marketing"))
                            * (allocation_ratio or ZERO)
                        ),
                        "rule": COMMON_EXPENSE_RULE,
                        "allocation_coefficient": _decimal_text(
                            allocation_ratio, Decimal("0.00000001")
                        ),
                        "source_digest": str(account_row["raw_source_digest"]),
                    },
                    "other_expense_internal": [
                        {
                            "source_category": category_key,
                            "direct_amount_rub": _decimal_text(
                                other_expense_categories[category_key]
                                - allocated_categories[category_key]
                            ),
                            "allocated_amount_rub": _decimal_text(
                                allocated_categories[category_key]
                            ),
                            "allocation_rule": COMMON_EXPENSE_RULE,
                            "direct_source_digest": str(
                                sku_row["raw_source_digest"]
                            ),
                            "allocated_source_digest": str(
                                account_row["raw_source_digest"]
                            ),
                            "total_amount_rub": _decimal_text(
                                other_expense_categories[category_key]
                            ),
                        }
                        for category_key, _label in OTHER_EXPENSE_CATEGORIES
                    ],
                    "allocated_main_expense_internal": [
                        {
                            "source_category": category_key,
                            "allocated_amount_rub": _decimal_text(
                                allocated_main_expenses[category_key]
                            ),
                            "allocation_rule": COMMON_EXPENSE_RULE,
                            "allocated_source_digest": str(
                                account_row["raw_source_digest"]
                            ),
                        }
                        for category_key in ALLOCATED_MAIN_EXPENSE_KEYS
                    ],
                }
            )

        totals = self._period_totals(week_records, params=params)
        source_manifest = {
            "settings_fingerprint": settings["fingerprint"],
            "formula_version": PARTNER_REPORT_FORMULA_VERSION,
            "finance": [
                {
                    "week_start": item["week_start"],
                    "digest": item["finance_source_digest"],
                    "week_content_hash": item["finance_week_content_hash"],
                    "formula_version": item["finance_formula_version"],
                }
                for item in provenance_weeks
            ],
            "ads": [
                {"week_start": item["week_start"], "digest": item["ads_source_digest"]}
                for item in provenance_weeks
            ],
            "cost": [
                {"week_start": item["week_start"], "digest": item["cost_source_digest"]}
                for item in provenance_weeks
            ],
        }
        source_digest = _sha256_json(source_manifest)
        report = {
            "status": "incomplete" if blockers else "ready",
            "contract_version": PARTNER_REPORT_SCHEMA_VERSION,
            "formula_version": PARTNER_REPORT_FORMULA_VERSION,
            "finance_cost_formula_version": COST_METHOD_VERSION,
            "finance_profit_formula_version": PROFIT_METHOD_VERSION,
            "finance_sku_formula_version": SKU_AGGREGATE_FORMULA_VERSION,
            "settings_version_id": settings["settings_version_id"],
            "settings_fingerprint": settings["fingerprint"],
            "nm_id": nm_id,
            "product_name": settings["product_name"],
            "parameters": params,
            "selected_weeks": selected_weeks,
            "weeks": week_records,
            "totals": totals,
            "other_expense_breakdown_total": _display_breakdown(
                period_other_expenses,
                target_total=_decimal(totals[OTHER_DIRECT_ALLOCATED_KEY]),
            ),
            "other_expense_category_definitions": [
                {"key": key, "label": label}
                for key, label in OTHER_EXPENSE_CATEGORIES
                if key in visible_other_expense_keys
            ],
            "other_expense_tooltip": OTHER_DIRECT_ALLOCATED_TOOLTIP,
            "source_coverage": {"complete": not blockers, "blocker_count": len(blockers)},
            "source_manifest": source_manifest,
            "source_digest": source_digest,
            "blockers": blockers,
            "finalized": False,
            "generated_at": _now_iso(self.now_factory),
            "annualized_return_formula": (
                "average weekly dividends × 52 / invested capital × 100%; "
                "calculated, not guaranteed"
            ),
            "preview_source": "indexed_per_sku_weekly_finance_aggregate",
        }
        provenance = {
            "schema_version": "partner_report_provenance_v3",
            "seller_id": self.seller_id,
            "nm_id": nm_id,
            "weeks": provenance_weeks,
            "source_manifest": source_manifest,
            "source_digest": source_digest,
        }
        return report, provenance

    @staticmethod
    def _other_expense_categories(
        components: Mapping[str, Any],
        allocated: Mapping[str, Decimal],
    ) -> dict[str, Decimal]:
        """Return exact direct + allocated values for public subrows."""

        direct = {
            "uncapitalized_transit_logistics": (
                _decimal(components.get("transit_logistics"))
                - _decimal(components.get("capitalized_transit_logistics"))
            ),
            "wb_jam_subscription": _decimal(components.get("subscriptions")),
            "wb_paid_services": _decimal(components.get("paid_services")),
            "review_points": _decimal(components.get("review_points")),
            "other_withholdings": _decimal(components.get("other_deductions")),
        }
        return {
            key: direct[key] + _decimal(allocated.get(key))
            for key, _label in OTHER_EXPENSE_CATEGORIES
        }

    @staticmethod
    def _allocated_account_expense_categories(
        account_metrics: Mapping[str, Any],
        *,
        allocation_ratio: Decimal,
    ) -> dict[str, Decimal]:
        """Allocate account-only subrow expenses once without a catch-all residual."""

        source = PartnerReportBlock._account_expense_category_source(account_metrics)
        allocated = {
            key: source[key] * allocation_ratio
            for key, _label in OTHER_EXPENSE_CATEGORIES
        }
        if sum(allocated.values(), ZERO).quantize(MONEY_QUANT) != (
            sum(source.values(), ZERO) * allocation_ratio
        ).quantize(MONEY_QUANT):
            raise PartnerReportError(
                "allocated Partner subrow categories do not conserve their source total",
                code="common_expense_conservation_failed",
            )
        return allocated

    @staticmethod
    def _allocated_account_main_expenses(
        account_metrics: Mapping[str, Any],
        *,
        allocation_ratio: Decimal,
    ) -> dict[str, Decimal]:
        source = PartnerReportBlock._account_main_expense_source(account_metrics)
        allocated = {
            key: source[key] * allocation_ratio for key in ALLOCATED_MAIN_EXPENSE_KEYS
        }
        if sum(allocated.values(), ZERO).quantize(MONEY_QUANT) != (
            sum(source.values(), ZERO) * allocation_ratio
        ).quantize(MONEY_QUANT):
            raise PartnerReportError(
                "allocated Partner main-row expenses do not conserve their source total",
                code="common_expense_conservation_failed",
            )
        return allocated

    @staticmethod
    def _account_expense_category_source(
        account_metrics: Mapping[str, Any],
    ) -> dict[str, Decimal]:
        """Return account expenses that belong under Partner subrows."""

        return {
            "uncapitalized_transit_logistics": (
                _decimal(account_metrics.get("transit_logistics"))
                - _decimal(account_metrics.get("capitalized_transit_logistics"))
            ),
            "wb_jam_subscription": _decimal(account_metrics.get("subscriptions")),
            "wb_paid_services": _decimal(account_metrics.get("paid_services")),
            "review_points": _decimal(account_metrics.get("review_points")),
            "other_withholdings": _decimal(account_metrics.get("other_deductions")),
        }

    @staticmethod
    def _account_main_expense_source(
        account_metrics: Mapping[str, Any],
    ) -> dict[str, Decimal]:
        """Return account expenses routed to existing Partner main rows."""

        return {
            "agent_remuneration": _decimal(account_metrics.get("agent_remuneration")),
            "acquiring": _decimal(account_metrics.get("acquiring")),
            "logistics": _decimal(account_metrics.get("logistics")),
            "storage": _decimal(account_metrics.get("storage")),
            "acceptance": (
                _decimal(account_metrics.get("acceptance"))
                - _decimal(account_metrics.get("capitalized_acceptance"))
            ),
            "penalties_and_adjustments": (
                _decimal(account_metrics.get("penalties"))
                + _decimal(account_metrics.get("corrections"))
                - _decimal(account_metrics.get("positive_adjustments"))
            ),
        }

    @staticmethod
    def _account_allocatable_expense_total(
        account_metrics: Mapping[str, Any],
    ) -> Decimal:
        """Prove explicit account categories and exclude Finance marketing."""

        main = PartnerReportBlock._account_main_expense_source(account_metrics)
        subrows = PartnerReportBlock._account_expense_category_source(account_metrics)
        classified_total = sum(main.values(), ZERO) + sum(subrows.values(), ZERO)
        expected_total = (
            _decimal(account_metrics.get("profit_period_expenses"))
            - _decimal(account_metrics.get("positive_adjustments"))
            - _decimal(account_metrics.get("marketing"))
        )
        if classified_total.quantize(MONEY_QUANT) != expected_total.quantize(MONEY_QUANT):
            raise PartnerReportError(
                "account expense categories do not reconcile after marketing exclusion",
                code="common_expense_conservation_failed",
                blockers=[
                    {
                        "code": "common_expense_category_reconciliation_failed",
                        "classified_total_rub": _decimal_text(classified_total),
                        "expected_without_marketing_rub": _decimal_text(expected_total),
                        "finance_marketing_excluded_rub": _decimal_text(
                            _decimal(account_metrics.get("marketing"))
                        ),
                    }
                ],
            )
        return classified_total

    def _week_formulas(
        self,
        *,
        components: Mapping[str, Decimal],
        cogs: Decimal | None,
        ads: Decimal | None,
        other_expense_categories: Mapping[str, Decimal],
        allocated_main_expenses: Mapping[str, Decimal] | None = None,
        other_direct_and_allocated_total: Decimal | None = None,
        params: Mapping[str, str],
    ) -> dict[str, str | None]:
        net_revenue = _decimal(components.get("net_revenue"))
        profit_revenue = (
            _decimal(components.get("profit_revenue_covered"))
            if "profit_revenue_covered" in components
            else net_revenue
        )
        allocated_main = allocated_main_expenses or {}
        agent = _decimal(components.get("agent_remuneration")) + _decimal(
            allocated_main.get("agent_remuneration")
        )
        acquiring = _decimal(components.get("acquiring")) + _decimal(
            allocated_main.get("acquiring")
        )
        logistics = _decimal(components.get("logistics")) + _decimal(
            allocated_main.get("logistics")
        )
        storage = _decimal(components.get("storage")) + _decimal(
            allocated_main.get("storage")
        )
        acceptance = (
            _decimal(components.get("acceptance"))
            - _decimal(components.get("capitalized_acceptance"))
        ) + _decimal(allocated_main.get("acceptance"))
        penalties_and_adjustments = (
            _decimal(components.get("penalties"))
            + _decimal(components.get("corrections"))
            - _decimal(components.get("positive_adjustments"))
            + _decimal(allocated_main.get("penalties_and_adjustments"))
        )
        other_direct_and_allocated = (
            other_direct_and_allocated_total
            if other_direct_and_allocated_total is not None
            else sum(
                (
                    _decimal(other_expense_categories.get(key))
                    for key, _label in OTHER_EXPENSE_CATEGORIES
                ),
                ZERO,
            )
        )
        margin = (
            profit_revenue
            - cogs
            - agent
            - acquiring
            - logistics
            - ads
            - storage
            - acceptance
            - penalties_and_adjustments
            - other_direct_and_allocated
            if cogs is not None and ads is not None
            else None
        )
        office = _decimal(params["weekly_office_expense_rub"])
        tax = profit_revenue * _decimal(params["tax_rate_pct"]) / HUNDRED
        reserve = (
            max(margin, ZERO) * _decimal(params["replenishment_reserve_pct"]) / HUNDRED
            if margin is not None
            else None
        )
        net_profit = margin - office - tax - reserve if margin is not None and reserve is not None else None
        dividends = (
            max(net_profit, ZERO) * _decimal(params["partner_share_pct"]) / HUNDRED
            if net_profit is not None
            else None
        )
        capital = _decimal(params["invested_capital_rub"])
        annualized = (
            dividends * Decimal("52") / capital * HUNDRED
            if dividends is not None and capital > ZERO
            else None
        )
        values: dict[str, Decimal | None] = {
            "net_revenue": net_revenue,
            "sales_without_cost_rub": _decimal(
                components.get("sales_without_cost_rub")
            ),
            "orders_without_cost": _decimal(
                components.get("orders_without_cost")
            ),
            "units_without_cost": _decimal(
                components.get("units_without_cost")
            ),
            "cogs": cogs,
            "agent_remuneration": agent,
            "acquiring": acquiring,
            "logistics": logistics,
            "ads": ads,
            "storage": storage,
            "acceptance": acceptance,
            "penalties_and_adjustments": penalties_and_adjustments,
            OTHER_DIRECT_ALLOCATED_KEY: other_direct_and_allocated,
            "finance_margin": margin,
            "office": office,
            "estimated_tax": tax,
            "replenishment_reserve": reserve,
            "net_profit": net_profit,
            "dividends": dividends,
            "annualized_return_pct": annualized,
        }
        return {
            key: _decimal_text(value, PERCENT_QUANT if key.endswith("_pct") else MONEY_QUANT)
            for key, value in values.items()
        }

    def _period_totals(
        self,
        weeks: list[Mapping[str, Any]],
        *,
        params: Mapping[str, str],
        loss_carry_in: Decimal = ZERO,
    ) -> dict[str, str | None]:
        del loss_carry_in
        sum_keys = [key for key, _label in REPORT_ROWS if key != "annualized_return_pct"]
        totals: dict[str, Decimal | None] = {}
        for key in sum_keys:
            source_values = [week["values"].get(key) for week in weeks]
            totals[key] = (
                None
                if any(value is None for value in source_values)
                else sum((_decimal(value) for value in source_values), ZERO)
            )
        capital = _decimal(params["invested_capital_rub"])
        total_dividends = totals.get("dividends")
        annualized = (
            total_dividends / Decimal(len(weeks)) * Decimal("52") / capital * HUNDRED
            if total_dividends is not None and weeks and capital > ZERO
            else None
        )
        result = {key: _decimal_text(value) for key, value in totals.items()}
        result["annualized_return_pct"] = _decimal_text(annualized, PERCENT_QUANT)
        return result

    def _loss_carry_context(
        self,
        conn: sqlite3.Connection,
        *,
        nm_id: str,
        selected_weeks: list[str],
        finalization: bool,
    ) -> tuple[Decimal, dict[str, Any], list[dict[str, Any]]]:
        rows = conn.execute(
            """SELECT report_id,selected_weeks_json,loss_carry_out_rub,source_digest,
                      report_json,finalized_at
               FROM partner_report_finalized_reports
               WHERE seller_id=? AND nm_id=? ORDER BY finalized_at,report_id""",
            (self.seller_id, nm_id),
        ).fetchall()
        selected_dates = [date.fromisoformat(item) for item in selected_weeks]
        selected_set = set(selected_dates)
        first_selected = selected_dates[0]
        last_selected = selected_dates[-1]
        prior_candidates: list[tuple[date, sqlite3.Row]] = []
        overlaps: list[str] = []
        later: list[str] = []
        same_period_rows: list[sqlite3.Row] = []
        for row in rows:
            try:
                report_dates = [
                    date.fromisoformat(str(item))
                    for item in json.loads(str(row["selected_weeks_json"] or "[]"))
                ]
            except (json.JSONDecodeError, ValueError, TypeError):
                if finalization:
                    overlaps.append(str(row["report_id"]))
                continue
            if not report_dates:
                continue
            if report_dates == selected_dates:
                same_period_rows.append(row)
                continue
            report_set = set(report_dates)
            if selected_set & report_set:
                overlaps.append(str(row["report_id"]))
            elif report_dates[-1] < first_selected:
                prior_candidates.append((report_dates[-1], row))
            elif report_dates[0] > last_selected:
                later.append(str(row["report_id"]))
            else:
                overlaps.append(str(row["report_id"]))
        if same_period_rows:
            stored_report = json.loads(str(same_period_rows[0]["report_json"] or "{}"))
            stored_source = dict(
                (stored_report.get("source_manifest") or {}).get("loss_carry") or {}
            )
            stored_carry = max(
                _decimal((stored_report.get("totals") or {}).get("loss_carry_in")),
                ZERO,
            )
            return (
                stored_carry,
                stored_source,
                [],
            )
        blockers: list[dict[str, Any]] = []
        if finalization and overlaps:
            blockers.append(
                {
                    "code": "finalized_period_overlap",
                    "report_ids": sorted(overlaps),
                }
            )
        if finalization and later:
            blockers.append(
                {
                    "code": "finalized_period_out_of_order",
                    "report_ids": sorted(later),
                }
            )
        carry = ZERO
        source: dict[str, Any] = {
            "source_report_id": "",
            "source_digest": "",
            "loss_carry_in_rub": "0.0000",
        }
        if prior_candidates:
            prior_end, prior = max(prior_candidates, key=lambda item: (item[0], item[1]["report_id"]))
            expected_start = prior_end + timedelta(days=7)
            if first_selected == expected_start:
                carry = max(_decimal(prior["loss_carry_out_rub"]), ZERO)
                source = {
                    "source_report_id": str(prior["report_id"]),
                    "source_digest": str(prior["source_digest"] or ""),
                    "source_period_end": prior_end.isoformat(),
                    "loss_carry_in_rub": _decimal_text(carry) or "0.0000",
                }
            elif finalization:
                blockers.append(
                    {
                        "code": "finalized_period_gap",
                        "previous_report_id": str(prior["report_id"]),
                        "required_week_start": expected_start.isoformat(),
                        "selected_week_start": first_selected.isoformat(),
                    }
                )
        return carry, source, blockers

    @staticmethod
    def _empty_components() -> dict[str, Decimal]:
        return {
            "net_revenue": ZERO,
            "commission": ZERO,
            "logistics": ZERO,
            "storage": ZERO,
            "other_direct_expenses": ZERO,
            "positive_adjustments": ZERO,
        }

    def _accumulate_components(
        self,
        values: dict[str, Decimal],
        row: Mapping[str, Any],
        week_start: date,
    ) -> None:
        values["net_revenue"] += self._row_net_revenue(row)
        doc = str(row.get("docTypeName") or "").casefold()
        revenue = _decimal(row.get("retailPriceWithDisc"))
        commission = revenue - _decimal(row.get("forPay"))
        if doc == "продажа":
            values["commission"] += commission
        elif doc == "возврат":
            values["commission"] -= commission
        values["logistics"] += _decimal(row.get("deliveryService"))
        values["storage"] += _decimal(row.get("paidStorage"))
        operation_date, _source = _operation_date(row, week_start)
        if operation_date < RETRO_COST_PERIOD_START:
            values["other_direct_expenses"] += _decimal(row.get("paidAcceptance"))
        values["other_direct_expenses"] += _decimal(row.get("penalty"))
        deduction = _decimal(row.get("deduction"))
        if deduction:
            bucket = classify_deduction(row)
            if bucket == "transit_logistics" and operation_date >= RETRO_COST_PERIOD_START:
                pass
            elif bucket == "marketing":
                pass
            else:
                values["other_direct_expenses"] += deduction
        additional = _decimal(row.get("additionalPayment"))
        if additional >= ZERO:
            values["positive_adjustments"] += additional
        else:
            values["other_direct_expenses"] += abs(additional)

    def _row_account_expense(self, row: Mapping[str, Any], week_start: date) -> Decimal:
        values = self._empty_components()
        self._accumulate_components(values, row, week_start)
        return (
            values["commission"]
            + values["logistics"]
            + values["storage"]
            + values["other_direct_expenses"]
            - values["positive_adjustments"]
        )

    @staticmethod
    def _row_net_revenue(row: Mapping[str, Any]) -> Decimal:
        doc = str(row.get("docTypeName") or "").casefold()
        revenue = _decimal(row.get("retailPriceWithDisc"))
        return revenue if doc == "продажа" else -revenue if doc == "возврат" else ZERO

    def _ads_for_week(
        self,
        conn: sqlite3.Connection,
        *,
        nm_id: str,
        week_start: date,
        week_end: date,
    ) -> tuple[Decimal | None, list[dict[str, Any]], list[dict[str, Any]]]:
        rows: list[dict[str, Any]] = []
        blockers: list[dict[str, Any]] = []
        total = ZERO
        cursor = week_start
        while cursor <= week_end:
            day = cursor.isoformat()
            source = conn.execute(
                """SELECT captured_at,payload_json FROM temporal_source_slot_snapshots
                   WHERE source_key=? AND snapshot_date=? AND snapshot_role=?""",
                (ADS_SOURCE_KEY, day, ADS_SOURCE_ROLE),
            ).fetchone()
            if source is None:
                blockers.append({"code": "ads_date_missing", "date": day, "nm_id": nm_id})
                cursor += timedelta(days=1)
                continue
            try:
                payload = json.loads(str(source["payload_json"] or ""))
            except json.JSONDecodeError:
                payload = {}
            result, envelope_origin = resolve_ads_snapshot_payload(payload)
            kind = str((result or {}).get("kind") or "invalid")
            matched: list[dict[str, Any]] = []
            for item in (result or {}).get("items") or []:
                if not isinstance(item, dict):
                    continue
                item_nm = str(item.get("nm_id", item.get("nmId", "")) or "")
                if item_nm == nm_id:
                    matched.append(item)
            source_digest = "sha256:" + hashlib.sha256(
                str(source["payload_json"]).encode("utf-8")
            ).hexdigest()
            if kind == "empty":
                ads_sum = ZERO
                coverage = "confirmed_empty"
            elif kind == "success" and matched:
                invalid_items = [
                    item
                    for item in matched
                    if item.get("ads_sum") in (None, "")
                    or not self._valid_decimal(item.get("ads_sum"))
                ]
                if invalid_items:
                    blockers.append(
                        {
                            "code": "ads_value_invalid",
                            "date": day,
                            "nm_id": nm_id,
                            "invalid_item_count": len(invalid_items),
                        }
                    )
                    cursor += timedelta(days=1)
                    continue
                ads_sum = sum((_decimal(item["ads_sum"]) for item in matched), ZERO)
                coverage = "covered"
            else:
                blockers.append(
                    {
                        "code": "ads_sku_coverage_missing",
                        "date": day,
                        "nm_id": nm_id,
                        "source_kind": kind,
                    }
                )
                cursor += timedelta(days=1)
                continue
            total += ads_sum
            detail_items = matched if kind == "success" else [{}]
            for item in detail_items:
                rows.append(
                    {
                        "date": day,
                        "nm_id": nm_id,
                        "advert_id": str(
                            item.get("advert_id", item.get("advertId", "")) or ""
                        ),
                        "campaign": str(
                            item.get(
                                "campaign",
                                item.get("campaign_name", item.get("name", "")),
                            )
                            or ""
                        ),
                        "placement": str(
                            item.get("placement", item.get("placement_name", "")) or ""
                        ),
                        "ads_sum": _decimal_text(
                            _decimal(item["ads_sum"]) if kind == "success" else ZERO
                        ),
                        "source_status": kind,
                        "envelope_origin": envelope_origin,
                        "coverage": coverage,
                        "source_digest": source_digest,
                    }
                )
            cursor += timedelta(days=1)
        return (total if not blockers else None), rows, blockers

    @staticmethod
    def _valid_decimal(value: Any) -> bool:
        try:
            return Decimal(str(value)).is_finite()
        except (InvalidOperation, ValueError, TypeError):
            return False

    def _build_package(
        self,
        conn: sqlite3.Connection,
        *,
        report: Mapping[str, Any],
        provenance: Mapping[str, Any],
    ) -> tuple[bytes, str, dict[str, Any]]:
        nm_id = str(report["nm_id"])
        product_slug = self._safe_filename(str(report["product_name"]))
        first_week = str(report["selected_weeks"][0])
        last_week = str(report["selected_weeks"][-1])
        period_slug = f"{first_week}_{last_week}"
        main_name = f"00_Партнёрский_отчёт_{nm_id}_{period_slug}.xlsx"
        entries: list[tuple[str, bytes]] = [
            (main_name, self._build_main_workbook(report))
        ]
        provenance_by_week = {
            str(item["week_start"]): item for item in provenance["weeks"]
        }
        finance_names: list[str] = []
        for index, week in enumerate(report["weeks"], start=1):
            week_start = str(week["week_start"])
            week_end = str(week["week_end"])
            name = f"{index:02d}_Финотчёт_WB_{week_start}_{week_end}_{nm_id}.xlsx"
            finance_names.append(name)
            entries.append(
                (
                    name,
                    self._build_finance_workbook(
                        nm_id=nm_id,
                        product_name=str(report["product_name"]),
                        week_start=week_start,
                        week_end=week_end,
                        rows=provenance_by_week[week_start]["finance_rows"],
                    ),
                )
            )
        ads_rows = [
            row
            for week in provenance["weeks"]
            for row in week["ads_rows"]
        ]
        cost_rows = [
            {**row, "week": str(week["week_start"])}
            for week in provenance["weeks"]
            for row in week["cost_rows"]
        ]
        entries.append(
            (
                f"Реклама_WB_{nm_id}_{period_slug}.xlsx",
                self._build_ads_workbook(report=report, rows=ads_rows),
            )
        )
        entries.append(
            (
                f"Расчёт_себестоимости_{nm_id}_{period_slug}.xlsx",
                self._build_cost_workbook(report=report, rows=cost_rows),
            )
        )
        common_rows = [
            {"week": str(week["week_start"]), **row}
            for week in provenance["weeks"]
            for row in week["common_expense_safe"]
        ]
        entries.append(
            (
                f"Общие_расходы_WB_{nm_id}_{period_slug}.xlsx",
                self._build_common_expense_workbook(report=report, rows=common_rows),
            )
        )
        entries.append(
            (
                "Методология_и_параметры.txt",
                self._methodology_text(report, provenance).encode("utf-8"),
            )
        )
        forbidden_tokens = self._other_sku_tokens(conn, selected_nm_id=nm_id)
        verification = self._verify_package(
            entries,
            report=report,
            provenance=provenance,
            finance_names=finance_names,
            forbidden_tokens=forbidden_tokens,
        )
        if not verification["passed"]:
            raise PartnerReportError(
                "partner package confidentiality or reconciliation verification failed",
                code="package_verification_failed",
                blockers=verification["findings"],
            )
        output = BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, body in entries:
                info = zipfile.ZipInfo(name)
                info.compress_type = zipfile.ZIP_DEFLATED
                archive.writestr(info, body)
        filename = f"Пакет_для_партнёра_{product_slug or nm_id}_{period_slug}.zip"
        verification["zip_sha256"] = "sha256:" + hashlib.sha256(output.getvalue()).hexdigest()
        verification["entry_names"] = [name for name, _body in entries]
        return output.getvalue(), filename, verification

    def _build_main_workbook(self, report: Mapping[str, Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Партнёрский отчёт"
        ws.sheet_view.showGridLines = False
        weeks = list(report["weeks"])
        total_col = 3 + len(weeks)
        ws.cell(1, 1, "")
        ws.cell(1, 2, "")
        for index, week in enumerate(weeks, start=3):
            ws.cell(1, index, str(week["label"]))
        ws.cell(1, total_col, "Итого за период")
        workbook_rows: list[tuple[str, str, bool]] = []
        visible_other_categories = [
            (str(item.get("key") or ""), str(item.get("label") or ""))
            for item in report.get("other_expense_category_definitions") or []
            if str(item.get("key") or "") and str(item.get("label") or "")
        ]
        for key, label in REPORT_ROWS:
            workbook_rows.append((key, label, False))
            if key == OTHER_DIRECT_ALLOCATED_KEY:
                workbook_rows.extend(
                    (f"other_expense:{category_key}", category_label, True)
                    for category_key, category_label in visible_other_categories
                )
        row_by_key: dict[str, int] = {}
        subordinate_rows: set[int] = set()
        for row_no, (key, label, subordinate) in enumerate(workbook_rows, start=2):
            row_by_key[key] = row_no
            if subordinate:
                subordinate_rows.add(row_no)
            coefficient: float | None = None
            if key == "estimated_tax":
                coefficient = float(_decimal(report["parameters"]["tax_rate_pct"]) / HUNDRED)
            elif key == "replenishment_reserve":
                coefficient = float(_decimal(report["parameters"]["replenishment_reserve_pct"]) / HUNDRED)
            elif key == "dividends":
                coefficient = float(_decimal(report["parameters"]["partner_share_pct"]) / HUNDRED)
            ws.cell(row_no, 1, coefficient)
            ws.cell(row_no, 2, label)
            for col_no, week in enumerate(weeks, start=3):
                if key.startswith("other_expense:"):
                    category_key = key.split(":", 1)[1]
                    value = _breakdown_map(
                        week.get("other_expense_breakdown") or []
                    ).get(category_key)
                else:
                    value = week["values"].get(key)
                ws.cell(row_no, col_no, None if value is None else float(Decimal(value)))
            if key.startswith("other_expense:"):
                category_key = key.split(":", 1)[1]
                total_value = _breakdown_map(
                    report.get("other_expense_breakdown_total") or []
                ).get(category_key)
            else:
                total_value = report["totals"].get(key)
            ws.cell(row_no, total_col, None if total_value is None else float(Decimal(total_value)))
        self._style_main_workbook(
            ws,
            row_by_key=row_by_key,
            total_col=total_col,
            max_row=1 + len(workbook_rows),
            subordinate_rows=subordinate_rows,
        )
        parameters = wb.create_sheet("Параметры")
        parameter_rows = [
            ("Отчёт", "Отчёт о доходности карточки"),
            ("Сформирован", str(report.get("generated_at") or "")),
            ("Карточка", str(report["product_name"])),
            ("nmId", str(report["nm_id"])),
            ("Доля инвестора, %", str(report["parameters"]["partner_share_pct"])),
            ("Вложенный капитал, ₽", str(report["parameters"]["invested_capital_rub"])),
            ("Пополнение товарных остатков, %", str(report["parameters"]["replenishment_reserve_pct"])),
            ("Офисные расходы, ₽/нед.", str(report["parameters"]["weekly_office_expense_rub"])),
            ("Налоговая ставка, %", str(report["parameters"]["tax_rate_pct"])),
            ("Недели", ", ".join(str(item) for item in report["selected_weeks"])),
            ("Formula version", str(report["formula_version"])),
            ("Source digest", str(report["source_digest"])),
            ("Формула доходности", str(report["annualized_return_formula"])),
        ]
        for row in parameter_rows:
            parameters.append(row)
        parameters.column_dimensions["A"].width = 38
        parameters.column_dimensions["B"].width = 92
        for row in parameters.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, color="000000")
                cell.border = Border(
                    bottom=Side(style="thin", color="D9D9D9")
                )
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        parameters.freeze_panes = "A2"
        wb.properties.creator = "WB Core"
        wb.properties.title = "Отчёт о доходности карточки"
        wb.properties.subject = f"nmId {report['nm_id']}"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        return self._save_workbook(wb)

    def _style_main_workbook(
        self,
        ws: Any,
        *,
        row_by_key: Mapping[str, int],
        total_col: int,
        max_row: int,
        subordinate_rows: set[int],
    ) -> None:
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=total_col):
            for cell in row:
                cell.font = Font(name="Arial", size=10, color="000000")
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="right")
        for cell in ws[1]:
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_no in range(2, max_row + 1):
            ws.cell(row_no, 2).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
            ws.cell(row_no, 1).number_format = "0.00"
            for col in range(3, total_col + 1):
                ws.cell(row_no, col).number_format = "#,##0.00;[Red](#,##0.00);-"
            if row_no in subordinate_rows:
                ws.cell(row_no, 2).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                    indent=1,
                )
                for col in range(1, total_col + 1):
                    ws.cell(row_no, col).font = Font(
                        name="Arial", size=9, color="666666"
                    )
        for key in ("annualized_return_pct",):
            row_no = row_by_key[key]
            for col in range(3, total_col + 1):
                ws.cell(row_no, col).number_format = "0.0%;[Red](0.0%);-"
                value = ws.cell(row_no, col).value
                if value is not None:
                    ws.cell(row_no, col).value = float(value) / 100
        payout_row = row_by_key["dividends"]
        for col in range(1, total_col + 1):
            ws.cell(payout_row, col).font = Font(name="Arial", size=10, bold=True, color="0070C0")
        roi_row = row_by_key["annualized_return_pct"]
        for col in range(1, total_col + 1):
            ws.cell(roi_row, col).font = Font(name="Arial", size=10, bold=True, color="0070C0")
        total_fill = PatternFill("solid", fgColor="F7F7F7")
        for row_no in range(1, max_row + 1):
            ws.cell(row_no, total_col).fill = total_fill
            ws.cell(row_no, total_col).font = Font(
                name="Arial",
                size=10,
                bold=True,
                color="0070C0" if row_no in {payout_row, roi_row} else "000000",
            )
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 36
        for col in range(3, total_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.row_dimensions[1].height = 30
        for row_no in range(2, max_row + 1):
            ws.row_dimensions[row_no].height = 21
        for key in (
            "replenishment_reserve",
            "annualized_return_pct",
        ):
            ws.row_dimensions[row_by_key[key]].height = 32
        ws.freeze_panes = "C2"
        ws.print_area = f"A1:{get_column_letter(total_col)}{max_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4)

    def _build_finance_workbook(
        self,
        *,
        nm_id: str,
        product_name: str,
        week_start: str,
        week_end: str,
        rows: list[Mapping[str, Any]],
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Выборка по SKU"
        ws.append(["Финотчёт WB — выборка по SKU"])
        ws.append([f"nmId {nm_id} · {product_name} · {week_start}–{week_end}"])
        headers = [label for _source, label in FINANCE_EXPORT_COLUMNS]
        ws.append(headers)
        if rows:
            for row in rows:
                ws.append([row.get(label) for _source, label in FINANCE_EXPORT_COLUMNS])
        else:
            ws.append(["Операции по выбранному SKU отсутствуют"] + [None] * (len(headers) - 1))
        self._style_detail_sheet(ws, header_row=3, freeze="A4")
        wb.properties.creator = "WB Core"
        wb.properties.title = "Финотчёт WB — выборка по SKU"
        return self._save_workbook(wb)

    def _build_ads_workbook(
        self,
        *,
        report: Mapping[str, Any],
        rows: list[Mapping[str, Any]],
    ) -> bytes:
        headers = [
            "date",
            "nm_id",
            "advert_id",
            "campaign",
            "placement",
            "ads_sum",
            "source_status",
            "coverage",
            "source_digest",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Реклама WB"
        ws.append([f"Реклама WB — только nmId {report['nm_id']}"])
        ws.append(headers)
        for row in rows:
            ws.append([row.get(key) for key in headers])
        self._style_detail_sheet(ws, header_row=2, freeze="A3")
        wb.properties.creator = "WB Core"
        wb.properties.title = "Реклама WB — выборка по SKU"
        return self._save_workbook(wb)

    def _build_cost_workbook(
        self,
        *,
        report: Mapping[str, Any],
        rows: list[Mapping[str, Any]],
    ) -> bytes:
        headers = [
            "week",
            "operation_date",
            "movement",
            "quantity",
            "unit_cost_rub",
            "cost_source",
            "source_date",
            "formula_version",
            "signed_cogs_rub",
            "weekly_total_rub",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Себестоимость"
        ws.append([f"Расчёт себестоимости — только nmId {report['nm_id']}"])
        ws.append(headers)
        for row in rows:
            ws.append([row.get(key) for key in headers])
        for week in report["weeks"]:
            ws.append(
                [
                    week["week_start"],
                    "ИТОГО НЕДЕЛИ",
                    None,
                    None,
                    None,
                    None,
                    None,
                    report["finance_cost_formula_version"],
                    None,
                    week["values"]["cogs"],
                ]
            )
        self._style_detail_sheet(ws, header_row=2, freeze="A3")
        wb.properties.creator = "WB Core"
        wb.properties.title = "Расчёт себестоимости — выборка по SKU"
        return self._save_workbook(wb)

    def _build_common_expense_workbook(
        self,
        *,
        report: Mapping[str, Any],
        rows: list[Mapping[str, Any]],
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Распределённые расходы"
        ws.append(["Общие расходы WB — безопасная расшифровка по выбранному SKU"])
        ws.append([f"nmId {report['nm_id']} · {report['product_name']}"])
        headers = (
            "week",
            "category",
            "allocated_amount_rub",
            "rule",
            "formula_version",
            "source_digest",
        )
        ws.append(list(headers))
        if rows:
            for row in rows:
                ws.append([row.get(key) for key in headers])
        else:
            ws.append(
                ["Общие расходы для выбранного SKU отсутствуют"]
                + [None] * (len(headers) - 1)
            )
        self._style_detail_sheet(ws, header_row=3, freeze="A4")
        wb.properties.creator = "WB Core"
        wb.properties.title = "Общие расходы WB — безопасная расшифровка"
        return self._save_workbook(wb)

    @staticmethod
    def _style_detail_sheet(ws: Any, *, header_row: int, freeze: str) -> None:
        ws.sheet_view.showGridLines = False
        light = Side(style="thin", color="D9D9D9")
        for cell in ws[header_row]:
            cell.font = Font(name="Arial", size=10, bold=True, color="000000")
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.border = Border(bottom=light)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = Font(name="Arial", size=9, color="000000")
                cell.border = Border(bottom=light)
                cell.alignment = Alignment(vertical="center")
        for index, column in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 11), 28)
        ws.freeze_panes = freeze
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    def _methodology_text(
        self,
        report: Mapping[str, Any],
        provenance: Mapping[str, Any],
    ) -> str:
        params = report["parameters"]
        lines = [
            "Партнёрский отчёт WB Core",
            f"Report ID: {report.get('report_id', 'preview')}",
            f"nmId: {report['nm_id']}",
            f"Товар: {report['product_name']}",
            f"Период: {report['selected_weeks'][0]} — {report['selected_weeks'][-1]}",
            f"Formula version: {report['formula_version']}",
            f"Source digest: {report['source_digest']}",
            "",
            f"Доля партнёра, %: {params['partner_share_pct']}",
            f"Вложенный капитал, ₽: {params['invested_capital_rub']}",
            f"Резерв ТО, %: {params['replenishment_reserve_pct']}",
            f"Офис, ₽/нед.: {params['weekly_office_expense_rub']}",
            f"Расчётная ставка налога, %: {params['tax_rate_pct']}",
            "Общие расходы WB: распределены по доле чистой выручки SKU; исходная account-level сумма не раскрывается.",
            "Реклама: persisted accepted closed-day ads_compact/fullstats, только выбранный nmId.",
            "Эквайринг не вычитается отдельно; он входит в комиссию WB.",
            "Платная приёмка и транзит с 01.05.2026 не вычитаются повторно: они капитализированы в себестоимости.",
            "Финотчёты в пакете являются конфиденциальной выборкой по SKU, а не неизменённым оригинальным отчётом WB.",
            "",
            f"Finance digests: {', '.join(item['finance_source_digest'] for item in provenance['weeks'])}",
            f"Ads digests: {', '.join(item['ads_source_digest'] for item in provenance['weeks'])}",
            f"Cost digests: {', '.join(item['cost_source_digest'] for item in provenance['weeks'])}",
        ]
        return "\n".join(lines) + "\n"

    def _verify_package(
        self,
        entries: list[tuple[str, bytes]],
        *,
        report: Mapping[str, Any],
        provenance: Mapping[str, Any],
        finance_names: list[str],
        forbidden_tokens: set[str],
    ) -> dict[str, Any]:
        findings: list[dict[str, Any]] = []
        workbook_values: dict[str, dict[str, list[list[Any]]]] = {}
        if len(finance_names) != len(report["selected_weeks"]):
            findings.append({"code": "finance_file_count_mismatch"})
        safe_names = {name for name, _body in entries}
        if len(safe_names) != len(entries):
            findings.append({"code": "duplicate_filename"})
        prohibited_fragments = (
            "/users/",
            "\\users\\",
            "/opt/",
            "file://",
            "bearer ",
            "authorization:",
            "password",
            "credential",
            "vbaProject".casefold(),
        )
        for name, body in entries:
            scan_blobs = [("filename", name.encode("utf-8")), ("body", body)]
            if name.casefold().endswith(".xlsx"):
                try:
                    workbook = load_workbook(BytesIO(body), read_only=False, data_only=False)
                    if any(sheet.sheet_state != "visible" for sheet in workbook.worksheets):
                        findings.append({"code": "hidden_sheet", "file": name})
                    if getattr(workbook, "vba_archive", None) is not None:
                        findings.append({"code": "macro_present", "file": name})
                    workbook_values[name] = {
                        sheet.title: [
                            [cell.value for cell in row]
                            for row in sheet.iter_rows()
                        ]
                        for sheet in workbook.worksheets
                    }
                    workbook.close()
                    with zipfile.ZipFile(BytesIO(body)) as archive:
                        for member in archive.namelist():
                            lowered = member.casefold()
                            if lowered.startswith("xl/externallinks/"):
                                findings.append({"code": "external_link", "file": name, "member": member})
                            if lowered.startswith(("xl/embeddings/", "xl/oleobjects/")):
                                findings.append(
                                    {"code": "embedded_object", "file": name, "member": member}
                                )
                            if "vbaproject" in lowered:
                                findings.append({"code": "macro_present", "file": name, "member": member})
                            scan_blobs.append((member, archive.read(member)))
                except Exception as exc:
                    findings.append({"code": "xlsx_invalid", "file": name, "reason": str(exc)})
            for location, blob in scan_blobs:
                text = blob.decode("utf-8", errors="ignore").casefold()
                for token in forbidden_tokens:
                    if token.casefold() in text:
                        findings.append(
                            {
                                "code": "other_sku_leak",
                                "file": name,
                                "location": location,
                                "token_sha256": "sha256:" + hashlib.sha256(token.encode("utf-8")).hexdigest(),
                            }
                        )
                for fragment in prohibited_fragments:
                    if fragment in text:
                        findings.append(
                            {"code": "sensitive_metadata", "file": name, "location": location}
                        )
        findings.extend(
            self._verify_artifact_values(
                entries=entries,
                workbook_values=workbook_values,
                report=report,
                finance_names=finance_names,
            )
        )
        report_week_by_start = {str(item["week_start"]): item for item in report["weeks"]}
        for source_week in provenance["weeks"]:
            week_start = str(source_week["week_start"])
            expected = report_week_by_start[week_start]["values"]
            direct = self._components_from_safe_rows(
                source_week["finance_rows"],
                week_start=date.fromisoformat(week_start),
            )
            for key in (
                "net_revenue",
                "commission",
                "logistics",
                "storage",
                "other_direct_expenses",
                "positive_adjustments",
            ):
                if _decimal(expected[key]) != direct[key]:
                    findings.append(
                        {"code": "finance_reconciliation_failed", "week_start": week_start, "metric": key}
                    )
            allocated = sum(
                (_decimal(item["allocated_amount_rub"]) for item in source_week["common_expense_safe"]),
                ZERO,
            )
            if _decimal(expected["allocated_common_expenses"]) != allocated:
                findings.append(
                    {"code": "common_expense_reconciliation_failed", "week_start": week_start}
                )
            ads_week_total = sum(
                (_decimal(item["ads_sum"]) for item in source_week["ads_rows"]), ZERO
            )
            if _decimal(expected["ads"]) != ads_week_total:
                findings.append({"code": "ads_reconciliation_failed", "week_start": week_start})
            cogs_week_total = sum(
                (_decimal(item["signed_cogs_rub"]) for item in source_week["cost_rows"]), ZERO
            )
            if _decimal(expected["cogs"]) != cogs_week_total:
                findings.append({"code": "cogs_reconciliation_failed", "week_start": week_start})
        finance_total = sum(
            (_decimal(week["values"]["net_revenue"]) for week in report["weeks"]), ZERO
        )
        ads_total = sum((_decimal(week["values"]["ads"]) for week in report["weeks"]), ZERO)
        cogs_total = sum((_decimal(week["values"]["cogs"]) for week in report["weeks"]), ZERO)
        if _decimal(report["totals"]["net_revenue"]) != finance_total:
            findings.append({"code": "finance_reconciliation_failed"})
        if _decimal(report["totals"]["ads"]) != ads_total:
            findings.append({"code": "ads_reconciliation_failed"})
        if _decimal(report["totals"]["cogs"]) != cogs_total:
            findings.append({"code": "cogs_reconciliation_failed"})
        return {
            "passed": not findings,
            "findings": findings,
            "finance_file_count": len(finance_names),
            "selected_week_count": len(report["selected_weeks"]),
            "other_sku_tokens_checked": len(forbidden_tokens),
            "hidden_sheets": 0 if not any(item.get("code") == "hidden_sheet" for item in findings) else None,
            "external_links": 0 if not any(item.get("code") == "external_link" for item in findings) else None,
            "macros": 0 if not any(item.get("code") == "macro_present" for item in findings) else None,
            "reconciliation": {
                "finance": "ok",
                "ads": "ok",
                "cogs": "ok",
            },
            "source_digest": report["source_digest"],
        }

    def _verify_artifact_values(
        self,
        *,
        entries: list[tuple[str, bytes]],
        workbook_values: Mapping[str, Mapping[str, list[list[Any]]]],
        report: Mapping[str, Any],
        finance_names: list[str],
    ) -> list[dict[str, Any]]:
        findings: list[dict[str, Any]] = []
        main_name = next((name for name, _body in entries if name.startswith("00_")), "")
        main_rows = workbook_values.get(main_name, {}).get("Партнёрский отчёт", [])
        if not main_rows:
            findings.append({"code": "main_workbook_missing"})
        else:
            labels = {
                str(row[1]): row
                for row in main_rows[1:]
                if len(row) > 1 and row[1] not in (None, "")
            }
            for key, label in REPORT_ROWS:
                row = labels.get(label)
                if row is None:
                    findings.append({"code": "main_metric_missing", "metric": key})
                    continue
                expected_values = [
                    week["values"].get(key) for week in report["weeks"]
                ] + [report["totals"].get(key)]
                actual_values = [
                    row[index] if index < len(row) else None
                    for index in range(2, 3 + len(report["weeks"]))
                ]
                percentage = key in {"period_roi_pct", "annualized_return_pct"}
                for position, (actual, expected) in enumerate(
                    zip(actual_values, expected_values, strict=True)
                ):
                    if not self._artifact_decimal_matches(
                        actual,
                        expected,
                        percentage_fraction=percentage,
                    ):
                        findings.append(
                            {
                                "code": "main_workbook_reconciliation_failed",
                                "metric": key,
                                "position": position,
                            }
                        )
        for name, week in zip(finance_names, report["weeks"], strict=True):
            rows = workbook_values.get(name, {}).get("Выборка по SKU", [])
            safe_rows = self._table_rows(rows, header_row=3)
            safe_rows = [row for row in safe_rows if row.get("report_id") not in (None, "")]
            if any(str(row.get("nm_id") or "") != str(report["nm_id"]) for row in safe_rows):
                findings.append(
                    {"code": "finance_workbook_wrong_sku", "week_start": week["week_start"]}
                )
            direct = self._components_from_safe_rows(
                safe_rows,
                week_start=date.fromisoformat(str(week["week_start"])),
            )
            for key in (
                "net_revenue",
                "commission",
                "logistics",
                "storage",
                "other_direct_expenses",
                "positive_adjustments",
            ):
                if not self._artifact_decimal_matches(direct[key], week["values"].get(key)):
                    findings.append(
                        {
                            "code": "finance_workbook_reconciliation_failed",
                            "week_start": week["week_start"],
                            "metric": key,
                        }
                    )
        ads_name = next((name for name, _body in entries if name.startswith("Реклама_WB_")), "")
        ads_rows = self._table_rows(
            workbook_values.get(ads_name, {}).get("Реклама WB", []),
            header_row=2,
        )
        if any(str(row.get("nm_id") or "") != str(report["nm_id"]) for row in ads_rows):
            findings.append({"code": "ads_workbook_wrong_sku"})
        ads_total = sum((_decimal(row.get("ads_sum")) for row in ads_rows), ZERO)
        if not self._artifact_decimal_matches(ads_total, report["totals"].get("ads")):
            findings.append({"code": "ads_workbook_reconciliation_failed"})

        cost_name = next(
            (name for name, _body in entries if name.startswith("Расчёт_себестоимости_")),
            "",
        )
        cost_rows = self._table_rows(
            workbook_values.get(cost_name, {}).get("Себестоимость", []),
            header_row=2,
        )
        cost_details = [row for row in cost_rows if row.get("operation_date") != "ИТОГО НЕДЕЛИ"]
        cost_total = sum((_decimal(row.get("signed_cogs_rub")) for row in cost_details), ZERO)
        if not self._artifact_decimal_matches(cost_total, report["totals"].get("cogs")):
            findings.append({"code": "cost_workbook_reconciliation_failed"})
        weekly_cost_totals = {
            str(row.get("week") or ""): row.get("weekly_total_rub")
            for row in cost_rows
            if row.get("operation_date") == "ИТОГО НЕДЕЛИ"
        }
        for week in report["weeks"]:
            week_start = str(week["week_start"])
            if not self._artifact_decimal_matches(
                weekly_cost_totals.get(week_start), week["values"].get("cogs")
            ):
                findings.append(
                    {"code": "cost_weekly_total_reconciliation_failed", "week_start": week_start}
                )

        common_name = next(
            (name for name, _body in entries if name.startswith("Общие_расходы_WB_")),
            "",
        )
        common_rows = self._table_rows(
            workbook_values.get(common_name, {}).get("Распределённые расходы", []),
            header_row=3,
        )
        common_total = sum(
            (_decimal(row.get("allocated_amount_rub")) for row in common_rows), ZERO
        )
        if not self._artifact_decimal_matches(
            common_total, report["totals"].get("allocated_common_expenses")
        ):
            findings.append({"code": "common_workbook_reconciliation_failed"})

        methodology = next(
            (body.decode("utf-8", errors="strict") for name, body in entries if name == "Методология_и_параметры.txt"),
            "",
        )
        for key in (
            "partner_share_pct",
            "invested_capital_rub",
            "replenishment_reserve_pct",
            "weekly_office_expense_rub",
            "tax_rate_pct",
        ):
            if str(report["parameters"][key]) not in methodology:
                findings.append({"code": "parameter_manifest_reconciliation_failed", "field": key})
        return findings

    @staticmethod
    def _table_rows(rows: list[list[Any]], *, header_row: int) -> list[dict[str, Any]]:
        if len(rows) < header_row:
            return []
        headers = [str(value or "") for value in rows[header_row - 1]]
        return [
            {
                header: (row[index] if index < len(row) else None)
                for index, header in enumerate(headers)
                if header
            }
            for row in rows[header_row:]
        ]

    @staticmethod
    def _artifact_decimal_matches(
        actual: Any,
        expected: Any,
        *,
        percentage_fraction: bool = False,
    ) -> bool:
        if actual is None or expected is None:
            return actual is None and expected is None
        actual_decimal = _decimal(actual)
        if percentage_fraction:
            actual_decimal *= HUNDRED
        return abs(actual_decimal - _decimal(expected)) <= MONEY_QUANT

    def _components_from_safe_rows(
        self,
        rows: Iterable[Mapping[str, Any]],
        *,
        week_start: date,
    ) -> dict[str, Decimal]:
        values = self._empty_components()
        for row in rows:
            remapped = {
                "docTypeName": row.get("operation_type"),
                "retailPriceWithDisc": row.get("revenue"),
                "forPay": row.get("for_pay"),
                "deliveryService": row.get("logistics"),
                "paidStorage": row.get("storage"),
                "paidAcceptance": row.get("paid_acceptance"),
                "penalty": row.get("penalty"),
                "deduction": row.get("deduction"),
                "additionalPayment": row.get("additional_payment"),
                "bonusTypeName": row.get("deduction_name"),
                "rrDate": row.get("operation_date"),
            }
            self._accumulate_components(values, remapped, week_start)
        return values

    def _validate_parameters(self, payload: Mapping[str, Any]) -> dict[str, str]:
        nm_id = str(payload.get("nm_id") or "").strip()
        if not nm_id.isdigit() or int(nm_id) <= 0:
            raise PartnerReportError("nm_id must be a positive integer", code="settings_invalid")
        partner_share = _strict_decimal(payload.get("partner_share_pct"), field="partner_share_pct")
        capital = _strict_decimal(payload.get("invested_capital_rub"), field="invested_capital_rub")
        reserve = _strict_decimal(
            payload.get("replenishment_reserve_pct"), field="replenishment_reserve_pct"
        )
        office = _strict_decimal(
            payload.get("weekly_office_expense_rub"), field="weekly_office_expense_rub"
        )
        tax = _strict_decimal(payload.get("tax_rate_pct"), field="tax_rate_pct")
        for field, value in (
            ("partner_share_pct", partner_share),
            ("replenishment_reserve_pct", reserve),
            ("tax_rate_pct", tax),
        ):
            if value < ZERO or value > HUNDRED:
                raise PartnerReportError(f"{field} must be between 0 and 100", code="settings_invalid")
        if capital <= ZERO:
            raise PartnerReportError("invested_capital_rub must be greater than zero", code="settings_invalid")
        if office < ZERO:
            raise PartnerReportError("weekly_office_expense_rub must be non-negative", code="settings_invalid")
        rule = str(payload.get("common_expense_rule") or "").strip()
        if rule != COMMON_EXPENSE_RULE:
            raise PartnerReportError(
                f"common_expense_rule must be {COMMON_EXPENSE_RULE}", code="settings_invalid"
            )
        return {
            "nm_id": nm_id,
            "partner_share_pct": _decimal_text(partner_share) or "0.0000",
            "invested_capital_rub": _decimal_text(capital) or "0.0000",
            "replenishment_reserve_pct": _decimal_text(reserve) or "0.0000",
            "weekly_office_expense_rub": _decimal_text(office) or "0.0000",
            "tax_rate_pct": _decimal_text(tax) or "0.0000",
            "common_expense_rule": rule,
        }

    def _load_settings(self, conn: sqlite3.Connection, *, nm_id: str) -> dict[str, Any]:
        row = conn.execute(
            """SELECT versions.* FROM partner_report_settings_current AS current
               JOIN partner_report_settings_versions AS versions
                 ON versions.settings_version_id=current.settings_version_id
               WHERE current.seller_id=? AND current.nm_id=?""",
            (self.seller_id, nm_id),
        ).fetchone()
        if row is None:
            raise PartnerReportError(
                f"server-owned settings are not saved for nmId {nm_id}",
                code="settings_missing",
            )
        return self._settings_payload(row)

    @staticmethod
    def _settings_payload(row: sqlite3.Row) -> dict[str, Any]:
        return {
            "settings_version_id": row["settings_version_id"],
            "seller_id": row["seller_id"],
            "nm_id": row["nm_id"],
            "product_name": row["product_name"],
            "parameters": json.loads(row["parameters_json"]),
            "fingerprint": row["fingerprint"],
            "created_at": row["created_at"],
            "created_by": row["created_by"],
        }

    @staticmethod
    def _validate_selected_weeks(
        value: Any, *, require_continuous: bool
    ) -> list[str]:
        if not isinstance(value, list) or not value:
            raise PartnerReportError("selected_weeks must be a non-empty list", code="weeks_invalid")
        parsed: list[date] = []
        for item in value:
            try:
                day = date.fromisoformat(str(item))
            except ValueError as exc:
                raise PartnerReportError("selected_weeks must contain ISO dates", code="weeks_invalid") from exc
            if day.weekday() != 0:
                raise PartnerReportError("selected weeks must start on Monday", code="weeks_invalid")
            parsed.append(day)
        if len(set(parsed)) != len(parsed):
            raise PartnerReportError("selected_weeks must not contain duplicates", code="weeks_invalid")
        parsed.sort()
        for previous, current in zip(parsed, parsed[1:]):
            if require_continuous and current - previous != timedelta(days=7):
                raise PartnerReportError(
                    "final payout periods must be continuous; profitable-only gaps are forbidden",
                    code="weeks_not_continuous",
                )
        return [item.isoformat() for item in parsed]

    def _nomenclature_items(self, conn: sqlite3.Connection) -> list[dict[str, Any]]:
        rows = conn.execute(
            """SELECT nm_id,nomenclature_name,wb_title,vendor_code,barcode,is_active,is_hidden,
                      created_at
               FROM sheet_vitrina_v1_nomenclature_items
               WHERE nm_id IS NOT NULL AND nm_id>0
               ORDER BY is_active DESC,is_hidden ASC,created_at,nm_id"""
        ).fetchall()
        by_nm: dict[str, dict[str, Any]] = {}
        for row in rows:
            nm_id = str(row["nm_id"])
            by_nm.setdefault(
                nm_id,
                {
                    "nm_id": nm_id,
                    "product_name": str(row["nomenclature_name"] or row["wb_title"] or f"nmId {nm_id}"),
                    "vendor_code": str(row["vendor_code"] or ""),
                    "barcode": str(row["barcode"] or ""),
                    "is_active": bool(row["is_active"]),
                    "is_hidden": bool(row["is_hidden"]),
                },
            )
        return list(by_nm.values())

    def _nomenclature_product(
        self, conn: sqlite3.Connection, nm_id: str
    ) -> dict[str, Any] | None:
        return next((item for item in self._nomenclature_items(conn) if item["nm_id"] == nm_id), None)

    @staticmethod
    def _safe_finance_row(
        row: Mapping[str, Any], *, selected_nm_id: str
    ) -> dict[str, Any]:
        safe: dict[str, Any] = {}
        for source, label in FINANCE_EXPORT_COLUMNS:
            value = row.get(source)
            if label == "nm_id":
                value = selected_nm_id
            safe[label] = value
        return safe

    def _other_sku_tokens(self, conn: sqlite3.Connection, *, selected_nm_id: str) -> set[str]:
        rows = conn.execute(
            """SELECT nm_id,vendor_code,barcode,barcodes_json,aliases_json,
                      nomenclature_name,wb_title,our_sku
               FROM sheet_vitrina_v1_nomenclature_items WHERE nm_id IS NOT NULL AND nm_id>0"""
        ).fetchall()
        tokens: set[str] = set()
        for row in rows:
            if str(row["nm_id"]) == selected_nm_id:
                continue
            values: list[str] = [
                str(row["nm_id"] or ""),
                str(row["vendor_code"] or ""),
                str(row["barcode"] or ""),
                str(row["nomenclature_name"] or ""),
                str(row["wb_title"] or ""),
                str(row["our_sku"] or ""),
            ]
            try:
                values.extend(str(item) for item in json.loads(row["barcodes_json"] or "[]"))
            except (json.JSONDecodeError, TypeError):
                pass
            try:
                values.extend(str(item) for item in json.loads(row["aliases_json"] or "[]"))
            except (json.JSONDecodeError, TypeError):
                pass
            for value in values:
                normalized = value.strip()
                if len(normalized) >= 4:
                    tokens.add(normalized)
        return tokens

    def _audit(
        self,
        conn: sqlite3.Connection,
        *,
        action: str,
        object_id: str,
        actor: str,
        payload_digest: str,
        created_at: str,
    ) -> None:
        audit_id = hashlib.sha256(
            f"{self.seller_id}|{action}|{object_id}|{actor}|{created_at}".encode("utf-8")
        ).hexdigest()
        conn.execute(
            """INSERT INTO partner_report_audit(
               audit_id,seller_id,action,object_id,actor,payload_digest,created_at
               ) VALUES(?,?,?,?,?,?,?)""",
            (audit_id, self.seller_id, action, object_id, actor, payload_digest, created_at),
        )

    @staticmethod
    def _save_workbook(workbook: Workbook) -> bytes:
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    @staticmethod
    def _safe_filename(value: str) -> str:
        cleaned = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._-]+", "_", value).strip("._")
        return cleaned[:80]

    def _connect(self) -> sqlite3.Connection:
        manifest = self.store_registry.load()
        conn = self.store_registry.connect(
            "operational",
            mode="rw",
            operation="partner_report",
            manifest=manifest,
        )
        if (
            manifest.state == "cutover"
            and manifest.canonical_source == "split"
        ):
            self.store_registry.attach_readonly(
                conn,
                "finance_raw",
                schema_name="finance_raw_store",
                operation="partner_report_finance_raw_read",
                manifest=manifest,
            )
            conn.execute(
                """CREATE TEMP VIEW wb_finance_weekly_raw_rows AS
                   SELECT seller_id,report_id,rrd_id,report_type,week_start,
                          week_end,nm_id,vendor_code,barcode,doc_type_name,
                          seller_oper_name,row_hash,raw_json,first_seen_at,
                          updated_at
                     FROM finance_raw_store.finance_raw_current_rows"""
            )
        return conn


def block_from_env(runtime_dir: Path) -> PartnerReportBlock:
    import os

    seller_id = os.environ.get("SELLER_PORTAL_CANONICAL_SUPPLIER_ID") or "canonical"
    return PartnerReportBlock(runtime_dir, seller_id=seller_id)
