"""Deterministic customs-declaration item matching and XLSX export."""

from __future__ import annotations

from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any, Iterable, Mapping

from packages.application.supplier_customs_dt_matching_policy import (
    DT_ANNEX_MATCHING_POLICY_VERSION,
    canonical_dt_series,
    normalized_model_key_set,
    resolve_dt_annex_series_model,
)
from packages.contracts.supplier_financial_documents import CUSTOMS_ANNEX_ITEMS_PARSER_VERSION


ZERO = Decimal("0")
MATCHED_STATUSES = {"matched", "matched_by_barcode", "matched_by_compatibility"}
DT_ANNEX_ITEMS_PARSER_VERSION = CUSTOMS_ANNEX_ITEMS_PARSER_VERSION
STRICT_ACCOUNTING_RECONCILIATION_VERSION = "supplier_customs_accounting_reconciliation_v2"
MAX_RECONCILIATION_SEARCH_STATES = 100_000
WORKBOOK_HEADERS = (
    "№ позиции ДТ",
    "№ строки приложения",
    "Наименование из ДТ",
    "Артикул из ДТ",
    "Модель из ДТ",
    "Определённая серия",
    "Количество",
    "Ед. изм.",
    "nmID",
    "Наша номенклатура",
    "Штрихкод",
    "Штрихкод из ДТ",
    "Статус сопоставления",
    "Основание сопоставления",
    "Код ТН ВЭД",
)


def build_customs_breakdown_xlsx(
    *,
    customs_document: Mapping[str, Any],
    shipment: Mapping[str, Any],
    shipment_lines: Iterable[Mapping[str, Any]],
    nomenclature_items: Iterable[Mapping[str, Any]],
    packing_documents: Iterable[Mapping[str, Any]] = (),
    matching_override: Mapping[str, Any] | None = None,
    customs_document_key: str = "",
) -> tuple[bytes, str, dict[str, Any]]:
    """Return a verified workbook, safe filename and bounded mapping receipt."""

    normalized = _mapping(customs_document.get("normalized_parse"))
    goods_items = [dict(item) for item in normalized.get("goods_items") or [] if isinstance(item, Mapping)]
    annex_items = [dict(item) for item in normalized.get("annex_items") or [] if isinstance(item, Mapping)]
    if matching_override is not None:
        matching = _matching_for_customs_document(
            matching_override,
            customs_document_key=customs_document_key,
        )
    elif annex_items:
        matching = match_customs_annex_items(
            annex_items=annex_items,
            goods_items=goods_items,
            shipment_lines=shipment_lines,
            nomenclature_items=nomenclature_items,
            expected_quantity_total=normalized.get("annex_quantity_total"),
            parser_quantity_conserved=normalized.get("annex_quantity_conserved"),
            shipment=shipment,
        )
    else:
        matching = match_customs_annex_items(
            annex_items=(),
            goods_items=goods_items,
            shipment_lines=shipment_lines,
            nomenclature_items=nomenclature_items,
            shipment=shipment,
        )
    if not matching.get("package_ready"):
        raise ValueError("customs breakdown is not fully reconciled")
    declaration_number = str(
        normalized.get("declaration_number")
        or normalized.get("document_number")
        or customs_document.get("document_number")
        or customs_document.get("document_id")
        or "document"
    ).strip()
    declaration_date = str(
        normalized.get("declaration_date")
        or normalized.get("document_date")
        or customs_document.get("document_date")
        or ""
    ).strip()
    header = _mapping(shipment.get("header")) if isinstance(shipment.get("header"), Mapping) else dict(shipment)
    metadata = {
        "declaration_number": declaration_number,
        "declaration_date": declaration_date,
        "supplier_order_id": str(header.get("shipment_id") or customs_document.get("supplier_order_id") or ""),
        "invoice_number": str(header.get("invoice_no") or ""),
        "invoice_date": str(header.get("invoice_date") or ""),
        "source_filename": str(
            customs_document.get("file_original_name")
            or customs_document.get("original_filename")
            or customs_document.get("source_filename")
            or ""
        ),
    }
    workbook_bytes = _render_workbook(metadata=metadata, matching=matching)
    validation = validate_customs_breakdown_workbook(
        workbook_bytes,
        expected_row_count=len(matching["rows"]),
        expected_quantity_total=matching.get("output_quantity_total"),
        require_owned_fields=True,
        expected_controls=matching,
    )
    if not validation["valid"]:
        raise ValueError("generated customs breakdown workbook is invalid: " + "; ".join(validation["errors"]))
    receipt = {
        **{key: value for key, value in matching.items() if key != "rows"},
        "workbook_valid": True,
        "workbook_control_valid": bool(validation.get("control_valid")),
        "workbook_row_count": validation["row_count"],
        "workbook_quantity_total": validation["quantity_total"],
        "declaration_number": declaration_number,
        "declaration_date": declaration_date,
        "requires_review": bool(matching["requires_review"]),
        "review_message": "Расшифровка ДТ требует проверки" if matching["requires_review"] else "",
    }
    return workbook_bytes, customs_breakdown_filename(declaration_number), receipt


def match_customs_annex_items(
    *,
    annex_items: Iterable[Mapping[str, Any]],
    goods_items: Iterable[Mapping[str, Any]],
    shipment_lines: Iterable[Mapping[str, Any]],
    nomenclature_items: Iterable[Mapping[str, Any]],
    expected_quantity_total: Any = None,
    parser_quantity_conserved: Any = None,
    shipment: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Strictly reconcile DT annex rows to current order-owned Invoice product lines."""

    annex_items = [dict(item) for item in annex_items if isinstance(item, Mapping)]
    goods_items = [dict(item) for item in goods_items if isinstance(item, Mapping)]
    all_lines = [dict(line) for line in shipment_lines if isinstance(line, Mapping)]
    product_lines = [line for line in all_lines if str(line.get("line_type") or "") == "product"]
    product_lines.sort(key=lambda item: (int(item.get("sort_order") or 0), str(item.get("line_id") or "")))
    nomenclature_by_nm: dict[int, dict[str, Any]] = {}
    for raw in nomenclature_items:
        item = dict(raw)
        if not bool(item.get("is_active", True)):
            continue
        nm_id = int(item.get("nm_id") or item.get("internal_nm_id") or 0)
        if nm_id > 0:
            nomenclature_by_nm[nm_id] = item

    line_candidates = [_dt_order_candidate(line, nomenclature_by_nm) for line in product_lines]
    owners = _aggregate_dt_order_owners(line_candidates)
    blockers: list[dict[str, Any]] = []
    for index, (line, candidate) in enumerate(zip(product_lines, line_candidates), start=1):
        quantity = _decimal_or_none(line.get("qty"))
        missing_fields = []
        if quantity is None or quantity <= ZERO:
            missing_fields.append("positive_quantity")
        if candidate.get("nm_id", 0) <= 0:
            missing_fields.append("internal_nm_id")
        if str(line.get("match_status") or "") not in MATCHED_STATUSES:
            missing_fields.append("barcode_only_identity")
        if not candidate.get("nomenclature_name"):
            missing_fields.append("nomenclature_name")
        if not candidate.get("canonical_barcode"):
            missing_fields.append("canonical_barcode")
        if not candidate.get("series"):
            missing_fields.append("canonical_series")
        if not candidate.get("model_key_sets"):
            missing_fields.append("compatible_model_keys")
        if missing_fields:
            blockers.append(_blocker("invoice_product_line_incomplete", invoice_line=index, fields=missing_fields))

    invoice_quantity_complete = all(
        (quantity := _decimal_or_none(line.get("qty"))) is not None and quantity > ZERO
        for line in product_lines
    )
    invoice_quantity_total = sum(
        (_decimal_or_none(line.get("qty")) or ZERO for line in product_lines),
        ZERO,
    )
    dt_quantities = [_decimal_or_none(item.get("quantity")) for item in annex_items]
    dt_quantity_complete = all(quantity is not None and quantity > ZERO for quantity in dt_quantities)
    dt_quantity_total = sum((quantity or ZERO for quantity in dt_quantities), ZERO)
    if not annex_items:
        blockers.append(_blocker("dt_annex_rows_missing"))
    for index, item in enumerate(annex_items, start=1):
        missing_fields = []
        quantity = dt_quantities[index - 1]
        if quantity is None or quantity <= ZERO:
            missing_fields.append("positive_quantity")
        if not str(item.get("unit") or "").strip():
            missing_fields.append("unit")
        elif not _is_piece_unit(item.get("unit")):
            missing_fields.append("piece_unit")
        if not str(item.get("parent_position_number") or "").strip():
            missing_fields.append("parent_position_number")
        if missing_fields:
            blockers.append(_blocker("dt_annex_row_incomplete", dt_position=_dt_position_ref(item), fields=missing_fields))

    exact_assignments: dict[int, dict[str, Any]] = {}
    unresolved: list[dict[str, Any]] = []
    owner_allocated: defaultdict[int, Decimal] = defaultdict(lambda: ZERO)
    for index, item in enumerate(annex_items):
        quantity = dt_quantities[index]
        policy = resolve_dt_annex_series_model(item)
        source_barcode = _digits(item.get("barcode") or _mapping(item.get("identifiers")).get("barcode"))
        exact_candidates: list[dict[str, Any]] = []
        exact_basis = ""
        if source_barcode:
            exact_candidates = [owner for owner in owners if source_barcode in owner.get("barcodes", ())]
            exact_basis = "matched_by_exact_dt_barcode"
        if not source_barcode and not exact_candidates and policy.get("status") == "confirmed":
            dt_keys = tuple(str(value) for value in policy.get("model_keys") or [])
            exact_candidates = [
                owner for owner in owners
                if owner.get("series") == policy.get("series")
                and dt_keys in owner.get("model_key_sets", ())
            ]
            exact_basis = "matched_by_exact_series_full_model_set"
        exact_candidates = _unique_dt_owners(exact_candidates)
        if len(exact_candidates) == 1 and quantity is not None and quantity > ZERO:
            owner = exact_candidates[0]
            exact_assignments[index] = owner
            owner_allocated[int(owner["nm_id"])] += quantity
            continue
        compatible = []
        if not source_barcode and policy.get("status") == "confirmed" and quantity is not None and quantity > ZERO:
            dt_keys = set(str(value) for value in policy.get("model_keys") or [])
            compatible = [
                owner for owner in owners
                if owner.get("series") == policy.get("series")
                and dt_keys
                and any(dt_keys.issubset(set(owner_keys)) for owner_keys in owner.get("model_key_sets", ()))
            ]
        unresolved.append({
            "index": index,
            "item": item,
            "quantity": quantity,
            "policy": policy,
            "source_barcode": source_barcode,
            "candidate_owners": _unique_dt_owners(compatible),
            "exact_candidate_count": len(exact_candidates),
            "exact_basis": exact_basis,
        })

    owner_remaining = {
        int(owner["nm_id"]): _decimal(owner.get("invoice_quantity_total")) - owner_allocated[int(owner["nm_id"])]
        for owner in owners
    }
    for nm_id, remaining in owner_remaining.items():
        if remaining < ZERO:
            blockers.append(_blocker("exact_owner_quantity_exceeded", nm_id=nm_id))
    solutions, search_exhausted = _unique_full_reconciliation(
        unresolved,
        owners=owners,
        owner_remaining=owner_remaining,
    )
    bounded_assignments = solutions[0] if len(solutions) == 1 else {}
    if search_exhausted:
        blockers.append(_blocker("bounded_reconciliation_search_limit"))
    elif unresolved and not solutions:
        blockers.append(_blocker("bounded_reconciliation_has_no_full_solution"))
    elif len(solutions) > 1:
        blockers.append(_blocker("bounded_reconciliation_has_multiple_full_solutions"))

    rows: list[dict[str, Any]] = []
    for index, item in enumerate(annex_items):
        quantity = dt_quantities[index]
        policy = resolve_dt_annex_series_model(item)
        source_barcode = _digits(item.get("barcode") or _mapping(item.get("identifiers")).get("barcode"))
        owner = exact_assignments.get(index)
        if owner is not None:
            rows.append(_annex_mapping_row(
                item=item,
                policy=policy,
                quantity=quantity,
                source_barcode=source_barcode,
                candidate=owner,
                status="matched",
                status_ru="Сопоставлено",
                basis=(
                    "matched_by_exact_dt_barcode"
                    if source_barcode and source_barcode in owner.get("barcodes", ())
                    else f"{DT_ANNEX_MATCHING_POLICY_VERSION}: matched_by_exact_series_full_model_set"
                ),
            ))
            continue
        owner = bounded_assignments.get(index)
        if owner is not None:
            rows.append(_annex_mapping_row(
                item=item,
                policy=policy,
                quantity=quantity,
                source_barcode=source_barcode,
                candidate=owner,
                status="matched_by_bounded_reconciliation",
                status_ru="Сопоставлено",
                basis=(
                    f"{DT_ANNEX_MATCHING_POLICY_VERSION}: matched_by_bounded_reconciliation; "
                    "exact series; model-key containment; exact quantity; unique full order solution"
                ),
            ))
        else:
            unresolved_item = next((entry for entry in unresolved if entry["index"] == index), {})
            ambiguous = bool(
                policy.get("status") == "ambiguous"
                or unresolved_item.get("exact_candidate_count", 0) > 1
                or unresolved_item.get("candidate_owners")
                or len(solutions) > 1
            )
            rows.append(_annex_mapping_row(
                item=item,
                policy=policy,
                quantity=quantity,
                source_barcode=source_barcode,
                status="ambiguous" if ambiguous else "unmatched",
                status_ru="Неоднозначно" if ambiguous else "Не сопоставлено",
                basis=(
                    "нет единственного полного bounded reconciliation решения"
                    if ambiguous
                    else "нет доказанного владельца внутри текущего заказа"
                ),
            ))

    status_counts = Counter(str(row.get("status") or "") for row in rows)
    matched_count = sum(status_counts.get(status, 0) for status in ("matched", "matched_by_bounded_reconciliation"))
    owner_dt_totals: defaultdict[int, Decimal] = defaultdict(lambda: ZERO)
    for row in rows:
        if int(row.get("nm_id") or 0) > 0 and row.get("status") in {"matched", "matched_by_bounded_reconciliation"}:
            owner_dt_totals[int(row["nm_id"])] += _decimal(row.get("quantity"))
    reconciled_owner_ids = {
        int(owner["nm_id"])
        for owner in owners
        if owner_dt_totals[int(owner["nm_id"])] == _decimal(owner.get("invoice_quantity_total"))
    }
    per_owner_quantity_reconciled = bool(owners) and len(reconciled_owner_ids) == len(owners)
    invoice_product_lines_covered = sum(
        len(owner.get("invoice_line_ids") or [])
        for owner in owners
        if int(owner["nm_id"]) in reconciled_owner_ids
    )
    expected_total = _decimal_or_none(expected_quantity_total)
    invoice_dt_totals_equal = bool(
        dt_quantity_complete
        and invoice_quantity_complete
        and dt_quantity_total == invoice_quantity_total
        and (expected_total is None or expected_total == dt_quantity_total)
    )
    quantity_conserved = invoice_dt_totals_equal and parser_quantity_conserved is not False
    header = _mapping((shipment or {}).get("header")) if isinstance((shipment or {}).get("header"), Mapping) else dict(shipment or {})
    header_total = _decimal_or_none(header.get("product_qty_total"))
    if len(product_lines) != len(annex_items):
        blockers.append(_blocker("invoice_dt_product_row_count_mismatch"))
    if not invoice_dt_totals_equal:
        blockers.append(_blocker("invoice_dt_product_quantity_total_mismatch"))
    if header_total is None or header_total != invoice_quantity_total:
        blockers.append(_blocker("shipment_header_product_quantity_total_mismatch"))
    if matched_count != len(annex_items):
        blockers.append(_blocker("dt_rows_not_fully_matched"))
    if invoice_product_lines_covered != len(product_lines):
        blockers.append(_blocker("invoice_product_lines_not_fully_covered"))
    if not per_owner_quantity_reconciled:
        blockers.append(_blocker("per_owner_quantity_not_reconciled"))
    if parser_quantity_conserved is False:
        blockers.append(_blocker("parser_annex_quantity_not_conserved"))
    blockers = _dedupe_blockers(blockers)
    package_ready = not blockers
    return {
        "rows": rows,
        "position_count": len(goods_items),
        "annex_item_count": len(annex_items),
        "output_row_count": len(rows),
        "status_counts": dict(sorted(status_counts.items())),
        "matched_count": matched_count,
        "ambiguous_count": status_counts.get("ambiguous", 0),
        "unmatched_count": status_counts.get("unmatched", 0),
        "invoice_product_row_count": len(product_lines),
        "dt_annex_row_count": len(annex_items),
        "invoice_product_quantity_total": _decimal_text(invoice_quantity_total) if invoice_quantity_complete else None,
        "dt_quantity_total": _decimal_text(dt_quantity_total) if dt_quantity_complete else None,
        "shipment_header_product_quantity_total": _decimal_text(header_total),
        "output_quantity_total": _decimal_text(dt_quantity_total) if dt_quantity_complete else None,
        "invoice_product_lines_covered": invoice_product_lines_covered,
        "invoice_product_line_count": len(product_lines),
        "per_owner_quantity_reconciled": per_owner_quantity_reconciled,
        "quantity_conserved": quantity_conserved,
        "matching_policy_version": DT_ANNEX_MATCHING_POLICY_VERSION,
        "reconciliation_version": STRICT_ACCOUNTING_RECONCILIATION_VERSION,
        "blocker_reasons": blockers,
        "reconciliation_status": "ready" if package_ready else "blocked",
        "package_ready": package_ready,
        "requires_review": not package_ready,
    }


def reconcile_customs_accounting_package(
    *,
    customs_documents: Iterable[Mapping[str, Any]],
    shipment: Mapping[str, Any],
    shipment_lines: Iterable[Mapping[str, Any]],
    nomenclature_items: Iterable[Mapping[str, Any]],
) -> dict[str, Any]:
    """Reconcile every active DT as one fail-closed accounting-package proof."""

    annex_items: list[dict[str, Any]] = []
    goods_items: list[dict[str, Any]] = []
    parser_blockers: list[dict[str, Any]] = []
    document_count = 0
    for index, raw_document in enumerate(customs_documents, start=1):
        document = dict(raw_document)
        document_count += 1
        document_key = str(document.get("_customs_document_key") or f"dt:{index}")
        normalized = _mapping(document.get("normalized_parse"))
        document_goods = [dict(item) for item in normalized.get("goods_items") or [] if isinstance(item, Mapping)]
        document_annex = [dict(item) for item in normalized.get("annex_items") or [] if isinstance(item, Mapping)]
        goods_items.extend(document_goods)
        for item in document_annex:
            item["_customs_document_key"] = document_key
            annex_items.append(item)
        if str(normalized.get("annex_items_parser_version") or "") != DT_ANNEX_ITEMS_PARSER_VERSION:
            parser_blockers.append(_blocker("dt_annex_projection_version_mismatch", document=document_key))
        if not document_annex:
            parser_blockers.append(_blocker("dt_annex_rows_missing", document=document_key))
        document_quantities = [_decimal_or_none(item.get("quantity")) for item in document_annex]
        normalized_document_total = _decimal_or_none(normalized.get("annex_quantity_total"))
        calculated_document_total = sum((quantity or ZERO for quantity in document_quantities), ZERO)
        if (
            normalized.get("annex_quantity_conserved") is not True
            or not document_annex
            or any(quantity is None for quantity in document_quantities)
            or normalized_document_total is None
            or normalized_document_total != calculated_document_total
        ):
            parser_blockers.append(_blocker("parser_annex_quantity_not_conserved", document=document_key))
        goods_positions = {
            str(item.get("position_number") or "").strip()
            for item in document_goods
            if str(item.get("position_number") or "").strip()
        }
        annex_positions = {
            str(item.get("parent_position_number") or "").strip()
            for item in document_annex
            if str(item.get("parent_position_number") or "").strip()
        }
        if (
            normalized.get("annex_parent_positions_complete") is not True
            or not goods_positions
            or annex_positions != goods_positions
            or any(str(item.get("parent_position_number") or "").strip() not in goods_positions for item in document_annex)
        ):
            parser_blockers.append(_blocker("dt_parent_position_coverage_incomplete", document=document_key))
    matching = match_customs_annex_items(
        annex_items=annex_items,
        goods_items=goods_items,
        shipment_lines=shipment_lines,
        nomenclature_items=nomenclature_items,
        expected_quantity_total=sum(
            (_decimal_or_none(item.get("quantity")) or ZERO for item in annex_items),
            ZERO,
        ) if annex_items else None,
        parser_quantity_conserved=not parser_blockers,
        shipment=shipment,
    )
    blockers = [*matching.get("blocker_reasons", []), *parser_blockers]
    unresolved_positions = [
        {
            "document": str(row.get("_customs_document_key") or ""),
            "position": f"{row.get('position_number') or '?'}:{row.get('annex_row_number') or '?'}",
        }
        for row in matching.get("rows") or []
        if str(row.get("status") or "") not in {"matched", "matched_by_bounded_reconciliation"}
    ][:50]
    if unresolved_positions:
        blockers.append(_blocker("unresolved_dt_positions", positions=unresolved_positions))
    blockers = _dedupe_blockers(blockers)
    matching.update({
        "position_count": len(goods_items),
        "customs_document_count": document_count,
        "blocker_reasons": blockers,
        "package_ready": not blockers,
        "requires_review": bool(blockers),
        "reconciliation_status": "ready" if not blockers else "blocked",
    })
    return matching


def match_customs_goods_items(
    *,
    goods_items: Iterable[Mapping[str, Any]],
    shipment_lines: Iterable[Mapping[str, Any]],
    nomenclature_items: Iterable[Mapping[str, Any]],
    packing_documents: Iterable[Mapping[str, Any]] = (),
) -> dict[str, Any]:
    goods_items = [dict(item) for item in goods_items if isinstance(item, Mapping)]
    product_lines = [
        dict(line)
        for line in shipment_lines
        if str(line.get("line_type") or "") == "product"
        and int(line.get("internal_nm_id") or 0) > 0
        and str(line.get("match_status") or "") in MATCHED_STATUSES
    ]
    product_lines.sort(key=lambda item: (int(item.get("sort_order") or 0), str(item.get("line_id") or "")))
    line_by_barcode: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    line_by_name: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for line in product_lines:
        barcode = _digits(line.get("barcode"))
        if barcode:
            line_by_barcode[barcode].append(line)
        for source_name in _line_source_names(line):
            line_by_name[source_name].append(line)

    nomenclature_by_barcode: defaultdict[str, set[int]] = defaultdict(set)
    nomenclature_name_by_nm: dict[int, str] = {}
    for item in nomenclature_items:
        if not bool(item.get("is_active", True)):
            continue
        nm_id = int(item.get("nm_id") or 0)
        if nm_id <= 0:
            continue
        nomenclature_name_by_nm[nm_id] = str(item.get("nomenclature_name") or "")
        barcode_values = [item.get("barcode"), *(item.get("barcodes") or [])]
        for value in barcode_values:
            barcode = _digits(value)
            if barcode:
                nomenclature_by_barcode[barcode].add(nm_id)

    _attach_packing_exact_names(line_by_name, product_lines, packing_documents)

    rows: list[dict[str, Any]] = []
    status_counts: defaultdict[str, int] = defaultdict(int)
    dt_quantity_total = ZERO
    output_quantity_total = ZERO
    quantity_complete = True
    for item in goods_items:
        position = str(item.get("position_number") or "")
        source_name = str(item.get("source_name") or "").strip()
        quantity = _decimal_or_none(item.get("quantity"))
        unit = str(item.get("unit") or "").strip()
        barcode = _digits(item.get("barcode") or _mapping(item.get("identifiers")).get("barcode"))
        if quantity is None:
            quantity_complete = False
        else:
            dt_quantity_total += quantity

        candidates = _unique_lines(line_by_barcode.get(barcode, [])) if barcode else []
        barcode_basis = ""
        if not candidates and barcode:
            owner_nm_ids = nomenclature_by_barcode.get(barcode, set())
            candidates = _unique_lines(
                line for line in product_lines if int(line.get("internal_nm_id") or 0) in owner_nm_ids
            )
            if candidates:
                barcode_basis = "точный barcode через server-owned номенклатуру"
        elif candidates:
            barcode_basis = "точный barcode строки заказа"

        if barcode and len(candidates) == 1:
            rows.append(
                _mapping_row(
                    item=item,
                    line=candidates[0],
                    quantity=quantity,
                    status="matched",
                    status_ru="Сопоставлено",
                    basis=barcode_basis,
                    barcode=barcode,
                    nomenclature_name_by_nm=nomenclature_name_by_nm,
                )
            )
            status_counts["matched"] += 1
            if quantity is not None:
                output_quantity_total += quantity
            continue
        if barcode and len(candidates) > 1:
            rows.append(_unmatched_row(item, quantity, "ambiguous", "Неоднозначно", "barcode принадлежит нескольким строкам заказа", barcode))
            status_counts["ambiguous"] += 1
            if quantity is not None:
                output_quantity_total += quantity
            continue

        source_model = str(_mapping(item.get("identifiers")).get("source_model") or "").strip()
        exact_source_keys = [
            key
            for key in (_name_key(source_name), _name_key(source_model))
            if key
        ]
        candidate_sets = [
            _unique_lines(line_by_name.get(key, []))
            for key in dict.fromkeys(exact_source_keys)
            if line_by_name.get(key)
        ]
        name_candidates = _unique_lines(
            line for candidates_for_key in candidate_sets for line in candidates_for_key
        )
        if len(name_candidates) == 1:
            rows.append(
                _mapping_row(
                    item=item,
                    line=name_candidates[0],
                    quantity=quantity,
                    status="matched",
                    status_ru="Сопоставлено",
                    basis="точное source model/name связанного invoice или packing list",
                    barcode=barcode,
                    nomenclature_name_by_nm=nomenclature_name_by_nm,
                )
            )
            status_counts["matched"] += 1
            if quantity is not None:
                output_quantity_total += quantity
            continue
        if len(name_candidates) > 1:
            group_total = sum((_decimal(line.get("qty")) for line in name_candidates), ZERO)
            candidate_keys_consistent = bool(
                candidate_sets
                and all(
                    {
                        str(line.get("line_id") or "")
                        for line in candidates_for_key
                    }
                    == {
                        str(line.get("line_id") or "")
                        for line in name_candidates
                    }
                    for candidates_for_key in candidate_sets
                )
            )
            if (
                candidate_keys_consistent
                and quantity is not None
                and quantity >= ZERO
                and group_total == quantity
            ):
                for line in name_candidates:
                    allocated_quantity = _decimal(line.get("qty"))
                    rows.append(
                        _mapping_row(
                            item=item,
                            line=line,
                            quantity=allocated_quantity,
                            status="reconciled_group",
                            status_ru="Сопоставлено группой",
                            basis=(
                                "точная reconciliation-группа: source name совпадает, "
                                f"сумма количеств строк заказа {group_total} = количеству позиции ДТ {quantity}"
                            ),
                            barcode=barcode,
                            nomenclature_name_by_nm=nomenclature_name_by_nm,
                            source_quantity=quantity,
                        )
                    )
                    output_quantity_total += allocated_quantity
                status_counts["reconciled_group"] += 1
                continue
            rows.append(
                _unmatched_row(
                    item,
                    quantity,
                    "ambiguous",
                    "Неоднозначно",
                    (
                        "точные source model/name дали противоречивые строки заказа"
                        if not candidate_keys_consistent
                        else "точное source model/name совпало с несколькими SKU, но контроль количества группы не сошёлся"
                    ),
                    barcode,
                )
            )
            status_counts["ambiguous"] += 1
            if quantity is not None:
                output_quantity_total += quantity
            continue

        rows.append(
            _unmatched_row(
                item,
                quantity,
                "unmatched",
                "Не сопоставлено",
                "нет точного barcode, source model/name или сходящейся reconciliation-группы",
                barcode,
            )
        )
        status_counts["unmatched"] += 1
        if quantity is not None:
            output_quantity_total += quantity

    quantity_conserved = quantity_complete and dt_quantity_total == output_quantity_total
    requires_review = bool(
        not goods_items
        or status_counts.get("ambiguous")
        or status_counts.get("unmatched")
        or not quantity_conserved
    )
    return {
        "rows": rows,
        "position_count": len(goods_items),
        "output_row_count": len(rows),
        "status_counts": dict(sorted(status_counts.items())),
        "dt_quantity_total": _decimal_text(dt_quantity_total) if quantity_complete else None,
        "output_quantity_total": _decimal_text(output_quantity_total) if quantity_complete else None,
        "quantity_conserved": quantity_conserved,
        "reconciliation_status": "ok" if not requires_review else "requires_review",
        "requires_review": requires_review,
    }


def _dt_order_candidate(
    line: Mapping[str, Any],
    nomenclature_by_nm: Mapping[int, Mapping[str, Any]],
) -> dict[str, Any]:
    nm_id = int(line.get("internal_nm_id") or 0)
    nomenclature = _mapping(nomenclature_by_nm.get(nm_id))
    group_key = str(
        line.get("group_key")
        or line.get("product_type")
        or nomenclature.get("group_key")
        or nomenclature.get("product_type")
        or ""
    ).strip()
    model_key_sets: list[tuple[str, ...]] = []

    def add_model_key_set(values: Iterable[Any]) -> None:
        keys = tuple(sorted({str(value).strip() for value in values if str(value or "").strip()}))
        if keys and keys not in model_key_sets:
            model_key_sets.append(keys)

    for raw_keys in (line.get("compatible_model_keys"), nomenclature.get("compatible_model_keys")):
        if isinstance(raw_keys, list):
            add_model_key_set(raw_keys)
    for source in (
        line.get("model_raw"),
        _mapping(line.get("raw")).get("model_raw"),
        nomenclature.get("compatible_models_text"),
        nomenclature.get("match_key"),
        nomenclature.get("nomenclature_name"),
    ):
        add_model_key_set(normalized_model_key_set(source))
    additional_barcodes = nomenclature.get("barcodes")
    if not isinstance(additional_barcodes, list):
        additional_barcodes = []
    barcodes = {
        barcode
        for value in (
            line.get("barcode"),
            nomenclature.get("barcode"),
            *additional_barcodes,
        )
        if (barcode := _digits(value))
    }
    canonical_barcode = _digits(line.get("barcode")) or _digits(nomenclature.get("barcode"))
    if not canonical_barcode and barcodes:
        canonical_barcode = sorted(barcodes)[0]
    return {
        "nm_id": nm_id,
        "line_id": str(line.get("line_id") or ""),
        "invoice_line_identity": str(line.get("line_id") or f"sort:{int(line.get('sort_order') or 0)}"),
        "sort_order": int(line.get("sort_order") or 0),
        "invoice_quantity": _decimal_text(_decimal_or_none(line.get("qty"))),
        "nomenclature_name": str(
            line.get("internal_name")
            or nomenclature.get("nomenclature_name")
            or nomenclature.get("internal_name")
            or ""
        ),
        "canonical_barcode": canonical_barcode,
        "barcodes": tuple(sorted(barcodes)),
        "group_key": group_key,
        "series": canonical_dt_series(group_key),
        "model_keys": model_key_sets[0] if model_key_sets else (),
        "model_key_sets": tuple(model_key_sets),
    }


def _aggregate_dt_order_owners(candidates: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    owners: dict[int, dict[str, Any]] = {}
    for raw in candidates:
        candidate = dict(raw)
        nm_id = int(candidate.get("nm_id") or 0)
        if nm_id <= 0:
            continue
        owner = owners.setdefault(
            nm_id,
            {
                **candidate,
                "invoice_quantity_total": "0",
                "invoice_line_ids": [],
                "barcodes": set(),
                "model_key_sets": [],
            },
        )
        owner["invoice_quantity_total"] = _decimal_text(
            _decimal(owner.get("invoice_quantity_total")) + _decimal(candidate.get("invoice_quantity"))
        )
        line_identity = str(candidate.get("invoice_line_identity") or "")
        if line_identity:
            owner["invoice_line_ids"].append(line_identity)
        owner["barcodes"].update(candidate.get("barcodes") or ())
        for model_keys in candidate.get("model_key_sets") or ():
            normalized_keys = tuple(str(value) for value in model_keys)
            if normalized_keys and normalized_keys not in owner["model_key_sets"]:
                owner["model_key_sets"].append(normalized_keys)
        if (int(candidate.get("sort_order") or 0), str(candidate.get("line_id") or "")) < (
            int(owner.get("sort_order") or 0), str(owner.get("line_id") or "")
        ):
            for key in ("line_id", "sort_order", "nomenclature_name", "canonical_barcode", "group_key", "series"):
                owner[key] = candidate.get(key)
    result: list[dict[str, Any]] = []
    for nm_id in sorted(owners):
        owner = owners[nm_id]
        owner["barcodes"] = tuple(sorted(owner["barcodes"]))
        owner["model_key_sets"] = tuple(owner["model_key_sets"])
        owner["invoice_line_ids"] = tuple(owner["invoice_line_ids"])
        result.append(owner)
    return result


def _unique_full_reconciliation(
    unresolved: Iterable[Mapping[str, Any]],
    *,
    owners: Iterable[Mapping[str, Any]],
    owner_remaining: Mapping[int, Decimal],
) -> tuple[list[dict[int, dict[str, Any]]], bool]:
    owner_by_nm = {int(owner.get("nm_id") or 0): dict(owner) for owner in owners if int(owner.get("nm_id") or 0) > 0}
    remaining = {nm_id: Decimal(value) for nm_id, value in owner_remaining.items()}
    entries: list[dict[str, Any]] = []
    for raw in unresolved:
        entry = dict(raw)
        quantity = entry.get("quantity")
        candidates = [
            owner_by_nm[int(owner.get("nm_id") or 0)]
            for owner in entry.get("candidate_owners") or []
            if int(owner.get("nm_id") or 0) in owner_by_nm
            and quantity is not None
            and quantity > ZERO
            and remaining.get(int(owner.get("nm_id") or 0), ZERO) >= quantity
        ]
        entry["candidate_owners"] = candidates
        entries.append(entry)
    if any(entry.get("quantity") is None or not entry.get("candidate_owners") for entry in entries):
        return [], False
    if sum((_decimal(entry.get("quantity")) for entry in entries), ZERO) != sum(remaining.values(), ZERO):
        return [], False
    entries.sort(key=lambda item: (len(item.get("candidate_owners") or []), int(item.get("index") or 0)))
    solutions: list[dict[int, dict[str, Any]]] = []
    assignments: dict[int, dict[str, Any]] = {}
    states = 0
    exhausted = False

    def visit(offset: int) -> None:
        nonlocal states, exhausted
        if len(solutions) >= 2 or exhausted:
            return
        states += 1
        if states > MAX_RECONCILIATION_SEARCH_STATES:
            exhausted = True
            return
        if offset >= len(entries):
            if all(value == ZERO for value in remaining.values()):
                solutions.append(dict(assignments))
            return
        entry = entries[offset]
        quantity = _decimal(entry.get("quantity"))
        for owner in entry.get("candidate_owners") or []:
            nm_id = int(owner.get("nm_id") or 0)
            if remaining.get(nm_id, ZERO) < quantity:
                continue
            remaining[nm_id] -= quantity
            assignments[int(entry["index"])] = owner
            visit(offset + 1)
            assignments.pop(int(entry["index"]), None)
            remaining[nm_id] += quantity

    visit(0)
    return solutions, exhausted


def _matching_for_customs_document(
    matching: Mapping[str, Any],
    *,
    customs_document_key: str,
) -> dict[str, Any]:
    result = dict(matching)
    rows = [
        dict(row)
        for row in matching.get("rows") or []
        if isinstance(row, Mapping)
        and (not customs_document_key or str(row.get("_customs_document_key") or "") == customs_document_key)
    ]
    result["rows"] = rows
    result["output_row_count"] = len(rows)
    quantities = [_decimal_or_none(row.get("quantity")) for row in rows]
    result["output_quantity_total"] = (
        _decimal_text(sum((quantity or ZERO for quantity in quantities), ZERO))
        if rows and all(quantity is not None for quantity in quantities)
        else None
    )
    result["document_dt_quantity_total"] = result["output_quantity_total"]
    return result


def _is_piece_unit(value: Any) -> bool:
    normalized = re.sub(r"[^a-zа-яё]+", "", str(value or "").casefold())
    return normalized in {"шт", "штука", "штук", "pcs", "pc", "piece", "pieces"}


def _dt_position_ref(item: Mapping[str, Any]) -> str:
    parent = str(item.get("parent_position_number") or item.get("position_number") or "?")
    row = str(item.get("annex_row_number") or "?")
    return f"{parent}:{row}"


def _blocker(code: str, **details: Any) -> dict[str, Any]:
    messages = {
        "invoice_product_line_incomplete": "Строка Invoice не имеет полного доверенного product-line evidence.",
        "dt_annex_rows_missing": "Строки приложения ДТ не распознаны.",
        "dt_annex_row_incomplete": "Строка приложения ДТ не содержит доказанного количества, единицы или позиции.",
        "exact_owner_quantity_exceeded": "Точное сопоставление превысило количество product line Invoice.",
        "bounded_reconciliation_search_limit": "Bounded reconciliation превысила безопасный лимит поиска.",
        "bounded_reconciliation_has_no_full_solution": "Для оставшихся строк нет полного решения reconciliation.",
        "bounded_reconciliation_has_multiple_full_solutions": "Для оставшихся строк существует несколько полных решений reconciliation.",
        "invoice_dt_product_row_count_mismatch": "Количество товарных строк Invoice и строк ДТ не совпало.",
        "invoice_dt_product_quantity_total_mismatch": "Итоговое количество товара в Invoice и ДТ не совпало.",
        "shipment_header_product_quantity_total_mismatch": "Итог shipment header не совпал с product lines Invoice.",
        "dt_rows_not_fully_matched": "Не все строки ДТ имеют единственного владельца.",
        "invoice_product_lines_not_fully_covered": "Не все product lines Invoice покрыты строками ДТ.",
        "per_owner_quantity_not_reconciled": "Количество не сошлось по владельцам nmID.",
        "parser_annex_quantity_not_conserved": "Parser не подтвердил сохранение количества приложения ДТ.",
        "dt_annex_projection_version_mismatch": "ДТ не имеет актуальной parser projection строк приложения.",
        "dt_parent_position_coverage_incomplete": "Строки приложения ДТ не покрывают canonical позиции товара.",
        "unresolved_dt_positions": "Некоторые позиции ДТ не имеют доказанного владельца.",
    }
    return {"code": code, "message": messages.get(code, code), **details}


def _dedupe_blockers(blockers: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for blocker in blockers:
        item = dict(blocker)
        key = repr(sorted(item.items()))
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


def _unique_dt_owners(candidates: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    owners: dict[int, dict[str, Any]] = {}
    for raw in candidates:
        candidate = dict(raw)
        nm_id = int(candidate.get("nm_id") or 0)
        if nm_id <= 0:
            continue
        current = owners.get(nm_id)
        candidate_key = (int(candidate.get("sort_order") or 0), str(candidate.get("line_id") or ""))
        current_key = (
            int((current or {}).get("sort_order") or 0),
            str((current or {}).get("line_id") or ""),
        )
        if current is None or candidate_key < current_key:
            owners[nm_id] = candidate
    return [owners[nm_id] for nm_id in sorted(owners)]


def _annex_mapping_row(
    *,
    item: Mapping[str, Any],
    policy: Mapping[str, Any],
    quantity: Decimal | None,
    source_barcode: str,
    status: str,
    status_ru: str,
    basis: str,
    candidate: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    candidate = dict(candidate or {})
    identifiers = _mapping(item.get("identifiers"))
    return {
        "_customs_document_key": str(item.get("_customs_document_key") or ""),
        "position_number": str(item.get("parent_position_number") or item.get("position_number") or ""),
        "annex_row_number": str(item.get("annex_row_number") or ""),
        "source_name": str(item.get("source_name") or ""),
        "source_article": str(item.get("article") or identifiers.get("article") or ""),
        "source_model": str(item.get("source_model") or identifiers.get("source_model") or ""),
        "determined_series": _dt_series_label(policy.get("series")),
        "canonical_group": str(candidate.get("group_key") or ""),
        "quantity": _decimal_text(quantity) if quantity is not None else None,
        "unit": str(item.get("unit") or ""),
        "nm_id": int(candidate.get("nm_id") or 0) or None,
        "nomenclature_name": str(candidate.get("nomenclature_name") or ""),
        "barcode": str(candidate.get("canonical_barcode") or ""),
        "source_barcode": source_barcode,
        "status": status,
        "status_ru": status_ru,
        "basis": basis,
        "customs_code": str(identifiers.get("customs_code") or item.get("customs_code") or ""),
    }


def _upgrade_legacy_matching(matching: Mapping[str, Any]) -> dict[str, Any]:
    upgraded = dict(matching)
    upgraded["rows"] = [
        {
            **dict(row),
            "annex_row_number": "",
            "source_article": "",
            "determined_series": "",
            "source_barcode": str(row.get("barcode") or ""),
        }
        for row in matching.get("rows") or []
        if isinstance(row, Mapping)
    ]
    upgraded["annex_item_count"] = 0
    upgraded["matched_count"] = sum(
        int(value)
        for key, value in dict(upgraded.get("status_counts") or {}).items()
        if key in {"matched", "reconciled_group"}
    )
    upgraded["ambiguous_count"] = int(dict(upgraded.get("status_counts") or {}).get("ambiguous") or 0)
    upgraded["unmatched_count"] = int(dict(upgraded.get("status_counts") or {}).get("unmatched") or 0)
    upgraded["matching_policy_version"] = "supplier_customs_legacy_exact_v1"
    return upgraded


def _dt_series_label(value: Any) -> str:
    return {
        "clean": "Clean",
        "anti_spy": "Anti-spy",
        "matte": "Matte",
    }.get(str(value or "").strip(), "")


def validate_customs_breakdown_workbook(
    workbook_bytes: bytes,
    *,
    expected_row_count: int,
    expected_quantity_total: Any,
    require_owned_fields: bool = False,
    expected_controls: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    errors: list[str] = []
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
        worksheet = workbook["Расшифровка ДТ"]
        control = workbook["Контроль"]
    except Exception as exc:
        return {"valid": False, "errors": [f"openpyxl readback failed: {exc}"], "row_count": 0, "quantity_total": None}
    header_row = None
    for row_index in range(1, min(worksheet.max_row, 40) + 1):
        values = tuple(worksheet.cell(row=row_index, column=index).value for index in range(1, len(WORKBOOK_HEADERS) + 1))
        if values == WORKBOOK_HEADERS:
            header_row = row_index
            break
    if header_row is None:
        errors.append("workbook header row is missing")
        return {"valid": False, "errors": errors, "row_count": 0, "quantity_total": None}
    if expected_row_count <= 0:
        errors.append("workbook has no customs goods rows")
    row_count = 0
    quantity_total = ZERO
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        position = worksheet.cell(row=row_index, column=1).value
        if position in (None, ""):
            continue
        row_count += 1
        quantity = worksheet.cell(row=row_index, column=7).value
        if quantity in (None, ""):
            errors.append(f"row {row_index} quantity is missing")
        elif not isinstance(quantity, int | float):
            errors.append(f"row {row_index} quantity is not numeric")
        else:
            quantity_total += Decimal(str(quantity))
        unit = worksheet.cell(row=row_index, column=8).value
        if unit in (None, ""):
            errors.append(f"row {row_index} quantity unit is missing")
        if require_owned_fields:
            for column, label in ((9, "nmID"), (10, "nomenclature name"), (11, "canonical barcode")):
                if worksheet.cell(row=row_index, column=column).value in (None, ""):
                    errors.append(f"row {row_index} {label} is missing")
        for column, label in ((11, "canonical barcode"), (12, "source barcode")):
            barcode_cell = worksheet.cell(row=row_index, column=column)
            if barcode_cell.value not in (None, "") and barcode_cell.data_type != "s":
                errors.append(f"row {row_index} {label} is not stored as text")
    if row_count != expected_row_count:
        errors.append(f"row count {row_count} != expected {expected_row_count}")
    expected_total = _decimal_or_none(expected_quantity_total)
    if expected_total is not None and quantity_total != expected_total:
        errors.append(f"quantity total {quantity_total} != expected {expected_total}")
    control_values = {
        str(control.cell(row=row_index, column=1).value or ""): control.cell(row=row_index, column=2).value
        for row_index in range(1, control.max_row + 1)
    }
    if expected_controls is not None:
        expected = dict(expected_controls)
        control_expectations = {
            "Готовность бухгалтерского пакета": "Да" if expected.get("package_ready") else "Нет",
            "Product rows Invoice": expected.get("invoice_product_row_count"),
            "DT annex rows": expected.get("dt_annex_row_count"),
            "Product quantity Invoice": _excel_number(expected.get("invoice_product_quantity_total")),
            "DT quantity": _excel_number(expected.get("dt_quantity_total")),
            "Сопоставлено": expected.get("matched_count"),
            "Неоднозначно": expected.get("ambiguous_count"),
            "Не сопоставлено": expected.get("unmatched_count"),
            "Покрытие product lines Invoice": (
                f"{expected.get('invoice_product_lines_covered') or 0}/{expected.get('invoice_product_line_count') or 0}"
            ),
            "Per-owner quantity reconciliation": "Да" if expected.get("per_owner_quantity_reconciled") else "Нет",
            "Version parser projection": DT_ANNEX_ITEMS_PARSER_VERSION,
            "Version matching policy": expected.get("matching_policy_version"),
            "Version reconciliation": expected.get("reconciliation_version"),
        }
        for label, expected_value in control_expectations.items():
            if control_values.get(label) != expected_value:
                errors.append(f"control {label} differs from expected reconciliation evidence")
    return {
        "valid": not errors,
        "errors": errors,
        "row_count": row_count,
        "quantity_total": _decimal_text(quantity_total),
        "control_valid": not any(error.startswith("control ") for error in errors),
    }


def customs_breakdown_filename(declaration_number: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9А-Яа-яЁё._-]+", "_", str(declaration_number or "document")).strip("._-")
    return f"DT_{safe or 'document'}_rasshifrovka.xlsx"


def _render_workbook(*, metadata: Mapping[str, Any], matching: Mapping[str, Any]) -> bytes:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Расшифровка ДТ"
    worksheet.append(["Расшифровка ДТ"])
    worksheet["A1"].font = Font(bold=True, size=14)
    metadata_rows = (
        ("Номер ДТ", metadata.get("declaration_number")),
        ("Дата ДТ", metadata.get("declaration_date")),
        ("Заказ", metadata.get("supplier_order_id")),
        ("Invoice", metadata.get("invoice_number")),
        ("Дата invoice", metadata.get("invoice_date")),
        ("Исходный файл", metadata.get("source_filename")),
        ("Готовность бухгалтерского пакета", "Да" if matching.get("package_ready") else "Нет"),
        ("Итого количество по ДТ", matching.get("document_dt_quantity_total") or matching.get("dt_quantity_total")),
        ("Итого количество в расшифровке", matching.get("output_quantity_total")),
        ("Version matching policy", matching.get("matching_policy_version")),
    )
    for label, value in metadata_rows:
        worksheet.append([label, value if value not in (None, "") else "—"])
    worksheet.append([])
    worksheet.append(list(WORKBOOK_HEADERS))
    header_row = worksheet.max_row
    for cell in worksheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4C1D95")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in matching.get("rows") or []:
        quantity = _excel_number(row.get("quantity"))
        worksheet.append(
            [
                str(row.get("position_number") or ""),
                str(row.get("annex_row_number") or ""),
                str(row.get("source_name") or ""),
                str(row.get("source_article") or ""),
                str(row.get("source_model") or ""),
                str(row.get("determined_series") or ""),
                quantity,
                str(row.get("unit") or ""),
                int(row["nm_id"]) if row.get("nm_id") not in (None, "") else None,
                str(row.get("nomenclature_name") or ""),
                str(row.get("barcode") or ""),
                str(row.get("source_barcode") or ""),
                str(row.get("status_ru") or ""),
                str(row.get("basis") or ""),
                str(row.get("customs_code") or ""),
            ]
        )
        worksheet.cell(row=worksheet.max_row, column=11).number_format = "@"
        worksheet.cell(row=worksheet.max_row, column=12).number_format = "@"
    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.auto_filter.ref = f"A{header_row}:O{worksheet.max_row}"
    widths = (16, 18, 42, 30, 30, 22, 15, 12, 15, 34, 24, 24, 24, 70, 18)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    for row in worksheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    control = workbook.create_sheet("Контроль")
    control.append(["Показатель", "Значение"])
    control.append(["Статус reconciliation", matching.get("reconciliation_status")])
    control.append(["Готовность бухгалтерского пакета", "Да" if matching.get("package_ready") else "Нет"])
    control.append(["Позиций ДТ", matching.get("position_count")])
    control.append(["Строк приложения", matching.get("annex_item_count")])
    control.append(["Строк в расшифровке", matching.get("output_row_count")])
    control.append(["Product rows Invoice", matching.get("invoice_product_row_count")])
    control.append(["DT annex rows", matching.get("dt_annex_row_count")])
    control.append(["Product quantity Invoice", _excel_number(matching.get("invoice_product_quantity_total"))])
    control.append(["DT quantity", _excel_number(matching.get("dt_quantity_total"))])
    control.append(["Сопоставлено", matching.get("matched_count")])
    control.append(["Неоднозначно", matching.get("ambiguous_count")])
    control.append(["Не сопоставлено", matching.get("unmatched_count")])
    control.append(["Итого количество ДТ", _excel_number(matching.get("dt_quantity_total"))])
    control.append(["Итого количество расшифровки", _excel_number(matching.get("output_quantity_total"))])
    control.append(["Количество сохранено", "Да" if matching.get("quantity_conserved") else "Нет"])
    control.append([
        "Покрытие product lines Invoice",
        f"{matching.get('invoice_product_lines_covered') or 0}/{matching.get('invoice_product_line_count') or 0}",
    ])
    control.append(["Per-owner quantity reconciliation", "Да" if matching.get("per_owner_quantity_reconciled") else "Нет"])
    control.append(["Version parser projection", DT_ANNEX_ITEMS_PARSER_VERSION])
    control.append(["Version matching policy", matching.get("matching_policy_version")])
    control.append(["Version reconciliation", matching.get("reconciliation_version")])
    for cell in control[1]:
        cell.font = Font(bold=True)
    control.column_dimensions["A"].width = 34
    control.column_dimensions["B"].width = 24
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _mapping_row(
    *,
    item: Mapping[str, Any],
    line: Mapping[str, Any],
    quantity: Decimal | None,
    status: str,
    status_ru: str,
    basis: str,
    barcode: str,
    nomenclature_name_by_nm: Mapping[int, str],
    source_quantity: Decimal | None = None,
) -> dict[str, Any]:
    nm_id = int(line.get("internal_nm_id") or 0)
    identifiers = _mapping(item.get("identifiers"))
    return {
        "position_number": str(item.get("position_number") or ""),
        "source_name": str(item.get("source_name") or ""),
        "quantity": _decimal_text(quantity) if quantity is not None else None,
        "source_quantity": _decimal_text(source_quantity if source_quantity is not None else quantity) if (source_quantity is not None or quantity is not None) else None,
        "unit": str(item.get("unit") or ""),
        "barcode": barcode,
        "source_model": str(identifiers.get("source_model") or ""),
        "customs_code": str(identifiers.get("customs_code") or ""),
        "nm_id": nm_id,
        "nomenclature_name": str(line.get("internal_name") or nomenclature_name_by_nm.get(nm_id) or ""),
        "status": status,
        "status_ru": status_ru,
        "basis": basis,
    }


def _unmatched_row(
    item: Mapping[str, Any],
    quantity: Decimal | None,
    status: str,
    status_ru: str,
    basis: str,
    barcode: str,
) -> dict[str, Any]:
    identifiers = _mapping(item.get("identifiers"))
    return {
        "position_number": str(item.get("position_number") or ""),
        "source_name": str(item.get("source_name") or ""),
        "quantity": _decimal_text(quantity) if quantity is not None else None,
        "source_quantity": _decimal_text(quantity) if quantity is not None else None,
        "unit": str(item.get("unit") or ""),
        "barcode": barcode,
        "source_model": str(identifiers.get("source_model") or ""),
        "customs_code": str(identifiers.get("customs_code") or ""),
        "nm_id": None,
        "nomenclature_name": "",
        "status": status,
        "status_ru": status_ru,
        "basis": basis,
    }


def _line_source_names(line: Mapping[str, Any]) -> set[str]:
    raw = _mapping(line.get("raw"))
    values = (
        line.get("model_raw"),
        raw.get("model_raw"),
        raw.get("name_spec"),
        raw.get("source_name"),
        raw.get("description"),
    )
    return {key for key in (_name_key(value) for value in values) if key}


def _attach_packing_exact_names(
    line_by_name: defaultdict[str, list[dict[str, Any]]],
    product_lines: list[dict[str, Any]],
    packing_documents: Iterable[Mapping[str, Any]],
) -> None:
    for document in packing_documents:
        normalized = _mapping(document.get("normalized_parse"))
        for item in normalized.get("line_items") or []:
            if not isinstance(item, Mapping):
                continue
            packing_names = {_name_key(item.get("model")), _name_key(item.get("description"))} - {""}
            for packing_name in packing_names:
                candidates = _unique_lines(
                    line for line in product_lines if packing_name in _line_source_names(line)
                )
                if candidates:
                    line_by_name[packing_name].extend(candidates)


def _unique_lines(lines: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in lines:
        line = dict(raw)
        key = str(line.get("line_id") or f"nm:{line.get('internal_nm_id')}:{line.get('sort_order')}")
        if key not in seen:
            seen.add(key)
            result.append(line)
    return result


def _mapping(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, Mapping) else {}


def _name_key(value: Any) -> str:
    return re.sub(r"[^0-9a-zа-яё]+", " ", str(value or "").casefold()).strip()


def _digits(value: Any) -> str:
    text = str(value or "").strip().replace(" ", "")
    return text if text.isascii() and text.isdigit() else ""


def _decimal(value: Any) -> Decimal:
    parsed = _decimal_or_none(value)
    return parsed if parsed is not None else ZERO


def _decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _excel_number(value: Any) -> int | float | None:
    parsed = _decimal_or_none(value)
    if parsed is None:
        return None
    return int(parsed) if parsed == parsed.to_integral_value() else float(parsed)
