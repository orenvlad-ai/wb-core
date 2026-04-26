"""Application-слой bounded promo XLSX collector блока."""

from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
import json
from pathlib import Path
import re
import shutil
import time
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from packages.adapters.promo_xlsx_collector_block import PromoCollectorDriver
from packages.application.promo_campaign_archive import resolve_reusable_campaign
from packages.contracts.promo_xlsx_collector_block import (
    CampaignManifestItem,
    CampaignManifestSnapshot,
    CollectorRunSummary,
    CollectorStateSnapshot,
    DownloadArtifact,
    ExportKind,
    HydrationAttemptSummary,
    PromoCardData,
    PromoMetadata,
    PromoOutcome,
    PromoXlsxCollectorRequest,
    TemporalClassification,
    TimelineBlockSnapshot,
    TimelineCandidate,
    WorkbookInspection,
)


BUSINESS_TIMEZONE = ZoneInfo("Asia/Yekaterinburg")
_MONTHS = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}
_VISIBLE_TABS = ("Доступные", "Участвую", "Не участвую", "Акции WB", "Мои акции", "Скидка лояльности")
_GATING_HEADERS = {
    "Минимальная цена для применения скидки по автоакции",
    "Минимальная цена: осталось дней",
    "Блокировка применения скидки по автоакции",
    "Блокировка изменения скидки для участия в автоакции: осталось дней",
}
_COMMON_EXPORT_HEADERS = {
    "Товар уже участвует в акции",
    "Бренд",
    "Предмет",
    "Наименование",
    "Артикул поставщика",
    "Артикул WB",
    "Статус",
}
COLLECTOR_UI_SCHEMA_VERSION = "promo_collector_ui_status_v1"
COLLECTOR_PREFLIGHT_SCHEMA_VERSION = "promo_collector_preflight_v1"
TIMELINE_CLASSIFIER_SCHEMA_VERSION = "promo_timeline_classifier_v1"
MANIFEST_SCHEMA_VERSION = "promo_campaign_manifest_v1"
_ENDED_LABEL_KEYWORDS = ("акция завершилась", "завершилась", "акция завершена", "завершена")
_ACTIVE_LABEL_KEYWORDS = ("акция идёт", "акция идет")
_FUTURE_LABEL_KEYWORDS = ("запланирована", "запланировано")
_PENDING_LABEL_KEYWORDS = ("ожидает", "на модерации")
_TIMELINE_NON_MATERIALIZABLE_STATUSES = {"ended"}
_MANIFEST_NON_MATERIALIZABLE_STATUSES = {"ended"}


class PromoXlsxCollectorBlock:
    def __init__(self, driver: PromoCollectorDriver) -> None:
        self._driver = driver

    def execute(self, request: PromoXlsxCollectorRequest) -> CollectorRunSummary:
        run_dir = Path(request.output_root)
        run_dir.mkdir(parents=True, exist_ok=True)
        summary = CollectorRunSummary(
            run_dir=str(run_dir),
            status="running",
            started_at=_now_iso(),
        )
        self._write_json(run_dir / "run_summary.json", asdict(summary))

        self._driver.start(request)
        try:
            hydrated = False
            for attempt_num in range(1, request.hydration_attempt_budget + 1):
                try:
                    attempt = self._driver.attempt_hydration(attempt_num)
                except Exception as exc:
                    attempt = HydrationAttemptSummary(
                        attempt_num=attempt_num,
                        entry_strategy="direct_open",
                        cookie_clicked=False,
                        hydrated_success=False,
                        title="",
                        url="",
                        timeline_count=0,
                        overlay_count=0,
                        time_to_hydrated_sec=None,
                        blocker=f"hydration_exception={exc}",
                    )
                summary.hydration_attempts.append(attempt)
                self._write_json(run_dir / "run_summary.json", asdict(summary))
                if attempt.hydrated_success:
                    hydrated = True
                    break
            if not hydrated:
                summary.status = "blocked"
                self._write_json(run_dir / "run_summary.json", asdict(summary))
                return summary

            blocks = self._driver.enumerate_timeline_blocks(request.max_candidates)
            candidates = [
                build_timeline_candidate(block)
                for block in blocks
            ]
            candidates = [candidate for candidate in candidates if candidate is not None]
            summary.timeline_candidates_found = len(candidates)
            summary.timeline_card_seen_count = len(candidates)
            manifest_snapshot = self._campaign_manifest_snapshot()
            self._apply_manifest_snapshot_summary(summary, manifest_snapshot)
            self._write_json(run_dir / "run_summary.json", asdict(summary))

            downloads_seen = 0
            recovery_available = True
            for candidate in candidates:
                if request.max_downloads is not None and downloads_seen >= request.max_downloads:
                    break

                outcome = self._process_candidate(
                    request=request,
                    candidate=candidate,
                    manifest_snapshot=manifest_snapshot,
                )
                if outcome.status == "blocked_before_card" and recovery_available:
                    current_timeline_count = self._driver.current_timeline_count()
                    if current_timeline_count <= 0:
                        recovery = self._driver.attempt_hydration(len(summary.hydration_attempts) + 1, label_prefix="recovery")
                        summary.hydration_attempts.append(recovery)
                        summary.hydration_recoveries_used += 1
                        recovery_available = False
                        self._write_json(run_dir / "run_summary.json", asdict(summary))
                        if recovery.hydrated_success:
                            outcome = self._process_candidate(
                                request=request,
                                candidate=candidate,
                                manifest_snapshot=manifest_snapshot,
                            )

                summary.promos.append(outcome)
                self._apply_outcome_preflight_summary(summary, outcome)
                summary.card_confirmed_count += 1 if outcome.card_path else 0
                if outcome.status == "downloaded":
                    summary.downloaded_count += 1
                    downloads_seen += 1
                elif outcome.status == "reused_archive":
                    summary.reused_archive_count += 1
                    downloads_seen += 1
                elif outcome.status == "skipped_past":
                    summary.skipped_past_count += 1
                elif outcome.status == "blocked_before_card":
                    summary.blocked_before_card_count += 1
                elif outcome.status == "blocked_after_card":
                    summary.blocked_after_card_count += 1
                elif outcome.status == "blocked_before_download":
                    summary.blocked_before_download_count += 1
                elif outcome.status == "ambiguous":
                    summary.ambiguous_count += 1
                if outcome.export_kind and outcome.export_kind not in summary.export_kinds:
                    summary.export_kinds.append(outcome.export_kind)
                self._write_json(run_dir / "run_summary.json", asdict(summary))

            if any(
                count > 0
                for count in [
                    summary.blocked_before_card_count,
                    summary.blocked_after_card_count,
                    summary.blocked_before_download_count,
                ]
            ):
                summary.status = "partial"
            else:
                summary.status = "success"
            self._write_json(run_dir / "run_summary.json", asdict(summary))
            return summary
        finally:
            self._driver.stop()

    def _process_candidate(
        self,
        *,
        request: PromoXlsxCollectorRequest,
        candidate: TimelineCandidate,
        manifest_snapshot: CampaignManifestSnapshot,
    ) -> PromoOutcome:
        run_dir = Path(request.output_root)
        timeline_started = time.perf_counter()
        timeline_decision = classify_timeline_preflight(candidate)
        timeline_shallow_duration_ms = _elapsed_ms(timeline_started)
        if timeline_decision["timeline_classification_decision"] == "timeline_non_materializable_expected":
            return self._timeline_non_materializable(
                run_dir=run_dir,
                candidate=candidate,
                request=request,
                timeline_decision=timeline_decision,
                timeline_shallow_duration_ms=timeline_shallow_duration_ms,
            )

        manifest_match_started = time.perf_counter()
        manifest_decision = classify_manifest_preflight(
            candidate=candidate,
            manifest_snapshot=manifest_snapshot,
        )
        manifest_match_duration_ms = _elapsed_ms(manifest_match_started)
        manifest_decision = {
            **manifest_decision,
            "manifest_match_duration_ms": manifest_match_duration_ms,
        }
        if manifest_decision["manifest_decision"] == "drawer_avoid_manifest_non_materializable":
            return self._manifest_non_materializable(
                run_dir=run_dir,
                candidate=candidate,
                request=request,
                timeline_decision=timeline_decision,
                manifest_decision=manifest_decision,
                timeline_shallow_duration_ms=timeline_shallow_duration_ms,
                manifest_match_duration_ms=manifest_match_duration_ms,
            )

        preflight_started = time.perf_counter()
        drawer_open_started = time.perf_counter()
        try:
            card_state = self._driver.open_timeline_candidate(candidate)
            drawer_open_duration_ms = _elapsed_ms(drawer_open_started)
        except Exception as exc:
            drawer_open_duration_ms = _elapsed_ms(drawer_open_started)
            promo_folder = run_dir / "promos" / build_promo_folder_name(None, None, candidate.title)
            promo_folder.mkdir(parents=True, exist_ok=True)
            failure_path = promo_folder / "failure.json"
            failure_png = promo_folder / "failure.png"
            last_state = self._driver.last_state_snapshot()
            if last_state:
                shutil.copy2(last_state.screenshot, failure_png)
            self._write_json(
                failure_path,
                {
                    "status": "blocked_before_card",
                    "candidate_title": candidate.title,
                    "candidate_index": candidate.index,
                    "blocker": str(exc),
                    "last_state": asdict(last_state) if last_state else None,
                },
            )
            metadata = PromoMetadata(
                collected_at=_now_iso(),
                trace_run_dir=request.output_root,
                source_tab=request.source_tab,
                source_filter_code=request.source_filter_code,
                calendar_url=self._driver.current_url(),
                promo_id=None,
                period_id=None,
                promo_title=candidate.title,
                promo_period_text=candidate.short_period_text or "",
                promo_start_at=None,
                promo_end_at=None,
                period_parse_confidence="low",
                temporal_classification="ambiguous",
                promo_status=None,
                promo_status_text=None,
                eligible_count=None,
                participating_count=None,
                excluded_count=None,
                export_kind=None,
                original_suggested_filename=None,
                saved_filename=None,
                saved_path=None,
                timeline_status=candidate.timeline_status,
                timeline_status_confidence=candidate.timeline_status_confidence,
                timeline_status_raw_labels=list(candidate.timeline_status_raw_labels),
                timeline_evidence_sources=list(candidate.timeline_evidence_sources),
                timeline_period_text=candidate.short_period_text,
                timeline_goods_count=candidate.timeline_goods_count,
                timeline_autoaction_marker=candidate.timeline_autoaction_marker,
                timeline_classification_decision=timeline_decision["timeline_classification_decision"],
                drawer_opened=False,
                drawer_open_reason=timeline_decision.get("drawer_open_reason"),
                drawer_skip_reason=None,
                fallback_to_full_flow_reason=timeline_decision.get("fallback_to_full_flow_reason"),
                timeline_classifier_schema_version=TIMELINE_CLASSIFIER_SCHEMA_VERSION,
                **_manifest_metadata_fields(manifest_snapshot, manifest_decision),
            )
            return PromoOutcome(
                promo_title=candidate.title,
                timeline_block_index=candidate.index,
                timeline_short_period_text=candidate.short_period_text,
                timeline_preliminary_classification=candidate.preliminary_classification,
                status="blocked_before_card",
                promo_id=None,
                period_id=None,
                promo_folder=str(promo_folder),
                blocker=str(exc),
                metadata=metadata,
                timeline_shallow_duration_ms=timeline_shallow_duration_ms,
                drawer_open_duration_ms=drawer_open_duration_ms,
                **_timeline_outcome_fields(
                    candidate,
                    timeline_decision,
                    drawer_opened=False,
                    drawer_open_duration_ms=drawer_open_duration_ms,
                ),
            )

        card = extract_card_data(
            snapshot=card_state,
            fallback_title=candidate.title,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            reference_year=datetime.now(BUSINESS_TIMEZONE).year,
        )
        preflight = classify_collector_preflight(card)
        early_preflight_duration_ms = _elapsed_ms(preflight_started)

        if preflight["early_preflight_decision"] == "early_non_materializable":
            promo_folder = run_dir / "promos" / build_promo_folder_name(card.promo_id, None, card.promo_title)
            promo_folder.mkdir(parents=True, exist_ok=True)
            card_path = promo_folder / "card.json"
            card_png = promo_folder / "card.png"
            shutil.copy2(card.state_snapshot.screenshot, card_png)
            self._write_json(card_path, asdict(card))
            metadata = build_metadata(
                card=card,
                trace_run_dir=request.output_root,
                source_tab=request.source_tab,
                source_filter_code=request.source_filter_code,
                export_kind=None,
                download=None,
                workbook=None,
                preflight=preflight,
                timeline_candidate=candidate,
                timeline_decision=timeline_decision,
                drawer_opened=True,
                manifest_snapshot=manifest_snapshot,
                manifest_decision=manifest_decision,
            )
            metadata_path = promo_folder / "metadata.json"
            self._write_json(metadata_path, asdict(metadata))
            drawer_reset = self._driver.reset_drawer(f"early_non_materializable__{slugify(card.promo_title)}")
            return PromoOutcome(
                promo_title=card.promo_title,
                timeline_block_index=candidate.index,
                timeline_short_period_text=candidate.short_period_text,
                timeline_preliminary_classification=candidate.preliminary_classification,
                status="skipped_past",
                promo_id=card.promo_id,
                period_id=None,
                promo_folder=str(promo_folder),
                blocker=None,
                metadata=metadata,
                card_path=str(card_path),
                metadata_path=str(metadata_path),
                drawer_reset=drawer_reset,
                early_preflight_duration_ms=early_preflight_duration_ms,
                **_preflight_outcome_fields(preflight),
                **_timeline_outcome_fields(
                    candidate,
                    timeline_decision,
                    drawer_opened=True,
                    drawer_open_duration_ms=drawer_open_duration_ms,
                ),
            )

        if card.temporal_classification == "past":
            promo_folder = run_dir / "promos" / build_promo_folder_name(card.promo_id, None, card.promo_title)
            promo_folder.mkdir(parents=True, exist_ok=True)
            card_path = promo_folder / "card.json"
            card_png = promo_folder / "card.png"
            shutil.copy2(card.state_snapshot.screenshot, card_png)
            self._write_json(card_path, asdict(card))
            legacy_preflight = {
                **preflight,
                "early_preflight_decision": "legacy_past_skip",
                "heavy_flow_required": False,
                "heavy_flow_reason": "past_temporal_classification",
                "non_materializable_reason": "past_campaign",
                "fallback_to_full_flow_reason": None,
            }
            metadata = build_metadata(
                card=card,
                trace_run_dir=request.output_root,
                source_tab=request.source_tab,
                source_filter_code=request.source_filter_code,
                export_kind=None,
                download=None,
                workbook=None,
                preflight=legacy_preflight,
                timeline_candidate=candidate,
                timeline_decision=timeline_decision,
                drawer_opened=True,
                manifest_snapshot=manifest_snapshot,
                manifest_decision=manifest_decision,
            )
            metadata_path = promo_folder / "metadata.json"
            self._write_json(metadata_path, asdict(metadata))
            drawer_reset = self._driver.reset_drawer(f"skipped_past__{slugify(card.promo_title)}")
            return PromoOutcome(
                promo_title=card.promo_title,
                timeline_block_index=candidate.index,
                timeline_short_period_text=candidate.short_period_text,
                timeline_preliminary_classification=candidate.preliminary_classification,
                status="skipped_past",
                promo_id=card.promo_id,
                period_id=None,
                promo_folder=str(promo_folder),
                blocker=None,
                metadata=metadata,
                card_path=str(card_path),
                metadata_path=str(metadata_path),
                drawer_reset=drawer_reset,
                early_preflight_duration_ms=early_preflight_duration_ms,
                **_preflight_outcome_fields(legacy_preflight),
                **_timeline_outcome_fields(
                    candidate,
                    timeline_decision,
                    drawer_opened=True,
                    drawer_open_duration_ms=drawer_open_duration_ms,
                ),
            )

        if card.temporal_classification == "ambiguous":
            promo_folder = run_dir / "promos" / build_promo_folder_name(card.promo_id, None, card.promo_title)
            promo_folder.mkdir(parents=True, exist_ok=True)
            card_path = promo_folder / "card.json"
            card_png = promo_folder / "card.png"
            shutil.copy2(card.state_snapshot.screenshot, card_png)
            self._write_json(card_path, asdict(card))
            metadata = build_metadata(
                card=card,
                trace_run_dir=request.output_root,
                source_tab=request.source_tab,
                source_filter_code=request.source_filter_code,
                export_kind=None,
                download=None,
                workbook=None,
                preflight={
                    **preflight,
                    "early_preflight_decision": "legacy_ambiguous_skip",
                    "heavy_flow_required": False,
                    "heavy_flow_reason": "ambiguous_temporal_classification",
                    "non_materializable_reason": "ambiguous_temporal_classification",
                },
                timeline_candidate=candidate,
                timeline_decision=timeline_decision,
                drawer_opened=True,
                manifest_snapshot=manifest_snapshot,
                manifest_decision=manifest_decision,
            )
            metadata_path = promo_folder / "metadata.json"
            self._write_json(metadata_path, asdict(metadata))
            drawer_reset = self._driver.reset_drawer(f"ambiguous__{slugify(card.promo_title)}")
            return PromoOutcome(
                promo_title=card.promo_title,
                timeline_block_index=candidate.index,
                timeline_short_period_text=candidate.short_period_text,
                timeline_preliminary_classification=candidate.preliminary_classification,
                status="ambiguous",
                promo_id=card.promo_id,
                period_id=None,
                promo_folder=str(promo_folder),
                blocker=None,
                metadata=metadata,
                card_path=str(card_path),
                metadata_path=str(metadata_path),
                drawer_reset=drawer_reset,
                early_preflight_duration_ms=early_preflight_duration_ms,
                **_preflight_outcome_fields(metadata.__dict__),
                **_timeline_outcome_fields(
                    candidate,
                    timeline_decision,
                    drawer_opened=True,
                    drawer_open_duration_ms=drawer_open_duration_ms,
                ),
            )

        deep_flow_started = time.perf_counter()
        reusable = self._resolve_reusable_archive_record(request=request, card=card)
        if reusable is not None:
            return self._reused_from_archive(
                run_dir=run_dir,
                candidate=candidate,
                card=card,
                request=request,
                record=reusable,
                preflight=preflight,
                early_preflight_duration_ms=early_preflight_duration_ms,
                deep_flow_duration_ms=_elapsed_ms(deep_flow_started),
                timeline_decision=timeline_decision,
                drawer_open_duration_ms=drawer_open_duration_ms,
                manifest_snapshot=manifest_snapshot,
                manifest_decision=manifest_decision,
            )

        try:
            generate_state = self._driver.open_generate_screen(slugify(card.promo_title))
        except Exception as exc:
            return self._blocked_after_card(
                run_dir=run_dir,
                candidate=candidate,
                card=card,
                blocker=str(exc),
                request=request,
                preflight=preflight,
                early_preflight_duration_ms=early_preflight_duration_ms,
                deep_flow_duration_ms=_elapsed_ms(deep_flow_started),
                generate_screen_attempted=True,
                timeline_decision=timeline_decision,
                drawer_open_duration_ms=drawer_open_duration_ms,
                manifest_snapshot=manifest_snapshot,
                manifest_decision=manifest_decision,
            )
        try:
            ready_state = self._driver.generate_file_and_wait_ready(slugify(card.promo_title))
        except Exception as exc:
            return self._blocked_before_download(
                run_dir=run_dir,
                candidate=candidate,
                card=card,
                request=request,
                generate_state=generate_state,
                blocker=str(exc),
                preflight=preflight,
                early_preflight_duration_ms=early_preflight_duration_ms,
                deep_flow_duration_ms=_elapsed_ms(deep_flow_started),
                generate_screen_attempted=True,
                timeline_decision=timeline_decision,
                drawer_open_duration_ms=drawer_open_duration_ms,
                manifest_snapshot=manifest_snapshot,
                manifest_decision=manifest_decision,
            )
        try:
            download = self._driver.download_current_workbook()
        except Exception as exc:
            return self._blocked_before_download(
                run_dir=run_dir,
                candidate=candidate,
                card=card,
                request=request,
                generate_state=ready_state,
                blocker=str(exc),
                preflight=preflight,
                early_preflight_duration_ms=early_preflight_duration_ms,
                deep_flow_duration_ms=_elapsed_ms(deep_flow_started),
                generate_screen_attempted=True,
                download_attempted=True,
                timeline_decision=timeline_decision,
                drawer_open_duration_ms=drawer_open_duration_ms,
            )

        promo_folder = run_dir / "promos" / build_promo_folder_name(card.promo_id, download.period_id, card.promo_title)
        promo_folder.mkdir(parents=True, exist_ok=True)
        card_path = promo_folder / "card.json"
        card_png = promo_folder / "card.png"
        generate_png = promo_folder / "generate_screen.png"
        ready_png = promo_folder / "ready_signal.png"
        workbook_path = promo_folder / "workbook.xlsx"
        shutil.copy2(card.state_snapshot.screenshot, card_png)
        shutil.copy2(generate_state.screenshot, generate_png)
        shutil.copy2(ready_state.screenshot, ready_png)
        shutil.move(download.saved_path, workbook_path)
        self._write_json(card_path, asdict(card))

        inspection = inspect_workbook(workbook_path)
        inspection_path = promo_folder / "workbook_inspection.json"
        self._write_json(inspection_path, asdict(inspection))
        export_kind = classify_export_kind(download.original_suggested_filename, inspection.workbook_header_summary)
        metadata = build_metadata(
            card=card,
            trace_run_dir=request.output_root,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            export_kind=export_kind,
            download=DownloadArtifact(
                original_suggested_filename=download.original_suggested_filename,
                saved_filename=workbook_path.name,
                saved_path=str(workbook_path),
                period_id=download.period_id,
            ),
            workbook=inspection,
            preflight=preflight,
            timeline_candidate=candidate,
            timeline_decision=timeline_decision,
            drawer_opened=True,
            manifest_snapshot=manifest_snapshot,
            manifest_decision=manifest_decision,
        )
        metadata_path = promo_folder / "metadata.json"
        self._write_json(metadata_path, asdict(metadata))
        drawer_reset = self._driver.reset_drawer(f"after_download__{slugify(card.promo_title)}")
        return PromoOutcome(
            promo_title=card.promo_title,
            timeline_block_index=candidate.index,
            timeline_short_period_text=candidate.short_period_text,
            timeline_preliminary_classification=candidate.preliminary_classification,
            status="downloaded",
            promo_id=card.promo_id,
            period_id=download.period_id,
            promo_folder=str(promo_folder),
            blocker=None,
            metadata=metadata,
            card_path=str(card_path),
            metadata_path=str(metadata_path),
            workbook_inspection_path=str(inspection_path),
            saved_path=str(workbook_path),
            original_suggested_filename=download.original_suggested_filename,
            export_kind=export_kind,
            drawer_reset=drawer_reset,
            early_preflight_duration_ms=early_preflight_duration_ms,
            deep_flow_duration_ms=_elapsed_ms(deep_flow_started),
            generate_screen_attempted=True,
            download_attempted=True,
            **_preflight_outcome_fields(preflight),
            **_timeline_outcome_fields(
                candidate,
                timeline_decision,
                drawer_opened=True,
                drawer_open_duration_ms=drawer_open_duration_ms,
            ),
        )

    def _campaign_manifest_snapshot(self) -> CampaignManifestSnapshot:
        snapshot_method = getattr(self._driver, "campaign_manifest_snapshot", None)
        if not callable(snapshot_method):
            return CampaignManifestSnapshot()
        try:
            snapshot = snapshot_method()
        except Exception as exc:
            return CampaignManifestSnapshot(
                manifest_source="unknown",
                manifest_loaded_success=False,
                manifest_error_kind=f"manifest_snapshot_error={type(exc).__name__}",
            )
        if isinstance(snapshot, CampaignManifestSnapshot):
            return snapshot
        return CampaignManifestSnapshot()

    @staticmethod
    def _apply_manifest_snapshot_summary(
        summary: CollectorRunSummary,
        manifest_snapshot: CampaignManifestSnapshot,
    ) -> None:
        summary.manifest_loaded_success = bool(manifest_snapshot.manifest_loaded_success)
        summary.manifest_source = manifest_snapshot.manifest_source
        summary.manifest_campaign_seen_count = int(manifest_snapshot.manifest_campaign_count or 0)
        summary.manifest_load_duration_ms = int(manifest_snapshot.manifest_load_duration_ms or 0)
        summary.manifest_parse_duration_ms = int(manifest_snapshot.manifest_parse_duration_ms or 0)

    @staticmethod
    def _apply_outcome_preflight_summary(summary: CollectorRunSummary, outcome: PromoOutcome) -> None:
        metadata = outcome.metadata
        if outcome.drawer_opened is True:
            summary.opened_drawer_count += 1
        manifest_decision = str(getattr(metadata, "manifest_decision", "") or "")
        manifest_match_confidence = str(getattr(metadata, "manifest_match_confidence", "") or "none")
        manifest_status = str(getattr(metadata, "manifest_status", "") or "unknown")
        manifest_downloadability = str(getattr(metadata, "manifest_downloadability", "") or "unknown")
        if manifest_match_confidence == "high":
            summary.manifest_timeline_match_count += 1
        elif manifest_match_confidence in {"medium", "low"}:
            summary.manifest_match_low_confidence_count += 1
        elif manifest_decision == "drawer_required_no_manifest":
            summary.manifest_missing_for_card_count += 1
        if manifest_status != "unknown":
            summary.manifest_status_classified_count += 1
        if manifest_downloadability != "unknown":
            summary.manifest_downloadability_classified_count += 1
        if manifest_decision == "drawer_avoid_manifest_non_materializable":
            summary.manifest_drawer_avoid_count += 1
            summary.drawer_open_avoided_count += 1
            summary.heavy_flow_avoided_count += 1
            summary.estimated_heavy_flow_avoided_count += 1
        elif manifest_decision:
            summary.manifest_drawer_required_count += 1
            if manifest_decision in {
                "drawer_required_manifest_unknown",
                "drawer_required_no_manifest",
            }:
                summary.manifest_unknown_full_flow_count += 1
            if manifest_decision == "drawer_required_manifest_low_confidence":
                summary.manifest_low_confidence_full_flow_count += 1
        summary.manifest_match_duration_ms += int(getattr(metadata, "manifest_match_duration_ms", 0) or 0)
        decision = str(outcome.timeline_classification_decision or "")
        if decision == "timeline_non_materializable_expected":
            summary.drawer_open_avoided_count += 1
            summary.timeline_non_materializable_count += 1
            summary.heavy_flow_avoided_count += 1
            summary.estimated_heavy_flow_avoided_count += 1
            _increment_count(summary.timeline_skip_reason_counts, str(outcome.drawer_skip_reason or "unknown"))
        elif decision in {"drawer_required", "unknown_full_flow"} and manifest_decision != "drawer_avoid_manifest_non_materializable":
            summary.drawer_open_required_count += 1
            if decision == "unknown_full_flow":
                summary.timeline_unknown_full_flow_count += 1
            _increment_count(summary.drawer_fallback_reason_counts, str(outcome.fallback_to_full_flow_reason or "unknown"))
        if str(outcome.timeline_status or "unknown") != "unknown":
            summary.timeline_status_classified_count += 1
        _increment_count(summary.timeline_evidence_confidence_counts, str(outcome.timeline_status_confidence or "low"))
        if getattr(metadata, "early_preflight_decision", None):
            summary.shallow_status_checked_count += 1
        if bool(getattr(outcome, "heavy_flow_required", False)):
            summary.deep_workbook_flow_count += 1
        if getattr(outcome, "non_materializable_reason", None):
            summary.non_materializable_expected_count += 1
        if getattr(outcome, "non_materializable_reason", None) == "ended_without_download":
            summary.early_ended_no_download_count += 1
        if getattr(outcome, "early_preflight_decision", None) == "early_non_materializable":
            summary.early_non_materializable_count += 1
            summary.heavy_flow_avoided_count += 1
            summary.estimated_heavy_flow_avoided_count += 1
        if bool(getattr(outcome, "heavy_flow_required", False)):
            if str(getattr(metadata, "ui_status", "") or "") == "unknown":
                summary.unknown_status_full_flow_count += 1
            if str(getattr(metadata, "download_action_state", "") or "") == "available":
                summary.active_downloadable_full_flow_count += 1
        if outcome.generate_screen_attempted:
            summary.generate_screen_attempt_count += 1
        if outcome.download_attempted:
            summary.download_attempt_count += 1
        summary.early_preflight_duration_ms += int(outcome.early_preflight_duration_ms or 0)
        summary.deep_flow_duration_ms += int(outcome.deep_flow_duration_ms or 0)
        summary.timeline_shallow_duration_ms += int(outcome.timeline_shallow_duration_ms or 0)
        summary.drawer_open_duration_ms += int(outcome.drawer_open_duration_ms or 0)

    def _timeline_non_materializable(
        self,
        *,
        run_dir: Path,
        candidate: TimelineCandidate,
        request: PromoXlsxCollectorRequest,
        timeline_decision: dict[str, Any],
        timeline_shallow_duration_ms: int,
    ) -> PromoOutcome:
        promo_folder = run_dir / "promos" / build_promo_folder_name(None, None, candidate.title)
        promo_folder.mkdir(parents=True, exist_ok=True)
        timeline_path = promo_folder / "timeline_card.json"
        timeline_payload = _timeline_candidate_diagnostic(candidate, timeline_decision, drawer_opened=False)
        self._write_json(timeline_path, timeline_payload)
        promo_status = candidate.timeline_status_raw_labels[0] if candidate.timeline_status_raw_labels else None
        promo_start_at, promo_end_at, confidence = parse_period_text(
            candidate.short_period_text or "",
            reference_year=datetime.now(BUSINESS_TIMEZONE).year,
        )
        metadata = PromoMetadata(
            collected_at=_now_iso(),
            trace_run_dir=request.output_root,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            calendar_url=self._driver.current_url(),
            promo_id=None,
            period_id=None,
            promo_title=candidate.title,
            promo_period_text=candidate.short_period_text or "",
            promo_start_at=promo_start_at,
            promo_end_at=promo_end_at,
            period_parse_confidence=confidence,
            temporal_classification=classify_temporal_status(promo_status),
            promo_status=promo_status,
            promo_status_text=None,
            eligible_count=candidate.timeline_goods_count,
            participating_count=None,
            excluded_count=None,
            export_kind=None,
            original_suggested_filename=None,
            saved_filename=None,
            saved_path=None,
            ui_status=candidate.timeline_status,
            ui_status_confidence=candidate.timeline_status_confidence,
            ui_status_raw_labels=list(candidate.timeline_status_raw_labels),
            download_action_state="unknown",
            download_action_evidence="timeline_shallow_no_drawer",
            status_evidence_sources=list(candidate.timeline_evidence_sources),
            ui_loaded_success=False,
            campaign_identity_match=False,
            collector_ui_schema_version=COLLECTOR_UI_SCHEMA_VERSION,
            early_preflight_decision=str(timeline_decision.get("timeline_classification_decision") or ""),
            heavy_flow_required=False,
            heavy_flow_reason="timeline_high_confidence_non_materializable",
            non_materializable_reason=str(timeline_decision.get("non_materializable_reason") or ""),
            fallback_to_full_flow_reason=None,
            collector_preflight_schema_version=COLLECTOR_PREFLIGHT_SCHEMA_VERSION,
            timeline_status=candidate.timeline_status,
            timeline_status_confidence=candidate.timeline_status_confidence,
            timeline_status_raw_labels=list(candidate.timeline_status_raw_labels),
            timeline_evidence_sources=list(candidate.timeline_evidence_sources),
            timeline_period_text=candidate.short_period_text,
            timeline_goods_count=candidate.timeline_goods_count,
            timeline_autoaction_marker=candidate.timeline_autoaction_marker,
            timeline_classification_decision="timeline_non_materializable_expected",
            drawer_opened=False,
            drawer_open_reason=None,
            drawer_skip_reason=str(timeline_decision.get("drawer_skip_reason") or "timeline_non_materializable"),
            timeline_classifier_schema_version=TIMELINE_CLASSIFIER_SCHEMA_VERSION,
        )
        metadata_path = promo_folder / "metadata.json"
        self._write_json(metadata_path, asdict(metadata))
        return PromoOutcome(
            promo_title=candidate.title,
            timeline_block_index=candidate.index,
            timeline_short_period_text=candidate.short_period_text,
            timeline_preliminary_classification=candidate.preliminary_classification,
            status="skipped_past",
            promo_id=None,
            period_id=None,
            promo_folder=str(promo_folder),
            blocker=None,
            metadata=metadata,
            metadata_path=str(metadata_path),
            early_preflight_decision="timeline_non_materializable_expected",
            heavy_flow_required=False,
            heavy_flow_reason="timeline_high_confidence_non_materializable",
            non_materializable_reason=str(timeline_decision.get("non_materializable_reason") or ""),
            fallback_to_full_flow_reason=None,
            timeline_shallow_duration_ms=timeline_shallow_duration_ms,
            **_timeline_outcome_fields(
                candidate,
                timeline_decision,
                drawer_opened=False,
                drawer_open_duration_ms=0,
            ),
        )

    def _manifest_non_materializable(
        self,
        *,
        run_dir: Path,
        candidate: TimelineCandidate,
        request: PromoXlsxCollectorRequest,
        timeline_decision: dict[str, Any],
        manifest_decision: dict[str, Any],
        timeline_shallow_duration_ms: int,
        manifest_match_duration_ms: int,
    ) -> PromoOutcome:
        manifest_item = manifest_decision.get("manifest_campaign")
        promo_id = getattr(manifest_item, "promo_id", None)
        promo_folder = run_dir / "promos" / build_promo_folder_name(promo_id, None, candidate.title)
        promo_folder.mkdir(parents=True, exist_ok=True)
        manifest_path = promo_folder / "manifest_card.json"
        manifest_payload = _manifest_candidate_diagnostic(
            candidate=candidate,
            manifest_decision=manifest_decision,
            drawer_opened=False,
        )
        self._write_json(manifest_path, manifest_payload)
        promo_title = str(getattr(manifest_item, "title", None) or candidate.title)
        promo_period_text = str(
            getattr(manifest_item, "period_text", None)
            or candidate.short_period_text
            or ""
        )
        promo_start_at = getattr(manifest_item, "start_at", None)
        promo_end_at = getattr(manifest_item, "end_at", None)
        promo_status = str(getattr(manifest_item, "lifecycle_status", "") or "ended")
        metadata = PromoMetadata(
            collected_at=_now_iso(),
            trace_run_dir=request.output_root,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            calendar_url=self._driver.current_url(),
            promo_id=promo_id,
            period_id=None,
            promo_title=promo_title,
            promo_period_text=promo_period_text,
            promo_start_at=promo_start_at,
            promo_end_at=promo_end_at,
            period_parse_confidence="high" if promo_start_at and promo_end_at else "low",
            temporal_classification="past",
            promo_status=promo_status,
            promo_status_text=str(getattr(manifest_item, "participation_status", "") or "") or None,
            eligible_count=getattr(manifest_item, "goods_count", None),
            participating_count=None,
            excluded_count=None,
            export_kind=None,
            original_suggested_filename=None,
            saved_filename=None,
            saved_path=None,
            ui_status="unknown",
            ui_status_confidence="low",
            ui_status_raw_labels=[],
            download_action_state="unknown",
            download_action_evidence="manifest_non_materializable_no_drawer",
            status_evidence_sources=list(getattr(manifest_item, "evidence_sources", []) or []),
            ui_loaded_success=False,
            campaign_identity_match=False,
            collector_ui_schema_version=COLLECTOR_UI_SCHEMA_VERSION,
            early_preflight_decision="manifest_non_materializable_expected",
            heavy_flow_required=False,
            heavy_flow_reason="manifest_high_confidence_non_materializable",
            non_materializable_reason=str(manifest_decision.get("non_materializable_reason") or ""),
            fallback_to_full_flow_reason=None,
            collector_preflight_schema_version=COLLECTOR_PREFLIGHT_SCHEMA_VERSION,
            timeline_status=candidate.timeline_status,
            timeline_status_confidence=candidate.timeline_status_confidence,
            timeline_status_raw_labels=list(candidate.timeline_status_raw_labels),
            timeline_evidence_sources=list(candidate.timeline_evidence_sources),
            timeline_period_text=candidate.short_period_text,
            timeline_goods_count=candidate.timeline_goods_count,
            timeline_autoaction_marker=candidate.timeline_autoaction_marker,
            timeline_classification_decision=timeline_decision.get("timeline_classification_decision"),
            drawer_opened=False,
            drawer_open_reason=None,
            drawer_skip_reason=str(manifest_decision.get("drawer_skip_reason") or "manifest_non_materializable"),
            timeline_classifier_schema_version=TIMELINE_CLASSIFIER_SCHEMA_VERSION,
            **_manifest_metadata_fields(
                manifest_decision.get("manifest_snapshot"),
                {
                    **manifest_decision,
                    "manifest_match_duration_ms": manifest_match_duration_ms,
                },
            ),
        )
        metadata_path = promo_folder / "metadata.json"
        self._write_json(metadata_path, asdict(metadata))
        return PromoOutcome(
            promo_title=promo_title,
            timeline_block_index=candidate.index,
            timeline_short_period_text=candidate.short_period_text,
            timeline_preliminary_classification=candidate.preliminary_classification,
            status="skipped_past",
            promo_id=promo_id,
            period_id=None,
            promo_folder=str(promo_folder),
            blocker=None,
            metadata=metadata,
            metadata_path=str(metadata_path),
            early_preflight_decision="manifest_non_materializable_expected",
            heavy_flow_required=False,
            heavy_flow_reason="manifest_high_confidence_non_materializable",
            non_materializable_reason=str(manifest_decision.get("non_materializable_reason") or ""),
            fallback_to_full_flow_reason=None,
            timeline_shallow_duration_ms=timeline_shallow_duration_ms,
            **_timeline_outcome_fields(
                candidate,
                timeline_decision,
                drawer_opened=False,
                drawer_open_duration_ms=0,
            ),
        )

    def _resolve_reusable_archive_record(
        self,
        *,
        request: PromoXlsxCollectorRequest,
        card: PromoCardData,
    ):
        archive_root_text = str(request.archive_root or "").strip()
        if not archive_root_text:
            return None
        archive_root = Path(archive_root_text)
        if not archive_root.exists():
            return None
        live_metadata = build_metadata(
            card=card,
            trace_run_dir=request.output_root,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            export_kind=None,
            download=None,
            workbook=None,
        )
        return resolve_reusable_campaign(
            archive_root=archive_root,
            metadata=live_metadata,
        )

    def _reused_from_archive(
        self,
        *,
        run_dir: Path,
        candidate: TimelineCandidate,
        card: PromoCardData,
        request: PromoXlsxCollectorRequest,
        record,
        preflight: dict[str, Any] | None = None,
        timeline_decision: dict[str, Any] | None = None,
        manifest_snapshot: CampaignManifestSnapshot | None = None,
        manifest_decision: dict[str, Any] | None = None,
        early_preflight_duration_ms: int = 0,
        deep_flow_duration_ms: int = 0,
        drawer_open_duration_ms: int = 0,
    ) -> PromoOutcome:
        promo_folder = run_dir / "promos" / build_promo_folder_name(
            card.promo_id,
            record.metadata.period_id,
            card.promo_title,
        )
        promo_folder.mkdir(parents=True, exist_ok=True)
        card_path = promo_folder / "card.json"
        card_png = promo_folder / "card.png"
        self._write_json(card_path, asdict(card))
        shutil.copy2(card.state_snapshot.screenshot, card_png)

        inspection = None
        inspection_path: Path | None = None
        if record.workbook_inspection_path and Path(record.workbook_inspection_path).exists():
            inspection = WorkbookInspection(
                **json.loads(Path(record.workbook_inspection_path).read_text(encoding="utf-8"))
            )
            inspection_path = promo_folder / "workbook_inspection.json"
            inspection_path.write_text(
                json.dumps(asdict(inspection), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        workbook_path = Path(str(record.workbook_path))
        metadata = build_metadata(
            card=card,
            trace_run_dir=request.output_root,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            export_kind=record.metadata.export_kind,
            download=DownloadArtifact(
                original_suggested_filename=(
                    record.metadata.original_suggested_filename
                    or record.metadata.saved_filename
                    or workbook_path.name
                ),
                saved_filename=workbook_path.name,
                saved_path=str(workbook_path),
                period_id=record.metadata.period_id,
            ),
            workbook=inspection,
            preflight=preflight,
            timeline_candidate=candidate,
            timeline_decision=timeline_decision,
            drawer_opened=True,
            manifest_snapshot=manifest_snapshot,
            manifest_decision=manifest_decision,
        )
        metadata_path = promo_folder / "metadata.json"
        self._write_json(metadata_path, asdict(metadata))
        (promo_folder / "archive_reuse.json").write_text(
            json.dumps(
                {
                    "archive_key": record.archive_key,
                    "archive_dir": record.archive_dir,
                    "reused_workbook_path": str(workbook_path),
                    "downloaded_at": record.downloaded_at,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        drawer_reset = self._driver.reset_drawer(f"after_archive_reuse__{slugify(card.promo_title)}")
        return PromoOutcome(
            promo_title=card.promo_title,
            timeline_block_index=candidate.index,
            timeline_short_period_text=candidate.short_period_text,
            timeline_preliminary_classification=candidate.preliminary_classification,
            status="reused_archive",
            promo_id=card.promo_id,
            period_id=record.metadata.period_id,
            promo_folder=str(promo_folder),
            blocker=None,
            metadata=metadata,
            card_path=str(card_path),
            metadata_path=str(metadata_path),
            workbook_inspection_path=str(inspection_path) if inspection_path is not None else record.workbook_inspection_path,
            saved_path=str(workbook_path),
            original_suggested_filename=metadata.original_suggested_filename,
            export_kind=metadata.export_kind,
            drawer_reset=drawer_reset,
            early_preflight_duration_ms=early_preflight_duration_ms,
            deep_flow_duration_ms=deep_flow_duration_ms,
            **_preflight_outcome_fields(preflight),
            **_timeline_outcome_fields(
                candidate,
                timeline_decision,
                drawer_opened=True,
                drawer_open_duration_ms=drawer_open_duration_ms,
            ),
        )

    def _blocked_after_card(
        self,
        *,
        run_dir: Path,
        candidate: TimelineCandidate,
        card: PromoCardData,
        blocker: str,
        request: PromoXlsxCollectorRequest,
        preflight: dict[str, Any] | None = None,
        timeline_decision: dict[str, Any] | None = None,
        manifest_snapshot: CampaignManifestSnapshot | None = None,
        manifest_decision: dict[str, Any] | None = None,
        early_preflight_duration_ms: int = 0,
        deep_flow_duration_ms: int = 0,
        drawer_open_duration_ms: int = 0,
        generate_screen_attempted: bool = False,
    ) -> PromoOutcome:
        promo_folder = run_dir / "promos" / build_promo_folder_name(card.promo_id, None, card.promo_title)
        promo_folder.mkdir(parents=True, exist_ok=True)
        card_path = promo_folder / "card.json"
        card_png = promo_folder / "card.png"
        failure_path = promo_folder / "failure.json"
        failure_png = promo_folder / "failure.png"
        shutil.copy2(card.state_snapshot.screenshot, card_png)
        shutil.copy2(card.state_snapshot.screenshot, failure_png)
        self._write_json(card_path, asdict(card))
        self._write_json(failure_path, {"status": "blocked_after_card", "blocker": blocker, "card": asdict(card)})
        metadata = build_metadata(
            card=card,
            trace_run_dir=request.output_root,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            export_kind=None,
            download=None,
            workbook=None,
            preflight=preflight,
            timeline_candidate=candidate,
            timeline_decision=timeline_decision,
            drawer_opened=True,
            manifest_snapshot=manifest_snapshot,
            manifest_decision=manifest_decision,
        )
        metadata_path = promo_folder / "metadata.json"
        self._write_json(metadata_path, asdict(metadata))
        drawer_reset = self._driver.reset_drawer(f"blocked_after_card__{slugify(card.promo_title)}")
        return PromoOutcome(
            promo_title=card.promo_title,
            timeline_block_index=candidate.index,
            timeline_short_period_text=candidate.short_period_text,
            timeline_preliminary_classification=candidate.preliminary_classification,
            status="blocked_after_card",
            promo_id=card.promo_id,
            period_id=None,
            promo_folder=str(promo_folder),
            blocker=blocker,
            metadata=metadata,
            card_path=str(card_path),
            metadata_path=str(metadata_path),
            drawer_reset=drawer_reset,
            early_preflight_duration_ms=early_preflight_duration_ms,
            deep_flow_duration_ms=deep_flow_duration_ms,
            generate_screen_attempted=generate_screen_attempted,
            **_preflight_outcome_fields(preflight),
            **_timeline_outcome_fields(
                candidate,
                timeline_decision,
                drawer_opened=True,
                drawer_open_duration_ms=drawer_open_duration_ms,
            ),
        )

    def _blocked_before_download(
        self,
        *,
        run_dir: Path,
        candidate: TimelineCandidate,
        card: PromoCardData,
        request: PromoXlsxCollectorRequest,
        generate_state: CollectorStateSnapshot,
        blocker: str,
        preflight: dict[str, Any] | None = None,
        timeline_decision: dict[str, Any] | None = None,
        manifest_snapshot: CampaignManifestSnapshot | None = None,
        manifest_decision: dict[str, Any] | None = None,
        early_preflight_duration_ms: int = 0,
        deep_flow_duration_ms: int = 0,
        drawer_open_duration_ms: int = 0,
        generate_screen_attempted: bool = False,
        download_attempted: bool = False,
    ) -> PromoOutcome:
        promo_folder = run_dir / "promos" / build_promo_folder_name(card.promo_id, None, card.promo_title)
        promo_folder.mkdir(parents=True, exist_ok=True)
        card_path = promo_folder / "card.json"
        card_png = promo_folder / "card.png"
        failure_path = promo_folder / "failure.json"
        failure_png = promo_folder / "failure.png"
        shutil.copy2(card.state_snapshot.screenshot, card_png)
        shutil.copy2(generate_state.screenshot, failure_png)
        self._write_json(card_path, asdict(card))
        self._write_json(
            failure_path,
            {
                "status": "blocked_before_download",
                "blocker": blocker,
                "card": asdict(card),
                "generate_state": asdict(generate_state),
            },
        )
        metadata = build_metadata(
            card=card,
            trace_run_dir=request.output_root,
            source_tab=request.source_tab,
            source_filter_code=request.source_filter_code,
            export_kind=None,
            download=None,
            workbook=None,
            preflight=preflight,
            timeline_candidate=candidate,
            timeline_decision=timeline_decision,
            drawer_opened=True,
            manifest_snapshot=manifest_snapshot,
            manifest_decision=manifest_decision,
        )
        metadata_path = promo_folder / "metadata.json"
        self._write_json(metadata_path, asdict(metadata))
        drawer_reset = self._driver.reset_drawer(f"blocked_before_download__{slugify(card.promo_title)}")
        return PromoOutcome(
            promo_title=card.promo_title,
            timeline_block_index=candidate.index,
            timeline_short_period_text=candidate.short_period_text,
            timeline_preliminary_classification=candidate.preliminary_classification,
            status="blocked_before_download",
            promo_id=card.promo_id,
            period_id=None,
            promo_folder=str(promo_folder),
            blocker=blocker,
            metadata=metadata,
            card_path=str(card_path),
            metadata_path=str(metadata_path),
            drawer_reset=drawer_reset,
            early_preflight_duration_ms=early_preflight_duration_ms,
            deep_flow_duration_ms=deep_flow_duration_ms,
            generate_screen_attempted=generate_screen_attempted,
            download_attempted=download_attempted,
            **_preflight_outcome_fields(preflight),
            **_timeline_outcome_fields(
                candidate,
                timeline_decision,
                drawer_opened=True,
                drawer_open_duration_ms=drawer_open_duration_ms,
            ),
        )

    @staticmethod
    def _write_json(path: Path, payload: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_timeline_candidate(block: TimelineBlockSnapshot) -> TimelineCandidate | None:
    lines = _clean_lines(block.raw_text)
    if not lines:
        return None
    filtered = [line for line in lines if line not in {"Автоакция.", "Акция."}]
    if not filtered:
        return None
    title = filtered[0]
    short_period_text = next((line for line in filtered[1:] if _looks_like_short_period(line)), None)
    preliminary = classify_timeline_period(short_period_text)
    evidence = classify_timeline_status(lines=filtered, title=title, period_text=short_period_text)
    return TimelineCandidate(
        index=block.index,
        title=title,
        short_period_text=short_period_text,
        preliminary_classification=preliminary,
        raw_text=block.raw_text,
        timeline_status=evidence["timeline_status"],
        timeline_status_confidence=evidence["timeline_status_confidence"],
        timeline_status_raw_labels=evidence["timeline_status_raw_labels"],
        timeline_evidence_sources=evidence["timeline_evidence_sources"],
        timeline_goods_count=evidence["timeline_goods_count"],
        timeline_autoaction_marker=evidence["timeline_autoaction_marker"],
        timeline_classifier_schema_version=TIMELINE_CLASSIFIER_SCHEMA_VERSION,
    )


def classify_timeline_status(*, lines: list[str], title: str, period_text: str | None) -> dict[str, Any]:
    normalized_lines = [_normalize_label(line) for line in lines]
    status_labels = _status_raw_labels(lines)
    sources: list[str] = []
    if title.strip():
        sources.append("timeline_title")
    if period_text:
        sources.append("timeline_period_text")
    status = "unknown"
    confidence = "low"
    if any(_label_contains(normalized, _ENDED_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "ended"
        confidence = "high" if title.strip() and period_text and status_labels else "medium"
        sources.append("timeline_status_label")
    elif any(_label_contains(normalized, _ACTIVE_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "active"
        confidence = "high" if title.strip() and status_labels else "medium"
        sources.append("timeline_status_label")
    elif any(_label_contains(normalized, _FUTURE_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "future"
        confidence = "high" if title.strip() and status_labels else "medium"
        sources.append("timeline_status_label")
    elif any(_label_contains(normalized, _PENDING_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "pending"
        confidence = "medium"
        sources.append("timeline_status_label")

    goods_count, _, _ = extract_counts(lines)
    if goods_count is not None:
        sources.append("timeline_goods_count")
    autoaction_marker = _timeline_autoaction_marker(lines)
    if autoaction_marker:
        sources.append("timeline_autoaction_marker")

    return {
        "timeline_status": status,
        "timeline_status_confidence": confidence,
        "timeline_status_raw_labels": status_labels,
        "timeline_evidence_sources": sorted(set(sources)),
        "timeline_goods_count": goods_count,
        "timeline_autoaction_marker": autoaction_marker,
    }


def classify_timeline_preflight(candidate: TimelineCandidate) -> dict[str, Any]:
    sources = set(candidate.timeline_evidence_sources or [])
    status = str(candidate.timeline_status or "unknown")
    confidence = str(candidate.timeline_status_confidence or "low")
    title_present = bool(str(candidate.title or "").strip())
    period_present = bool(str(candidate.short_period_text or "").strip())
    has_status_evidence = "timeline_status_label" in sources
    has_identity_evidence = title_present and "timeline_title" in sources
    has_period_evidence = period_present and "timeline_period_text" in sources
    high_confidence_non_materializable = (
        status in _TIMELINE_NON_MATERIALIZABLE_STATUSES
        and confidence == "high"
        and has_status_evidence
        and has_identity_evidence
        and has_period_evidence
    )
    if high_confidence_non_materializable:
        return {
            "timeline_classification_decision": "timeline_non_materializable_expected",
            "drawer_open_reason": None,
            "drawer_skip_reason": "timeline_ended_non_materializable",
            "non_materializable_reason": "ended_without_download",
            "fallback_to_full_flow_reason": None,
            "timeline_classifier_schema_version": TIMELINE_CLASSIFIER_SCHEMA_VERSION,
        }

    if status in {"active"} and confidence == "high" and has_status_evidence:
        return _timeline_drawer_decision("drawer_required", "active_or_materializable")
    if status in {"future", "pending"} and confidence == "high" and has_status_evidence:
        return _timeline_drawer_decision("drawer_required", "future_archive_capture_policy")
    if status == "unknown":
        return _timeline_drawer_decision("unknown_full_flow", "timeline_status_unknown")
    if confidence != "high":
        return _timeline_drawer_decision("unknown_full_flow", "low_confidence_timeline_status")
    if not has_status_evidence:
        return _timeline_drawer_decision("unknown_full_flow", "missing_timeline_status_evidence")
    if not has_identity_evidence:
        return _timeline_drawer_decision("unknown_full_flow", "missing_timeline_identity_evidence")
    if not has_period_evidence:
        return _timeline_drawer_decision("unknown_full_flow", "missing_timeline_period_evidence")
    return _timeline_drawer_decision("unknown_full_flow", "conservative_timeline_full_flow")


def classify_manifest_preflight(
    *,
    candidate: TimelineCandidate,
    manifest_snapshot: CampaignManifestSnapshot,
) -> dict[str, Any]:
    base = {
        "manifest_snapshot": manifest_snapshot,
        "manifest_campaign": None,
        "manifest_match_confidence": "none",
        "manifest_decision": "drawer_required_no_manifest",
        "drawer_skip_reason": None,
        "drawer_required_reason": "manifest_not_loaded",
        "non_materializable_reason": None,
        "evidence_sources": [],
        "manifest_match_duration_ms": 0,
    }
    if not manifest_snapshot.manifest_loaded_success or not manifest_snapshot.campaigns:
        return base

    matches = [
        _match_manifest_campaign(candidate=candidate, item=item)
        for item in manifest_snapshot.campaigns
    ]
    matches = [match for match in matches if match["manifest_match_confidence"] != "none"]
    if not matches:
        return {
            **base,
            "manifest_decision": "drawer_required_no_manifest",
            "drawer_required_reason": "manifest_missing_for_card",
        }

    confidence_rank = {"high": 3, "medium": 2, "low": 1, "none": 0}
    matches.sort(
        key=lambda match: (
            confidence_rank.get(str(match["manifest_match_confidence"]), 0),
            len(match.get("evidence_sources") or []),
        ),
        reverse=True,
    )
    best = matches[0]
    same_rank_count = sum(
        1
        for match in matches
        if confidence_rank.get(str(match["manifest_match_confidence"]), 0)
        == confidence_rank.get(str(best["manifest_match_confidence"]), 0)
    )
    if same_rank_count > 1:
        return {
            **base,
            **best,
            "manifest_match_confidence": "low",
            "manifest_decision": "drawer_required_manifest_low_confidence",
            "drawer_required_reason": "ambiguous_manifest_match",
        }

    item = best["manifest_campaign"]
    match_confidence = str(best["manifest_match_confidence"] or "none")
    status = str(getattr(item, "lifecycle_status", "") or "unknown")
    status_confidence = str(getattr(item, "lifecycle_status_confidence", "") or "low")
    downloadability = str(getattr(item, "downloadability", "") or "unknown")
    downloadability_confidence = str(getattr(item, "downloadability_confidence", "") or "low")
    high_confidence_non_materializable = (
        match_confidence == "high"
        and status in _MANIFEST_NON_MATERIALIZABLE_STATUSES
        and status_confidence == "high"
        and downloadability == "not_available"
        and downloadability_confidence == "high"
    )
    if high_confidence_non_materializable:
        return {
            **base,
            **best,
            "manifest_decision": "drawer_avoid_manifest_non_materializable",
            "drawer_skip_reason": "manifest_ended_non_materializable",
            "drawer_required_reason": None,
            "non_materializable_reason": "ended_without_download",
        }

    if match_confidence != "high":
        return {
            **base,
            **best,
            "manifest_decision": "drawer_required_manifest_low_confidence",
            "drawer_required_reason": "low_confidence_manifest_match",
        }
    if status in {"active", "future", "pending"} or downloadability == "available":
        return {
            **base,
            **best,
            "manifest_decision": "drawer_required_manifest_active_or_downloadable",
            "drawer_required_reason": "manifest_active_or_downloadable",
        }
    return {
        **base,
        **best,
        "manifest_decision": "drawer_required_manifest_unknown",
        "drawer_required_reason": "manifest_status_or_downloadability_unknown",
    }


def _match_manifest_campaign(
    *,
    candidate: TimelineCandidate,
    item: CampaignManifestItem,
) -> dict[str, Any]:
    evidence_sources: list[str] = []
    candidate_title_key = _title_match_key(candidate.title)
    manifest_title_key = _title_match_key(item.title)
    title_matches = bool(candidate_title_key and candidate_title_key == manifest_title_key)
    if title_matches:
        evidence_sources.append("manifest_title_exact")
    candidate_period_key = _period_match_key(candidate.short_period_text)
    manifest_period_key = _period_match_key(item.period_text)
    period_matches = bool(candidate_period_key and candidate_period_key == manifest_period_key)
    if period_matches:
        evidence_sources.append("manifest_period_exact")
    confidence = "none"
    if title_matches and period_matches:
        confidence = "high"
    elif title_matches:
        confidence = "medium"
    elif period_matches and _title_token_overlap(candidate.title, item.title) >= 2:
        confidence = "low"
        evidence_sources.append("manifest_title_token_overlap")
    return {
        "manifest_campaign": item,
        "manifest_match_confidence": confidence,
        "evidence_sources": evidence_sources,
    }


def _timeline_drawer_decision(decision: str, reason: str) -> dict[str, Any]:
    return {
        "timeline_classification_decision": decision,
        "drawer_open_reason": reason,
        "drawer_skip_reason": None,
        "non_materializable_reason": None,
        "fallback_to_full_flow_reason": reason,
        "timeline_classifier_schema_version": TIMELINE_CLASSIFIER_SCHEMA_VERSION,
    }


def classify_timeline_period(period_text: str | None) -> TemporalClassification:
    if not period_text:
        return "ambiguous"
    if any(keyword in period_text for keyword in ("мая", "апреля", "марта", "февраля", "января", "декабря")):
        return "ambiguous"
    return "ambiguous"


def extract_card_data(
    *,
    snapshot: CollectorStateSnapshot,
    fallback_title: str,
    source_tab: str,
    source_filter_code: str,
    reference_year: int,
) -> PromoCardData:
    del source_tab
    del source_filter_code
    lines = _clean_lines(snapshot.body_excerpt)
    promo_title = fallback_title
    matching_indexes = [idx for idx, line in enumerate(lines) if line == fallback_title]
    title_matched = bool(matching_indexes) or _title_matches_card(fallback_title, lines)
    if matching_indexes:
        title_index = matching_indexes[-1]
    else:
        title_index = next((idx for idx, line in enumerate(lines) if line == fallback_title), -1)
    if title_index < 0:
        title_index = 0
    card_lines = lines[title_index : title_index + 80]
    promo_period_text = next((line for line in card_lines[1:] if "→" in line), "")
    ui_status_evidence = classify_card_ui_status(
        card_lines=card_lines,
        snapshot=snapshot,
        title_matched=title_matched,
    )
    promo_status = (
        ui_status_evidence["ui_status_raw_labels"][0]
        if ui_status_evidence["ui_status_raw_labels"]
        else next((line for line in card_lines[1:] if line in {"Запланирована", "Акция идёт", "Завершилась"}), None)
    )
    promo_status_text = next((line for line in card_lines[1:] if line.startswith("Автоакция:")), None)
    promo_start_at, promo_end_at, confidence = parse_period_text(promo_period_text, reference_year=reference_year)
    temporal_classification = classify_temporal_status(promo_status)
    eligible_count, participating_count, excluded_count = extract_counts(card_lines)
    raw_card_excerpt = "\n".join(card_lines)
    promo_id = parse_promo_id(snapshot.url)
    return PromoCardData(
        calendar_url=snapshot.url,
        promo_id=promo_id,
        promo_title=promo_title,
        promo_period_text=promo_period_text,
        promo_start_at=promo_start_at,
        promo_end_at=promo_end_at,
        period_parse_confidence=confidence,
        temporal_classification=temporal_classification,
        temporal_confidence="high" if temporal_classification != "ambiguous" else "low",
        promo_status=promo_status,
        promo_status_text=promo_status_text,
        eligible_count=eligible_count,
        participating_count=participating_count,
        excluded_count=excluded_count,
        raw_card_excerpt=raw_card_excerpt,
        state_snapshot=snapshot,
        ui_status=ui_status_evidence["ui_status"],
        ui_status_confidence=ui_status_evidence["ui_status_confidence"],
        ui_status_raw_labels=ui_status_evidence["ui_status_raw_labels"],
        download_action_state=ui_status_evidence["download_action_state"],
        download_action_evidence=ui_status_evidence["download_action_evidence"],
        status_evidence_sources=ui_status_evidence["status_evidence_sources"],
        ui_loaded_success=ui_status_evidence["ui_loaded_success"],
        campaign_identity_match=ui_status_evidence["campaign_identity_match"],
        collector_ui_schema_version=COLLECTOR_UI_SCHEMA_VERSION,
    )


def classify_temporal_status(status: str | None) -> TemporalClassification:
    if not status:
        return "ambiguous"
    normalized = status.strip().lower()
    if "заверш" in normalized:
        return "past"
    if "заплан" in normalized:
        return "future"
    if "ид" in normalized:
        return "current"
    return "ambiguous"


def classify_card_ui_status(
    *,
    card_lines: list[str],
    snapshot: CollectorStateSnapshot,
    title_matched: bool,
) -> dict[str, Any]:
    ui_loaded_success = bool(title_matched and card_lines)
    normalized_lines = [_normalize_label(line) for line in card_lines]
    status_labels = _status_raw_labels(card_lines)
    status = "unknown"
    confidence = "low"
    sources: list[str] = []
    if any(_label_contains(normalized, _ENDED_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "ended"
        confidence = "high"
        sources.append("footer_label")
    elif any(_label_contains(normalized, _ACTIVE_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "active"
        confidence = "high"
        sources.append("footer_label")
    elif any(_label_contains(normalized, _FUTURE_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "future"
        confidence = "high"
        sources.append("footer_label")
    elif any(_label_contains(normalized, _PENDING_LABEL_KEYWORDS) for normalized in normalized_lines):
        status = "pending"
        confidence = "medium"
        sources.append("footer_label")

    if ui_loaded_success:
        sources.append("drawer_loaded")
    if title_matched:
        sources.append("title_match")

    if not ui_loaded_success:
        download_state = "ui_not_loaded"
        download_evidence = "drawer_not_loaded"
    elif snapshot.has_download:
        download_state = "available"
        download_evidence = "download_button_visible"
    elif snapshot.has_generate:
        download_state = "available"
        download_evidence = "generate_button_visible"
    elif snapshot.has_configure:
        download_state = "available"
        download_evidence = "configure_button_visible"
    else:
        download_state = "absent"
        download_evidence = "configure_generate_download_buttons_absent"
        sources.append("download_button_absent")

    return {
        "ui_status": status,
        "ui_status_confidence": confidence,
        "ui_status_raw_labels": status_labels,
        "download_action_state": download_state,
        "download_action_evidence": download_evidence,
        "status_evidence_sources": sorted(set(sources)),
        "ui_loaded_success": ui_loaded_success,
        "campaign_identity_match": bool(title_matched),
    }


def classify_collector_preflight(card: PromoCardData) -> dict[str, Any]:
    ui_status = str(card.ui_status or "unknown")
    confidence = str(card.ui_status_confidence or "low")
    download_action_state = str(card.download_action_state or "unknown")
    loaded = bool(card.ui_loaded_success)
    identity_matched = bool(card.campaign_identity_match)
    status_sources = set(card.status_evidence_sources or [])
    has_status_evidence = bool(status_sources & {"footer_label", "badge"})

    high_confidence_non_materializable = (
        ui_status == "ended"
        and confidence == "high"
        and download_action_state in {"absent", "disabled"}
        and loaded
        and identity_matched
        and has_status_evidence
    )
    if high_confidence_non_materializable:
        return {
            "early_preflight_decision": "early_non_materializable",
            "heavy_flow_required": False,
            "heavy_flow_reason": "high_confidence_non_materializable",
            "non_materializable_reason": "ended_without_download",
            "fallback_to_full_flow_reason": None,
            "collector_preflight_schema_version": COLLECTOR_PREFLIGHT_SCHEMA_VERSION,
        }

    fallback_reason = None
    if not loaded:
        fallback_reason = "ui_not_loaded"
    elif not identity_matched:
        fallback_reason = "campaign_identity_mismatch"
    elif confidence != "high":
        fallback_reason = "low_confidence_status"
    elif download_action_state in {"unknown", "ui_not_loaded"}:
        fallback_reason = "unclear_download_action"
    elif not has_status_evidence:
        fallback_reason = "missing_status_evidence"
    elif download_action_state == "available":
        fallback_reason = "download_action_available"
    else:
        fallback_reason = "conservative_full_flow"

    return {
        "early_preflight_decision": "full_flow",
        "heavy_flow_required": True,
        "heavy_flow_reason": "materializable_or_unclear",
        "non_materializable_reason": None,
        "fallback_to_full_flow_reason": fallback_reason,
        "collector_preflight_schema_version": COLLECTOR_PREFLIGHT_SCHEMA_VERSION,
    }


def parse_period_text(period_text: str, *, reference_year: int) -> tuple[str | None, str | None, str]:
    match = re.search(
        r"(?P<start_day>\d{1,2})\s+(?P<start_month>[А-Яа-яё]+)\s+(?P<start_time>\d{2}:\d{2})\s*[→-]\s*(?P<end_day>\d{1,2})\s+(?P<end_month>[А-Яа-яё]+)\s+(?P<end_time>\d{2}:\d{2})",
        period_text,
    )
    if not match:
        return None, None, "low"
    start_month = _MONTHS.get(match.group("start_month").lower())
    end_month = _MONTHS.get(match.group("end_month").lower())
    if start_month is None or end_month is None:
        return None, None, "low"
    if end_month < start_month:
        return None, None, "low"
    start_at = datetime(
        reference_year,
        start_month,
        int(match.group("start_day")),
        int(match.group("start_time").split(":")[0]),
        int(match.group("start_time").split(":")[1]),
    )
    end_at = datetime(
        reference_year,
        end_month,
        int(match.group("end_day")),
        int(match.group("end_time").split(":")[0]),
        int(match.group("end_time").split(":")[1]),
    )
    return start_at.strftime("%Y-%m-%dT%H:%M"), end_at.strftime("%Y-%m-%dT%H:%M"), "high"


def extract_counts(lines: list[str]) -> tuple[int | None, int | None, int | None]:
    text = "\n".join(lines)
    eligible_count = _extract_int(text, r"(\d+)\s+подходящих товара")
    participating_count = _extract_int(text, r"(\d+)\s+будут участвовать")
    excluded_count = _extract_int(text, r"(\d+)\s+исключено")
    participating_pair = re.search(r"Участвует\s+(\d+)\s+из\s+(\d+)\s+товаров", text)
    if participating_pair:
        participating_count = int(participating_pair.group(1))
        eligible_count = int(participating_pair.group(2))
    added_pair = re.search(r"Добавлено\s+(\d+)\s+из\s+(\d+)\s+товаров", text)
    if added_pair:
        participating_count = int(added_pair.group(1))
        eligible_count = int(added_pair.group(2))
    return eligible_count, participating_count, excluded_count


def _status_raw_labels(lines: list[str], *, limit: int = 5) -> list[str]:
    labels: list[str] = []
    for line in lines:
        normalized = _normalize_label(line)
        if not normalized:
            continue
        if (
            _label_contains(normalized, _ENDED_LABEL_KEYWORDS)
            or _label_contains(normalized, _ACTIVE_LABEL_KEYWORDS)
            or _label_contains(normalized, _FUTURE_LABEL_KEYWORDS)
            or _label_contains(normalized, _PENDING_LABEL_KEYWORDS)
        ):
            labels.append(_sanitize_label(line))
        if len(labels) >= limit:
            break
    return labels


def _label_contains(normalized: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in normalized for keyword in keywords)


def _normalize_label(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip().lower()


def _title_match_key(value: str | None) -> str:
    normalized = _normalize_label(str(value or "").replace("ё", "е"))
    normalized = re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE)
    tokens = [token for token in normalized.split() if token]
    return " ".join(tokens)


def _period_match_key(value: str | None) -> str:
    normalized = _normalize_label(str(value or "").replace("ё", "е"))
    normalized = normalized.replace("→", "-").replace("—", "-")
    normalized = re.sub(r"\s*-\s*", "-", normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def _title_token_overlap(left: str | None, right: str | None) -> int:
    left_tokens = {
        token
        for token in _title_match_key(left).split()
        if len(token) >= 4 and token not in {"акция", "автоакция"}
    }
    right_tokens = {
        token
        for token in _title_match_key(right).split()
        if len(token) >= 4 and token not in {"акция", "автоакция"}
    }
    return len(left_tokens & right_tokens)


def _sanitize_label(value: str, *, limit: int = 80) -> str:
    sanitized = re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()
    return sanitized[:limit]


def _title_matches_card(fallback_title: str, lines: list[str]) -> bool:
    title = _normalize_label(fallback_title)
    if not title:
        return False
    joined = _normalize_label(" ".join(lines[:20]))
    if title in joined:
        return True
    tokens = [
        token
        for token in re.split(r"\W+", title, flags=re.UNICODE)
        if len(token) >= 4 and token not in {"акция", "автоакция"}
    ]
    if not tokens:
        return False
    matched = sum(1 for token in tokens if token in joined)
    return matched >= max(1, min(len(tokens), 2))


def classify_export_kind(original_filename: str | None, headers: list[str]) -> ExportKind:
    normalized_name = (original_filename or "").strip().lower()
    header_set = {header.strip() for header in headers if header and header.strip()}
    has_gating = bool(_GATING_HEADERS & header_set)
    if normalized_name.startswith("товары для исключения из акции") or has_gating:
        return "exclude_list_template"
    if normalized_name.startswith("все товары подходящие для акции") and header_set:
        return "eligible_items_report"
    if _COMMON_EXPORT_HEADERS.issubset(header_set) and not has_gating:
        return "eligible_items_report"
    return "unknown"


def inspect_workbook(path: Path) -> WorkbookInspection:
    workbook = load_workbook(filename=str(path), read_only=False, data_only=False)
    try:
        sheet_names = workbook.sheetnames
        hidden_sheets = any(workbook[sheet_name].sheet_state != "visible" for sheet_name in sheet_names)
        target_sheet, header_row_index, header_summary = _find_workbook_data_sheet(workbook)
        row_count = target_sheet.max_row
        col_count = target_sheet.max_column
        formulas_present = False
        distinct_statuses: set[str] = set()
        status_column = None
        for idx, header in enumerate(header_summary, start=1):
            if header == "Статус":
                status_column = idx
                break
        non_empty_rows = 0
        for row_index, row in enumerate(target_sheet.iter_rows(), start=1):
            values = [cell.value for cell in row]
            if any(value not in (None, "") for value in values):
                non_empty_rows += 1
            if any(getattr(cell, "data_type", None) == "f" for cell in row):
                formulas_present = True
            if row_index <= header_row_index:
                continue
            if status_column is not None and len(row) >= status_column:
                value = row[status_column - 1].value
                if isinstance(value, str) and value.strip():
                    distinct_statuses.add(value.strip())
        merged_cells_present = bool(target_sheet.merged_cells.ranges)
        workbook_has_date_fields = any(
            any(token in header.lower() for token in ("дата", "date", "period", "период"))
            for header in header_summary
        )
        return WorkbookInspection(
            workbook_sheet_names=sheet_names,
            workbook_row_count=row_count,
            workbook_col_count=col_count,
            workbook_header_summary=header_summary,
            workbook_has_date_fields=workbook_has_date_fields,
            workbook_item_status_distinct_values=sorted(distinct_statuses),
            hidden_sheets=hidden_sheets,
            formulas_present=formulas_present,
            merged_cells_present=merged_cells_present,
            rough_data_completeness_summary=(
                f"non_empty_rows={non_empty_rows}; data_rows={max(row_count - 1, 0)}; "
                f"header_columns={len(header_summary)}"
            ),
        )
    finally:
        workbook.close()


def _find_workbook_data_sheet(workbook: Any) -> tuple[Any, int, list[str]]:
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_index, header_summary = _find_workbook_header_row(sheet)
        if header_summary and _COMMON_EXPORT_HEADERS.issubset(set(header_summary)):
            return sheet, row_index, header_summary
    first_sheet = workbook[workbook.sheetnames[0]]
    row_index, header_summary = _find_workbook_header_row(first_sheet)
    return first_sheet, row_index, header_summary


def _find_workbook_header_row(sheet: Any) -> tuple[int, list[str]]:
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=True), start=1):
        header_summary = [str(value).strip() for value in row if value not in (None, "")]
        header_set = set(header_summary)
        if _COMMON_EXPORT_HEADERS.issubset(header_set):
            return row_index, header_summary
    fallback = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return 1, [str(value).strip() for value in fallback if value not in (None, "")]


def build_metadata(
    *,
    card: PromoCardData,
    trace_run_dir: str,
    source_tab: str,
    source_filter_code: str,
    export_kind: ExportKind | None,
    download: DownloadArtifact | None,
    workbook: WorkbookInspection | None,
    preflight: dict[str, Any] | None = None,
    timeline_candidate: TimelineCandidate | None = None,
    timeline_decision: dict[str, Any] | None = None,
    drawer_opened: bool | None = None,
    manifest_snapshot: CampaignManifestSnapshot | None = None,
    manifest_decision: dict[str, Any] | None = None,
) -> PromoMetadata:
    preflight = preflight or {}
    timeline_decision = timeline_decision or {}
    return PromoMetadata(
        collected_at=_now_iso(),
        trace_run_dir=trace_run_dir,
        source_tab=source_tab,
        source_filter_code=source_filter_code,
        calendar_url=card.calendar_url,
        promo_id=card.promo_id,
        period_id=download.period_id if download else None,
        promo_title=card.promo_title,
        promo_period_text=card.promo_period_text,
        promo_start_at=card.promo_start_at,
        promo_end_at=card.promo_end_at,
        period_parse_confidence=card.period_parse_confidence,
        temporal_classification=card.temporal_classification,
        promo_status=card.promo_status,
        promo_status_text=card.promo_status_text,
        eligible_count=card.eligible_count,
        participating_count=card.participating_count,
        excluded_count=card.excluded_count,
        export_kind=export_kind,
        original_suggested_filename=download.original_suggested_filename if download else None,
        saved_filename=download.saved_filename if download else None,
        saved_path=download.saved_path if download else None,
        workbook_sheet_names=workbook.workbook_sheet_names if workbook else [],
        workbook_row_count=workbook.workbook_row_count if workbook else 0,
        workbook_col_count=workbook.workbook_col_count if workbook else 0,
        workbook_header_summary=workbook.workbook_header_summary if workbook else [],
        workbook_has_date_fields=workbook.workbook_has_date_fields if workbook else False,
        workbook_item_status_distinct_values=workbook.workbook_item_status_distinct_values if workbook else [],
        ui_status=card.ui_status,
        ui_status_confidence=card.ui_status_confidence,
        ui_status_raw_labels=list(card.ui_status_raw_labels),
        download_action_state=card.download_action_state,
        download_action_evidence=card.download_action_evidence,
        status_evidence_sources=list(card.status_evidence_sources),
        ui_loaded_success=card.ui_loaded_success,
        campaign_identity_match=card.campaign_identity_match,
        collector_ui_schema_version=card.collector_ui_schema_version,
        early_preflight_decision=preflight.get("early_preflight_decision"),
        heavy_flow_required=preflight.get("heavy_flow_required"),
        heavy_flow_reason=preflight.get("heavy_flow_reason"),
        non_materializable_reason=preflight.get("non_materializable_reason"),
        fallback_to_full_flow_reason=preflight.get("fallback_to_full_flow_reason"),
        collector_preflight_schema_version=str(
            preflight.get("collector_preflight_schema_version")
            or COLLECTOR_PREFLIGHT_SCHEMA_VERSION
        ),
        timeline_status=(
            timeline_candidate.timeline_status
            if timeline_candidate is not None
            else "unknown"
        ),
        timeline_status_confidence=(
            timeline_candidate.timeline_status_confidence
            if timeline_candidate is not None
            else "low"
        ),
        timeline_status_raw_labels=(
            list(timeline_candidate.timeline_status_raw_labels)
            if timeline_candidate is not None
            else []
        ),
        timeline_evidence_sources=(
            list(timeline_candidate.timeline_evidence_sources)
            if timeline_candidate is not None
            else []
        ),
        timeline_period_text=timeline_candidate.short_period_text if timeline_candidate is not None else None,
        timeline_goods_count=timeline_candidate.timeline_goods_count if timeline_candidate is not None else None,
        timeline_autoaction_marker=(
            timeline_candidate.timeline_autoaction_marker
            if timeline_candidate is not None
            else None
        ),
        timeline_classification_decision=timeline_decision.get("timeline_classification_decision"),
        drawer_opened=drawer_opened,
        drawer_open_reason=timeline_decision.get("drawer_open_reason"),
        drawer_skip_reason=timeline_decision.get("drawer_skip_reason"),
        timeline_classifier_schema_version=str(
            timeline_decision.get("timeline_classifier_schema_version")
            or TIMELINE_CLASSIFIER_SCHEMA_VERSION
        ),
        **_manifest_metadata_fields(manifest_snapshot, manifest_decision),
    )


def build_promo_folder_name(promo_id: int | None, period_id: int | None, title: str) -> str:
    promo_part = str(promo_id) if promo_id is not None else "pending"
    period_part = str(period_id) if period_id is not None else "pending"
    return f"{promo_part}__{period_part}__{slugify(title)}"


def slugify(value: str) -> str:
    normalized = re.sub(r"[^\w]+", "-", value.lower(), flags=re.UNICODE)
    normalized = normalized.strip("-")
    normalized = re.sub(r"-{2,}", "-", normalized)
    return normalized or "pending"


def parse_promo_id(url: str) -> int | None:
    match = re.search(r"[?&]action=(\d+)", url)
    return int(match.group(1)) if match else None


def _looks_like_short_period(line: str) -> bool:
    return bool(re.search(r"\d", line)) and (" - " in line or "→" in line)


def _extract_int(text: str, pattern: str) -> int | None:
    match = re.search(pattern, text)
    return int(match.group(1)) if match else None


def _clean_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


def _timeline_autoaction_marker(lines: list[str]) -> str | None:
    normalized = _normalize_label(" ".join(lines))
    if "автоматические скидки" in normalized:
        return "auto_discount"
    if "автоакц" in normalized:
        return "auto_promo"
    return None


def _now_iso() -> str:
    return datetime.now(BUSINESS_TIMEZONE).isoformat(timespec="seconds")


def _preflight_outcome_fields(preflight: dict[str, Any] | None) -> dict[str, Any]:
    preflight = preflight or {}
    return {
        "early_preflight_decision": preflight.get("early_preflight_decision"),
        "heavy_flow_required": preflight.get("heavy_flow_required"),
        "heavy_flow_reason": preflight.get("heavy_flow_reason"),
        "non_materializable_reason": preflight.get("non_materializable_reason"),
        "fallback_to_full_flow_reason": preflight.get("fallback_to_full_flow_reason"),
    }


def _timeline_outcome_fields(
    candidate: TimelineCandidate,
    timeline_decision: dict[str, Any] | None,
    *,
    drawer_opened: bool,
    drawer_open_duration_ms: int,
) -> dict[str, Any]:
    timeline_decision = timeline_decision or _timeline_drawer_decision("unknown_full_flow", "timeline_decision_missing")
    return {
        "timeline_status": candidate.timeline_status,
        "timeline_status_confidence": candidate.timeline_status_confidence,
        "timeline_classification_decision": timeline_decision.get("timeline_classification_decision"),
        "drawer_opened": drawer_opened,
        "drawer_open_reason": timeline_decision.get("drawer_open_reason"),
        "drawer_skip_reason": timeline_decision.get("drawer_skip_reason"),
        "drawer_open_duration_ms": drawer_open_duration_ms,
    }


def _manifest_metadata_fields(
    manifest_snapshot: CampaignManifestSnapshot | None,
    manifest_decision: dict[str, Any] | None,
) -> dict[str, Any]:
    manifest_snapshot = manifest_snapshot or CampaignManifestSnapshot()
    manifest_decision = manifest_decision or {}
    item = manifest_decision.get("manifest_campaign")
    evidence_sources = list(manifest_decision.get("evidence_sources") or [])
    if item is not None:
        evidence_sources.extend(list(getattr(item, "evidence_sources", []) or []))
    evidence_sources = sorted({str(source) for source in evidence_sources if str(source or "").strip()})
    return {
        "manifest_schema_version": str(
            getattr(manifest_snapshot, "manifest_schema_version", None)
            or MANIFEST_SCHEMA_VERSION
        ),
        "manifest_source": str(getattr(manifest_snapshot, "manifest_source", None) or "none"),
        "manifest_loaded_success": bool(getattr(manifest_snapshot, "manifest_loaded_success", False)),
        "manifest_campaign_count": int(getattr(manifest_snapshot, "manifest_campaign_count", 0) or 0),
        "manifest_campaign_id": getattr(item, "campaign_id", None),
        "manifest_promo_id": getattr(item, "promo_id", None),
        "manifest_title": _sanitize_label(str(getattr(item, "title", "") or ""), limit=160) or None,
        "manifest_period_text": _sanitize_label(str(getattr(item, "period_text", "") or ""), limit=120) or None,
        "manifest_status": str(getattr(item, "lifecycle_status", None) or "unknown"),
        "manifest_status_confidence": str(getattr(item, "lifecycle_status_confidence", None) or "low"),
        "manifest_downloadability": str(getattr(item, "downloadability", None) or "unknown"),
        "manifest_downloadability_confidence": str(
            getattr(item, "downloadability_confidence", None)
            or "low"
        ),
        "manifest_match_confidence": str(
            manifest_decision.get("manifest_match_confidence")
            or "none"
        ),
        "manifest_decision": manifest_decision.get("manifest_decision"),
        "manifest_evidence_sources": evidence_sources,
        "manifest_loaded_at": getattr(manifest_snapshot, "manifest_loaded_at", None),
        "manifest_drawer_skip_reason": manifest_decision.get("drawer_skip_reason"),
        "manifest_drawer_required_reason": manifest_decision.get("drawer_required_reason"),
        "manifest_participation_status": getattr(item, "participation_status", None),
        "manifest_match_duration_ms": int(manifest_decision.get("manifest_match_duration_ms", 0) or 0),
    }


def _manifest_candidate_diagnostic(
    *,
    candidate: TimelineCandidate,
    manifest_decision: dict[str, Any],
    drawer_opened: bool,
) -> dict[str, Any]:
    item = manifest_decision.get("manifest_campaign")
    return {
        "manifest_classifier_schema_version": MANIFEST_SCHEMA_VERSION,
        "timeline_block_index": candidate.index,
        "timeline_title": _sanitize_label(candidate.title, limit=160),
        "timeline_period_text": _sanitize_label(candidate.short_period_text or "", limit=120) or None,
        "campaign_id": getattr(item, "campaign_id", None),
        "promo_id": getattr(item, "promo_id", None),
        "manifest_title": _sanitize_label(str(getattr(item, "title", "") or ""), limit=160) or None,
        "manifest_period_text": _sanitize_label(str(getattr(item, "period_text", "") or ""), limit=120) or None,
        "manifest_status": str(getattr(item, "lifecycle_status", "") or "unknown"),
        "manifest_status_confidence": str(getattr(item, "lifecycle_status_confidence", "") or "low"),
        "manifest_downloadability": str(getattr(item, "downloadability", "") or "unknown"),
        "manifest_match_confidence": str(manifest_decision.get("manifest_match_confidence") or "none"),
        "manifest_decision": manifest_decision.get("manifest_decision"),
        "drawer_opened": drawer_opened,
        "drawer_skip_reason": manifest_decision.get("drawer_skip_reason"),
        "drawer_required_reason": manifest_decision.get("drawer_required_reason"),
        "evidence_sources": list(manifest_decision.get("evidence_sources") or []),
    }


def _timeline_candidate_diagnostic(
    candidate: TimelineCandidate,
    timeline_decision: dict[str, Any],
    *,
    drawer_opened: bool,
) -> dict[str, Any]:
    return {
        "timeline_classifier_schema_version": TIMELINE_CLASSIFIER_SCHEMA_VERSION,
        "campaign_id": None,
        "promo_id": None,
        "timeline_block_index": candidate.index,
        "title": _sanitize_label(candidate.title, limit=160),
        "timeline_period_text": _sanitize_label(candidate.short_period_text or "", limit=120) or None,
        "timeline_status": candidate.timeline_status,
        "timeline_status_confidence": candidate.timeline_status_confidence,
        "timeline_status_raw_labels": list(candidate.timeline_status_raw_labels),
        "timeline_goods_count": candidate.timeline_goods_count,
        "timeline_autoaction_marker": candidate.timeline_autoaction_marker,
        "classification_decision": timeline_decision.get("timeline_classification_decision"),
        "drawer_opened": drawer_opened,
        "drawer_open_reason": timeline_decision.get("drawer_open_reason"),
        "drawer_skip_reason": timeline_decision.get("drawer_skip_reason"),
        "fallback_to_full_flow_reason": timeline_decision.get("fallback_to_full_flow_reason"),
        "evidence_sources": list(candidate.timeline_evidence_sources),
    }


def _increment_count(target: dict[str, int], key: str) -> None:
    target[key] = int(target.get(key, 0) or 0) + 1


def _elapsed_ms(started_perf: float) -> int:
    return max(0, int(round((time.perf_counter() - started_perf) * 1000)))
