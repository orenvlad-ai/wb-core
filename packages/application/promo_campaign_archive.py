"""Archive-first promo campaign storage and interval-based replay."""

from __future__ import annotations

from dataclasses import asdict
from dataclasses import dataclass
from datetime import date, datetime
import hashlib
import json
from pathlib import Path
import re
import shutil
import time
from typing import Any, Iterable

from openpyxl import load_workbook

from packages.application.promo_metric_truth import (
    PromoCandidateRow,
    PromoEligibilityEvaluation,
    evaluate_candidate_rows,
    find_workbook_data_sheet,
    iter_workbook_plan_rows,
)
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime
from packages.contracts.promo_live_source import (
    PromoLiveSourceIncomplete,
    PromoLiveSourceItem,
    PromoLiveSourceSuccess,
)
from packages.contracts.promo_xlsx_collector_block import PromoMetadata


PROMO_CAMPAIGN_ARCHIVE_DIRNAME = "promo_campaign_archive"
ARCHIVE_RECORD_FILENAME = "archive_record.json"
ARCHIVE_METADATA_FILENAME = "metadata.json"
ARCHIVE_WORKBOOK_FILENAME = "workbook.xlsx"
ARCHIVE_WORKBOOK_INSPECTION_FILENAME = "workbook_inspection.json"
PRICES_SOURCE_KEY = "prices_snapshot"
PRICES_ACCEPTED_CURRENT_ROLE = "accepted_current_snapshot"
ARTIFACT_VALIDATION_SCHEMA_VERSION = "promo_artifact_validation_v1"
ARTIFACT_STATE_COMPLETE = "complete"
ARTIFACT_STATE_INCOMPLETE = "incomplete"
ARTIFACT_STATE_STALE = "stale"
ARTIFACT_STATE_CORRUPTED = "corrupted"
ARTIFACT_STATE_MISSING_WORKBOOK = "missing_workbook"
ARTIFACT_STATE_METADATA_ONLY = "metadata_only"
ARTIFACT_STATE_WORKBOOK_WITHOUT_METADATA = "workbook_without_metadata"
ARTIFACT_STATE_AMBIGUOUS_DATE = "ambiguous_date"
ARTIFACT_STATE_UNUSABLE = "unusable"
ARTIFACT_STATE_ENDED_WITHOUT_DOWNLOAD = "ended_without_download"
ARTIFACT_REASON_METADATA_ONLY_ENDED_WITHOUT_DOWNLOAD = "metadata_only_ended_without_download"
ARTIFACT_REASON_METADATA_ONLY_TRUE_ARTIFACT_LOSS = "metadata_only_true_artifact_loss"
ARTIFACT_STATES = (
    ARTIFACT_STATE_COMPLETE,
    ARTIFACT_STATE_INCOMPLETE,
    ARTIFACT_STATE_STALE,
    ARTIFACT_STATE_CORRUPTED,
    ARTIFACT_STATE_MISSING_WORKBOOK,
    ARTIFACT_STATE_METADATA_ONLY,
    ARTIFACT_STATE_WORKBOOK_WITHOUT_METADATA,
    ARTIFACT_STATE_AMBIGUOUS_DATE,
    ARTIFACT_STATE_UNUSABLE,
    ARTIFACT_STATE_ENDED_WITHOUT_DOWNLOAD,
)


@dataclass(frozen=True)
class PromoCampaignArchiveRecord:
    archive_key: str
    archive_dir: str
    metadata_fingerprint: str
    workbook_fingerprint: str | None
    workbook_present: bool
    workbook_path: str | None
    workbook_inspection_path: str | None
    collected_at: str
    downloaded_at: str | None
    metadata: PromoMetadata


@dataclass(frozen=True)
class PromoCampaignArchiveSyncSummary:
    scanned_promo_dirs: int = 0
    created_records: int = 0
    updated_records: int = 0
    unchanged_records: int = 0

    def to_note(self) -> str:
        return (
            f"archive_scanned={self.scanned_promo_dirs}; "
            f"archive_created={self.created_records}; "
            f"archive_updated={self.updated_records}; "
            f"archive_unchanged={self.unchanged_records}"
        )


@dataclass(frozen=True)
class DailyPriceTruthResolution:
    price_by_nm_id: dict[int, float]
    source_note: str
    captured_at: str | None = None
    fingerprint: str | None = None


@dataclass(frozen=True)
class PromoCampaignArtifactValidation:
    archive_key: str
    campaign_id: str | None
    promo_id: int | None
    period_id: int | None
    normalized_artifact_key: str | None
    artifact_state: str
    validation_failure_reason: str | None
    expected_workbook_path: str | None
    actual_candidate_paths: list[str]
    metadata_exists: bool
    workbook_exists: bool
    coverage_dates: dict[str, str | None]
    date_confidence: str | None
    temporal_classification: str | None
    export_kind: str | None
    file_size: int | None
    file_mtime: int | None
    checksum_or_fingerprint: str | None
    workbook_required: bool
    non_materializable_reason: str | None
    ui_status: str | None
    ui_status_confidence: str | None
    ui_status_raw_labels: list[str]
    download_action_state: str | None
    download_action_evidence: str | None
    status_evidence_sources: list[str]
    ui_loaded_success: bool
    campaign_identity_match: bool

    @property
    def is_complete(self) -> bool:
        return self.artifact_state == ARTIFACT_STATE_COMPLETE

    @property
    def is_expected_non_materializable(self) -> bool:
        return self.artifact_state == ARTIFACT_STATE_ENDED_WITHOUT_DOWNLOAD

    def to_diagnostic(self) -> dict[str, Any]:
        return {
            "campaign_id": self.campaign_id,
            "promo_id": self.promo_id,
            "period_id": self.period_id,
            "normalized_artifact_key": self.normalized_artifact_key,
            "artifact_state": self.artifact_state,
            "validation_failure_reason": self.validation_failure_reason,
            "expected_workbook_path": self.expected_workbook_path,
            "actual_candidate_paths": list(self.actual_candidate_paths),
            "metadata_exists": self.metadata_exists,
            "workbook_exists": self.workbook_exists,
            "coverage_dates": dict(self.coverage_dates),
            "date_confidence": self.date_confidence,
            "temporal_classification": self.temporal_classification,
            "export_kind": self.export_kind,
            "file_size": self.file_size,
            "file_mtime": self.file_mtime,
            "checksum_or_fingerprint": self.checksum_or_fingerprint,
            "workbook_required": self.workbook_required,
            "non_materializable_reason": self.non_materializable_reason,
            "ui_status": self.ui_status,
            "ui_status_confidence": self.ui_status_confidence,
            "ui_status_raw_labels": list(self.ui_status_raw_labels),
            "download_action_state": self.download_action_state,
            "download_action_evidence": self.download_action_evidence,
            "status_evidence_sources": list(self.status_evidence_sources),
            "ui_loaded_success": self.ui_loaded_success,
            "campaign_identity_match": self.campaign_identity_match,
        }


def promo_campaign_archive_root(runtime_dir: Path) -> Path:
    return runtime_dir / PROMO_CAMPAIGN_ARCHIVE_DIRNAME


def load_promo_campaign_archive(runtime_dir: Path) -> list[PromoCampaignArchiveRecord]:
    archive_root = promo_campaign_archive_root(runtime_dir)
    if not archive_root.exists():
        return []
    records: list[PromoCampaignArchiveRecord] = []
    for record_path in sorted(archive_root.glob(f"*/{ARCHIVE_RECORD_FILENAME}")):
        payload = json.loads(record_path.read_text(encoding="utf-8"))
        metadata_payload = payload.get("metadata", {})
        records.append(
            PromoCampaignArchiveRecord(
                archive_key=str(payload.get("archive_key", "") or ""),
                archive_dir=str(payload.get("archive_dir", "") or record_path.parent),
                metadata_fingerprint=str(payload.get("metadata_fingerprint", "") or ""),
                workbook_fingerprint=_normalize_optional_text(payload.get("workbook_fingerprint")),
                workbook_present=bool(payload.get("workbook_present", False)),
                workbook_path=_normalize_optional_text(payload.get("workbook_path")),
                workbook_inspection_path=_normalize_optional_text(payload.get("workbook_inspection_path")),
                collected_at=str(payload.get("collected_at", "") or ""),
                downloaded_at=_normalize_optional_text(payload.get("downloaded_at")),
                metadata=PromoMetadata(**metadata_payload),
            )
        )
    return records


def resolve_reusable_campaign(
    *,
    archive_root: Path,
    metadata: PromoMetadata,
) -> PromoCampaignArchiveRecord | None:
    matches: list[PromoCampaignArchiveRecord] = []
    for record in load_promo_campaign_archive(archive_root.parent):
        workbook_path = Path(record.workbook_path) if record.workbook_path else None
        if not record.workbook_present or workbook_path is None or not workbook_path.exists():
            continue
        if _record_matches_live_metadata(record=record, metadata=metadata):
            matches.append(record)
    if not matches:
        return None
    matches.sort(
        key=lambda record: (
            record.downloaded_at or "",
            record.collected_at,
            record.archive_key,
        ),
        reverse=True,
    )
    return matches[0]


def validate_promo_campaign_artifact(
    record: PromoCampaignArchiveRecord,
    *,
    requested_slot_date: date | None = None,
    deep_workbook_check: bool = False,
) -> PromoCampaignArtifactValidation:
    archive_dir = Path(record.archive_dir)
    archive_key = record.archive_key or archive_dir.name
    metadata_path = archive_dir / ARCHIVE_METADATA_FILENAME
    record_workbook_path = Path(record.workbook_path) if record.workbook_path else None
    expected_workbook_path = archive_dir / ARCHIVE_WORKBOOK_FILENAME
    actual_candidate_paths = _candidate_workbook_paths(
        archive_dir=archive_dir,
        record_workbook_path=record_workbook_path,
        expected_workbook_path=expected_workbook_path,
    )
    workbook_path = _resolve_record_workbook_path(
        record_workbook_path=record_workbook_path,
        expected_workbook_path=expected_workbook_path,
    )
    workbook_exists = bool(workbook_path and workbook_path.exists())
    file_size = _file_size(workbook_path) if workbook_exists else None
    file_mtime = _file_mtime(workbook_path) if workbook_exists else None
    checksum_or_fingerprint = record.workbook_fingerprint
    if workbook_exists and not checksum_or_fingerprint and workbook_path is not None:
        checksum_or_fingerprint = _sha256_path(workbook_path)

    state = ARTIFACT_STATE_COMPLETE
    reason: str | None = None
    workbook_required = True
    non_materializable_reason: str | None = None
    metadata_exists = metadata_path.exists()
    ended_without_download = _metadata_indicates_ended_without_download(record.metadata)
    coverage_dates = {
        "promo_start_at": record.metadata.promo_start_at,
        "promo_end_at": record.metadata.promo_end_at,
    }

    if not metadata_exists:
        state = ARTIFACT_STATE_WORKBOOK_WITHOUT_METADATA if workbook_exists else ARTIFACT_STATE_INCOMPLETE
        reason = "metadata_sidecar_missing"
    elif not record.metadata.promo_start_at or not record.metadata.promo_end_at:
        state = ARTIFACT_STATE_AMBIGUOUS_DATE
        reason = "coverage_dates_missing"
    elif not _record_dates_parseable(record):
        state = ARTIFACT_STATE_AMBIGUOUS_DATE
        reason = "coverage_dates_unparseable"
    elif str(record.metadata.period_parse_confidence or "") != "high":
        state = ARTIFACT_STATE_AMBIGUOUS_DATE
        reason = "coverage_dates_not_high_confidence"
    elif requested_slot_date is not None and not _record_covers_date_safe(record, requested_slot_date):
        state = ARTIFACT_STATE_STALE
        reason = "coverage_does_not_include_requested_slot"
    elif not record_workbook_path:
        if ended_without_download:
            state = ARTIFACT_STATE_ENDED_WITHOUT_DOWNLOAD
            reason = ARTIFACT_REASON_METADATA_ONLY_ENDED_WITHOUT_DOWNLOAD
            workbook_required = False
            non_materializable_reason = "ended_without_download"
        else:
            state = ARTIFACT_STATE_METADATA_ONLY
            reason = ARTIFACT_REASON_METADATA_ONLY_TRUE_ARTIFACT_LOSS
    elif _normalized_path(record_workbook_path) != _normalized_path(expected_workbook_path):
        state = ARTIFACT_STATE_STALE
        reason = "workbook_path_mismatch"
    elif not workbook_exists:
        if ended_without_download:
            state = ARTIFACT_STATE_ENDED_WITHOUT_DOWNLOAD
            reason = ARTIFACT_REASON_METADATA_ONLY_ENDED_WITHOUT_DOWNLOAD
            workbook_required = False
            non_materializable_reason = "ended_without_download"
        else:
            state = (
                ARTIFACT_STATE_MISSING_WORKBOOK
                if record.workbook_present
                else ARTIFACT_STATE_METADATA_ONLY
            )
            reason = "workbook_file_missing" if record.workbook_present else ARTIFACT_REASON_METADATA_ONLY_TRUE_ARTIFACT_LOSS
    elif file_size is None or file_size <= 0:
        state = ARTIFACT_STATE_CORRUPTED
        reason = "workbook_file_empty"
    else:
        inspection_reason = _workbook_inspection_failure_reason(record)
        if inspection_reason:
            state = ARTIFACT_STATE_CORRUPTED
            reason = inspection_reason

    if (
        state == ARTIFACT_STATE_COMPLETE
        and deep_workbook_check
        and workbook_path is not None
    ):
        workbook_reason = _workbook_open_failure_reason(workbook_path)
        if workbook_reason:
            state = ARTIFACT_STATE_CORRUPTED
            reason = workbook_reason

    return PromoCampaignArtifactValidation(
        archive_key=archive_key,
        campaign_id=_campaign_identity(record) if archive_key else None,
        promo_id=record.metadata.promo_id,
        period_id=record.metadata.period_id,
        normalized_artifact_key=archive_key,
        artifact_state=state,
        validation_failure_reason=reason,
        expected_workbook_path=str(expected_workbook_path),
        actual_candidate_paths=actual_candidate_paths,
        metadata_exists=metadata_exists,
        workbook_exists=workbook_exists,
        coverage_dates=coverage_dates,
        date_confidence=record.metadata.period_parse_confidence,
        temporal_classification=record.metadata.temporal_classification,
        export_kind=record.metadata.export_kind,
        file_size=file_size,
        file_mtime=file_mtime,
        checksum_or_fingerprint=checksum_or_fingerprint,
        workbook_required=workbook_required,
        non_materializable_reason=non_materializable_reason,
        ui_status=record.metadata.ui_status,
        ui_status_confidence=record.metadata.ui_status_confidence,
        ui_status_raw_labels=list(record.metadata.ui_status_raw_labels),
        download_action_state=record.metadata.download_action_state,
        download_action_evidence=record.metadata.download_action_evidence,
        status_evidence_sources=list(record.metadata.status_evidence_sources),
        ui_loaded_success=record.metadata.ui_loaded_success,
        campaign_identity_match=record.metadata.campaign_identity_match,
    )


def audit_promo_campaign_archive(
    runtime_dir: Path,
    *,
    snapshot_date: str | None = None,
    max_examples: int = 20,
    deep_workbook_check: bool = False,
) -> dict[str, Any]:
    archive_root = promo_campaign_archive_root(runtime_dir)
    slot_date = date.fromisoformat(snapshot_date) if snapshot_date else None
    records = load_promo_campaign_archive(runtime_dir)
    validations = [
        validate_promo_campaign_artifact(
            record,
            requested_slot_date=slot_date if slot_date and _record_covers_date_safe(record, slot_date) else None,
            deep_workbook_check=deep_workbook_check,
        )
        for record in records
    ]
    known_dirs = {_normalized_path(Path(record.archive_dir)) for record in records}
    if archive_root.exists():
        for archive_dir in sorted(path for path in archive_root.iterdir() if path.is_dir()):
            if _normalized_path(archive_dir) in known_dirs:
                continue
            validations.append(_validate_orphan_archive_dir(archive_dir))
    state_counts = _artifact_state_counts(validations)
    failures = [validation for validation in validations if not validation.is_complete]
    expected_non_materializable = [validation for validation in validations if validation.is_expected_non_materializable]
    metadata_only_true_artifact_loss_count = sum(
        1
        for validation in validations
        if validation.validation_failure_reason == ARTIFACT_REASON_METADATA_ONLY_TRUE_ARTIFACT_LOSS
    )
    covering_validations = (
        [
            validation
            for record, validation in zip(records, validations)
            if slot_date is not None and _record_covers_date_safe(record, slot_date)
        ]
        if slot_date is not None
        else []
    )
    return {
        "artifact_validation_schema_version": ARTIFACT_VALIDATION_SCHEMA_VERSION,
        "archive_root": str(archive_root),
        "snapshot_date": snapshot_date,
        "record_count": len(records),
        "archive_dir_count": len(list(archive_root.iterdir())) if archive_root.exists() else 0,
        "artifact_state_counts": state_counts,
        "complete_artifact_count": state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
        "incomplete_artifact_count": len(failures),
        "validation_failed_count": len(failures),
        "validated_workbook_usable_count": state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
        "ended_without_download_count": state_counts.get(ARTIFACT_STATE_ENDED_WITHOUT_DOWNLOAD, 0),
        "metadata_only_true_artifact_loss_count": metadata_only_true_artifact_loss_count,
        "non_materializable_expected_count": len(expected_non_materializable),
        "ui_status_counts": _validation_field_counts(validations, "ui_status"),
        "download_action_state_counts": _validation_field_counts(validations, "download_action_state"),
        "covering_campaign_count": len(covering_validations),
        "covering_artifact_state_counts": _artifact_state_counts(covering_validations),
        "failure_examples": [
            validation.to_diagnostic()
            for validation in failures[:max(0, int(max_examples))]
        ],
    }


def sync_promo_campaign_archive(runtime_dir: Path) -> PromoCampaignArchiveSyncSummary:
    runs_root = runtime_dir / "promo_xlsx_collector_runs"
    archive_root = promo_campaign_archive_root(runtime_dir)
    archive_root.mkdir(parents=True, exist_ok=True)
    summary = PromoCampaignArchiveSyncSummary()
    if not runs_root.exists():
        return summary
    created = 0
    updated = 0
    unchanged = 0
    scanned = 0
    for metadata_path in sorted(runs_root.glob("*/promos/*/metadata.json")):
        scanned += 1
        changed = _sync_archive_record_from_metadata(
            metadata_path=metadata_path,
            archive_root=archive_root,
        )
        if changed == "created":
            created += 1
        elif changed == "updated":
            updated += 1
        else:
            unchanged += 1
    return PromoCampaignArchiveSyncSummary(
        scanned_promo_dirs=scanned,
        created_records=created,
        updated_records=updated,
        unchanged_records=unchanged,
    )


def materialize_promo_result_from_archive(
    *,
    runtime_dir: Path,
    snapshot_date: str,
    requested_nm_ids: Iterable[int],
    sync_summary: PromoCampaignArchiveSyncSummary | None = None,
    trace_run_dir: str | None = None,
    detail_prefix: str | None = None,
    diagnostics: dict[str, Any] | None = None,
) -> PromoLiveSourceSuccess | PromoLiveSourceIncomplete:
    if sync_summary is None:
        sync_phase = _start_promo_diag_phase(diagnostics, "archive_sync")
        sync_state = sync_promo_campaign_archive(runtime_dir)
        _finish_promo_diag_phase(diagnostics, sync_phase, status="success")
    else:
        sync_state = sync_summary

    archive_lookup_phase = _start_promo_diag_phase(diagnostics, "archive_lookup")
    archive_root = promo_campaign_archive_root(runtime_dir)
    requested = sorted({int(nm_id) for nm_id in requested_nm_ids})
    slot_date = date.fromisoformat(snapshot_date)
    records = load_promo_campaign_archive(runtime_dir)
    covering = [record for record in records if _record_covers_date_safe(record, slot_date)]
    validation_pairs = [
        (
            record,
            validate_promo_campaign_artifact(
                record,
                requested_slot_date=slot_date,
                deep_workbook_check=True,
            ),
        )
        for record in covering
    ]
    usable = [
        record
        for record, validation in validation_pairs
        if validation.is_complete
    ]
    failed_artifacts = [
        (record, validation)
        for record, validation in validation_pairs
        if not validation.is_complete
    ]
    missing_artifacts = [record for record, _validation in failed_artifacts]
    _apply_artifact_validation_diagnostics(
        diagnostics=diagnostics,
        validations=[validation for _record, validation in validation_pairs],
        snapshot_date=snapshot_date,
    )
    _set_promo_diag_counter(diagnostics, "campaign_count", len(records), overwrite=False)
    _set_promo_diag_counter(diagnostics, "current_promo_count", len(covering))
    _set_promo_diag_counter(diagnostics, "archive_hit_count", len(usable))
    _set_promo_diag_counter(diagnostics, "archive_miss_count", len(missing_artifacts))
    _set_promo_diag_fingerprint(
        diagnostics,
        "promo_archive_fingerprint",
        _archive_records_fingerprint(covering),
    )
    _set_promo_diag_fingerprint(
        diagnostics,
        "workbook_metadata_fingerprint",
        _record_field_fingerprint(usable, "metadata_fingerprint"),
    )
    _set_promo_diag_fingerprint(
        diagnostics,
        "workbook_file_fingerprint",
        _record_field_fingerprint(usable, "workbook_fingerprint"),
    )
    _set_promo_diag_fingerprint(
        diagnostics,
        "workbook_file_size_total",
        _workbook_file_size_total(usable),
    )
    _set_promo_diag_fingerprint(
        diagnostics,
        "workbook_file_mtime_max",
        _workbook_file_mtime_max(usable),
    )
    _finish_promo_diag_phase(diagnostics, archive_lookup_phase, status="success")

    metadata_phase = _start_promo_diag_phase(diagnostics, "metadata_validation")
    metadata_valid_count = sum(1 for record in records if record.metadata.promo_start_at and record.metadata.promo_end_at)
    _set_promo_diag_counter(diagnostics, "metadata_valid_count", metadata_valid_count)
    _set_promo_diag_counter(diagnostics, "metadata_invalid_count", max(0, len(records) - metadata_valid_count))
    _finish_promo_diag_phase(diagnostics, metadata_phase, status="success")

    common = dict(
        snapshot_date=snapshot_date,
        date_from=snapshot_date,
        date_to=snapshot_date,
        requested_count=len(requested),
        trace_run_dir=trace_run_dir or str(archive_root),
        current_promos=len(covering),
        current_promos_downloaded=len(usable),
        current_promos_blocked=len(missing_artifacts),
        future_promos=0,
        skipped_past_promos=0,
        ambiguous_promos=0,
        current_download_export_kinds=sorted(
            {
                str(record.metadata.export_kind)
                for record in usable
                if record.metadata.export_kind
            }
        ),
    )

    detail_parts = [
        "archive_mode=interval_replay",
        sync_state.to_note(),
        f"covering_campaigns={len(covering)}",
        f"usable_campaigns={len(usable)}",
        f"artifact_validation_schema={ARTIFACT_VALIDATION_SCHEMA_VERSION}",
    ]
    if detail_prefix:
        detail_parts.insert(0, detail_prefix)

    if missing_artifacts:
        missing_keys = ",".join(record.archive_key for record in missing_artifacts[:8])
        failure_reasons = ",".join(
            sorted(
                {
                    str(validation.validation_failure_reason or validation.artifact_state)
                    for _record, validation in failed_artifacts
                }
            )
        )
        _set_promo_diag_counter(diagnostics, "candidate_row_count", 0)
        _set_promo_diag_counter(diagnostics, "eligible_row_count", 0)
        _set_promo_diag_counter(diagnostics, "accepted_row_count", 0)
        _set_promo_diag_counter(diagnostics, "skipped_row_count", len(requested))
        _set_promo_diag_counter(diagnostics, "price_truth_missing_count", None)
        _set_promo_diag_counter(diagnostics, "price_truth_available_count", None)
        _finish_source_payload_build_phase(diagnostics, status="incomplete", note_kind="missing_campaign_artifacts")
        return PromoLiveSourceIncomplete(
            kind="incomplete",
            covered_count=0,
            items=[],
            missing_nm_ids=requested,
            detail="; ".join(
                detail_parts
                + [
                    f"missing_campaign_artifacts={missing_keys}",
                    f"artifact_validation_failed={failure_reasons}",
                ]
            ),
            diagnostics=_snapshot_promo_diagnostics(diagnostics),
            **common,
        )

    candidate_rows_by_nm_id: dict[int, list[PromoCandidateRow]] = {
        nm_id: []
        for nm_id in requested
    }
    workbook_phase = _start_promo_diag_phase(diagnostics, "workbook_inspection")
    candidate_row_count = 0
    for record in usable:
        candidate_row_count += _merge_archive_workbook_rows(
            candidate_rows_by_nm_id=candidate_rows_by_nm_id,
            workbook_path=Path(str(record.workbook_path)),
            requested_nm_ids=requested,
            campaign_identity=_campaign_identity(record),
        )
    _set_promo_diag_counter(diagnostics, "candidate_row_count", candidate_row_count)
    _finish_promo_diag_phase(diagnostics, workbook_phase, status="success")
    detail_parts.append(f"requested_candidate_rows={candidate_row_count}")

    requested_nm_ids_with_candidates = sorted(
        nm_id
        for nm_id, candidate_rows in candidate_rows_by_nm_id.items()
        if candidate_rows
    )
    if not requested_nm_ids_with_candidates:
        _set_promo_diag_counter(diagnostics, "eligible_row_count", 0)
        _set_promo_diag_counter(diagnostics, "accepted_row_count", len(requested))
        _set_promo_diag_counter(diagnostics, "skipped_row_count", len(requested))
        _set_promo_diag_counter(diagnostics, "price_truth_missing_count", 0)
        _set_promo_diag_counter(diagnostics, "price_truth_available_count", 0)
        _finish_source_payload_build_phase(diagnostics, status="success", note_kind="no_covering_campaign_rows")
        return PromoLiveSourceSuccess(
            kind="success",
            covered_count=len(requested),
            items=_build_zero_items(snapshot_date=snapshot_date, requested_nm_ids=requested),
            detail="; ".join(
                detail_parts
                + [
                    "daily_price_source=not_needed",
                    "no_covering_campaign_rows_for_requested_nm_ids=true",
                ]
            ),
            diagnostics=_snapshot_promo_diagnostics(diagnostics),
            **common,
        )

    price_truth_phase = _start_promo_diag_phase(diagnostics, "price_truth_lookup")
    price_truth = _load_daily_price_truth(
        runtime_dir=runtime_dir,
        snapshot_date=snapshot_date,
        requested_nm_ids=requested_nm_ids_with_candidates,
    )
    _set_promo_diag_fingerprint(diagnostics, "accepted_price_truth_date", snapshot_date if price_truth.price_by_nm_id else None)
    _set_promo_diag_fingerprint(diagnostics, "accepted_price_truth_version", price_truth.captured_at)
    _set_promo_diag_fingerprint(diagnostics, "accepted_price_truth_fingerprint", price_truth.fingerprint)
    _finish_promo_diag_phase(
        diagnostics,
        price_truth_phase,
        status="success" if price_truth.price_by_nm_id else "missing",
        note_kind="accepted_price_truth" if price_truth.price_by_nm_id else "missing_price_truth",
    )
    detail_parts.append(price_truth.source_note)
    missing_price_truth_nm_ids = sorted(
        nm_id
        for nm_id in requested_nm_ids_with_candidates
        if nm_id not in price_truth.price_by_nm_id
    )

    materialized_items: list[PromoLiveSourceItem] = []
    truthful_zero_no_candidate_nm_ids: list[int] = []
    truthful_zero_ineligible_nm_ids: list[int] = []
    eligible_nm_ids: list[int] = []
    eligible_row_count = 0
    price_join_phase = _start_promo_diag_phase(diagnostics, "price_truth_join")
    for nm_id in requested:
        candidate_rows = candidate_rows_by_nm_id[nm_id]
        item, evaluation = _build_item_from_candidate_rows(
            snapshot_date=snapshot_date,
            nm_id=nm_id,
            candidate_rows=candidate_rows,
            price_seller_discounted=price_truth.price_by_nm_id.get(nm_id),
        )
        materialized_items.append(item)
        if not candidate_rows:
            truthful_zero_no_candidate_nm_ids.append(nm_id)
        elif evaluation.eligible_campaign_identities:
            eligible_nm_ids.append(nm_id)
            eligible_row_count += len(evaluation.eligible_campaign_identities)
        elif nm_id not in missing_price_truth_nm_ids:
            truthful_zero_ineligible_nm_ids.append(nm_id)
    _set_promo_diag_counter(diagnostics, "eligible_row_count", eligible_row_count)
    _set_promo_diag_counter(diagnostics, "price_truth_available_count", len(price_truth.price_by_nm_id))
    _set_promo_diag_counter(diagnostics, "price_truth_missing_count", len(missing_price_truth_nm_ids))
    _set_promo_diag_counter(
        diagnostics,
        "skipped_row_count",
        len(truthful_zero_no_candidate_nm_ids) + len(truthful_zero_ineligible_nm_ids) + len(missing_price_truth_nm_ids),
    )
    _finish_promo_diag_phase(diagnostics, price_join_phase, status="success")

    detail_parts.extend(
        _promo_metric_coverage_detail_parts(
            eligible_nm_ids=eligible_nm_ids,
            truthful_zero_no_candidate_nm_ids=truthful_zero_no_candidate_nm_ids,
            truthful_zero_ineligible_nm_ids=truthful_zero_ineligible_nm_ids,
            missing_price_truth_nm_ids=missing_price_truth_nm_ids,
        )
    )
    archive_keys = ",".join(record.archive_key for record in usable[:8])
    if missing_price_truth_nm_ids:
        _set_promo_diag_counter(diagnostics, "accepted_row_count", 0)
        _finish_source_payload_build_phase(diagnostics, status="incomplete", note_kind="missing_price_truth")
        return PromoLiveSourceIncomplete(
            kind="incomplete",
            covered_count=len(requested) - len(missing_price_truth_nm_ids),
            items=materialized_items,
            missing_nm_ids=missing_price_truth_nm_ids,
            detail="; ".join(
                detail_parts
                + [
                    "daily_price_truth_required=true",
                    "missing_daily_price_truth_nm_ids="
                    + ",".join(str(nm_id) for nm_id in missing_price_truth_nm_ids),
                    f"archive_keys={archive_keys}",
                ]
            ),
            diagnostics=_snapshot_promo_diagnostics(diagnostics),
            **common,
        )
    _set_promo_diag_counter(diagnostics, "accepted_row_count", len(requested))
    _finish_source_payload_build_phase(diagnostics, status="success")
    return PromoLiveSourceSuccess(
        kind="success",
        covered_count=len(requested),
        items=materialized_items,
        detail="; ".join(detail_parts + [f"archive_keys={archive_keys}"]),
        diagnostics=_snapshot_promo_diagnostics(diagnostics),
        **common,
    )


def _apply_artifact_validation_diagnostics(
    *,
    diagnostics: dict[str, Any] | None,
    validations: list[PromoCampaignArtifactValidation],
    snapshot_date: str,
) -> None:
    if diagnostics is None:
        return
    state_counts = _artifact_state_counts(validations)
    failures = [validation for validation in validations if not validation.is_complete]
    true_failures = [validation for validation in failures if not validation.is_expected_non_materializable]
    expected_non_materializable = [
        validation
        for validation in validations
        if validation.is_expected_non_materializable
    ]
    missing_workbook_count = (
        state_counts.get(ARTIFACT_STATE_MISSING_WORKBOOK, 0)
        + state_counts.get(ARTIFACT_STATE_METADATA_ONLY, 0)
    )
    ui_status_counts = _validation_field_counts(validations, "ui_status")
    download_action_state_counts = _validation_field_counts(validations, "download_action_state")
    ended_without_download_count = state_counts.get(ARTIFACT_STATE_ENDED_WITHOUT_DOWNLOAD, 0)
    metadata_only_true_artifact_loss_count = sum(
        1
        for validation in validations
        if validation.validation_failure_reason == ARTIFACT_REASON_METADATA_ONLY_TRUE_ARTIFACT_LOSS
    )
    diagnostics["artifact_validation_schema_version"] = ARTIFACT_VALIDATION_SCHEMA_VERSION
    diagnostics["artifact_state_counts"] = state_counts
    diagnostics["ui_status_counts"] = ui_status_counts
    diagnostics["download_action_state_counts"] = download_action_state_counts
    diagnostics["missing_campaign_artifacts"] = [
        validation.to_diagnostic()
        for validation in failures[:20]
    ]
    diagnostics["artifact_validation_summary"] = {
        "schema_version": ARTIFACT_VALIDATION_SCHEMA_VERSION,
        "snapshot_date": snapshot_date,
        "covering_campaign_count": len(validations),
        "complete_artifact_count": state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
        "incomplete_artifact_count": len(failures),
        "true_validation_failed_count": len(true_failures),
        "validated_workbook_usable_count": state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
        "materializer_usable_count": state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
        "validation_failed_count": len(failures),
        "workbook_missing_count": missing_workbook_count,
        "metadata_only_count": state_counts.get(ARTIFACT_STATE_METADATA_ONLY, 0),
        "workbook_without_metadata_count": state_counts.get(ARTIFACT_STATE_WORKBOOK_WITHOUT_METADATA, 0),
        "corrupted_count": state_counts.get(ARTIFACT_STATE_CORRUPTED, 0),
        "ambiguous_date_count": state_counts.get(ARTIFACT_STATE_AMBIGUOUS_DATE, 0),
        "ended_without_download_count": ended_without_download_count,
        "metadata_only_true_artifact_loss_count": metadata_only_true_artifact_loss_count,
        "non_materializable_expected_count": len(expected_non_materializable),
        "ui_status_counts": ui_status_counts,
        "download_action_state_counts": download_action_state_counts,
    }
    _set_promo_diag_counter(
        diagnostics,
        "validated_workbook_usable_count",
        state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
    )
    _set_promo_diag_counter(
        diagnostics,
        "materializer_usable_count",
        state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
    )
    _set_promo_diag_counter(diagnostics, "validation_failed_count", len(failures))
    _set_promo_diag_counter(diagnostics, "workbook_missing_count", missing_workbook_count)
    _set_promo_diag_counter(
        diagnostics,
        "metadata_only_count",
        state_counts.get(ARTIFACT_STATE_METADATA_ONLY, 0),
    )
    _set_promo_diag_counter(
        diagnostics,
        "workbook_without_metadata_count",
        state_counts.get(ARTIFACT_STATE_WORKBOOK_WITHOUT_METADATA, 0),
    )
    _set_promo_diag_counter(
        diagnostics,
        "corrupted_count",
        state_counts.get(ARTIFACT_STATE_CORRUPTED, 0),
    )
    _set_promo_diag_counter(
        diagnostics,
        "ambiguous_date_count",
        state_counts.get(ARTIFACT_STATE_AMBIGUOUS_DATE, 0),
    )
    _set_promo_diag_counter(
        diagnostics,
        "ended_without_download_count",
        ended_without_download_count,
    )
    _set_promo_diag_counter(
        diagnostics,
        "metadata_only_true_artifact_loss_count",
        metadata_only_true_artifact_loss_count,
    )
    _set_promo_diag_counter(
        diagnostics,
        "non_materializable_expected_count",
        len(expected_non_materializable),
    )
    _set_promo_diag_counter(
        diagnostics,
        "complete_artifact_count",
        state_counts.get(ARTIFACT_STATE_COMPLETE, 0),
    )
    _set_promo_diag_counter(diagnostics, "incomplete_artifact_count", len(failures))


def _sync_archive_record_from_metadata(
    *,
    metadata_path: Path,
    archive_root: Path,
) -> str:
    promo_dir = metadata_path.parent
    metadata = PromoMetadata(**json.loads(metadata_path.read_text(encoding="utf-8")))
    archive_key = _archive_key(metadata)
    archive_dir = archive_root / archive_key
    archive_dir.mkdir(parents=True, exist_ok=True)

    existing_record = _load_archive_record(archive_dir / ARCHIVE_RECORD_FILENAME)
    workbook_src = promo_dir / "workbook.xlsx"
    inspection_src = promo_dir / "workbook_inspection.json"
    workbook_present = workbook_src.exists()
    workbook_fingerprint = _sha256_path(workbook_src) if workbook_present else None

    archive_workbook_path = archive_dir / ARCHIVE_WORKBOOK_FILENAME
    archive_inspection_path = archive_dir / ARCHIVE_WORKBOOK_INSPECTION_FILENAME
    record_changed = existing_record is None

    if workbook_present:
        if (
            existing_record is None
            or existing_record.workbook_fingerprint != workbook_fingerprint
            or not archive_workbook_path.exists()
        ):
            shutil.copy2(workbook_src, archive_workbook_path)
            record_changed = True
        if inspection_src.exists():
            if not archive_inspection_path.exists() or archive_inspection_path.read_bytes() != inspection_src.read_bytes():
                shutil.copy2(inspection_src, archive_inspection_path)
                record_changed = True
    elif existing_record and existing_record.workbook_present and existing_record.workbook_path:
        archive_workbook_path = Path(existing_record.workbook_path)
        if existing_record.workbook_inspection_path:
            archive_inspection_path = Path(existing_record.workbook_inspection_path)
        workbook_fingerprint = existing_record.workbook_fingerprint
    else:
        archive_workbook_path = archive_dir / ARCHIVE_WORKBOOK_FILENAME
        archive_inspection_path = archive_dir / ARCHIVE_WORKBOOK_INSPECTION_FILENAME

    normalized_metadata = _normalize_metadata_for_archive(
        metadata=metadata,
        workbook_path=archive_workbook_path if archive_workbook_path.exists() else None,
    )
    metadata_fingerprint = _metadata_fingerprint(normalized_metadata)
    if existing_record is None or existing_record.metadata_fingerprint != metadata_fingerprint:
        record_changed = True

    archive_metadata_path = archive_dir / ARCHIVE_METADATA_FILENAME
    metadata_payload = asdict(normalized_metadata)
    existing_metadata_bytes = archive_metadata_path.read_bytes() if archive_metadata_path.exists() else None
    new_metadata_bytes = json.dumps(metadata_payload, ensure_ascii=False, indent=2).encode("utf-8")
    if existing_metadata_bytes != new_metadata_bytes:
        archive_metadata_path.write_bytes(new_metadata_bytes)
        record_changed = True

    record = PromoCampaignArchiveRecord(
        archive_key=archive_key,
        archive_dir=str(archive_dir),
        metadata_fingerprint=metadata_fingerprint,
        workbook_fingerprint=workbook_fingerprint,
        workbook_present=bool(archive_workbook_path.exists()),
        workbook_path=str(archive_workbook_path) if archive_workbook_path.exists() else None,
        workbook_inspection_path=(
            str(archive_inspection_path)
            if archive_inspection_path.exists()
            else None
        ),
        collected_at=normalized_metadata.collected_at,
        downloaded_at=(
            normalized_metadata.collected_at if workbook_src.exists() else (
                existing_record.downloaded_at if existing_record is not None else None
            )
        ),
        metadata=normalized_metadata,
    )
    record_payload = json.dumps(asdict(record), ensure_ascii=False, indent=2).encode("utf-8")
    record_path = archive_dir / ARCHIVE_RECORD_FILENAME
    existing_record_bytes = record_path.read_bytes() if record_path.exists() else None
    if existing_record_bytes != record_payload:
        record_path.write_bytes(record_payload)
        record_changed = True

    if existing_record is None:
        return "created"
    if record_changed:
        return "updated"
    return "unchanged"


def _load_archive_record(record_path: Path) -> PromoCampaignArchiveRecord | None:
    if not record_path.exists():
        return None
    payload = json.loads(record_path.read_text(encoding="utf-8"))
    return PromoCampaignArchiveRecord(
        archive_key=str(payload.get("archive_key", "") or ""),
        archive_dir=str(payload.get("archive_dir", "") or record_path.parent),
        metadata_fingerprint=str(payload.get("metadata_fingerprint", "") or ""),
        workbook_fingerprint=_normalize_optional_text(payload.get("workbook_fingerprint")),
        workbook_present=bool(payload.get("workbook_present", False)),
        workbook_path=_normalize_optional_text(payload.get("workbook_path")),
        workbook_inspection_path=_normalize_optional_text(payload.get("workbook_inspection_path")),
        collected_at=str(payload.get("collected_at", "") or ""),
        downloaded_at=_normalize_optional_text(payload.get("downloaded_at")),
        metadata=PromoMetadata(**payload.get("metadata", {})),
    )


def _normalize_metadata_for_archive(
    *,
    metadata: PromoMetadata,
    workbook_path: Path | None,
) -> PromoMetadata:
    payload = asdict(metadata)
    if workbook_path is not None and workbook_path.exists():
        payload["saved_path"] = str(workbook_path)
        payload["saved_filename"] = workbook_path.name
    return PromoMetadata(**payload)


def _record_covers_date(record: PromoCampaignArchiveRecord, slot_date: date) -> bool:
    if not record.metadata.promo_start_at or not record.metadata.promo_end_at:
        return False
    start_date = date.fromisoformat(record.metadata.promo_start_at[:10])
    end_date = date.fromisoformat(record.metadata.promo_end_at[:10])
    return start_date <= slot_date <= end_date


def _record_covers_date_safe(record: PromoCampaignArchiveRecord, slot_date: date) -> bool:
    try:
        return _record_covers_date(record, slot_date)
    except Exception:
        return False


def _record_dates_parseable(record: PromoCampaignArchiveRecord) -> bool:
    try:
        if not record.metadata.promo_start_at or not record.metadata.promo_end_at:
            return False
        date.fromisoformat(record.metadata.promo_start_at[:10])
        date.fromisoformat(record.metadata.promo_end_at[:10])
    except Exception:
        return False
    return True


def _candidate_workbook_paths(
    *,
    archive_dir: Path,
    record_workbook_path: Path | None,
    expected_workbook_path: Path,
) -> list[str]:
    candidates: list[Path] = []
    if record_workbook_path is not None:
        candidates.append(record_workbook_path)
    candidates.append(expected_workbook_path)
    if archive_dir.exists():
        candidates.extend(sorted(archive_dir.glob("*.xlsx")))
    seen: set[str] = set()
    result: list[str] = []
    for path in candidates:
        normalized = _normalized_path(path)
        if normalized in seen:
            continue
        seen.add(normalized)
        if path.exists() or path == expected_workbook_path or path == record_workbook_path:
            result.append(str(path))
    return result[:8]


def _resolve_record_workbook_path(
    *,
    record_workbook_path: Path | None,
    expected_workbook_path: Path,
) -> Path | None:
    if record_workbook_path is not None and record_workbook_path.exists():
        return record_workbook_path
    if expected_workbook_path.exists():
        return expected_workbook_path
    return record_workbook_path


def _file_size(path: Path | None) -> int | None:
    try:
        return int(path.stat().st_size) if path is not None else None
    except OSError:
        return None


def _file_mtime(path: Path | None) -> int | None:
    try:
        return int(path.stat().st_mtime) if path is not None else None
    except OSError:
        return None


def _workbook_inspection_failure_reason(record: PromoCampaignArchiveRecord) -> str | None:
    if not record.workbook_inspection_path:
        return None
    inspection_path = Path(record.workbook_inspection_path)
    if not inspection_path.exists():
        return "workbook_inspection_missing"
    try:
        payload = json.loads(inspection_path.read_text(encoding="utf-8"))
    except Exception:
        return "workbook_inspection_corrupted"
    if not isinstance(payload, dict):
        return "workbook_inspection_not_object"
    return None


def _workbook_open_failure_reason(workbook_path: Path) -> str | None:
    workbook = None
    try:
        workbook = load_workbook(filename=str(workbook_path), read_only=True, data_only=True)
        find_workbook_data_sheet(workbook)
    except Exception:
        return "workbook_unusable"
    finally:
        if workbook is not None:
            workbook.close()
    return None


def _artifact_state_counts(
    validations: list[PromoCampaignArtifactValidation],
) -> dict[str, int]:
    counts = {state: 0 for state in ARTIFACT_STATES}
    for validation in validations:
        state = validation.artifact_state or ARTIFACT_STATE_UNUSABLE
        counts[state] = counts.get(state, 0) + 1
    return counts


def _validation_field_counts(
    validations: list[PromoCampaignArtifactValidation],
    field_name: str,
) -> dict[str, int]:
    counts: dict[str, int] = {}
    for validation in validations:
        value = str(getattr(validation, field_name, None) or "unknown")
        counts[value] = counts.get(value, 0) + 1
    return counts


def _validate_orphan_archive_dir(archive_dir: Path) -> PromoCampaignArtifactValidation:
    metadata_path = archive_dir / ARCHIVE_METADATA_FILENAME
    expected_workbook_path = archive_dir / ARCHIVE_WORKBOOK_FILENAME
    workbook_exists = expected_workbook_path.exists()
    metadata_payload: dict[str, Any] = {}
    if metadata_path.exists():
        try:
            raw_payload = json.loads(metadata_path.read_text(encoding="utf-8"))
            metadata_payload = raw_payload if isinstance(raw_payload, dict) else {}
        except Exception:
            metadata_payload = {}
    if not metadata_path.exists() and workbook_exists:
        state = ARTIFACT_STATE_WORKBOOK_WITHOUT_METADATA
        reason = "archive_record_and_metadata_missing"
    elif metadata_path.exists() and not workbook_exists:
        state = ARTIFACT_STATE_METADATA_ONLY
        reason = "archive_record_and_workbook_missing"
    else:
        state = ARTIFACT_STATE_INCOMPLETE
        reason = "archive_record_missing"
    file_size = _file_size(expected_workbook_path) if workbook_exists else None
    file_mtime = _file_mtime(expected_workbook_path) if workbook_exists else None
    checksum = _sha256_path(expected_workbook_path) if workbook_exists and file_size else None
    return PromoCampaignArtifactValidation(
        archive_key=archive_dir.name,
        campaign_id=None,
        promo_id=_safe_int_or_none(metadata_payload.get("promo_id")),
        period_id=_safe_int_or_none(metadata_payload.get("period_id")),
        normalized_artifact_key=archive_dir.name,
        artifact_state=state,
        validation_failure_reason=reason,
        expected_workbook_path=str(expected_workbook_path),
        actual_candidate_paths=_candidate_workbook_paths(
            archive_dir=archive_dir,
            record_workbook_path=None,
            expected_workbook_path=expected_workbook_path,
        ),
        metadata_exists=metadata_path.exists(),
        workbook_exists=workbook_exists,
        coverage_dates={
            "promo_start_at": _normalize_optional_text(metadata_payload.get("promo_start_at")),
            "promo_end_at": _normalize_optional_text(metadata_payload.get("promo_end_at")),
        },
        date_confidence=_normalize_optional_text(metadata_payload.get("period_parse_confidence")),
        temporal_classification=_normalize_optional_text(metadata_payload.get("temporal_classification")),
        export_kind=_normalize_optional_text(metadata_payload.get("export_kind")),
        file_size=file_size,
        file_mtime=file_mtime,
        checksum_or_fingerprint=checksum,
        workbook_required=True,
        non_materializable_reason=None,
        ui_status=_normalize_optional_text(metadata_payload.get("ui_status")),
        ui_status_confidence=_normalize_optional_text(metadata_payload.get("ui_status_confidence")),
        ui_status_raw_labels=_normalize_text_list(metadata_payload.get("ui_status_raw_labels")),
        download_action_state=_normalize_optional_text(metadata_payload.get("download_action_state")),
        download_action_evidence=_normalize_optional_text(metadata_payload.get("download_action_evidence")),
        status_evidence_sources=_normalize_text_list(metadata_payload.get("status_evidence_sources")),
        ui_loaded_success=bool(metadata_payload.get("ui_loaded_success", False)),
        campaign_identity_match=bool(metadata_payload.get("campaign_identity_match", False)),
    )


def _safe_int_or_none(value: object) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _metadata_indicates_ended_without_download(metadata: PromoMetadata) -> bool:
    drawer_evidence = (
        str(metadata.ui_status or "") == "ended"
        and str(metadata.ui_status_confidence or "") == "high"
        and str(metadata.download_action_state or "") in {"absent", "disabled"}
        and bool(metadata.ui_loaded_success)
        and bool(metadata.campaign_identity_match)
    )
    timeline_sources = set(metadata.timeline_evidence_sources or [])
    timeline_evidence = (
        str(metadata.timeline_classification_decision or "") == "timeline_non_materializable_expected"
        and str(metadata.timeline_status or "") == "ended"
        and str(metadata.timeline_status_confidence or "") == "high"
        and metadata.drawer_opened is False
        and str(metadata.drawer_skip_reason or "") == "timeline_ended_non_materializable"
        and "timeline_status_label" in timeline_sources
        and "timeline_title" in timeline_sources
        and "timeline_period_text" in timeline_sources
    )
    return drawer_evidence or timeline_evidence


def _normalize_text_list(value: object) -> list[str]:
    if not isinstance(value, list):
        return []
    result: list[str] = []
    for item in value:
        text = str(item or "").strip()
        if text:
            result.append(text)
    return result


def _record_matches_live_metadata(
    *,
    record: PromoCampaignArchiveRecord,
    metadata: PromoMetadata,
) -> bool:
    if metadata.promo_id is not None and record.metadata.promo_id != metadata.promo_id:
        return False
    comparable_fields = (
        "promo_title",
        "promo_period_text",
        "promo_start_at",
        "promo_end_at",
        "period_parse_confidence",
        "temporal_classification",
        "promo_status",
        "promo_status_text",
        "eligible_count",
        "participating_count",
        "excluded_count",
        "source_tab",
        "source_filter_code",
    )
    return all(
        getattr(record.metadata, field_name) == getattr(metadata, field_name)
        for field_name in comparable_fields
    )


def _merge_archive_workbook_rows(
    *,
    candidate_rows_by_nm_id: dict[int, list[PromoCandidateRow]],
    workbook_path: Path,
    requested_nm_ids: Iterable[int],
    campaign_identity: str,
) -> int:
    requested = {int(nm_id) for nm_id in requested_nm_ids}
    workbook = load_workbook(filename=str(workbook_path), read_only=False, data_only=True)
    row_count = 0
    try:
        sheet, header_row_index = find_workbook_data_sheet(workbook)
        for row in iter_workbook_plan_rows(
            sheet=sheet,
            header_row_index=header_row_index,
            requested_nm_ids=requested,
        ):
            candidate_rows_by_nm_id.setdefault(row.nm_id, []).append(
                PromoCandidateRow(
                    nm_id=row.nm_id,
                    campaign_identity=campaign_identity,
                    plan_price=row.plan_price,
                )
            )
            row_count += 1
    finally:
        workbook.close()
    return row_count


def _load_daily_price_truth(
    *,
    runtime_dir: Path,
    snapshot_date: str,
    requested_nm_ids: Iterable[int],
) -> DailyPriceTruthResolution:
    runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
    payload, captured_at = runtime.load_temporal_source_slot_snapshot(
        source_key=PRICES_SOURCE_KEY,
        snapshot_date=snapshot_date,
        snapshot_role=PRICES_ACCEPTED_CURRENT_ROLE,
    )
    if _is_exact_prices_payload(payload, snapshot_date):
        note = "daily_price_source=prices_snapshot.accepted_current_snapshot"
        if captured_at:
            note = f"{note}; daily_price_captured_at={captured_at}"
        price_by_nm_id = _extract_price_by_nm_id(
            payload=payload,
            requested_nm_ids=requested_nm_ids,
        )
        return DailyPriceTruthResolution(
            price_by_nm_id=price_by_nm_id,
            source_note=note,
            captured_at=captured_at,
            fingerprint=_json_fingerprint(price_by_nm_id) if price_by_nm_id else None,
        )

    return DailyPriceTruthResolution(
        price_by_nm_id={},
        source_note="daily_price_source=missing",
    )


def _extract_price_by_nm_id(
    *,
    payload: object,
    requested_nm_ids: Iterable[int],
) -> dict[int, float]:
    requested = {int(nm_id) for nm_id in requested_nm_ids}
    price_by_nm_id: dict[int, float] = {}
    for item in list(getattr(payload, "items", []) or []):
        nm_id = getattr(item, "nm_id", None)
        price = getattr(item, "price_seller_discounted", None)
        if not isinstance(nm_id, int) or nm_id not in requested:
            continue
        if not isinstance(price, (int, float)):
            continue
        price_by_nm_id[nm_id] = float(price)
    return price_by_nm_id


def _is_exact_prices_payload(payload: object | None, snapshot_date: str) -> bool:
    if payload is None:
        return False
    if str(getattr(payload, "snapshot_date", "") or "") != snapshot_date:
        return False
    items = getattr(payload, "items", None)
    return isinstance(items, list)


def _campaign_identity(record: PromoCampaignArchiveRecord) -> str:
    if record.metadata.promo_id is not None and record.metadata.period_id is not None:
        return f"{record.metadata.promo_id}:{record.metadata.period_id}"
    if record.metadata.promo_id is not None:
        return str(record.metadata.promo_id)
    return record.archive_key


def _build_zero_items(
    *,
    snapshot_date: str,
    requested_nm_ids: Iterable[int],
) -> list[PromoLiveSourceItem]:
    return [
        PromoLiveSourceItem(
            snapshot_date=snapshot_date,
            nm_id=nm_id,
            promo_count_by_price=0.0,
            promo_entry_price_best=0.0,
            promo_participation=0.0,
        )
        for nm_id in sorted({int(nm_id) for nm_id in requested_nm_ids})
    ]


def _build_item_from_candidate_rows(
    *,
    snapshot_date: str,
    nm_id: int,
    candidate_rows: list[PromoCandidateRow],
    price_seller_discounted: float | None,
) -> tuple[PromoLiveSourceItem, PromoEligibilityEvaluation]:
    evaluation = evaluate_candidate_rows(
        candidate_rows=candidate_rows,
        price_seller_discounted=price_seller_discounted,
    )
    return (
        PromoLiveSourceItem(
            snapshot_date=snapshot_date,
            nm_id=nm_id,
            promo_count_by_price=round(evaluation.promo_count_by_price, 6),
            promo_entry_price_best=round(evaluation.promo_entry_price_best, 6),
            promo_participation=round(evaluation.promo_participation, 6),
        ),
        evaluation,
    )


def _promo_metric_coverage_detail_parts(
    *,
    eligible_nm_ids: list[int],
    truthful_zero_no_candidate_nm_ids: list[int],
    truthful_zero_ineligible_nm_ids: list[int],
    missing_price_truth_nm_ids: list[int],
) -> list[str]:
    return [
        f"promo_metric_eligible_nm_ids={_format_nm_id_sample(eligible_nm_ids)}",
        (
            "promo_metric_truthful_zero_no_candidate_nm_ids="
            f"{_format_nm_id_sample(truthful_zero_no_candidate_nm_ids)}"
        ),
        (
            "promo_metric_truthful_zero_ineligible_nm_ids="
            f"{_format_nm_id_sample(truthful_zero_ineligible_nm_ids)}"
        ),
        (
            "promo_metric_missing_price_truth_nm_ids="
            f"{_format_nm_id_sample(missing_price_truth_nm_ids)}"
        ),
    ]


def _format_nm_id_sample(nm_ids: list[int], *, limit: int = 20) -> str:
    unique = sorted({int(nm_id) for nm_id in nm_ids})
    if not unique:
        return ""
    sample = ",".join(str(nm_id) for nm_id in unique[:limit])
    if len(unique) > limit:
        sample = f"{sample},...(+{len(unique) - limit})"
    return sample


def _start_promo_diag_phase(
    diagnostics: dict[str, Any] | None,
    phase_key: str,
) -> dict[str, Any] | None:
    if diagnostics is None:
        return None
    return {
        "phase_key": phase_key,
        "started_at": _diag_now_iso(),
        "started_perf": time.perf_counter(),
    }


def _finish_promo_diag_phase(
    diagnostics: dict[str, Any] | None,
    phase: dict[str, Any] | None,
    *,
    status: str,
    note_kind: str | None = None,
    error_kind: str | None = None,
) -> None:
    if diagnostics is None or phase is None:
        return
    item = {
        "phase_key": str(phase.get("phase_key") or ""),
        "started_at": str(phase.get("started_at") or ""),
        "finished_at": _diag_now_iso(),
        "duration_ms": max(0, int(round((time.perf_counter() - float(phase.get("started_perf") or time.perf_counter())) * 1000))),
        "status": status,
    }
    if note_kind:
        item["note_kind"] = note_kind
    if error_kind:
        item["error_kind"] = error_kind
    diagnostics.setdefault("phase_summary", []).append(item)


def _finish_source_payload_build_phase(
    diagnostics: dict[str, Any] | None,
    *,
    status: str,
    note_kind: str | None = None,
) -> None:
    phase = _start_promo_diag_phase(diagnostics, "source_payload_build")
    _finish_promo_diag_phase(diagnostics, phase, status=status, note_kind=note_kind)


def _set_promo_diag_counter(
    diagnostics: dict[str, Any] | None,
    key: str,
    value: Any,
    *,
    overwrite: bool = True,
) -> None:
    if diagnostics is None:
        return
    counters = diagnostics.setdefault("counters", {})
    if not isinstance(counters, dict):
        return
    if overwrite or key not in counters or counters.get(key) is None:
        counters[key] = value


def _set_promo_diag_fingerprint(
    diagnostics: dict[str, Any] | None,
    key: str,
    value: Any,
) -> None:
    if diagnostics is None:
        return
    fingerprints = diagnostics.setdefault("fingerprints", {})
    if not isinstance(fingerprints, dict):
        return
    fingerprints[key] = value


def _snapshot_promo_diagnostics(diagnostics: dict[str, Any] | None) -> dict[str, Any]:
    if diagnostics is None:
        return {}
    return dict(diagnostics)


def _archive_records_fingerprint(records: list[PromoCampaignArchiveRecord]) -> str | None:
    payload = [
        {
            "archive_key": record.archive_key,
            "metadata_fingerprint": record.metadata_fingerprint,
            "workbook_fingerprint": record.workbook_fingerprint,
            "workbook_present": record.workbook_present,
        }
        for record in records
    ]
    return _json_fingerprint(payload) if payload else None


def _record_field_fingerprint(records: list[PromoCampaignArchiveRecord], field_name: str) -> str | None:
    values = [
        str(getattr(record, field_name, "") or "")
        for record in records
        if str(getattr(record, field_name, "") or "").strip()
    ]
    return _json_fingerprint(sorted(values)) if values else None


def _workbook_file_size_total(records: list[PromoCampaignArchiveRecord]) -> int | None:
    total = 0
    seen = False
    for record in records:
        path = Path(record.workbook_path) if record.workbook_path else None
        if path is None or not path.exists():
            continue
        seen = True
        total += path.stat().st_size
    return total if seen else None


def _workbook_file_mtime_max(records: list[PromoCampaignArchiveRecord]) -> int | None:
    mtimes: list[int] = []
    for record in records:
        path = Path(record.workbook_path) if record.workbook_path else None
        if path is None or not path.exists():
            continue
        mtimes.append(int(path.stat().st_mtime))
    return max(mtimes) if mtimes else None


def _json_fingerprint(value: Any) -> str:
    raw = json.dumps(value, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _diag_now_iso() -> str:
    return datetime.now().astimezone().replace(microsecond=0).isoformat()


def _metadata_fingerprint(metadata: PromoMetadata) -> str:
    stable_payload = {
        "promo_id": metadata.promo_id,
        "period_id": metadata.period_id,
        "promo_title": metadata.promo_title,
        "promo_period_text": metadata.promo_period_text,
        "promo_start_at": metadata.promo_start_at,
        "promo_end_at": metadata.promo_end_at,
        "period_parse_confidence": metadata.period_parse_confidence,
        "temporal_classification": metadata.temporal_classification,
        "promo_status": metadata.promo_status,
        "promo_status_text": metadata.promo_status_text,
        "eligible_count": metadata.eligible_count,
        "participating_count": metadata.participating_count,
        "excluded_count": metadata.excluded_count,
        "source_tab": metadata.source_tab,
        "source_filter_code": metadata.source_filter_code,
        "export_kind": metadata.export_kind,
        "ui_status": metadata.ui_status,
        "ui_status_confidence": metadata.ui_status_confidence,
        "ui_status_raw_labels": metadata.ui_status_raw_labels,
        "download_action_state": metadata.download_action_state,
        "download_action_evidence": metadata.download_action_evidence,
        "status_evidence_sources": metadata.status_evidence_sources,
        "ui_loaded_success": metadata.ui_loaded_success,
        "campaign_identity_match": metadata.campaign_identity_match,
        "collector_ui_schema_version": metadata.collector_ui_schema_version,
        "timeline_status": metadata.timeline_status,
        "timeline_status_confidence": metadata.timeline_status_confidence,
        "timeline_status_raw_labels": metadata.timeline_status_raw_labels,
        "timeline_evidence_sources": metadata.timeline_evidence_sources,
        "timeline_period_text": metadata.timeline_period_text,
        "timeline_goods_count": metadata.timeline_goods_count,
        "timeline_autoaction_marker": metadata.timeline_autoaction_marker,
        "timeline_classification_decision": metadata.timeline_classification_decision,
        "drawer_opened": metadata.drawer_opened,
        "drawer_open_reason": metadata.drawer_open_reason,
        "drawer_skip_reason": metadata.drawer_skip_reason,
        "timeline_classifier_schema_version": metadata.timeline_classifier_schema_version,
    }
    raw = json.dumps(stable_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _archive_key(metadata: PromoMetadata) -> str:
    promo_part = str(metadata.promo_id) if metadata.promo_id is not None else "pending"
    period_part = str(metadata.period_id) if metadata.period_id is not None else "pending"
    return f"{promo_part}__{period_part}__{_slugify(metadata.promo_title)}"


def _slugify(value: str) -> str:
    normalized = re.sub(r"[^\w]+", "-", str(value or "").lower(), flags=re.UNICODE)
    normalized = normalized.strip("-")
    normalized = re.sub(r"-{2,}", "-", normalized)
    return normalized or "pending"


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalized_path(path: Path) -> str:
    try:
        return str(path.resolve(strict=False))
    except OSError:
        return str(path)


def _normalize_optional_text(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None
