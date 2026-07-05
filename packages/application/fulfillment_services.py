"""Fulfillment service XLSX upload/payment-validation block for supply operator UI."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
import hashlib
import json
from pathlib import Path
import re
import sqlite3
from typing import Any, Mapping
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime


CONTRACT_NAME = "sheet_vitrina_v1_fulfillment_services"
CONTRACT_VERSION = "v1"
FULFILLMENT_SERVICES_BASE_PATH = "/v1/sheet-vitrina-v1/supply/fulfillment-services"
FULFILLMENT_TEMPLATE_FILENAME = "sheet-vitrina-v1-fulfillment-services-template.xlsx"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_CONTENT_TYPE = "application/pdf"
UPLOADS_TABLE = "sheet_vitrina_v1_fulfillment_service_uploads"
LINES_TABLE = "sheet_vitrina_v1_fulfillment_service_lines"

TEMPLATE_HEADERS = [
    "Номер поставки",
    "Стоимость услуг",
    "Кол-во коробов",
    "Цена",
    "Кол-во паллет",
    "Цена",
    "",
    "Выезд",
    "Итого",
    "НДС 5%",
]

HEADER_SUPPLY_ID = "Номер поставки"
HEADER_SERVICE_NAME = "Стоимость услуг"
HEADER_BOXES_QTY = "Кол-во коробов"
HEADER_PALLETS_QTY = "Кол-во паллет"
HEADER_DEPARTURE = "Выезд"
HEADER_TOTAL = "Итого"
HEADER_VAT = "НДС 5%"

VALIDATION_OK = "ok"
VALIDATION_FAILED = "failed"
MATCH_OK = "ok"
MATCH_UNMATCHED = "unmatched"
MATCH_DUPLICATE = "duplicate"
MATCH_INVALID = "invalid"


class FulfillmentServicesError(RuntimeError):
    """Controlled fulfillment-services application error."""


@dataclass
class _ParsedLine:
    row_index: int
    supply_id_input: str
    service_name: str
    route: str
    boxes_qty: Decimal | None
    box_price: Decimal | None
    boxes_amount: Decimal | None
    pallets_qty: Decimal | None
    pallet_price: Decimal | None
    pallets_amount: Decimal | None
    services_subtotal_amount: Decimal | None
    departure_amount: Decimal | None
    amount_without_vat: Decimal | None
    vat_amount: Decimal | None
    amount_with_vat: Decimal | None
    raw_row: dict[str, Any]
    row_error: str
    row_warnings: list[str]
    match_status: str = MATCH_INVALID
    matched_wb_supply_id: str = ""
    matched_wb_cache_key: str = ""


class FulfillmentServicesBlock:
    """Server-owned Fulfillment services runtime block."""

    def __init__(
        self,
        *,
        runtime: RegistryUploadDbBackedRuntime,
        timestamp_factory: Any | None = None,
    ) -> None:
        self.runtime = runtime
        self.timestamp_factory = timestamp_factory or _default_timestamp_factory

    def build_template(self) -> tuple[bytes, str, str]:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Fulfillment"
        header_fill = PatternFill("solid", fgColor="E8C9D7")
        header_font = Font(bold=True, italic=True, color="1F1F23")
        for index, header in enumerate(TEMPLATE_HEADERS, start=1):
            cell = sheet.cell(row=1, column=index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [18, 28, 16, 12, 16, 12, 14, 12, 14, 12]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = True
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue(), FULFILLMENT_TEMPLATE_FILENAME, XLSX_CONTENT_TYPE

    def upload_xlsx(
        self,
        workbook_bytes: bytes,
        *,
        uploaded_filename: str | None = None,
        uploaded_content_type: str | None = None,
    ) -> dict[str, Any]:
        del uploaded_content_type
        if not workbook_bytes:
            raise ValueError("Fulfillment XLSX file is empty")
        filename = _safe_filename(uploaded_filename or "fulfillment-services.xlsx")
        if Path(filename).suffix.lower() != ".xlsx":
            raise ValueError("Fulfillment upload accepts .xlsx files only")
        now = self.timestamp_factory()
        upload_id = "ffu_" + uuid4().hex[:16]
        file_sha256 = hashlib.sha256(workbook_bytes).hexdigest()
        stored_file_path = self._store_uploaded_file(upload_id, filename, workbook_bytes)
        parsed_lines, parse_errors = self._parse_workbook(workbook_bytes)
        validated_lines, validation_errors = self._validate_lines(parsed_lines)
        all_errors = [*parse_errors, *validation_errors]
        rows_total = len(validated_lines)
        rows_matched = sum(1 for line in validated_lines if line.match_status == MATCH_OK)
        totals = _line_totals(validated_lines)
        validation_status = VALIDATION_OK if rows_total > 0 and not all_errors else VALIDATION_FAILED
        if rows_total <= 0:
            validation_status = VALIDATION_FAILED
            if not any("detail row" in item for item in all_errors):
                all_errors.append("XLSX must contain at least one detail row")

        payment_validation_id = ""
        pdf_file_path = ""
        if validation_status == VALIDATION_OK:
            payment_validation_id = self._new_payment_validation_id()
            pdf_bytes = self._build_payment_validation_pdf(
                upload_id=upload_id,
                payment_validation_id=payment_validation_id,
                source_filename=filename,
                file_sha256=file_sha256,
                generated_at=now,
                rows=validated_lines,
                totals=totals,
            )
            pdf_file_path = self._store_pdf_file(upload_id, payment_validation_id, pdf_bytes)

        self._save_upload(
            upload_id=upload_id,
            filename=filename,
            stored_file_path=stored_file_path,
            file_sha256=file_sha256,
            uploaded_at=now,
            validation_status=validation_status,
            validation_error_summary="; ".join(all_errors[:12]),
            rows_total=rows_total,
            rows_matched=rows_matched,
            totals=totals,
            payment_validation_id=payment_validation_id,
            pdf_file_path=pdf_file_path,
            created_at=now,
            updated_at=now,
            lines=validated_lines,
        )
        return self.get_upload(upload_id)

    def list_uploads(self, *, limit: int = 20) -> dict[str, Any]:
        normalized_limit = max(1, min(int(limit or 20), 100))
        with self._connect() as conn:
            _ensure_schema(conn)
            rows = conn.execute(
                f"""
                SELECT *
                FROM {UPLOADS_TABLE}
                WHERE validation_status = ?
                  AND deleted_at IS NULL
                ORDER BY uploaded_at DESC, created_at DESC, upload_id DESC
                LIMIT ?
                """,
                (VALIDATION_OK, normalized_limit),
            ).fetchall()
        uploads = [_upload_row_to_dict(row, include_links=True) for row in rows]
        latest = uploads[0] if uploads else None
        return {
            "contract_name": CONTRACT_NAME,
            "contract_version": CONTRACT_VERSION,
            "uploads": uploads,
            "latest_upload": latest,
            "summary": {
                "returned_uploads": len(uploads),
                "latest_status": latest.get("validation_status") if latest else "",
                "latest_upload_id": latest.get("upload_id") if latest else "",
                "accepted_uploads": len(uploads),
            },
        }

    def get_upload(self, upload_id: str) -> dict[str, Any]:
        normalized_id = str(upload_id or "").strip()
        if not normalized_id:
            raise ValueError("upload_id is required")
        with self._connect() as conn:
            _ensure_schema(conn)
            upload_row = conn.execute(
                f"SELECT * FROM {UPLOADS_TABLE} WHERE upload_id = ? AND deleted_at IS NULL",
                (normalized_id,),
            ).fetchone()
            if upload_row is None:
                raise ValueError(f"Fulfillment upload not found: {normalized_id}")
            line_rows = conn.execute(
                f"""
                SELECT *
                FROM {LINES_TABLE}
                WHERE upload_id = ?
                ORDER BY row_index ASC
                """,
                (normalized_id,),
            ).fetchall()
        upload = _upload_row_to_dict(upload_row, include_links=True)
        lines = [_line_row_to_dict(row) for row in line_rows]
        return {
            "contract_name": CONTRACT_NAME,
            "contract_version": CONTRACT_VERSION,
            "upload": upload,
            "lines": lines,
            "validation_status": upload["validation_status"],
            "row_errors": [
                {
                    "row_index": line["row_index"],
                    "supply_id_input": line["supply_id_input"],
                    "match_status": line["match_status"],
                    "row_error": line["row_error"],
                }
                for line in lines
                if line.get("row_error")
            ],
            "totals": {
                "amount_without_vat_total": upload["amount_without_vat_total"],
                "vat_total": upload["vat_total"],
                "amount_with_vat_total": upload["amount_with_vat_total"],
                "rows_total": upload["rows_total"],
                "rows_matched": upload["rows_matched"],
            },
            "pdf_available": bool(upload.get("pdf_available")),
            "pdf_path": upload.get("pdf_path") or "",
        }

    def download_pdf(self, upload_id: str) -> tuple[bytes, str, str]:
        payload = self.get_upload(upload_id)
        upload = payload["upload"]
        if upload.get("validation_status") != VALIDATION_OK or not upload.get("pdf_file_path"):
            raise ValueError("PDF payment validation is available only for fully valid uploads")
        pdf_path = self._runtime_path(str(upload.get("pdf_file_path") or ""))
        if not pdf_path.is_file():
            raise ValueError("PDF payment validation file is missing in runtime storage")
        filename = f"fulfillment-payment-validation-{upload['upload_id']}.pdf"
        return pdf_path.read_bytes(), filename, PDF_CONTENT_TYPE

    def delete_upload(
        self,
        upload_id: str,
        *,
        deleted_by: str = "",
        delete_reason: str = "operator_delete",
    ) -> dict[str, Any]:
        normalized_id = str(upload_id or "").strip()
        if not normalized_id:
            raise ValueError("upload_id is required")
        now = self.timestamp_factory()
        pdf_file_path = ""
        already_deleted = False
        with self._connect() as conn:
            _ensure_schema(conn)
            upload_row = conn.execute(
                f"SELECT * FROM {UPLOADS_TABLE} WHERE upload_id = ?",
                (normalized_id,),
            ).fetchone()
            if upload_row is None:
                raise ValueError(f"Fulfillment upload not found: {normalized_id}")
            pdf_file_path = str(upload_row["pdf_file_path"] or "")
            already_deleted = bool(upload_row["deleted_at"])
            if already_deleted:
                deleted_at = str(upload_row["deleted_at"] or "")
                if pdf_file_path:
                    conn.execute(
                        f"""
                        UPDATE {UPLOADS_TABLE}
                        SET pdf_file_path = '',
                            updated_at = ?
                        WHERE upload_id = ?
                        """,
                        (now, normalized_id),
                    )
            else:
                deleted_at = now
                conn.execute(
                    f"""
                    UPDATE {UPLOADS_TABLE}
                    SET deleted_at = ?,
                        deleted_by = ?,
                        delete_reason = ?,
                        pdf_file_path = '',
                        updated_at = ?
                    WHERE upload_id = ?
                    """,
                    (
                        deleted_at,
                        str(deleted_by or "").strip(),
                        str(delete_reason or "").strip(),
                        now,
                        normalized_id,
                    ),
                )
            conn.commit()
        if pdf_file_path:
            try:
                self._runtime_path(pdf_file_path).unlink(missing_ok=True)
            except OSError:
                pass
        return {
            "contract_name": CONTRACT_NAME,
            "contract_version": CONTRACT_VERSION,
            "upload_id": normalized_id,
            "deleted": True,
            "already_deleted": already_deleted,
            "deleted_at": deleted_at,
            "soft_deleted": True,
            "lines_deactivated": True,
            "pdf_available": False,
            "message": "Документ удалён. Данные услуг фулфилмента удалены из overlay WB-поставок.",
        }

    def approved_overlay_by_supply(self) -> dict[str, dict[str, Any]]:
        with self._connect() as conn:
            _ensure_schema(conn)
            rows = conn.execute(
                f"""
                SELECT line.*, upload.payment_validation_id, upload.uploaded_at
                FROM {LINES_TABLE} AS line
                JOIN {UPLOADS_TABLE} AS upload ON upload.upload_id = line.upload_id
                WHERE upload.validation_status = ?
                  AND upload.deleted_at IS NULL
                  AND line.match_status = ?
                ORDER BY upload.uploaded_at ASC, line.row_index ASC
                """,
                (VALIDATION_OK, MATCH_OK),
            ).fetchall()
        grouped: dict[str, dict[str, Any]] = {}
        by_canonical: dict[str, dict[str, Any]] = {}
        for row in rows:
            canonical = str(row["matched_wb_cache_key"] or row["matched_wb_supply_id"] or row["supply_id_input"] or "").strip()
            if not canonical:
                continue
            item = by_canonical.setdefault(
                canonical,
                {
                    "amount_without_vat_total": 0.0,
                    "vat_total": 0.0,
                    "amount_with_vat_total": 0.0,
                    "upload_ids": [],
                    "payment_validation_ids": [],
                    "service_names": [],
                    "line_count": 0,
                    "uploaded_at": str(row["uploaded_at"] or ""),
                },
            )
            item["amount_without_vat_total"] += float(row["amount_without_vat"] or 0)
            item["vat_total"] += float(row["vat_amount"] or 0)
            item["amount_with_vat_total"] += float(row["amount_with_vat"] or 0)
            item["line_count"] += 1
            _append_unique(item["upload_ids"], str(row["upload_id"] or ""))
            _append_unique(item["payment_validation_ids"], str(row["payment_validation_id"] or ""))
            _append_unique(item["service_names"], str(row["service_name"] or ""))
            for identity in _line_identity_values(row):
                grouped[identity] = item
        return grouped

    def _parse_workbook(self, workbook_bytes: bytes) -> tuple[list[_ParsedLine], list[str]]:
        try:
            workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl owns exact exception types
            return [], [f"XLSX parse failed: {exc}"]
        sheet = workbook.worksheets[0]
        header_row, columns = _detect_header_row(sheet)
        if header_row <= 0:
            return [], ["Header row with Номер поставки, Итого and НДС 5% was not found"]
        parsed: list[_ParsedLine] = []
        errors: list[str] = []
        for row_index in range(header_row + 1, sheet.max_row + 1):
            values = [sheet.cell(row=row_index, column=column).value for column in range(1, sheet.max_column + 1)]
            if _row_is_empty(values):
                continue
            raw_row = _raw_row_payload(row_index, values, columns)
            supply_id = _cell_text(_value_at(values, columns["supply_id"]))
            if not supply_id and _row_looks_like_footer(values, columns):
                continue
            line = _line_from_values(row_index, values, columns, raw_row)
            parsed.append(line)
        return parsed, errors

    def _validate_lines(self, lines: list[_ParsedLine]) -> tuple[list[_ParsedLine], list[str]]:
        errors: list[str] = []
        seen_supply_ids: set[str] = set()
        for line in lines:
            line_errors: list[str] = []
            if not line.supply_id_input:
                line.match_status = MATCH_INVALID
                line_errors.append("Номер поставки is required")
            if line.amount_without_vat is None:
                line_errors.append("Итого must be numeric")
            elif line.amount_without_vat < 0:
                line_errors.append("Итого must be >= 0")
            if line.vat_amount is None:
                line_errors.append("НДС 5% must be numeric")
            elif line.vat_amount < 0:
                line_errors.append("НДС 5% must be >= 0")
            if line.amount_without_vat is not None and line.vat_amount is not None:
                line.amount_with_vat = _money(line.amount_without_vat + line.vat_amount)
            normalized_supply = _normalize_supply_identity(line.supply_id_input)
            if normalized_supply and normalized_supply in seen_supply_ids:
                line.match_status = MATCH_DUPLICATE
                line_errors.append("Duplicate Номер поставки inside upload")
            elif normalized_supply:
                seen_supply_ids.add(normalized_supply)

            if line.supply_id_input and line.match_status != MATCH_DUPLICATE:
                match = self.runtime.load_wb_supply_record(line.supply_id_input)
                if match is None:
                    line.match_status = MATCH_UNMATCHED
                    line_errors.append("Номер поставки does not match cached WB supply")
                else:
                    normalized = match.get("normalized") if isinstance(match.get("normalized"), Mapping) else {}
                    line.matched_wb_supply_id = str(
                        normalized.get("wb_supply_id")
                        or match.get("wb_supply_id")
                        or normalized.get("visible_number")
                        or normalized.get("supply_id")
                        or line.supply_id_input
                    ).strip()
                    line.matched_wb_cache_key = str(
                        match.get("cache_key")
                        or normalized.get("cache_key")
                        or normalized.get("supply_id")
                        or line.matched_wb_supply_id
                    ).strip()
                    if not line_errors:
                        line.match_status = MATCH_OK
            if line_errors:
                if line.match_status == MATCH_OK:
                    line.match_status = MATCH_INVALID
                line.row_error = "; ".join(line_errors)
                errors.append(f"row {line.row_index}: {line.row_error}")
            elif line.match_status != MATCH_OK:
                line.row_error = f"match_status={line.match_status}"
                errors.append(f"row {line.row_index}: {line.row_error}")
        return lines, errors

    def _save_upload(
        self,
        *,
        upload_id: str,
        filename: str,
        stored_file_path: str,
        file_sha256: str,
        uploaded_at: str,
        validation_status: str,
        validation_error_summary: str,
        rows_total: int,
        rows_matched: int,
        totals: Mapping[str, Decimal],
        payment_validation_id: str,
        pdf_file_path: str,
        created_at: str,
        updated_at: str,
        lines: list[_ParsedLine],
    ) -> None:
        with self._connect() as conn:
            _ensure_schema(conn)
            conn.execute(
                f"""
                INSERT INTO {UPLOADS_TABLE}(
                    upload_id,
                    original_filename,
                    stored_file_path,
                    file_sha256,
                    uploaded_at,
                    validation_status,
                    validation_error_summary,
                    rows_total,
                    rows_matched,
                    amount_without_vat_total,
                    vat_total,
                    amount_with_vat_total,
                    payment_validation_id,
                    pdf_file_path,
                    created_at,
                    updated_at
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    upload_id,
                    filename,
                    stored_file_path,
                    file_sha256,
                    uploaded_at,
                    validation_status,
                    validation_error_summary,
                    rows_total,
                    rows_matched,
                    _decimal_to_float(totals["amount_without_vat_total"]),
                    _decimal_to_float(totals["vat_total"]),
                    _decimal_to_float(totals["amount_with_vat_total"]),
                    payment_validation_id,
                    pdf_file_path,
                    created_at,
                    updated_at,
                ),
            )
            conn.executemany(
                f"""
                INSERT INTO {LINES_TABLE}(
                    upload_id,
                    row_index,
                    supply_id_input,
                    matched_wb_supply_id,
                    matched_wb_cache_key,
                    match_status,
                    service_name,
                    route,
                    boxes_qty,
                    box_price,
                    boxes_amount,
                    pallets_qty,
                    pallet_price,
                    pallets_amount,
                    services_subtotal_amount,
                    departure_amount,
                    amount_without_vat,
                    vat_amount,
                    amount_with_vat,
                    raw_row_json,
                    row_error,
                    row_warnings_json,
                    created_at
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [_line_insert_values(upload_id, line, created_at) for line in lines],
            )
            conn.commit()

    def _new_payment_validation_id(self) -> str:
        with self._connect() as conn:
            _ensure_schema(conn)
            for _ in range(100):
                candidate = "FF-" + uuid4().hex[:8].upper()
                row = conn.execute(
                    f"SELECT 1 FROM {UPLOADS_TABLE} WHERE payment_validation_id = ? LIMIT 1",
                    (candidate,),
                ).fetchone()
                if row is None:
                    return candidate
        raise RuntimeError("failed to allocate unique payment_validation_id")

    def _store_uploaded_file(self, upload_id: str, filename: str, payload: bytes) -> str:
        target_dir = self.runtime.runtime_dir / "fulfillment_services" / "uploads" / upload_id
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / filename
        target_path.write_bytes(payload)
        return _relative_to_runtime(self.runtime.runtime_dir, target_path)

    def _store_pdf_file(self, upload_id: str, payment_validation_id: str, payload: bytes) -> str:
        target_dir = self.runtime.runtime_dir / "fulfillment_services" / "payment_validations" / upload_id
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / f"{payment_validation_id}.pdf"
        target_path.write_bytes(payload)
        return _relative_to_runtime(self.runtime.runtime_dir, target_path)

    def _runtime_path(self, relative_path: str) -> Path:
        root = self.runtime.runtime_dir.resolve()
        target = (root / relative_path).resolve()
        if root not in target.parents and target != root:
            raise ValueError("runtime file path escapes runtime dir")
        return target

    def _connect(self) -> sqlite3.Connection:
        self.runtime.runtime_dir.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(self.runtime.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def _build_payment_validation_pdf(
        self,
        *,
        upload_id: str,
        payment_validation_id: str,
        source_filename: str,
        file_sha256: str,
        generated_at: str,
        rows: list[_ParsedLine],
        totals: Mapping[str, Decimal],
    ) -> bytes:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except Exception as exc:  # pragma: no cover - dependency boundary
            raise RuntimeError("reportlab is required to generate Fulfillment payment validation PDF") from exc

        font_name = _register_pdf_font(pdfmetrics, TTFont)
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
            title="Виза на оплату Fulfillment-услуг",
        )
        styles = getSampleStyleSheet()
        for style in styles.byName.values():
            style.fontName = font_name
        title_style = styles["Title"]
        title_style.fontName = font_name
        body_style = styles["BodyText"]
        body_style.fontName = font_name
        story: list[Any] = [
            Paragraph("Виза на оплату Fulfillment-услуг", title_style),
            Spacer(1, 4 * mm),
        ]
        meta_rows = [
            ("Статус", "Проверено системой / OK"),
            ("Generated at", generated_at),
            ("upload_id", upload_id),
            ("payment_validation_id", payment_validation_id),
            ("Source filename", source_filename),
            ("Short file hash", file_sha256[:12]),
            ("Rows total", str(len(rows))),
            ("Rows matched", str(len(rows))),
            ("Total Итого", _format_money(totals["amount_without_vat_total"])),
            ("Total НДС 5%", _format_money(totals["vat_total"])),
            ("К оплате = Итого + НДС 5%", _format_money(totals["amount_with_vat_total"])),
        ]
        meta_table = Table([[Paragraph(k, body_style), Paragraph(v, body_style)] for k, v in meta_rows], colWidths=[55 * mm, 115 * mm])
        meta_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7BEC8")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.extend([meta_table, Spacer(1, 5 * mm)])
        table_rows = [["Номер поставки", "Стоимость услуг / направление", "Итого", "НДС 5%", "К оплате"]]
        for line in rows:
            table_rows.append(
                [
                    line.supply_id_input,
                    line.service_name or line.route or "—",
                    _format_money(line.amount_without_vat),
                    _format_money(line.vat_amount),
                    _format_money(line.amount_with_vat),
                ]
            )
        supplies_table = Table(table_rows, colWidths=[28 * mm, 70 * mm, 28 * mm, 25 * mm, 30 * mm], repeatRows=1)
        supplies_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7BEC8")),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8C9D7")),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.extend(
            [
                supplies_table,
                Spacer(1, 5 * mm),
                Paragraph(
                    "PDF действителен только для загруженного файла с указанным upload_id и hash",
                    body_style,
                ),
            ]
        )
        doc.build(story)
        return buffer.getvalue()


def _detect_header_row(sheet: Any) -> tuple[int, dict[str, int]]:
    for row_index in range(1, min(sheet.max_row, 50) + 1):
        values = [sheet.cell(row=row_index, column=column).value for column in range(1, sheet.max_column + 1)]
        normalized = [_normalize_header(value) for value in values]
        if not (
            _normalize_header(HEADER_SUPPLY_ID) in normalized
            and _normalize_header(HEADER_TOTAL) in normalized
            and _normalize_header(HEADER_VAT) in normalized
        ):
            continue
        columns = _header_columns(normalized)
        if columns:
            return row_index, columns
    return 0, {}


def _header_columns(normalized_headers: list[str]) -> dict[str, int]:
    def first(label: str) -> int:
        normalized = _normalize_header(label)
        try:
            return normalized_headers.index(normalized) + 1
        except ValueError:
            return 0

    supply_col = first(HEADER_SUPPLY_ID)
    total_col = first(HEADER_TOTAL)
    vat_col = first(HEADER_VAT)
    service_col = first(HEADER_SERVICE_NAME)
    boxes_col = first(HEADER_BOXES_QTY)
    pallets_col = first(HEADER_PALLETS_QTY)
    departure_col = first(HEADER_DEPARTURE)
    price_cols = [index + 1 for index, value in enumerate(normalized_headers) if value == _normalize_header("Цена")]
    box_price_col = next((col for col in price_cols if boxes_col and col > boxes_col), price_cols[0] if price_cols else 0)
    pallet_price_col = next((col for col in price_cols if pallets_col and col > pallets_col), price_cols[1] if len(price_cols) > 1 else 0)
    subtotal_col = 0
    if pallet_price_col and departure_col:
        for col in range(pallet_price_col + 1, departure_col):
            subtotal_col = col
            break
    if not (supply_col and total_col and vat_col):
        return {}
    return {
        "supply_id": supply_col,
        "service_name": service_col,
        "boxes_qty": boxes_col,
        "box_price": box_price_col,
        "pallets_qty": pallets_col,
        "pallet_price": pallet_price_col,
        "services_subtotal_amount": subtotal_col,
        "departure_amount": departure_col,
        "amount_without_vat": total_col,
        "vat_amount": vat_col,
    }


def _line_from_values(row_index: int, values: list[Any], columns: Mapping[str, int], raw_row: dict[str, Any]) -> _ParsedLine:
    supply_id = _cell_text(_value_at(values, columns["supply_id"]))
    service_name = _cell_text(_value_at(values, columns.get("service_name", 0)))
    boxes_qty = _parse_decimal(_value_at(values, columns.get("boxes_qty", 0)))
    box_price = _parse_decimal(_value_at(values, columns.get("box_price", 0)))
    pallets_qty = _parse_decimal(_value_at(values, columns.get("pallets_qty", 0)))
    pallet_price = _parse_decimal(_value_at(values, columns.get("pallet_price", 0)))
    subtotal_amount = _parse_decimal(_value_at(values, columns.get("services_subtotal_amount", 0)))
    departure_amount = _parse_decimal(_value_at(values, columns.get("departure_amount", 0)))
    amount_without_vat = _parse_decimal(_value_at(values, columns["amount_without_vat"]))
    vat_amount = _parse_decimal(_value_at(values, columns["vat_amount"]))
    boxes_amount = _money(boxes_qty * box_price) if boxes_qty is not None and box_price is not None else None
    pallets_amount = _money(pallets_qty * pallet_price) if pallets_qty is not None and pallet_price is not None else None
    amount_with_vat = _money(amount_without_vat + vat_amount) if amount_without_vat is not None and vat_amount is not None else None
    return _ParsedLine(
        row_index=row_index,
        supply_id_input=supply_id,
        service_name=service_name,
        route=service_name,
        boxes_qty=boxes_qty,
        box_price=box_price,
        boxes_amount=boxes_amount,
        pallets_qty=pallets_qty,
        pallet_price=pallet_price,
        pallets_amount=pallets_amount,
        services_subtotal_amount=subtotal_amount,
        departure_amount=departure_amount,
        amount_without_vat=amount_without_vat,
        vat_amount=vat_amount,
        amount_with_vat=amount_with_vat,
        raw_row=raw_row,
        row_error="",
        row_warnings=[],
    )


def _raw_row_payload(row_index: int, values: list[Any], columns: Mapping[str, int]) -> dict[str, Any]:
    inverse = {index: key for key, index in columns.items() if index}
    cells: list[dict[str, Any]] = []
    for index, value in enumerate(values, start=1):
        if value is None or str(value).strip() == "":
            continue
        cells.append(
            {
                "column_index": index,
                "column_key": inverse.get(index, f"col_{index}"),
                "value": _json_value(value),
            }
        )
    return {"row_index": row_index, "cells": cells}


def _line_totals(lines: list[_ParsedLine]) -> dict[str, Decimal]:
    amount_without_vat_total = sum((line.amount_without_vat or Decimal("0")) for line in lines)
    vat_total = sum((line.vat_amount or Decimal("0")) for line in lines)
    amount_with_vat_total = sum((line.amount_with_vat or Decimal("0")) for line in lines)
    return {
        "amount_without_vat_total": _money(amount_without_vat_total),
        "vat_total": _money(vat_total),
        "amount_with_vat_total": _money(amount_with_vat_total),
    }


def _line_insert_values(upload_id: str, line: _ParsedLine, created_at: str) -> tuple[Any, ...]:
    return (
        upload_id,
        line.row_index,
        line.supply_id_input,
        line.matched_wb_supply_id,
        line.matched_wb_cache_key,
        line.match_status,
        line.service_name,
        line.route,
        _decimal_to_float(line.boxes_qty),
        _decimal_to_float(line.box_price),
        _decimal_to_float(line.boxes_amount),
        _decimal_to_float(line.pallets_qty),
        _decimal_to_float(line.pallet_price),
        _decimal_to_float(line.pallets_amount),
        _decimal_to_float(line.services_subtotal_amount),
        _decimal_to_float(line.departure_amount),
        _decimal_to_float(line.amount_without_vat),
        _decimal_to_float(line.vat_amount),
        _decimal_to_float(line.amount_with_vat),
        json.dumps(line.raw_row, ensure_ascii=False),
        line.row_error,
        json.dumps(line.row_warnings, ensure_ascii=False),
        created_at,
    )


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        f"""
        CREATE TABLE IF NOT EXISTS {UPLOADS_TABLE} (
            upload_id TEXT PRIMARY KEY,
            original_filename TEXT NOT NULL,
            stored_file_path TEXT NOT NULL,
            file_sha256 TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            validation_status TEXT NOT NULL CHECK (validation_status IN ('ok', 'failed')),
            validation_error_summary TEXT NOT NULL DEFAULT '',
            rows_total INTEGER NOT NULL DEFAULT 0,
            rows_matched INTEGER NOT NULL DEFAULT 0,
            amount_without_vat_total REAL NOT NULL DEFAULT 0,
            vat_total REAL NOT NULL DEFAULT 0,
            amount_with_vat_total REAL NOT NULL DEFAULT 0,
            payment_validation_id TEXT,
            pdf_file_path TEXT,
            deleted_at TEXT,
            deleted_by TEXT,
            delete_reason TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE UNIQUE INDEX IF NOT EXISTS idx_fulfillment_upload_payment_validation_id
        ON {UPLOADS_TABLE}(payment_validation_id)
        WHERE payment_validation_id IS NOT NULL AND payment_validation_id <> '';

        CREATE INDEX IF NOT EXISTS idx_fulfillment_uploads_uploaded_at
        ON {UPLOADS_TABLE}(uploaded_at DESC, upload_id DESC);

        CREATE TABLE IF NOT EXISTS {LINES_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id TEXT NOT NULL REFERENCES {UPLOADS_TABLE}(upload_id) ON DELETE CASCADE,
            row_index INTEGER NOT NULL,
            supply_id_input TEXT NOT NULL DEFAULT '',
            matched_wb_supply_id TEXT NOT NULL DEFAULT '',
            matched_wb_cache_key TEXT NOT NULL DEFAULT '',
            match_status TEXT NOT NULL CHECK (match_status IN ('ok', 'unmatched', 'duplicate', 'invalid')),
            service_name TEXT NOT NULL DEFAULT '',
            route TEXT NOT NULL DEFAULT '',
            boxes_qty REAL,
            box_price REAL,
            boxes_amount REAL,
            pallets_qty REAL,
            pallet_price REAL,
            pallets_amount REAL,
            services_subtotal_amount REAL,
            departure_amount REAL,
            amount_without_vat REAL,
            vat_amount REAL,
            amount_with_vat REAL,
            raw_row_json TEXT NOT NULL,
            row_error TEXT NOT NULL DEFAULT '',
            row_warnings_json TEXT NOT NULL DEFAULT '[]',
            created_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_fulfillment_lines_upload
        ON {LINES_TABLE}(upload_id, row_index);

        CREATE INDEX IF NOT EXISTS idx_fulfillment_lines_matched_supply
        ON {LINES_TABLE}(matched_wb_supply_id, matched_wb_cache_key, match_status);
        """
    )
    existing_upload_columns = {
        str(row["name"])
        for row in conn.execute(f"PRAGMA table_info({UPLOADS_TABLE})").fetchall()
    }
    for column_name, column_sql in {
        "deleted_at": "TEXT",
        "deleted_by": "TEXT",
        "delete_reason": "TEXT",
    }.items():
        if column_name not in existing_upload_columns:
            conn.execute(f"ALTER TABLE {UPLOADS_TABLE} ADD COLUMN {column_name} {column_sql}")


def _upload_row_to_dict(row: sqlite3.Row, *, include_links: bool = False) -> dict[str, Any]:
    payload = {
        "upload_id": row["upload_id"],
        "original_filename": row["original_filename"],
        "stored_file_path": row["stored_file_path"],
        "file_sha256": row["file_sha256"],
        "short_file_hash": str(row["file_sha256"] or "")[:12],
        "uploaded_at": row["uploaded_at"],
        "validation_status": row["validation_status"],
        "validation_error_summary": row["validation_error_summary"],
        "rows_total": int(row["rows_total"] or 0),
        "rows_matched": int(row["rows_matched"] or 0),
        "amount_without_vat_total": row["amount_without_vat_total"],
        "vat_total": row["vat_total"],
        "amount_with_vat_total": row["amount_with_vat_total"],
        "payment_validation_id": row["payment_validation_id"] or "",
        "pdf_file_path": row["pdf_file_path"] or "",
        "deleted_at": row["deleted_at"] or "",
        "deleted_by": row["deleted_by"] or "",
        "delete_reason": row["delete_reason"] or "",
        "pdf_available": bool(row["pdf_file_path"] and row["validation_status"] == VALIDATION_OK and not row["deleted_at"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }
    if include_links:
        payload["pdf_path"] = (
            f"{FULFILLMENT_SERVICES_BASE_PATH}/uploads/{row['upload_id']}/payment-validation.pdf"
            if payload["pdf_available"]
            else ""
        )
    return payload


def _line_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "upload_id": row["upload_id"],
        "row_index": int(row["row_index"] or 0),
        "supply_id_input": row["supply_id_input"] or "",
        "matched_wb_supply_id": row["matched_wb_supply_id"] or "",
        "matched_wb_cache_key": row["matched_wb_cache_key"] or "",
        "match_status": row["match_status"] or "",
        "service_name": row["service_name"] or "",
        "route": row["route"] or "",
        "boxes_qty": row["boxes_qty"],
        "box_price": row["box_price"],
        "boxes_amount": row["boxes_amount"],
        "pallets_qty": row["pallets_qty"],
        "pallet_price": row["pallet_price"],
        "pallets_amount": row["pallets_amount"],
        "services_subtotal_amount": row["services_subtotal_amount"],
        "departure_amount": row["departure_amount"],
        "amount_without_vat": row["amount_without_vat"],
        "vat_amount": row["vat_amount"],
        "amount_with_vat": row["amount_with_vat"],
        "raw_row": _loads_json_dict(row["raw_row_json"]),
        "row_error": row["row_error"] or "",
        "row_warnings": _loads_json_list(row["row_warnings_json"]),
        "created_at": row["created_at"] or "",
    }


def _line_identity_values(row: Mapping[str, Any]) -> set[str]:
    values = {
        str(row["supply_id_input"] or "").strip(),
        str(row["matched_wb_supply_id"] or "").strip(),
        str(row["matched_wb_cache_key"] or "").strip(),
    }
    expanded: set[str] = set()
    for value in values:
        if not value:
            continue
        expanded.add(value)
        expanded.add(value.removeprefix("supply:"))
        expanded.add(value.removeprefix("preorder:"))
        if not value.startswith(("supply:", "preorder:")):
            expanded.add(f"supply:{value}")
    return {item for item in expanded if item}


def _append_unique(items: list[str], value: str) -> None:
    if value and value not in items:
        items.append(value)


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().replace("ё", "е").lower())


def _normalize_supply_identity(value: Any) -> str:
    return str(value or "").strip().casefold()


def _value_at(values: list[Any], column: int | None) -> Any:
    if not column or column <= 0:
        return None
    index = column - 1
    if index < 0 or index >= len(values):
        return None
    return values[index]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and abs(value - round(value)) < 0.000001:
        return str(int(round(value)))
    return str(value).strip()


def _row_is_empty(values: list[Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _row_looks_like_footer(values: list[Any], columns: Mapping[str, int]) -> bool:
    supply_id = _cell_text(_value_at(values, columns["supply_id"]))
    if supply_id:
        return False
    service_name = _cell_text(_value_at(values, columns.get("service_name", 0))).casefold()
    if service_name in {"итого", "всего", "total", "footer"}:
        return True
    non_empty_by_key = []
    for key, column in columns.items():
        if not column:
            continue
        value = _value_at(values, column)
        if value is not None and str(value).strip() != "":
            non_empty_by_key.append(key)
    return bool(non_empty_by_key) and set(non_empty_by_key).issubset(
        {"amount_without_vat", "vat_amount", "services_subtotal_amount"}
    )


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = (
        text.replace("\u00a0", " ")
        .replace("\u202f", " ")
        .replace("₽", "")
        .replace("руб.", "")
        .replace("руб", "")
        .strip()
    )
    text = re.sub(r"\s+", "", text)
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_to_float(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _format_money(value: Decimal | None) -> str:
    if value is None:
        return "—"
    amount = _money(value)
    return f"{amount:,.2f}".replace(",", " ").replace(".", ",") + " ₽"


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _loads_json_dict(value: str | None) -> dict[str, Any]:
    try:
        parsed = json.loads(value or "{}")
    except (TypeError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _loads_json_list(value: str | None) -> list[Any]:
    try:
        parsed = json.loads(value or "[]")
    except (TypeError, json.JSONDecodeError):
        return []
    return parsed if isinstance(parsed, list) else []


def _safe_filename(filename: str) -> str:
    normalized = Path(str(filename or "fulfillment-services.xlsx")).name.strip()
    normalized = re.sub(r"[^\w.\- ()а-яА-ЯёЁ]+", "_", normalized, flags=re.UNICODE).strip("._ ")
    return normalized[:180] or "fulfillment-services.xlsx"


def _relative_to_runtime(runtime_dir: Path, path: Path) -> str:
    return str(path.resolve().relative_to(runtime_dir.resolve()))


def _default_timestamp_factory() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _register_pdf_font(pdfmetrics: Any, TTFont: Any) -> str:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]
    for candidate in candidates:
        path = Path(candidate)
        if not path.is_file():
            continue
        try:
            pdfmetrics.registerFont(TTFont("WbCoreUnicode", str(path)))
            return "WbCoreUnicode"
        except Exception:
            continue
    return "Helvetica"
