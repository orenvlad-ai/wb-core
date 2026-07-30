#!/usr/bin/env python3
"""Fail-closed Partner preview/XLSX production-acceptance regressions."""

from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import zipfile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.finance_partner_production_ui_flow import (  # noqa: E402
    _negative_profit_dividends_valid,
    _partner_acceptance_passed,
    _partner_preview_api_evidence,
    _protected_json_post,
    _validate_finance_storage_card,
    _validate_finance_storage_health,
    _verify_partner_xlsx,
    _xlsx_evidence,
)
from packages.application.partner_report import (  # noqa: E402
    OTHER_DIRECT_ALLOCATED_KEY,
    OTHER_EXPENSE_CATEGORIES,
    PartnerReportBlock,
    REPORT_ROWS,
)


VISIBLE_EXPENSE_CATEGORIES = OTHER_EXPENSE_CATEGORIES[:-1]


def main() -> None:
    _assert(
        _validate_finance_storage_health(_monolith_storage_health())
        == "implicit_monolith",
        "implicit monolith storage phase was not accepted",
    )
    _assert(
        _validate_finance_storage_health(
            _rollback_monolith_storage_health()
        )
        == "selected_rollback_monolith",
        "explicit rollback monolith is accepted as one healthy selected file",
    )
    split_health = _split_storage_health()
    _assert(
        _validate_finance_storage_health(split_health) == "selected_split",
        "selected split storage phase was not accepted",
    )
    _validate_finance_storage_card(
        _monolith_storage_card(),
        storage_health=_monolith_storage_health(),
        storage_phase="implicit_monolith",
    )
    _validate_finance_storage_card(
        _split_storage_card(split_health),
        storage_health=split_health,
        storage_phase="selected_split",
    )
    _assert_storage_card_rejected(
        _split_storage_card(split_health).replace(
            "operational-" + str(split_health["generation_epoch"]),
            "operational-" + "f" * 20,
        ),
        storage_health=split_health,
        storage_phase="selected_split",
        label="mixed rendered split generation",
    )
    _assert_storage_card_rejected(
        _split_storage_card(split_health).replace(
            "consumer lag: 0",
            "consumer lag: 1",
        ),
        storage_health=split_health,
        storage_phase="selected_split",
        label="rendered split consumer lag",
    )
    invalid_split_generation = deepcopy(split_health)
    invalid_split_generation["operational"]["generation_epoch"] = "f" * 20
    _assert_storage_rejected(
        invalid_split_generation,
        "mixed split generation",
    )
    invalid_split_lag = deepcopy(split_health)
    invalid_split_lag["consumer_lag_events"] = 1
    _assert_storage_rejected(invalid_split_lag, "split consumer lag")
    invalid_split_rollback = deepcopy(split_health)
    invalid_split_rollback["rollback_generation_id"] = "unknown"
    _assert_storage_rejected(
        invalid_split_rollback,
        "missing retained monolith rollback",
    )
    invalid_split_dead_letter = deepcopy(split_health)
    invalid_split_dead_letter["actionable_dead_letters"] = 1
    _assert_storage_rejected(
        invalid_split_dead_letter,
        "split actionable dead letter",
    )

    class NonJsonResponse:
        status = 502

        @staticmethod
        def json() -> object:
            raise ValueError("not JSON")

    class NonJsonRequest:
        @staticmethod
        def post(*_args: object, **_kwargs: object) -> NonJsonResponse:
            return NonJsonResponse()

    class NonJsonContext:
        request = NonJsonRequest()

    non_json_status, non_json_payload, non_json_object = _protected_json_post(
        NonJsonContext(),
        "https://example.invalid/preview",
        {},
        label="Partner preview API",
    )
    non_json_evidence = _partner_preview_api_evidence(
        non_json_status,
        non_json_payload,
    )
    _assert(
        not non_json_object
        and non_json_evidence["api_http_status"] == 502
        and non_json_evidence["api_code"] == "response_not_json"
        and non_json_evidence["blockers"]
        == [{"code": "response_not_json", "http_status": 502}],
        f"non-JSON preview evidence was not preserved: {non_json_evidence}",
    )
    rejected = _partner_preview_api_evidence(
        422,
        {
            "error": "Derived Finance rows require canonical apply",
            "code": "source_coverage_incomplete",
            "blockers": [{"code": "stale_finance_projection"}],
        },
    )
    _assert(
        rejected
        == {
            "api_http_status": 422,
            "api_status": "",
            "api_code": "source_coverage_incomplete",
            "blockers": [{"code": "stale_finance_projection"}],
            "reason": "Derived Finance rows require canonical apply",
            "source_digest": "",
        },
        f"non-200 preview evidence was not preserved: {rejected}",
    )
    preview = _preview()
    profitable_preview = _preview()
    profitable_preview["weeks"][0]["values"]["net_profit"] = "5.0000"
    profitable_preview["weeks"][0]["values"]["dividends"] = "2.0000"
    _assert(
        _negative_profit_dividends_valid(profitable_preview),
        "all-profitable production data must remain acceptable",
    )
    invalid_loss_preview = _preview()
    invalid_loss_preview["weeks"][0]["values"]["dividends"] = "1.0000"
    _assert(
        not _negative_profit_dividends_valid(invalid_loss_preview),
        "loss week with dividends must fail acceptance",
    )
    ui = _ui_table(preview)
    with TemporaryDirectory(prefix="finance-partner-ui-flow-smoke-") as tmp:
        root = Path(tmp)
        workbook_path = root / "partner.xlsx"
        workbook_path.write_bytes(
            PartnerReportBlock(root)._build_main_workbook(preview)  # noqa: SLF001
        )
        valid = _verify_partner_xlsx(workbook_path, preview=preview, ui_table=ui)
        _assert(valid["passed"], f"valid Partner XLSX rejected: {valid}")

        percentage_preview = _preview()
        for week in percentage_preview["weeks"]:
            week["values"]["net_profit"] = "2000.0000"
            week["values"]["dividends"] = "1000.0000"
            week["values"]["annualized_return_pct"] = "5.2000"
        percentage_preview["totals"]["net_profit"] = "4000.0000"
        percentage_preview["totals"]["dividends"] = "2000.0000"
        percentage_preview["totals"]["annualized_return_pct"] = "5.2000"
        percentage_path = root / "non-zero-annualized-return.xlsx"
        percentage_path.write_bytes(
            PartnerReportBlock(root)._build_main_workbook(percentage_preview)  # noqa: SLF001
        )
        percentage_result = _verify_partner_xlsx(
            percentage_path,
            preview=percentage_preview,
            ui_table=_ui_table(percentage_preview),
        )
        _assert(
            percentage_result["passed"],
            f"Excel fraction for non-zero annualized return was rejected: {percentage_result}",
        )

        wrong_nm = root / "wrong-nm.xlsx"
        _mutate_workbook(
            workbook_path,
            wrong_nm,
            lambda workbook: setattr(
                next(
                    row[1]
                    for row in workbook["Параметры"].iter_rows(min_col=1, max_col=2)
                    if row[0].value == "nmId"
                ),
                "value",
                "999999",
            ),
        )
        _assert_rejected(wrong_nm, preview, ui, "wrong_nm_id")

        wrong_weeks = root / "wrong-weeks.xlsx"
        _mutate_workbook(
            workbook_path,
            wrong_weeks,
            lambda workbook: setattr(
                next(
                    row[1]
                    for row in workbook["Параметры"].iter_rows(min_col=1, max_col=2)
                    if row[0].value == "Недели"
                ),
                "value",
                "2026-07-06",
            ),
        )
        _assert_rejected(wrong_weeks, preview, ui, "wrong_weeks")

        hidden = root / "hidden.xlsx"
        _mutate_workbook(
            workbook_path,
            hidden,
            lambda workbook: setattr(
                workbook["Параметры"], "sheet_state", "hidden"
            ),
        )
        _assert_rejected(hidden, preview, ui, "hidden_sheets")

        mismatch = root / "mismatch.xlsx"
        def break_category_sum(workbook: object) -> None:
            sheet = workbook["Партнёрский отчёт"]  # type: ignore[index]
            for row in sheet.iter_rows(min_row=2):
                if row[1].value == OTHER_EXPENSE_CATEGORIES[0][1]:
                    row[2].value = 999
                    return
            raise AssertionError("category row not found")

        _mutate_workbook(workbook_path, mismatch, break_category_sum)
        _assert_rejected(mismatch, preview, ui, "category_values_mismatch")

        wrong_cogs = root / "wrong-cogs.xlsx"
        def break_cogs(workbook: object) -> None:
            sheet = workbook["Партнёрский отчёт"]  # type: ignore[index]
            for row in sheet.iter_rows(min_row=2):
                if row[1].value == "Себестоимость":
                    row[2].value = 999
                    return
            raise AssertionError("COGS row not found")

        _mutate_workbook(workbook_path, wrong_cogs, break_cogs)
        _assert_rejected(wrong_cogs, preview, ui, "metric_values_mismatch")

        external = root / "external.xlsx"
        with zipfile.ZipFile(workbook_path) as source, zipfile.ZipFile(
            external, "w", compression=zipfile.ZIP_DEFLATED
        ) as target:
            for member in source.infolist():
                target.writestr(member, source.read(member.filename))
            target.writestr(
                "xl/externalLinks/externalLink1.xml",
                "<externalLink xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'/>",
            )
        _assert_rejected(external, preview, ui, "external_links")

        invalid = root / "not-a-workbook.xlsx"
        invalid.write_bytes(b"not empty but not a workbook")
        _assert_rejected(invalid, preview, ui, "workbook_open_or_verify_failed")
        invalid_verification = _verify_partner_xlsx(
            invalid,
            preview=preview,
            ui_table=ui,
        )
        invalid_evidence = _xlsx_evidence(
            invalid,
            filename=invalid.name,
            verification=invalid_verification,
        )
        _assert(
            invalid_evidence["downloaded"]
            and invalid_evidence["size_bytes"] == invalid.stat().st_size
            and invalid_evidence["verification"] == invalid_verification,
            "semantic failure must retain downloaded XLSX path/hash/size/findings",
        )

        ready_evidence = {
            "attempted": True,
            "ready": True,
            "api_status": "ready",
            "blockers": [],
            "table_visible": True,
            "download_enabled": True,
            "xlsx": {
                "downloaded": True,
                "path": str(workbook_path),
                "sha256": "sha256:fixture",
                "size_bytes": workbook_path.stat().st_size,
                "verification": {"passed": True, "ui_xlsx_reconciled": True},
            },
        }
        _assert(_partner_acceptance_passed(ready_evidence), "complete evidence did not pass")
        for mutation in (
            {"ready": False},
            {"blockers": [{"code": "fixture"}]},
            {"xlsx": {"downloaded": False}},
            {
                "xlsx": {
                    **ready_evidence["xlsx"],
                    "verification": {"passed": False, "ui_xlsx_reconciled": False},
                }
            },
        ):
            candidate = {**ready_evidence, **mutation}
            _assert(
                not _partner_acceptance_passed(candidate),
                f"fail-closed evidence was accepted: {mutation}",
            )

    print(
        "finance_partner_production_ui_flow_smoke: ok -> preview ready/blockers, "
        "non-200 API evidence, mandatory XLSX, nmID/weeks, percent units, retained failure evidence, "
        "hidden sheet, external link, semantic sums, exact monolith/split storage phases"
    )


def _monolith_storage_health() -> dict:
    return {
        "state": "monolith",
        "canonical_source": "monolith",
        "implicit_manifest": True,
        "raw": {"generation_id": "monolith"},
        "operational": {"generation_id": "monolith"},
        "rollback_ready": True,
        "cutover_ready": False,
    }


def _rollback_monolith_storage_health() -> dict:
    generation_epoch = "rollback-" + ("a" * 20)
    relative_path = (
        f"generations/{generation_epoch}/monolith.sqlite3"
    )
    return {
        "state": "monolith",
        "canonical_source": "monolith",
        "implicit_manifest": False,
        "generation_epoch": generation_epoch,
        "rollback_generation_id": "a" * 20,
        "raw": {
            "exists": True,
            "generation_epoch": generation_epoch,
            "generation_id": generation_epoch,
            "relative_path": relative_path,
        },
        "operational": {
            "exists": True,
            "generation_epoch": generation_epoch,
            "generation_id": generation_epoch,
            "relative_path": relative_path,
        },
        "raw_schema_ready": True,
        "raw_schema_mode": "legacy_monolith",
        "operational_schema_ready": True,
        "cursor_contract": "not_applicable_monolith",
        "cursor_mismatch": False,
        "rollback_ready": True,
        "cutover_ready": False,
    }


def _split_storage_health() -> dict:
    generation_epoch = "a" * 20
    return {
        "state": "cutover",
        "canonical_source": "split",
        "implicit_manifest": False,
        "generation_epoch": generation_epoch,
        "rollback_generation_id": "monolith",
        "raw": {
            "exists": True,
            "generation_epoch": generation_epoch,
            "generation_id": f"finance-raw-{generation_epoch}",
            "schema_revision": "finance_raw_v1",
            "size_bytes": 10_000,
        },
        "operational": {
            "exists": True,
            "generation_epoch": generation_epoch,
            "generation_id": f"operational-{generation_epoch}",
            "schema_revision": "operational_v1",
            "size_bytes": 5_000,
        },
        "raw_schema_ready": True,
        "operational_schema_ready": True,
        "latest_outbox_sequence": 7,
        "raw_ack_cursor": 7,
        "operational_cursor": 7,
        "consumer_lag_events": 0,
        "cursor_mismatch": False,
        "shadow_mismatch_count": 0,
        "actionable_dead_letters": 0,
        "raw_counts": {"pending_outbox": 0},
        "operational_counts": {"dead_letters": 0},
        "rollback_ready": False,
        "cutover_ready": False,
    }


def _monolith_storage_card() -> str:
    return (
        "Состояние: monolith; canonical: monolith; implicit monolith.\n"
        "rollback: готов; cutover: не разрешён/не готов."
    )


def _split_storage_card(storage_health: dict) -> str:
    raw = storage_health["raw"]
    operational = storage_health["operational"]
    return (
        "Состояние: cutover; canonical: split.\n"
        f"Raw generation: {raw['generation_id']}, "
        f"schema {raw['schema_revision']}, 1,00 ГБ.\n"
        f"Operational generation: {operational['generation_id']}, "
        f"schema {operational['schema_revision']}, 512,0 МБ.\n"
        "Cursor raw/operational: "
        f"{storage_health['raw_ack_cursor']} / "
        f"{storage_health['operational_cursor']}; "
        f"consumer lag: {storage_health['consumer_lag_events']}; "
        "live-tail cursor/lag: "
        f"{storage_health.get('live_tail_cursor', 0)} / "
        f"{storage_health.get('live_tail_lag_events', 0)}; "
        f"mismatches: {storage_health['shadow_mismatch_count']}; "
        f"dead letters: {storage_health['actionable_dead_letters']}.\n"
        "rollback: не доказан; cutover: не разрешён/не готов."
    )


def _assert_storage_rejected(storage_health: dict, label: str) -> None:
    try:
        _validate_finance_storage_health(storage_health)
    except AssertionError:
        return
    raise AssertionError(f"Finance UI accepted invalid storage health: {label}")


def _assert_storage_card_rejected(
    storage_text: str,
    *,
    storage_health: dict,
    storage_phase: str,
    label: str,
) -> None:
    try:
        _validate_finance_storage_card(
            storage_text,
            storage_health=storage_health,
            storage_phase=storage_phase,
        )
    except AssertionError:
        return
    raise AssertionError(f"Finance UI accepted invalid storage card: {label}")


def _preview() -> dict:
    values = {key: "0.0000" for key, _label in REPORT_ROWS}
    values[OTHER_DIRECT_ALLOCATED_KEY] = "10.0000"
    values["net_profit"] = "-5.0000"
    values["dividends"] = "0.0000"
    breakdown = [
        {"key": key, "label": label, "amount_rub": amount}
        for (key, label), amount in zip(
            OTHER_EXPENSE_CATEGORIES,
            ("3.33", "3.33", "3.33", "0.01", "0.00"),
            strict=True,
        )
    ]
    return {
        "status": "ready",
        "contract_version": "partner_report_v4",
        "formula_version": "partner_report_profitability_ui_first_v4",
        "source_digest": "sha256:" + "a" * 64,
        "nm_id": "101101",
        "product_name": "Контрольный SKU",
        "parameters": {
            "partner_share_pct": "40",
            "invested_capital_rub": "1000000",
            "replenishment_reserve_pct": "20",
            "weekly_office_expense_rub": "10000",
            "tax_rate_pct": "6",
        },
        "selected_weeks": ["2026-07-06", "2026-07-13"],
        "weeks": [
            {"label": "06.07–12.07", "values": dict(values), "other_expense_breakdown": list(breakdown)},
            {"label": "13.07–19.07", "values": dict(values), "other_expense_breakdown": list(breakdown)},
        ],
        "totals": {
            **{key: "0.0000" for key, _label in REPORT_ROWS},
            OTHER_DIRECT_ALLOCATED_KEY: "20.0000",
            "net_profit": "-10.0000",
            "dividends": "0.0000",
            "annualized_return_pct": "0.0000",
        },
        "other_expense_breakdown_total": [
            {"key": key, "label": label, "amount_rub": amount}
            for (key, label), amount in zip(
                OTHER_EXPENSE_CATEGORIES,
                ("6.66", "6.66", "6.66", "0.02", "0.00"),
                strict=True,
            )
        ],
        "other_expense_category_definitions": [
            {"key": key, "label": label}
            for key, label in VISIBLE_EXPENSE_CATEGORIES
        ],
        "annualized_return_formula": "average weekly dividends × 52 / invested capital × 100%",
        "generated_at": "2026-07-22T00:00:00Z",
    }


def _ui_table(preview: dict) -> dict:
    def money(value: str) -> str:
        return str(value).replace(".", ",") + " ₽"

    weeks = list(preview["weeks"])
    return {
        "rows": [
            {
                "key": key,
                "values": [
                    (
                        str(week["values"][key]).replace(".", ",") + "%"
                        if key == "annualized_return_pct"
                        else money(week["values"][key])
                    )
                    for week in weeks
                ]
                + [
                    (
                        str(preview["totals"][key]).replace(".", ",") + "%"
                        if key == "annualized_return_pct"
                        else money(preview["totals"][key])
                    )
                ],
            }
            for key, _label in REPORT_ROWS
        ],
        "main_values": [
            money(week["values"][OTHER_DIRECT_ALLOCATED_KEY]) for week in weeks
        ] + [money(preview["totals"][OTHER_DIRECT_ALLOCATED_KEY])],
        "categories": [
            {
                "key": key,
                "label": label,
                "values": [
                    money(next(item["amount_rub"] for item in week["other_expense_breakdown"] if item["key"] == key))
                    for week in weeks
                ] + [money(next(item["amount_rub"] for item in preview["other_expense_breakdown_total"] if item["key"] == key))],
            }
            for key, label in VISIBLE_EXPENSE_CATEGORIES
        ],
    }


def _mutate_workbook(source: Path, target: Path, mutation: object) -> None:
    workbook = load_workbook(source)
    mutation(workbook)  # type: ignore[operator]
    workbook.save(target)
    workbook.close()


def _assert_rejected(
    path: Path,
    preview: dict,
    ui: dict,
    expected_code: str,
) -> None:
    result = _verify_partner_xlsx(path, preview=preview, ui_table=ui)
    codes = {str(item.get("code") or "") for item in result["findings"]}
    _assert(not result["passed"] and expected_code in codes, f"{expected_code}: {result}")


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


if __name__ == "__main__":
    main()
