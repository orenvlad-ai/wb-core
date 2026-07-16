"""Smoke-check supplier invoice barcode detection and barcode-only matching."""

from __future__ import annotations

from collections import Counter
from copy import deepcopy
from io import BytesIO
from pathlib import Path
import sys
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.application.supplier_invoice_parser import (  # noqa: E402
    normalize_barcode_value,
    parse_supplier_invoice_xlsx,
)
from packages.application.supplier_shipments import (  # noqa: E402
    _apply_nomenclature_matches,
    _assert_atomic_supplier_product_matching,
    _bind_source_owned_line_identity,
)


PRIMARY_BARCODE = "0123456789012"
SECONDARY_BARCODE = "1234567890123"
THIRD_BARCODE = "1234567890124"


def main() -> None:
    _assert_header_aliases()
    _assert_structural_detection()
    _assert_conflict_and_rejection_paths()
    _assert_lossless_barcode_values()
    _assert_barcode_only_matching()
    _assert_authoritative_projection_and_model_diagnostics()
    _assert_source_identity_and_atomic_numbers()
    _assert_synthetic_34_of_34_mapping()
    _assert_metadata_totals_and_extras()
    print("supplier_invoice_parser_smoke: OK")


def _assert_header_aliases() -> None:
    aliases = [
        "Braocde\n(条形码）",
        "Barcode",
        "Bar code",
        "条形码",
        "條形碼",
        "ШК",
        "Штрихкод",
        "Штрих-код",
        "  bAr \n CoDe  ",
    ]
    for header in aliases:
        payload = parse_supplier_invoice_xlsx(_build_invoice_fixture(barcode_header=header), filename="sanitized.xlsx")
        diagnostics = payload.get("diagnostics", {}).get("barcode_column", {})
        if diagnostics.get("method") != "header_alias" or diagnostics.get("column_index") != 4:
            raise AssertionError(f"barcode header alias {header!r} was not detected semantically: {diagnostics}")
        product_lines = [line for line in payload["lines"] if line["line_type"] == "product"]
        if [line.get("barcode") for line in product_lines] != [PRIMARY_BARCODE, SECONDARY_BARCODE, THIRD_BARCODE]:
            raise AssertionError(f"barcode alias {header!r} changed extracted identities")


def _assert_structural_detection() -> None:
    moved_layout = ["no", "models", "name_spec", "comment", "barcode", "qty", "unit_price", "amount"]
    moved_payload = parse_supplier_invoice_xlsx(
        _build_invoice_fixture(barcode_header="Barcode", layout=moved_layout),
        filename="moved.xlsx",
    )
    moved_diagnostics = moved_payload["diagnostics"]["barcode_column"]
    if moved_diagnostics.get("method") != "header_alias" or moved_diagnostics.get("column_index") != 5:
        raise AssertionError(f"semantic barcode column must survive movement away from D: {moved_diagnostics}")

    relative_payload = parse_supplier_invoice_xlsx(
        _build_invoice_fixture(barcode_header="Factory identifier", layout=moved_layout),
        filename="relative.xlsx",
    )
    relative_diagnostics = relative_payload["diagnostics"]["barcode_column"]
    if relative_diagnostics.get("method") != "relative_structure" or relative_diagnostics.get("column_index") != 5:
        raise AssertionError(f"unknown moved barcode column must use relative structure: {relative_diagnostics}")

    positional_payload = parse_supplier_invoice_xlsx(
        _build_invoice_fixture(barcode_header="Factory identifier"),
        filename="positional.xlsx",
    )
    positional_diagnostics = positional_payload["diagnostics"]["barcode_column"]
    if positional_diagnostics.get("method") != "positional_d" or positional_diagnostics.get("column_index") != 4:
        raise AssertionError(f"confirmed current-template D fallback changed: {positional_diagnostics}")


def _assert_conflict_and_rejection_paths() -> None:
    invalid_structure = _build_invoice_fixture(
        barcode_header="Factory identifier",
        layout=["no", "models", "comment", "barcode", "qty", "unit_price", "amount", "name_spec"],
    )
    _assert_parse_error(invalid_structure, "barcode column not found")

    no_barcode_column = _build_invoice_fixture(
        layout=["no", "models", "name_spec", "qty", "unit_price", "amount", "comment"],
    )
    _assert_parse_error(no_barcode_column, "barcode column not found")

    ambiguous = _build_invoice_fixture(
        barcode_header="Factory identifier A",
        layout=["no", "models", "name_spec", "barcode", "barcode_copy", "qty", "unit_price", "amount"],
    )
    _assert_parse_error(ambiguous, "ambiguous barcode column")

    missing_one = parse_supplier_invoice_xlsx(
        _build_invoice_fixture(barcodes=[PRIMARY_BARCODE, "", THIRD_BARCODE]),
        filename="missing-one.xlsx",
    )
    missing_lines = [line for line in missing_one["lines"] if line["line_type"] == "product"]
    if missing_lines[1].get("barcode") or missing_lines[1].get("match_status") != "unmatched":
        raise AssertionError("a missing product barcode must stay unmatched without a model fallback")


def _assert_lossless_barcode_values() -> None:
    payload = parse_supplier_invoice_xlsx(
        _build_invoice_fixture(barcodes=[PRIMARY_BARCODE, 1234567890123, "'1234567890124"]),
        filename="lossless.xlsx",
    )
    product_lines = [line for line in payload["lines"] if line["line_type"] == "product"]
    actual = [line.get("barcode") for line in product_lines]
    if actual != [PRIMARY_BARCODE, SECONDARY_BARCODE, THIRD_BARCODE]:
        raise AssertionError(f"text, numeric, apostrophe or leading-zero barcode normalization is lossy: {actual}")
    if normalize_barcode_value(1234567890123.0) != SECONDARY_BARCODE:
        raise AssertionError("an exact Excel numeric barcode must normalize without scientific notation")
    if normalize_barcode_value(" 0123\u00a0456789012 ") != PRIMARY_BARCODE:
        raise AssertionError("safe whitespace cleanup must preserve a leading zero")
    try:
        normalize_barcode_value("1.234567890123E+12")
    except ValueError as exc:
        if "scientific" not in str(exc).lower():
            raise
    else:
        raise AssertionError("scientific-notation text must be rejected instead of guessed")
    for unsafe_value in (2**53 + 1, -12345678, "123.0", "ABC123"):
        try:
            normalize_barcode_value(unsafe_value)
        except ValueError:
            continue
        raise AssertionError(f"unsafe or ambiguous barcode must be rejected: {unsafe_value!r}")


def _assert_barcode_only_matching() -> None:
    parsed = parse_supplier_invoice_xlsx(_build_invoice_fixture(), filename="matching.xlsx")
    product_lines = [dict(line) for line in parsed["lines"] if line["line_type"] == "product"]
    for line in product_lines:
        line.update(
            {
                "product_type": "intentionally_wrong",
                "model_raw": "intentionally wrong model",
                "model_normalized": "intentionally_wrong_model",
                "match_key": "intentionally_wrong|model",
            }
        )
    matched = _apply_nomenclature_matches(
        product_lines,
        [
            _nomenclature_item("nom-primary", 101, PRIMARY_BARCODE, []),
            _nomenclature_item("nom-inactive-duplicate", 102, PRIMARY_BARCODE, [], active=False),
            _nomenclature_item("nom-secondary", 202, "9999999999999", [SECONDARY_BARCODE]),
            _nomenclature_item("nom-third", 303, THIRD_BARCODE, [], hidden=True),
        ],
    )
    if [line.get("internal_nm_id") for line in matched] != [101, 202, 303]:
        raise AssertionError(f"primary/all/hidden barcode owners were not resolved: {matched}")
    if any(line.get("match_status") != "matched_by_barcode" for line in matched):
        raise AssertionError(f"barcode matches must expose matched_by_barcode evidence: {matched}")

    unknown = _apply_nomenclature_matches(
        [{**product_lines[0], "barcode": "8888888888888"}],
        [_nomenclature_item("nom-model-fallback", 404, PRIMARY_BARCODE, [])],
    )[0]
    if unknown.get("match_status") != "unmatched" or unknown.get("internal_nm_id") is not None:
        raise AssertionError(f"unknown barcode received a forbidden model/type fallback: {unknown}")
    if unknown.get("match_evidence", {}).get("reason") != "barcode_unknown":
        raise AssertionError(f"unknown barcode diagnostic is unclear: {unknown}")

    missing = _apply_nomenclature_matches(
        [{**product_lines[0], "barcode": ""}],
        [_nomenclature_item("nom-missing-fallback", 505, PRIMARY_BARCODE, [])],
    )[0]
    if missing.get("match_status") != "unmatched" or missing.get("internal_nm_id") is not None:
        raise AssertionError(f"missing barcode received a forbidden fallback: {missing}")
    if missing.get("match_evidence", {}).get("reason") != "barcode_missing":
        raise AssertionError(f"missing barcode diagnostic is unclear: {missing}")

    ambiguous = _apply_nomenclature_matches(
        [product_lines[0]],
        [
            _nomenclature_item("nom-duplicate-a", 606, PRIMARY_BARCODE, []),
            _nomenclature_item("nom-duplicate-b", 607, PRIMARY_BARCODE, []),
        ],
    )[0]
    if ambiguous.get("match_status") != "ambiguous" or ambiguous.get("internal_nm_id") is not None:
        raise AssertionError(f"duplicate active barcode owners must be ambiguous: {ambiguous}")

    missing_nmid = _apply_nomenclature_matches(
        [product_lines[0]],
        [_nomenclature_item("nom-missing-nmid", None, PRIMARY_BARCODE, [])],
    )[0]
    if (
        missing_nmid.get("match_status") != "unmatched"
        or missing_nmid.get("internal_nm_id") is not None
        or missing_nmid.get("match_evidence", {}).get("reason") != "barcode_owner_missing_nmid"
    ):
        raise AssertionError(f"barcode owner without nmID must stay explicit and unmatched: {missing_nmid}")
    for rejected in (unknown, missing, ambiguous, missing_nmid):
        try:
            _assert_atomic_supplier_product_matching([rejected])
        except ValueError:
            continue
        raise AssertionError(f"unresolved barcode identity must reject the whole save: {rejected}")


def _assert_source_identity_and_atomic_numbers() -> None:
    parsed = parse_supplier_invoice_xlsx(_build_invoice_fixture(), filename="source-owned.xlsx")
    swapped = deepcopy(parsed)
    first_raw = deepcopy(swapped["lines"][0]["raw"])
    swapped["lines"][0]["raw"] = deepcopy(swapped["lines"][1]["raw"])
    swapped["lines"][1]["raw"] = first_raw
    try:
        _bind_source_owned_line_identity(
            swapped,
            trusted_payload=parsed,
            context="source identity smoke",
        )
    except ValueError as exc:
        if "identity/order" not in str(exc):
            raise
    else:
        raise AssertionError("client must not reassign source invoice row identities")

    product_lines = [dict(line) for line in parsed["lines"] if line["line_type"] == "product"]
    matched = _apply_nomenclature_matches(
        product_lines,
        [
            _nomenclature_item("atomic-a", 1001, PRIMARY_BARCODE, []),
            _nomenclature_item("atomic-b", 1002, SECONDARY_BARCODE, []),
            _nomenclature_item("atomic-c", 1003, THIRD_BARCODE, []),
        ],
    )
    for field, invalid_value in (("qty", float("nan")), ("unit_price", 0), ("amount", -1)):
        invalid_lines = deepcopy(matched)
        invalid_lines[0][field] = invalid_value
        try:
            _assert_atomic_supplier_product_matching(invalid_lines)
        except ValueError:
            continue
        raise AssertionError(f"atomic save must reject invalid {field}")


def _assert_authoritative_projection_and_model_diagnostics() -> None:
    parsed = parse_supplier_invoice_xlsx(_build_invoice_fixture(), filename="authoritative-projection.xlsx")
    products = [dict(line) for line in parsed["lines"] if line["line_type"] == "product"]
    owners = [
        _nomenclature_item("owner-clean", 701, PRIMARY_BARCODE, []),
        _nomenclature_item("owner-matte", 702, SECONDARY_BARCODE, []),
        _nomenclature_item("owner-anti-spy", 703, THIRD_BARCODE, []),
    ]
    owner_fields = [
        ("no_frame_clean", "no_frame_clean|iphone_14_pro", ["iphone_14_pro"]),
        ("no_frame_matte", "no_frame_matte|iphone_16", ["iphone_16"]),
        ("no_frame_anti_spy", "no_frame_anti_spy|iphone_18", ["iphone_18"]),
    ]
    for owner, (group_key, match_key, model_keys) in zip(owners, owner_fields, strict=True):
        owner["product_type"] = group_key
        owner["match_key"] = match_key
        owner["compatible_model_keys"] = model_keys
    sku_groups = [
        {"group_key": "no_frame_clean", "label": "No Frame Clean", "is_active": True},
        {"group_key": "no_frame_matte", "label": "No Frame Matte", "is_active": True},
        {"group_key": "no_frame_anti_spy", "label": "No Frame Anti-spy", "is_active": True},
    ]
    matched = _apply_nomenclature_matches(products, owners, sku_groups)
    if [line.get("source_product_type") for line in matched] != ["clear", "clear", "anti_spy"]:
        raise AssertionError(f"parser classification must remain source-only evidence: {matched}")
    if [line.get("product_type") for line in matched] != [
        "no_frame_clean",
        "no_frame_matte",
        "no_frame_anti_spy",
    ]:
        raise AssertionError(f"barcode owners must replace source groups: {matched}")
    if [line.get("match_key") for line in matched] != [value[1] for value in owner_fields]:
        raise AssertionError(f"authoritative match keys must come from the same barcode owners: {matched}")
    if [line.get("group_label") for line in matched] != [
        "No Frame Clean",
        "No Frame Matte",
        "No Frame Anti-spy",
    ]:
        raise AssertionError(f"group labels must resolve through server-owned SKU groups: {matched}")
    if [line.get("model_diagnostic", {}).get("status") for line in matched] != [
        "consistent",
        "consistent",
        "mismatch",
    ]:
        raise AssertionError(f"model checks must be non-blocking canonical diagnostics: {matched}")
    if matched[2].get("internal_nm_id") != 703 or matched[2].get("match_status") != "matched_by_barcode":
        raise AssertionError("model mismatch must not change the barcode owner or match status")

    uncheckable_line = {**products[0], "model_raw": "unrecognized factory model"}
    uncheckable = _apply_nomenclature_matches([uncheckable_line], [owners[0]], sku_groups)[0]
    if uncheckable.get("model_diagnostic", {}).get("status") != "not_checkable":
        raise AssertionError(f"unrecognized models must not be guessed: {uncheckable}")

    renamed_groups = [{**group, "label": "Renamed runtime label"} for group in sku_groups if group["group_key"] == "no_frame_clean"]
    renamed = _apply_nomenclature_matches([products[0]], [owners[0]], renamed_groups)[0]
    if renamed.get("group_label") != "Renamed runtime label":
        raise AssertionError("invoice group labels must follow the runtime SKU-group resolver without code changes")


def _assert_synthetic_34_of_34_mapping() -> None:
    categories = ["No Frame Clean"] * 10 + ["No Frame Anti-spy"] * 12 + ["No Frame Matte"] * 12
    barcodes = [f"3{index:012d}" for index in range(1, 35)]
    parsed = parse_supplier_invoice_xlsx(
        _build_34_row_invoice_fixture(barcodes),
        filename="synthetic-26GN583-structure.xlsx",
    )
    diagnostics = parsed.get("diagnostics") or {}
    products = [dict(line) for line in parsed.get("lines") or [] if line.get("line_type") == "product"]
    if (
        diagnostics.get("worksheet") != "PI"
        or diagnostics.get("header_row") != 6
        or diagnostics.get("barcode_column", {}).get("method") != "header_alias"
        or diagnostics.get("barcode_column", {}).get("column_letter") != "D"
        or len(products) != 34
        or len({line.get("barcode") for line in products}) != 34
    ):
        raise AssertionError(f"synthetic 26GN583 structure changed: {diagnostics}, products={len(products)}")
    nomenclature = []
    category_keys = {
        "No Frame Clean": "no_frame_clean",
        "No Frame Anti-spy": "no_frame_anti_spy",
        "No Frame Matte": "no_frame_matte",
    }
    for index, (barcode, category) in enumerate(zip(barcodes, categories, strict=True), start=1):
        item = _nomenclature_item(
            f"synthetic-{index}",
            900000000 + index,
            barcode if index % 2 else "9999999999999",
            [] if index % 2 else [barcode],
        )
        item["nomenclature_name"] = f"{category} synthetic {index:02d}"
        item["product_type"] = category_keys[category]
        item["match_key"] = f"{category_keys[category]}|synthetic_{index:02d}"
        nomenclature.append(item)
    sku_groups = [
        {"group_key": key, "label": label, "is_active": True}
        for label, key in category_keys.items()
    ]
    matched = _apply_nomenclature_matches(products, nomenclature, sku_groups)
    matched_categories = Counter(
        str(line.get("internal_name") or "").rsplit(" synthetic ", 1)[0]
        for line in matched
    )
    if (
        len(matched) != 34
        or any(line.get("match_status") != "matched_by_barcode" for line in matched)
        or len({line.get("internal_nm_id") for line in matched}) != 34
        or Counter(line.get("product_type") for line in matched)
        != Counter({"no_frame_clean": 10, "no_frame_anti_spy": 12, "no_frame_matte": 12})
        or matched_categories
        != Counter({"No Frame Clean": 10, "No Frame Anti-spy": 12, "No Frame Matte": 12})
    ):
        raise AssertionError(f"synthetic barcode mapping must close 34/34 exactly: {matched_categories}")


def _assert_metadata_totals_and_extras() -> None:
    payload = parse_supplier_invoice_xlsx(_build_invoice_fixture(), filename="PI-test 26GN390 (14.5.2026).xlsx")
    if payload["diagnostics"].get("worksheet") != "PI" or payload["diagnostics"].get("header_row") != 6:
        raise AssertionError(f"sanitized current-template fixture shape changed: {payload['diagnostics']}")
    metadata = payload["metadata"]
    if metadata.get("contract_no") != "CNT-2026-0513" or metadata.get("contract_date") != "2026-05-13":
        raise AssertionError(f"cell metadata extraction regressed: {metadata}")
    product_lines = [line for line in payload["lines"] if line["line_type"] == "product"]
    extra_lines = [line for line in payload["lines"] if line["line_type"] == "extra"]
    if len(product_lines) != 3 or len(extra_lines) != 1 or extra_lines[0].get("barcode"):
        raise AssertionError("product/extra separation or optional extra barcode regressed")
    summary = payload["summary"]
    if (
        summary.get("product_qty_total") != 18.0
        or summary.get("product_amount_total") != 26.0
        or summary.get("extras_amount_total") != 5.0
        or summary.get("invoice_amount_total") != 31.0
        or summary.get("checksum_error")
    ):
        raise AssertionError(f"invoice totals/checksum regressed: {summary}")

    drawing_payload = parse_supplier_invoice_xlsx(
        _with_drawing_text_fixture(_build_invoice_fixture(contract_cells=False)),
        filename="PI-drawing 26GN391 (15.5.2026).xlsx",
    )
    drawing_metadata = drawing_payload["metadata"]
    if drawing_metadata.get("contract_no") != "26DRAW001" or drawing_metadata.get("contract_date") != "2026-05-13":
        raise AssertionError(f"drawing XML metadata extraction regressed: {drawing_metadata}")


def _assert_parse_error(workbook_bytes: bytes, expected: str) -> None:
    try:
        parse_supplier_invoice_xlsx(workbook_bytes, filename="rejected.xlsx")
    except ValueError as exc:
        if expected not in str(exc).lower():
            raise AssertionError(f"expected {expected!r}, got {exc!r}") from exc
    else:
        raise AssertionError(f"invalid invoice must fail with {expected!r}")


def _nomenclature_item(
    item_id: str,
    nm_id: int | None,
    barcode: str,
    barcodes: list[str],
    *,
    hidden: bool = False,
    active: bool = True,
) -> dict[str, object]:
    return {
        "item_id": item_id,
        "is_active": active,
        "is_hidden": hidden,
        "our_sku": f"SKU-{nm_id}",
        "nm_id": nm_id,
        "barcode": barcode,
        "barcodes": barcodes,
        "nomenclature_name": f"Barcode item {nm_id}",
        "product_type": "wrong_type",
        "match_key": "wrong|model",
        "aliases": ["also wrong"],
        "compatible_model_keys": ["iphone_99"],
    }


def _build_34_row_invoice_fixture(barcodes: list[str]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    sheet.append(["Invoice No:", "SYNTHETIC-34"])
    sheet.append(["Invoice Date:", "2026-07-13"])
    sheet.append(["Contract No.", "SYNTHETIC-CONTRACT"])
    sheet.append(["Date of Contract", "2026-06-08"])
    sheet.append(["Supplier:", "Sanitized Supplier"])
    sheet.append(
        [
            "NO.",
            "MODELS / （型号）",
            "NAME & SPECIFICATION / （品名规格）",
            "Braocde\n(条形码）",
            "QTY (PCS) / （数量）",
            "U.PRICE / （单价） (RMB/PCS)",
            "AMOUNT / （总价） (RMB)",
            "备注",
        ]
    )
    for index, barcode in enumerate(barcodes, start=1):
        marker = "高清膜 smk" if index <= 10 else "防窥膜 (Anti-Spy)" if index <= 22 else "磨砂膜 (Matte)"
        sheet.append([index, f"Synthetic Model {index:02d}", marker, barcode, 1, 1, 1, ""])
    sheet.append(["（总值）Total:", "", "", "", "", "", 34, ""])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_invoice_fixture(
    *,
    barcode_header: str = "Braocde\n(条形码）",
    layout: list[str] | None = None,
    barcodes: list[object] | None = None,
    contract_cells: bool = True,
) -> bytes:
    layout = layout or ["no", "models", "name_spec", "barcode", "qty", "unit_price", "amount", "comment"]
    barcodes = barcodes or [PRIMARY_BARCODE, SECONDARY_BARCODE, THIRD_BARCODE]
    headers = {
        "no": "NO.",
        "models": "MODELS / （型号）",
        "name_spec": "NAME & SPECIFICATION / （品名规格）",
        "barcode": barcode_header,
        "barcode_copy": "Factory identifier B",
        "qty": "QTY (PCS) / （数量）",
        "unit_price": "U.PRICE / （单价） (RMB/PCS)",
        "amount": "AMOUNT / （总价） (RMB)",
        "comment": "备注",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    sheet["A1"] = "Invoice No:"
    sheet["B1"] = "26GN390"
    sheet["A2"] = "Invoice Date:"
    sheet["B2"] = "14.5.2026"
    if contract_cells:
        sheet["A3"] = "Contract No."
        sheet["B3"] = "CNT-2026-0513"
        sheet["A4"] = "Date of Contract"
        sheet["B4"] = "2026.5.13"
    sheet["A5"] = "Supplier:"
    sheet["B5"] = "Sanitized Supplier"
    sheet.append([headers[role] for role in layout])

    product_rows = [
        {
            "no": 1,
            "models": "iPhone 14 Pro",
            "name_spec": "高清膜 smk",
            "barcode": barcodes[0],
            "barcode_copy": barcodes[0],
            "qty": 10,
            "unit_price": 1,
            "amount": 10,
            "comment": "packaging instruction",
        },
        {
            "no": 2,
            "models": "iPhone 15 / 16",
            "name_spec": "高清膜 smk",
            "barcode": barcodes[1],
            "barcode_copy": barcodes[1],
            "qty": 5,
            "unit_price": 2,
            "amount": 10,
            "comment": "",
        },
        {
            "no": 3,
            "models": "iPhone 16 Pro/17",
            "name_spec": "防窥膜 (Anti-Spy)",
            "barcode": barcodes[2],
            "barcode_copy": barcodes[2],
            "qty": 3,
            "unit_price": 2,
            "amount": 6,
            "comment": "",
        },
        {
            "no": 4,
            "models": "OPP bag packets",
            "name_spec": "OPP bag packets",
            "barcode": "",
            "barcode_copy": "",
            "qty": 100,
            "unit_price": 0.05,
            "amount": 5,
            "comment": "OPP packets",
        },
    ]
    for values in product_rows:
        sheet.append([values.get(role, "") for role in layout])
    sheet.append(["（总值）Total:" if role == "no" else 31 if role == "amount" else "" for role in layout])

    if "name_spec" in layout:
        name_column = layout.index("name_spec") + 1
        sheet.merge_cells(start_row=7, start_column=name_column, end_row=8, end_column=name_column)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _with_drawing_text_fixture(workbook_bytes: bytes) -> bytes:
    drawing_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
  xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:sp>
      <xdr:txBody>
        <a:bodyPr/><a:lstStyle/>
        <a:p><a:r><a:t>合同号 (Invoice No.)：26DRAW001</a:t></a:r></a:p>
        <a:p><a:r><a:t>下单日期 (Date)：2026.5.13</a:t></a:r></a:p>
      </xdr:txBody>
    </xdr:sp>
  </xdr:twoCellAnchor>
</xdr:wsDr>"""
    source = BytesIO(workbook_bytes)
    target = BytesIO()
    with ZipFile(source, "r") as zin, ZipFile(target, "w", ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        zout.writestr("xl/drawings/drawing1.xml", drawing_xml)
    return target.getvalue()


if __name__ == "__main__":
    main()
