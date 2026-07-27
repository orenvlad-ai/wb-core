"""Regression smoke for atomic operator + WB bulk recommendation export."""

from __future__ import annotations

from hashlib import sha256
from io import BytesIO
from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import zipfile

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime
from packages.application.simple_xlsx import read_first_sheet_rows
from packages.application.wb_regional_supply import WbRegionalSupplyBlock
from packages.application import wb_regional_supply_export as export_module
from packages.application.wb_regional_supply_export import (
    WB_UPLOAD_HEADERS,
    WB_UPLOAD_SHEET_NAME,
    WB_UPLOAD_TEMPLATE_PATH,
    WbUploadRow,
    build_wb_upload_workbook_bytes,
    recommendation_identity,
    recommendation_prefix,
    validate_raw_district_quantities,
)
from packages.contracts.factory_order_supply import STOCK_FF_SOURCE_MANUAL_EXCEL
from packages.contracts.wb_supply_planning_zones import (
    PLANNING_ZONE_CENTRAL_NORTH,
    PLANNING_ZONE_CENTRAL_SOUTH,
)


CALCULATED_AT = "2026-07-20T09:00:00Z"
REPORT_DATE = "2026-07-20"
CALCULATION_ID = "export-run-001"


def main() -> None:
    _check_quantity_validation()
    _check_safe_unique_names()
    _check_structured_table_expansion()
    with TemporaryDirectory(prefix="wb-regional-export-") as tmp:
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=Path(tmp) / "runtime")
        _seed_nomenclature(runtime)
        payload = _result_payload()
        runtime.save_wb_regional_supply_result_state(calculated_at=CALCULATED_AT, payload=payload)
        block = WbRegionalSupplyBlock(runtime=runtime)

        archive_bytes, archive_name = block.download_all_recommendations_archive()
        if archive_name != "Рекомендации_поставок_2026-07-20_14-00_export-run-001.zip":
            raise AssertionError(f"archive filename must carry stable run identity and business time: {archive_name}")
        _check_archive(archive_bytes, payload)
        _check_single_recommendation_archive(runtime, block)
        _check_same_destination_archive(runtime, block)
        runtime.save_wb_regional_supply_result_state(calculated_at=CALCULATED_AT, payload=payload)

        direct_bytes, _ = block.download_district_recommendation(PLANNING_ZONE_CENTRAL_NORTH)
        direct_rows = read_first_sheet_rows(direct_bytes)
        if direct_rows[2] != ["nmId", "SKU", "Количество к поставке"]:
            raise AssertionError(f"direct operator export must remove deficit column: {direct_rows[2]}")
        if any(cell == "Дефицит" for row in direct_rows for cell in row):
            raise AssertionError("direct operator export must not contain deficit header or values")

        broken = _result_payload(calculation_id="export-run-broken")
        broken["districts"][1]["rows"] = [
            _row(999999, "SKU без баркода", allocated_qty=4),
        ]
        broken["districts"][1]["total_qty"] = 4
        runtime.save_wb_regional_supply_result_state(calculated_at=CALCULATED_AT, payload=broken)
        try:
            block.download_all_recommendations_archive()
        except ValueError as exc:
            message = str(exc)
            if (
                "атомарно отменена" not in message
                or "ЦФО Юг" not in message
                or "R260720-export-run-broken-002" not in message
                or "SKU nmId=999999" not in message
                or "отсутствует баркод" not in message
            ):
                raise AssertionError(f"atomic error must identify recommendation, destination and SKU: {message}") from exc
        else:
            raise AssertionError("one invalid recommendation must prevent the complete archive response")

        runtime.save_wb_regional_supply_result_state(
            calculated_at=CALCULATED_AT,
            payload=_result_payload(),
        )
        ambiguous_item = _nomenclature(103, "4600000000003", "Anti-spy")
        ambiguous_item["barcodes"] = ["4600000000003", "4600000000999"]
        ambiguous_item["barcode_status"] = "multiple"
        runtime.save_nomenclature_item(ambiguous_item)
        try:
            block.download_all_recommendations_archive()
        except ValueError as exc:
            message = str(exc)
            if "несколько неоднозначных баркодов" not in message or "SKU nmId=103" not in message:
                raise AssertionError(f"ambiguous canonical barcode evidence must fail closed: {message}") from exc
        else:
            raise AssertionError("ambiguous barcode must cancel the complete archive")

    print("wb_regional_supply_export: ok")


def _check_archive(archive_bytes: bytes, payload: dict[str, object]) -> None:
    if not archive_bytes:
        raise AssertionError("archive must not be empty")
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        names = archive.namelist()
        if len(names) != 4 or len(names) != len(set(names)):
            raise AssertionError(f"two recommendations must produce four unique files: {names}")
        folders: dict[str, list[str]] = {}
        for name in names:
            folder, filename = name.split("/", 1)
            folders.setdefault(folder, []).append(filename)
            if not filename.startswith(folder + "__"):
                raise AssertionError(f"folder and both files must share one prefix: {name}")
        if len(folders) != 2 or any(len(files) != 2 for files in folders.values()):
            raise AssertionError(f"each recommendation folder must contain exactly two files: {folders}")

        ordered_folders = list(folders)
        if not ordered_folders[0].startswith("01_R260720-export-run-001-001_ЦФО_Север"):
            raise AssertionError(f"first folder must match UI recommendation order: {ordered_folders}")
        if not ordered_folders[1].startswith("02_R260720-export-run-001-002_ЦФО_Юг"):
            raise AssertionError(f"second folder must match UI recommendation order: {ordered_folders}")

        first_operator_name = next(name for name in names if name.endswith("__01_РЕКОМЕНДАЦИЯ.xlsx"))
        first_wb_name = next(name for name in names if name.endswith("__02_ЗАГРУЗКА_WB.xlsx"))
        operator_rows = read_first_sheet_rows(archive.read(first_operator_name))
        if operator_rows[2] != ["nmId", "SKU", "Количество к поставке"]:
            raise AssertionError(f"operator headers must be rebuilt after deficit removal: {operator_rows[2]}")
        if any(cell == "Дефицит" for row in operator_rows for cell in row):
            raise AssertionError("operator workbook must contain no deficit header/value")
        if sum(int(row[2]) for row in operator_rows[3:] if len(row) > 2) != 5:
            raise AssertionError("operator workbook quantity must equal recommendation total")

        template = load_workbook(WB_UPLOAD_TEMPLATE_PATH)
        wb_file = load_workbook(BytesIO(archive.read(first_wb_name)))
        try:
            if wb_file.sheetnames != template.sheetnames or wb_file.sheetnames != [WB_UPLOAD_SHEET_NAME]:
                raise AssertionError("WB workbook must preserve the canonical sheet topology")
            template_sheet = template[WB_UPLOAD_SHEET_NAME]
            sheet = wb_file[WB_UPLOAD_SHEET_NAME]
            if (sheet["A1"].value, sheet["B1"].value) != WB_UPLOAD_HEADERS:
                raise AssertionError("WB workbook must preserve canonical headers")
            if list(sheet.tables) != list(template_sheet.tables):
                raise AssertionError("WB workbook must preserve canonical Excel-table identities")
            if sheet["A2"].value != "0012345678901" or sheet["A2"].data_type != "s":
                raise AssertionError("leading-zero barcode must remain exact text")
            if sheet["A2"].number_format != "@" or sheet["B2"].value != 5:
                raise AssertionError("duplicate barcode quantities must be merged into one positive integer row")
            if sheet.max_row != 2 or sheet.max_column != 2:
                raise AssertionError("WB workbook must not add service rows, sheets or columns")
        finally:
            wb_file.close()
            template.close()

        second_wb_name = [name for name in names if name.endswith("__02_ЗАГРУЗКА_WB.xlsx")][1]
        second_wb = load_workbook(BytesIO(archive.read(second_wb_name)))
        try:
            second_sheet = second_wb[WB_UPLOAD_SHEET_NAME]
            if second_sheet["A2"].value != "4600000000003" or second_sheet["B2"].value != 4:
                raise AssertionError("each WB workbook must contain only its own recommendation rows")
        finally:
            second_wb.close()


def _check_single_recommendation_archive(
    runtime: RegistryUploadDbBackedRuntime,
    block: WbRegionalSupplyBlock,
) -> None:
    payload = _result_payload(calculation_id="export-run-single")
    payload["districts"] = [payload["districts"][0]]
    payload["settings"]["included_district_keys"] = [PLANNING_ZONE_CENTRAL_NORTH]
    payload["summary"] = {"total_qty": 5, "estimated_weight": 0.43, "estimated_volume": 0.0}
    runtime.save_wb_regional_supply_result_state(calculated_at=CALCULATED_AT, payload=payload)
    archive_bytes, _ = block.download_all_recommendations_archive()
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        names = archive.namelist()
        if len(names) != 2 or len({name.split("/", 1)[0] for name in names}) != 1:
            raise AssertionError(f"one multi-SKU recommendation must produce one paired folder: {names}")


def _check_same_destination_archive(
    runtime: RegistryUploadDbBackedRuntime,
    block: WbRegionalSupplyBlock,
) -> None:
    payload = _result_payload(calculation_id="export-run-same-destination")
    payload["districts"][1]["planning_zone_label"] = "ЦФО Север"
    runtime.save_wb_regional_supply_result_state(calculated_at=CALCULATED_AT, payload=payload)
    archive_bytes, _ = block.download_all_recommendations_archive()
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        folders = [name.split("/", 1)[0] for name in archive.namelist()[::2]]
        if len(folders) != 2 or len(folders) != len(set(folders)):
            raise AssertionError(f"same-destination recommendations must retain unique folders: {folders}")
        if not all(folder.endswith("_ЦФО_Север") for folder in folders):
            raise AssertionError(f"destination labels must remain visible in both prefixes: {folders}")


def _check_quantity_validation() -> None:
    zero_payload = {"total_qty": 0, "rows": [_row(1, "zero", allocated_qty=0)]}
    if validate_raw_district_quantities(zero_payload):
        raise AssertionError("zero quantity must be excluded without an export error")
    negative_payload = {"total_qty": -1, "rows": [_row(1, "negative", allocated_qty=-1)]}
    negative_issues = validate_raw_district_quantities(negative_payload)
    if not any("отрицательным" in issue for issue in negative_issues):
        raise AssertionError(f"negative quantity must be rejected: {negative_issues}")
    fractional_payload = {"total_qty": 1, "rows": [_row(1, "fractional", allocated_qty=1.5)]}
    fractional_issues = validate_raw_district_quantities(fractional_payload)
    if not any("целым" in issue for issue in fractional_issues):
        raise AssertionError(f"fractional quantity must be rejected: {fractional_issues}")


def _check_safe_unique_names() -> None:
    first_id = recommendation_identity(report_date=REPORT_DATE, calculation_id=CALCULATION_ID, ordinal=1)
    second_id = recommendation_identity(report_date=REPORT_DATE, calculation_id=CALCULATION_ID, ordinal=2)
    first = recommendation_prefix(
        ordinal=1,
        recommendation_id=first_id,
        destination_name='Казань / СЦ:*? "тест".',
    )
    second = recommendation_prefix(
        ordinal=2,
        recommendation_id=second_id,
        destination_name='Казань / СЦ:*? "тест".',
    )
    if first == second or any(character in first + second for character in '<>:"/\\|?*'):
        raise AssertionError(f"same-destination names must stay unique and cross-platform safe: {first!r}, {second!r}")
    longest_member = f"{first}/{first}__02_ЗАГРУЗКА_WB.xlsx"
    if len(longest_member) > 220:
        raise AssertionError(f"ZIP member path must stay bounded for Windows extraction: {longest_member!r}")


def _check_structured_table_expansion() -> None:
    original_path = export_module.WB_UPLOAD_TEMPLATE_PATH
    original_sha256 = export_module.WB_UPLOAD_TEMPLATE_SHA256
    with TemporaryDirectory(prefix="wb-table-template-") as tmp:
        synthetic_path = Path(tmp) / "template-with-table.xlsx"
        workbook = load_workbook(WB_UPLOAD_TEMPLATE_PATH)
        sheet = workbook[WB_UPLOAD_SHEET_NAME]
        sheet.append(["0000000000000", 1])
        table = Table(displayName="WbSupplyUpload", ref="A1:B2")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        workbook.save(synthetic_path)
        workbook.close()
        export_module.WB_UPLOAD_TEMPLATE_PATH = synthetic_path
        export_module.WB_UPLOAD_TEMPLATE_SHA256 = sha256(synthetic_path.read_bytes()).hexdigest()
        try:
            generated = build_wb_upload_workbook_bytes(
                [WbUploadRow("0000000000001", 2), WbUploadRow("0000000000002", 3)],
                expected_total=5,
            )
            result = load_workbook(BytesIO(generated))
            try:
                if result[WB_UPLOAD_SHEET_NAME].tables["WbSupplyUpload"].ref != "A1:B3":
                    raise AssertionError("canonical structured table range must expand over all generated rows")
            finally:
                result.close()
            empty_generated = build_wb_upload_workbook_bytes([], expected_total=0)
            empty_result = load_workbook(BytesIO(empty_generated))
            try:
                if empty_result[WB_UPLOAD_SHEET_NAME].tables["WbSupplyUpload"].ref != "A1:B1":
                    raise AssertionError("header-only recommendation must keep a valid bounded table range")
            finally:
                empty_result.close()
        finally:
            export_module.WB_UPLOAD_TEMPLATE_PATH = original_path
            export_module.WB_UPLOAD_TEMPLATE_SHA256 = original_sha256


def _seed_nomenclature(runtime: RegistryUploadDbBackedRuntime) -> None:
    rows = [
        _nomenclature(101, "0012345678901", "Clean"),
        _nomenclature(102, "0012345678901", "Matte"),
        _nomenclature(103, "4600000000003", "Anti-spy"),
    ]
    runtime.save_nomenclature_items_atomic(rows)


def _nomenclature(nm_id: int, barcode: str, name: str) -> dict[str, object]:
    return {
        "item_id": f"export-{nm_id}",
        "is_active": True,
        "nm_id": nm_id,
        "barcode": barcode,
        "barcodes": [barcode],
        "barcode_source": "manual",
        "barcode_status": "manual",
        "nomenclature_name": name,
        "product_type": "clear",
        "match_key": f"export-{nm_id}",
        "created_at": CALCULATED_AT,
        "updated_at": CALCULATED_AT,
    }


def _result_payload(
    *,
    calculation_id: str = CALCULATION_ID,
) -> dict[str, object]:
    districts = [
        {
            "district_key": PLANNING_ZONE_CENTRAL_NORTH,
            "district_name_ru": "ЦФО Север",
            "planning_zone_key": PLANNING_ZONE_CENTRAL_NORTH,
            "planning_zone_label": "ЦФО Север",
            "total_qty": 5,
            "deficit_qty": 2,
            "rows": [
                _row(101, "Clean", allocated_qty=2),
                _row(102, "Matte", allocated_qty=3),
                _row(104, "Нулевая позиция", allocated_qty=0, deficit_qty=2),
            ],
        },
        {
            "district_key": PLANNING_ZONE_CENTRAL_SOUTH,
            "district_name_ru": "ЦФО Юг",
            "planning_zone_key": PLANNING_ZONE_CENTRAL_SOUTH,
            "planning_zone_label": "ЦФО Юг",
            "total_qty": 4,
            "deficit_qty": 0,
            "rows": [_row(103, "Anti-spy", allocated_qty=4)],
        },
    ]
    return {
        "payload_version": "v2_planning_zones",
        "status": "success",
        "calculation_id": calculation_id,
        "calculated_at": CALCULATED_AT,
        "report_date": REPORT_DATE,
        "horizon_days": 7,
        "active_sku_count": 4,
        "methodology_note": "test",
        "settings": {
            "sales_avg_period_days": 14,
            "cycle_supply_days": 7,
            "lead_time_to_region_days": 15,
            "lead_time_to_region_days_by_district": {
                PLANNING_ZONE_CENTRAL_NORTH: 15,
                PLANNING_ZONE_CENTRAL_SOUTH: 15,
            },
            "safety_days": 15,
            "order_batch_qty": 1,
            "stock_ff_source": STOCK_FF_SOURCE_MANUAL_EXCEL,
            "included_district_keys": [PLANNING_ZONE_CENTRAL_NORTH, PLANNING_ZONE_CENTRAL_SOUTH],
        },
        "stock_ff_source": STOCK_FF_SOURCE_MANUAL_EXCEL,
        "shared_datasets": {},
        "summary": {"total_qty": 9, "estimated_weight": 0.77, "estimated_volume": 0.0},
        "districts": districts,
    }


def _row(
    nm_id: int,
    sku_comment: str,
    *,
    allocated_qty: int | float,
    deficit_qty: int = 0,
) -> dict[str, object]:
    full = allocated_qty + deficit_qty
    return {
        "nm_id": nm_id,
        "sku_comment": sku_comment,
        "full_recommendation_qty": full,
        "allocated_qty": allocated_qty,
        "deficit_qty": deficit_qty,
        "current_stock": 0,
        "projected_stock_on_eta": 0,
        "target_stock_after_arrival": full,
        "daily_demand_total": 0,
        "district_daily_demand": 0,
    }


if __name__ == "__main__":
    main()
