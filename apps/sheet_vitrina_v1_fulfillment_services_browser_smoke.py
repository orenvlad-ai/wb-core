"""Browser live-flow smoke for Fulfillment services operator UI."""

from __future__ import annotations

from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import threading

from openpyxl import load_workbook
from playwright.sync_api import expect, sync_playwright

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.sheet_vitrina_v1_fulfillment_services_smoke import (  # noqa: E402
    NOW,
    _get_bytes,
    _get_json,
    _pdf_text,
    _reserve_free_port,
    _seed_wb_supplies,
    _valid_row,
    _wb_supply_row,
)
from packages.adapters.registry_upload_http_entrypoint import (  # noqa: E402
    DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH,
    DEFAULT_SHEET_OPERATOR_UI_PATH,
    DEFAULT_SHEET_PLAN_PATH,
    DEFAULT_SHEET_STATUS_PATH,
    DEFAULT_SHEET_WEB_VITRINA_UI_PATH,
    DEFAULT_UPLOAD_PATH,
    DEFAULT_WB_SUPPLIES_PATH,
    build_registry_upload_http_server,
)
from packages.application.fulfillment_services import TEMPLATE_HEADERS  # noqa: E402
from packages.application.registry_upload_db_backed_runtime import DB_FILENAME, RegistryUploadDbBackedRuntime  # noqa: E402
from packages.application.registry_upload_http_entrypoint import RegistryUploadHttpEntrypoint  # noqa: E402
from packages.contracts.registry_upload_http_entrypoint import RegistryUploadHttpEntrypointConfig  # noqa: E402


def main() -> None:
    with TemporaryDirectory(prefix="fulfillment-services-browser-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        _seed_wb_supplies(runtime)
        real_xlsx_path = _find_real_fulfillment_xlsx()
        real_xlsx_info = _inspect_real_fulfillment_xlsx(real_xlsx_path)
        _seed_real_wb_supplies(runtime, real_xlsx_info)
        port = _reserve_free_port()
        entrypoint = RegistryUploadHttpEntrypoint(
            runtime_dir=runtime_dir,
            runtime=runtime,
            activated_at_factory=lambda: NOW,
        )
        cfg = RegistryUploadHttpEntrypointConfig(
            host="127.0.0.1",
            port=port,
            upload_path=DEFAULT_UPLOAD_PATH,
            sheet_plan_path=DEFAULT_SHEET_PLAN_PATH,
            sheet_refresh_path="/v1/sheet-vitrina-v1/refresh",
            sheet_status_path=DEFAULT_SHEET_STATUS_PATH,
            sheet_operator_ui_path=DEFAULT_SHEET_OPERATOR_UI_PATH,
            runtime_dir=runtime_dir,
        )
        server = build_registry_upload_http_server(cfg, entrypoint=entrypoint)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            base_url = f"http://127.0.0.1:{cfg.port}"
            with sync_playwright() as playwright:
                browser = playwright.chromium.launch()
                page = browser.new_page(viewport={"width": 1440, "height": 1000}, accept_downloads=True)
                page.goto(f"{base_url}{DEFAULT_SHEET_WEB_VITRINA_UI_PATH}", wait_until="domcontentloaded")
                page.locator("[data-unified-tab-button='factory-order']").click()
                operator_frame = page.frame_locator("iframe[title='Поставки']")
                ff_tab = operator_frame.get_by_role("button", name="ФФ", exact=True)
                expect(ff_tab).to_be_visible(timeout=10000)
                ff_tab.click()
                fulfillment_tab = operator_frame.get_by_role("button", name="Услуги ФФ", exact=True)
                expect(fulfillment_tab).to_be_visible(timeout=10000)
                fulfillment_tab.click()
                expect(operator_frame.locator("#fulfillmentServicesTitle")).to_contain_text("Услуги ФФ")
                expect(operator_frame.locator(".fulfillment-services-block")).to_contain_text("Загруженные документы")
                if operator_frame.locator("#fulfillmentLatestBlock").count() != 0:
                    raise AssertionError("Fulfillment UI must not render the old latest-upload block")
                if operator_frame.locator("#fulfillmentServicesSummary").count() != 0:
                    raise AssertionError("Fulfillment UI must not render technical summary chips")
                expect(operator_frame.locator("#fulfillmentUploadsBody")).to_contain_text("Загруженных документов пока нет")

                template_path = Path(tmp) / "downloaded-fulfillment-template.xlsx"
                with page.expect_download() as download_info:
                    operator_frame.locator("#fulfillmentTemplateButton").click()
                download = download_info.value
                if not download.suggested_filename.endswith(".xlsx"):
                    raise AssertionError(f"template download must be xlsx, got {download.suggested_filename!r}")
                download.save_as(str(template_path))
                template_headers = _read_headers(template_path)
                if template_headers[:2] != ["Номер поставки", "Склад"]:
                    raise AssertionError(f"downloaded template must expose Номер поставки/Склад, got {template_headers[:2]}")

                operator_frame.locator("#fulfillmentFileInput").set_input_files(str(real_xlsx_path))
                expect(operator_frame.locator("#fulfillmentServicesMessage")).to_contain_text(
                    "Документ принят. Все строки смэтчены, PDF-виза сформирована.",
                    timeout=10000,
                )
                list_status, list_payload = _get_json(f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}")
                latest_upload = (list_payload.get("uploads") or [{}])[0]
                upload_id = str(latest_upload.get("upload_id") or "")
                if list_status != 200 or latest_upload.get("validation_status") != "ok" or not upload_id:
                    raise AssertionError(f"real XLSX upload must be latest OK upload, got {list_status} {list_payload}")
                fulfillment_section_text = _normalize_ui_text(operator_frame.locator(".fulfillment-services-block").inner_text())
                for expected in (
                    "Дата загрузки",
                    real_xlsx_path.name,
                    f"{latest_upload.get('rows_matched')}/{latest_upload.get('rows_total')}",
                    _format_ru_rub(latest_upload.get("amount_without_vat_total")),
                    _format_ru_rub(latest_upload.get("vat_total")),
                    _format_ru_rub(latest_upload.get("amount_with_vat_total")),
                    "Скачать PDF-визу",
                    "Удалить",
                ):
                    if expected not in fulfillment_section_text:
                        raise AssertionError(f"accepted table must expose {expected!r}: {fulfillment_section_text}")
                for forbidden in ("Последняя загрузка", "upload_id:", "hash:", "status:", "payment_validation_id:"):
                    if forbidden in fulfillment_section_text:
                        raise AssertionError(f"Fulfillment main screen must not expose technical token {forbidden!r}: {fulfillment_section_text}")

                pdf_status, pdf_body, _ = _get_bytes(
                    f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}/{upload_id}/payment-validation.pdf"
                )
                pdf_text = _pdf_text(pdf_body)
                for expected in (
                    "Виза на оплату услуг ФФ",
                    upload_id,
                    str(latest_upload.get("payment_validation_id") or ""),
                    str(latest_upload.get("short_file_hash") or ""),
                    "Хранение",
                    str(real_xlsx_info["ordinary_supply_ids"][0]),
                ):
                    if pdf_status != 200 or expected not in pdf_text:
                        raise AssertionError(f"PDF validation text missing {expected!r}: {pdf_text!r}")

                operator_frame.get_by_role("button", name="Wildberries", exact=True).click()
                expect(operator_frame.locator("#wbSuppliesTitle")).to_contain_text("Все поставки", timeout=10000)
                operator_frame.locator("#wbSuppliesSizeFilterSelect").select_option("all")
                real_supply_id = str(real_xlsx_info["ordinary_supply_ids"][0])
                operator_frame.locator("#wbSuppliesSearchInput").fill(real_supply_id)
                expect(operator_frame.locator("#wbSuppliesTableBody")).to_contain_text(real_supply_id, timeout=10000)
                actual_columns = operator_frame.locator("#wbSuppliesTableBody").locator("xpath=ancestor::table[1]//thead//th").evaluate_all(
                    "(nodes) => nodes.map((node) => node.textContent.trim())"
                )
                if "Стоимость" in actual_columns or "Транзит" not in actual_columns or "Услуги ФФ" not in actual_columns:
                    raise AssertionError(f"WB supplies columns must expose new overlay contract, got {actual_columns}")
                row_real_text = _normalize_ui_text(
                    operator_frame.locator("#wbSuppliesTableBody tr", has_text=real_supply_id).inner_text()
                )
                for expected in ("₽/шт", "в т.ч. хранение"):
                    if expected not in row_real_text:
                        raise AssertionError(f"WB row {real_supply_id} must show FF amount/per-unit/storage note {expected!r}: {row_real_text}")
                wb_real_status, wb_real_payload = _get_json(f"{base_url}{DEFAULT_WB_SUPPLIES_PATH}?search={real_supply_id}&size_filter=all")
                wb_real_row = (wb_real_payload.get("rows") or [{}])[0]
                if (
                    wb_real_status != 200
                    or wb_real_row.get("fulfillment_amount_with_vat_total") is None
                    or not wb_real_row.get("fulfillment_storage_allocated_amount_with_vat_total")
                    or "₽/шт" not in str(wb_real_row.get("fulfillment_storage_per_unit_display") or "")
                ):
                    raise AssertionError(f"WB API overlay must include storage allocation, got {wb_real_status} {wb_real_payload}")
                if "Seller Portal" in row_real_text:
                    raise AssertionError(f"transit cell must show per-unit instead of Seller Portal source label: {row_real_text}")

                ff_tab.click()
                operator_frame.locator("#fulfillmentUploadsBody [data-fulfillment-delete]").first.click()
                expect(operator_frame.locator("#fulfillmentUploadsBody")).to_contain_text(
                    "Удалить документ? Данные услуг ФФ по связанным WB-поставкам будут удалены из системы.",
                    timeout=10000,
                )
                operator_frame.locator("#fulfillmentUploadsBody [data-fulfillment-delete-cancel]").first.click()
                expect(operator_frame.locator("#fulfillmentUploadsBody")).to_contain_text(real_xlsx_path.name)
                if operator_frame.locator("#fulfillmentUploadsBody [data-fulfillment-delete-confirm]").count() != 0:
                    raise AssertionError("cancel must close Fulfillment delete confirmation")
                operator_frame.locator("#fulfillmentUploadsBody [data-fulfillment-delete]").first.click()
                operator_frame.locator("#fulfillmentUploadsBody [data-fulfillment-delete-confirm]").first.click()
                expect(operator_frame.locator("#fulfillmentUploadsBody")).to_contain_text("Загруженных документов пока нет", timeout=10000)
                if "fulfillment-ok.xlsx" in operator_frame.locator("#fulfillmentUploadsBody").inner_text():
                    raise AssertionError("deleted accepted document must disappear from accepted table")
                if real_xlsx_path.name in operator_frame.locator("#fulfillmentUploadsBody").inner_text():
                    raise AssertionError("deleted real accepted document must disappear from accepted table")
                pdf_after_delete_status, _, _ = _get_bytes(
                    f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}/{upload_id}/payment-validation.pdf"
                )
                if pdf_after_delete_status != 404:
                    raise AssertionError(f"deleted upload PDF must be unavailable, got HTTP {pdf_after_delete_status}")
                operator_frame.get_by_role("button", name="Wildberries", exact=True).click()
                operator_frame.locator("#wbSuppliesSizeFilterSelect").select_option("all")
                operator_frame.locator("#wbSuppliesSearchInput").fill(real_supply_id)
                expect(operator_frame.locator("#wbSuppliesTableBody")).to_contain_text(real_supply_id, timeout=10000)
                row_real_after_delete = _normalize_ui_text(
                    operator_frame.locator("#wbSuppliesTableBody tr", has_text=real_supply_id).inner_text()
                )
                if "в т.ч. хранение" in row_real_after_delete:
                    raise AssertionError(f"deleted upload storage allocation must disappear from WB overlay: {row_real_after_delete}")
                wb_real_after_status, wb_real_after_payload = _get_json(f"{base_url}{DEFAULT_WB_SUPPLIES_PATH}?search={real_supply_id}&size_filter=all")
                wb_real_after_row = (wb_real_after_payload.get("rows") or [{}])[0]
                if wb_real_after_status != 200 or wb_real_after_row.get("fulfillment_amount_with_vat_total") is not None:
                    raise AssertionError(f"deleted upload must disappear from WB API overlay, got {wb_real_after_status} {wb_real_after_payload}")

                ff_tab.click()
                unmatched_xlsx_path = Path(tmp) / "fulfillment-unmatched.xlsx"
                _fill_downloaded_template(template_path, unmatched_xlsx_path, [_valid_row("9999")])
                operator_frame.locator("#fulfillmentFileInput").set_input_files(str(unmatched_xlsx_path))
                expect(operator_frame.locator("#fulfillmentServicesMessage")).to_contain_text(
                    "Документ не принят. Исправьте ошибки и загрузите файл заново.",
                    timeout=10000,
                )
                expect(operator_frame.locator("#fulfillmentErrorsBody")).to_contain_text("9999")
                expect(operator_frame.locator("#fulfillmentErrorsBody")).to_contain_text("не найдена в WB-поставках")
                if "fulfillment-unmatched.xlsx" in operator_frame.locator("#fulfillmentUploadsBody").inner_text():
                    raise AssertionError("failed unmatched upload must not enter accepted table")

                duplicate_xlsx_path = Path(tmp) / "fulfillment-duplicate.xlsx"
                _fill_downloaded_template(template_path, duplicate_xlsx_path, [_valid_row("1001"), _valid_row("1001")])
                operator_frame.locator("#fulfillmentFileInput").set_input_files(str(duplicate_xlsx_path))
                expect(operator_frame.locator("#fulfillmentServicesMessage")).to_contain_text(
                    "Документ не принят. Исправьте ошибки и загрузите файл заново.",
                    timeout=10000,
                )
                expect(operator_frame.locator("#fulfillmentErrorsBody")).to_contain_text("дублируется номер поставки 1001")
                if "fulfillment-duplicate.xlsx" in operator_frame.locator("#fulfillmentUploadsBody").inner_text():
                    raise AssertionError("failed duplicate upload must not enter accepted table")

                wb_status, wb_payload = _get_json(f"{base_url}{DEFAULT_WB_SUPPLIES_PATH}?search=1001&size_filter=all")
                wb_row = (wb_payload.get("rows") or [{}])[0]
                if wb_status != 200 or wb_row.get("fulfillment_amount_with_vat_total") is not None:
                    raise AssertionError(f"failed/deleted uploads must not alter approved WB overlay, got {wb_status} {wb_payload}")

                browser.close()
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)

        print(f"runtime_db: ok -> {runtime_dir / DB_FILENAME}")

    print("sheet_vitrina_v1_fulfillment_services_browser_smoke: OK")


def _fill_downloaded_template(template_path: Path, output_path: Path, rows: list[list[object]]) -> None:
    workbook = load_workbook(template_path)
    sheet = workbook.worksheets[0]
    headers = _read_headers(template_path)
    if headers != TEMPLATE_HEADERS:
        raise AssertionError(f"downloaded template headers mismatch: {headers}")
    for row in rows:
        sheet.append(row)
    if rows:
        total = sum(float(row[8]) for row in rows)
        vat = sum(float(row[9]) for row in rows)
        sheet.append(["", "", "", "", "", "", "", "", total, vat])
    workbook.save(output_path)


def _find_real_fulfillment_xlsx() -> Path:
    downloads = Path.home() / "Downloads"
    patterns = [
        "fulfillment_services_filled_2026-07-07_batch*.xlsx",
        "fulfillment_services_filled_2026-07-06_second_batch*.xlsx",
        "fulfillment_services_filled*.xlsx",
    ]
    checked: list[str] = []
    errors: list[str] = []
    for pattern in patterns:
        for path in sorted(downloads.glob(pattern)):
            checked.append(str(path))
            try:
                _inspect_real_fulfillment_xlsx(path)
            except AssertionError as exc:
                errors.append(f"{path}: {exc}")
                continue
            return path
    raise AssertionError(
        "real Fulfillment XLSX with STORAGE was not found; checked paths: "
        + ", ".join(checked or [str(downloads / pattern) for pattern in patterns])
        + ("; errors: " + " | ".join(errors) if errors else "")
    )


def _inspect_real_fulfillment_xlsx(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header_row = 0
    columns: dict[str, int] = {}
    for row_index in range(1, min(sheet.max_row or 0, 20) + 1):
        values = [sheet.cell(row_index, column).value for column in range(1, (sheet.max_column or 0) + 1)]
        normalized = [_normalize_header(value) for value in values]
        if {"номер поставки", "итого", "ндс 5%"}.issubset(set(normalized)):
            header_row = row_index
            columns = {
                "supply_id": normalized.index("номер поставки") + 1,
                "boxes_qty": normalized.index("кол-во коробов") + 1 if "кол-во коробов" in normalized else 0,
                "total": normalized.index("итого") + 1,
                "vat": normalized.index("ндс 5%") + 1,
            }
            break
    if not header_row:
        raise AssertionError("header row with Номер поставки/Итого/НДС 5% was not found")
    ordinary_rows: list[dict] = []
    storage_total = 0.0
    for row_index in range(header_row + 1, (sheet.max_row or header_row) + 1):
        supply_id = _cell_text(sheet.cell(row_index, columns["supply_id"]).value)
        if not supply_id:
            continue
        total = _number_or_none(sheet.cell(row_index, columns["total"]).value)
        vat = _number_or_none(sheet.cell(row_index, columns["vat"]).value)
        if supply_id.casefold() == "storage":
            if total is None or vat is None:
                raise AssertionError(f"STORAGE row {row_index} has nonnumeric totals")
            storage_total += total + vat
            continue
        boxes = _number_or_none(sheet.cell(row_index, columns["boxes_qty"]).value) if columns.get("boxes_qty") else None
        ordinary_rows.append({"supply_id": supply_id, "boxes_qty": boxes})
    if not ordinary_rows:
        raise AssertionError("no ordinary WB supply rows found")
    if storage_total <= 0:
        raise AssertionError("no positive STORAGE total found")
    return {
        "path": str(path),
        "ordinary_rows": ordinary_rows,
        "ordinary_supply_ids": [item["supply_id"] for item in ordinary_rows],
        "storage_total_with_vat": storage_total,
    }


def _seed_real_wb_supplies(runtime: RegistryUploadDbBackedRuntime, info: dict) -> None:
    rows = []
    for item in info.get("ordinary_rows") or []:
        supply_id = str(item.get("supply_id") or "").strip()
        if not supply_id:
            continue
        boxes = _number_or_none(item.get("boxes_qty"))
        quantity = max(1, int(round(boxes or 100)))
        rows.append(_wb_supply_row(supply_id, accepted_quantity=quantity, quantity_added=quantity, cost_total=200))
    if rows:
        runtime.save_wb_supply_rows(
            rows=rows,
            warehouses=[{"ID": 777, "name": "Электросталь"}],
            synced_at=NOW,
        )


def _read_headers(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    return ["" if value is None else str(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def _normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().replace("ё", "е").lower().split())


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and abs(value - round(value)) < 0.000001:
        return str(int(round(value)))
    return str(value).strip()


def _number_or_none(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace("\u00a0", "").replace(" ", "").replace(",", ".").replace("₽", ""))
    except (TypeError, ValueError):
        return None


def _format_ru_rub(value: object) -> str:
    numeric = _number_or_none(value)
    if numeric is None:
        return "—"
    if abs(numeric) < 0.005:
        return "0 ₽"
    if abs(numeric - round(numeric)) < 0.005:
        return f"{int(round(numeric)):,}".replace(",", " ") + " ₽"
    integer, fractional = f"{numeric:.2f}".split(".")
    fractional = fractional.rstrip("0")
    return f"{int(integer):,}".replace(",", " ") + f",{fractional} ₽"


def _normalize_ui_text(value: str) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").replace("\u202f", " ").split())


if __name__ == "__main__":
    main()
