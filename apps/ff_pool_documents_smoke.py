"""Fixture smoke for Stage 2 facility × pool documents and XLSX contracts."""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from dataclasses import replace
from datetime import datetime, timedelta, timezone
from decimal import Decimal, localcontext
from io import BytesIO
import json
from pathlib import Path
import sqlite3
import sys
from tempfile import TemporaryDirectory
import threading
import zipfile

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from packages.application.ff_pool_documents import (  # noqa: E402
    DOCUMENTS_TABLE,
    DOCUMENT_LINES_TABLE,
    DOCUMENT_RELATIONS_TABLE,
    EXPENSE_LINES_TABLE,
    REQUESTS_TABLE,
    FfPoolDocumentError,
    FfPoolDocumentService,
    _allocate_cents,
    _apply_balance_movement,
    _apply_guided_aggregate_projection,
    _build_posting_plan,
    _component_share,
    _posting_plan_preview,
)
from packages.application.ff_pool_documents_xlsx import (  # noqa: E402
    DEFAULT_LIMITS,
    INVENTORY_SHEET,
    CHINA_SHEET,
    XLSX_CONTENT_TYPE,
    FfPoolXlsxError,
    parse_inventory_workbook,
    validate_xlsx_request_seam,
)
from packages.application.ff_pool_foundation import (  # noqa: E402
    FACILITY_CHANGES_TABLE,
    BALANCES_TABLE,
    FACILITIES_TABLE,
    FEATURE_EPOCHS_TABLE,
    LINES_TABLE,
    OPERATIONS_TABLE,
    evaluate_ff_pool_aggregate_parity,
)
from packages.application.warehouse_functional import STAGES  # noqa: E402
from packages.application.warehouse_recovery_policy import (  # noqa: E402
    DOMAIN_EXACT_TABLES,
    WarehouseRecoveryRegistry,
)
from packages.contracts.ff_pool_documents import DocumentIdentity  # noqa: E402


class Clock:
    def __init__(self) -> None:
        self.value = datetime(2026, 8, 11, 8, 0, tzinfo=timezone.utc)
        self.lock = threading.Lock()

    def __call__(self) -> str:
        with self.lock:
            result = self.value.isoformat(timespec="seconds").replace("+00:00", "Z")
            self.value += timedelta(seconds=1)
            return result


def identity(name: str, *, revision: str = "fixture-v1", epoch: int = 1) -> DocumentIdentity:
    return DocumentIdentity(
        request_id=f"fixture:{name}:request",
        source_system="fixture",
        source_type="ff_pool_stage2_smoke",
        source_id=name,
        source_revision=revision,
        idempotency_epoch=epoch,
        actor="fixture-operator",
        business_date="2026-08-11",
    )


def main() -> None:
    assert STAGES == (
        "production",
        "china_to_ff",
        "ff",
        "ff_to_wb",
        "wb",
        "wb_acceptance_discrepancy",
    )
    _default_off_and_empty_schema()
    with TemporaryDirectory(prefix="ff-pool-documents-") as directory:
        root = Path(directory)
        db_path = root / "state.sqlite3"
        clock = Clock()
        service = FfPoolDocumentService(
            db_path=db_path,
            runtime_dir=root,
            timestamp_factory=clock,
        )
        _seed(service, clock)
        catalog, shipment_lines = _xlsx_contracts(service)
        _production_shaped_26gn527(service)
        opening = _opening(service)
        _inventory_missing_fbs_zero(service, catalog)
        _idempotent_repeat_and_recovery(service, opening)
        _transfer_and_late_expense(service)
        _mis_sort(service)
        _reallocation_inventory_overhead(service, catalog)
        _property_checks()
        _relations_and_indexes(db_path)
        _assert_recovery(service)
        _assert_no_aggregate_or_legacy_writes(db_path)
    print("ff_pool_documents_smoke: OK")


def _production_shaped_26gn527(service: FfPoolDocumentService) -> None:
    rows = [
        (210183142, 3250, 750, 2500, "321477.9087517898"),
        (210183919, 7000, 2500, 4500, "692413.9573115471"),
        (210184534, 3000, 500, 2500, "296748.8388478059"),
        (245720334, 1250, 750, 500, "128000.62296944665"),
        (259460529, 6000, 5000, 1000, "614402.990253344"),
        (259465495, 5750, 1500, 4250, "608837.1235272811"),
        (259473237, 6000, 5500, 500, "614402.990253344"),
        (391659990, 250, 250, 0, "31697.50742322784"),
        (428850065, 250, 250, 0, "30826.452733322334"),
        (428853741, 1500, 500, 1000, "176248.169500879"),
        (428854140, 500, 500, 0, "60491.49921343734"),
        (428854299, 1000, 500, 500, "120982.99842687468"),
        (428855306, 2000, 750, 1250, "234997.55933450535"),
        (428855560, 1500, 500, 1000, "176248.169500879"),
        (428855758, 1750, 500, 1250, "199525.48158835364"),
        (428855978, 1750, 250, 1500, "199525.48158835364"),
        (497414010, 6500, 1500, 5000, "688250.6613786656"),
        (497414624, 3000, 1500, 1500, "317654.15140553797"),
        (497416271, 4750, 750, 4000, "618802.6801495334"),
        (497417163, 3750, 1250, 2500, "453686.24410078005"),
        (497417474, 5250, 1250, 4000, "635160.7417410921"),
    ]
    shipment_lines = [
        {
            "nm_id": nm_id,
            "barcode": str(nm_id),
            "sku": f"26GN527-{nm_id}",
            "accepted_quantity": quantity,
            "accepted_capital_rub": capital,
            "identity_evidence_digest": "sha256:" + f"{nm_id:064d}"[-64:],
        }
        for nm_id, quantity, _fbo, _fbs, capital in rows
    ]
    source_revision = "26gn527-production-shaped-v1"
    template = service.generate_china_acceptance_template(
        shipment_lines=shipment_lines,
        source_revision=source_revision,
        selected_facility_id="fac_msk",
    )
    workbook = load_workbook(BytesIO(template))
    sheet = workbook[CHINA_SHEET]
    for row_no, (_nm_id, _quantity, fbo, fbs, _capital) in enumerate(rows, start=6):
        sheet.cell(row=row_no, column=7, value=fbo)
        sheet.cell(row=row_no, column=8, value=fbs)
    filled = BytesIO()
    workbook.save(filled)
    evidence = filled.getvalue()
    request_identity = identity("26gn527", revision=source_revision)
    preview = service.preview_china_acceptance_workbook(
        identity=request_identity,
        source_bytes=evidence,
        source_filename="FF_приёмка_26GN527_заполнено.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        shipment_lines=shipment_lines,
        expenses=[],
        template_source_revision=source_revision,
    )
    assert preview["state"] == "ready", preview
    manifest = preview["preview_manifest"]
    assert len(manifest["allocations"]) == 21
    assert sum(int(item["expected_quantity"]) for item in manifest["allocations"]) == 66_000
    assert sum(int(item["accepted_quantity"]) for item in manifest["allocations"]) == 66_000
    assert sum(int(item["quantity_fbs"]) for item in manifest["allocations"]) == 39_250
    assert sum(int(item["quantity_fbo"]) for item in manifest["allocations"]) == 26_750
    assert manifest["expenses"] == []
    normalization = manifest["capital_normalization"]
    assert normalization["exact_total_rub"] == "7220382.230000000494"
    assert normalization["canonical_total_rub"] == "7220382.23"
    assert normalization["total_residual_rub"] == "-0.000000000494"
    assert normalization["residual_owner_nm_ids"] == [
        428853741, 428855560, 428855306, 428854140, 210184534,
        210183142, 428854299, 391659990, 210183919, 497417163,
    ]
    with sqlite3.connect(service.db_path) as conn:
        conn.row_factory = sqlite3.Row
        request = conn.execute(
            f"SELECT * FROM {REQUESTS_TABLE} WHERE request_id=?",
            (str(preview["request_id"]),),
        ).fetchone()
        plan = _build_posting_plan(
            conn,
            request=request,
            manifest=manifest,
            epoch=1,
        )
    movements = [
        movement
        for document in plan["documents"]
        for movement in document.get("movements") or []
    ]
    assert sum(int(item["quantity_delta"]) for item in movements) == 66_000
    assert sum(int(item["capital_delta_cents"]) for item in movements) == 722_038_223
    assert sum(int(item["quantity_delta"]) for item in movements if item["pool"] == "FBS") == 39_250
    assert sum(int(item["quantity_delta"]) for item in movements if item["pool"] == "FBO") == 26_750
    per_nm = {}
    for item in movements:
        per_nm[int(item["nm_id"])] = per_nm.get(int(item["nm_id"]), 0) + int(
            item["capital_delta_cents"]
        )
    assert per_nm == {
        int(key): int(value)
        for key, value in normalization["capital_cents_by_nm"].items()
    }
    # Production opening balances retain authoritative exact Decimal capital.
    # A guided receipt adds only its canonical kopeck delta; it must not reject
    # or silently normalize the pre-existing fractional-kopeck capital.  The
    # exact inverse movement must return every affected balance byte-for-
    # semantic-byte to the original quantity/capital pair.
    fractional_opening = {
        210183142: (2249, "205231.4311365185716975379064"),
        210184534: (250, "22685.48291654259178871196266"),
        245720334: (743, "72996.70782552040825768721842"),
        259460529: (1228, "117162.7582051580009068280460"),
        259465495: (3132, "315544.9438867871117622305145"),
        259473237: (286, "27090.34575724342461900569472"),
        428853741: (500, "57862.8139406905458181875208"),
        428854299: (2946, "357448.9272079002304083502190"),
        428855306: (998, "114141.4070768659663193726124"),
        428855978: (205, "22944.60690012700219304333326"),
        497414010: (3889, "392247.7763148711604622173653"),
        497414624: (9746, "982128.3651914405693875570705"),
        497416271: (747, "99225.61859946938832759534936"),
        497417163: (664, "80565.54453038891819115565094"),
        497417474: (3248, "392768.2532584173349552632159"),
    }
    with sqlite3.connect(service.db_path) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("SAVEPOINT production_decimal_pool_apply")
        for nm_id, (quantity, capital) in fractional_opening.items():
            with localcontext() as context:
                context.prec = 38
                opening_wac = str(Decimal(capital) / Decimal(quantity))
            conn.execute(
                f"""INSERT INTO {BALANCES_TABLE}(
                       facility_id,pool,nm_id,projection_epoch,quantity,capital_rub,
                       wac_rub,source_watermark,updated_at
                   ) VALUES('fac_msk','FBS',?,1,?,?,?,?,?)""",
                (
                    nm_id,
                    quantity,
                    capital,
                    opening_wac,
                    "production-shaped-opening",
                    "2026-08-15T00:00:00Z",
                ),
            )
        movement_by_key = {
            (str(item["facility_id"]), str(item["pool"]), int(item["nm_id"])): item
            for item in movements
        }
        for line_no, movement in enumerate(movements, start=1):
            _apply_balance_movement(
                conn,
                movement=movement,
                operation_id="production-shaped-guided-forward",
                line_no=line_no,
                epoch=1,
                posted_at="2026-08-15T01:00:00Z",
            )
        for nm_id, (quantity, capital) in fractional_opening.items():
            row = conn.execute(
                f"SELECT quantity,capital_rub FROM {BALANCES_TABLE} "
                "WHERE facility_id='fac_msk' AND pool='FBS' AND nm_id=?",
                (nm_id,),
            ).fetchone()
            movement = movement_by_key[("fac_msk", "FBS", nm_id)]
            assert int(row[0]) == quantity + int(movement["quantity_delta"])
            with localcontext() as context:
                context.prec = 160
                expected_capital = Decimal(capital) + (
                    Decimal(int(movement["capital_delta_cents"])) / Decimal(100)
                )
            assert str(row[1]) == format(expected_capital, "f")
        for line_no, movement in enumerate(movements, start=1):
            inverse = dict(movement)
            inverse["quantity_delta"] = -int(movement["quantity_delta"])
            inverse["capital_delta_cents"] = -int(
                movement["capital_delta_cents"]
            )
            _apply_balance_movement(
                conn,
                movement=inverse,
                operation_id="production-shaped-guided-storno",
                line_no=line_no,
                epoch=1,
                posted_at="2026-08-15T02:00:00Z",
            )
        for nm_id, (quantity, capital) in fractional_opening.items():
            row = conn.execute(
                f"SELECT quantity,capital_rub FROM {BALANCES_TABLE} "
                "WHERE facility_id='fac_msk' AND pool='FBS' AND nm_id=?",
                (nm_id,),
            ).fetchone()
            assert int(row[0]) == quantity
            assert str(row[1]) == capital
        movement_capital = conn.execute(
            f"SELECT capital_delta_rub FROM {LINES_TABLE} "
            "WHERE operation_id IN ('production-shaped-guided-forward',"
            "'production-shaped-guided-storno')"
        ).fetchall()
        assert len(movement_capital) == 78
        assert all(
            Decimal(str(row[0])).quantize(Decimal("0.01"))
            == Decimal(str(row[0]))
            for row in movement_capital
        )
        conn.execute("ROLLBACK TO production_decimal_pool_apply")
        conn.execute("RELEASE production_decimal_pool_apply")
    # Production-shaped guided planning must treat an inbound SKU absent from
    # the current aggregate FF snapshot as exact semantic zero, not as an
    # error.  The three missing rows mirror the real 26GN527 preview evidence.
    missing_aggregate_nm_ids = {210183919, 428855560, 428855758}
    with sqlite3.connect(service.db_path) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute(
            """CREATE TABLE sheet_vitrina_v1_supplier_shipments(
                   shipment_id TEXT PRIMARY KEY,actual_ff_acceptance_date TEXT,
                   order_status TEXT NOT NULL,updated_at TEXT NOT NULL
               )"""
        )
        conn.execute(
            """INSERT INTO sheet_vitrina_v1_supplier_shipments
               VALUES('26gn527',NULL,'in_transit','2026-08-15T00:00:00Z')"""
        )
        conn.execute(
            """CREATE TABLE sheet_vitrina_v1_warehouse_functional_active(
                   slot INTEGER PRIMARY KEY,version_id TEXT NOT NULL
               )"""
        )
        conn.execute(
            """INSERT INTO sheet_vitrina_v1_warehouse_functional_active
               VALUES(1,'26gn527-before')"""
        )
        conn.execute(
            """CREATE TABLE sheet_vitrina_v1_warehouse_functional_balances(
                   version_id TEXT NOT NULL,warehouse_key TEXT NOT NULL,nm_id INTEGER NOT NULL,
                   quantity TEXT NOT NULL,wac_rub TEXT,capital_rub TEXT NOT NULL,
                   cost_covered_quantity TEXT NOT NULL,quality TEXT NOT NULL,
                   certified INTEGER NOT NULL,wb_quantity TEXT NOT NULL,
                   wb_in_way_to_client TEXT NOT NULL,wb_in_way_from_client TEXT NOT NULL,
                   provenance_json TEXT NOT NULL,
                   PRIMARY KEY(version_id,warehouse_key,nm_id)
               )"""
        )
        exact_aggregate_opening = {
            210184534: (250, "22685.48291654259178871196266"),
            245720334: (742, "72898.46191996789088452747788"),
            259473237: (161, "15250.15967453213763517453472"),
            497414010: (3714, "374597.1306848628156226987078"),
            497414624: (9744, "981926.8202775904892378777031"),
            497416271: (747, "99225.61859946938832759534936"),
            497417163: (605, "73406.85909771881853260417024"),
            497417474: (3247, "392647.3270720693000614962014"),
        }
        aggregate_opening_quantity = 0
        for nm_id, quantity, _fbo, _fbs, capital in rows:
            if nm_id in missing_aggregate_nm_ids:
                continue
            quantity, capital = exact_aggregate_opening.get(
                nm_id, (quantity, capital)
            )
            aggregate_opening_quantity += quantity
            conn.execute(
                """INSERT INTO sheet_vitrina_v1_warehouse_functional_balances
                   VALUES('26gn527-before','ff',?,?,?,?,?,'fixture',0,'0','0','0','{}')""",
                (
                    nm_id,
                    str(quantity),
                    str(Decimal(capital) / Decimal(quantity)),
                    capital,
                    str(quantity),
                ),
            )
            conn.execute(
                f"""INSERT INTO {BALANCES_TABLE}(
                       facility_id,pool,nm_id,projection_epoch,quantity,capital_rub,
                       wac_rub,source_watermark,updated_at
                   ) VALUES('fac_msk','FBS',?,1,?,?,?,?,?)""",
                (
                    nm_id,
                    quantity,
                    capital,
                    str(Decimal(capital) / Decimal(quantity)),
                    "production-shaped-global-parity",
                    "2026-08-15T00:00:00Z",
                ),
            )
        guided_request = dict(request)
        guided_request["source_type"] = "china_acceptance_workbook"
        guided_plan = _build_posting_plan(
            conn,
            request=guided_request,
            manifest=manifest,
            epoch=1,
        )
        readiness = _posting_plan_preview(plan=guided_plan, epoch=1)
        assert readiness["confirm_plan_ready"] is True
        assert readiness["quantity_delta"] == 66_000
        assert readiness["capital_delta_rub"] == "7220382.23"
        assert readiness["aggregate_semantic_zero_nm_ids"] == sorted(
            missing_aggregate_nm_ids
        )
        assert readiness["aggregate_pool_parity"]["status"] == "pass"
        assert readiness["aggregate_pool_parity"]["detail_quantity"] == (
            aggregate_opening_quantity
        )
        assert readiness["aggregate_pool_parity"]["aggregate_quantity"] == (
            aggregate_opening_quantity
        )
        assert readiness["aggregate_pool_parity"]["detail_fingerprint"].startswith(
            "sha256:"
        )
        conn.execute("SAVEPOINT guided_exact_aggregate")
        guided_movements = [
            item
            for document in guided_plan["documents"]
            for item in document.get("movements") or []
        ]
        for line_no, movement in enumerate(guided_movements, start=1):
            _apply_balance_movement(
                conn,
                movement=movement,
                operation_id="production-shaped-guided-exact-aggregate",
                line_no=line_no,
                epoch=1,
                posted_at="2026-08-15T03:00:00Z",
            )
        _apply_guided_aggregate_projection(
            conn,
            plan=guided_plan,
            request=guided_request,
            posted_at="2026-08-15T03:00:00Z",
        )
        aggregate_rows = [
            {
                "nm_id": int(row[0]),
                "quantity": int(row[1]),
                "capital_rub": str(row[2]),
            }
            for row in conn.execute(
                """SELECT nm_id,quantity,capital_rub
                   FROM sheet_vitrina_v1_warehouse_functional_balances
                   WHERE version_id='26gn527-before' AND warehouse_key='ff'
                   ORDER BY nm_id"""
            ).fetchall()
        ]
        exact_parity = evaluate_ff_pool_aggregate_parity(conn, aggregate_rows)
        assert exact_parity.status == "pass", exact_parity
        for nm_id, (_quantity, opening_capital) in exact_aggregate_opening.items():
            aggregate_capital = conn.execute(
                """SELECT capital_rub
                   FROM sheet_vitrina_v1_warehouse_functional_balances
                   WHERE version_id='26gn527-before' AND warehouse_key='ff'
                     AND nm_id=?""",
                (nm_id,),
            ).fetchone()[0]
            pool_capitals = conn.execute(
                f"""SELECT capital_rub FROM {BALANCES_TABLE}
                    WHERE facility_id='fac_msk' AND nm_id=? ORDER BY pool""",
                (nm_id,),
            ).fetchall()
            with localcontext() as context:
                context.prec = 160
                expected = sum(
                    (Decimal(str(item[0])) for item in pool_capitals),
                    Decimal("0"),
                )
            assert str(aggregate_capital) == format(expected, "f")
            assert Decimal(str(aggregate_capital)) > Decimal(opening_capital)
        conn.execute("ROLLBACK TO guided_exact_aggregate")
        conn.execute("RELEASE guided_exact_aggregate")
        conn.execute(
            f"DELETE FROM {BALANCES_TABLE} "
            "WHERE source_watermark='production-shaped-global-parity'"
        )
        conn.execute("DROP TABLE sheet_vitrina_v1_warehouse_functional_balances")
        conn.execute("DROP TABLE sheet_vitrina_v1_warehouse_functional_active")
        conn.execute("DROP TABLE sheet_vitrina_v1_supplier_shipments")
        conn.commit()
    repeated = service.preview_china_acceptance_workbook(
        identity=replace(
            request_identity,
            request_id="fixture:26gn527:response-loss-retry",
        ),
        source_bytes=evidence,
        source_filename="renamed-26GN527.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        shipment_lines=shipment_lines,
        expenses=[],
        template_source_revision=source_revision,
    )
    assert repeated["request_id"] == preview["request_id"]
    assert repeated["idempotent"] is True
    stale = service.preview_china_acceptance_workbook(
        identity=identity("26gn527-stale", revision="26gn527-stale-v2"),
        source_bytes=evidence,
        source_filename="stale-26GN527.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        shipment_lines=shipment_lines,
        expenses=[],
        template_source_revision="26gn527-stale-v2",
    )
    assert stale["state"] == "blocked"
    assert stale["error"]["code"] == "template_fingerprint_mismatch"


def _default_off_and_empty_schema() -> None:
    with TemporaryDirectory(prefix="ff-pool-default-off-") as directory:
        root = Path(directory)
        service = FfPoolDocumentService(db_path=root / "state.sqlite3", runtime_dir=root)
        with sqlite3.connect(service.db_path) as conn:
            for table in (
                FACILITIES_TABLE,
                FEATURE_EPOCHS_TABLE,
                BALANCES_TABLE,
                REQUESTS_TABLE,
                DOCUMENTS_TABLE,
                DOCUMENT_LINES_TABLE,
                EXPENSE_LINES_TABLE,
                DOCUMENT_RELATIONS_TABLE,
            ):
                assert conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0] == 0
        try:
            service.generate_inventory_template(
                facility_id="missing",
                scope="FBS",
                catalog=[{"nm_id": 1, "barcode": "001", "sku": "x"}],
                source_revision="empty",
            )
        except FfPoolXlsxError as exc:
            assert exc.code == "no_active_facilities"
        else:
            raise AssertionError("empty facility registry must fail closed")
        preview = service.accept_preview(
            identity=identity("default-off"),
            document_kind="transfer_root",
            manifest={
                "source": {"facility_id": "missing", "pool": "FBS"},
                "destination": {"facility_id": "missing", "pool": "FBO"},
            },
        )
        assert preview["state"] == "ready"
        blocked = service.post(preview["request_id"])
        assert blocked["state"] == "blocked"
        assert blocked["error"]["code"] == "feature_writer_disabled"


def _seed(service: FfPoolDocumentService, clock: Clock) -> None:
    with sqlite3.connect(service.db_path) as conn:
        conn.execute("PRAGMA foreign_keys=ON")
        for facility_id, code, name in (
            ("fac_msk", "MSK", "Москва"),
            ("fac_orb", "ORB", "Оренбург"),
            ("fac_single", "SINGLE", "Только FBS"),
        ):
            now = clock()
            conn.execute(
                f"INSERT INTO {FACILITIES_TABLE}(facility_id,code,name,active,display_timezone,created_at,updated_at) "
                "VALUES(?,?,?,?,?,?,?)",
                (facility_id, code, name, 1, "Asia/Yekaterinburg", now, now),
            )
        conn.execute(
            f"INSERT INTO {FEATURE_EPOCHS_TABLE}(epoch,writer_enabled,reader_enabled,source_revision,created_at,metadata_json) "
            "VALUES(1,1,0,'fixture-writer-v1',?,'{}')",
            (clock(),),
        )
        conn.commit()


def _xlsx_contracts(
    service: FfPoolDocumentService,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    catalog = [
        {
            "nm_id": 101,
            "barcode": "0000000000101",
            "barcodes": ["0000000000101", "101-A"],
            "sku": "SKU-101",
            "active": True,
            "identity_revision": "catalog-v1",
        },
        {
            "nm_id": 102,
            "barcode": "0000000000102",
            "barcodes": ["0000000000102"],
            "sku": "=HYPERLINK(\"https://invalid.example\",\"SKU-102\")",
            "active": True,
            "identity_revision": "catalog-v1",
        },
    ]
    shipment_lines = [
        {
            **catalog[0],
            "accepted_quantity": 10,
            "accepted_capital_rub": "1000.00",
        },
        {
            **catalog[1],
            "accepted_quantity": 5,
            "accepted_capital_rub": "1000.00",
        },
    ]
    china = service.generate_china_acceptance_template(
        shipment_lines=shipment_lines,
        source_revision="china-source-v1",
        selected_facility_id="fac_msk",
    )
    workbook = load_workbook(BytesIO(china))
    sheet = workbook[CHINA_SHEET]
    assert sheet["B2"].value == "MSK — Москва"
    assert len(sheet.data_validations.dataValidation) == 1
    assert sheet["B6"].number_format == "@"
    assert set(sheet.tables) == {"ChinaAcceptanceTable"}
    sheet["G6"], sheet["H6"] = 4, 6
    sheet["G7"], sheet["H7"] = 5, 0
    buffer = BytesIO()
    workbook.save(buffer)
    china_ready = service.preview_china_acceptance_workbook(
        identity=identity("china-xlsx", revision="china-source-v1"),
        source_bytes=buffer.getvalue(),
        source_filename="china-acceptance.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        shipment_lines=shipment_lines,
        expenses=[{"amount_rub": "15.00", "basis": "Общая приёмка"}],
    )
    assert china_ready["state"] == "ready"
    # Stage 2 preview is durable, but the current supplier acceptance trigger is not switched.
    assert china_ready["preview_manifest"]["facility_id"] == "fac_msk"
    renamed = service.preview_china_acceptance_workbook(
        identity=replace(
            identity("china-xlsx", revision="china-source-v1"),
            request_id="fixture:china-xlsx-renamed:request",
        ),
        source_bytes=buffer.getvalue(),
        source_filename="renamed-china-acceptance.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        shipment_lines=shipment_lines,
        expenses=[{"amount_rub": "15.00", "basis": "Общая приёмка"}],
    )
    assert renamed["request_id"] == china_ready["request_id"]
    assert renamed["idempotent"] is True

    inventory = service.generate_inventory_template(
        facility_id="fac_msk",
        scope="both",
        catalog=catalog,
        source_revision="catalog-v1",
    )
    inventory_template = load_workbook(BytesIO(inventory))
    inventory_sheet = inventory_template[INVENTORY_SHEET]
    assert len(inventory_sheet.data_validations.dataValidation) == 2
    assert inventory_sheet["B6"].number_format == "@"
    assert inventory_sheet["C7"].data_type == "s"
    assert set(inventory_sheet.tables) == {"FacilityPoolInventoryTable"}
    parsed = parse_inventory_workbook(
        inventory,
        filename="inventory.xlsx",
        content_type=XLSX_CONTENT_TYPE,
        facilities=service.active_facilities(),
        catalog=catalog,
        source_revision="catalog-v1",
    )
    assert len(parsed["targets"]) == 2
    assert all(item["target_fbo"] == 0 and item["target_fbs"] == 0 for item in parsed["targets"])
    validate_xlsx_request_seam(
        content_length=len(inventory),
        filename="inventory.xlsx",
        content_type=XLSX_CONTENT_TYPE,
    )
    try:
        validate_xlsx_request_seam(
            content_length=DEFAULT_LIMITS.max_request_bytes + 1,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
        )
    except FfPoolXlsxError as exc:
        assert exc.code == "request_too_large"
    else:
        raise AssertionError("pre-buffering request limit must fail closed")

    formula = load_workbook(BytesIO(inventory))
    formula[INVENTORY_SHEET]["D6"] = "=1+1"
    formula_bytes = BytesIO()
    formula.save(formula_bytes)
    _expect_xlsx_error(
        "formulas_forbidden",
        lambda: parse_inventory_workbook(
            formula_bytes.getvalue(),
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    namespaced_formula = _replace_zip_entry(
        inventory,
        "xl/worksheets/sheet1.xml",
        lambda payload: payload.replace(
            b"<worksheet ",
            (
                b'<worksheet xmlns:x-foo="http://schemas.openxmlformats.org/'
                b'spreadsheetml/2006/main" '
            ),
            1,
        ).replace(b"</c>", b"<x-foo:f>1</x-foo:f></c>", 1),
    )
    _expect_xlsx_error(
        "formulas_forbidden",
        lambda: parse_inventory_workbook(
            namespaced_formula,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    numeric_barcode = load_workbook(BytesIO(inventory))
    numeric_barcode[INVENTORY_SHEET]["B6"] = 101
    numeric_bytes = BytesIO()
    numeric_barcode.save(numeric_bytes)
    _expect_xlsx_error(
        "invalid_inventory_rows",
        lambda: parse_inventory_workbook(
            numeric_bytes.getvalue(),
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    incomplete = load_workbook(BytesIO(inventory))
    incomplete[INVENTORY_SHEET].delete_rows(7, 1)
    incomplete_bytes = BytesIO()
    incomplete.save(incomplete_bytes)
    _expect_xlsx_error(
        "invalid_inventory_rows",
        lambda: parse_inventory_workbook(
            incomplete_bytes.getvalue(),
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    external = _add_zip_entry(inventory, "xl/externalLinks/externalLink1.xml", b"<externalLink/>")
    _expect_xlsx_error(
        "external_links_forbidden",
        lambda: parse_inventory_workbook(
            external,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    macro = _add_zip_entry(inventory, "xl/vbaProject.bin", b"macro")
    _expect_xlsx_error(
        "macros_forbidden",
        lambda: parse_inventory_workbook(
            macro,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    _expect_xlsx_error(
        "zip_uncompressed_limit_exceeded",
        lambda: parse_inventory_workbook(
            inventory,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
            limits=replace(DEFAULT_LIMITS, max_uncompressed_bytes=1),
        ),
    )
    _expect_xlsx_error(
        "malformed_ooxml",
        lambda: parse_inventory_workbook(
            b"not-an-ooxml-archive",
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    huge_integer = load_workbook(BytesIO(inventory))
    huge_integer[INVENTORY_SHEET]["D6"] = "9" * 5000
    huge_integer_bytes = BytesIO()
    huge_integer.save(huge_integer_bytes)
    _expect_xlsx_error(
        "invalid_inventory_rows",
        lambda: parse_inventory_workbook(
            huge_integer_bytes.getvalue(),
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    oversized_rows = _replace_zip_entry(
        inventory,
        "xl/worksheets/sheet1.xml",
        lambda payload: payload.replace(
            b"</sheetData>",
            b"<row/>" * (DEFAULT_LIMITS.max_rows + 1) + b"</sheetData>",
        ),
    )
    _expect_xlsx_error(
        "worksheet_size_limit_exceeded",
        lambda: parse_inventory_workbook(
            oversized_rows,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    utf16_sheet = _replace_zip_entry(
        inventory,
        "xl/worksheets/sheet1.xml",
        lambda payload: (
            '<?xml version="1.0" encoding="UTF-16"?>'
            + payload.decode("utf-8").removeprefix('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
        ).encode("utf-16"),
    )
    _expect_xlsx_error(
        "unsupported_xml_encoding",
        lambda: parse_inventory_workbook(
            utf16_sheet,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    bomless_utf16_sheet = _replace_zip_entry(
        inventory,
        "xl/worksheets/sheet1.xml",
        lambda payload: payload.decode("utf-8")
        .replace('encoding="UTF-8"', 'encoding="UTF-16LE"', 1)
        .encode("utf-16le"),
    )
    _expect_xlsx_error(
        "unsupported_xml_encoding",
        lambda: parse_inventory_workbook(
            bomless_utf16_sheet,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    cells_without_references = _replace_zip_entry(
        inventory,
        "xl/worksheets/sheet1.xml",
        lambda payload: payload.replace(
            b"</sheetData>",
            b"<row>" + (b"<c/>" * (DEFAULT_LIMITS.max_columns + 1)) + b"</row></sheetData>",
        ),
    )
    _expect_xlsx_error(
        "worksheet_dimension_limit_exceeded",
        lambda: parse_inventory_workbook(
            cells_without_references,
            filename="inventory.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            facilities=service.active_facilities(),
            catalog=catalog,
            source_revision="catalog-v1",
        ),
    )
    return catalog, shipment_lines


def _opening(service: FfPoolDocumentService) -> dict[str, object]:
    manifest = {
        "aggregate_rows": [
            {"nm_id": 101, "quantity": 30, "capital_rub": "3000.00"},
            {"nm_id": 102, "quantity": 15, "capital_rub": "2500.00"},
        ],
        "allocations": [
            {"facility_id": "fac_msk", "pool": "FBS", "nm_id": 101, "quantity": 20, "capital_rub": "2000.00"},
            {"facility_id": "fac_msk", "pool": "FBO", "nm_id": 101, "quantity": 10, "capital_rub": "1000.00"},
            {"facility_id": "fac_msk", "pool": "FBS", "nm_id": 102, "quantity": 5, "capital_rub": "500.00"},
            {"facility_id": "fac_orb", "pool": "FBS", "nm_id": 102, "quantity": 10, "capital_rub": "2000.00"},
        ],
    }
    preview = service.accept_preview(
        identity=identity("opening"),
        document_kind="facility_pool_opening",
        manifest=manifest,
    )
    assert preview["state"] == "ready"
    result = service.post(preview["request_id"])
    assert result["state"] == "complete", result
    with sqlite3.connect(service.db_path) as conn:
        conn.row_factory = sqlite3.Row
        assert conn.execute(f"SELECT SUM(quantity) FROM {BALANCES_TABLE}").fetchone()[0] == 45
        assert conn.execute(f"SELECT SUM(CAST(capital_rub AS NUMERIC)) FROM {BALANCES_TABLE}").fetchone()[0] == 5500
    second_preview = service.accept_preview(
        identity=identity("second-opening"),
        document_kind="facility_pool_opening",
        manifest={
            "aggregate_rows": [{"nm_id": 101, "quantity": 1, "capital_rub": "100.00"}],
            "allocations": [
                {
                    "facility_id": "fac_single",
                    "pool": "FBS",
                    "nm_id": 101,
                    "quantity": 1,
                    "capital_rub": "100.00",
                }
            ],
        },
    )
    second = service.post(str(second_preview["request_id"]))
    assert second["state"] == "blocked"
    assert second["error"]["code"] == "opening_requires_empty_detail"
    return result


def _idempotent_repeat_and_recovery(
    service: FfPoolDocumentService,
    opening: dict[str, object],
) -> None:
    before = _table_count(service.db_path, LINES_TABLE)
    repeat_identity = replace(identity("opening"), request_id="fixture:opening:response-loss-retry")
    repeated = service.accept_preview(
        identity=repeat_identity,
        document_kind="facility_pool_opening",
        manifest=opening["preview_manifest"],
    )
    assert repeated["state"] == "complete"
    assert repeated["idempotent"] is True
    assert _table_count(service.db_path, LINES_TABLE) == before

    root_preview = service.accept_preview(
        identity=identity("restart-root"),
        document_kind="transfer_root",
        manifest={
            "source": {"facility_id": "fac_msk", "pool": "FBS"},
            "destination": {"facility_id": "fac_orb", "pool": "FBO"},
        },
    )
    canonical = str(root_preview["request_id"])
    with sqlite3.connect(service.db_path) as conn:
        conn.execute(f"UPDATE {REQUESTS_TABLE} SET state='processing' WHERE request_id=?", (canonical,))
        conn.commit()
    restarted = FfPoolDocumentService(
        db_path=service.db_path,
        runtime_dir=service.runtime_dir,
        timestamp_factory=service.timestamp_factory,
    )
    assert restarted.status(request_id=canonical)["state"] == "accepted"
    assert restarted.process_request(canonical)["state"] == "ready"
    posted = restarted.post(canonical)
    assert posted["state"] == "complete"
    recovery_id = str(posted["recovery_operation_id"])
    with sqlite3.connect(service.db_path) as conn:
        conn.execute(
            f"UPDATE {REQUESTS_TABLE} SET state='replay',completed_at='' WHERE request_id=?",
            (canonical,),
        )
        conn.execute(
            "UPDATE sheet_vitrina_v1_recovery_operations "
            "SET lifecycle_state='mutation_running',next_action='reconcile_or_complete_business_mutation' "
            "WHERE operation_id=?",
            (recovery_id,),
        )
        conn.commit()
    resumed = FfPoolDocumentService(
        db_path=service.db_path,
        runtime_dir=service.runtime_dir,
        timestamp_factory=service.timestamp_factory,
    )
    assert resumed.status(request_id=canonical)["state"] == "complete"
    assert WarehouseRecoveryRegistry(
        runtime_dir=service.runtime_dir,
        db_path=service.db_path,
    ).get_operation(recovery_id)["lifecycle"] == "retained"
    operation_count = _table_count(service.db_path, OPERATIONS_TABLE)
    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(lambda _item: restarted.post(canonical), range(2)))
    assert all(item["state"] == "complete" for item in results)
    assert _table_count(service.db_path, OPERATIONS_TABLE) == operation_count

    concurrent = restarted.accept_preview(
        identity=identity("concurrent-root"),
        document_kind="transfer_root",
        manifest={
            "source": {"facility_id": "fac_msk", "pool": "FBS"},
            "destination": {"facility_id": "fac_orb", "pool": "FBO"},
        },
    )
    concurrent_request = str(concurrent["request_id"])
    before_concurrent = _table_count(service.db_path, OPERATIONS_TABLE)
    with ThreadPoolExecutor(max_workers=2) as pool:
        concurrent_results = list(
            pool.map(lambda _item: restarted.post(concurrent_request), range(2))
        )
    assert all(item["state"] == "complete" for item in concurrent_results), concurrent_results
    assert _table_count(service.db_path, OPERATIONS_TABLE) == before_concurrent + 1

    conflicting_source = identity("source-revision-conflict")
    first_source = restarted.accept_preview(
        identity=conflicting_source,
        document_kind="transfer_root",
        manifest={
            "source": {"facility_id": "fac_msk", "pool": "FBS"},
            "destination": {"facility_id": "fac_orb", "pool": "FBO"},
        },
    )
    assert first_source["state"] == "ready"
    try:
        restarted.accept_preview(
            identity=replace(
                conflicting_source,
                request_id="fixture:source-revision-conflict-second:request",
            ),
            document_kind="transfer_root",
            manifest={
                "source": {"facility_id": "fac_msk", "pool": "FBO"},
                "destination": {"facility_id": "fac_orb", "pool": "FBS"},
            },
        )
    except FfPoolDocumentError as exc:
        assert exc.code == "source_revision_identity_conflict"
    else:
        raise AssertionError("one immutable source revision must bind one semantic document")
    try:
        restarted.accept_preview(
            identity=replace(
                conflicting_source,
                request_id="fixture:source-revision-cross-kind:request",
            ),
            document_kind="pool_inventory",
            manifest={
                "facility_id": "fac_msk",
                "scope": "FBS",
                "targets": [{"pool": "FBS", "nm_id": 101, "quantity": 1}],
            },
        )
    except FfPoolDocumentError as exc:
        assert exc.code == "source_revision_identity_conflict"
    else:
        raise AssertionError("source revision identity must be unique across document kinds")

    race_identity = identity("client-request-race")
    before_race = _table_count(service.db_path, REQUESTS_TABLE)

    def accept_race(destination_pool: str) -> tuple[str, str]:
        try:
            result = restarted.accept_preview(
                identity=race_identity,
                document_kind="transfer_root",
                manifest={
                    "source": {"facility_id": "fac_msk", "pool": "FBS"},
                    "destination": {"facility_id": "fac_orb", "pool": destination_pool},
                },
            )
            return "ok", str(result["request_id"])
        except FfPoolDocumentError as exc:
            return "error", exc.code

    with ThreadPoolExecutor(max_workers=2) as pool:
        race_results = list(pool.map(accept_race, ("FBO", "FBS")))
    assert sorted(item[0] for item in race_results) == ["error", "ok"], race_results
    assert {item[1] for item in race_results if item[0] == "error"} == {
        "request_id_identity_conflict"
    }
    assert _table_count(service.db_path, REQUESTS_TABLE) == before_race + 1


def _inventory_missing_fbs_zero(
    service: FfPoolDocumentService,
    catalog: list[dict[str, object]],
) -> None:
    """Routine full-scope inventory makes an absent FBS zero explicit."""

    target_key = ("fac_orb", "FBS", 101)
    with sqlite3.connect(service.db_path) as conn:
        assert conn.execute(
            f"SELECT 1 FROM {BALANCES_TABLE} WHERE facility_id=? AND pool=? AND nm_id=?",
            target_key,
        ).fetchone() is None
        non_target_before = conn.execute(
            f"SELECT facility_id,pool,nm_id,projection_epoch,quantity,capital_rub,wac_rub,"
            f"source_watermark,updated_at FROM {BALANCES_TABLE} "
            "WHERE NOT (facility_id=? AND pool=? AND nm_id=?) "
            "ORDER BY facility_id,pool,nm_id",
            target_key,
        ).fetchall()
        movement_count_before = int(
            conn.execute(f"SELECT COUNT(*) FROM {LINES_TABLE}").fetchone()[0]
        )

    targets = {
        (101, "FBS"): 0,
        (102, "FBS"): _balance(service.db_path, "fac_orb", "FBS", 102)[0],
    }
    workbook_bytes = service.generate_inventory_template(
        facility_id="fac_orb",
        scope="FBS",
        catalog=catalog,
        source_revision="inventory-explicit-zero-v1",
        targets=targets,
    )
    inventory_identity = replace(
        identity("inventory-explicit-zero", revision="inventory-explicit-zero-v1"),
        source_system="operator_http",
        source_type="pool_inventory_workbook",
    )
    preview = service.preview_inventory_workbook(
        identity=inventory_identity,
        source_bytes=workbook_bytes,
        source_filename="inventory-explicit-zero.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        catalog=catalog,
    )
    assert preview["state"] == "ready", preview
    result = service.post(str(preview["request_id"]))
    assert result["state"] == "complete", result
    root_document_id = str(result["document"]["document_id"])

    with sqlite3.connect(service.db_path) as conn:
        conn.row_factory = sqlite3.Row
        target = conn.execute(
            f"SELECT quantity,capital_rub,wac_rub,source_watermark FROM {BALANCES_TABLE} "
            "WHERE facility_id=? AND pool=? AND nm_id=?",
            target_key,
        ).fetchone()
        assert target is not None
        assert dict(target) == {
            "quantity": 0,
            "capital_rub": "0",
            "wac_rub": None,
            "source_watermark": root_document_id,
        }
        assert conn.execute(
            f"SELECT 1 FROM {BALANCES_TABLE} "
            "WHERE facility_id='fac_orb' AND pool='FBO' AND nm_id=101"
        ).fetchone() is None
        non_target_after = conn.execute(
            f"SELECT facility_id,pool,nm_id,projection_epoch,quantity,capital_rub,wac_rub,"
            f"source_watermark,updated_at FROM {BALANCES_TABLE} "
            "WHERE NOT (facility_id=? AND pool=? AND nm_id=?) "
            "ORDER BY facility_id,pool,nm_id",
            target_key,
        ).fetchall()
        assert [tuple(row) for row in non_target_after] == [
            tuple(row) for row in non_target_before
        ]
        assert (
            int(conn.execute(f"SELECT COUNT(*) FROM {LINES_TABLE}").fetchone()[0])
            == movement_count_before
        )
        evidence = conn.execute(
            f"SELECT quantity,capital_rub,metadata_json FROM {DOCUMENT_LINES_TABLE} "
            "WHERE document_id=? AND line_role='absolute_target' "
            "AND facility_id=? AND pool=? AND nm_id=?",
            (root_document_id, *target_key),
        ).fetchone()
        assert evidence is not None
        assert int(evidence["quantity"]) == 0
        assert Decimal(str(evidence["capital_rub"])) == 0
        assert json.loads(str(evidence["metadata_json"])) == {
            "before_quantity": 0,
            "explicit_physical_zero": True,
            "selected_pool": True,
        }
        posted_manifest = json.loads(
            str(
                conn.execute(
                    f"SELECT posted_manifest_json FROM {DOCUMENTS_TABLE} WHERE document_id=?",
                    (root_document_id,),
                ).fetchone()[0]
            )
        )
        assert posted_manifest["domain"]["explicit_zero_fbs_balance_count"] == 1
        assert posted_manifest["domain"]["explicit_zero_fbs_nm_ids"] == [101]

    repeated = service.preview_inventory_workbook(
        identity=replace(
            inventory_identity,
            request_id="fixture:inventory-explicit-zero:response-loss-retry",
        ),
        source_bytes=workbook_bytes,
        source_filename="renamed-inventory-explicit-zero.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        catalog=catalog,
    )
    assert repeated["request_id"] == result["request_id"]
    assert repeated["state"] == "complete" and repeated["idempotent"] is True
    with sqlite3.connect(service.db_path) as conn:
        assert conn.execute(
            f"SELECT COUNT(*) FROM {BALANCES_TABLE} "
            "WHERE facility_id=? AND pool=? AND nm_id=?",
            target_key,
        ).fetchone()[0] == 1
        assert (
            int(conn.execute(f"SELECT COUNT(*) FROM {LINES_TABLE}").fetchone()[0])
            == movement_count_before
        )


def _transfer_and_late_expense(service: FfPoolDocumentService) -> None:
    root = _accept_post(
        service,
        "transfer-main-root",
        "transfer_root",
        {
            "source": {"facility_id": "fac_msk", "pool": "FBS"},
            "destination": {"facility_id": "fac_orb", "pool": "FBO"},
        },
    )
    root_id = str(root["document"]["document_id"])
    shipment = _accept_post(
        service,
        "transfer-main-shipment",
        "transfer_shipment",
        {
            "root_document_id": root_id,
            "items": [{"nm_id": 101, "quantity": 10}],
            "expenses": [
                {"amount_rub": "60.00", "basis": "Доставка"},
                {"amount_rub": "40.00", "basis": "Погрузка"},
            ],
        },
    )
    assert shipment["state"] == "complete"
    receipt = _accept_post(
        service,
        "transfer-main-receipt",
        "transfer_receipt",
        {"root_document_id": root_id, "items": [{"nm_id": 101, "quantity": 6}]},
    )
    loss = _accept_post(
        service,
        "transfer-main-loss",
        "transfer_loss",
        {"root_document_id": root_id, "items": [{"nm_id": 101, "quantity": 2}]},
    )
    projection = service.open_transfer_projection(root_id)
    assert projection["state"] == "open"
    assert projection["lines"][0]["open_quantity"] == 2
    _accept_post(
        service,
        "transfer-main-loss-storno",
        "storno",
        {"target_document_id": str(loss["document"]["document_id"])},
    )
    projection = service.open_transfer_projection(root_id)
    assert projection["lines"][0]["open_quantity"] == 4
    _accept_post(
        service,
        "transfer-main-loss-replacement",
        "transfer_loss",
        {"root_document_id": root_id, "items": [{"nm_id": 101, "quantity": 2}]},
    )
    _accept_post(
        service,
        "transfer-main-cancel",
        "transfer_cancellation",
        {"root_document_id": root_id},
    )
    projection = service.open_transfer_projection(root_id)
    assert projection["state"] == "closed"
    assert projection["quantity_conserved"]
    before_destination = _balance(service.db_path, "fac_orb", "FBO", 101)
    late = _accept_post(
        service,
        "transfer-main-late-expense",
        "late_expense",
        {
            "root_document_id": root_id,
            "expenses": [{"amount_rub": "40.00", "basis": "Поздний счёт", "source_file_sha256": "sha256:" + "a" * 64}],
        },
    )
    assert late["state"] == "complete"
    after_destination = _balance(service.db_path, "fac_orb", "FBO", 101)
    assert after_destination[0] == before_destination[0]
    assert after_destination[1] - before_destination[1] == 2400
    receipt_id = str(receipt["document"]["document_id"])
    blocked_receipt_storno_preview = service.accept_preview(
        identity=identity("transfer-main-receipt-storno-blocked"),
        document_kind="storno",
        manifest={"target_document_id": receipt_id},
    )
    blocked_receipt_storno = service.post(
        str(blocked_receipt_storno_preview["request_id"])
    )
    assert blocked_receipt_storno["state"] == "blocked"
    assert blocked_receipt_storno["error"]["code"] == "receipt_storno_has_active_late_expense"
    transfer_correction_preview = service.accept_preview(
        identity=identity("transfer-main-correction-blocked"),
        document_kind="correction",
        manifest={
            "target_document_id": root_id,
            "movements": [
                {
                    "facility_id": "fac_msk",
                    "pool": "FBS",
                    "nm_id": 101,
                    "quantity_delta": 1,
                    "capital_delta_rub": "100.00",
                }
            ],
        },
    )
    transfer_correction = service.post(str(transfer_correction_preview["request_id"]))
    assert transfer_correction["state"] == "blocked"
    assert transfer_correction["error"]["code"] == "transfer_correction_requires_typed_outcome"
    _accept_post(
        service,
        "transfer-main-late-expense-storno",
        "storno",
        {"target_document_id": str(late["document"]["document_id"])},
    )
    receipt_storno = _accept_post(
        service,
        "transfer-main-receipt-storno",
        "storno",
        {"target_document_id": receipt_id},
    )
    assert receipt_storno["state"] == "complete"
    descendant_correction_preview = service.accept_preview(
        identity=identity("transfer-main-descendant-correction-blocked"),
        document_kind="correction",
        manifest={
            "target_document_id": str(receipt_storno["document"]["document_id"]),
            "movements": [
                {
                    "facility_id": "fac_msk",
                    "pool": "FBS",
                    "nm_id": 101,
                    "quantity_delta": 1,
                    "capital_delta_rub": "100.00",
                }
            ],
        },
    )
    descendant_correction = service.post(
        str(descendant_correction_preview["request_id"])
    )
    assert descendant_correction["state"] == "blocked"
    assert descendant_correction["error"]["code"] == "transfer_correction_requires_typed_outcome"


def _mis_sort(service: FfPoolDocumentService) -> None:
    root = _accept_post(
        service,
        "mis-sort-root",
        "transfer_root",
        {
            "source": {"facility_id": "fac_msk", "pool": "FBS"},
            "destination": {"facility_id": "fac_orb", "pool": "FBO"},
        },
    )
    root_id = str(root["document"]["document_id"])
    _accept_post(
        service,
        "mis-sort-shipment",
        "transfer_shipment",
        {
            "root_document_id": root_id,
            "items": [{"nm_id": 101, "quantity": 4}],
            "expenses": [{"amount_rub": "4.00", "basis": "Доставка пересорта"}],
        },
    )
    receipt = _accept_post(
        service,
        "mis-sort-correct-receipt",
        "transfer_receipt",
        {"root_document_id": root_id, "items": [{"nm_id": 101, "quantity": 2}]},
    )
    receipt_id = str(receipt["document"]["document_id"])
    source_102_before = _balance(service.db_path, "fac_msk", "FBS", 102)
    destination_102_before = _balance(service.db_path, "fac_orb", "FBO", 102)
    success = _accept_post(
        service,
        "mis-sort-success",
        "transfer_discrepancy",
        {
            "root_document_id": root_id,
            "expected_not_sent": [{"nm_id": 101, "quantity": 1}],
            "unexpected": [{"nm_id": 102, "quantity": 2}],
        },
    )
    assert success["state"] == "complete"
    source_102_after = _balance(service.db_path, "fac_msk", "FBS", 102)
    destination_102_after = _balance(service.db_path, "fac_orb", "FBO", 102)
    moved_source_capital = source_102_before[1] - source_102_after[1]
    assert destination_102_after[1] - destination_102_before[1] == moved_source_capital + 100
    before_ops = _table_count(service.db_path, OPERATIONS_TABLE)
    blocked_preview = service.accept_preview(
        identity=identity("mis-sort-blocked"),
        document_kind="transfer_discrepancy",
        manifest={
            "root_document_id": root_id,
            "unexpected": [{"nm_id": 102, "quantity": 999}],
        },
    )
    blocked = service.post(str(blocked_preview["request_id"]))
    assert blocked["state"] == "blocked"
    assert blocked["error"]["code"] == "mis_sort_insufficient_source"
    assert _table_count(service.db_path, OPERATIONS_TABLE) == before_ops
    assert _table_count_where(service.db_path, DOCUMENTS_TABLE, "document_id=?", (receipt_id,)) == 1


def _reallocation_inventory_overhead(
    service: FfPoolDocumentService,
    catalog: list[dict[str, object]],
) -> None:
    _accept_post(
        service,
        "single-pool-acceptance",
        "china_acceptance",
        {
            "facility_id": "fac_single",
            "allocations": [
                {
                    "nm_id": 101,
                    "quantity_fbs": 2,
                    "quantity_fbo": 0,
                    "accepted_quantity": 2,
                    "accepted_capital_rub": "200.00",
                    "identity_evidence_digest": "sha256:" + "1" * 64,
                }
            ],
        },
    )
    single_pool_overhead = _accept_post(
        service,
        "single-pool-overhead-both",
        "pool_overhead",
        {
            "facility_id": "fac_single",
            "scope": "both",
            "amount_rub": "5.00",
            "reason": "Расход при пустом FBO",
        },
    )
    assert single_pool_overhead["preview_manifest"]["scope"] == "both"
    assert _balance(service.db_path, "fac_single", "FBO", 101) == (0, 0)

    physical_before = _facility_total(service.db_path, "fac_msk")
    _accept_post(
        service,
        "pool-reallocation",
        "pool_reallocation",
        {
            "facility_id": "fac_msk",
            "source_pool": "FBS",
            "destination_pool": "FBO",
            "items": [{"nm_id": 101, "quantity": 2}],
            "expenses": [{"amount_rub": "10.00", "basis": "Внутренняя обработка"}],
        },
    )
    assert _facility_total(service.db_path, "fac_msk") == physical_before

    current = {
        (nm_id, pool): _balance(service.db_path, "fac_msk", pool, nm_id)[0]
        for nm_id in (101, 102)
        for pool in ("FBO", "FBS")
    }
    targets = dict(current)
    targets[(101, "FBS")] += 1
    targets[(102, "FBS")] -= 1
    workbook_bytes = service.generate_inventory_template(
        facility_id="fac_msk",
        scope="FBS",
        catalog=catalog,
        source_revision="inventory-fixture-v1",
        targets=targets,
    )
    fbo_before = {
        nm_id: _balance(service.db_path, "fac_msk", "FBO", nm_id)
        for nm_id in (101, 102)
    }
    preview = service.preview_inventory_workbook(
        identity=identity("inventory", revision="inventory-fixture-v1"),
        source_bytes=workbook_bytes,
        source_filename="inventory.xlsx",
        source_content_type=XLSX_CONTENT_TYPE,
        catalog=catalog,
    )
    assert preview["state"] == "ready", preview
    inventory = service.post(str(preview["request_id"]))
    assert inventory["state"] == "complete", inventory
    assert _balance(service.db_path, "fac_msk", "FBS", 101)[0] == targets[(101, "FBS")]
    assert _balance(service.db_path, "fac_msk", "FBS", 102)[0] == targets[(102, "FBS")]
    assert {
        nm_id: _balance(service.db_path, "fac_msk", "FBO", nm_id)
        for nm_id in (101, 102)
    } == fbo_before
    inventory_root = str(inventory["document"]["document_id"])
    with sqlite3.connect(service.db_path) as conn:
        roles = {
            row[0]
            for row in conn.execute(
                f"SELECT relation_type FROM {DOCUMENT_RELATIONS_TABLE} WHERE root_document_id=?",
                (inventory_root,),
            ).fetchall()
        }
    assert roles == {"inventory_surplus_of", "inventory_shortage_of"}
    inventory_storno_preview = service.accept_preview(
        identity=identity("inventory-root-storno-blocked"),
        document_kind="storno",
        manifest={"target_document_id": inventory_root},
    )
    inventory_storno = service.post(str(inventory_storno_preview["request_id"]))
    assert inventory_storno["state"] == "blocked"
    assert inventory_storno["error"]["code"] == "storno_target_has_no_direct_effect"

    quantities_before = _all_quantities(service.db_path, "fac_msk")
    overhead = _accept_post(
        service,
        "overhead-both",
        "pool_overhead",
        {"facility_id": "fac_msk", "scope": "both", "amount_rub": "1.01", "reason": "Общая уборка"},
    )
    assert _all_quantities(service.db_path, "fac_msk") == quantities_before
    overhead_id = str(overhead["document"]["document_id"])
    with sqlite3.connect(service.db_path) as conn:
        allocation = conn.execute(
            f"SELECT SUM(CAST(expense_rub AS NUMERIC)) FROM {DOCUMENT_LINES_TABLE} WHERE document_id=?",
            (overhead_id,),
        ).fetchone()[0]
    assert round(float(allocation), 2) == 1.01
    storno = _accept_post(
        service,
        "overhead-storno",
        "storno",
        {"target_document_id": overhead_id},
    )
    assert storno["state"] == "complete"
    assert _all_quantities(service.db_path, "fac_msk") == quantities_before


def _relations_and_indexes(db_path: Path) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        relation = conn.execute(
            f"SELECT * FROM {DOCUMENT_RELATIONS_TABLE} ORDER BY created_at LIMIT 1"
        ).fetchone()
        assert relation is not None
        try:
            conn.execute(
                f"INSERT INTO {DOCUMENT_RELATIONS_TABLE}(parent_document_id,child_document_id,root_document_id,relation_type,created_at) "
                "VALUES(?,?,?,?,?)",
                (
                    str(relation["child_document_id"]),
                    str(relation["parent_document_id"]),
                    str(relation["root_document_id"]),
                    "storno_of",
                    "2026-08-11T12:00:00Z",
                ),
            )
        except sqlite3.IntegrityError:
            pass
        else:
            raise AssertionError("cycle/invalid child relation must fail closed")
        balance_plan = " ".join(
            str(row[3])
            for row in conn.execute(
                f"EXPLAIN QUERY PLAN SELECT * FROM {BALANCES_TABLE} WHERE facility_id=? AND pool=? AND nm_id=?",
                ("fac_msk", "FBS", 101),
            ).fetchall()
        )
        assert "INDEX" in balance_plan.upper()
        line_plan = " ".join(
            str(row[3])
            for row in conn.execute(
                f"EXPLAIN QUERY PLAN SELECT * FROM {DOCUMENT_LINES_TABLE} WHERE root_document_id=? AND nm_id=?",
                (str(relation["root_document_id"]), 101),
            ).fetchall()
        )
        assert "ff_pool_document_lines_by_root_nm" in line_plan


def _property_checks() -> None:
    for total in range(0, 250):
        for weights in (
            [("a", 1)],
            [("a", 1), ("b", 1)],
            [("a", 1), ("b", 2), ("c", 7)],
            [("z", 11), ("a", 3), ("m", 5)],
        ):
            allocated = _allocate_cents(total, weights)
            assert sum(allocated.values()) == total
            assert all(value >= 0 for value in allocated.values())
            assert allocated == _allocate_cents(total, list(reversed(weights)))
    partitions = ((1, 2, 7), (3, 3, 4), (6, 2, 2), (10,))
    for total_cents in range(0, 1000, 7):
        for parts in partitions:
            prior = 0
            shares = []
            for quantity in parts:
                shares.append(_component_share(total_cents, 10, prior, quantity))
                prior += quantity
            assert prior == 10
            assert sum(shares) == total_cents


def _assert_recovery(service: FfPoolDocumentService) -> None:
    assert FACILITIES_TABLE in DOMAIN_EXACT_TABLES
    assert FACILITY_CHANGES_TABLE in DOMAIN_EXACT_TABLES
    registry = WarehouseRecoveryRegistry(runtime_dir=service.runtime_dir, db_path=service.db_path)
    rows = registry.list_operations(limit=500)
    pool_rows = [item for item in rows if item["operation_kind"] == "ff_pool_document_posting"]
    assert pool_rows
    assert all(item["tier"] == "T1" for item in pool_rows)
    assert all(item["actual_bytes"] > 0 for item in pool_rows)
    assert all(
        all(artifact.get("artifact_kind") == "undo" for artifact in item.get("artifacts", []))
        for item in pool_rows
    )
    retained = next(item for item in pool_rows if item["lifecycle"] == "retained")
    operation_count = _table_count(service.db_path, OPERATIONS_TABLE)
    line_count = _table_count(service.db_path, LINES_TABLE)
    try:
        registry.rollback_t1(str(retained["operation_id"]), reason="fixture append-only guard")
    except sqlite3.IntegrityError as exc:
        assert "immutable" in str(exc) or "append-only" in str(exc)
    else:
        raise AssertionError("generic T1 rollback must not partially delete append-only documents")
    assert _table_count(service.db_path, OPERATIONS_TABLE) == operation_count
    assert _table_count(service.db_path, LINES_TABLE) == line_count
    assert registry.get_operation(str(retained["operation_id"]))["lifecycle"] == "retained"


def _assert_no_aggregate_or_legacy_writes(db_path: Path) -> None:
    with sqlite3.connect(db_path) as conn:
        tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        assert "sheet_vitrina_v1_ff_stock_operations" not in tables
        assert "sheet_vitrina_v1_warehouse_functional_balances" not in tables
        assert conn.execute(f"SELECT COUNT(*) FROM {FEATURE_EPOCHS_TABLE}").fetchone()[0] == 1
        assert conn.execute(f"SELECT reader_enabled FROM {FEATURE_EPOCHS_TABLE} WHERE epoch=1").fetchone()[0] == 0


def _accept_post(
    service: FfPoolDocumentService,
    name: str,
    kind: str,
    manifest: dict[str, object],
) -> dict[str, object]:
    preview = service.accept_preview(
        identity=identity(name),
        document_kind=kind,
        manifest=manifest,
    )
    assert preview["state"] == "ready", preview
    result = service.post(str(preview["request_id"]))
    assert result["state"] == "complete", result
    return result


def _expect_xlsx_error(code: str, action: object) -> None:
    try:
        action()  # type: ignore[operator]
    except FfPoolXlsxError as exc:
        assert exc.code == code, (exc.code, code, exc.details)
    else:
        raise AssertionError(f"expected XLSX error: {code}")


def _add_zip_entry(source: bytes, name: str, payload: bytes) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source), "r") as original, zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        for item in original.infolist():
            target.writestr(item, original.read(item.filename))
        target.writestr(name, payload)
    return output.getvalue()


def _replace_zip_entry(source: bytes, name: str, transform: object) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source), "r") as original, zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_STORED
    ) as target:
        found = False
        for item in original.infolist():
            payload = original.read(item.filename)
            if item.filename == name:
                payload = transform(payload)  # type: ignore[operator]
                found = True
            target.writestr(item.filename, payload)
    assert found, name
    return output.getvalue()


def _table_count(db_path: Path, table: str) -> int:
    with sqlite3.connect(db_path) as conn:
        return int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])


def _table_count_where(db_path: Path, table: str, where: str, params: tuple[object, ...]) -> int:
    with sqlite3.connect(db_path) as conn:
        return int(conn.execute(f"SELECT COUNT(*) FROM {table} WHERE {where}", params).fetchone()[0])


def _balance(db_path: Path, facility_id: str, pool: str, nm_id: int) -> tuple[int, int]:
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            f"SELECT quantity,capital_rub FROM {BALANCES_TABLE} WHERE facility_id=? AND pool=? AND nm_id=?",
            (facility_id, pool, nm_id),
        ).fetchone()
    if row is None:
        return 0, 0
    return int(row[0]), int(round(float(row[1]) * 100))


def _facility_total(db_path: Path, facility_id: str) -> int:
    with sqlite3.connect(db_path) as conn:
        return int(
            conn.execute(
                f"SELECT COALESCE(SUM(quantity),0) FROM {BALANCES_TABLE} WHERE facility_id=?",
                (facility_id,),
            ).fetchone()[0]
        )


def _all_quantities(db_path: Path, facility_id: str) -> list[tuple[str, int, int]]:
    with sqlite3.connect(db_path) as conn:
        return [
            (str(row[0]), int(row[1]), int(row[2]))
            for row in conn.execute(
                f"SELECT pool,nm_id,quantity FROM {BALANCES_TABLE} WHERE facility_id=? ORDER BY pool,nm_id",
                (facility_id,),
            ).fetchall()
        ]


if __name__ == "__main__":
    main()
