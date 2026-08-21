"""Focused Stage 3 checks for protected facility/pool read and mutation models."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path
from io import BytesIO
import json
import sqlite3
import sys
from tempfile import TemporaryDirectory


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from packages.application.ff_pool_documents import (  # noqa: E402
    DOCUMENTS_TABLE,
    FfPoolDocumentService,
    REQUESTS_TABLE,
)
from packages.application.ff_pool_documents_xlsx import (  # noqa: E402
    CHINA_SHEET,
    XLSX_CONTENT_TYPE,
    generate_china_acceptance_workbook,
)
from packages.application.ff_pool_foundation import (  # noqa: E402
    BALANCES_TABLE,
    FACILITIES_TABLE,
    FACILITY_CHANGES_TABLE,
    FACILITY_PROFILES_TABLE,
    FEATURE_EPOCHS_TABLE,
    evaluate_ff_pool_aggregate_parity,
    record_ff_pool_parity_diagnostic,
)
from packages.application.ff_pool_surfaces import (  # noqa: E402
    FfPoolSurface,
    FfPoolSurfaceError,
    _resolve_supplier_lines_with_canonical_nomenclature,
)
from packages.contracts.ff_pool_documents import DocumentIdentity  # noqa: E402
from apps.russian_payment_orders_smoke import _fixture, _render_pdf  # noqa: E402


class Clock:
    def __init__(self) -> None:
        self.current = datetime(2026, 8, 12, 6, 0, tzinfo=timezone.utc)

    def __call__(self) -> str:
        value = self.current.isoformat(timespec="seconds").replace("+00:00", "Z")
        self.current += timedelta(seconds=1)
        return value


def main() -> None:
    _schema_absence_is_controlled()
    _guided_preview_is_default_off()
    _guided_source_uses_canonical_nomenclature()
    _overhead_operator_workflows()
    with TemporaryDirectory(prefix="ff-pool-surfaces-") as directory:
        root = Path(directory)
        clock = Clock()
        service = FfPoolDocumentService(
            db_path=root / "state.sqlite3",
            runtime_dir=root,
            timestamp_factory=clock,
            resume=False,
        )
        surface = FfPoolSurface(
            db_path=service.db_path,
            runtime_dir=root,
            timestamp_factory=clock,
        )
        _default_off_reads_do_not_write(surface, service.db_path)
        facilities = _facility_management(surface, service.db_path, clock)
        request_id = _document_workflow(surface, facilities)
        _exact_reader_and_pagination(surface, service.db_path, clock, facilities)
        _deactivation_with_dependencies_is_blocked(surface, facilities)
        _read_models(surface, request_id, facilities)
    print("ff_pool_surfaces_smoke: OK")


def _overhead_operator_workflows() -> None:
    """Manual and synthetic-PDF overheads share one durable pool workflow."""

    with TemporaryDirectory(prefix="ff-pool-overhead-surfaces-") as directory:
        root = Path(directory)
        clock = Clock()
        clock.current = datetime(2026, 8, 12, 21, 30, tzinfo=timezone.utc)
        service = FfPoolDocumentService(
            db_path=root / "state.sqlite3",
            runtime_dir=root,
            timestamp_factory=clock,
            resume=False,
        )
        surface = FfPoolSurface(
            db_path=service.db_path,
            runtime_dir=root,
            timestamp_factory=clock,
        )
        with sqlite3.connect(service.db_path) as conn:
            conn.execute(
                f"INSERT INTO {FEATURE_EPOCHS_TABLE}(epoch,writer_enabled,reader_enabled,source_revision,created_at,metadata_json) "
                "VALUES(1,1,0,'overhead-surface-v1',?,'{}')",
                (clock(),),
            )
            for facility_id, code in (("fac_overhead", "OH"), ("fac_control", "CTL")):
                now = clock()
                conn.execute(
                    f"INSERT INTO {FACILITIES_TABLE}(facility_id,code,name,active,display_timezone,created_at,updated_at) "
                    "VALUES(?,?,?,?,?,?,?)",
                    (facility_id, code, f"Синтетический FF {code}", 1, "Asia/Yekaterinburg", now, now),
                )
                conn.execute(
                    f"INSERT INTO {FACILITY_PROFILES_TABLE}(facility_id,city,future_fields_json,created_at,updated_at) "
                    "VALUES(?,?,'{}',?,?)",
                    (facility_id, "Тестоград", now, now),
                )
            conn.executemany(
                f"INSERT INTO {BALANCES_TABLE}(facility_id,pool,nm_id,projection_epoch,quantity,capital_rub,wac_rub,source_watermark,updated_at) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                [
                    ("fac_overhead", "FBS", 101, 1, 2, "20.00", "10.00", "fixture", clock()),
                    ("fac_overhead", "FBO", 102, 1, 3, "30.00", "10.00", "fixture", clock()),
                    ("fac_control", "FBS", 101, 1, 4, "40.00", "10.00", "fixture", clock()),
                ],
            )
            conn.commit()

        facilities = surface.facilities_page()["facilities"]
        overhead_facility = next(item for item in facilities if item["facility_id"] == "fac_overhead")
        assert overhead_facility["current_business_date"] == "2026-08-13"
        base_payload = {
            "facility_id": "fac_overhead",
            "scope": "both",
            "category": "storage",
            "comment": "",
        }
        for missing_field, expected_code in (
            ("facility_id", "invalid_facility_id"),
            ("scope", "invalid_pool_scope"),
            ("category", "invalid_overhead_category"),
        ):
            payload = {**base_payload, "request_id": f"overhead:missing:{missing_field}", "amount_rub": "1.00"}
            payload[missing_field] = ""
            try:
                surface.accept_pool_overhead_preview(payload, actor="fixture-operator")
            except FfPoolSurfaceError as exc:
                assert exc.code == expected_code, (missing_field, exc.code)
            else:
                raise AssertionError(f"{missing_field} must be selected explicitly")
        try:
            surface.accept_pool_overhead_preview(
                {
                    **base_payload,
                    "request_id": "overhead:other:missing-comment",
                    "category": "other",
                    "amount_rub": "1.00",
                },
                actor="fixture-operator",
            )
        except FfPoolSurfaceError as exc:
            assert exc.code == "overhead_other_comment_required"
        else:
            raise AssertionError("other overhead category must require a comment")
        try:
            surface.accept_pool_overhead_preview(
                {
                    **base_payload,
                    "request_id": "overhead:backdate",
                    "business_date": "2026-08-12",
                    "amount_rub": "1.00",
                },
                actor="fixture-operator",
            )
        except FfPoolSurfaceError as exc:
            assert exc.code == "overhead_business_date_mismatch"
            assert exc.details["business_date"] == "2026-08-13"
        else:
            raise AssertionError("pool overhead must not accept a browser-selected date")

        quantities_before = _surface_quantities(service.db_path)
        control_before = _surface_balance(service.db_path, "fac_control", "FBS", 101)
        manual = surface.accept_pool_overhead_preview(
            {
                **base_payload,
                "request_id": "overhead:manual",
                "amount_rub": "0.01",
            },
            actor="fixture-operator",
        )
        manual_summary = manual["preview"]["summary"]
        assert manual["state"] == "ready" and manual["confirm_allowed"]
        assert manual_summary["business_date"] == "2026-08-13"
        assert manual_summary["source_mode"] == "manual"
        assert manual_summary["denominator_quantity"] == 5, manual_summary
        manual_complete = surface.confirm_document(str(manual["request_id"]))
        assert manual_complete["state"] == "posted"
        assert manual_complete["publication"]["status"] == "queued"
        assert _surface_quantities(service.db_path) == quantities_before
        assert _surface_balance(service.db_path, "fac_control", "FBS", 101) == control_before
        manual_document_id = str(manual_complete["document"]["document_id"])
        manual_detail = surface.document_detail(manual_document_id)["documents"][0]
        assert manual_detail["overhead"]["source_mode"] == "manual"
        assert manual_detail["overhead"]["category"] == "storage"
        assert manual_detail["actor"] == "fixture-operator"
        assert manual_detail["source_type"] == "ff_pool_overhead_manual"
        assert not manual_detail["source_file_available"]

        wb_text = _fixture("wb_bank_0401060.txt")
        wb_equivalent_text = _fixture("wb_bank_0401060_equivalent_layout.txt")
        vtb_text = _fixture("vtb_0401060.txt")
        wb_pdf = _render_pdf(wb_text, title="overhead-wb-a", x_offset=0)
        wb_equivalent_pdf = _render_pdf(wb_equivalent_text, title="overhead-wb-b", x_offset=24)
        vtb_pdf = _render_pdf(vtb_text, title="overhead-vtb", x_offset=8)

        try:
            surface.accept_pool_overhead_preview(
                {
                    **base_payload,
                    "request_id": "overhead:pdf:mismatch",
                    "amount_rub": "1.00",
                },
                actor="fixture-operator",
                source_bytes=wb_pdf,
                filename="synthetic-wb.pdf",
                content_type="application/pdf",
            )
        except FfPoolSurfaceError as exc:
            assert exc.code == "overhead_amount_mismatch"
            assert "account" not in exc.details["payment_evidence"]["payer"]
        else:
            raise AssertionError("attached PDF amount must be authoritative")

        wb_preview = surface.accept_pool_overhead_preview(
            {
                **base_payload,
                "request_id": "overhead:pdf:wb",
                "amount_rub": "",
                "category": "fbs_order_processing",
            },
            actor="fixture-operator",
            source_bytes=wb_pdf,
            filename="synthetic-wb.pdf",
            content_type="application/pdf",
        )
        wb_summary = wb_preview["preview"]["summary"]
        assert wb_preview["state"] == "ready" and wb_summary["amount_rub"] == "12345.67"
        assert wb_summary["business_date"] == "2026-08-13"
        assert wb_summary["payment_evidence"]["adapter"] == "wb_bank_0401060_v1"
        assert "account" not in wb_summary["payment_evidence"]["payer"]
        assert "inn" not in wb_summary["payment_evidence"]["beneficiary"]
        with sqlite3.connect(service.db_path) as conn:
            stored_manifest = json.loads(
                str(
                    conn.execute(
                        f"SELECT request_payload_json FROM {REQUESTS_TABLE} WHERE request_id=?",
                        (str(wb_preview["request_id"]),),
                    ).fetchone()[0]
                )
            )
        assert stored_manifest["payment_evidence"]["payer"]["account"]
        wb_complete = surface.confirm_document(str(wb_preview["request_id"]))
        wb_document_id = str(wb_complete["document"]["document_id"])
        wb_detail = surface.document_detail(wb_document_id)["documents"][0]
        assert wb_detail["source_file_available"]
        assert wb_detail["overhead"]["filename"] == "synthetic-wb.pdf"
        assert wb_detail["overhead"]["payment_evidence"]["payment_fingerprint"]
        source_bytes, source_filename, source_type = surface.source_file(wb_document_id)
        assert source_bytes == wb_pdf and source_filename == "synthetic-wb.pdf"
        assert source_type == "application/pdf"

        wb_storno_preview = surface.accept_document_preview(
            {
                "request_id": "overhead:pdf:wb:storno",
                "document_kind": "storno",
                "business_date": "2026-08-13",
                "manifest": {"target_document_id": wb_document_id},
            },
            actor="fixture-operator",
        )
        wb_storno = surface.confirm_document(str(wb_storno_preview["request_id"]))
        with sqlite3.connect(service.db_path) as conn:
            storno_manifest = json.loads(
                str(
                    conn.execute(
                        f"SELECT posted_manifest_json FROM {DOCUMENTS_TABLE} WHERE document_id=?",
                        (str(wb_storno["document"]["document_id"]),),
                    ).fetchone()[0]
                )
            )
        assert (
            storno_manifest["domain"]["overhead_evidence_link"]["payment_evidence"]
            ["payment_fingerprint"]
            == wb_detail["overhead"]["payment_evidence"]["payment_fingerprint"]
        )

        duplicate = surface.accept_pool_overhead_preview(
            {
                **base_payload,
                "request_id": "overhead:pdf:wb:renamed",
                "amount_rub": "",
                "category": "other",
                "comment": "Эквивалентная синтетическая копия",
            },
            actor="fixture-operator",
            source_bytes=wb_equivalent_pdf,
            filename="renamed-synthetic.pdf",
            content_type="application/pdf",
        )
        assert duplicate["payment_duplicate"]
        assert duplicate["request_id"] == wb_complete["request_id"]
        assert duplicate["duplicate_link"]["document_id"] == wb_document_id

        vtb_preview = surface.accept_pool_overhead_preview(
            {
                **base_payload,
                "request_id": "overhead:pdf:vtb",
                "amount_rub": "",
                "scope": "FBS",
                "category": "inbound_logistics_to_ff",
            },
            actor="fixture-operator",
            source_bytes=vtb_pdf,
            filename="synthetic-vtb.pdf",
            content_type="application/pdf",
        )
        assert not vtb_preview["payment_duplicate"]
        assert vtb_preview["preview"]["summary"]["payment_evidence"]["adapter"] == "vtb_0401060_v2"
        assert surface.confirm_document(str(vtb_preview["request_id"]))["state"] == "posted"

        not_executed_pdf = _render_pdf(
            wb_text.replace("ИСПОЛНЕН\n19.08.2026 10:11:12", "НЕ ИСПОЛНЕН"),
            title="overhead-not-executed",
            x_offset=0,
        )
        blocked = surface.accept_pool_overhead_preview(
            {
                **base_payload,
                "request_id": "overhead:pdf:not-executed",
                "amount_rub": "",
                "category": "returns_processing",
            },
            actor="fixture-operator",
            source_bytes=not_executed_pdf,
            filename="synthetic-not-executed.pdf",
            content_type="application/pdf",
        )
        assert blocked["state"] == "blocked" and not blocked["confirm_allowed"]
        assert blocked["error"]["code"] == "overhead_payment_order_not_eligible"
        assert blocked["preview"]["summary"]["payment_evidence"]["execution_status"] == "not_executed"
        assert blocked["source"]["file_available"]

        blank_pdf = _render_pdf("", title="overhead-ocr-only", x_offset=0)
        ocr_only = surface.accept_pool_overhead_preview(
            {
                **base_payload,
                "request_id": "overhead:pdf:ocr-only",
                "amount_rub": "",
                "category": "packaging_labeling_consumables",
            },
            actor="fixture-operator",
            source_bytes=blank_pdf,
            filename="synthetic-ocr-only.pdf",
            content_type="application/pdf",
        )
        assert ocr_only["state"] == "blocked" and not ocr_only["confirm_allowed"]
        assert ocr_only["preview"]["summary"]["payment_evidence"]["parse_status"] == "parse_error"

        documents = surface.documents_page(document_kind="pool_overhead")["documents"]
        assert len(documents) == 3
        assert all(item["overhead"] for item in documents)


def _surface_balance(db_path: Path, facility_id: str, pool: str, nm_id: int) -> tuple[int, int]:
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            f"SELECT quantity,capital_rub FROM {BALANCES_TABLE} WHERE facility_id=? AND pool=? AND nm_id=?",
            (facility_id, pool, nm_id),
        ).fetchone()
    assert row is not None
    return int(row[0]), int(round(float(row[1]) * 100))


def _surface_quantities(db_path: Path) -> list[tuple[str, str, int, int]]:
    with sqlite3.connect(db_path) as conn:
        return [
            (str(row[0]), str(row[1]), int(row[2]), int(row[3]))
            for row in conn.execute(
                f"SELECT facility_id,pool,nm_id,quantity FROM {BALANCES_TABLE} ORDER BY facility_id,pool,nm_id"
            ).fetchall()
        ]


def _guided_source_uses_canonical_nomenclature() -> None:
    source_rows: list[dict[str, object]] = []
    nomenclature_rows: list[dict[str, object]] = []
    expected_barcodes: dict[int, str] = {}
    for position in range(1, 22):
        nm_id = 700_000_000 + position
        primary = f"0460000000{position:03d}"
        additional = f"1460000000{position:03d}"
        expected_barcodes[nm_id] = primary
        source_rows.append(
            {
                "line_id": f"guided-line-{position}",
                "barcode": "",
                "internal_sku": f"GUIDED-{position:02d}",
                "internal_nm_id": nm_id,
                "internal_name": f"Guided SKU {position:02d}",
                "qty": 6_000 if position == 21 else 3_000,
            }
        )
        nomenclature_rows.append(
            {
                "item_id": f"guided-nm-{nm_id}",
                "nm_id": nm_id,
                "barcode": primary,
                "barcodes_json": json.dumps([primary, additional]),
            }
        )

    resolved = _resolve_supplier_lines_with_canonical_nomenclature(
        source_rows,
        nomenclature_rows,
    )
    repeated = _resolve_supplier_lines_with_canonical_nomenclature(
        source_rows,
        nomenclature_rows,
    )
    assert len(resolved) == 21 and resolved == repeated
    assert sum(int(item["quantity"]) for item in resolved) == 66_000
    assert all(
        item["barcode"] == expected_barcodes[int(item["nm_id"])]
        for item in resolved
    )
    workbook = load_workbook(
        BytesIO(
            generate_china_acceptance_workbook(
                facilities=[
                    {
                        "facility_id": "fac_guided",
                        "code": "GUIDED",
                        "name": "Guided FF",
                        "active": True,
                    }
                ],
                shipment_lines=[
                    {**item, "capital_rub": str(item["quantity"])}
                    for item in resolved
                ],
                source_revision="sha256:" + "a" * 64,
                selected_facility_id="fac_guided",
            )
        ),
        data_only=False,
    )
    sheet = workbook[CHINA_SHEET]
    assert sheet.max_row == 26
    for row_number, item in enumerate(resolved, start=6):
        barcode_cell = sheet.cell(row_number, 2)
        assert barcode_cell.data_type == "s"
        assert barcode_cell.value == item["barcode"]

    _assert_guided_source_error(
        source_rows,
        nomenclature_rows[1:],
        "exact_identity_evidence_missing",
    )
    drifted_source = [dict(item) for item in source_rows]
    drifted_source[0]["barcode"] = "9999999999999"
    _assert_guided_source_error(
        drifted_source,
        nomenclature_rows,
        "supplier_identity_drift",
    )
    shared_barcode = [dict(item) for item in nomenclature_rows]
    shared_barcode[1]["barcode"] = expected_barcodes[700_000_001]
    shared_barcode[1]["barcodes_json"] = json.dumps(
        [expected_barcodes[700_000_001]]
    )
    _assert_guided_source_error(
        source_rows,
        shared_barcode,
        "ambiguous_nomenclature_barcode",
    )
    duplicate_nm = [dict(item) for item in nomenclature_rows]
    duplicate_nm.append(
        {
            "item_id": "guided-duplicate-nm",
            "nm_id": 700_000_001,
            "barcode": "0469999999999",
            "barcodes_json": "[]",
        }
    )
    _assert_guided_source_error(
        source_rows,
        duplicate_nm,
        "ambiguous_nomenclature",
    )
    changed_nomenclature = [dict(item) for item in nomenclature_rows]
    changed_nomenclature[0]["barcode"] = "0468888888888"
    changed_nomenclature[0]["barcodes_json"] = json.dumps(["0468888888888"])
    changed = _resolve_supplier_lines_with_canonical_nomenclature(
        source_rows,
        changed_nomenclature,
    )
    assert changed[0]["identity_revision"] != resolved[0]["identity_revision"]


def _assert_guided_source_error(
    source_rows: list[dict[str, object]],
    nomenclature_rows: list[dict[str, object]],
    expected_code: str,
) -> None:
    try:
        _resolve_supplier_lines_with_canonical_nomenclature(
            source_rows,
            nomenclature_rows,
        )
    except FfPoolSurfaceError as exc:
        assert exc.code == expected_code, (expected_code, exc.code)
    else:
        raise AssertionError(f"guided source must fail closed with {expected_code}")


def _guided_preview_is_default_off() -> None:
    with TemporaryDirectory(prefix="ff-guided-default-off-") as directory:
        root = Path(directory)
        clock = Clock()
        service = FfPoolDocumentService(
            db_path=root / "state.sqlite3", runtime_dir=root, timestamp_factory=clock, resume=False
        )
        with sqlite3.connect(service.db_path) as conn:
            conn.execute(
                f"INSERT INTO {FACILITIES_TABLE} VALUES(?,?,?,?,?,?,?)",
                ("fac_preview", "PREVIEW", "Preview FF", 1, "Europe/Moscow", clock(), clock()),
            )
            conn.execute(
                f"INSERT INTO {FACILITY_PROFILES_TABLE} VALUES(?,?,?, ?,?)",
                ("fac_preview", "Москва", "{}", clock(), clock()),
            )
            conn.commit()
        lines = [{
            "nm_id": 101, "barcode": "0000000000101", "sku": "SKU-101",
            "accepted_quantity": 2, "accepted_capital_rub": "200.00",
        }]
        workbook_bytes = service.generate_china_acceptance_template(
            shipment_lines=lines, source_revision="supplier-revision-v1", selected_facility_id="fac_preview"
        )
        workbook = load_workbook(BytesIO(workbook_bytes))
        workbook[CHINA_SHEET]["G6"], workbook[CHINA_SHEET]["H6"] = 1, 1
        output = BytesIO()
        workbook.save(output)
        preview = service.preview_china_acceptance_workbook(
            identity=DocumentIdentity(
                request_id="fixture:guided-preview:request",
                source_system="supplier_registry",
                source_type="china_acceptance_workbook",
                source_id="shipment-preview",
                source_revision="supplier-revision-v1",
                idempotency_epoch=1,
                actor="fixture",
                business_date="2026-08-12",
            ),
            source_bytes=output.getvalue(),
            source_filename="guided.xlsx",
            source_content_type=XLSX_CONTENT_TYPE,
            shipment_lines=lines,
            template_source_revision="supplier-revision-v1",
        )
        assert preview["state"] == "ready"
        surface = FfPoolSurface(db_path=service.db_path, runtime_dir=root, timestamp_factory=clock)
        status = surface.request_status(str(preview["request_id"]))
        assert not status["confirm_allowed"]
        assert status["guided_acceptance_activation"]["reason"] == "writer_epoch_off"
        try:
            surface.confirm_document(str(preview["request_id"]))
            raise AssertionError("default-off guided acceptance must fail closed")
        except FfPoolSurfaceError as exc:
            assert exc.code == "guided_acceptance_not_activated"
        with sqlite3.connect(service.db_path) as conn:
            assert conn.execute(f"SELECT COUNT(*) FROM {DOCUMENTS_TABLE}").fetchone()[0] == 0


def _schema_absence_is_controlled() -> None:
    with TemporaryDirectory(prefix="ff-pool-surface-absent-") as directory:
        db_path = Path(directory) / "empty.sqlite3"
        db_path.touch()
        surface = FfPoolSurface(db_path=db_path, runtime_dir=Path(directory))
        payload = surface.capabilities()
        assert payload["status"] == "schema_absent"
        assert payload["feature"]["reason"] == "schema_absent_default_off"
        assert payload["hidden_actions"] == ["facility_pool_opening"]


def _counts(db_path: Path) -> dict[str, int]:
    with sqlite3.connect(db_path) as conn:
        tables = [
            str(row[0])
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'sheet_vitrina_v1_ff_%'"
            ).fetchall()
        ]
        return {table: int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]) for table in tables}


def _default_off_reads_do_not_write(surface: FfPoolSurface, db_path: Path) -> None:
    before = _counts(db_path)
    first = surface.capabilities(aggregate_revision="aggregate-empty")
    page = surface.facilities_page(aggregate_revision="aggregate-empty")
    documents = surface.documents_page()
    after = _counts(db_path)
    assert before == after
    assert not first["feature"]["writer_effective"]
    assert not first["feature"]["reader_effective"]
    assert page["facilities"] == [] and documents["documents"] == []
    assert first["etag"] == surface.capabilities(aggregate_revision="aggregate-empty")["etag"]
    try:
        surface.create_facility(
            {"request_id": "fixture:facility:off", "name": "Blocked", "active": True},
            actor="fixture",
        )
    except FfPoolSurfaceError as exc:
        assert exc.code == "facility_pool_feature_off" and exc.http_status == 409
    else:
        raise AssertionError("feature-off facility mutation must fail closed")
    assert _counts(db_path) == before


def _facility_management(surface: FfPoolSurface, db_path: Path, clock: Clock) -> list[dict[str, object]]:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            f"INSERT INTO {FEATURE_EPOCHS_TABLE}(epoch,writer_enabled,reader_enabled,source_revision,created_at,metadata_json) "
            "VALUES(1,1,0,'stage3-fixture-writer',?,'{}')",
            (clock(),),
        )
        conn.commit()
    first_payload = {
        "request_id": "fixture:facility:moscow",
        "name": "Москва FF",
        "city": "Москва",
        "active": True,
        "display_timezone": "Europe/Moscow",
    }
    first = surface.create_facility(first_payload, actor="fixture-operator")
    repeated = surface.create_facility(first_payload, actor="fixture-operator")
    assert not first["idempotent"] and repeated["idempotent"]
    facility = dict(first["facility"])
    stable_id, stable_code = facility["facility_id"], facility["code"]
    updated = surface.update_facility(
        str(stable_id),
        {
            "request_id": "fixture:facility:moscow:update",
            "expected_updated_at": facility["updated_at"],
            "name": "Москва Север",
            "active": False,
            "display_timezone": "Asia/Yekaterinburg",
        },
        actor="fixture-operator",
    )
    assert updated["facility"]["facility_id"] == stable_id
    assert updated["facility"]["code"] == stable_code
    assert len(updated["audit"]) == 4
    second = surface.create_facility(
        {
            "request_id": "fixture:facility:orenburg",
            "name": "Москва Юг",
            "city": "Москва",
            "active": True,
            "display_timezone": "Asia/Yekaterinburg",
        },
        actor="fixture-operator",
    )
    with sqlite3.connect(db_path) as conn:
        assert conn.execute(f"SELECT COUNT(*) FROM {FACILITY_CHANGES_TABLE}").fetchone()[0] == 5
        try:
            conn.execute(f"DELETE FROM {FACILITIES_TABLE} WHERE facility_id=?", (stable_id,))
        except sqlite3.IntegrityError as exc:
            assert "retained" in str(exc)
        else:
            raise AssertionError("facility physical delete must be blocked")
    return [dict(updated["facility"]), dict(second["facility"])]


def _deactivation_with_dependencies_is_blocked(
    surface: FfPoolSurface,
    facilities: list[dict[str, object]],
) -> None:
    facility_id = str(facilities[1]["facility_id"])
    detail = surface.facility_detail(facility_id)
    try:
        surface.update_facility(
            facility_id,
            {
                "request_id": "fixture:facility:deactivation-blocked",
                "expected_updated_at": detail["facility"]["updated_at"],
                "active": False,
            },
            actor="fixture-operator",
        )
    except FfPoolSurfaceError as exc:
        assert exc.code == "facility_deactivation_blocked"
        assert exc.details["nonzero_balance_count"] == 2
    else:
        raise AssertionError("facility with non-zero balances must stay active")


def _document_workflow(surface: FfPoolSurface, facilities: list[dict[str, object]]) -> str:
    request = surface.accept_document_preview(
        {
            "request_id": "fixture:transfer:root",
            "document_kind": "transfer_root",
            "business_date": "2026-08-12",
            "manifest": {
                "source": {"facility_id": facilities[1]["facility_id"], "pool": "FBS"},
                "destination": {"facility_id": facilities[1]["facility_id"], "pool": "FBO"},
            },
        },
        actor="fixture-operator",
    )
    assert request["state"] == "ready" and request["confirm_allowed"]
    canonical = str(request["request_id"])
    confirmed = surface.confirm_document(canonical)
    assert confirmed["state"] == "complete"
    repeat = surface.accept_document_preview(
        {
            "request_id": "fixture:transfer:root",
            "document_kind": "transfer_root",
            "business_date": "2026-08-12",
            "manifest": {
                "source": {"facility_id": facilities[1]["facility_id"], "pool": "FBS"},
                "destination": {"facility_id": facilities[1]["facility_id"], "pool": "FBO"},
            },
        },
        actor="fixture-operator",
    )
    assert repeat["request_id"] == canonical and repeat["state"] == "complete"
    return canonical


def _exact_reader_and_pagination(
    surface: FfPoolSurface,
    db_path: Path,
    clock: Clock,
    facilities: list[dict[str, object]],
) -> None:
    facility_id = str(facilities[1]["facility_id"])
    aggregate = [
        {"nm_id": 101, "quantity": 1, "capital_rub": "0.1"},
        {"nm_id": 102, "quantity": 2, "capital_rub": "0.2"},
    ]
    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute(
            f"INSERT INTO {FEATURE_EPOCHS_TABLE}(epoch,writer_enabled,reader_enabled,source_revision,created_at,metadata_json) "
            "VALUES(2,1,1,'stage3-fixture-reader',?,'{}')",
            (clock(),),
        )
        conn.executemany(
            f"INSERT INTO {BALANCES_TABLE}(facility_id,pool,nm_id,projection_epoch,quantity,capital_rub,wac_rub,source_watermark,updated_at) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            [
                (facility_id, "FBS", 101, 2, 1, "0.1", "0.1", "fixture", clock()),
                (facility_id, "FBS", 102, 2, 2, "0.2", "0.1", "fixture", clock()),
            ],
        )
        parity = evaluate_ff_pool_aggregate_parity(conn, aggregate)
        assert parity.status == "pass"
        record_ff_pool_parity_diagnostic(
            conn,
            diagnostic_id="stage3-fixture-pass",
            aggregate_revision="aggregate-v2",
            checked_at=clock(),
            result=parity,
        )
        conn.commit()
    page = surface.facilities_page(page=1, limit=1, aggregate_revision="aggregate-v2")
    assert page["feature"]["reader_effective"] and page["page"]["has_next"]
    second_page = surface.facilities_page(page=2, limit=1, aggregate_revision="aggregate-v2")
    assert len(second_page["facilities"]) == 1
    detail = surface.facility_detail(facility_id, aggregate_revision="aggregate-v2")
    fbs = next(item for item in detail["pools"] if item["pool"] == "FBS")
    assert fbs["quantity"] == 3 and fbs["capital_rub"] == "0.3" and fbs["wac_rub"] == "0.1"
    pool = surface.pool_detail(facility_id, "FBS", limit=1, aggregate_revision="aggregate-v2")
    assert len(pool["balances"]) == 1 and pool["page"]["has_next"]


def _read_models(surface: FfPoolSurface, request_id: str, facilities: list[dict[str, object]]) -> None:
    status = surface.request_status(request_id)
    assert status["state"] == "complete" and len(status["steps"]) == 6
    documents = surface.documents_page(page=1, limit=10)
    assert documents["page"]["total_count"] == 1
    assert all(item["document_kind"] != "facility_pool_opening" for item in documents["documents"])
    root_id = str(documents["documents"][0]["root_document_id"])
    detail = surface.document_detail(root_id)
    assert detail["root_document_id"] == root_id and len(detail["documents"]) == 1
    assert surface.document_lines(root_id)["lines"] == []
    assert surface.document_expenses(root_id)["expenses"] == []
    graph = surface.document_graph(root_id)
    assert len(graph["nodes"]) == 1
    filtered = surface.documents_page(facility_id=str(facilities[1]["facility_id"]))
    assert filtered["documents"] == []  # root has no movement lines; context is never inferred.
    assert documents["payload_bytes"] < 128 * 1024


if __name__ == "__main__":
    main()
