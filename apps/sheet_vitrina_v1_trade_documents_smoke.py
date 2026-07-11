"""Smoke-check trade document registry for supplier invoices/contracts."""

from __future__ import annotations

from contextlib import contextmanager
from http.cookiejar import CookieJar
from io import BytesIO
import base64
import hashlib
import json
import os
from pathlib import Path
import socket
import sys
from tempfile import TemporaryDirectory
import threading
from urllib import error as urllib_error, parse as urllib_parse, request as urllib_request
from uuid import uuid4

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.adapters.registry_upload_http_entrypoint import (  # noqa: E402
    DEFAULT_SHEET_OPERATOR_UI_PATH,
    DEFAULT_SHEET_PLAN_PATH,
    DEFAULT_SHEET_STATUS_PATH,
    DEFAULT_SUPPLIER_SHIPMENTS_PARSE_PATH,
    DEFAULT_SUPPLIER_SHIPMENTS_PATH,
    DEFAULT_TRADE_DOCUMENTS_PATH,
    DEFAULT_UPLOAD_PATH,
    build_registry_upload_http_server,
)
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime  # noqa: E402
from packages.application.registry_upload_http_entrypoint import RegistryUploadHttpEntrypoint  # noqa: E402
from packages.application import supplier_shipments as supplier_shipments_module  # noqa: E402
from packages.application.supplier_shipments import SupplierShipmentsBlock  # noqa: E402
from packages.contracts.supplier_shipments import (  # noqa: E402
    DEFAULT_SUPPLIER_NAME,
    SUPPLIER_INVOICE_CONTENT_TYPE,
    TRADE_DOCUMENT_CONTRACT_PARSER_VERSION,
    TRADE_DOCUMENT_SOURCE_SETTINGS_UPLOAD,
    TRADE_DOCUMENT_STATUS_ACTIVE,
    TRADE_DOCUMENT_TYPE_CONTRACT,
)
from packages.contracts.registry_upload_http_entrypoint import RegistryUploadHttpEntrypointConfig  # noqa: E402


def main() -> None:
    _assert_contract_parser_rules()
    _assert_contract_ocr_parser_with_fake_tools()
    _assert_contract_ocr_missing_tools_warning()
    owner_password = "owner-password-not-secret"
    supplier_password = "supplier-password-not-secret"
    invoice_bytes = _build_invoice_fixture()
    auto_contract_bytes = _build_contract_xlsx_fixture("AUTO-2026-0601", "2026.06.01")
    override_contract_bytes = _build_contract_xlsx_fixture("PARSER-2026-0703", "2026年7月3日")
    chinese_contract_bytes = _build_chinese_contract_xlsx_fixture("CN-2026-0910", "合同日期 2026年9月10日")
    pdf_contract_bytes = _build_contract_pdf_fixture("PDF-2026-0602", "Contract Date: 06/02/2026")
    contract_bytes = b"%PDF-1.4\n% wb-core trade document smoke\n"
    with TemporaryDirectory(prefix="trade-documents-smoke-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        runtime.save_nomenclature_item(
            {
                "item_id": "trade-documents-clear-iphone-14-pro",
                "is_active": True,
                "our_sku": "SMOKE-IP14PRO-CLEAR",
                "nm_id": 210183919,
                "vendor_code": "SMOKE-IP14PRO-CLEAR",
                "barcode": "210183919",
                "nomenclature_name": "Clear iPhone 14 Pro",
                "product_type": "clear",
                "match_key": "clear|iphone_14_pro",
                "purchase_price_yuan": 1,
                "aliases": [],
                "compatible_models_text": "iPhone 14 Pro",
                "compatible_model_keys": ["iphone_14_pro"],
                "comment": "trade documents smoke",
                "created_at": "2026-05-30T08:00:00Z",
                "updated_at": "2026-05-30T08:00:00Z",
            }
        )
        config = RegistryUploadHttpEntrypointConfig(
            host="127.0.0.1",
            port=_reserve_free_port(),
            upload_path=DEFAULT_UPLOAD_PATH,
            sheet_plan_path=DEFAULT_SHEET_PLAN_PATH,
            sheet_refresh_path="/v1/sheet-vitrina-v1/refresh",
            sheet_status_path=DEFAULT_SHEET_STATUS_PATH,
            sheet_operator_ui_path=DEFAULT_SHEET_OPERATOR_UI_PATH,
            runtime_dir=runtime_dir,
        )
        clock = _Clock()
        with _patched_env(
            {
                "WB_CORE_WEB_AUTH_REQUIRED": "1",
                "WB_CORE_WEB_AUTH_USERNAME": "owner",
                "WB_CORE_WEB_AUTH_PASSWORD_HASH": _password_hash(owner_password),
                "WB_CORE_WEB_AUTH_SESSION_SECRET": "trade-documents-smoke-session-secret",
                "WB_CORE_SUPPLIER_AUTH_USERNAME": "supplier",
                "WB_CORE_SUPPLIER_AUTH_PASSWORD_HASH": _password_hash(supplier_password),
                "WB_CORE_SUPPLIER_AUTH_DISPLAY_NAME": "Supplier",
            }
        ):
            entrypoint = RegistryUploadHttpEntrypoint(
                runtime_dir=runtime_dir,
                runtime=runtime,
                activated_at_factory=clock.next,
            )
            server = build_registry_upload_http_server(config, entrypoint=entrypoint)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                base_url = f"http://127.0.0.1:{config.port}"
                operator = urllib_request.build_opener(urllib_request.HTTPCookieProcessor(CookieJar()))
                supplier = urllib_request.build_opener(urllib_request.HTTPCookieProcessor(CookieJar()))
                _login(operator, base_url, "owner", owner_password)
                _login(supplier, base_url, "supplier", supplier_password)

                backfill_block = SupplierShipmentsBlock(runtime=runtime, timestamp_factory=clock.next)
                backfill_contract_bytes = _build_contract_xlsx_fixture("BACKFILL-2026-0801", "Contract Date: 2026-08-01")
                backfill_contract_path = _write_runtime_contract_file(
                    runtime_dir,
                    document_id="tdoc_backfill_empty_contract",
                    filename="backfill-empty-contract.xlsx",
                    body=backfill_contract_bytes,
                )
                runtime.save_trade_document(
                    {
                        "document_id": "tdoc_backfill_empty_contract",
                        "document_type": TRADE_DOCUMENT_TYPE_CONTRACT,
                        "number": "",
                        "document_date": "",
                        "supplier_name": "",
                        "currency": "",
                        "amount_total": None,
                        "source": TRADE_DOCUMENT_SOURCE_SETTINGS_UPLOAD,
                        "source_shipment_id": "",
                        "source_upload_id": "",
                        "file_original_name": "backfill-empty-contract.xlsx",
                        "file_content_type": SUPPLIER_INVOICE_CONTENT_TYPE,
                        "file_sha256": hashlib.sha256(backfill_contract_bytes).hexdigest(),
                        "file_path": backfill_contract_path,
                        "parser_version": "",
                        "parsed_metadata": {},
                        "warnings": ["contract parser skipped OCR: OCR tools missing (pdftoppm/tesseract)"],
                        "errors": [],
                        "status": TRADE_DOCUMENT_STATUS_ACTIVE,
                        "created_at": clock.next(),
                        "updated_at": clock.next(),
                    }
                )
                manual_contract_bytes = _build_contract_xlsx_fixture("PARSER-SHOULD-NOT-WIN", "2026-08-02")
                manual_contract_path = _write_runtime_contract_file(
                    runtime_dir,
                    document_id="tdoc_backfill_manual_contract",
                    filename="backfill-manual-contract.xlsx",
                    body=manual_contract_bytes,
                )
                runtime.save_trade_document(
                    {
                        "document_id": "tdoc_backfill_manual_contract",
                        "document_type": TRADE_DOCUMENT_TYPE_CONTRACT,
                        "number": "MANUAL-KEEP-0802",
                        "document_date": "2026-08-03",
                        "supplier_name": "",
                        "currency": "",
                        "amount_total": None,
                        "source": TRADE_DOCUMENT_SOURCE_SETTINGS_UPLOAD,
                        "source_shipment_id": "",
                        "source_upload_id": "",
                        "file_original_name": "backfill-manual-contract.xlsx",
                        "file_content_type": SUPPLIER_INVOICE_CONTENT_TYPE,
                        "file_sha256": hashlib.sha256(manual_contract_bytes).hexdigest(),
                        "file_path": manual_contract_path,
                        "parser_version": "",
                        "parsed_metadata": {},
                        "warnings": [],
                        "errors": [],
                        "status": TRADE_DOCUMENT_STATUS_ACTIVE,
                        "created_at": clock.next(),
                        "updated_at": clock.next(),
                    }
                )
                filled_warning_contract_bytes = _build_contract_xlsx_fixture("FILLED-WARN-0804", "2026-08-04")
                filled_warning_contract_path = _write_runtime_contract_file(
                    runtime_dir,
                    document_id="tdoc_backfill_filled_warning_contract",
                    filename="backfill-filled-warning-contract.xlsx",
                    body=filled_warning_contract_bytes,
                )
                runtime.save_trade_document(
                    {
                        "document_id": "tdoc_backfill_filled_warning_contract",
                        "document_type": TRADE_DOCUMENT_TYPE_CONTRACT,
                        "number": "FILLED-WARN-0804",
                        "document_date": "2026-08-04",
                        "supplier_name": DEFAULT_SUPPLIER_NAME,
                        "currency": "",
                        "amount_total": None,
                        "source": TRADE_DOCUMENT_SOURCE_SETTINGS_UPLOAD,
                        "source_shipment_id": "",
                        "source_upload_id": "",
                        "file_original_name": "backfill-filled-warning-contract.xlsx",
                        "file_content_type": SUPPLIER_INVOICE_CONTENT_TYPE,
                        "file_sha256": hashlib.sha256(filled_warning_contract_bytes).hexdigest(),
                        "file_path": filled_warning_contract_path,
                        "parser_version": TRADE_DOCUMENT_CONTRACT_PARSER_VERSION,
                        "parsed_metadata": {},
                        "warnings": ["contract parser skipped OCR: OCR tools missing (pdftoppm/tesseract)"],
                        "errors": [],
                        "status": TRADE_DOCUMENT_STATUS_ACTIVE,
                        "created_at": clock.next(),
                        "updated_at": clock.next(),
                    }
                )
                backfill_result = backfill_block.backfill_trade_document_metadata()
                empty_backfilled = runtime.load_trade_document("tdoc_backfill_empty_contract") or {}
                manual_backfilled = runtime.load_trade_document("tdoc_backfill_manual_contract") or {}
                filled_warning_backfilled = runtime.load_trade_document("tdoc_backfill_filled_warning_contract") or {}
                if (
                    backfill_result.get("updated_documents") != 3
                    or empty_backfilled.get("number") != "BACKFILL-2026-0801"
                    or empty_backfilled.get("document_date") != "2026-08-01"
                    or empty_backfilled.get("supplier_name") != DEFAULT_SUPPLIER_NAME
                    or empty_backfilled.get("parser_version") != TRADE_DOCUMENT_CONTRACT_PARSER_VERSION
                    or empty_backfilled.get("warnings") != []
                    or manual_backfilled.get("number") != "MANUAL-KEEP-0802"
                    or manual_backfilled.get("document_date") != "2026-08-03"
                    or manual_backfilled.get("supplier_name") != DEFAULT_SUPPLIER_NAME
                    or filled_warning_backfilled.get("warnings") != []
                ):
                    raise AssertionError(
                        "trade document metadata backfill failed: "
                        f"{backfill_result} {empty_backfilled} {manual_backfilled} {filled_warning_backfilled}"
                    )
                second_backfill_result = backfill_block.backfill_trade_document_metadata()
                if second_backfill_result.get("updated_documents") != 0:
                    raise AssertionError(f"trade document metadata backfill must be idempotent, got {second_backfill_result}")

                auto_contract_code, auto_contract_payload = _opener_post_multipart(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                    auto_contract_bytes,
                    filename="auto-contract.xlsx",
                    fields={"document_type": "contract"},
                )
                auto_contract_doc = auto_contract_payload.get("document") or {}
                auto_contract_id = str(auto_contract_doc.get("document_id") or "")
                if (
                    auto_contract_code != 200
                    or not auto_contract_id
                    or auto_contract_doc.get("number") != "AUTO-2026-0601"
                    or auto_contract_doc.get("document_date") != "2026-06-01"
                    or auto_contract_doc.get("parsed_number") != "AUTO-2026-0601"
                    or auto_contract_doc.get("parsed_document_date") != "2026-06-01"
                    or auto_contract_doc.get("supplier_name") != DEFAULT_SUPPLIER_NAME
                ):
                    raise AssertionError(f"contract XLSX upload must parse number/date, got {auto_contract_code} {auto_contract_payload}")

                edit_contract_code, edit_contract_payload = _opener_patch_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{auto_contract_id}",
                    {"number": "AUTO-EDITED-0603", "document_date": "2026-06-03", "supplier_name": ""},
                )
                edit_contract_doc = edit_contract_payload.get("document") or {}
                if (
                    edit_contract_code != 200
                    or edit_contract_doc.get("number") != "AUTO-EDITED-0603"
                    or edit_contract_doc.get("document_date") != "2026-06-03"
                    or edit_contract_doc.get("supplier_name") != DEFAULT_SUPPLIER_NAME
                ):
                    raise AssertionError(f"contract metadata PATCH must edit allowed fields and default supplier, got {edit_contract_code} {edit_contract_payload}")
                invalid_edit_code, invalid_edit_payload = _opener_patch_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{auto_contract_id}",
                    {"document_date": "03.06.2026"},
                )
                if invalid_edit_code != 400 or "YYYY-MM-DD" not in str(invalid_edit_payload.get("error", "")):
                    raise AssertionError(f"invalid document_date PATCH must return 400, got {invalid_edit_code} {invalid_edit_payload}")
                unsupported_edit_code, unsupported_edit_payload = _opener_patch_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{auto_contract_id}",
                    {"amount_total": 99},
                )
                if unsupported_edit_code != 400 or "unsupported" not in str(unsupported_edit_payload.get("error", "")).lower():
                    raise AssertionError(f"unsupported document metadata PATCH must return 400, got {unsupported_edit_code} {unsupported_edit_payload}")

                chinese_contract_code, chinese_contract_payload = _opener_post_multipart(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                    chinese_contract_bytes,
                    filename="chinese-contract.xlsx",
                    fields={"document_type": "contract"},
                )
                chinese_contract_doc = chinese_contract_payload.get("document") or {}
                if (
                    chinese_contract_code != 200
                    or chinese_contract_doc.get("number") != "CN-2026-0910"
                    or chinese_contract_doc.get("document_date") != "2026-09-10"
                    or chinese_contract_doc.get("supplier_name") != DEFAULT_SUPPLIER_NAME
                ):
                    raise AssertionError(f"contract Chinese labels must parse number/date, got {chinese_contract_code} {chinese_contract_payload}")

                pdf_contract_code, pdf_contract_payload = _opener_post_multipart(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                    pdf_contract_bytes,
                    filename="text-contract.pdf",
                    content_type="application/pdf",
                    fields={"document_type": "contract"},
                )
                pdf_contract_doc = pdf_contract_payload.get("document") or {}
                if (
                    pdf_contract_code != 200
                    or pdf_contract_doc.get("number") != "PDF-2026-0602"
                    or pdf_contract_doc.get("document_date") != "2026-06-02"
                    or pdf_contract_doc.get("supplier_name") != DEFAULT_SUPPLIER_NAME
                ):
                    raise AssertionError(f"contract PDF text upload must parse number/date, got {pdf_contract_code} {pdf_contract_payload}")

                override_contract_code, override_contract_payload = _opener_post_multipart(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                    override_contract_bytes,
                    filename="override-contract.xlsx",
                    fields={
                        "document_type": "contract",
                        "number": "MANUAL-2026-0704",
                        "document_date": "2026-07-04",
                    },
                )
                override_contract_doc = override_contract_payload.get("document") or {}
                override_warnings = " ".join(str(item) for item in override_contract_doc.get("parser_warnings") or [])
                if (
                    override_contract_code != 200
                    or override_contract_doc.get("number") != "MANUAL-2026-0704"
                    or override_contract_doc.get("document_date") != "2026-07-04"
                    or override_contract_doc.get("parsed_number") != "PARSER-2026-0703"
                    or override_contract_doc.get("parsed_document_date") != "2026-07-03"
                    or override_contract_doc.get("supplier_name") != DEFAULT_SUPPLIER_NAME
                    or "manual" not in override_warnings.lower()
                ):
                    raise AssertionError(f"manual contract metadata must override parser values, got {override_contract_code} {override_contract_payload}")

                contract_code, contract_payload = _opener_post_multipart(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                    contract_bytes,
                    filename="contract-cnt-2026-0513.pdf",
                    content_type="application/pdf",
                    fields={
                        "document_type": "contract",
                        "number": "CNT-2026-0513",
                        "document_date": "2026-05-13",
                        "supplier_name": "HanShang Technology",
                    },
                )
                contract_doc = contract_payload.get("document") or {}
                contract_id = str(contract_doc.get("document_id") or "")
                if contract_code != 200 or not contract_id or contract_doc.get("document_type") != "contract":
                    raise AssertionError(f"contract upload failed: {contract_code} {contract_payload}")

                duplicate_code, duplicate_payload = _opener_post_multipart(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                    contract_bytes,
                    filename="contract-cnt-2026-0513.pdf",
                    content_type="application/pdf",
                    fields={"document_type": "contract", "number": "CNT-2026-0513"},
                )
                if (
                    duplicate_code != 200
                    or duplicate_payload.get("status") != "duplicate_existing"
                    or (duplicate_payload.get("document") or {}).get("document_id") != contract_id
                ):
                    raise AssertionError(f"settings duplicate upload must return existing document, got {duplicate_code} {duplicate_payload}")

                invoice_code, invoice_payload = _opener_post_multipart(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                    invoice_bytes,
                    filename="PI-test 26GN390.xlsx",
                    fields={"document_type": "invoice"},
                )
                invoice_doc = invoice_payload.get("document") or {}
                invoice_id = str(invoice_doc.get("document_id") or "")
                if (
                    invoice_code != 200
                    or invoice_doc.get("number") != "26GN390"
                    or invoice_doc.get("document_date") != "2026-05-14"
                    or invoice_doc.get("parsed_metadata", {}).get("contract_no") != "CNT-2026-0513"
                ):
                    raise AssertionError(f"settings invoice upload must parse metadata, got {invoice_code} {invoice_payload}")

                empty_shipments_code, empty_shipments_payload = _opener_json(
                    operator,
                    f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}",
                )
                if empty_shipments_code != 200 or empty_shipments_payload.get("shipments") != []:
                    raise AssertionError("settings invoice upload must not create supplier shipment")

                link_code, link_payload = _opener_patch_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{invoice_id}/contract",
                    {"contract_document_id": contract_id},
                )
                if link_code != 200 or link_payload.get("link", {}).get("contract_document_id") != contract_id:
                    raise AssertionError(f"invoice->contract link failed: {link_code} {link_payload}")

                for document, expected_sha in ((contract_doc, hashlib.sha256(contract_bytes).hexdigest()), (invoice_doc, hashlib.sha256(invoice_bytes).hexdigest())):
                    path = str(document.get("download_path") or "")
                    status, body, headers = _opener_bytes(operator, f"{base_url}{path}")
                    if status != 200 or hashlib.sha256(body).hexdigest() != expected_sha:
                        raise AssertionError(f"document download mismatch for {path}: {status} {headers}")

                archive_contract_code, archive_contract_payload = _opener_delete_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{contract_id}",
                )
                if archive_contract_code != 400 or "linked invoice" not in str(archive_contract_payload.get("error", "")).lower():
                    raise AssertionError(f"linked contract archive must be rejected, got {archive_contract_code} {archive_contract_payload}")

                unlink_code, unlink_payload = _opener_delete_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{invoice_id}/contract",
                )
                if (
                    unlink_code != 200
                    or unlink_payload.get("deleted") is not True
                    or (unlink_payload.get("invoice") or {}).get("linked_contract_document_id")
                ):
                    raise AssertionError(f"invoice->contract unlink failed: {unlink_code} {unlink_payload}")
                second_unlink_code, second_unlink_payload = _opener_delete_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{invoice_id}/contract",
                )
                if second_unlink_code != 200 or second_unlink_payload.get("deleted") is not False:
                    raise AssertionError(f"invoice->contract unlink must be idempotent, got {second_unlink_code} {second_unlink_payload}")
                relink_code, relink_payload = _opener_patch_json(
                    operator,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{invoice_id}/contract",
                    {"contract_document_id": contract_id},
                )
                if relink_code != 200 or relink_payload.get("link", {}).get("contract_document_id") != contract_id:
                    raise AssertionError(f"invoice->contract relink failed: {relink_code} {relink_payload}")

                supplier_settings_code, supplier_settings_payload = _opener_json(
                    supplier,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}",
                )
                if supplier_settings_code != 403 or supplier_settings_payload.get("error") != "forbidden":
                    raise AssertionError("supplier role must not list settings documents")
                supplier_unlink_code, supplier_unlink_payload = _opener_delete_json(
                    supplier,
                    f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}/{invoice_id}/contract",
                )
                if supplier_unlink_code != 403 or supplier_unlink_payload.get("error") != "forbidden":
                    raise AssertionError("supplier role must not unlink settings document contracts")
                supplier_arbitrary_file_code, _, _ = _opener_bytes(
                    supplier,
                    f"{base_url}{contract_doc.get('download_path')}",
                )
                if supplier_arbitrary_file_code != 403:
                    raise AssertionError("supplier role must not download arbitrary settings document file")

                parse_code, parse_payload = _opener_post_multipart(
                    supplier,
                    f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PARSE_PATH}",
                    invoice_bytes,
                    filename="supplier-flow.xlsx",
                )
                if parse_code != 200 or not parse_payload.get("upload_id"):
                    raise AssertionError(f"supplier parse failed: {parse_code} {parse_payload}")
                create_code, create_payload = _opener_post_json(
                    supplier,
                    f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}",
                    {"upload_id": parse_payload["upload_id"], "shipment_date": "2026-05-14", "payload": parse_payload},
                )
                shipment_id = str(create_payload.get("shipment_id") or "")
                if (
                    create_code != 200
                    or not shipment_id
                    or not create_payload.get("invoice_document_id")
                    or create_payload.get("contract_document_id") != contract_id
                    or not create_payload.get("contract_download_path")
                ):
                    raise AssertionError(f"supplier shipment must auto-create/link documents, got {create_code} {create_payload}")
                supplier_contract_code, supplier_contract_bytes, _ = _opener_bytes(
                    supplier,
                    f"{base_url}{create_payload.get('contract_download_path')}",
                )
                if supplier_contract_code != 200 or supplier_contract_bytes != contract_bytes:
                    raise AssertionError("supplier role must download shipment-linked contract")

                legacy_file = runtime_dir / "supplier_invoices" / "files" / "sup_legacy_doc" / "legacy.xlsx"
                legacy_file.parent.mkdir(parents=True, exist_ok=True)
                legacy_file.write_bytes(invoice_bytes)
                runtime.save_supplier_shipment(
                    header={
                        "shipment_id": "sup_legacy_doc",
                        "created_at": "2026-05-30T08:10:00Z",
                        "updated_at": "2026-05-30T08:10:00Z",
                        "shipment_date": "2026-05-16",
                        "invoice_no": "LEGACY-DOC",
                        "invoice_date": "2026-05-15",
                        "contract_no": "CNT-2026-0513",
                        "contract_date": "2026-05-13",
                        "supplier_name": "",
                        "customer_name": "",
                        "currency": "RMB",
                        "product_qty_total": 0,
                        "product_amount_total": 0,
                        "extras_amount_total": 0,
                        "invoice_amount_total": 0,
                        "declared_invoice_total": 0,
                        "match_status": "all_matched",
                        "source_filename": "legacy.xlsx",
                        "source_file_sha256": hashlib.sha256(invoice_bytes).hexdigest(),
                        "source_file_path": legacy_file.relative_to(runtime_dir).as_posix(),
                        "parser_version": "legacy",
                        "warnings": [],
                        "errors": [],
                    },
                    lines=[],
                )
                before_docs_code, before_docs = _opener_json(operator, f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}")
                before_count = len(before_docs.get("documents", []))
                legacy_detail_code, legacy_detail = _opener_json(
                    operator,
                    f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/sup_legacy_doc",
                )
                if (
                    legacy_detail_code != 200
                    or not legacy_detail.get("invoice_document_id")
                    or legacy_detail.get("contract_document_id") != contract_id
                ):
                    raise AssertionError(f"legacy migration must create and link invoice document, got {legacy_detail_code} {legacy_detail}")
                after_docs_code, after_docs = _opener_json(operator, f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}")
                second_legacy_code, _ = _opener_json(operator, f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/sup_legacy_doc")
                final_docs_code, final_docs = _opener_json(operator, f"{base_url}{DEFAULT_TRADE_DOCUMENTS_PATH}")
                if (
                    after_docs_code != 200
                    or second_legacy_code != 200
                    or final_docs_code != 200
                    or len(after_docs.get("documents", [])) != before_count + 1
                    or len(final_docs.get("documents", [])) != len(after_docs.get("documents", []))
                ):
                    raise AssertionError("legacy migration must be idempotent")
                delete_legacy_code, _ = _opener_delete_json(
                    operator,
                    f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/sup_legacy_doc",
                )
                if delete_legacy_code != 200 or not legacy_file.exists():
                    raise AssertionError("supplier shipment delete must not remove physical legacy invoice file")
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=5)
    print("sheet_vitrina_v1_trade_documents_smoke: OK")


class _Clock:
    def __init__(self) -> None:
        self.value = 0

    def next(self) -> str:
        self.value += 1
        return f"2026-05-30T08:{self.value:02d}:00Z"


def _build_invoice_fixture() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"
    sheet.append(["Invoice No:", "26GN390"])
    sheet.append(["Invoice Date:", "14.5.2026"])
    sheet.append(["Contract No.", "CNT-2026-0513"])
    sheet.append(["Date of Contract", "2026.5.13"])
    sheet.append(["Supplier:", "HanShang Technology", "", "Currency:", "RMB"])
    sheet.append(["Invoice Total:", 15])
    sheet.append(["NO.", "NAME & SPECIFICATION", "MODELS", "QTY", "U.PRICE", "AMOUNT", "COMMENT"])
    sheet.append([1, "高清膜 smk", "iPhone 14 Pro", 10, 1, 10, ""])
    sheet.append([2, "OPP bag packets", "", 100, 0.05, 5, "OPP packets"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_contract_xlsx_fixture(number: str, date_text: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contract"
    sheet.append([f"Contract No. {number}"])
    sheet.append(["Contract Date", date_text])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_chinese_contract_xlsx_fixture(number: str, date_text: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contract"
    sheet.append([f"合同编号：{number}"])
    sheet.append([date_text])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_runtime_contract_file(runtime_dir: Path, *, document_id: str, filename: str, body: bytes) -> str:
    path = runtime_dir / "trade_documents" / "files" / "contract" / document_id / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(body)
    return path.relative_to(runtime_dir).as_posix()


def _build_contract_pdf_fixture(number: str, date_line: str) -> bytes:
    stream = (
        "BT\n"
        "/F1 12 Tf\n"
        "72 760 Td\n"
        f"(Contract No. {number}) Tj\n"
        "0 -18 Td\n"
        f"({date_line}) Tj\n"
        "ET\n"
    ).encode("utf-8")
    return (
        b"%PDF-1.4\n"
        b"1 0 obj\n"
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\n"
        b"stream\n" + stream + b"endstream\n"
        b"endobj\n"
        b"%%EOF\n"
    )


def _assert_contract_parser_rules() -> None:
    number_cases = {
        "Contract No. CNT-2026-0513": "CNT-2026-0513",
        "合同编号：CN-2026-0910": "CN-2026-0910",
        "№ 223/26": "223/26",
        "KOHTPAKT Nb 223/26 CONTRACT J 223/26": "223/26",
    }
    for line, expected in number_cases.items():
        got = supplier_shipments_module._extract_contract_number(line, [line])
        if got != expected:
            raise AssertionError(f"contract number rule failed for {line!r}: {got!r} != {expected!r}")
    date_cases = {
        "Contract Date 2026-05-13": "2026-05-13",
        "合同日期 2026年5月13日": "2026-05-13",
        "Date: 2026/05/13": "2026-05-13",
        "Date: May 13, 2026": "2026-05-13",
        "Date: 13 May 2026": "2026-05-13",
    }
    for text, expected in date_cases.items():
        got = supplier_shipments_module._extract_contract_document_date(text)
        if got != expected:
            raise AssertionError(f"contract date rule failed for {text!r}: {got!r} != {expected!r}")


def _assert_contract_ocr_parser_with_fake_tools() -> None:
    with TemporaryDirectory(prefix="trade-documents-fake-ocr-") as tmp:
        tmp_path = Path(tmp)
        bin_dir = tmp_path / "bin"
        bin_dir.mkdir()
        _write_executable(
            bin_dir / "pdfinfo",
            "#!/usr/bin/env python3\n"
            "print('Pages: 1')\n"
            "print('Page size: 1526 x 2190 pts')\n",
        )
        _write_executable(
            bin_dir / "pdftoppm",
            "#!/usr/bin/env python3\n"
            "import pathlib, sys\n"
            "pathlib.Path(sys.argv[-1] + '.png').write_bytes(b'fake-image')\n",
        )
        _write_executable(
            bin_dir / "tesseract",
            "#!/usr/bin/env python3\n"
            "import sys\n"
            "if '--list-langs' in sys.argv:\n"
            "    print('List of available languages in fake tessdata (3):')\n"
            "    print('eng')\n"
            "    print('chi_sim')\n"
            "    print('rus')\n"
            "    raise SystemExit(0)\n"
            "psm = sys.argv[sys.argv.index('--psm') + 1] if '--psm' in sys.argv else ''\n"
            "if psm == '6':\n"
            "    raise SystemExit(0)\n"
            "print('KOHTPAKT Nb 223/26 CONTRACT J 223/26')\n"
            "print('Date: May 13, 2026')\n",
        )
        supplier_shipments_module._TESSERACT_LANGUAGES_CACHE = None
        with _patched_env({"PATH": str(bin_dir) + os.pathsep + os.environ.get("PATH", "")}):
            block = SupplierShipmentsBlock(runtime=RegistryUploadDbBackedRuntime(runtime_dir=tmp_path / "runtime"))
            metadata, warnings, errors, version = block._parse_contract_document_metadata(
                b"%PDF-1.3\n% fake scanned contract\n",
                filename="scan-contract.pdf",
            )
        supplier_shipments_module._TESSERACT_LANGUAGES_CACHE = None
        diagnostics = metadata.get("diagnostics") if isinstance(metadata.get("diagnostics"), dict) else {}
        if (
            metadata.get("parsed_number") != "223/26"
            or metadata.get("parsed_document_date") != "2026-05-13"
            or diagnostics.get("ocr_available") is not True
            or diagnostics.get("ocr_languages") != ["eng", "chi_sim", "rus"]
            or "psm11" not in str(diagnostics.get("ocr_strategy_used") or "")
            or warnings
            or errors
            or version != TRADE_DOCUMENT_CONTRACT_PARSER_VERSION
        ):
            raise AssertionError(f"fake OCR contract parser failed: {metadata} {warnings} {errors} {version}")


def _assert_contract_ocr_missing_tools_warning() -> None:
    with TemporaryDirectory(prefix="trade-documents-no-ocr-") as tmp:
        tmp_path = Path(tmp)
        supplier_shipments_module._TESSERACT_LANGUAGES_CACHE = None
        with _patched_env({"PATH": str(tmp_path)}):
            block = SupplierShipmentsBlock(runtime=RegistryUploadDbBackedRuntime(runtime_dir=tmp_path / "runtime"))
            metadata, warnings, errors, _ = block._parse_contract_document_metadata(
                b"%PDF-1.3\n% fake scanned contract\n",
                filename="scan-contract.pdf",
            )
        supplier_shipments_module._TESSERACT_LANGUAGES_CACHE = None
        warning_text = " ".join(warnings)
        diagnostics = metadata.get("diagnostics") if isinstance(metadata.get("diagnostics"), dict) else {}
        if diagnostics.get("ocr_available") is not False or "OCR tools missing" not in warning_text or errors:
            raise AssertionError(f"missing OCR tools must produce controlled warning, got {metadata} {warnings} {errors}")


def _write_executable(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")
    path.chmod(0o755)


def _login(opener: urllib_request.OpenerDirector, base_url: str, username: str, password: str) -> None:
    body = urllib_parse.urlencode({"username": username, "password": password, "next": "/sheet-vitrina-v1/supplier"}).encode("utf-8")
    request = urllib_request.Request(
        f"{base_url}/login",
        data=body,
        headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "text/html"},
        method="POST",
    )
    with opener.open(request, timeout=5) as response:
        response.read()


def _opener_post_multipart(
    opener: urllib_request.OpenerDirector,
    url: str,
    body_bytes: bytes,
    *,
    filename: str,
    content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    fields: dict[str, str] | None = None,
) -> tuple[int, dict[str, object]]:
    boundary = "----wbcore-trade-documents" + uuid4().hex
    chunks: list[bytes] = []
    for key, value in (fields or {}).items():
        chunks.extend(
            [
                f"--{boundary}\r\n".encode("utf-8"),
                f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"),
                str(value).encode("utf-8"),
                b"\r\n",
            ]
        )
    chunks.extend(
        [
            f"--{boundary}\r\n".encode("utf-8"),
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'.encode("utf-8"),
            f"Content-Type: {content_type}\r\n\r\n".encode("utf-8"),
            body_bytes,
            f"\r\n--{boundary}--\r\n".encode("utf-8"),
        ]
    )
    request = urllib_request.Request(
        url,
        data=b"".join(chunks),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}", "Accept": "application/json"},
        method="POST",
    )
    return _opener_json_request(opener, request)


def _opener_post_json(opener: urllib_request.OpenerDirector, url: str, payload: dict[str, object]) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8", "Accept": "application/json"},
        method="POST",
    )
    return _opener_json_request(opener, request)


def _opener_patch_json(opener: urllib_request.OpenerDirector, url: str, payload: dict[str, object]) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8", "Accept": "application/json"},
        method="PATCH",
    )
    return _opener_json_request(opener, request)


def _opener_delete_json(opener: urllib_request.OpenerDirector, url: str) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(url, headers={"Accept": "application/json"}, method="DELETE")
    return _opener_json_request(opener, request)


def _opener_json(opener: urllib_request.OpenerDirector, url: str) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(url, headers={"Accept": "application/json"}, method="GET")
    return _opener_json_request(opener, request)


def _opener_json_request(opener: urllib_request.OpenerDirector, request: urllib_request.Request) -> tuple[int, dict[str, object]]:
    try:
        with opener.open(request, timeout=5) as response:
            return response.status, json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        return exc.code, json.loads(exc.read().decode("utf-8"))


def _opener_bytes(opener: urllib_request.OpenerDirector, url: str) -> tuple[int, bytes, dict[str, str]]:
    request = urllib_request.Request(url, headers={"Accept": "*/*"}, method="GET")
    try:
        with opener.open(request, timeout=5) as response:
            return response.status, response.read(), dict(response.headers.items())
    except urllib_error.HTTPError as exc:
        return exc.code, exc.read(), dict(exc.headers.items())


def _reserve_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


@contextmanager
def _patched_env(values: dict[str, str]):
    previous = {key: os.environ.get(key) for key in values}
    os.environ.update(values)
    try:
        yield
    finally:
        for key, value in previous.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


def _password_hash(password: str) -> str:
    salt = b"trade-documents-smoke-salt"
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 100_000)
    return "pbkdf2_sha256$100000$" + base64.b64encode(salt).decode("ascii") + "$" + base64.b64encode(digest).decode("ascii")


if __name__ == "__main__":
    main()
