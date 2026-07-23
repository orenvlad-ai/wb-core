"""Bounded fallback for one new simple tabular XLSX outside the repository.

The primary path remains the active Spreadsheets skill with
``@oai/artifact-tool``.  This helper never edits an existing workbook because a
fallback rewrite could lose formulas, charts, styles, or relationships.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
import json
import math
import os
from pathlib import Path
import sys
import tempfile
from typing import Any, Iterable, Mapping
from xml.etree import ElementTree as ET
import zipfile


REPO_ROOT = Path(__file__).resolve().parents[1]
FALLBACK_BACKENDS = ("openpyxl", "xlsxwriter", "stdlib")
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
REQUIRED_MEMBERS = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/styles.xml",
    }
)


def xlsx_creation_plan(
    *, artifact_tool_import_error: BaseException | None
) -> tuple[str, ...]:
    """Keep the primary path on success; expose the bounded fallback after failure."""

    return ("artifact-tool",) if artifact_tool_import_error is None else FALLBACK_BACKENDS


@dataclass(frozen=True)
class TabularXlsxSpec:
    sheet_name: str
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]
    text_columns: frozenset[int]
    freeze_header: bool = True
    auto_filter: bool = True
    column_widths: tuple[float, ...] = ()

    @classmethod
    def from_mapping(cls, payload: Mapping[str, Any]) -> "TabularXlsxSpec":
        spec = cls(
            sheet_name=str(payload.get("sheet_name") or "Sheet1").strip(),
            headers=tuple(str(value).strip() for value in payload.get("headers") or []),
            rows=tuple(tuple(row) for row in payload.get("rows") or []),
            text_columns=frozenset(int(value) for value in payload.get("text_columns") or []),
            freeze_header=bool(payload.get("freeze_header", True)),
            auto_filter=bool(payload.get("auto_filter", True)),
            column_widths=tuple(float(value) for value in payload.get("column_widths") or []),
        )
        spec.validate()
        return spec

    def validate(self) -> None:
        if not self.sheet_name or len(self.sheet_name) > 31 or any(
            char in self.sheet_name for char in "[]:*?/\\"
        ):
            raise ValueError("invalid Excel sheet_name")
        if not self.headers or any(not value for value in self.headers):
            raise ValueError("non-empty headers are required")
        width = len(self.headers)
        if any(len(row) != width for row in self.rows):
            raise ValueError("every row must match the header width")
        if any(index < 0 or index >= width for index in self.text_columns):
            raise ValueError("text_columns contains an invalid zero-based index")
        if self.column_widths and len(self.column_widths) != width:
            raise ValueError("column_widths must match the header width")
        if any(not math.isfinite(value) or value <= 0 for value in self.column_widths):
            raise ValueError("column widths must be finite and positive")
        for row in self.rows:
            for value in row:
                if value is not None and not isinstance(value, (str, int, float, bool)):
                    raise ValueError("cells support only string, number, boolean or null")
                if isinstance(value, float) and not math.isfinite(value):
                    raise ValueError("non-finite cell number")

    def widths(self) -> tuple[float, ...]:
        if self.column_widths:
            return self.column_widths
        return tuple(
            float(
                min(
                    80,
                    max(
                        10,
                        max(
                            [
                                len(self.headers[index]),
                                *[
                                    len("" if row[index] is None else str(row[index]))
                                    for row in self.rows
                                ],
                            ]
                        )
                        + 2,
                    ),
                )
            )
            for index in range(len(self.headers))
        )


def create_tabular_xlsx(
    output_path: str | Path,
    spec: TabularXlsxSpec,
    *,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Write the dependency-free final fallback atomically outside the repo."""

    spec.validate()
    output = Path(output_path).expanduser().resolve(strict=False)
    if output.suffix.casefold() != ".xlsx":
        raise ValueError("output path must end with .xlsx")
    if output == REPO_ROOT or REPO_ROOT in output.parents:
        raise ValueError("user-artifact output must be outside the repository")
    if output.exists() and not overwrite:
        raise FileExistsError(f"refusing to overwrite {output}")
    output.parent.mkdir(parents=True, exist_ok=True)

    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.name}.", suffix=".xlsx", dir=output.parent
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        _write_stdlib(temporary, spec)
        verification = verify_tabular_xlsx(temporary, spec)
        os.replace(temporary, output)
        verify_tabular_xlsx(output, spec)
    finally:
        temporary.unlink(missing_ok=True)
    return {
        "path": str(output),
        "backend": "stdlib",
        **verification,
    }


def verify_tabular_xlsx(path: str | Path, spec: TabularXlsxSpec) -> dict[str, Any]:
    """Check ZIP/XML/shape/format and reopen with openpyxl when available."""

    source = Path(path)
    if not source.is_file() or source.stat().st_size <= 0:
        raise ValueError("XLSX is missing or empty")
    with zipfile.ZipFile(source) as archive:
        if archive.testzip() is not None:
            raise ValueError("XLSX ZIP integrity failed")
        missing = REQUIRED_MEMBERS - frozenset(archive.namelist())
        if missing:
            raise ValueError("missing OOXML members: " + ", ".join(sorted(missing)))
        parsed = {name: ET.fromstring(archive.read(name)) for name in REQUIRED_MEMBERS}
        sheets = parsed["xl/workbook.xml"].findall(
            f"{{{MAIN_NS}}}sheets/{{{MAIN_NS}}}sheet"
        )
        if [sheet.attrib.get("name") for sheet in sheets] != [spec.sheet_name]:
            raise ValueError("unexpected worksheets")
        worksheet = parsed["xl/worksheets/sheet1.xml"]
        rows = worksheet.findall(f"{{{MAIN_NS}}}sheetData/{{{MAIN_NS}}}row")
        if len(rows) != len(spec.rows) + 1 or len(rows[0]) != len(spec.headers):
            raise ValueError("unexpected worksheet dimensions")
        _verify_worksheet_values(rows, spec)
        if spec.freeze_header:
            pane = worksheet.find(
                f"{{{MAIN_NS}}}sheetViews/{{{MAIN_NS}}}sheetView/{{{MAIN_NS}}}pane"
            )
            if (
                pane is None
                or pane.attrib.get("state") != "frozen"
                or pane.attrib.get("ySplit") != "1"
                or pane.attrib.get("topLeftCell") != "A2"
            ):
                raise ValueError("frozen header is missing")
        if spec.auto_filter:
            auto_filter = worksheet.find(f"{{{MAIN_NS}}}autoFilter")
            if auto_filter is None or auto_filter.attrib.get("ref") != _table_ref(spec):
                raise ValueError("autofilter is missing")
        columns = worksheet.findall(f"{{{MAIN_NS}}}cols/{{{MAIN_NS}}}col")
        if len(columns) != len(spec.headers):
            raise ValueError("column widths are missing")
        for column, expected_width in zip(columns, spec.widths(), strict=True):
            if column.attrib.get("customWidth") != "1" or not math.isclose(
                float(column.attrib.get("width", "nan")), expected_width
            ):
                raise ValueError("unexpected column width")

    reader = "structural-ooxml"
    try:
        import openpyxl
    except ModuleNotFoundError:
        pass
    else:
        workbook = openpyxl.load_workbook(source, data_only=False, read_only=False)
        try:
            if workbook.sheetnames != [spec.sheet_name]:
                raise ValueError("independent reader found unexpected worksheets")
            sheet = workbook[spec.sheet_name]
            if sheet.max_row != len(spec.rows) + 1 or sheet.max_column != len(spec.headers):
                raise ValueError("independent reader found unexpected dimensions")
            actual_headers = tuple(
                sheet.cell(1, column_number).value
                for column_number in range(1, len(spec.headers) + 1)
            )
            if actual_headers != spec.headers:
                raise ValueError("independent reader found unexpected headers")
            for row_number, expected_row in enumerate(spec.rows, start=2):
                for column_number, expected in enumerate(expected_row, start=1):
                    cell = sheet.cell(row_number, column_number)
                    wanted = (
                        str(expected)
                        if column_number - 1 in spec.text_columns and expected is not None
                        else expected
                    )
                    if not (wanted == "" and cell.value is None) and cell.value != wanted:
                        raise ValueError(f"saved value mismatch in {cell.coordinate}")
                    if (
                        column_number - 1 in spec.text_columns
                        and expected not in (None, "")
                        and (cell.data_type != "s" or cell.number_format != "@")
                    ):
                        raise ValueError(f"identifier is not text in {cell.coordinate}")
        finally:
            workbook.close()
        reader = "openpyxl"
    return {
        "size_bytes": source.stat().st_size,
        "zip_integrity": "ok",
        "xml_integrity": "ok",
        "reader": reader,
        "sheet_names": [spec.sheet_name],
        "data_rows": len(spec.rows),
        "columns": len(spec.headers),
    }


def _write_stdlib(path: Path, spec: TabularXlsxSpec) -> None:
    members = {
        "[Content_Types].xml": _CONTENT_TYPES,
        "_rels/.rels": _ROOT_RELS,
        "xl/_rels/workbook.xml.rels": _WORKBOOK_RELS,
        "xl/styles.xml": _STYLES,
        "xl/workbook.xml": _workbook_xml(spec.sheet_name),
        "xl/worksheets/sheet1.xml": _worksheet_xml(spec),
    }
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in members.items():
            archive.writestr(name, content)


def _verify_worksheet_values(rows: list[ET.Element], spec: TabularXlsxSpec) -> None:
    expected_rows: tuple[tuple[Any, ...], ...] = (spec.headers, *spec.rows)
    for row_number, (row, expected_row) in enumerate(
        zip(rows, expected_rows, strict=True), start=1
    ):
        cells = {cell.attrib.get("r"): cell for cell in row.findall(f"{{{MAIN_NS}}}c")}
        for index, expected in enumerate(expected_row):
            coordinate = f"{_column_name(index + 1)}{row_number}"
            cell = cells.get(coordinate)
            if expected is None:
                if cell is not None:
                    raise ValueError(f"unexpected value in {coordinate}")
                continue
            if cell is None:
                raise ValueError(f"missing value in {coordinate}")
            text_value = row_number == 1 or index in spec.text_columns or isinstance(expected, str)
            if text_value:
                if cell.attrib.get("t") != "inlineStr":
                    raise ValueError(f"text value is not stored as text in {coordinate}")
                actual = "".join(
                    node.text or "" for node in cell.findall(f".//{{{MAIN_NS}}}t")
                )
                if actual != str(expected):
                    raise ValueError(f"saved value mismatch in {coordinate}")
                if row_number > 1 and index in spec.text_columns and cell.attrib.get("s") != "2":
                    raise ValueError(f"identifier format is missing in {coordinate}")
            else:
                actual = cell.findtext(f"{{{MAIN_NS}}}v")
                wanted = (
                    "1"
                    if expected is True
                    else "0"
                    if expected is False
                    else str(expected)
                )
                expected_type = "b" if isinstance(expected, bool) else "n"
                if cell.attrib.get("t") != expected_type or actual != wanted:
                    raise ValueError(f"saved value mismatch in {coordinate}")


def _workbook_xml(sheet_name: str) -> bytes:
    ET.register_namespace("", MAIN_NS)
    ET.register_namespace("r", REL_NS)
    root = ET.Element(f"{{{MAIN_NS}}}workbook")
    sheets = ET.SubElement(root, f"{{{MAIN_NS}}}sheets")
    sheet = ET.SubElement(sheets, f"{{{MAIN_NS}}}sheet", name=sheet_name, sheetId="1")
    sheet.set(f"{{{REL_NS}}}id", "rId1")
    return _xml(root)


def _worksheet_xml(spec: TabularXlsxSpec) -> bytes:
    ET.register_namespace("", MAIN_NS)
    root = ET.Element(f"{{{MAIN_NS}}}worksheet")
    views = ET.SubElement(root, f"{{{MAIN_NS}}}sheetViews")
    view = ET.SubElement(views, f"{{{MAIN_NS}}}sheetView", workbookViewId="0")
    if spec.freeze_header:
        ET.SubElement(view, f"{{{MAIN_NS}}}pane", ySplit="1", topLeftCell="A2", state="frozen")
    columns = ET.SubElement(root, f"{{{MAIN_NS}}}cols")
    for index, width in enumerate(spec.widths(), start=1):
        ET.SubElement(columns, f"{{{MAIN_NS}}}col", min=str(index), max=str(index), width=str(width), customWidth="1")
    sheet_data = ET.SubElement(root, f"{{{MAIN_NS}}}sheetData")
    _xml_row(sheet_data, 1, spec.headers, frozenset(), header=True)
    for row_number, row in enumerate(spec.rows, start=2):
        _xml_row(sheet_data, row_number, row, spec.text_columns, header=False)
    if spec.auto_filter:
        ET.SubElement(root, f"{{{MAIN_NS}}}autoFilter", ref=_table_ref(spec))
    return _xml(root)


def _xml_row(parent: ET.Element, number: int, values: tuple[Any, ...], text_columns: frozenset[int], *, header: bool) -> None:
    row = ET.SubElement(parent, f"{{{MAIN_NS}}}row", r=str(number))
    for index, raw in enumerate(values):
        if raw is None:
            continue
        style = "1" if header else ("2" if index in text_columns else "0")
        value = str(raw) if index in text_columns else raw
        cell = ET.SubElement(row, f"{{{MAIN_NS}}}c", r=f"{_column_name(index + 1)}{number}", s=style)
        if isinstance(value, str):
            cell.set("t", "inlineStr")
            inline = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
            text = ET.SubElement(inline, f"{{{MAIN_NS}}}t")
            if value != value.strip():
                text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            text.text = value
        else:
            cell.set("t", "b" if isinstance(value, bool) else "n")
            ET.SubElement(cell, f"{{{MAIN_NS}}}v").text = (
                "1" if value is True else "0" if value is False else str(value)
            )


def _column_name(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _table_ref(spec: TabularXlsxSpec) -> str:
    return f"A1:{_column_name(len(spec.headers))}{len(spec.rows) + 1}"


def _xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


_CONTENT_TYPES = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>'''
_ROOT_RELS = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''
_WORKBOOK_RELS = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''
_STYLES = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><color rgb="FFFFFFFF"/><sz val="11"/><name val="Calibri"/></font></fonts>
<fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF1F4E78"/><bgColor indexed="64"/></patternFill></fill></fills>
<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
<cellXfs count="3"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1"/><xf numFmtId="49" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/></cellXfs>
<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>'''


def _load_spec(path: str) -> TabularXlsxSpec:
    payload = json.load(sys.stdin) if path == "-" else json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(payload, Mapping):
        raise ValueError("input JSON must be an object")
    return TabularXlsxSpec.from_mapping(payload)


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="JSON spec path or - for stdin")
    parser.add_argument("--output", required=True)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args(argv)
    result = create_tabular_xlsx(
        args.output,
        _load_spec(args.input),
        overwrite=args.overwrite,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
