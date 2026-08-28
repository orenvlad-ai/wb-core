#!/usr/bin/env python3
"""Focused smoke for immutable inventory-balance calculations and dry-run jobs."""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import json
from pathlib import Path
import socket
import sqlite3
import sys
import tempfile
import threading
import time
from types import SimpleNamespace
from urllib import request as urllib_request
import zipfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.application.sku_inventory_balance import (  # noqa: E402
    LiveWbInventoryBalanceApplyAdapter,
    SkuInventoryBalanceBlock,
    SkuInventoryBalanceError,
    calculate_inventory_balance_row,
)
from packages.application.sku_management import (  # noqa: E402
    SkuManagementBlock,
    _inventory_balance_aggregate_wb_evidence,
    _inventory_balance_wb_stock_from_row,
)
from packages.adapters.registry_upload_http_entrypoint import (  # noqa: E402
    DEFAULT_SHEET_OPERATOR_UI_PATH,
    DEFAULT_SHEET_PLAN_PATH,
    DEFAULT_SHEET_REFRESH_PATH,
    DEFAULT_SHEET_STATUS_PATH,
    DEFAULT_SKU_INVENTORY_BALANCE_CALCULATE_PATH,
    DEFAULT_SKU_INVENTORY_BALANCE_OPERATIONS_PATH,
    DEFAULT_SKU_INVENTORY_BALANCE_PATH,
    DEFAULT_UPLOAD_PATH,
    build_registry_upload_http_server,
)
from packages.contracts.registry_upload_http_entrypoint import (  # noqa: E402
    RegistryUploadHttpEntrypointConfig,
)


class FakeRuntime:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.runtime_dir = db_path.parent
        sqlite3.connect(db_path).close()
        self.configs: dict[tuple[str, str], dict] = {}

    def load_sheet_vitrina_user_config(self, *, user_key: str, config_key: str) -> dict:
        return self.configs.get((user_key, config_key), {"status": "missing", "revision": 0})

    def save_sheet_vitrina_user_config(
        self,
        *,
        user_key: str,
        config_key: str,
        schema_version: int,
        payload: dict,
        updated_at: str,
        expected_revision: int | None,
    ) -> dict:
        current = self.configs.get((user_key, config_key), {"revision": 0})
        if expected_revision is not None and int(expected_revision) != int(current.get("revision") or 0):
            return {"status": "conflict", "revision": int(current.get("revision") or 0)}
        row = {
            "status": "ok",
            "revision": int(current.get("revision") or 0) + 1,
            "schema_version": schema_version,
            "config": dict(payload),
            "updated_at": updated_at,
        }
        self.configs[(user_key, config_key)] = row
        return row


class FakeAdsBlock:
    def build_sku_detail(self, nm_id: int, params: dict | None = None) -> dict:
        del params
        return {
            "rows": [
                {
                    "nm_id": nm_id,
                    "advert_id": 9001,
                    "campaign_name": "new cpc recommendations",
                    "payment_type": "cpc",
                    "placement": "recommendations",
                    "status": 9,
                    "current_bid_rub": 5,
                    "min_bid_rub": 2,
                    "orders": 10,
                    "spend_rub": 300,
                },
                {
                    "nm_id": nm_id,
                    "advert_id": 8001,
                    "campaign_name": "old cpm search",
                    "payment_type": "cpm",
                    "placement": "search",
                    "status": 9,
                    "current_bid_rub": 1000,
                    "min_bid_rub": 300,
                    "orders": 5,
                    "spend_rub": 400,
                },
            ],
            "meta": {"stats_status": "ok", "stats_scope": "campaign_sku_aggregate"},
        }


class FakeSkuManagement:
    def __init__(self) -> None:
        self.ads_block = FakeAdsBlock()
        self.preview_calls = 0
        self.commit_calls = 0
        self.balance_sales_period_requests: list[int] = []

    def build_table(self, *, user_key: str) -> dict:
        del user_key
        start = datetime(2026, 8, 26, tzinfo=timezone.utc)

        def timeline(opening: float, demand: float, inbound_day: int, inbound: float) -> list[dict]:
            result = []
            balance = opening
            for index in range(1, 21):
                quantity = inbound if index == inbound_day else 0.0
                balance += quantity - demand
                result.append(
                    {
                        "date": (start + timedelta(days=index - 1)).date().isoformat(),
                        "inbound_qty": quantity,
                        "inbound_sources": ["supplier_shipment"] if quantity else [],
                        "demand_qty": demand,
                        "ending_stock": balance,
                    }
                )
            return result

        return {
            "contract_name": "sheet_vitrina_v1_sku_management_table",
            "generated_at": "2026-08-26T08:00:00+00:00",
            "meta": {"metric_policy": {"business_date": "2026-08-26"}},
            "rows": [
                {
                    "nm_id": 101,
                    "name": "Deficit SKU",
                    "our_sku": "DEF",
                    "stock_wb": 100,
                    "stock_ff": 0,
                    "daily_demand": 10,
                    "quality": "complete",
                    "quality_warnings": [],
                    "timeline": timeline(100, 10, 10, 100),
                    "inventory_balance_as_of_date": "2026-08-26",
                    "inventory_balance_inbounds": [
                        {
                            "date": "2026-09-04",
                            "quantity": 100,
                            "source": "supplier_shipment",
                            "source_id": "shipment-1",
                            "consumes_current_ff": False,
                        }
                    ],
                },
                {
                    "nm_id": 202,
                    "name": "Excess SKU",
                    "our_sku": "EXC",
                    "stock_wb": 1000,
                    "stock_ff": 0,
                    "daily_demand": 10,
                    "quality": "complete",
                    "quality_warnings": [],
                    "timeline": timeline(1000, 10, 20, 0),
                    "inventory_balance_as_of_date": "2026-08-26",
                    "inventory_balance_inbounds": [
                        {
                            "date": "2026-09-14",
                            "quantity": 100,
                            "source": "supplier_shipment",
                            "source_id": "shipment-2",
                            "consumes_current_ff": False,
                        }
                    ],
                },
                {
                    "nm_id": 303,
                    "name": "No supply SKU",
                    "our_sku": "NOSUP",
                    "stock_wb": 100,
                    "stock_ff": 20,
                    "daily_demand": 5,
                    "quality": "complete",
                    "quality_warnings": [],
                    "timeline": timeline(100, 5, 20, 0),
                    "inventory_balance_as_of_date": "2026-08-26",
                    "inventory_balance_inbounds": [],
                },
                *[
                    {
                        "nm_id": nm_id,
                        "name": "Explicitly excluded glass",
                        "our_sku": f"AIR-{nm_id}",
                        "stock_wb": 50,
                        "stock_ff": 20,
                        "daily_demand": 1,
                        "quality": "complete",
                        "quality_warnings": [],
                        "timeline": timeline(50, 1, 3, 20),
                        "inventory_balance_as_of_date": "2026-08-26",
                        "inventory_balance_inbounds": [],
                    }
                    for nm_id in (497413772, 497415593, 497416931)
                ],
            ],
        }

    def build_inventory_balance_evidence(self, *, user_key: str, sales_period_days: int) -> dict:
        source = self.build_table(user_key=user_key)
        self.balance_sales_period_requests.append(int(sales_period_days))
        for row in source["rows"]:
            if int(row["nm_id"]) in {101, 202}:
                row["daily_demand"] = 10 if int(sales_period_days) == 7 else 20
                row["demand_evidence"] = {
                    "sales_avg_period_days": int(sales_period_days),
                    "daily_demand": row["daily_demand"],
                    "quality": "complete",
                }
        source["contract_name"] = "sheet_vitrina_v1_sku_inventory_balance_evidence/v2"
        source["meta"]["inventory_balance_evidence"] = {
            "sales_period_days": int(sales_period_days),
            "sales_lookup_days": int(sales_period_days) * 4,
            "date_from": "2026-07-29" if int(sales_period_days) == 7 else "2026-07-01",
            "date_to": "2026-08-25",
            "demand_mode": "availability_adjusted",
            "supplier_eta": {
                "method": "empirical_last_completed_shipments",
                "quality": "complete",
                "sample_count": 3,
                "samples": [
                    {"shipment_id": "done-1", "lead_days": 20},
                    {"shipment_id": "done-2", "lead_days": 22},
                    {"shipment_id": "done-3", "lead_days": 21},
                ],
                "mean_days_exact": 21,
                "applied_days": 21,
            },
        }
        return source

    def preview_bid(self, payload: dict, *, actor: str) -> dict:
        del payload, actor
        self.preview_calls += 1
        return {"preview": {"preview_id": "preview"}}

    def commit_bid(self, payload: dict, *, actor: str) -> dict:
        del payload, actor
        self.commit_calls += 1
        return {"status": "success"}


class FakeInventoryBalanceHttpEntrypoint:
    def __init__(self, block: SkuInventoryBalanceBlock, runtime: FakeRuntime) -> None:
        self.sku_inventory_balance_block = block
        self.runtime = runtime

    def handle_sku_inventory_balance_calculate_request(
        self,
        payload: dict,
        *,
        user_key: str,
        actor: str,
    ) -> dict:
        return self.sku_inventory_balance_block.start_calculation_operation(
            payload,
            user_key=user_key,
            actor=actor,
        )

    def handle_sku_inventory_balance_operation_request(
        self,
        operation_id: str,
        *,
        user_key: str,
    ) -> dict:
        return self.sku_inventory_balance_block.get_calculation_operation(
            operation_id,
            user_key=user_key,
        )

    def handle_sku_inventory_balance_request(self, *, user_key: str) -> dict:
        return self.sku_inventory_balance_block.latest(user_key=user_key)


class FakeBalanceSalesHistory:
    def load_order_count_samples_by_date(self, **kwargs) -> dict:
        del kwargs
        samples = []
        for offset in range(14, 0, -1):
            observed = date(2026, 8, 26) - timedelta(days=offset)
            samples.append((observed.isoformat(), 20.0 if offset > 7 else 10.0))
        return {101: samples}


class FakeBalanceEvidenceRuntime:
    def __init__(self) -> None:
        self.details = {
            "done-10": {"header": {"actual_shipment_date": "2026-06-01", "actual_ff_acceptance_date": "2026-06-11"}, "lines": []},
            "done-12": {"header": {"actual_shipment_date": "2026-06-10", "actual_ff_acceptance_date": "2026-06-22"}, "lines": []},
            "done-14": {"header": {"actual_shipment_date": "2026-07-01", "actual_ff_acceptance_date": "2026-07-15"}, "lines": []},
            "done-16": {"header": {"actual_shipment_date": "2026-07-20", "actual_ff_acceptance_date": "2026-08-05"}, "lines": []},
            "transit": {
                "header": {"order_status": "in_transit", "actual_shipment_date": "2026-08-20"},
                "lines": [{"line_type": "product", "match_status": "matched_by_barcode", "internal_nm_id": 101, "qty": 30}],
            },
            "production": {
                "header": {"order_status": "production", "planned_shipment_date": "2026-08-25"},
                "lines": [{"line_type": "product", "match_status": "matched_by_barcode", "internal_nm_id": 101, "qty": 40}],
            },
        }

    def list_supplier_shipments(self) -> list[dict]:
        return [{"shipment_id": key} for key in self.details]

    def load_supplier_shipment(self, shipment_id: str) -> dict:
        return self.details[shipment_id]


def _aggregate_wb_balance_fallback_checks() -> None:
    digest = "sha256:" + "a" * 64
    aggregate_result = SimpleNamespace(
        kind="success",
        snapshot_date="2026-08-26",
        fetched_at="2026-08-26T08:00:00Z",
        pagination_complete=True,
        raw_rows_digest=digest,
        warehouse_granularity_complete=False,
        count=1,
        items=[SimpleNamespace(nm_id=101, stock_total=80.0)],
        warehouse_rows=[],
    )
    aggregate = _inventory_balance_aggregate_wb_evidence(
        aggregate_result,
        requested_nm_ids=[101],
        expected_snapshot_date="2026-08-26",
    )
    collector = object.__new__(SkuManagementBlock)
    collector.now_factory = lambda: datetime(2026, 8, 26, 8, tzinfo=timezone.utc)
    collector.stocks_block = SimpleNamespace(
        execute=lambda request: SimpleNamespace(result=aggregate_result)
    )
    collector.runtime = SimpleNamespace(
        load_ff_stock_activation_operation=lambda: {"status": "active"},
        list_ff_stock_balances=lambda: [{"nm_id": 101, "quantity": 10}],
    )
    collector.sales_history = None
    collector._append_supplier_inbounds = lambda result, settings: None
    collector._append_factory_order_inbounds = lambda result, settings: None
    collector._append_wb_supply_inbounds = lambda result, settings: None
    collector._append_regional_demand = lambda result, settings, as_of_date: None
    collected = collector._collect_forecast_evidence(
        active=[{"nm_id": 101}],
        settings=SimpleNamespace(sales_avg_period_days=7),
    )
    general_row = collected[101]
    assert general_row["stock_wb"] is None
    assert general_row["districts"] == {}
    assert general_row["inventory_balance_wb_stock_evidence"] == aggregate[101]
    selected_stock, selected_evidence, warning = _inventory_balance_wb_stock_from_row(
        general_row,
        expected_snapshot_date="2026-08-26",
    )
    assert selected_stock == 80
    assert selected_evidence["mode"] == "aggregate_per_sku_total"
    assert selected_evidence["incident_projection_applied"] is False
    assert selected_evidence["used_by_inventory_balance"] is True
    assert "без раскладки" in warning

    exact_zero_result = SimpleNamespace(
        **{
            **vars(aggregate_result),
            "items": [SimpleNamespace(nm_id=101, stock_total=0.0)],
        }
    )
    exact_zero = _inventory_balance_aggregate_wb_evidence(
        exact_zero_result,
        requested_nm_ids=[101],
        expected_snapshot_date="2026-08-26",
    )
    assert exact_zero[101]["stock_wb_units"] == 0.0

    aggregate_row = calculate_inventory_balance_row(
        {
            "nm_id": 101,
            "name": "Aggregate-only WB SKU",
            "stock_wb": selected_stock,
            "stock_ff": 10,
            "daily_demand": 8,
            "quality": "complete",
            "quality_warnings": [warning],
            "wb_stock_evidence": selected_evidence,
            "inventory_balance_as_of_date": "2026-08-26",
            "inventory_balance_inbounds": [
                {
                    "date": "2026-09-04",
                    "quantity": 30,
                    "source": "supplier_shipment",
                    "source_id": "aggregate-next",
                    "consumes_current_ff": False,
                },
                {
                    "date": "2026-09-14",
                    "quantity": 40,
                    "source": "supplier_shipment",
                    "source_id": "aggregate-subsequent",
                    "consumes_current_ff": False,
                },
            ],
        },
        settings={
            "wb_confidence_coefficient": 0.25,
            "safety_stock_days": 10,
            "bid_scale_max": 2.0,
        },
    )
    assert aggregate_row["known_stock_units"] == 30
    assert aggregate_row["confidence_adjusted_wb_units"] == 20
    assert aggregate_row["milestones"][0]["available_before_arrival"] == 30
    assert aggregate_row["next_inbound"]["source_ids"] == ["aggregate-next"]
    assert aggregate_row["subsequent_inbound"]["source_ids"] == [
        "aggregate-subsequent"
    ]
    assert aggregate_row["target_daily_sales"] is not None
    assert aggregate_row["pace_change_pct"] is not None
    assert aggregate_row["days_cover"] == 3.75
    assert aggregate_row["wb_stock_evidence"]["wb_confidence_coefficient"] == 0.25

    zero_coefficient = calculate_inventory_balance_row(
        {
            "nm_id": 101,
            "name": "Zero WB coefficient",
            "stock_wb": selected_stock,
            "stock_ff": 10,
            "daily_demand": 8,
            "quality": "complete",
            "quality_warnings": [warning],
            "wb_stock_evidence": selected_evidence,
            "inventory_balance_as_of_date": "2026-08-26",
            "inventory_balance_inbounds": [
                {
                    "date": "2026-09-04",
                    "quantity": 30,
                    "source": "supplier_shipment",
                    "source_id": "aggregate-next-zero",
                    "consumes_current_ff": False,
                }
            ],
        },
        settings={
            "wb_confidence_coefficient": 0.0,
            "safety_stock_days": 10,
            "bid_scale_max": 2.0,
        },
    )
    assert zero_coefficient["known_stock_units"] == 10
    assert zero_coefficient["confidence_adjusted_wb_units"] == 0

    invalid_results = [
        SimpleNamespace(**{**vars(aggregate_result), "pagination_complete": False}),
        SimpleNamespace(**{**vars(aggregate_result), "snapshot_date": "2026-08-25"}),
        SimpleNamespace(**{**vars(aggregate_result), "raw_rows_digest": "sha256:invalid"}),
        SimpleNamespace(**{**vars(aggregate_result), "count": True}),
        SimpleNamespace(**{**vars(aggregate_result), "count": 0, "items": []}),
        SimpleNamespace(
            **{
                **vars(aggregate_result),
                "items": [SimpleNamespace(nm_id=101, stock_total=float("nan"))],
            }
        ),
        SimpleNamespace(
            **{
                **vars(aggregate_result),
                "items": [SimpleNamespace(nm_id=101, stock_total=float("inf"))],
            }
        ),
        SimpleNamespace(
            **{
                **vars(aggregate_result),
                "items": [SimpleNamespace(nm_id=101, stock_total=-1.0)],
            }
        ),
        SimpleNamespace(
            **{
                **vars(aggregate_result),
                "count": 2,
                "items": [
                    SimpleNamespace(nm_id=101, stock_total=80.0),
                    SimpleNamespace(nm_id=101, stock_total=0.0),
                ],
            }
        ),
    ]
    for invalid in invalid_results:
        try:
            _inventory_balance_aggregate_wb_evidence(
                invalid,
                requested_nm_ids=(
                    [101, 202]
                    if getattr(invalid, "count", None) == 2
                    else [101]
                ),
                expected_snapshot_date="2026-08-26",
            )
        except ValueError:
            pass
        else:  # pragma: no cover
            raise AssertionError("partial/missing/malformed aggregate WB evidence was accepted")

    malformed_row = {
        "stock_wb": None,
        "inventory_balance_wb_stock_evidence": {
            **aggregate[101],
            "raw_rows_digest": "",
        },
    }
    missing_stock, malformed_evidence, malformed_warning = (
        _inventory_balance_wb_stock_from_row(
            malformed_row,
            expected_snapshot_date="2026-08-26",
        )
    )
    assert missing_stock is None
    assert malformed_evidence["quality"] == "invalid_aggregate_evidence"
    assert "остаётся неизвестным" in malformed_warning


def main() -> None:
    _aggregate_wb_balance_fallback_checks()
    evidence_block = object.__new__(SkuManagementBlock)
    evidence_block.sales_history = FakeBalanceSalesHistory()
    evidence_block.runtime = FakeBalanceEvidenceRuntime()
    evidence_block.now_factory = lambda: datetime(2026, 8, 26, 8, tzinfo=timezone.utc)
    evidence_block.build_table = lambda user_key: {  # type: ignore[method-assign]
        "contract_name": "sheet_vitrina_v1_sku_management_table",
        "generated_at": "2026-08-26T08:00:00+00:00",
        "settings": {"forecast": {"factory_to_ff_lead_days": 99}},
        "meta": {"metric_policy": {"business_date": "2026-08-26"}},
        "rows": [
            {
                "nm_id": 101,
                "stock_wb": 50,
                "stock_ff": 20,
                "quality_warnings": [],
            }
        ],
    }
    seven_day_evidence = evidence_block.build_inventory_balance_evidence(
        user_key="operator",
        sales_period_days=7,
    )
    fourteen_day_evidence = evidence_block.build_inventory_balance_evidence(
        user_key="operator",
        sales_period_days=14,
    )
    assert seven_day_evidence["rows"][0]["daily_demand"] == 10
    assert fourteen_day_evidence["rows"][0]["daily_demand"] == 15
    eta = seven_day_evidence["meta"]["inventory_balance_evidence"]["supplier_eta"]
    assert eta["sample_count"] == 4
    assert eta["mean_days_exact"] == 13
    assert eta["quality"] == "complete"
    evidence_inbounds = seven_day_evidence["rows"][0]["inventory_balance_inbounds"]
    assert [(item["status"], item["date"], item["quantity"]) for item in evidence_inbounds] == [
        ("in_transit", "2026-09-02", 30.0),
        ("production", "2026-09-07", 40.0),
    ]
    fallback_runtime = FakeBalanceEvidenceRuntime()
    del fallback_runtime.details["done-14"]
    del fallback_runtime.details["done-16"]
    evidence_block.runtime = fallback_runtime
    fallback_evidence = evidence_block.build_inventory_balance_evidence(
        user_key="operator",
        sales_period_days=7,
    )
    fallback_eta = fallback_evidence["meta"]["inventory_balance_evidence"]["supplier_eta"]
    assert fallback_eta["quality"] == "partial"
    assert fallback_eta["method"] == "configured_fallback_insufficient_completed_samples"
    assert fallback_eta["applied_days"] == 99

    with tempfile.TemporaryDirectory(prefix="sku-inventory-balance-") as tmp:
        runtime = FakeRuntime(Path(tmp) / "runtime.sqlite3")
        sku = FakeSkuManagement()
        timestamps = iter(
            [
                (datetime(2026, 8, 26, 8, tzinfo=timezone.utc) + timedelta(seconds=index)).isoformat()
                for index in range(60)
            ]
        )
        block = SkuInventoryBalanceBlock(
            runtime=runtime,
            sku_management_block=sku,
            now_factory=lambda: datetime(2026, 8, 26, 8, tzinfo=timezone.utc),
            timestamp_factory=lambda: next(timestamps),
        )
        with sqlite3.connect(runtime.db_path) as conn:
            outcome_columns = [
                str(row[1])
                for row in conn.execute(
                    "PRAGMA table_info(sheet_vitrina_v1_inventory_balance_outcomes)"
                ).fetchall()
            ]
            assert outcome_columns.count("observed_spend_rub") == 1
            assert len(outcome_columns) == len(set(outcome_columns))
        empty = block.latest(user_key="operator")
        assert empty["calculation"] is None
        assert empty["apply_capability"]["wb_patch_reachable"] is False
        assert empty["settings"]["calculation"]["wb_confidence_coefficient"] == 0.5

        settings = block.save_settings(
            {
                "base_revision": 0,
                "calculation": {
                    "wb_confidence_coefficient": 0.5,
                    "safety_stock_days": 10,
                    "sales_period_days": 7,
                },
                "table": {
                    "visible_columns": ["product", "status"],
                    "column_order": ["status", "product"],
                    "preset": "deficit",
                },
            },
            user_key="operator",
        )
        assert settings["revision"] == 1
        assert settings["table"]["column_order"][:2] == ["select", "product"]
        assert settings["table"]["preset"] == "deficit"

        all_fronts = calculate_inventory_balance_row(
            {
                "nm_id": 303,
                "name": "All-fronts stock semantics",
                "stock_wb": 50,
                "stock_ff": 20,
                "daily_demand": 5,
                "quality": "complete",
                "quality_warnings": [],
                "inventory_balance_as_of_date": "2026-08-26",
                "timeline": [
                    {"date": f"2026-08-{day:02d}", "ending_stock": 0}
                    for day in range(26, 32)
                ],
                "inventory_balance_inbounds": [
                    {
                        "date": "2026-08-30",
                        "quantity": 30,
                        "source": "supplier_shipment",
                        "source_id": "future-30",
                        "district_key": "central",
                        "consumes_current_ff": False,
                    },
                    {
                        "date": "2026-08-30",
                        "quantity": 30,
                        "source": "supplier_shipment",
                        "source_id": "future-30",
                        "district_key": "central",
                        "consumes_current_ff": False,
                    },
                    {
                        "date": "2026-08-29",
                        "quantity": 20,
                        "source": "current_ff_stock",
                        "source_id": "already-in-opening",
                        "consumes_current_ff": True,
                    },
                    {
                        "date": "2027-03-01",
                        "quantity": 10,
                        "source": "supplier_shipment",
                        "source_id": "beyond-incumbent-forecast",
                        "consumes_current_ff": False,
                    },
                ],
            },
            settings=settings["calculation"],
        )
        assert all_fronts["known_stock_units"] == 45
        assert all_fronts["confidence_adjusted_wb_units"] == 25
        assert all_fronts["stock_ff_units"] == 20
        assert all_fronts["next_inbound"]["quantity"] == 30
        assert all_fronts["next_inbound"]["sources"] == ["supplier_shipment"]
        assert all_fronts["subsequent_inbound"]["date"] == "2027-03-01"
        assert all_fronts["horizon_date"] == "2027-03-01"
        assert any("Duplicate registry inbound identity" in item for item in all_fronts["quality_warnings"])

        zero_sales = calculate_inventory_balance_row(
            {
                "nm_id": 304,
                "name": "Zero-sales overstock",
                "stock_wb": 100,
                "stock_ff": 20,
                "daily_demand": 0,
                "quality": "complete",
                "quality_warnings": [],
                "inventory_balance_as_of_date": "2026-08-26",
                "timeline": [
                    {"date": f"2026-08-{day:02d}", "ending_stock": 0}
                    for day in range(26, 32)
                ],
                "inventory_balance_inbounds": [
                    {
                        "date": "2026-08-30",
                        "quantity": 1,
                        "source": "supplier_shipment",
                        "source_id": "zero-launch-supply",
                        "consumes_current_ff": False,
                    }
                ],
            },
            settings=settings["calculation"],
        )
        assert zero_sales["status"] == "Переизбыток"
        assert zero_sales["target_daily_sales"] > 0
        assert zero_sales["pace_ratio"] == settings["calculation"]["bid_scale_max"]
        assert zero_sales["pace_ratio_basis"] == "zero_sales_bounded_launch"
        assert any("Нет наблюдаемой эластичности" in item for item in zero_sales["quality_warnings"])
        no_supply = calculate_inventory_balance_row(
            {
                "nm_id": 305,
                "name": "No fixed supply",
                "stock_wb": 100,
                "stock_ff": 20,
                "daily_demand": 5,
                "quality": "complete",
                "quality_warnings": [],
                "inventory_balance_as_of_date": "2026-08-26",
                "inventory_balance_inbounds": [],
            },
            settings=settings["calculation"],
        )
        assert no_supply["status"] == "Недостаточно данных"
        assert no_supply["pace_ratio_basis"] == "no_supply_unknown"
        assert no_supply["target_daily_sales"] is None
        before_arrival_constraint = calculate_inventory_balance_row(
            {
                "nm_id": 306,
                "name": "Before-arrival constraint",
                "stock_wb": 10,
                "stock_ff": 0,
                "daily_demand": 2,
                "quality": "complete",
                "quality_warnings": [],
                "inventory_balance_as_of_date": "2026-08-26",
                "inventory_balance_inbounds": [
                    {
                        "date": "2026-09-04",
                        "quantity": 100,
                        "source": "supplier_shipment",
                        "source_id": "day-10",
                        "consumes_current_ff": False,
                    }
                ],
            },
            settings={**settings["calculation"], "wb_confidence_coefficient": 1.0},
        )
        assert before_arrival_constraint["milestones"][0]["available_before_arrival"] == 10
        assert before_arrival_constraint["hard_daily_sales"] == 1
        assert before_arrival_constraint["reserve_daily_sales"] == 0.5

        calculation = block.calculate({}, user_key="operator", actor="operator")
        assert calculation["registry_immutable"] is True
        assert calculation["automatic_ml_or_training"] is False
        assert calculation["formula_version"] == "sku_inventory_balance_conservative_pace_v2"
        assert calculation["previous_calculation_id"] is None
        assert calculation["source_digest"].startswith("sha256:")
        assert {row["nm_id"] for row in calculation["rows"]}.isdisjoint(
            {497413772, 497415593, 497416931}
        )
        assert [row["nm_id"] for row in calculation["excluded_rows"]] == [
            497413772,
            497415593,
            497416931,
        ]
        assert calculation["lineage"]["exclusion_policy"]["identity_rule"] == "exact_nm_id_only"
        assert calculation["lineage"]["sales_evidence_window"]["sales_period_days"] == 7
        assert calculation["lineage"]["supplier_eta_evidence"]["mean_days_exact"] == 21
        assert [row["status"] for row in calculation["rows"]] == [
            "Дефицит",
            "Переизбыток",
            "Недостаточно данных",
        ]
        deficit = calculation["rows"][0]
        assert deficit["raw_opening_stock_units"] == 100
        assert deficit["known_stock_units"] == 50
        assert deficit["wb_confidence_coefficient"] == 0.5
        assert calculation["lineage"]["wb_stock_evidence"][0][
            "wb_confidence_coefficient"
        ] == 0.5
        assert len(deficit["new_cpc_campaigns"]) == 1
        assert len(deficit["old_cpm_campaigns"]) == 1
        cpc = deficit["new_cpc_campaigns"][0]
        assert cpc["cpo_rub"] == 30
        assert cpc["can_apply"] is False
        assert cpc["allocation_action"] == "hold_other_group"
        cpm = deficit["old_cpm_campaigns"][0]
        calculated_bid = cpm["calculated_target_bid_rub"]
        assert cpm["cpo_rub"] == 80
        assert cpm["can_apply"] is True
        assert cpm["allocation_action"] == "decrease_less_efficient_group"
        assert cpm["relative_efficiency"]["selected_group"] == "old_cpm"
        excess = calculation["rows"][1]
        excess_cpc = excess["new_cpc_campaigns"][0]
        excess_cpm = excess["old_cpm_campaigns"][0]
        assert excess_cpc["allocation_action"] == "increase_more_efficient_group"
        assert excess_cpc["can_apply"] is True
        assert excess_cpm["allocation_action"] == "hold_other_group"
        assert excess_cpm["can_apply"] is False
        no_supply_campaign = calculation["rows"][2]["new_cpc_campaigns"][0]
        assert no_supply_campaign["manual_override_allowed"] is False
        assert no_supply_campaign["can_apply"] is False
        try:
            block.save_override(
                calculation["calculation_id"],
                {"target_key": no_supply_campaign["target_key"], "manual_target_bid_rub": 4.25},
                actor="operator",
            )
        except SkuInventoryBalanceError as exc:
            assert exc.http_status == 422
        else:  # pragma: no cover
            raise AssertionError("manual override unexpectedly bypassed unknown inventory evidence")

        overridden = block.save_override(
            calculation["calculation_id"],
            {"target_key": cpm["target_key"], "manual_target_bid_rub": 725.25},
            actor="operator",
        )
        overridden_cpm = overridden["rows"][0]["old_cpm_campaigns"][0]
        assert overridden_cpm["calculated_target_bid_rub"] == calculated_bid
        assert overridden_cpm["manual_target_bid_rub"] == 725.25
        assert overridden_cpm["final_target_bid_rub"] == 725.25

        with sqlite3.connect(runtime.db_path) as conn:
            stored_override = conn.execute(
                """SELECT calculation_id,target_key,nm_id,advert_id,placement,
                          calculated_target_bid_rub,manual_target_bid_rub,updated_at,updated_by
                   FROM sheet_vitrina_v1_inventory_balance_overrides
                   WHERE calculation_id=? AND target_key=?""",
                (calculation["calculation_id"], cpm["target_key"]),
            ).fetchone()
            assert stored_override is not None
            assert stored_override[:5] == (
                calculation["calculation_id"],
                cpm["target_key"],
                101,
                8001,
                "search",
            )
            assert float(stored_override[5]) == calculated_bid
            assert float(stored_override[6]) == 725.25
            assert stored_override[8] == "operator"
            try:
                conn.execute(
                    "UPDATE sheet_vitrina_v1_inventory_balance_calculations SET created_by='tamper' WHERE calculation_id=?",
                    (calculation["calculation_id"],),
                )
            except sqlite3.IntegrityError as exc:
                assert "immutable" in str(exc)
            else:  # pragma: no cover - must remain fail closed
                raise AssertionError("immutable calculation update unexpectedly succeeded")

        second_calculation = block.calculate(
            {"calculation": {"sales_period_days": 14}},
            user_key="operator",
            actor="operator",
        )
        assert second_calculation["previous_calculation_id"] == calculation["calculation_id"]
        assert second_calculation["lineage"]["comparison_status"] == "available"
        assert second_calculation["lineage"]["sales_evidence_window"]["sales_period_days"] == 14
        assert next(row for row in second_calculation["rows"] if row["nm_id"] == 101)["current_daily_sales"] == 20
        assert sku.balance_sales_period_requests == [7, 14]

        workbook, filename = block.build_workbook(calculation["calculation_id"])
        assert filename.endswith(".xlsx")
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            assert archive.testzip() is None
        book = load_workbook(BytesIO(workbook), read_only=True, data_only=False)
        assert book.sheetnames[0] == "Решения"
        assert book.sheetnames == ["Решения", "Расчёт", "Кампании", "Поставки", "Источники", "История расчётов"]
        decision_headers = [cell.value for cell in next(book["Решения"].iter_rows(min_row=1, max_row=1))]
        assert "Следующая поставка" in decision_headers
        assert "Последующая поставка" in decision_headers
        assert "WB источник" in decision_headers
        next_column = decision_headers.index("Следующая поставка") + 1
        subsequent_column = decision_headers.index("Последующая поставка") + 1
        assert book["Решения"].cell(row=2, column=next_column).value == "2026-09-04: 100.0"
        assert book["Решения"].cell(row=2, column=subsequent_column).value in {None, ""}
        assert book["Кампании"].max_row == 7
        inbound_headers = [cell.value for cell in next(book["Поставки"].iter_rows(min_row=1, max_row=1))]
        assert inbound_headers == [
            "Товар", "nmID", "Дата", "Количество", "Доступно до прибытия",
            "Накопительно после прибытия", "Source", "Source ID", "Роль",
            "ETA method", "ETA samples", "ETA mean days", "ETA applied days",
            "ETA quality", "Row quality",
        ]
        first_inbound = [cell.value for cell in next(book["Поставки"].iter_rows(min_row=2, max_row=2))]
        assert first_inbound[1:9] == [101, "2026-09-04", 100, 50, 150, "supplier_shipment", "shipment-1", "next"]
        assert first_inbound[9:14] == ["empirical_last_completed_shipments", 3, 21, 21, "complete"]
        source_fields = {
            row[0].value: row[1].value
            for row in book["Источники"].iter_rows(min_row=2, max_col=2)
        }
        assert "wb_stock_evidence" in source_fields
        assert '"wb_confidence_coefficient":0.5' in source_fields["wb_stock_evidence"]
        assert book["История расчётов"].max_row == 3

        job = block.start_apply(
            {
                "calculation_id": calculation["calculation_id"],
                "nm_ids": [101],
                "mode": "dry_run",
                "confirmed": True,
            },
            actor="operator",
        )
        assert job["state"] == "pending"
        assert job["external_writes"] is False
        assert job["apply_manifest_digest"].startswith("sha256:")
        assert job["apply_manifest"]["targets"][0]["final_target_bid_rub"] == 725.25
        finished = block.resume_apply(job["job_id"], actor="operator", limit=10)
        assert finished["state"] == "completed"
        assert finished["progress"]["percent"] == 100.0
        assert finished["sku_states"] == [{"nm_id": 101, "state": "succeeded", "target_count": 1}]
        assert all(item["result"]["wb_patch_called"] is False for item in finished["items"])
        assert sku.preview_calls == 0
        assert sku.commit_calls == 0
        reloaded = block.latest(user_key="operator")
        assert reloaded["apply_job"]["job_id"] == job["job_id"]
        assert reloaded["apply_job"]["state"] == "completed"
        same_manifest = block.start_apply(
            {
                "calculation_id": calculation["calculation_id"],
                "nm_ids": [101],
                "mode": "dry_run",
                "confirmed": True,
            },
            actor="operator",
        )
        assert same_manifest["job_id"] == job["job_id"]
        changed = block.save_override(
            calculation["calculation_id"],
            {"target_key": cpm["target_key"], "manual_target_bid_rub": 700},
            actor="operator",
        )
        changed_job = block.start_apply(
            {
                "calculation_id": calculation["calculation_id"],
                "nm_ids": [101],
                "mode": "dry_run",
                "confirmed": True,
            },
            actor="operator",
        )
        assert changed_job["job_id"] != job["job_id"]
        assert changed_job["apply_manifest_digest"] != job["apply_manifest_digest"]
        assert changed_job["apply_manifest"]["targets"][0]["final_target_bid_rub"] == 700
        assert changed["rows"][0]["old_cpm_campaigns"][0]["calculated_target_bid_rub"] == calculated_bid
        registry = block.list_registry(limit=1)
        assert registry["items"][0]["calculation_id"] == second_calculation["calculation_id"]
        first_registry = block.list_registry(limit=20)["items"][1]
        assert first_registry["calculation_id"] == calculation["calculation_id"]
        assert [item["job_id"] for item in first_registry["apply_jobs"]] == [
            changed_job["job_id"],
            job["job_id"],
        ]
        assert first_registry["apply_jobs"][0]["apply_manifest_digest"] == changed_job["apply_manifest_digest"]

        try:
            block.start_apply(
                {
                    "calculation_id": calculation["calculation_id"],
                    "nm_ids": [101],
                    "mode": "live_wb",
                    "confirmed": True,
                },
                actor="operator",
            )
        except SkuInventoryBalanceError as exc:
            assert exc.http_status == 403
        else:  # pragma: no cover
            raise AssertionError("live mode unexpectedly accepted")

        live = LiveWbInventoryBalanceApplyAdapter(sku_management_block=sku)
        try:
            live.apply(overridden_cpm, actor="operator")
        except SkuInventoryBalanceError as exc:
            assert exc.http_status == 403
        else:  # pragma: no cover
            raise AssertionError("disabled live adapter unexpectedly called")
        assert sku.preview_calls == 0
        assert sku.commit_calls == 0

        operation_payload = {
            "operation_id": "ibop_unit_operation_0001",
            "idempotency_key": "ibkey_unit_operation_0001",
            "calculation": {"sales_period_days": 7},
        }
        acceptance = block.start_calculation_operation(
            operation_payload,
            user_key="operator",
            actor="operator",
        )
        repeated_acceptance = block.start_calculation_operation(
            operation_payload,
            user_key="operator",
            actor="operator",
        )
        acceptance_bytes = json.dumps(
            acceptance,
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8") + b"\n"
        repeated_bytes = json.dumps(
            repeated_acceptance,
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8") + b"\n"
        assert acceptance_bytes == repeated_bytes
        assert acceptance["state"] == "accepted"
        operation = _wait_operation(block, "ibop_unit_operation_0001")
        assert operation["state"] == "succeeded"
        assert operation["progress"] == {"percent": 100, "terminal": True}
        assert operation["result"]["operation_id"] == "ibop_unit_operation_0001"
        with sqlite3.connect(runtime.db_path) as conn:
            linked = conn.execute(
                """SELECT COUNT(*),COUNT(DISTINCT calculation_id)
                   FROM sheet_vitrina_v1_inventory_balance_calculations
                   WHERE operation_id=?""",
                ("ibop_unit_operation_0001",),
            ).fetchone()
            assert linked == (1, 1)

        original_evidence_builder = sku.build_inventory_balance_evidence

        def fail_before_calculation(*, user_key: str, sales_period_days: int) -> dict:
            del user_key, sales_period_days
            raise RuntimeError("synthetic upstream failure")

        sku.build_inventory_balance_evidence = fail_before_calculation  # type: ignore[method-assign]
        block.start_calculation_operation(
            {
                "operation_id": "ibop_unit_operation_0002",
                "idempotency_key": "ibkey_unit_operation_0002",
                "calculation": {"sales_period_days": 7},
            },
            user_key="operator",
            actor="operator",
        )
        failed_operation = _wait_operation(block, "ibop_unit_operation_0002")
        sku.build_inventory_balance_evidence = original_evidence_builder  # type: ignore[method-assign]
        assert failed_operation["state"] == "failed"
        assert failed_operation["calculation_id"] is None
        assert failed_operation["retryable_by_new_operation"] is True
        assert failed_operation["blind_resubmit_allowed"] is False
        assert failed_operation["outcome"]["durable_outcome"] == "no_calculation_created"

    _http_operation_disconnect_smoke()
    print("sku_inventory_balance_smoke: ok")


def _wait_operation(block: SkuInventoryBalanceBlock, operation_id: str) -> dict:
    for _attempt in range(200):
        operation = block.get_calculation_operation(operation_id, user_key="operator")
        if operation["state"] in {"succeeded", "failed"}:
            return operation
        time.sleep(0.01)
    raise AssertionError(f"operation did not finish: {operation_id}")


def _http_operation_disconnect_smoke() -> None:
    with tempfile.TemporaryDirectory(prefix="sku-inventory-balance-http-") as tmp:
        runtime = FakeRuntime(Path(tmp) / "runtime.sqlite3")
        sku = FakeSkuManagement()
        worker_started = threading.Event()
        worker_release = threading.Event()
        original_builder = sku.build_inventory_balance_evidence

        def blocked_builder(*, user_key: str, sales_period_days: int) -> dict:
            worker_started.set()
            if not worker_release.wait(timeout=5):
                raise RuntimeError("bounded test worker release timed out")
            return original_builder(
                user_key=user_key,
                sales_period_days=sales_period_days,
            )

        sku.build_inventory_balance_evidence = blocked_builder  # type: ignore[method-assign]
        block = SkuInventoryBalanceBlock(
            runtime=runtime,
            sku_management_block=sku,
        )
        entrypoint = FakeInventoryBalanceHttpEntrypoint(block, runtime)
        config = RegistryUploadHttpEntrypointConfig(
            host="127.0.0.1",
            port=0,
            upload_path=DEFAULT_UPLOAD_PATH,
            sheet_plan_path=DEFAULT_SHEET_PLAN_PATH,
            sheet_refresh_path=DEFAULT_SHEET_REFRESH_PATH,
            sheet_status_path=DEFAULT_SHEET_STATUS_PATH,
            sheet_operator_ui_path=DEFAULT_SHEET_OPERATOR_UI_PATH,
            runtime_dir=Path(tmp),
        )
        server = build_registry_upload_http_server(config, entrypoint=entrypoint)
        server_thread = threading.Thread(target=server.serve_forever, daemon=True)
        server_thread.start()
        port = int(server.server_address[1])
        base_url = f"http://127.0.0.1:{port}"
        operation_id = "ibop_http_disconnect_0001"
        payload = {
            "operation_id": operation_id,
            "idempotency_key": "ibkey_http_disconnect_0001",
            "calculation": {"sales_period_days": 7},
        }
        body = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        raw_request = (
            f"POST {DEFAULT_SKU_INVENTORY_BALANCE_CALCULATE_PATH} HTTP/1.1\r\n"
            f"Host: 127.0.0.1:{port}\r\n"
            "Content-Type: application/json\r\n"
            "Accept: application/json\r\n"
            f"Content-Length: {len(body)}\r\n"
            "Connection: close\r\n\r\n"
        ).encode("ascii") + body
        client = socket.create_connection(("127.0.0.1", port), timeout=2)
        try:
            client.sendall(raw_request)
        finally:
            client.close()
        try:
            assert worker_started.wait(timeout=2), "disconnected POST did not start durable operation"
            neighbor_started = time.monotonic()
            neighbor_status, neighbor_body = _http_get_json(
                base_url + DEFAULT_SKU_INVENTORY_BALANCE_PATH
            )
            neighbor_duration = time.monotonic() - neighbor_started
            assert neighbor_status == 200
            assert neighbor_duration < 1.0, neighbor_duration
            assert neighbor_body["calculation_operation"]["operation_id"] == operation_id
            assert neighbor_body["calculation_operation"]["state"] == "running"

            first_status, first_bytes = _http_post_json_bytes(
                base_url + DEFAULT_SKU_INVENTORY_BALANCE_CALCULATE_PATH,
                payload,
            )
            second_status, second_bytes = _http_post_json_bytes(
                base_url + DEFAULT_SKU_INVENTORY_BALANCE_CALCULATE_PATH,
                payload,
            )
            assert first_status == second_status == 202
            assert first_bytes == second_bytes

            worker_release.set()
            operation = None
            for _attempt in range(200):
                status, candidate = _http_get_json(
                    base_url
                    + DEFAULT_SKU_INVENTORY_BALANCE_OPERATIONS_PATH
                    + "/"
                    + operation_id
                )
                assert status == 200
                operation = candidate
                if operation["state"] in {"succeeded", "failed"}:
                    break
                time.sleep(0.01)
            assert operation is not None and operation["state"] == "succeeded", operation
            assert operation["result"]["operation_id"] == operation_id
            with sqlite3.connect(runtime.db_path) as conn:
                assert conn.execute(
                    """SELECT COUNT(*) FROM sheet_vitrina_v1_inventory_balance_operations
                       WHERE operation_id=?""",
                    (operation_id,),
                ).fetchone()[0] == 1
                assert conn.execute(
                    """SELECT COUNT(*) FROM sheet_vitrina_v1_inventory_balance_calculations
                       WHERE operation_id=?""",
                    (operation_id,),
                ).fetchone()[0] == 1
        finally:
            worker_release.set()
            server.shutdown()
            server_thread.join(timeout=5)
            server.server_close()


def _http_post_json_bytes(url: str, payload: dict) -> tuple[int, bytes]:
    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    request = urllib_request.Request(
        url,
        data=body,
        method="POST",
        headers={"Accept": "application/json", "Content-Type": "application/json"},
    )
    with urllib_request.urlopen(request, timeout=2) as response:
        return int(response.status), response.read()


def _http_get_json(url: str) -> tuple[int, dict]:
    request = urllib_request.Request(url, headers={"Accept": "application/json"})
    with urllib_request.urlopen(request, timeout=2) as response:
        return int(response.status), json.loads(response.read().decode("utf-8"))


if __name__ == "__main__":
    main()
