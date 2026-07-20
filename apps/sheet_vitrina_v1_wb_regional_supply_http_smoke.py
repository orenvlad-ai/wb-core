"""HTTP end-to-end smoke-check for the WB regional supply operator flow."""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import json
from pathlib import Path
import socket
import sys
from tempfile import TemporaryDirectory
import threading
from types import SimpleNamespace
from urllib import error, parse as urllib_parse, request as urllib_request
import zipfile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.adapters.registry_upload_http_entrypoint import (
    DEFAULT_FACTORY_ORDER_DELETE_STOCK_FF_PATH,
    DEFAULT_FACTORY_ORDER_TEMPLATE_STOCK_FF_PATH,
    DEFAULT_FACTORY_ORDER_UPLOAD_STOCK_FF_PATH,
    DEFAULT_SHEET_OPERATOR_UI_PATH,
    DEFAULT_SHEET_PLAN_PATH,
    DEFAULT_SHEET_REFRESH_PATH,
    DEFAULT_SHEET_STATUS_PATH,
    DEFAULT_UPLOAD_PATH,
    DEFAULT_WB_REGIONAL_CALCULATE_PATH,
    DEFAULT_WB_REGIONAL_PLANNING_OPTIONS_PATH,
    DEFAULT_WB_REGIONAL_RECOMMENDATIONS_ZIP_PATH,
    DEFAULT_WB_REGIONAL_STATUS_PATH,
    build_registry_upload_http_server,
)
from packages.application.factory_order_sales_history import persist_sales_history_result_exact_dates
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime
from packages.application.registry_upload_http_entrypoint import RegistryUploadHttpEntrypoint
from packages.application.simple_xlsx import build_single_sheet_workbook_bytes, read_first_sheet_rows
from packages.application.wb_regional_supply_export import recommendation_identity, recommendation_prefix
from packages.contracts.registry_upload_http_entrypoint import RegistryUploadHttpEntrypointConfig
from packages.contracts.sales_funnel_history_block import SalesFunnelHistoryItem, SalesFunnelHistorySuccess
from packages.contracts.supplier_shipments import (
    NOMENCLATURE_BARCODE_SOURCE_MANUAL,
    NOMENCLATURE_BARCODE_STATUS_MANUAL,
)
from packages.contracts.stocks_block import StocksEnvelope, StocksItem, StocksSuccess
from packages.contracts.wb_regional_supply import (
    DISTRICT_FAR_SIBERIA,
    DISTRICT_NORTHWEST,
)
from packages.contracts.wb_supply_planning_zones import (
    PLANNING_ZONE_CENTRAL_NORTH,
    PLANNING_ZONE_CENTRAL_SOUTH,
    SUPPLY_PLANNING_ZONE_KEYS,
)

INPUT_BUNDLE_FIXTURE = (
    ROOT / "artifacts" / "registry_upload_http_entrypoint" / "input" / "registry_upload_bundle__fixture.json"
)
NOW = datetime(2026, 4, 18, 9, 0, tzinfo=timezone.utc)
ACTIVATED_AT = "2026-04-18T09:00:00Z"
MAIN_NM_ID = 210183919


class FakeStocksBlock:
    def __init__(self, nm_ids: list[int]) -> None:
        self.nm_ids = list(nm_ids)

    def execute(self, request_obj: object) -> SimpleNamespace:
        items = []
        for nm_id in self.nm_ids:
            central = 0.0
            northwest = 0.0
            if nm_id == MAIN_NM_ID:
                central = 100.0
                northwest = 100.0
            items.append(
                SimpleNamespace(
                    nm_id=nm_id,
                    stock_total=central + northwest,
                    stock_ru_central=central,
                    stock_ru_northwest=northwest,
                    stock_ru_volga=0.0,
                    stock_ru_ural=0.0,
                    stock_ru_south_caucasus=0.0,
                    stock_ru_far_siberia=0.0,
                    stock_ru_central_north=central,
                    stock_ru_central_east=500.0 if nm_id == MAIN_NM_ID else 0.0,
                    stock_ru_central_south=500.0 if nm_id == MAIN_NM_ID else 0.0,
                )
            )
        return SimpleNamespace(result=SimpleNamespace(kind="success", items=items))


class NoopSalesHistoryBlock:
    def execute(self, request_obj: object) -> SimpleNamespace:  # pragma: no cover - should not be called
        raise AssertionError("runtime coverage is fully seeded; live fetch must not be called in HTTP smoke")


class FakePlanningSource:
    def __init__(self) -> None:
        self.acceptance_requests = []

    def fetch_acceptance_options(self, *, products, warehouse_id=None):
        self.acceptance_requests.append({"products": list(products), "warehouse_id": warehouse_id})
        return {
            "result": [
                {
                    "barcode": str((products or [{}])[0].get("barcode") or ""),
                    "warehouses": [{"warehouseID": 301806, "canBox": True}],
                }
            ]
        }

    def fetch_warehouses(self):
        return [{"ID": 301806, "name": "Тверь", "isActive": True}]

    def fetch_marketplace_offices(self):
        return []

    def fetch_box_tariffs(self, *, tariff_date=None):
        return [{"warehouseName": "Тверь", "boxDeliveryBase": "5"}]

    def fetch_transit_tariffs(self):
        return []

    def fetch_acceptance_coefficients(self, *, warehouse_ids=None):
        return [{"warehouseID": 301806, "warehouseName": "Тверь", "date": "2026-07-01", "coefficient": 1, "allowUnload": True, "boxTypeID": 1, "isSortingCenter": False}]


def main() -> None:
    bundle = json.loads(INPUT_BUNDLE_FIXTURE.read_text(encoding="utf-8"))
    with TemporaryDirectory(prefix="sheet-vitrina-wb-regional-http-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        port = _reserve_free_port()
        entrypoint = RegistryUploadHttpEntrypoint(
            runtime_dir=runtime_dir,
            runtime=runtime,
            activated_at_factory=lambda: ACTIVATED_AT,
            now_factory=lambda: NOW,
        )
        cfg = RegistryUploadHttpEntrypointConfig(
            host="127.0.0.1",
            port=port,
            upload_path=DEFAULT_UPLOAD_PATH,
            sheet_plan_path=DEFAULT_SHEET_PLAN_PATH,
            sheet_refresh_path=DEFAULT_SHEET_REFRESH_PATH,
            sheet_status_path=DEFAULT_SHEET_STATUS_PATH,
            sheet_operator_ui_path=DEFAULT_SHEET_OPERATOR_UI_PATH,
            runtime_dir=runtime_dir,
        )
        server = build_registry_upload_http_server(cfg, entrypoint=entrypoint)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            base_url = f"http://127.0.0.1:{cfg.port}"

            upload_status, upload_payload = _post_json(f"{base_url}{DEFAULT_UPLOAD_PATH}", bundle)
            if upload_status != 200 or upload_payload.get("status") != "accepted":
                raise AssertionError(f"bundle upload must be accepted, got {upload_status} {upload_payload}")

            active_nm_ids = [item.nm_id for item in runtime.load_current_state().config_v2 if item.enabled]
            _seed_runtime_sales_history(runtime, active_nm_ids=active_nm_ids)
            _seed_runtime_stock_history(runtime, active_nm_ids=active_nm_ids)
            entrypoint.wb_regional_supply_block.stocks_block = FakeStocksBlock(active_nm_ids)
            entrypoint.wb_regional_supply_block.sales_funnel_history_block = NoopSalesHistoryBlock()
            entrypoint.wb_regional_supply_block.sales_history.sales_funnel_history_block = NoopSalesHistoryBlock()
            planning_source = FakePlanningSource()
            entrypoint.wb_regional_supply_planning_block.source = planning_source

            operator_status, operator_html = _get_text(f"{base_url}{DEFAULT_SHEET_OPERATOR_UI_PATH}?embedded_tab=factory-order")
            if operator_status != 200:
                raise AssertionError(f"operator page must return 200, got {operator_status}")
            for expected in (
                "Общий вход для двух расчётов",
                "Поставка на Wildberries",
                "Цикл поставок, дней",
                "Рассчитать поставку на Wildberries",
                "Сводка по направлениям поставки",
                "Диагностика методологии появится после расчёта.",
                "Направления для расчёта поставки",
                "Доставка, дней",
                "Без ДВ/Сибирь",
                "Скачать все рекомендации",
                "Формируем архив...",
                "Рекомендации_поставок.zip",
                "Скачать Excel",
                "<th>Рекомендовано / к поставке</th>",
                "<th>Дефицит</th>",
                "<th>Скачать XLSX</th>",
                "<th>Подбор WB</th>",
                "Подобрать склады WB",
                "regionalPlanningPanel",
                "copyRegionalPlanningPayloadButton",
                "wb_regional_planning_options_path",
                "data-regional-district-download",
                "data-regional-planning-district",
                "data-regional-lead-time-district",
                "lead_time_to_region_days_by_district",
            ):
                if expected not in operator_html:
                    raise AssertionError(f"operator page must expose {expected!r}")
            for abbreviation in ("ЦФО Север", "ЦФО Восток", "ЦФО Юг", "СЗФО", "ПФО", "УФО", "ЮФО/СКФО", "ДВФО/СФО"):
                if abbreviation not in operator_html:
                    raise AssertionError(f"operator page must expose district abbreviation {abbreviation!r}")
            if "leadTimeToRegionDays" in operator_html or "Доставка до склада Wildberries, дней" in operator_html:
                raise AssertionError("operator UI must not expose global lead-time field as an editable source of truth")
            for removed in ("XLSX по округам", "Excel по округам", "regionalDistrictDownloads", "district-download-list"):
                if removed in operator_html:
                    raise AssertionError(f"regional UI must not render duplicated district download block token {removed!r}")
            if "https://docs.google.com/spreadsheets/d/" in operator_html:
                raise AssertionError("wb-regional supply surface must not expose legacy Google Sheets as an active link")
            if "value=\"14\"" not in operator_html or "value=\"15\"" not in operator_html or "value=\"250\"" not in operator_html:
                raise AssertionError("operator page must prefill the WB defaults directly in the form")

            stock_template_status, stock_template_bytes, _ = _get_bytes(
                f"{base_url}{DEFAULT_FACTORY_ORDER_TEMPLATE_STOCK_FF_PATH}"
            )
            if stock_template_status != 200:
                raise AssertionError("shared stock_ff template route must stay available")
            stock_rows = read_first_sheet_rows(stock_template_bytes)
            stock_upload_rows = [list(row) for row in stock_rows]
            for row in stock_upload_rows[1:]:
                row[2] = 0
            for row in stock_upload_rows[1:]:
                if int(row[0]) == MAIN_NM_ID:
                    row[2] = 120
                    break
            stock_upload_status, stock_upload_payload = _post_multipart(
                f"{base_url}{DEFAULT_FACTORY_ORDER_UPLOAD_STOCK_FF_PATH}",
                build_single_sheet_workbook_bytes("Остатки ФФ", stock_upload_rows),
                filename="shared-stock-ff.xlsx",
            )
            if stock_upload_status != 200 or stock_upload_payload.get("dataset", {}).get("uploaded_filename") != "shared-stock-ff.xlsx":
                raise AssertionError("stock_ff upload must stay shared and downloadable")

            regional_status_code, regional_status_payload = _get_json(f"{base_url}{DEFAULT_WB_REGIONAL_STATUS_PATH}")
            if regional_status_code != 200:
                raise AssertionError("regional status route must return 200")
            shared_dataset = regional_status_payload.get("shared_datasets", {}).get("stock_ff", {})
            if shared_dataset.get("uploaded_filename") != "shared-stock-ff.xlsx":
                raise AssertionError("regional status must expose the shared stock_ff filename")
            if shared_dataset.get("download_path") != "/v1/sheet-vitrina-v1/supply/factory-order/uploaded/stock-ff.xlsx":
                raise AssertionError("regional status must point to the shared stock_ff download route")
            if len(regional_status_payload.get("district_options") or []) != 8:
                raise AssertionError("regional status must expose district options")
            if DISTRICT_FAR_SIBERIA not in regional_status_payload.get("default_included_district_keys", []):
                raise AssertionError("regional status must default to all districts")

            calc_status, calc_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                },
            )
            if calc_status != 200:
                raise AssertionError(f"regional calculate route must succeed, got {calc_status} {calc_payload}")
            diagnostics = calc_payload.get("diagnostics") or {}
            if diagnostics.get("regional_demand_method") not in {
                "full_clean_days",
                "regional_share_ladder",
            }:
                raise AssertionError(f"regional diagnostics must expose share ladder methodology, got {diagnostics}")
            if diagnostics.get("fallback_sku_count") != 0:
                raise AssertionError(f"share ladder fixture must not use current-stock fallback, got {diagnostics}")
            if not diagnostics.get("share_source_counts"):
                raise AssertionError(f"share source counts must be exposed, got {diagnostics}")
            if diagnostics.get("requested_valid_day_count") != 14:
                raise AssertionError("regional diagnostics must expose requested depletion day count")
            if diagnostics.get("district_selection_mode") != "all_districts":
                raise AssertionError(f"old calculate payload must default to all districts, got {diagnostics}")
            if calc_payload.get("settings", {}).get("lead_time_to_region_days_by_district") != {key: 2 for key in SUPPLY_PLANNING_ZONE_KEYS}:
                raise AssertionError("old calculate payload must expand scalar lead time to all district keys")
            districts = {item["district_key"]: item for item in calc_payload.get("districts", [])}
            if PLANNING_ZONE_CENTRAL_NORTH not in districts:
                raise AssertionError(f"HTTP calculation must expose planning zones, got {sorted(districts)}; payload={calc_payload}")
            if districts[PLANNING_ZONE_CENTRAL_NORTH]["total_qty"] != 50 or districts[PLANNING_ZONE_CENTRAL_NORTH]["deficit_qty"] != 100:
                raise AssertionError("regional summary must expose truthful central-north allocation and deficit")
            if districts["northwest"]["total_qty"] != 50 or districts["northwest"]["deficit_qty"] != 100:
                raise AssertionError("regional summary must expose truthful northwest allocation and deficit")
            if sum(int(item.get("total_qty", 0)) for item in calc_payload.get("districts", [])) != int(calc_payload.get("summary", {}).get("total_qty", 0)):
                raise AssertionError("regional HTTP summary total must equal the sum of district totals")
            central_main_row = next(row for row in districts[PLANNING_ZONE_CENTRAL_NORTH]["rows"] if int(row["nm_id"]) == MAIN_NM_ID)
            row_diagnostics = central_main_row.get("demand_diagnostics") or {}
            if row_diagnostics.get("regional_demand_method") != "full_clean_days":
                raise AssertionError(f"main SKU must use full-clean diagnostics, got {row_diagnostics}")
            if row_diagnostics.get("selected_valid_day_count") != 14:
                raise AssertionError("main SKU must use 14 selected stock-depletion days")
            if abs(float(central_main_row.get("daily_demand_total", 0.0)) - 60.0) > 1e-9:
                raise AssertionError("main SKU total daily demand must remain based on orderCount")
            if int(central_main_row.get("lead_time_to_region_days", 0)) != 2:
                raise AssertionError("legacy scalar lead time must be exposed on HTTP district rows")

            _seed_planning_nomenclature(runtime, active_nm_ids=active_nm_ids)
            planning_status, planning_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_PLANNING_OPTIONS_PATH}",
                {
                    "district_key": PLANNING_ZONE_CENTRAL_NORTH,
                    "planning_zone_key": PLANNING_ZONE_CENTRAL_NORTH,
                    "calculation_id": calc_payload["calculation_id"],
                    "package_type": "box",
                },
            )
            if planning_status != 200:
                raise AssertionError(f"planning-options route must return 200, got {planning_status} {planning_payload}")
            if planning_payload.get("status") != "ready":
                raise AssertionError(f"planning-options must return ready payload, got {planning_payload}")
            if planning_payload.get("options", [{}])[0].get("warehouse_scope") != "same_district":
                raise AssertionError(f"planning-options must map warehouse to selected district, got {planning_payload}")
            if planning_source.acceptance_requests[0]["products"][0].get("barcode") != f"46{active_nm_ids[0]}":
                raise AssertionError(f"planning-options must call acceptance/options with barcode evidence, got {planning_source.acceptance_requests}")
            acceptance_request_count_before_mismatch = len(planning_source.acceptance_requests)
            mismatch_status, mismatch_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_PLANNING_OPTIONS_PATH}",
                {
                    "district_key": PLANNING_ZONE_CENTRAL_NORTH,
                    "planning_zone_key": PLANNING_ZONE_CENTRAL_NORTH,
                    "calculation_id": "stale-calculation-id",
                    "package_type": "box",
                },
            )
            if mismatch_status != 200:
                raise AssertionError(f"planning-options mismatch path must return 200 controlled payload, got {mismatch_status} {mismatch_payload}")
            if mismatch_payload.get("status") != "blocked" or not any(
                item.get("code") == "calculation_id_mismatch" for item in mismatch_payload.get("blockers", [])
            ):
                raise AssertionError(f"planning-options mismatch path must be a structured blocker, got {mismatch_payload}")
            if len(planning_source.acceptance_requests) != acceptance_request_count_before_mismatch:
                raise AssertionError("planning-options mismatch path must not call acceptance/options")

            mixed_lead_times = {key: 2 for key in SUPPLY_PLANNING_ZONE_KEYS}
            mixed_lead_times[DISTRICT_NORTHWEST] = 10
            mixed_status, mixed_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days_by_district": mixed_lead_times,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                    "included_district_keys": [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST],
                },
            )
            if mixed_status != 200:
                raise AssertionError(f"regional per-district lead-time calculate must succeed, got {mixed_status} {mixed_payload}")
            mixed_districts = {item["district_key"]: item for item in mixed_payload.get("districts", [])}
            mixed_central_row = next(row for row in mixed_districts[PLANNING_ZONE_CENTRAL_NORTH]["rows"] if int(row["nm_id"]) == MAIN_NM_ID)
            mixed_northwest_row = next(row for row in mixed_districts[DISTRICT_NORTHWEST]["rows"] if int(row["nm_id"]) == MAIN_NM_ID)
            if int(mixed_central_row.get("lead_time_to_region_days", 0)) != 2 or int(mixed_northwest_row.get("lead_time_to_region_days", 0)) != 10:
                raise AssertionError("HTTP rows must expose the district-specific lead times")
            if float(mixed_central_row.get("projected_stock_on_eta", 0.0)) <= float(mixed_northwest_row.get("projected_stock_on_eta", 0.0)):
                raise AssertionError("different lead times must change projected_stock_on_eta by district")
            if (mixed_payload.get("diagnostics") or {}).get("lead_time_to_region_days_by_district") != mixed_lead_times:
                raise AssertionError("HTTP diagnostics must expose applied lead-time map")

            _seed_wb_regional_overlay_fixture(
                runtime,
                supply_id="wb-http-regional-central",
                nm_id=MAIN_NM_ID,
                quantity=50.0,
                supply_date="2026-04-20",
                warehouse_name="Коледино",
                district_key="central",
                warehouse_id=507,
            )
            overlay_status, overlay_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                    "selected_wb_supply_ids": ["wb-http-regional-central"],
                },
            )
            if overlay_status != 200:
                raise AssertionError(f"regional selected WB supply HTTP calc must succeed, got {overlay_status} {overlay_payload}")
            overlay_diag = overlay_payload.get("wb_supply_overlay") or {}
            overlay_stock = overlay_diag.get("stock_ff", {})
            overlay_regional = overlay_diag.get("wb_regional", {})
            if overlay_stock.get("by_nm_id", {}).get(str(MAIN_NM_ID), {}).get("effective_stock_ff") != 70.0:
                raise AssertionError(f"regional HTTP overlay must reduce FF pool, got {overlay_stock}")
            added_by_district = overlay_regional.get("added_qty_by_district", {})
            if added_by_district.get(PLANNING_ZONE_CENTRAL_SOUTH) != 50.0:
                raise AssertionError(f"regional HTTP overlay must add qty to mapped district only, got {overlay_regional}")
            if any(float(qty or 0) for key, qty in added_by_district.items() if key != PLANNING_ZONE_CENTRAL_SOUTH):
                raise AssertionError(f"regional HTTP overlay must not spread selected qty to other districts, got {added_by_district}")
            overlay_districts = {item["district_key"]: item for item in overlay_payload.get("districts", [])}
            overlay_central_row = next(row for row in overlay_districts[PLANNING_ZONE_CENTRAL_SOUTH]["rows"] if int(row["nm_id"]) == MAIN_NM_ID)
            if overlay_central_row.get("demand_diagnostics", {}).get("selected_wb_supply_qty") != 50.0:
                raise AssertionError("regional HTTP row diagnostics must expose selected WB supply qty")

            selected_status, selected_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                    "included_district_keys": [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST],
                },
            )
            if selected_status != 200:
                raise AssertionError(f"regional selected-district calculate must succeed, got {selected_status} {selected_payload}")
            selected_diagnostics = selected_payload.get("diagnostics") or {}
            if selected_diagnostics.get("included_district_keys") != [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST]:
                raise AssertionError(f"selected district diagnostics must be returned, got {selected_diagnostics}")
            if DISTRICT_FAR_SIBERIA not in selected_diagnostics.get("excluded_district_keys", []):
                raise AssertionError("selected district diagnostics must include excluded far/siberia")
            selected_districts = {item["district_key"]: item for item in selected_payload.get("districts", [])}
            if sorted(selected_districts) != [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST]:
                raise AssertionError(f"selected district response must include selected districts only, got {sorted(selected_districts)}")

            unknown_lead_status, unknown_lead_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days_by_district": {
                        **{key: 2 for key in SUPPLY_PLANNING_ZONE_KEYS},
                        "unknown": 3,
                    },
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                },
            )
            if unknown_lead_status != 422 or "Неизвестный федеральный округ в сроках доставки" not in str(unknown_lead_payload.get("error", "")):
                raise AssertionError(f"unknown lead-time key must return controlled 422, got {unknown_lead_status} {unknown_lead_payload}")

            incomplete_lead_times = {key: 2 for key in SUPPLY_PLANNING_ZONE_KEYS}
            incomplete_lead_times.pop(DISTRICT_NORTHWEST)
            missing_lead_status, missing_lead_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days_by_district": incomplete_lead_times,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                },
            )
            if missing_lead_status != 422 or "Не задан срок доставки для федерального округа" not in str(missing_lead_payload.get("error", "")):
                raise AssertionError(f"missing lead-time key must return controlled 422, got {missing_lead_status} {missing_lead_payload}")

            invalid_status, invalid_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                    "included_district_keys": [],
                },
            )
            if invalid_status != 422 or "Выберите хотя бы один округ" not in str(invalid_payload.get("error", "")):
                raise AssertionError(f"empty district selection must return controlled 422, got {invalid_status} {invalid_payload}")

            central_download_path = districts[PLANNING_ZONE_CENTRAL_NORTH].get("download_path")
            if central_download_path != "/v1/sheet-vitrina-v1/supply/wb-regional/district/central_north.xlsx":
                raise AssertionError("regional district route must use narrow server-owned download path")
            district_status, district_bytes, district_headers = _get_bytes(f"{base_url}{central_download_path}")
            if district_status != 200 or "spreadsheetml.sheet" not in str(district_headers.get("Content-Type", "")):
                raise AssertionError("district download route must return XLSX")
            central_disposition = str(district_headers.get("Content-Disposition", ""))
            if (
                'filename="wb_regional_central_north_fo.xlsx"' not in central_disposition
                or not _is_ascii(central_disposition)
            ):
                raise AssertionError(f"central district download filename must be ASCII translit, got {central_disposition!r}")
            load_workbook(BytesIO(district_bytes), data_only=True)
            district_rows = read_first_sheet_rows(district_bytes)
            if district_rows[2] != ["nmId", "SKU", "Количество к поставке"]:
                raise AssertionError(f"district XLSX must remove the deficit column, got {district_rows[2]}")
            if any(cell == "Дефицит" for row in district_rows for cell in row):
                raise AssertionError("district XLSX must not contain deficit header or values")
            district_qty_sum = sum(int(row[2]) for row in district_rows[3:] if len(row) >= 3 and str(row[2]).strip())
            if district_qty_sum != districts[PLANNING_ZONE_CENTRAL_NORTH]["total_qty"]:
                raise AssertionError("district XLSX must match the regional summary total for the same district")

            northwest_download_path = selected_districts[DISTRICT_NORTHWEST].get("download_path")
            northwest_status, northwest_bytes, northwest_headers = _get_bytes(f"{base_url}{northwest_download_path}")
            northwest_disposition = str(northwest_headers.get("Content-Disposition", ""))
            if northwest_status != 200 or "spreadsheetml.sheet" not in str(northwest_headers.get("Content-Type", "")):
                raise AssertionError("included northwest district download route must return XLSX")
            if (
                'filename="wb_regional_northwest_fo.xlsx"' not in northwest_disposition
                or not _is_ascii(northwest_disposition)
            ):
                raise AssertionError(f"northwest district filename must be ASCII translit, got {northwest_disposition!r}")
            load_workbook(BytesIO(northwest_bytes), data_only=True)

            far_status, far_payload = _get_json(
                f"{base_url}/v1/sheet-vitrina-v1/supply/wb-regional/district/{DISTRICT_FAR_SIBERIA}.xlsx"
            )
            if far_status != 422 or "Округ не участвовал в последнем расчёте: far_siberia" not in str(far_payload.get("error", "")):
                raise AssertionError(f"excluded district direct download must return controlled 422, got {far_status} {far_payload}")

            zip_status, zip_bytes, zip_headers = _get_bytes(f"{base_url}{DEFAULT_WB_REGIONAL_RECOMMENDATIONS_ZIP_PATH}")
            zip_disposition = str(zip_headers.get("Content-Disposition", ""))
            if zip_status != 200 or str(zip_headers.get("Content-Type", "")).split(";")[0] != "application/zip":
                raise AssertionError(f"ZIP route must return application/zip, got {zip_status} {zip_headers}")
            zip_filename = _filename_from_content_disposition(zip_disposition)
            if (
                not zip_filename.startswith("Рекомендации_поставок_2026-04-18_14-00_")
                or str(selected_payload.get("calculation_id") or "") not in zip_filename
                or not zip_filename.endswith(".zip")
            ):
                raise AssertionError(f"ZIP content-disposition must expose the Unicode run filename, got {zip_disposition!r}")
            expected_archive_names: list[str] = []
            for ordinal, district in enumerate(selected_payload.get("districts") or [], start=1):
                recommendation_id = recommendation_identity(
                    report_date=str(selected_payload.get("report_date") or ""),
                    calculation_id=str(selected_payload.get("calculation_id") or ""),
                    ordinal=ordinal,
                )
                prefix = recommendation_prefix(
                    ordinal=ordinal,
                    recommendation_id=recommendation_id,
                    destination_name=str(
                        district.get("planning_zone_label")
                        or district.get("district_name_ru")
                        or ""
                    ),
                )
                expected_archive_names.extend(
                    [
                        f"{prefix}/{prefix}__01_РЕКОМЕНДАЦИЯ.xlsx",
                        f"{prefix}/{prefix}__02_ЗАГРУЗКА_WB.xlsx",
                    ]
                )
            with zipfile.ZipFile(BytesIO(zip_bytes)) as archive:
                archive_names = archive.namelist()
                if archive_names != expected_archive_names:
                    raise AssertionError(f"ZIP must contain paired XLSX folders in UI order, got {archive_names}")
                for name in archive_names:
                    workbook_bytes = archive.read(name)
                    load_workbook(BytesIO(workbook_bytes), data_only=True)
                    if name.endswith("__01_РЕКОМЕНДАЦИЯ.xlsx"):
                        rows = read_first_sheet_rows(workbook_bytes)
                        if rows[2] != ["nmId", "SKU", "Количество к поставке"]:
                            raise AssertionError(f"ZIP operator workbook {name} must remove deficit, got {rows[2]}")
                    elif name.endswith("__02_ЗАГРУЗКА_WB.xlsx"):
                        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
                        try:
                            sheet = workbook["Sheet1"]
                            if (sheet["A1"].value, sheet["B1"].value) != ("Баркод", "Количество"):
                                raise AssertionError(f"ZIP WB workbook {name} must preserve canonical headers")
                            if sum(int(sheet.cell(row=index, column=2).value) for index in range(2, sheet.max_row + 1)) <= 0:
                                raise AssertionError(f"ZIP WB workbook {name} must contain positive recommendation rows")
                        finally:
                            workbook.close()

            _seed_runtime_sales_history(runtime, active_nm_ids=active_nm_ids, all_active_signal=False)
            _seed_runtime_stock_history(
                runtime,
                active_nm_ids=active_nm_ids,
                all_active_signal=False,
                persistent_zero_south_for_main=True,
            )
            seed_stock_upload_rows = [list(row) for row in stock_rows]
            for row in seed_stock_upload_rows[1:]:
                row[2] = 0
            for row in seed_stock_upload_rows[1:]:
                if int(row[0]) == MAIN_NM_ID:
                    row[2] = 400
                    break
            seed_upload_status, _ = _post_multipart(
                f"{base_url}{DEFAULT_FACTORY_ORDER_UPLOAD_STOCK_FF_PATH}",
                build_single_sheet_workbook_bytes("Остатки ФФ", seed_stock_upload_rows),
                filename="shared-stock-ff-seed.xlsx",
            )
            if seed_upload_status != 200:
                raise AssertionError("seed stock_ff upload must succeed")
            seed_status, seed_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                    "included_district_keys": [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST, "south_caucasus"],
                },
            )
            if seed_status != 200:
                raise AssertionError(f"seed-floor calculate must succeed, got {seed_status} {seed_payload}")
            seed_diagnostics = seed_payload.get("diagnostics") or {}
            if seed_diagnostics.get("fallback_sku_count") != 0:
                raise AssertionError(f"seed-floor fixture must not fallback, got {seed_diagnostics}")
            if seed_diagnostics.get("seed_floor_sku_district_count", 0) < 1:
                raise AssertionError(f"HTTP seed-floor diagnostics must count affected directions, got {seed_diagnostics}")
            if seed_diagnostics.get("seed_sku_count") != 1 or seed_diagnostics.get("seed_allocated_qty_total") != 50:
                raise AssertionError(f"HTTP seed diagnostics must expose one allocated test box, got {seed_diagnostics}")
            seed_districts = {item["district_key"]: item for item in seed_payload.get("districts", [])}
            south_seed_row = next(row for row in seed_districts["south_caucasus"]["rows"] if int(row["nm_id"]) == MAIN_NM_ID)
            if int(south_seed_row.get("seed_qty", 0)) != 50 or not bool(south_seed_row.get("seed_floor_applied")):
                raise AssertionError(f"HTTP row must expose seed fields, got {south_seed_row}")
            if int(south_seed_row.get("demand_allocated_qty", -1)) != 0 or int(south_seed_row.get("allocated_qty", 0)) != 50:
                raise AssertionError("HTTP row must separate seed qty from demand allocation")
            if south_seed_row.get("share_source") != "seed_floor":
                raise AssertionError(f"HTTP row must expose seed_floor share source, got {south_seed_row}")

            delete_status, delete_payload = _delete_json(f"{base_url}{DEFAULT_FACTORY_ORDER_DELETE_STOCK_FF_PATH}")
            if delete_status != 200 or delete_payload.get("status") != "deleted":
                raise AssertionError("shared stock_ff delete route must still work")
            blocked_status, blocked_payload = _post_json(
                f"{base_url}{DEFAULT_WB_REGIONAL_CALCULATE_PATH}",
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                },
            )
            if blocked_status != 422 or "Остатки ФФ" not in str(blocked_payload.get("error", "")):
                raise AssertionError("regional calculate must truthfully block when shared stock_ff is missing")

            print(f"regional_status_shared_stock: ok -> {shared_dataset.get('uploaded_filename')}")
            print(f"regional_summary_total: ok -> {calc_payload.get('summary', {}).get('total_qty')}")
            print(f"regional_central_north_deficit: ok -> {districts[PLANNING_ZONE_CENTRAL_NORTH]['deficit_qty']}")
            print(f"regional_lead_times: ok -> central {mixed_central_row.get('lead_time_to_region_days')}, northwest {mixed_northwest_row.get('lead_time_to_region_days')}")
            print(f"regional_district_xlsx_sum: ok -> {district_qty_sum}")
            print(f"regional_recommendations_zip: ok -> {archive_names}")
            print(f"regional_seed_floor: ok -> {seed_diagnostics.get('seed_allocated_qty_total')}")
            print(f"regional_missing_shared_blocker: ok -> {blocked_payload.get('error')}")
        finally:
            server.shutdown()
            thread.join(timeout=5)
            server.server_close()


def _seed_runtime_sales_history(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    active_nm_ids: list[int],
    all_active_signal: bool = False,
) -> None:
    report_date = date(2026, 4, 18)
    items: list[SalesFunnelHistoryItem] = []
    for offset in range(14, 0, -1):
        snapshot_date = (report_date - timedelta(days=offset)).isoformat()
        for nm_id in active_nm_ids:
            value = 60.0 if nm_id == MAIN_NM_ID or all_active_signal else 0.0
            items.append(
                SalesFunnelHistoryItem(
                    date=snapshot_date,
                    nm_id=int(nm_id),
                    metric="orderCount",
                    value=value,
                )
            )
    persist_sales_history_result_exact_dates(
        runtime=runtime,
        payload=SalesFunnelHistorySuccess(
            kind="success",
            date_from="2026-04-04",
            date_to="2026-04-17",
            count=len(items),
            items=items,
        ),
        captured_at=ACTIVATED_AT,
    )


def _seed_runtime_stock_history(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    active_nm_ids: list[int],
    all_active_signal: bool = False,
    persistent_zero_south_for_main: bool = False,
) -> None:
    first_snapshot = date(2026, 4, 3)
    for index in range(15):
        snapshot_date = first_snapshot + timedelta(days=index)
        items: list[StocksItem] = []
        for nm_id in active_nm_ids:
            central = 0.0
            northwest = 0.0
            volga = 0.0
            ural = 0.0
            south = 0.0
            far = 0.0
            if nm_id == MAIN_NM_ID or all_active_signal:
                central = 1000.0 - (10.0 * index)
                northwest = 1000.0 - (10.0 * index)
                volga = 500.0
                ural = 500.0
                south = 0.0 if persistent_zero_south_for_main and nm_id == MAIN_NM_ID else 500.0
                far = 500.0
            items.append(
                StocksItem(
                    nm_id=int(nm_id),
                    stock_total=central + northwest + volga + ural + south + far,
                    stock_ru_central=central,
                    stock_ru_northwest=northwest,
                    stock_ru_volga=volga,
                    stock_ru_ural=ural,
                    stock_ru_south_caucasus=south,
                    stock_ru_far_siberia=far,
                    stock_ru_central_north=central,
                    stock_ru_central_east=500.0 if nm_id == MAIN_NM_ID or all_active_signal else 0.0,
                    stock_ru_central_south=500.0 if nm_id == MAIN_NM_ID or all_active_signal else 0.0,
                )
            )
        runtime.save_temporal_source_snapshot(
            source_key="stocks",
            snapshot_date=snapshot_date.isoformat(),
            captured_at=ACTIVATED_AT,
            payload=StocksEnvelope(
                result=StocksSuccess(
                    kind="success",
                    snapshot_date=snapshot_date.isoformat(),
                    count=len(items),
                    items=items,
                )
            ),
        )


def _seed_wb_regional_overlay_fixture(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    supply_id: str,
    nm_id: int,
    quantity: float,
    supply_date: str,
    warehouse_name: str,
    district_key: str,
    warehouse_id: int | str | None = None,
) -> None:
    effective_warehouse_id = str(warehouse_id or supply_id)
    runtime.save_wb_supply_rows(
        rows=[
            {
                "supply_id": supply_id,
                "cache_key": supply_id,
                "wb_supply_id": supply_id,
                "preorder_id": "pre-" + supply_id,
                "number_label": supply_id,
                "status_id": 3,
                "status_label": "Отгрузка разрешена",
                "warehouse_id": effective_warehouse_id,
                "warehouse_name": warehouse_name,
                "planned_warehouse_id": effective_warehouse_id,
                "planned_warehouse_name": warehouse_name,
                "target_warehouse_id": effective_warehouse_id,
                "target_warehouse_name": warehouse_name,
                "warehouse_display": warehouse_name,
                "district_source_warehouse_id": effective_warehouse_id,
                "district_source_warehouse_name": warehouse_name,
                "supply_date": supply_date,
                "district_key": district_key,
                "district_label_ru": "",
                "quantity_for_size_filter": quantity,
                "raw_list": {"supplyID": supply_id, "statusID": 3, "supplyDate": supply_date},
                "raw_detail": {"warehouseName": warehouse_name},
                "raw_goods": [{"nmID": int(nm_id), "quantity": float(quantity)}],
                "raw_package": [],
            }
        ],
        warehouses=[{"warehouse_id": effective_warehouse_id, "warehouse_name": warehouse_name}],
        synced_at=ACTIVATED_AT,
    )


def _seed_planning_nomenclature(runtime: RegistryUploadDbBackedRuntime, *, active_nm_ids: list[int]) -> None:
    rows: list[dict[str, object]] = []
    for nm_id in active_nm_ids:
        barcode = f"46{int(nm_id)}"
        rows.append(
            {
                "item_id": f"planning-item-{int(nm_id)}",
                "is_active": True,
                "nm_id": int(nm_id),
                "barcode": barcode,
                "barcodes": [barcode],
                "barcode_source": NOMENCLATURE_BARCODE_SOURCE_MANUAL,
                "barcode_status": NOMENCLATURE_BARCODE_STATUS_MANUAL,
                "barcode_updated_at": ACTIVATED_AT,
                "nomenclature_name": f"Planning SKU {int(nm_id)}",
                "product_type": "clear",
                "match_key": f"planning-{int(nm_id)}",
                "created_at": ACTIVATED_AT,
                "updated_at": ACTIVATED_AT,
            }
        )
    runtime.save_nomenclature_items_atomic(rows)


def _filename_from_content_disposition(value: str) -> str:
    for part in str(value or "").split(";"):
        normalized = part.strip()
        if normalized.lower().startswith("filename*="):
            encoded = normalized.split("=", 1)[1].strip().strip('"')
            if "''" in encoded:
                encoded = encoded.split("''", 1)[1]
            return urllib_parse.unquote(encoded)
    return ""


def _reserve_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def _request(url: str, *, method: str = "GET", body: bytes | None = None, headers: dict[str, str] | None = None):
    req = urllib_request.Request(url, data=body, method=method)
    for key, value in (headers or {}).items():
        req.add_header(key, value)
    try:
        with urllib_request.urlopen(req) as response:
            return response.status, response.read(), response.headers
    except error.HTTPError as exc:
        return exc.code, exc.read(), exc.headers


def _post_json(url: str, payload: dict[str, object]) -> tuple[int, dict[str, object]]:
    status, body, _ = _request(
        url,
        method="POST",
        body=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8", "Accept": "application/json"},
    )
    return status, json.loads(body.decode("utf-8"))


def _delete_json(url: str) -> tuple[int, dict[str, object]]:
    status, body, _ = _request(url, method="DELETE", headers={"Accept": "application/json"})
    return status, json.loads(body.decode("utf-8"))


def _get_json(url: str) -> tuple[int, dict[str, object]]:
    status, body, _ = _request(url, headers={"Accept": "application/json"})
    return status, json.loads(body.decode("utf-8"))


def _get_text(url: str) -> tuple[int, str]:
    status, body, _ = _request(url)
    return status, body.decode("utf-8")


def _get_bytes(url: str) -> tuple[int, bytes, dict[str, str]]:
    status, body, headers = _request(url)
    return status, body, dict(headers.items())


def _is_ascii(value: str) -> bool:
    try:
        value.encode("ascii")
    except UnicodeEncodeError:
        return False
    return True


def _post_multipart(url: str, workbook_bytes: bytes, *, filename: str) -> tuple[int, dict[str, object]]:
    boundary = "----wb-core-regional-smoke-boundary"
    body = (
        (
            f"--{boundary}\r\n"
            f"Content-Disposition: form-data; name=\"file\"; filename=\"{filename}\"\r\n"
            "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n"
            "\r\n"
        ).encode("utf-8")
        + workbook_bytes
        + f"\r\n--{boundary}--\r\n".encode("utf-8")
    )
    status, response_body, _ = _request(
        url,
        method="POST",
        body=body,
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Accept": "application/json",
        },
    )
    return status, json.loads(response_body.decode("utf-8"))


if __name__ == "__main__":
    main()
