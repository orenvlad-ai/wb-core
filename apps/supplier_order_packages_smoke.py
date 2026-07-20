"""Smoke-check deterministic supplier package assembly and DT XLSX mapping."""

from __future__ import annotations

import hashlib
from io import BytesIO
import json
from pathlib import Path
import sys
import zipfile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.application.registry_upload_http_entrypoint import (  # noqa: E402
    SupplierAccountingPackageBlockedError,
    _build_supplier_order_documents_archive,
)
from packages.application.supplier_customs_breakdown import (  # noqa: E402
    DT_ANNEX_ITEMS_PARSER_VERSION,
    STRICT_ACCOUNTING_RECONCILIATION_VERSION,
    WORKBOOK_HEADERS,
    build_customs_breakdown_xlsx,
    match_customs_annex_items,
    match_customs_goods_items,
)
from packages.application.supplier_financial_documents import (  # noqa: E402
    _enrich_customs_goods_items_from_annex_rows,
    _extract_customs_annex_rows_from_layout_pages,
)
from packages.application.supplier_customs_dt_matching_policy import (  # noqa: E402
    DT_ANNEX_MATCHING_POLICY_VERSION,
    resolve_dt_annex_series_model,
)


SHIPMENT_LINES = [
    {
        "line_id": "line-101",
        "line_type": "product",
        "sort_order": 1,
        "internal_nm_id": 101,
        "internal_name": "Our Alpha",
        "model_raw": "Exact Model Alpha",
        "barcode": "4600000000001",
        "qty": 10,
        "match_status": "matched_by_barcode",
        "raw": {},
    },
    {
        "line_id": "line-102",
        "line_type": "product",
        "sort_order": 2,
        "internal_nm_id": 102,
        "internal_name": "Our Group One",
        "model_raw": "Exact Model Group",
        "barcode": "4600000000002",
        "qty": 2,
        "match_status": "matched",
        "raw": {},
    },
    {
        "line_id": "line-103",
        "line_type": "product",
        "sort_order": 3,
        "internal_nm_id": 103,
        "internal_name": "Our Group Two",
        "model_raw": "Exact Model Group",
        "barcode": "4600000000003",
        "qty": 3,
        "match_status": "matched",
        "raw": {},
    },
]

NOMENCLATURE = [
    {
        "is_active": True,
        "nm_id": item["internal_nm_id"],
        "barcode": item["barcode"],
        "barcodes": [item["barcode"]],
        "nomenclature_name": item["internal_name"],
    }
    for item in SHIPMENT_LINES
]


def main() -> None:
    _assert_dt_annex_policy()
    _assert_annex_parser_boundaries()
    _assert_strict_reconciliation_and_workbook()
    _assert_strict_reconciliation_blockers()
    _assert_logistics_assemblies()
    _assert_strict_accounting_package()
    _assert_strict_accounting_multiple_customs()
    _assert_strict_accounting_reparse_is_read_only()
    print("supplier_order_packages_smoke: OK")


def _assert_dt_annex_policy() -> None:
    cases = (
        ({"article": "Anti-spy 13 Pro", "source_model": "Anti-spy iPhone 13 Pro"}, "confirmed", "anti_spy", ["iphone_13_pro"]),
        ({"article": "Matte 14 Pro Max", "source_model": "matte iPhone 14ProMax"}, "confirmed", "matte", ["iphone_14_pro_max"]),
        ({"article": "13 Pro", "source_model": "iPhone 13 Pro"}, "confirmed", "clean", ["iphone_13_pro"]),
        ({"article": "14-13", "source_model": "iPhone 14 / iPhone 13"}, "confirmed", "clean", ["iphone_13", "iphone_14"]),
        ({"article": "15/16", "source_model": "iPhone 15 - 16"}, "confirmed", "clean", ["iphone_15", "iphone_16"]),
        ({"article": "Anti-spy 13 Pro", "source_model": "Anti-spy 13 Pro"}, "confirmed", "anti_spy", ["iphone_13_pro"]),
        ({"article": "Anti Spy iPhone 13Pro", "source_model": "anti-spy 13 Pro"}, "confirmed", "anti_spy", ["iphone_13_pro"]),
        ({"article": "防窥 13", "source_model": "防窥 iPhone 13"}, "confirmed", "anti_spy", ["iphone_13"]),
        ({"article": "磨砂 14", "source_model": "磨砂 iPhone 14"}, "confirmed", "matte", ["iphone_14"]),
        ({"article": "高清 15", "source_model": "高清 iPhone 15"}, "confirmed", "clean", ["iphone_15"]),
        ({"article": "", "source_model": "Matte 15 Pro"}, "confirmed", "matte", ["iphone_15_pro"]),
        ({"article": "Anti-spy Matte 13 Pro", "source_model": "Anti-spy Matte 13 Pro"}, "ambiguous", "", []),
        ({"article": "Anti-spy 13 Pro", "source_model": "Anti-spy 14 Pro"}, "ambiguous", "", []),
    )
    for item, expected_status, expected_series, expected_keys in cases:
        result = resolve_dt_annex_series_model(item)
        actual = (result.get("status"), result.get("series"), result.get("model_keys"))
        expected = (expected_status, expected_series, expected_keys)
        if actual != expected or result.get("policy_version") != DT_ANNEX_MATCHING_POLICY_VERSION:
            raise AssertionError(f"versioned DT annex policy changed: {actual} != {expected}")


def _assert_dt_annex_order_scoped_matching_and_workbook() -> None:
    lines = [
        _dt_line("clean-13p-a", 201, "clean", ["iphone_13_pro"], "0012345678901", "Our Clean 13 Pro", 1),
        _dt_line("clean-13p-b", 201, "clean", ["iphone_13_pro"], "0012345678901", "Our Clean 13 Pro", 2),
        _dt_line("anti-13p", 202, "anti_spy", ["iphone_13_pro"], "0022345678901", "Our Anti 13 Pro", 3),
        _dt_line("matte-14pm", 203, "matte", ["iphone_14_pro_max"], "0032345678901", "Our Matte 14 Pro Max", 4),
        _dt_line("clean-14-13", 204, "clean", ["iphone_14", "iphone_13"], "0042345678901", "Our Clean 14-13", 5),
    ]
    nomenclature = [_dt_nomenclature(line) for line in lines if line["line_id"] != "clean-13p-b"]
    annex_items = [
        _annex("1", "1", "Anti-spy 13 Pro", "Anti-spy 13 Pro", 2),
        _annex("1", "2", "Matte 14 Pro Max", "matte iPhone 14ProMax", 3),
        _annex("1", "3", "13 Pro", "iPhone 13 Pro", 4),
        _annex("1", "4", "14-13", "iPhone 14 / iPhone 13", 5),
    ]
    matching = match_customs_annex_items(
        annex_items=annex_items,
        goods_items=[_goods("1", "Sanitized aggregate", 14)],
        shipment_lines=lines,
        nomenclature_items=nomenclature,
        expected_quantity_total=14,
        parser_quantity_conserved=True,
    )
    if (
        matching.get("annex_item_count") != 4
        or matching.get("output_row_count") != 4
        or matching.get("status_counts") != {"matched": 4}
        or matching.get("matched_count") != 4
        or not matching.get("quantity_conserved")
        or matching.get("requires_review")
        or matching.get("matching_policy_version") != DT_ANNEX_MATCHING_POLICY_VERSION
        or [row.get("nm_id") for row in matching["rows"]] != [202, 203, 201, 204]
        or any(not row.get("nomenclature_name") or not row.get("barcode") for row in matching["rows"])
    ):
        raise AssertionError(f"order-scoped DT annex matching changed: {matching}")
    if matching["rows"][2].get("determined_series") != "Clean":
        raise AssertionError("bare iPhone model must resolve only to Clean in DT policy")

    barcode_item = _annex("1", "5", "Anti-spy 13 Pro", "Matte 14 Pro Max", 1)
    barcode_item["barcode"] = "0012345678901"
    barcode_item["identifiers"]["barcode"] = "0012345678901"
    barcode_match = match_customs_annex_items(
        annex_items=[barcode_item],
        goods_items=[_goods("1", "Sanitized aggregate", 1)],
        shipment_lines=lines,
        nomenclature_items=nomenclature,
        expected_quantity_total=1,
    )
    if (
        barcode_match.get("status_counts") != {"matched": 1}
        or barcode_match["rows"][0].get("nm_id") != 201
        or barcode_match["rows"][0].get("source_barcode") != "0012345678901"
        or barcode_match["rows"][0].get("barcode") != "0012345678901"
    ):
        raise AssertionError(f"exact source barcode must precede DT series/model policy: {barcode_match}")

    ambiguous = match_customs_annex_items(
        annex_items=[annex_items[0]],
        goods_items=[_goods("1", "Sanitized aggregate", 2)],
        shipment_lines=[*lines, _dt_line("anti-13p-other", 299, "anti_spy", ["iphone_13_pro"], "0092345678901", "Other Anti 13 Pro", 6)],
        nomenclature_items=[*nomenclature, {**_dt_nomenclature(_dt_line("anti-13p-other", 299, "anti_spy", ["iphone_13_pro"], "0092345678901", "Other Anti 13 Pro", 6))}],
        expected_quantity_total=2,
    )
    if ambiguous.get("status_counts") != {"ambiguous": 1} or ambiguous["rows"][0].get("nm_id") is not None:
        raise AssertionError(f"different nmID owners must remain ambiguous: {ambiguous}")
    unmatched = match_customs_annex_items(
        annex_items=[_annex("1", "9", "iPhone 19", "19", 1)],
        goods_items=[_goods("1", "Sanitized aggregate", 1)],
        shipment_lines=lines,
        nomenclature_items=nomenclature,
        expected_quantity_total=1,
    )
    if unmatched.get("status_counts") != {"unmatched": 1} or unmatched["rows"][0].get("nm_id") is not None:
        raise AssertionError(f"missing order owner must remain unmatched: {unmatched}")

    body, _, receipt = build_customs_breakdown_xlsx(
        customs_document={
            "document_id": "dt-annex-safe",
            "file_original_name": "sanitized-customs.pdf",
            "normalized_parse": {
                "declaration_number": "10131010/100626/5187132",
                "declaration_date": "2026-06-10",
                "goods_items": [_goods("1", "Sanitized aggregate", 14)],
                "annex_items": annex_items,
                "annex_item_count": 4,
                "annex_quantity_total": 14,
                "annex_quantity_conserved": True,
                "annex_items_parser_version": DT_ANNEX_ITEMS_PARSER_VERSION,
            },
        },
        shipment={"header": {"shipment_id": "order-safe", "invoice_no": "SAFE-1"}},
        shipment_lines=lines,
        nomenclature_items=nomenclature,
    )
    if receipt.get("requires_review") or receipt.get("workbook_row_count") != 4:
        raise AssertionError(f"fully matched annex workbook must be accepted: {receipt}")
    workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
    sheet = workbook["Расшифровка ДТ"]
    header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "№ позиции ДТ")
    actual_headers = tuple(sheet.cell(header_row, column).value for column in range(1, len(WORKBOOK_HEADERS) + 1))
    if actual_headers != WORKBOOK_HEADERS:
        raise AssertionError(f"DT annex workbook columns changed: {actual_headers}")
    for row in range(header_row + 1, sheet.max_row + 1):
        if not isinstance(sheet.cell(row, 7).value, (int, float)):
            raise AssertionError("DT annex quantity must be a numeric Excel value")
        if any(sheet.cell(row, column).value in (None, "") for column in (2, 4, 5, 6, 9, 10, 11)):
            raise AssertionError("matched DT annex row lost row/article/model/series/nmID/name/barcode")
        if sheet.cell(row, 11).data_type != "s":
            raise AssertionError("canonical barcode must be stored losslessly as text")
    control = workbook["Контроль"]
    control_values = {control.cell(row, 1).value: control.cell(row, 2).value for row in range(1, control.max_row + 1)}
    expected_control = {
        "Строк приложения": 4,
        "Сопоставлено": 4,
        "Неоднозначно": 0,
        "Не сопоставлено": 0,
        "Количество сохранено": "Да",
        "Version matching policy": DT_ANNEX_MATCHING_POLICY_VERSION,
    }
    if any(control_values.get(key) != value for key, value in expected_control.items()):
        raise AssertionError(f"DT annex workbook control sheet changed: {control_values}")


def _assert_matching_and_workbook() -> None:
    model_only = _goods("5", "", 10)
    model_only["identifiers"] = {"source_model": "Exact Model Alpha"}
    goods_items = [
        _goods("1", "Exact Model Alpha", 10, barcode="4600000000001"),
        _goods("2", "Exact Model Group", 5),
        _goods("3", "Exact Model Group", 4),
        _goods("4", "Unknown exact source", 7),
        model_only,
    ]
    goods_items[0]["identifiers"].update(
        {"source_model": "DT-MODEL-ALPHA", "customs_code": "7020008000"}
    )
    matching = match_customs_goods_items(
        goods_items=goods_items,
        shipment_lines=SHIPMENT_LINES,
        nomenclature_items=NOMENCLATURE,
    )
    if (
        matching.get("position_count") != 5
        or matching.get("output_row_count") != 6
        or matching.get("status_counts") != {
            "ambiguous": 1,
            "matched": 2,
            "reconciled_group": 1,
            "unmatched": 1,
        }
        or not matching.get("quantity_conserved")
        or not matching.get("requires_review")
    ):
        raise AssertionError(f"DT deterministic matching contract changed: {matching}")
    matched = matching["rows"][0]
    ambiguous = next(item for item in matching["rows"] if item.get("status") == "ambiguous")
    unmatched = next(item for item in matching["rows"] if item.get("status") == "unmatched")
    if matched.get("nm_id") != 101 or ambiguous.get("nm_id") is not None or unmatched.get("nm_id") is not None:
        raise AssertionError(f"DT matcher guessed an nmID or lost exact barcode: {matching}")

    body, filename, receipt = build_customs_breakdown_xlsx(
        customs_document={
            "document_id": "dt-matching",
            "document_number": "10131010/100626/5187132",
            "file_original_name": "sanitized-customs.pdf",
            "normalized_parse": {
                "declaration_number": "10131010/100626/5187132",
                "declaration_date": "2026-06-10",
                "goods_items": goods_items,
            },
        },
        shipment={"header": {"shipment_id": "order-safe", "invoice_no": "SAFE-1"}},
        shipment_lines=SHIPMENT_LINES,
        nomenclature_items=NOMENCLATURE,
    )
    if filename != "DT_10131010_100626_5187132_rasshifrovka.xlsx" or not receipt.get("workbook_valid"):
        raise AssertionError(f"DT workbook filename/validation mismatch: {filename} {receipt}")
    if "rows" in receipt:
        raise AssertionError("bounded DT generation receipt must not expose row-level product evidence")
    workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
    sheet = workbook["Расшифровка ДТ"]
    header_row = next(
        row_index
        for row_index in range(1, sheet.max_row + 1)
        if sheet.cell(row=row_index, column=1).value == "№ позиции ДТ"
    )
    quantities = [sheet.cell(row=row_index, column=7).value for row_index in range(header_row + 1, sheet.max_row + 1)]
    if not quantities or not all(isinstance(value, (int, float)) for value in quantities):
        raise AssertionError(f"DT workbook quantities must be numeric Excel cells: {quantities}")
    barcode = sheet.cell(row=header_row + 1, column=11)
    if barcode.value != "4600000000001" or barcode.data_type != "s":
        raise AssertionError(f"DT workbook barcode must be lossless text: {barcode.value} {barcode.data_type}")
    if (
        sheet.cell(row=header_row + 1, column=5).value != "DT-MODEL-ALPHA"
        or sheet.cell(row=header_row + 1, column=15).value != "7020008000"
    ):
        raise AssertionError("DT workbook must retain available deterministic identifiers")


def _assert_incomplete_customs_workbook_fails_closed() -> None:
    try:
        build_customs_breakdown_xlsx(
            customs_document={
                "document_id": "dt-incomplete",
                "normalized_parse": {
                    "goods_items": [
                        {
                            "position_number": "1",
                            "source_name": "",
                            "quantity": None,
                            "unit": "",
                            "identifiers": {},
                        }
                    ]
                },
            },
            shipment={"header": {"shipment_id": "order-safe"}},
            shipment_lines=SHIPMENT_LINES,
            nomenclature_items=NOMENCLATURE,
        )
    except ValueError as exc:
        if "quantity is missing" not in str(exc):
            raise AssertionError(f"invalid DT workbook returned incomplete diagnostics: {exc}") from exc
    else:
        raise AssertionError("DT workbook with blank quantity/unit must fail closed")


def _assert_logistics_assemblies() -> None:
    rows = [
        _document("contract-1", "contract", "contract.pdf"),
        _document("payment-1", "bank_transfer_application", "payment.pdf"),
        _document("payment-2", "bank_transfer_application", "payment.pdf"),
        _document("control-1", "bank_control_statement", "control.pdf"),
    ]
    payload = {"supplier_order_id": "order-safe", "required_documents": rows, "shipment": {}}

    archive_bytes, receipt = _build_supplier_order_documents_archive(
        payload,
        package_type="logistics",
        file_loader=_fixture_loader,
    )
    manifest, names = _manifest_and_names(archive_bytes)
    if (
        receipt.get("status") != "complete"
        or receipt.get("counts", {}).get("included") != 4
        or len(names) != len(set(names))
        or manifest.get("included") != receipt.get("included")
        or sorted(names) != sorted(item["archive_name"] for item in manifest["included"])
    ):
        raise AssertionError(f"complete logistics package mismatch: {receipt} {names}")
    reversed_bytes, _ = _build_supplier_order_documents_archive(
        {**payload, "required_documents": list(reversed(rows))},
        package_type="logistics",
        file_loader=_fixture_loader,
    )
    reversed_manifest, _ = _manifest_and_names(reversed_bytes)
    names_by_document = {
        str(item.get("document_id") or ""): str(item.get("archive_name") or "")
        for item in manifest.get("included") or []
    }
    reversed_names_by_document = {
        str(item.get("document_id") or ""): str(item.get("archive_name") or "")
        for item in reversed_manifest.get("included") or []
    }
    if reversed_names_by_document != names_by_document:
        raise AssertionError("archive names must remain deterministic when source row order changes")

    failed_bytes, failed_receipt = _build_supplier_order_documents_archive(
        payload,
        package_type="logistics",
        file_loader=lambda row: (_fixture_loader(row) if row.get("document_id") != "payment-2" else _raise_unreadable()),
    )
    failed_manifest, failed_names = _manifest_and_names(failed_bytes)
    if (
        failed_receipt.get("status") != "error"
        or failed_receipt.get("counts", {}).get("expected") != 4
        or failed_receipt.get("counts", {}).get("included") != 3
        or failed_receipt.get("counts", {}).get("failed") != 1
        or failed_manifest.get("included") != failed_receipt.get("included")
        or len(failed_names) != 3
    ):
        raise AssertionError(f"unreadable package member must be a red exact-assembly receipt: {failed_receipt}")

    damaged_rows = [dict(row) for row in rows]
    damaged_rows[1]["file_sha256"] = hashlib.sha256(b"different stored upload bytes").hexdigest()
    _, damaged_receipt = _build_supplier_order_documents_archive(
        {**payload, "required_documents": damaged_rows},
        package_type="logistics",
        file_loader=_fixture_loader,
    )
    if (
        damaged_receipt.get("status") != "error"
        or damaged_receipt.get("counts", {}).get("failed") != 1
        or "checksum" not in str(damaged_receipt.get("failed", [{}])[0].get("reason") or "")
    ):
        raise AssertionError(f"damaged package member must fail stored-upload integrity: {damaged_receipt}")

    partial_payload = {
        **payload,
        "required_documents": [row for row in rows if row["document_type"] != "bank_control_statement"],
    }
    _, partial_receipt = _build_supplier_order_documents_archive(
        partial_payload,
        package_type="logistics",
        file_loader=_fixture_loader,
    )
    if partial_receipt.get("status") != "partial" or partial_receipt.get("counts", {}).get("missing") != 1:
        raise AssertionError(f"missing logistics type must remain distinct from read failure: {partial_receipt}")


def _assert_accounting_complete_annex_package() -> None:
    line = _dt_line("clean-13p", 201, "clean", ["iphone_13_pro"], "0012345678901", "Our Clean 13 Pro", 1)
    dt = _document("dt-annex", "customs_declaration", "dt.pdf")
    dt["normalized_parse"] = {
        "declaration_number": "10131010/100626/5187132",
        "declaration_date": "2026-06-10",
        "goods_items": [_goods("1", "Sanitized aggregate", 10)],
        "annex_items": [
            _annex("1", "1", "13 Pro", "iPhone 13 Pro", 4),
            _annex("1", "2", "13Pro", "13 Pro", 6),
        ],
        "annex_item_count": 2,
        "annex_quantity_total": 10,
        "annex_quantity_conserved": True,
        "annex_items_parser_version": DT_ANNEX_ITEMS_PARSER_VERSION,
    }
    payload = {
        "supplier_order_id": "order-safe",
        "shipment": {"header": {"shipment_id": "order-safe", "invoice_no": "SAFE-1"}, "lines": [line]},
        "required_documents": [
            _document("invoice-1", "invoice", "invoice.xlsx"),
            _document("contract-1", "contract", "contract.pdf"),
            dt,
        ],
    }
    archive_bytes, receipt = _build_supplier_order_documents_archive(
        payload,
        package_type="accounting",
        file_loader=_fixture_loader,
        nomenclature_items=[_dt_nomenclature(line)],
    )
    manifest, names = _manifest_and_names(archive_bytes)
    generated = manifest.get("generated_files") or []
    if (
        receipt.get("status") != "complete"
        or receipt.get("counts", {}).get("expected") != 4
        or receipt.get("counts", {}).get("included") != 4
        or receipt.get("requires_review")
        or len(generated) != 1
        or generated[0].get("validation", {}).get("annex_item_count") != 2
        or generated[0].get("validation", {}).get("matched_count") != 2
        or generated[0].get("validation", {}).get("matching_policy_version") != DT_ANNEX_MATCHING_POLICY_VERSION
        or sorted(names) != sorted(item["archive_name"] for item in manifest["included"])
    ):
        raise AssertionError(f"complete annex accounting package mismatch: {receipt} {manifest}")
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        workbook = load_workbook(BytesIO(archive.read(generated[0]["archive_name"])), read_only=True, data_only=True)
        sheet = workbook["Расшифровка ДТ"]
        header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "№ позиции ДТ")
        if sheet.max_row - header_row != 2:
            raise AssertionError("accounting XLSX must contain every annex row, not one aggregate")
        if any(sheet.cell(row, 11).value != "0012345678901" for row in range(header_row + 1, sheet.max_row + 1)):
            raise AssertionError("matched accounting XLSX rows must carry canonical order barcode")


def _assert_accounting_reparses_stored_customs_read_only() -> None:
    stale_dt = _document("dt-stale-v1", "customs_declaration", "dt.pdf")
    stale_dt["normalized_parse"] = {
        "declaration_number": "10131010/100626/5187132",
        "declaration_date": "2026-06-10",
        "goods_items_parser_version": "supplier_customs_goods_items_v2",
        "goods_items": [
            {
                "position_number": "1",
                "source_name": "Sanitized unknown item",
                "quantity": 12.5,
                "unit": "ШТ",
                "barcode": "",
                "identifiers": {"customs_code": "7020008000"},
            }
        ],
    }
    payload = {
        "supplier_order_id": "order-safe",
        "shipment": {"header": {"shipment_id": "order-safe", "invoice_no": "SAFE-1"}, "lines": SHIPMENT_LINES},
        "required_documents": [
            _document("invoice-1", "invoice", "invoice.xlsx"),
            _document("contract-1", "contract", "contract.pdf"),
            stale_dt,
        ],
    }
    parser_calls: list[str] = []

    def refreshed_parser(_: bytes, filename: str) -> dict[str, object]:
        parser_calls.append(filename)
        return {
            "normalized_parse": {
                "goods_items_parser_version": "supplier_customs_goods_items_v2",
                "goods_items": [
                    {
                        "position_number": "1",
                        "source_name": "Sanitized unknown item",
                        "quantity": 12.5,
                        "unit": "ШТ",
                        "barcode": "",
                        "identifiers": {"customs_code": "7020008000"},
                        "quantity_evidence": "dt_box_38_net_weight_kg",
                    }
                ],
                "annex_items_parser_version": DT_ANNEX_ITEMS_PARSER_VERSION,
                "annex_item_count": 2,
                "annex_quantity_total": 12.5,
                "annex_quantity_conserved": True,
                "annex_items": [
                    _annex("1", "1", "13 Pro", "iPhone 13 Pro", 5),
                    {**_annex("1", "2", "Matte 14 Pro Max", "Matte 14 Pro Max", 7), "quantity": 7.5},
                ],
            }
        }

    archive_bytes, receipt = _build_supplier_order_documents_archive(
        payload,
        package_type="accounting",
        file_loader=_fixture_loader,
        nomenclature_items=NOMENCLATURE,
        customs_parser=refreshed_parser,
    )
    manifest, _ = _manifest_and_names(archive_bytes)
    generated = manifest.get("generated_files") or []
    if (
        parser_calls != ["dt.pdf"]
        or receipt.get("status") != "complete"
        or len(generated) != 1
        or "annex_items" in stale_dt["normalized_parse"]
    ):
        raise AssertionError(f"incomplete stored DT item evidence was not refreshed read-only: {receipt}")
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        workbook = load_workbook(BytesIO(archive.read(generated[0]["archive_name"])), read_only=True, data_only=True)
        sheet = workbook["Расшифровка ДТ"]
        header_row = next(
            row_index
            for row_index in range(1, sheet.max_row + 1)
            if sheet.cell(row=row_index, column=1).value == "№ позиции ДТ"
        )
        quantities = [sheet.cell(row=row, column=7).value for row in range(header_row + 1, sheet.max_row + 1)]
        if quantities != [5, 7.5] or any(sheet.cell(row=row, column=8).value != "ШТ" for row in range(header_row + 1, sheet.max_row + 1)):
            raise AssertionError("read-only reparsed DT workbook did not preserve each annex row quantity/unit")


def _layout_header() -> str:
    return (
        f"{'№':<4}{'НАИМЕНОВАНИЕ':<34}{'ПРОИЗВОДИТЕЛЬ':<20}{'МАРКА':<16}"
        f"{'МОДЕЛЬ':<28}{'КОЛ-ВО':<16}{'АРТИКУЛ':<28}"
    )


def _layout_row(row_number: int, article: str, quantity: int) -> str:
    return (
        f"{row_number:<4}{'Sanitized protective glass':<34}{'Sanitized maker':<20}{'Safe':<16}"
        f"{article:<28}{f'{quantity} ШТ':<16}{article:<28}"
    )


def _assert_annex_parser_boundaries() -> None:
    multi_quantities = [5500, *([4100] * 26), 4150]
    multi_pages = [
        "\n".join(
            line
            for position, quantity in enumerate(multi_quantities, start=1)
            for line in (
                f"31 Товар № {position}",
                _layout_header(),
                _layout_row(1, f"Safe {13 + position % 4} Pro", quantity),
            )
        )
    ]
    multi_rows = _extract_customs_annex_rows_from_layout_pages(multi_pages)
    if (
        len(multi_rows) != 28
        or sum(int(item["quantity"]) for item in multi_rows) != 116250
        or [item.get("annex_row_number") for item in multi_rows] != ["1"] * 28
        or [item.get("parent_position_number") for item in multi_rows] != [str(index) for index in range(1, 29)]
    ):
        raise AssertionError(f"multi-position box-31 projection changed: {multi_rows}")
    canonical_goods = [
        {
            "position_number": str(index),
            "source_name": f"Canonical safe position {index}",
            "quantity": float(index),
            "unit": "кг",
            "identifiers": {"customs_code": "7000000000"},
        }
        for index in range(1, 29)
    ]
    original_goods = json.loads(json.dumps(canonical_goods))
    enriched = _enrich_customs_goods_items_from_annex_rows(
        {
            "normalized_parse": {
                "document_type": "customs_declaration",
                "goods_items": canonical_goods,
                "total_goods_count": 28,
                "gross_weight_kg": 999.5,
            }
        },
        multi_rows,
    )
    normalized = enriched["normalized_parse"]
    if (
        normalized.get("goods_items") != original_goods
        or normalized.get("total_goods_count") != 28
        or normalized.get("gross_weight_kg") != 999.5
        or normalized.get("annex_item_count") != 28
        or normalized.get("annex_quantity_total") != 116250
        or normalized.get("annex_items_parser_version") != DT_ANNEX_ITEMS_PARSER_VERSION
        or normalized.get("annex_parent_positions_complete") is not True
        or any(item.get("unit") != "ШТ" for item in normalized.get("annex_items") or [])
    ):
        raise AssertionError(f"multi-position enrichment changed canonical aggregates: {normalized}")

    for row_count, quantities, expected_total in (
        (27, [750, *([3000] * 25), 4500], 80250),
        (22, [250, *([2500] * 20), 5000], 55250),
    ):
        split = row_count // 2
        pages = [
            "\n".join(["Товар № 1", _layout_header(), *[
                _layout_row(index, f"Safe {13 + index % 4}", quantity)
                for index, quantity in enumerate(quantities[:split], start=1)
            ]]),
            "\n".join([_layout_header(), *[
                _layout_row(index, f"Safe {13 + index % 4}", quantity)
                for index, quantity in enumerate(quantities[split:], start=split + 1)
            ]]),
        ]
        rows = _extract_customs_annex_rows_from_layout_pages(pages)
        if len(rows) != row_count or sum(int(item["quantity"]) for item in rows) != expected_total:
            raise AssertionError(f"single-position multi-row annex changed: {row_count} {rows}")
        if any(item.get("parent_position_number") != "1" for item in rows):
            raise AssertionError("parent position was not carried through a page break")

    page_break_rows = _extract_customs_annex_rows_from_layout_pages([
        "Товар № 7",
        "\n".join([_layout_header(), _layout_row(1, "Safe 13 Pro", 125)]),
    ])
    if len(page_break_rows) != 1 or page_break_rows[0].get("parent_position_number") != "7":
        raise AssertionError(f"position marker was not carried to the following page: {page_break_rows}")


def _strict_fixture(first_quantity: int = 5500) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], dict[str, object]]:
    lines = [
        _strict_line("line-clean-wide", 501, "clean", ["iphone_17e", "iphone_16e", "iphone_14", "iphone_13", "iphone_13_pro"], "0000000000501", "Safe Clean Wide", first_quantity, 1),
        _strict_line("line-anti", 502, "anti_spy", ["iphone_13_pro"], "0000000000502", "Safe Anti 13 Pro", 2000, 2),
        _strict_line("line-matte", 503, "matte", ["iphone_14_pro_max"], "0000000000503", "Safe Matte 14 Pro Max", 3000, 3),
        {
            "line_id": "line-extra",
            "line_type": "extra",
            "sort_order": 4,
            "qty": 999,
            "match_status": "extra",
        },
    ]
    nomenclature = [_dt_nomenclature(line) for line in lines if line.get("line_type") == "product"]
    annex = [
        _annex("1", "1", "14-13-13PRO", "14-13-13PRO", first_quantity),
        _annex("1", "2", "Anti-spy 13 Pro", "Anti-spy 13 Pro", 2000),
        _annex("1", "3", "Matte 14 Pro Max", "Matte 14 Pro Max", 3000),
    ]
    shipment = {
        "shipment_id": "order-safe",
        "invoice_no": "SAFE-1",
        "product_qty_total": first_quantity + 5000,
        "match_status": "all_matched",
        "errors": [],
        "lines": lines,
    }
    return lines, nomenclature, annex, shipment


def _strict_line(
    line_id: str,
    nm_id: int,
    product_type: str,
    model_keys: list[str],
    barcode: str,
    name: str,
    quantity: int,
    sort_order: int,
) -> dict[str, object]:
    line = _dt_line(line_id, nm_id, product_type, model_keys, barcode, name, sort_order)
    line["qty"] = quantity
    return line


def _assert_strict_reconciliation_and_workbook() -> None:
    lines, nomenclature, annex, shipment = _strict_fixture()
    matching = match_customs_annex_items(
        annex_items=annex,
        goods_items=[_goods("1", "Canonical aggregate", 1)],
        shipment_lines=lines,
        nomenclature_items=nomenclature,
        expected_quantity_total=10500,
        parser_quantity_conserved=True,
        shipment=shipment,
    )
    if (
        not matching.get("package_ready")
        or matching.get("matched_count") != 3
        or matching.get("ambiguous_count") != 0
        or matching.get("unmatched_count") != 0
        or matching.get("invoice_product_row_count") != 3
        or matching.get("dt_annex_row_count") != 3
        or matching.get("invoice_product_quantity_total") != "10500"
        or matching.get("dt_quantity_total") != "10500"
        or matching.get("invoice_product_lines_covered") != 3
        or not matching.get("per_owner_quantity_reconciled")
        or matching.get("reconciliation_version") != STRICT_ACCOUNTING_RECONCILIATION_VERSION
    ):
        raise AssertionError(f"strict three-way reconciliation changed: {matching}")
    bounded = [row for row in matching["rows"] if row.get("status") == "matched_by_bounded_reconciliation"]
    if len(bounded) != 1 or bounded[0].get("nm_id") != 501 or "model-key containment" not in str(bounded[0].get("basis")):
        raise AssertionError(f"bounded 14-13-13PRO proof changed: {bounded}")
    if any(not row.get("nomenclature_name") or not row.get("barcode") for row in matching["rows"]):
        raise AssertionError("strict matcher lost server-owned nmID/name/barcode evidence")

    for first_quantity in (5500, 750, 250):
        quantity_lines, quantity_nomenclature, quantity_annex, quantity_shipment = _strict_fixture(first_quantity)
        quantity_matching = match_customs_annex_items(
            annex_items=quantity_annex,
            goods_items=[_goods("1", "Canonical aggregate", 1)],
            shipment_lines=quantity_lines,
            nomenclature_items=quantity_nomenclature,
            parser_quantity_conserved=True,
            shipment=quantity_shipment,
        )
        bounded_rows = [row for row in quantity_matching["rows"] if row.get("status") == "matched_by_bounded_reconciliation"]
        if not quantity_matching.get("package_ready") or len(bounded_rows) != 1 or bounded_rows[0].get("quantity") != str(first_quantity):
            raise AssertionError(f"bounded first-row quantity {first_quantity} changed: {quantity_matching}")

    barcode_item = _annex("1", "1", "Anti-spy 13 Pro", "Matte 14 Pro Max", 5500)
    barcode_item["barcode"] = "0000000000501"
    barcode_item["identifiers"]["barcode"] = "0000000000501"
    barcode_match = match_customs_annex_items(
        annex_items=[barcode_item],
        goods_items=[_goods("1", "Canonical aggregate", 1)],
        shipment_lines=[lines[0]],
        nomenclature_items=[nomenclature[0]],
        parser_quantity_conserved=True,
        shipment={"product_qty_total": 5500},
    )
    if not barcode_match.get("package_ready") or barcode_match["rows"][0].get("nm_id") != 501 or barcode_match["rows"][0].get("source_barcode") != "0000000000501":
        raise AssertionError(f"exact DT barcode priority changed: {barcode_match}")

    workbook_bytes, _, receipt = build_customs_breakdown_xlsx(
        customs_document={
            "document_id": "dt-safe",
            "file_original_name": "safe.pdf",
            "normalized_parse": {
                "declaration_number": "SAFE-DT",
                "goods_items": [_goods("1", "Canonical aggregate", 1)],
                "annex_items": annex,
                "annex_quantity_total": 10500,
                "annex_quantity_conserved": True,
            },
        },
        shipment=shipment,
        shipment_lines=lines,
        nomenclature_items=nomenclature,
    )
    if receipt.get("requires_review") or not receipt.get("workbook_valid") or not receipt.get("workbook_control_valid"):
        raise AssertionError(f"strict workbook validation changed: {receipt}")
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook["Расшифровка ДТ"]
    header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "№ позиции ДТ")
    if tuple(sheet.cell(header_row, column).value for column in range(1, len(WORKBOOK_HEADERS) + 1)) != WORKBOOK_HEADERS:
        raise AssertionError("strict workbook headers changed")
    for row in range(header_row + 1, sheet.max_row + 1):
        if not isinstance(sheet.cell(row, 7).value, (int, float)):
            raise AssertionError("strict workbook quantity is not numeric")
        if any(sheet.cell(row, column).value in (None, "") for column in (9, 10, 11)):
            raise AssertionError("strict workbook lost nmID, nomenclature or canonical barcode")
        if sheet.cell(row, 11).data_type != "s":
            raise AssertionError("canonical barcode is not stored losslessly as text")
    control = workbook["Контроль"]
    values = {control.cell(row, 1).value: control.cell(row, 2).value for row in range(1, control.max_row + 1)}
    expected = {
        "Готовность бухгалтерского пакета": "Да",
        "Product rows Invoice": 3,
        "DT annex rows": 3,
        "Product quantity Invoice": 10500,
        "DT quantity": 10500,
        "Сопоставлено": 3,
        "Неоднозначно": 0,
        "Не сопоставлено": 0,
        "Покрытие product lines Invoice": "3/3",
        "Per-owner quantity reconciliation": "Да",
        "Version parser projection": DT_ANNEX_ITEMS_PARSER_VERSION,
        "Version matching policy": DT_ANNEX_MATCHING_POLICY_VERSION,
        "Version reconciliation": STRICT_ACCOUNTING_RECONCILIATION_VERSION,
    }
    if any(values.get(key) != value for key, value in expected.items()):
        raise AssertionError(f"strict workbook control sheet changed: {values}")


def _assert_strict_reconciliation_blockers() -> None:
    first = _strict_line("one", 601, "clean", ["iphone_13"], "0000000000601", "Safe one", 5, 1)
    second = _strict_line("two", 602, "clean", ["iphone_13"], "0000000000602", "Safe two", 5, 2)
    ambiguous = match_customs_annex_items(
        annex_items=[_annex("1", "1", "13", "13", 5), _annex("1", "2", "13", "13", 5)],
        goods_items=[_goods("1", "Safe", 1)],
        shipment_lines=[first, second],
        nomenclature_items=[_dt_nomenclature(first), _dt_nomenclature(second)],
        parser_quantity_conserved=True,
        shipment={"product_qty_total": 10},
    )
    _require_blocker(ambiguous, "bounded_reconciliation_has_multiple_full_solutions")

    line13 = _strict_line("thirteen", 611, "clean", ["iphone_13"], "0000000000611", "Safe 13", 5, 1)
    line14 = _strict_line("fourteen", 612, "clean", ["iphone_14"], "0000000000612", "Safe 14", 5, 2)
    wrong_owner = match_customs_annex_items(
        annex_items=[_annex("1", "1", "13", "13", 4), _annex("1", "2", "14", "14", 6)],
        goods_items=[_goods("1", "Safe", 1)],
        shipment_lines=[line13, line14],
        nomenclature_items=[_dt_nomenclature(line13), _dt_nomenclature(line14)],
        parser_quantity_conserved=True,
        shipment={"product_qty_total": 10},
    )
    if wrong_owner.get("invoice_product_quantity_total") != wrong_owner.get("dt_quantity_total"):
        raise AssertionError("wrong-owner fixture must keep the same grand total")
    _require_blocker(wrong_owner, "per_owner_quantity_not_reconciled")

    lines, nomenclature, annex, shipment = _strict_fixture()
    row_mismatch = match_customs_annex_items(
        annex_items=annex[:-1],
        goods_items=[_goods("1", "Safe", 1)],
        shipment_lines=lines,
        nomenclature_items=nomenclature,
        parser_quantity_conserved=True,
        shipment=shipment,
    )
    _require_blocker(row_mismatch, "invoice_dt_product_row_count_mismatch")
    missing_quantity = [dict(item) for item in annex]
    missing_quantity[0]["quantity"] = None
    missing_quantity[0]["unit"] = ""
    incomplete = match_customs_annex_items(
        annex_items=missing_quantity,
        goods_items=[_goods("1", "Safe", 1)],
        shipment_lines=lines,
        nomenclature_items=nomenclature,
        parser_quantity_conserved=False,
        shipment=shipment,
    )
    _require_blocker(incomplete, "dt_annex_row_incomplete")
    absent = match_customs_annex_items(
        annex_items=[_annex("1", "1", "19", "19", 10500)],
        goods_items=[_goods("1", "Safe", 1)],
        shipment_lines=lines,
        nomenclature_items=nomenclature,
        parser_quantity_conserved=True,
        shipment=shipment,
    )
    _require_blocker(absent, "dt_rows_not_fully_matched")

    for missing_field in ("internal_nm_id", "internal_name", "canonical_barcode"):
        broken_lines, broken_nomenclature, broken_annex, broken_shipment = _strict_fixture()
        if missing_field == "internal_nm_id":
            broken_lines[0]["internal_nm_id"] = None
        elif missing_field == "internal_name":
            broken_lines[0]["internal_name"] = ""
            broken_nomenclature[0]["nomenclature_name"] = ""
        else:
            broken_lines[0]["barcode"] = ""
            broken_nomenclature[0]["barcode"] = ""
            broken_nomenclature[0]["barcodes"] = []
        broken = match_customs_annex_items(
            annex_items=broken_annex,
            goods_items=[_goods("1", "Safe", 1)],
            shipment_lines=broken_lines,
            nomenclature_items=broken_nomenclature,
            parser_quantity_conserved=True,
            shipment=broken_shipment,
        )
        _require_blocker(broken, "invoice_product_line_incomplete")


def _require_blocker(result: dict[str, object], code: str) -> None:
    codes = {str(item.get("code") or "") for item in result.get("blocker_reasons") or []}
    if result.get("package_ready") or code not in codes:
        raise AssertionError(f"strict blocker {code} was not enforced: {result}")


def _strict_package_payload() -> tuple[dict[str, object], list[dict[str, object]]]:
    lines, nomenclature, annex, shipment = _strict_fixture()
    invoice = _document("invoice-1", "invoice", "invoice.xlsx")
    invoice["document_number"] = "SAFE-1"
    invoice["file_sha256"] = hashlib.sha256(b"sanitized:invoice-1").hexdigest()
    shipment.update({
        "invoice_document_id": "invoice-1",
        "source_file_sha256": invoice["file_sha256"],
    })
    customs = _document("dt-safe", "customs_declaration", "dt.pdf")
    customs["normalized_parse"] = {
        "declaration_number": "SAFE-DT",
        "goods_items": [_goods("1", "Canonical aggregate", 1)],
        "annex_items": annex,
        "annex_item_count": 3,
        "annex_quantity_total": 10500,
        "annex_quantity_conserved": True,
        "annex_parent_position_count": 1,
        "annex_parent_positions_complete": True,
        "annex_items_parser_version": DT_ANNEX_ITEMS_PARSER_VERSION,
    }
    payload = {
        "supplier_order_id": "order-safe",
        "shipment": shipment,
        "required_documents": [invoice, _document("contract-1", "contract", "contract.pdf"), customs],
    }
    return payload, nomenclature


def _assert_strict_accounting_package() -> None:
    payload, nomenclature = _strict_package_payload()
    archive_bytes, receipt = _build_supplier_order_documents_archive(
        payload,
        package_type="accounting",
        file_loader=_fixture_loader,
        nomenclature_items=nomenclature,
    )
    manifest, names = _manifest_and_names(archive_bytes)
    if (
        receipt.get("status") != "complete"
        or receipt.get("requires_review")
        or receipt.get("counts", {}).get("included") != 4
        or receipt.get("accounting_reconciliation", {}).get("matched_count") != 3
        or receipt.get("accounting_reconciliation", {}).get("package_ready") is not True
        or sorted(names) != sorted(item["archive_name"] for item in manifest.get("included") or [])
    ):
        raise AssertionError(f"strict accounting package changed: {receipt} {manifest}")
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        for item in manifest.get("included") or []:
            body = archive.read(item["archive_name"])
            if len(body) != item.get("size_bytes") or "sha256:" + hashlib.sha256(body).hexdigest() != item.get("sha256"):
                raise AssertionError("strict ZIP member differs from manifest size/hash")
        generated = (manifest.get("generated_files") or [])[0]
        workbook = load_workbook(BytesIO(archive.read(generated["archive_name"])), read_only=True, data_only=True)
        if set(workbook.sheetnames) != {"Расшифровка ДТ", "Контроль"}:
            raise AssertionError("strict accounting workbook readback changed")

    blocked_payload = json.loads(json.dumps(payload))
    blocked_payload["required_documents"][2]["normalized_parse"]["annex_items"][0]["quantity"] = 5499
    blocked_payload["required_documents"][2]["normalized_parse"]["annex_quantity_total"] = 10499
    try:
        _build_supplier_order_documents_archive(
            blocked_payload,
            package_type="accounting",
            file_loader=_fixture_loader,
            nomenclature_items=nomenclature,
        )
    except SupplierAccountingPackageBlockedError as exc:
        diagnostics = exc.diagnostics
        if (
            diagnostics.get("status") != "blocked"
            or diagnostics.get("matched") == diagnostics.get("dt_row_count")
            or not diagnostics.get("blocker_reasons")
            or diagnostics.get("requires_review") is not True
        ):
            raise AssertionError(f"controlled accounting blocker diagnostics changed: {diagnostics}")
    else:
        raise AssertionError("accounting package returned ZIP for an unreconciled quantity")

    try:
        _build_supplier_order_documents_archive(
            payload,
            package_type="accounting",
            file_loader=lambda row: _raise_unreadable() if row.get("document_type") == "customs_declaration" else _fixture_loader(row),
            nomenclature_items=nomenclature,
        )
    except SupplierAccountingPackageBlockedError as exc:
        if "customs_declaration" not in exc.diagnostics.get("failed", []):
            raise AssertionError(f"unreadable DT was not exposed as a controlled blocker: {exc.diagnostics}")
    else:
        raise AssertionError("accounting package returned ZIP when a source document was unreadable")

    invalid_workbook_payload = json.loads(json.dumps(payload))
    invalid_workbook_payload["required_documents"][2]["normalized_parse"]["annex_items"][0]["source_name"] = (
        "sanitized\u0000invalid workbook cell"
    )
    try:
        _build_supplier_order_documents_archive(
            invalid_workbook_payload,
            package_type="accounting",
            file_loader=_fixture_loader,
            nomenclature_items=nomenclature,
        )
    except SupplierAccountingPackageBlockedError as exc:
        codes = {str(item.get("code") or "") for item in exc.diagnostics.get("blocker_reasons") or []}
        if "accounting_workbook_readback_failed" not in codes:
            raise AssertionError(f"workbook failure did not use controlled accounting blocker: {exc.diagnostics}")
    else:
        raise AssertionError("accounting package returned ZIP after workbook generation/readback failure")


def _assert_strict_accounting_multiple_customs() -> None:
    first = _strict_line("split-a", 701, "clean", ["iphone_13"], "0000000000701", "Safe split", 4, 1)
    second = _strict_line("split-b", 701, "clean", ["iphone_13"], "0000000000701", "Safe split", 6, 2)
    nomenclature = [_dt_nomenclature(first)]
    invoice = _document("invoice-multi", "invoice", "invoice.xlsx")
    invoice["document_number"] = "SAFE-MULTI"
    invoice["file_sha256"] = hashlib.sha256(b"sanitized:invoice-multi").hexdigest()
    documents = []
    for index, quantity in enumerate((4, 6), start=1):
        customs = _document(f"dt-{index}", "customs_declaration", f"dt-{index}.pdf")
        customs["normalized_parse"] = {
            "declaration_number": f"SAFE-DT-{index}",
            "goods_items": [_goods("1", "Safe", 1)],
            "annex_items": [_annex("1", "1", "13", "13", quantity)],
            "annex_item_count": 1,
            "annex_quantity_total": quantity,
            "annex_quantity_conserved": True,
            "annex_parent_position_count": 1,
            "annex_parent_positions_complete": True,
            "annex_items_parser_version": DT_ANNEX_ITEMS_PARSER_VERSION,
        }
        documents.append(customs)
    payload = {
        "supplier_order_id": "order-multi",
        "shipment": {
            "shipment_id": "order-multi",
            "invoice_no": "SAFE-MULTI",
            "invoice_document_id": "invoice-multi",
            "source_file_sha256": invoice["file_sha256"],
            "product_qty_total": 10,
            "match_status": "all_matched",
            "errors": [],
            "lines": [first, second],
        },
        "required_documents": [invoice, _document("contract-multi", "contract", "contract.pdf"), *documents],
    }
    archive_bytes, receipt = _build_supplier_order_documents_archive(
        payload,
        package_type="accounting",
        file_loader=_fixture_loader,
        nomenclature_items=nomenclature,
    )
    manifest, _ = _manifest_and_names(archive_bytes)
    if (
        receipt.get("status") != "complete"
        or receipt.get("counts", {}).get("generated_files") != 2
        or receipt.get("accounting_reconciliation", {}).get("invoice_product_lines_covered") != 2
        or not receipt.get("accounting_reconciliation", {}).get("per_owner_quantity_reconciled")
        or len(manifest.get("generated_files") or []) != 2
    ):
        raise AssertionError(f"multi-DT package aggregation changed: {receipt}")


def _assert_strict_accounting_reparse_is_read_only() -> None:
    payload, nomenclature = _strict_package_payload()
    stale_customs = payload["required_documents"][2]
    fresh_normalized = json.loads(json.dumps(stale_customs["normalized_parse"]))
    stale_customs["normalized_parse"] = {
        "declaration_number": "SAFE-DT",
        "goods_items": [_goods("1", "Canonical aggregate", 1)],
    }
    parser_calls: list[str] = []

    def parser(_: bytes, filename: str) -> dict[str, object]:
        parser_calls.append(filename)
        return {"normalized_parse": fresh_normalized}

    archive_bytes, receipt = _build_supplier_order_documents_archive(
        payload,
        package_type="accounting",
        file_loader=_fixture_loader,
        nomenclature_items=nomenclature,
        customs_parser=parser,
    )
    if (
        parser_calls != ["dt.pdf"]
        or receipt.get("status") != "complete"
        or "annex_items" in stale_customs["normalized_parse"]
        or not archive_bytes
    ):
        raise AssertionError("stored DT was not reparsed read-only under the strict projection")


def _document(document_id: str, document_type: str, filename: str) -> dict[str, object]:
    return {
        "document_id": document_id,
        "document_type": document_type,
        "document_name": document_type,
        "document_number": document_id,
        "file_original_name": filename,
        "is_uploaded": True,
        "status": "uploaded",
        "warnings": [],
    }


def _goods(position: str, name: str, quantity: int, *, barcode: str = "") -> dict[str, object]:
    return {
        "position_number": position,
        "source_name": name,
        "quantity": quantity,
        "unit": "ШТ",
        "barcode": barcode,
        "identifiers": {"barcode": barcode} if barcode else {},
    }


def _annex(position: str, row_number: str, article: str, model: str, quantity: int) -> dict[str, object]:
    return {
        "parent_position_number": position,
        "annex_row_number": row_number,
        "source_name": "Sanitized protective glass",
        "article": article,
        "source_model": model,
        "quantity": quantity,
        "unit": "ШТ",
        "barcode": "",
        "identifiers": {"article": article, "source_model": model, "customs_code": "7020008000"},
    }


def _dt_line(
    line_id: str,
    nm_id: int,
    product_type: str,
    model_keys: list[str],
    barcode: str,
    name: str,
    sort_order: int,
) -> dict[str, object]:
    return {
        "line_id": line_id,
        "line_type": "product",
        "sort_order": sort_order,
        "internal_nm_id": nm_id,
        "internal_name": name,
        "product_type": product_type,
        "group_key": product_type,
        "compatible_model_keys": model_keys,
        "barcode": barcode,
        "qty": 1,
        "match_status": "matched_by_barcode",
        "raw": {},
    }


def _dt_nomenclature(line: dict[str, object]) -> dict[str, object]:
    return {
        "is_active": True,
        "nm_id": line["internal_nm_id"],
        "barcode": line["barcode"],
        "barcodes": [line["barcode"]],
        "nomenclature_name": line["internal_name"],
        "product_type": line["product_type"],
        "compatible_model_keys": list(line["compatible_model_keys"]),
    }


def _fixture_loader(row: dict[str, object]) -> tuple[bytes, str, str]:
    filename = str(row.get("file_original_name") or "document.bin")
    return f"sanitized:{row.get('document_id')}".encode(), filename, "application/octet-stream"


def _raise_unreadable() -> tuple[bytes, str, str]:
    raise OSError("sanitized source file is unreadable")


def _manifest_and_names(archive_bytes: bytes) -> tuple[dict[str, object], list[str]]:
    with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
        return manifest, [name for name in archive.namelist() if name != "manifest.json"]


if __name__ == "__main__":
    main()
