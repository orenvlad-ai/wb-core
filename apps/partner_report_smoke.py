#!/usr/bin/env python3
"""UI-first indexed Partner Report and XLSX regression smoke."""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
import json
import os
from pathlib import Path
import sqlite3
import sys
from tempfile import TemporaryDirectory
import time

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.application.partner_report import (  # noqa: E402
    COMMON_EXPENSE_RULE,
    OTHER_DIRECT_ALLOCATED_KEY,
    OTHER_DIRECT_ALLOCATED_LABEL,
    OTHER_EXPENSE_CATEGORIES,
    REPORT_ROWS,
    PartnerReportBlock,
    PartnerReportError,
    _display_breakdown,
)

WEEK_ONE = date(2026, 7, 6)
WEEK_TWO = date(2026, 7, 13)
TARGET_NM = 101101
OTHER_NM = 202202


def main() -> None:
    with TemporaryDirectory(prefix="partner-report-ui-first-") as tmp:
        block = PartnerReportBlock(
            Path(tmp),
            seller_id="seller-1",
            now_factory=lambda: datetime(2026, 7, 20, 12, 0, tzinfo=timezone.utc),
        )
        block.ensure_schema()
        _seed_sources(block.db_path)
        _seed_finance(block)
        _seed_ads(block.db_path)
        _assert_server_owned_settings(block)
        _assert_expense_category_math(block)
        report = _assert_preview(block)
        _assert_workbook(block, report)
        _assert_only_marketing_workbook(block)
        _assert_partial_cost_surface(block)
        _assert_incomplete_and_stale_states(block)
        _assert_negative_profit_and_validation(block)
        performance = _assert_indexed_performance(block)
        _assert_removed_package_surface(block)
    print(
        "partner_report: ok -> server-owned parameters, indexed UI preview, root/nested ads, "
        "blockers without zero masking, Finance formulas, dividends/annualized return, "
        "reference-like XLSX, no ZIP/raw export, production-like preview performance; "
        f"raw_scan_ms={performance['raw_scan_ms']}, indexed_preview_ms={performance['indexed_preview_ms']}, "
        f"raw_rows={performance['raw_rows']}"
    )


def _assert_server_owned_settings(block: PartnerReportBlock) -> None:
    options = block.options()
    cards = {item["nm_id"]: item for item in options["cards"]}
    if str(TARGET_NM) not in cards or str(OTHER_NM) not in cards:
        raise AssertionError(f"canonical nomenclature options incomplete: {options}")
    if "XLSX only" not in options["export_contract"]:
        raise AssertionError(f"current export scope not disclosed: {options}")
    saved = block.save_settings(_settings(), actor="operator@example.test")
    repeated = block.save_settings(_settings(), actor="operator@example.test")
    if saved["settings_version_id"] != repeated["settings_version_id"]:
        raise AssertionError("unchanged settings created an unnecessary version")
    updated_payload = _settings()
    updated_payload["partner_share_pct"] = "41"
    updated = block.save_settings(updated_payload, actor="operator@example.test")
    if updated["settings_version_id"] == saved["settings_version_id"]:
        raise AssertionError("changed server settings did not create an audited version")
    block.save_settings(_settings(), actor="operator@example.test")
    with sqlite3.connect(block.db_path) as conn:
        versions = conn.execute(
            "SELECT COUNT(*) FROM partner_report_settings_versions WHERE nm_id=?",
            (str(TARGET_NM),),
        ).fetchone()[0]
        audit = conn.execute(
            "SELECT COUNT(*) FROM partner_report_audit WHERE action='settings_saved'"
        ).fetchone()[0]
    if versions != 3 or audit != 3:
        raise AssertionError("settings audit/version provenance mismatch")


def _assert_expense_category_math(block: PartnerReportBlock) -> None:
    account_metrics = {
        "profit_period_expenses": "42",
        "positive_adjustments": "1",
        "marketing": "10",
        "agent_remuneration": "2",
        "acquiring": "1",
        "logistics": "3",
        "storage": "4",
        "acceptance": "5",
        "capitalized_acceptance": "1",
        "penalties": "2",
        "corrections": "1",
        "transit_logistics": "4",
        "capitalized_transit_logistics": "1",
        "subscriptions": "2",
        "paid_services": "1",
        "review_points": "6",
        "other_deductions": "3",
    }
    if block._account_allocatable_expense_total(account_metrics) != Decimal("31"):  # noqa: SLF001
        raise AssertionError("Finance marketing was not excluded from account allocation")
    allocated = block._allocated_account_expense_categories(  # noqa: SLF001
        account_metrics,
        allocation_ratio=Decimal("0.5"),
    )
    if allocated != {
        "uncapitalized_transit_logistics": Decimal("1.5"),
        "wb_jam_subscription": Decimal("1.0"),
        "wb_paid_services": Decimal("0.5"),
        "review_points": Decimal("3.0"),
        "other_withholdings": Decimal("1.5"),
    }:
        raise AssertionError(f"allocated subrow category mismatch: {allocated}")
    allocated_main = block._allocated_account_main_expenses(  # noqa: SLF001
        account_metrics,
        allocation_ratio=Decimal("0.5"),
    )
    if allocated_main != {
        "agent_remuneration": Decimal("1.0"),
        "acquiring": Decimal("0.5"),
        "logistics": Decimal("1.5"),
        "storage": Decimal("2.0"),
        "acceptance": Decimal("2.0"),
        "penalties_and_adjustments": Decimal("1.0"),
    }:
        raise AssertionError(f"allocated main-row category mismatch: {allocated_main}")
    if sum(allocated.values(), Decimal()) + sum(
        allocated_main.values(), Decimal()
    ) != Decimal("15.5"):
        raise AssertionError("allocated account categories do not conserve without marketing")
    formula = block._week_formulas(  # noqa: SLF001
        components={
            "net_revenue": "100",
            "agent_remuneration": "1",
            "acquiring": "2",
            "logistics": "3",
            "storage": "4",
            "acceptance": "5",
            "capitalized_acceptance": "1",
            "penalties": "6",
            "corrections": "2",
            "positive_adjustments": "1",
            "marketing": "999",
        },
        cogs=Decimal("10"),
        ads=Decimal("11"),
        other_expense_categories={
            key: Decimal("0") for key, _label in OTHER_EXPENSE_CATEGORIES
        },
        allocated_main_expenses=allocated_main,
        other_direct_and_allocated_total=Decimal("12"),
        params={
            "partner_share_pct": "40",
            "invested_capital_rub": "1000",
            "replenishment_reserve_pct": "0",
            "weekly_office_expense_rub": "0",
            "tax_rate_pct": "0",
        },
    )
    expected_formula = {
        "agent_remuneration": "2.0000",
        "acquiring": "2.5000",
        "logistics": "4.5000",
        "storage": "6.0000",
        "acceptance": "6.0000",
        "penalties_and_adjustments": "8.0000",
        OTHER_DIRECT_ALLOCATED_KEY: "12.0000",
        "finance_margin": "38.0000",
        "net_profit": "38.0000",
        "dividends": "15.2000",
    }
    for key, expected_value in expected_formula.items():
        if formula.get(key) != expected_value:
            raise AssertionError(
                f"allocated main-row/margin {key}: {formula.get(key)!r} != {expected_value!r}"
            )
    exact = block._other_expense_categories(  # noqa: SLF001
        {
            "transit_logistics": "2",
            "capitalized_transit_logistics": "3",
            "subscriptions": "4",
            "paid_services": "5",
            "other_deductions": "6",
        },
        allocated,
    )
    if exact["uncapitalized_transit_logistics"] != Decimal("0.5"):
        raise AssertionError("signed transit correction was clamped or double-counted")
    if sum(exact.values(), Decimal("0")) != Decimal("21.5"):
        raise AssertionError(f"direct + allocated categories do not conserve: {exact}")
    rounded = _display_breakdown(
        {key: Decimal("0.005") for key, _label in OTHER_EXPENSE_CATEGORIES}
    )
    rounded_values = [Decimal(item["amount_rub"]) for item in rounded]
    if sum(rounded_values, Decimal("0")) != Decimal("0.03"):
        raise AssertionError(f"kopeck residual is not deterministic/conserved: {rounded}")
    serialized_edge = _display_breakdown(
        {
            "other_withholdings": Decimal("0.004951"),
        }
    )
    if sum(
        (Decimal(item["amount_rub"]) for item in serialized_edge), Decimal("0")
    ) != Decimal("0.01"):
        raise AssertionError(
            "category cents must reconcile to the four-decimal serialized main row"
        )
    period_edge = _display_breakdown(
        {
            "uncapitalized_transit_logistics": Decimal("73.942707"),
            "wb_jam_subscription": Decimal("44.097056"),
            "wb_paid_services": Decimal("142.926849"),
            "review_points": Decimal("0"),
            "other_withholdings": Decimal("95.388299"),
        },
        target_total=Decimal("356.3550"),
    )
    if sum(
        (Decimal(item["amount_rub"]) for item in period_edge), Decimal("0")
    ) != Decimal("356.36"):
        raise AssertionError(
            "period category cents must reconcile to the sum of serialized weeks"
        )


def _assert_preview(block: PartnerReportBlock) -> dict:
    report = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat(), WEEK_TWO.isoformat()]}
    )
    if report["status"] != "ready" or report["blockers"]:
        raise AssertionError(f"complete preview was blocked: {report}")
    if report["performance"]["raw_finance_full_scan"]:
        raise AssertionError(f"preview regressed to a raw Finance scan: {report}")
    if report["preview_source"] != "indexed_per_sku_weekly_finance_aggregate":
        raise AssertionError(f"indexed projection is not the report source: {report}")
    first = report["weeks"][0]["values"]
    expected_first = {
        "net_revenue": "476034.0000",
        "cogs": "83837.0000",
        "agent_remuneration": "160000.0000",
        "acquiring": "14797.0000",
        "ads": "30904.0000",
        "finance_margin": "186496.0000",
        "office": "10000.0000",
        "estimated_tax": "28562.0400",
        "replenishment_reserve": "37299.2000",
        "net_profit": "110634.7600",
        "dividends": "44253.9040",
        "annualized_return_pct": "230.1203",
    }
    for key, expected in expected_first.items():
        if first.get(key) != expected:
            raise AssertionError(f"control fixture {key}: {first.get(key)!r} != {expected!r}")
    second = report["weeks"][1]["values"]
    if second[OTHER_DIRECT_ALLOCATED_KEY] != "2500.0000":
        raise AssertionError(f"approved revenue-share allocation mismatch: {second}")
    second_breakdown = {
        item["key"]: Decimal(item["amount_rub"])
        for item in report["weeks"][1]["other_expense_breakdown"]
    }
    if second_breakdown != {
        "uncapitalized_transit_logistics": Decimal("0.00"),
        "wb_jam_subscription": Decimal("0.00"),
        "wb_paid_services": Decimal("0.00"),
        "review_points": Decimal("0.00"),
        "other_withholdings": Decimal("2500.00"),
    }:
        raise AssertionError(f"expense breakdown mismatch: {second_breakdown}")
    if sum(second_breakdown.values(), Decimal("0")) != Decimal("2500.00"):
        raise AssertionError("displayed category amounts do not conserve the main row")
    if report["other_expense_category_definitions"] != [
        {"key": "other_withholdings", "label": "Прочие удержания"}
    ]:
        raise AssertionError("zero-only Partner subrows were not omitted deterministically")
    with sqlite3.connect(block.db_path) as conn:
        rows = conn.execute(
            """SELECT nm_id,metrics_json FROM wb_finance_weekly_sku_aggregates
               WHERE seller_id=? AND week_start=? AND nm_id IN (?, '__account__')""",
            (block.seller_id, WEEK_TWO.isoformat(), str(TARGET_NM)),
        ).fetchall()
    metrics = {str(nm_id): json.loads(payload) for nm_id, payload in rows}
    if (
        metrics[str(TARGET_NM)]["marketing"] != "3000.0000"
        or metrics["__account__"]["marketing"] != "7000.0000"
        or metrics["__account__"]["review_points"] != "0.0000"
        or metrics["__account__"]["other_deductions"] != "5000.0000"
        or second["ads"] != "10000.0000"
        or second[OTHER_DIRECT_ALLOCATED_KEY] != "2500.0000"
    ):
        raise AssertionError(
            "Finance marketing/review storno was not separated from Partner ads/other"
        )
    if second["dividends"] != "4800.0000":
        raise AssertionError(f"second-week dividends mismatch: {second}")
    expected_period_annualized = (
        (Decimal("44253.9040") + Decimal("4800"))
        / Decimal("2")
        * Decimal("52")
        / Decimal("1000000")
        * Decimal("100")
    ).quantize(Decimal("0.0001"))
    if Decimal(str(report["totals"]["annualized_return_pct"])) != expected_period_annualized:
        raise AssertionError(f"period annualized return must use average weekly dividends: {report}")
    if "not guaranteed" not in report["annualized_return_formula"]:
        raise AssertionError("annualized return disclosure is absent")
    return report


def _assert_workbook(block: PartnerReportBlock, report: dict) -> None:
    body, filename, evidence = block.build_preview_workbook(
        {"nm_id": str(TARGET_NM), "selected_weeks": report["selected_weeks"]},
        expected_source_digest=report["source_digest"],
    )
    if not filename.endswith(".xlsx") or str(TARGET_NM) not in filename:
        raise AssertionError(f"XLSX filename does not bind SKU and period: {filename}")
    if evidence["source_digest"] != report["source_digest"]:
        raise AssertionError("UI and Excel source digests diverged")
    wb = load_workbook(BytesIO(body), data_only=False)
    if wb.sheetnames != ["Партнёрский отчёт", "Параметры"]:
        raise AssertionError(f"unexpected workbook structure: {wb.sheetnames}")
    if any(sheet.sheet_state != "visible" for sheet in wb.worksheets):
        raise AssertionError("workbook contains a hidden evidence surface")
    if getattr(wb, "vba_archive", None) is not None or getattr(wb, "_external_links", []):
        raise AssertionError("workbook contains macros or external links")
    ws = wb["Партнёрский отчёт"]
    if ws.freeze_panes != "C2" or not ws.print_area:
        raise AssertionError("reference-like freeze/print contract missing")
    labels = {str(ws.cell(row, 2).value): row for row in range(2, ws.max_row + 1)}
    for required in (
        "Агентское вознаграждение WB",
        "Эквайринг",
        OTHER_DIRECT_ALLOCATED_LABEL,
        "Дивиденды",
        "Расчётная годовая доходность инвестора, %",
    ):
        if required not in labels:
            raise AssertionError(f"required report row absent from XLSX: {required}")
    dividends_row = labels["Дивиденды"]
    roi_row = labels["Расчётная годовая доходность инвестора, %"]
    if ws.cell(dividends_row, 2).font.color.rgb[-6:] != "0070C0":
        raise AssertionError("dividends row lost the reference blue accent")
    if ws.cell(roi_row, 2).font.color.rgb[-6:] != "0070C0":
        raise AssertionError("annualized return row lost the reference blue accent")
    if ws.cell(labels["Налог"], 1).value != 0.06:
        raise AssertionError("actual tax coefficient is absent from the compact left column")
    if ws.cell(labels["На пополнение товарных остатков"], 1).value != 0.2:
        raise AssertionError("actual replenishment coefficient is absent from XLSX")
    if ws.cell(dividends_row, 1).value != 0.4:
        raise AssertionError("actual investor share is absent from XLSX")
    visible_categories = [
        (item["key"], item["label"])
        for item in report["other_expense_category_definitions"]
    ]
    category_labels = dict(visible_categories)
    for category_key, category_label in visible_categories:
        row_no = labels.get(category_label)
        if row_no is None:
            raise AssertionError(f"Partner XLSX category missing: {category_key}")
        if ws.cell(row_no, 1).value is not None:
            raise AssertionError(f"Partner XLSX exposed a category coefficient: {category_key}")
    hidden_labels = {
        label for key, label in OTHER_EXPENSE_CATEGORIES if key not in category_labels
    }
    if hidden_labels.intersection(labels):
        raise AssertionError("Partner XLSX rendered a zero-only expense subrow")
    main_row = labels[OTHER_DIRECT_ALLOCATED_LABEL]
    for column in range(3, 3 + len(report["weeks"]) + 1):
        category_sum = sum(
            Decimal(str(ws.cell(labels[label], column).value or 0))
            for label in category_labels.values()
        )
        if category_sum != Decimal(str(ws.cell(main_row, column).value or 0)):
            raise AssertionError(f"Partner XLSX displayed kopecks do not reconcile in column {column}")
    for key, label in REPORT_ROWS:
        ui_value = report["weeks"][0]["values"][key]
        excel_value = ws.cell(labels[label], 3).value
        if ui_value is None:
            if excel_value is not None:
                raise AssertionError(f"Excel masked missing {key} as a value")
        elif abs(
            (Decimal(str(excel_value)) * (Decimal("100") if key.endswith("_pct") else Decimal("1")))
            - Decimal(ui_value)
        ) > Decimal("0.0001"):
            raise AssertionError(f"UI/XLSX mismatch for {key}: {ui_value} vs {excel_value}")
    serialized_cells = "\n".join(
        str(cell.value or "") for sheet in wb.worksheets for row in sheet.iter_rows() for cell in row
    )
    for forbidden in (str(OTHER_NM), "Другой секретный товар", "VC202", "BAR202"):
        if forbidden in serialized_cells:
            raise AssertionError(f"other-SKU data leaked into Partner XLSX: {forbidden}")
    if body[:2] != b"PK":
        raise AssertionError("generated Excel is not a valid OOXML archive")
    artifact_path = os.environ.get("PARTNER_REPORT_SMOKE_XLSX", "").strip()
    if artifact_path:
        Path(artifact_path).write_bytes(body)

    try:
        block.build_preview_workbook(
            {"nm_id": str(TARGET_NM), "selected_weeks": report["selected_weeks"]},
            expected_source_digest="sha256:stale-ui",
        )
    except PartnerReportError as exc:
        if exc.code != "preview_source_digest_changed":
            raise
    else:
        raise AssertionError("Excel export ignored UI/source digest drift")


def _assert_only_marketing_workbook(block: PartnerReportBlock) -> None:
    preview = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]}
    )
    if preview["other_expense_category_definitions"]:
        raise AssertionError("marketing-only Partner week exposed a zero subcategory")
    if preview["weeks"][0]["values"][OTHER_DIRECT_ALLOCATED_KEY] != "0.0000":
        raise AssertionError("Finance marketing leaked into Partner other expenses")
    body, _filename, _evidence = block.build_preview_workbook(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]},
        expected_source_digest=preview["source_digest"],
    )
    workbook = load_workbook(BytesIO(body), data_only=False)
    labels = {
        str(row[1].value)
        for row in workbook["Партнёрский отчёт"].iter_rows(min_row=2)
        if len(row) > 1 and row[1].value not in (None, "")
    }
    forbidden = {label for _key, label in OTHER_EXPENSE_CATEGORIES}
    workbook.close()
    if labels.intersection(forbidden):
        raise AssertionError("marketing-only Partner XLSX rendered a zero expense subrow")


def _assert_partial_cost_surface(block: PartnerReportBlock) -> None:
    missing = _sale(
        29,
        WEEK_TWO + timedelta(days=1),
        TARGET_NM,
        revenue="500",
        for_pay="500",
        acquiring="0",
    )
    missing.update(
        {
            "reportId": 29,
            "rrdId": 29,
            "deliveryType": "fbs",
            "rid": "synthetic-partner-unresolved-fbs",
        }
    )
    block.finance.ingest_week(
        WEEK_TWO,
        WEEK_TWO + timedelta(days=6),
        [*_week_two_finance_rows(), missing],
    )
    partial = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_TWO.isoformat()]}
    )
    if partial["status"] != "ready" or any(
        item["code"] == "partner_cost_coverage_incomplete"
        for item in partial["blockers"]
    ):
        raise AssertionError(f"partial cost coverage incorrectly blocked Partner: {partial}")
    week = partial["weeks"][0]
    values = week["values"]
    if (
        values["net_revenue"] != "200500.0000"
        or values["sales_without_cost_rub"] != "500.0000"
        or values["orders_without_cost"] != "1.0000"
        or values["units_without_cost"] != "1.0000"
        or values["cogs"] != "100000.0000"
        or values["estimated_tax"] != "12000.0000"
    ):
        raise AssertionError(f"Partner partial metrics are not coverage-bound: {partial}")
    cost = week["coverage"]["cost"]
    if (
        cost["profit_coverage_status"] != "partial"
        or cost["covered_sales_revenue_rub"] != "200000.0000"
        or cost["uncovered_sales_revenue_rub"] != "500.0000"
        or cost["uncovered_sales_order_count"] != 1
    ):
        raise AssertionError(f"Partner coverage evidence missing: {partial}")


def _assert_incomplete_and_stale_states(block: PartnerReportBlock) -> None:
    missing_day = (WEEK_ONE + timedelta(days=3)).isoformat()
    with sqlite3.connect(block.db_path) as conn:
        saved = conn.execute(
            """SELECT payload_json FROM temporal_source_slot_snapshots
               WHERE source_key='ads_compact' AND snapshot_date=?""",
            (missing_day,),
        ).fetchone()[0]
        conn.execute(
            "DELETE FROM temporal_source_slot_snapshots WHERE source_key='ads_compact' AND snapshot_date=?",
            (missing_day,),
        )
        conn.commit()
    incomplete = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]}
    )
    if incomplete["status"] != "incomplete":
        raise AssertionError(f"missing ads date did not block preview: {incomplete}")
    if not any(item["code"] == "ads_date_missing" and item["date"] == missing_day for item in incomplete["blockers"]):
        raise AssertionError(f"missing ads blocker is not specific: {incomplete}")
    values = incomplete["weeks"][0]["values"]
    if values["net_revenue"] is None or values["ads"] is not None or values["finance_margin"] is not None:
        raise AssertionError("partial preview did not preserve available values / missing state")
    try:
        block.build_preview_workbook(
            {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]}
        )
    except PartnerReportError as exc:
        if exc.code != "source_coverage_incomplete":
            raise
    else:
        raise AssertionError("Excel remained available for incomplete preview")
    with sqlite3.connect(block.db_path) as conn:
        conn.execute(
            "INSERT INTO temporal_source_slot_snapshots VALUES(?,?,?,?,?)",
            ("ads_compact", missing_day, "accepted_closed_day_snapshot", missing_day + "T23:00:00Z", saved),
        )
        conn.execute(
            """UPDATE sheet_vitrina_v1_warehouse_wb_daily_cost
               SET fingerprint='sha256:corrected-cost'
               WHERE as_of_date='2026-07-07' AND nm_id=?""",
            (TARGET_NM,),
        )
        conn.commit()
    stale = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]}
    )
    if not any(item["code"] == "finance_sku_aggregate_cost_stale" for item in stale["blockers"]):
        raise AssertionError(f"canonical cost correction did not invalidate aggregate: {stale}")
    block.finance.recalculate_week(WEEK_ONE, WEEK_ONE + timedelta(days=6))
    with sqlite3.connect(block.db_path) as conn:
        row = conn.execute(
            """SELECT coverage_json FROM wb_finance_weekly_sku_aggregates
               WHERE nm_id=? AND week_start=?""",
            (str(TARGET_NM), WEEK_ONE.isoformat()),
        ).fetchone()
        coverage = json.loads(str(row[0]))
        coverage["detail_rows"][0]["formula_version"] = "obsolete_cost_formula"
        conn.execute(
            """UPDATE wb_finance_weekly_sku_aggregates SET coverage_json=?
               WHERE nm_id=? AND week_start=?""",
            (
                json.dumps(coverage, ensure_ascii=False, sort_keys=True),
                str(TARGET_NM),
                WEEK_ONE.isoformat(),
            ),
        )
        conn.commit()
    formula_stale = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]}
    )
    if not any(
        item["code"] == "finance_sku_aggregate_cost_stale"
        for item in formula_stale["blockers"]
    ):
        raise AssertionError(
            f"old canonical cost formula remained ready: {formula_stale}"
        )
    block.finance.recalculate_week(WEEK_ONE, WEEK_ONE + timedelta(days=6))


def _assert_negative_profit_and_validation(block: PartnerReportBlock) -> None:
    negative_settings = _settings()
    negative_settings["weekly_office_expense_rub"] = "500000"
    block.save_settings(negative_settings, actor="operator@example.test")
    negative = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]}
    )
    values = negative["weeks"][0]["values"]
    if Decimal(str(values["net_profit"])) >= 0 or values["dividends"] != "0.0000":
        raise AssertionError(f"loss must remain visible without negative dividends: {negative}")
    invalid = _settings()
    invalid["invested_capital_rub"] = "0"
    try:
        block.save_settings(invalid, actor="operator@example.test")
    except PartnerReportError as exc:
        if exc.code != "settings_invalid" or "greater than zero" not in str(exc):
            raise
    else:
        raise AssertionError("zero invested capital was accepted")
    block.save_settings(_settings(), actor="operator@example.test")


def _assert_indexed_performance(block: PartnerReportBlock) -> dict[str, int]:
    with sqlite3.connect(block.db_path) as conn:
        conn.execute(
            """WITH RECURSIVE seq(n) AS (
                   SELECT 1 UNION ALL SELECT n+1 FROM seq WHERE n<295919
               )
               INSERT INTO wb_finance_weekly_raw_rows(
                   seller_id,report_id,rrd_id,report_type,week_start,week_end,nm_id,
                   vendor_code,barcode,doc_type_name,seller_oper_name,row_hash,raw_json,
                   first_seen_at,updated_at
               )
               SELECT 'seller-1','perf-' || n,'perf-' || n,1,'2025-01-06','2025-01-12',
                      '202202','VC202','BAR202','Продажа','Продажа','hash-' || n,
                      '{"nmId":202202}','2026-07-20T00:00:00Z','2026-07-20T00:00:00Z'
               FROM seq"""
        )
        conn.commit()
    raw_scan_started = time.perf_counter()
    with sqlite3.connect(block.db_path) as conn:
        raw_rows = conn.execute(
            "SELECT raw_json FROM wb_finance_weekly_raw_rows WHERE seller_id='seller-1'"
        ).fetchall()
        decoded_nm_ids = [
            str(json.loads(str(row[0] or "{}" )).get("nmId") or "") for row in raw_rows
        ]
    raw_scan_ms = int((time.perf_counter() - raw_scan_started) * 1000)
    if len(raw_rows) < 295919 or not decoded_nm_ids:
        raise AssertionError("production-like full-scan baseline was not exercised")
    started = time.perf_counter()
    report = block.preview(
        {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat(), WEEK_TWO.isoformat()]}
    )
    elapsed = time.perf_counter() - started
    if report["status"] != "ready" or elapsed >= 2.0 or report["performance"]["duration_ms"] >= 2000:
        raise AssertionError(f"indexed preview performance regression: elapsed={elapsed}, report={report}")
    return {
        "raw_scan_ms": raw_scan_ms,
        "indexed_preview_ms": int(report["performance"]["duration_ms"]),
        "raw_rows": len(raw_rows),
    }


def _assert_removed_package_surface(block: PartnerReportBlock) -> None:
    for action in (
        lambda: block.build_preview_package(
            {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]}
        ),
        lambda: block.build_finalized_package("any"),
        lambda: block.finalize(
            {"nm_id": str(TARGET_NM), "selected_weeks": [WEEK_ONE.isoformat()]},
            actor="operator@example.test",
        ),
    ):
        try:
            action()
        except PartnerReportError as exc:
            if exc.code not in {"partner_package_removed", "partner_finalization_removed"}:
                raise
        else:
            raise AssertionError("removed ZIP/raw/finalization surface remained callable")


def _seed_sources(db_path: Path) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE sheet_vitrina_v1_nomenclature_items(
                is_active INTEGER,nm_id INTEGER,vendor_code TEXT,barcode TEXT,
                barcodes_json TEXT,product_type TEXT,nomenclature_name TEXT,
                wb_title TEXT,is_hidden INTEGER,created_at TEXT,our_sku TEXT,
                aliases_json TEXT,match_key TEXT
            );
            INSERT INTO sheet_vitrina_v1_nomenclature_items VALUES
                (1,101101,'VC101','BAR101','["BAR101"]','other','Выбранный товар',
                 'Выбранный товар WB',0,'2026-01-01','OUR101','[]','target'),
                (0,202202,'VC202','BAR202','["BAR202"]','other','Другой секретный товар',
                 'Другой секретный товар WB',1,'2026-01-01','OUR202','[]','other');
            CREATE TABLE sheet_vitrina_v1_warehouse_functional_cutovers(
                cutover_id TEXT PRIMARY KEY,cutover_at TEXT,status TEXT,
                plan_fingerprint TEXT,source_watermarks_json TEXT,
                absorbed_supply_revisions_json TEXT,backup_json TEXT,
                created_at TEXT,updated_at TEXT
            );
            INSERT INTO sheet_vitrina_v1_warehouse_functional_cutovers VALUES(
                'warehouse_functional_cutover_v1','2026-07-01T00:00:00Z','posted',
                'sha256:cutover','{}','[]','{}','2026-07-01T00:00:00Z','2026-07-01T00:00:00Z'
            );
            CREATE TABLE sheet_vitrina_v1_warehouse_wb_daily_cost(
                cutover_id TEXT,as_of_date TEXT,nm_id INTEGER,quantity TEXT,wac_rub TEXT,
                capital_rub TEXT,quality TEXT,provenance_json TEXT,fingerprint TEXT,
                created_at TEXT,PRIMARY KEY(cutover_id,as_of_date,nm_id)
            );
            INSERT INTO sheet_vitrina_v1_warehouse_wb_daily_cost VALUES
                ('warehouse_functional_cutover_v1','2026-07-07',101101,'10','83837','838370','certified','{}','sha256:target-jul7','2026-07-07T00:00:00Z'),
                ('warehouse_functional_cutover_v1','2026-07-14',101101,'10','100000','1000000','certified','{}','sha256:target-jul14','2026-07-14T00:00:00Z'),
                ('warehouse_functional_cutover_v1','2026-07-14',202202,'10','50000','500000','certified','{}','sha256:other-jul14','2026-07-14T00:00:00Z');
            CREATE TABLE temporal_source_slot_snapshots(
                source_key TEXT NOT NULL,snapshot_date TEXT NOT NULL,
                snapshot_role TEXT NOT NULL,captured_at TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                PRIMARY KEY(source_key,snapshot_date,snapshot_role)
            );
            """
        )
        conn.commit()


def _seed_finance(block: PartnerReportBlock) -> None:
    block.finance.ingest_week(
        WEEK_ONE,
        WEEK_ONE + timedelta(days=6),
        [
            _sale(
                1,
                WEEK_ONE + timedelta(days=1),
                TARGET_NM,
                revenue="476034",
                for_pay="301237",
                acquiring="14797",
            ),
            _deduction(101, WEEK_ONE + timedelta(days=2), TARGET_NM, "500", "WB Продвижение"),
            _deduction(102, WEEK_ONE + timedelta(days=2), 0, "700", "Оказание услуг «WB Продвижение»"),
        ],
    )
    block.finance.ingest_week(
        WEEK_TWO,
        WEEK_TWO + timedelta(days=6),
        _week_two_finance_rows(),
    )


def _week_two_finance_rows() -> list[dict]:
    return [
        _sale(2, WEEK_TWO + timedelta(days=1), TARGET_NM, revenue="200000", for_pay="155000", acquiring="5000"),
        _sale(3, WEEK_TWO + timedelta(days=1), OTHER_NM, revenue="200000", for_pay="170000", acquiring="4000"),
        {
            **_sale(4, WEEK_TWO + timedelta(days=1), 0, revenue="0", for_pay="0", acquiring="0"),
            "nmId": 0,
            "vendorCode": "",
            "sku": "",
            "docTypeName": "",
            "sellerOperName": "Удержание",
            "quantity": 0,
            "deduction": "5000",
            "bonusTypeName": "Общее удержание продавца",
        },
        _deduction(5, WEEK_TWO + timedelta(days=2), TARGET_NM, "3000", "WB Продвижение"),
        _deduction(6, WEEK_TWO + timedelta(days=2), 0, "7000", "Оказание услуг «WB Продвижение»"),
        _deduction(7, WEEK_TWO + timedelta(days=2), 0, "1000", "Аванс за услугу \"Баллы за отзывы\""),
        _deduction(8, WEEK_TWO + timedelta(days=2), 0, "-1000", "Возврат неиспользованного остатка аванса за услугу \"Баллы за отзывы\""),
    ]


def _seed_ads(db_path: Path) -> None:
    with sqlite3.connect(db_path) as conn:
        for week, total, nested in ((WEEK_ONE, Decimal("30904"), False), (WEEK_TWO, Decimal("10000"), True)):
            for offset in range(7):
                day = (week + timedelta(days=offset)).isoformat()
                value = total if offset == 0 else Decimal("0")
                result = {
                    "kind": "success",
                    "snapshot_date": day,
                    "items": [{"nm_id": TARGET_NM, "ads_sum": str(value), "advert_id": "safe-target"}],
                }
                payload = {"result": result} if nested else result
                conn.execute(
                    "INSERT INTO temporal_source_slot_snapshots VALUES(?,?,?,?,?)",
                    (
                        "ads_compact",
                        day,
                        "accepted_closed_day_snapshot",
                        day + "T23:00:00Z",
                        json.dumps(payload, separators=(",", ":")),
                    ),
                )
        conn.commit()


def _sale(
    rrd_id: int,
    operation_date: date,
    nm_id: int,
    *,
    revenue: str,
    for_pay: str,
    acquiring: str,
) -> dict:
    return {
        "dateFrom": operation_date.isoformat(),
        "dateTo": operation_date.isoformat(),
        "reportId": rrd_id,
        "reportType": 1,
        "rrdId": rrd_id,
        "nmId": nm_id,
        "vendorCode": f"VC{str(nm_id)[:3]}" if nm_id else "",
        "sku": f"BAR{str(nm_id)[:3]}" if nm_id else "",
        "rrDate": operation_date.isoformat(),
        "saleDt": operation_date.isoformat(),
        "docTypeName": "Продажа",
        "sellerOperName": "Продажа",
        "quantity": 1,
        "retailPriceWithDisc": revenue,
        "forPay": for_pay,
        "acquiringFee": acquiring,
    }


def _deduction(
    rrd_id: int,
    operation_date: date,
    nm_id: int,
    amount: str,
    name: str,
) -> dict:
    return {
        **_sale(
            rrd_id,
            operation_date,
            nm_id,
            revenue="0",
            for_pay="0",
            acquiring="0",
        ),
        "docTypeName": "",
        "sellerOperName": "Удержание",
        "quantity": 0,
        "deduction": amount,
        "bonusTypeName": name,
    }


def _settings() -> dict[str, str]:
    return {
        "nm_id": str(TARGET_NM),
        "partner_share_pct": "40",
        "invested_capital_rub": "1000000",
        "replenishment_reserve_pct": "20",
        "weekly_office_expense_rub": "10000",
        "tax_rate_pct": "6",
        "common_expense_rule": COMMON_EXPENSE_RULE,
    }


if __name__ == "__main__":
    main()
