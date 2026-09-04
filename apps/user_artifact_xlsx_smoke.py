"""Regression coverage for the user-artifact XLSX execution contract."""

from __future__ import annotations

import json
from pathlib import Path
import sys
import tempfile


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.user_artifact_xlsx import (  # noqa: E402
    TabularXlsxSpec,
    create_tabular_xlsx,
    verify_tabular_xlsx,
    xlsx_creation_plan,
)


def _assert_artifact_tool_failure_falls_back_without_data_loss() -> None:
    synthetic = {
        "sheet_name": "Номенклатура",
        "headers": ["Артикул", "nmID", "Штрихкод"],
        "rows": [
            ["Тест A | Barcode: 001234 | nmID: 0042", "0042", "001234"],
            ["Тест B | Barcode: 9000000000007 | nmID: 100", "100", "9000000000007"],
        ],
        "text_columns": [0, 1, 2],
        "freeze_header": True,
        "auto_filter": True,
        "column_widths": [52, 16, 20],
    }
    frozen_source = json.dumps(synthetic, ensure_ascii=False, sort_keys=True)
    simulated_error = ModuleNotFoundError("No module named '@oai/artifact-tool'")
    assert xlsx_creation_plan(artifact_tool_import_error=None) == ("artifact-tool",)
    plan = xlsx_creation_plan(artifact_tool_import_error=simulated_error)
    assert plan == ("openpyxl", "xlsxwriter", "stdlib")

    spec = TabularXlsxSpec.from_mapping(synthetic)
    try:
        create_tabular_xlsx(ROOT / "forbidden.xlsx", spec)
    except ValueError as exc:
        assert "outside the repository" in str(exc)
    else:
        raise AssertionError("repo output boundary was not enforced")
    with tempfile.TemporaryDirectory(prefix="wb-core-user-artifact-") as directory:
        output = Path(directory) / "fallback.xlsx"
        result = create_tabular_xlsx(output, spec)
        assert result["backend"] == "stdlib"
        assert result["data_rows"] == 2 and result["columns"] == 3
        assert result["zip_integrity"] == "ok" and result["xml_integrity"] == "ok"
        assert output.is_file() and output.stat().st_size > 0
        verify_tabular_xlsx(output, spec)

        import openpyxl

        workbook = openpyxl.load_workbook(output, data_only=False, read_only=False)
        assert workbook.sheetnames == ["Номенклатура"]
        sheet = workbook["Номенклатура"]
        assert sheet.max_row == 3 and sheet.max_column == 3
        assert sheet.freeze_panes == "A2" and sheet.auto_filter.ref == "A1:C3"
        assert sheet["B2"].value == "0042" and sheet["C2"].value == "001234"
        assert sheet["B2"].data_type == "s" and sheet["C2"].data_type == "s"
        assert sheet["B2"].number_format == "@" and sheet["C2"].number_format == "@"

    assert json.dumps(synthetic, ensure_ascii=False, sort_keys=True) == frozen_source


def main() -> int:
    _assert_artifact_tool_failure_falls_back_without_data_loss()
    print("user_artifact_xlsx_smoke: ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
