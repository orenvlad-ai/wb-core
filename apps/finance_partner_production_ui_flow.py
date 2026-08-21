#!/usr/bin/env python3
"""Read-only authenticated production acceptance for Finance and Partner reports."""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
from io import BytesIO
import json
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
import zipfile

from openpyxl import load_workbook
from playwright.sync_api import expect, sync_playwright

from packages.application.partner_report import (
    OTHER_DIRECT_ALLOCATED_KEY,
    OTHER_DIRECT_ALLOCATED_LABEL,
    OTHER_EXPENSE_CATEGORIES,
    REPORT_ROWS,
)


REPORTS_PATH = "/sheet-vitrina-v1/vitrina?tab=reports"
FINANCE_API_PATH = "/v1/sheet-vitrina-v1/wb-finance-report"
PARTNER_OPTIONS_API_PATH = "/v1/sheet-vitrina-v1/partner-report/options"
PARTNER_PREVIEW_API_PATH = "/v1/sheet-vitrina-v1/partner-report/preview"
PARTNER_PREVIEW_XLSX_API_PATH = "/v1/sheet-vitrina-v1/partner-report/preview.xlsx"


def _validate_finance_storage_health(
    storage_health: dict[str, Any],
) -> str:
    state = str(storage_health.get("state") or "")
    canonical_source = str(storage_health.get("canonical_source") or "")
    raw = dict(storage_health.get("raw") or {})
    operational = dict(storage_health.get("operational") or {})
    if state == "monolith" and canonical_source == "monolith":
        implicit = storage_health.get("implicit_manifest") is True
        if implicit:
            _assert(
                raw.get("generation_id") == "monolith"
                and operational.get("generation_id") == "monolith",
                "Finance logical stores expose the exact implicit monolith",
            )
        else:
            generation_epoch = str(
                storage_health.get("generation_epoch") or ""
            )
            _assert(
                generation_epoch.startswith("rollback-")
                and raw.get("generation_epoch") == generation_epoch
                and operational.get("generation_epoch")
                == generation_epoch
                and raw.get("generation_id")
                == operational.get("generation_id")
                and raw.get("relative_path")
                == operational.get("relative_path")
                and raw.get("exists") is True
                and operational.get("exists") is True,
                "Finance rollback monolith binds one exact selected file",
            )
            _assert(
                storage_health.get("raw_schema_ready") is True
                and storage_health.get("raw_schema_mode")
                == "legacy_monolith"
                and storage_health.get("operational_schema_ready") is True
                and storage_health.get("cursor_contract")
                == "not_applicable_monolith"
                and storage_health.get("cursor_mismatch") is False,
                "Finance rollback monolith exposes healthy legacy raw and "
                "non-applicable split cursors",
            )
        _assert(
            storage_health.get("rollback_ready") is True
            and storage_health.get("cutover_ready") is False,
            "Finance monolith rollback is ready and cutover is not selected",
        )
        return (
            "implicit_monolith"
            if implicit
            else "selected_rollback_monolith"
        )

    _assert(
        state == "cutover"
        and canonical_source == "split"
        and storage_health.get("implicit_manifest") is False,
        "Finance storage is either the implicit monolith or an exact selected split",
    )
    generation_epoch = str(storage_health.get("generation_epoch") or "")
    _assert(
        len(generation_epoch) == 20
        and all(character in "0123456789abcdef" for character in generation_epoch),
        "Finance split exposes an exact generation epoch",
    )
    _assert(
        storage_health.get("rollback_generation_id") == "monolith",
        "Finance split retains the monolith rollback generation",
    )
    _assert(
        raw.get("exists") is True
        and raw.get("generation_epoch") == generation_epoch
        and raw.get("generation_id") == f"finance-raw-{generation_epoch}"
        and raw.get("schema_revision") == "finance_raw_v1"
        and int(raw.get("size_bytes") or 0) > 0,
        "Finance split raw store matches the selected generation",
    )
    _assert(
        operational.get("exists") is True
        and operational.get("generation_epoch") == generation_epoch
        and operational.get("generation_id")
        == f"operational-{generation_epoch}"
        and operational.get("schema_revision") == "operational_v1"
        and int(operational.get("size_bytes") or 0) > 0,
        "Finance split operational store matches the selected generation",
    )
    _assert(
        storage_health.get("raw_schema_ready") is True
        and storage_health.get("operational_schema_ready") is True
        and not storage_health.get("raw_health_error")
        and not storage_health.get("operational_health_error"),
        "Finance split schemas are healthy",
    )
    _assert(
        int(storage_health.get("latest_outbox_sequence") or 0)
        == int(storage_health.get("raw_ack_cursor") or 0)
        == int(storage_health.get("operational_cursor") or 0)
        and int(storage_health.get("consumer_lag_events") or 0) == 0
        and storage_health.get("cursor_mismatch") is False,
        "Finance split consumers are caught up on one exact cursor",
    )
    _assert(
        int(storage_health.get("shadow_mismatch_count") or 0) == 0
        and int(storage_health.get("actionable_dead_letters") or 0) == 0
        and int(
            (storage_health.get("raw_counts") or {}).get("pending_outbox")
            or 0
        )
        == 0
        and int(
            (storage_health.get("operational_counts") or {}).get(
                "dead_letters"
            )
            or 0
        )
        == 0,
        "Finance split has no mismatch, pending outbox, or dead letter",
    )
    _assert(
        storage_health.get("rollback_ready") is False
        and storage_health.get("cutover_ready") is False,
        "Finance split is selected and awaits the separate rollback candidate gate",
    )
    return "selected_split"


def _validate_finance_storage_card(
    storage_text: str,
    *,
    storage_health: dict[str, Any],
    storage_phase: str,
) -> None:
    visible = str(storage_text or "")
    if storage_phase in {
        "implicit_monolith",
        "selected_rollback_monolith",
    }:
        required = (
            "canonical: monolith",
            "rollback: готов",
            "cutover: не разрешён/не готов",
        )
    else:
        _assert(
            storage_phase == "selected_split",
            "Finance storage card uses a validated lifecycle phase",
        )
        raw = dict(storage_health.get("raw") or {})
        operational = dict(storage_health.get("operational") or {})
        required = (
            "canonical: split",
            f"Raw generation: {raw.get('generation_id')}",
            f"schema {raw.get('schema_revision')}",
            f"Operational generation: {operational.get('generation_id')}",
            f"schema {operational.get('schema_revision')}",
            "Cursor raw/operational: "
            f"{int(storage_health.get('raw_ack_cursor') or 0)} / "
            f"{int(storage_health.get('operational_cursor') or 0)}",
            f"consumer lag: {int(storage_health.get('consumer_lag_events') or 0)}",
            (
                "live-tail cursor/lag: "
                f"{int(storage_health.get('live_tail_cursor') or 0)} / "
                f"{int(storage_health.get('live_tail_lag_events') or 0)}"
                if storage_health.get("live_tail_applicable") is not False
                else "live-tail: не применяется после cutover "
                "(исторический cursor: "
                f"{int(storage_health.get('live_tail_cursor') or 0)})"
            ),
            f"mismatches: {int(storage_health.get('shadow_mismatch_count') or 0)}",
            f"dead letters: {int(storage_health.get('actionable_dead_letters') or 0)}",
            "rollback: не доказан",
            "cutover: не разрешён/не готов",
        )
    missing = [item for item in required if item not in visible]
    _assert(
        not missing,
        "Finance storage health card is missing validated lifecycle evidence: "
        + ", ".join(missing),
    )


def run_finance_partner_ui_flow(
    *,
    base_url: str,
    auth_cookie: str,
    evidence_dir: Path,
    headless: bool = True,
    deployed_sha: str = "",
) -> dict[str, Any]:
    normalized_base_url = str(base_url or "").strip().rstrip("/")
    parsed = urlparse(normalized_base_url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("Finance UI flow requires an absolute http(s) base URL")
    cookie_name, separator, cookie_value = str(auth_cookie or "").partition("=")
    if separator != "=" or cookie_name != "wb_core_web_session" or not cookie_value:
        raise ValueError("Finance UI flow requires a valid app-session cookie")
    if deployed_sha and (
        len(deployed_sha) != 40
        or any(character not in "0123456789abcdef" for character in deployed_sha.casefold())
    ):
        raise ValueError("deployed_sha must be an exact 40-character Git commit")

    evidence_dir = evidence_dir.resolve()
    evidence_dir.mkdir(parents=True, exist_ok=True)
    requested_url = normalized_base_url + REPORTS_PATH
    report_path = evidence_dir / "finance_partner_ui_flow_report.json"
    page_errors: list[str] = []
    console_errors: list[str] = []
    server_errors: list[dict[str, Any]] = []
    client_errors: list[dict[str, Any]] = []
    failed_requests: list[dict[str, Any]] = []
    document_responses: list[dict[str, Any]] = []
    non_read_requests: list[dict[str, str]] = []
    screenshots: list[str] = []
    preview_evidence: dict[str, Any] = {
        "attempted": False,
        "ready": False,
        "status": "not_attempted",
        "reason": "Partner preview has not been attempted",
        "blockers": [],
        "xlsx": {
            "downloaded": False,
            "path": "",
            "sha256": "",
            "size_bytes": 0,
            "verification": {"passed": False, "findings": ["not_downloaded"]},
        },
    }

    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=headless)
            context = browser.new_context(viewport={"width": 1600, "height": 1000})
            context.add_cookies(
                [
                    {
                        "name": cookie_name,
                        "value": cookie_value,
                        "url": normalized_base_url,
                        "httpOnly": True,
                        "sameSite": "Lax",
                    }
                ]
            )
            page = context.new_page()
            page.on("pageerror", lambda error: page_errors.append(str(error)))
            page.on(
                "console",
                lambda message: console_errors.append(message.text)
                if message.type == "error"
                else None,
            )
            page.on(
                "response",
                lambda response: server_errors.append(
                    {
                        "status": response.status,
                        "url": response.url,
                        "resource_type": response.request.resource_type,
                    }
                )
                if response.status >= 500
                else None,
            )
            page.on(
                "response",
                lambda response: client_errors.append(
                    {
                        "status": response.status,
                        "url": response.url,
                        "resource_type": response.request.resource_type,
                    }
                )
                if 400 <= response.status < 500
                else None,
            )
            page.on(
                "requestfailed",
                lambda request: failed_requests.append(
                    {
                        "method": request.method,
                        "url": request.url,
                        "failure": str(request.failure or ""),
                    }
                ),
            )
            page.on(
                "response",
                lambda response: document_responses.append(
                    {
                        "status": response.status,
                        "url": response.url,
                        "redirected_from": (
                            response.request.redirected_from.url
                            if response.request.redirected_from is not None
                            else ""
                        ),
                    }
                )
                if response.request.resource_type == "document"
                else None,
            )
            page.on(
                "request",
                lambda request: non_read_requests.append(
                    {"method": request.method, "url": request.url}
                )
                if request.method not in {"GET", "HEAD", "OPTIONS"}
                else None,
            )

            response = page.goto(
                requested_url,
                wait_until="domcontentloaded",
                timeout=120_000,
            )
            _assert(response is not None and response.status == 200, "reports page HTTP 200")
            _assert(page.url == requested_url, "reports page has no unexpected redirect")
            page.locator('[data-unified-tab-panel="reports"]:not([hidden])').wait_for(
                timeout=60_000
            )
            _assert(bool(page.title().strip()), "reports document title is non-empty")
            _assert(bool(page.locator("body").inner_text().strip()), "reports body is non-empty")

            reports = page.frame_locator('iframe[data-operator-embed-frame="reports"]')
            reports.locator('[data-report-section-button="wb-finance"]').click()
            reports.locator("#wbFinanceReportTableWrap table").wait_for(timeout=120_000)

            finance_payload = _protected_json_get(
                context,
                normalized_base_url + FINANCE_API_PATH,
                label="Finance weekly API",
            )
            finance_weeks = list(finance_payload.get("weeks") or [])
            _assert(finance_payload.get("status") == "ok", "Finance API status")
            _assert(bool(finance_weeks), "Finance API contains persisted weeks")
            storage_health = dict(finance_payload.get("storage_health") or {})
            storage_phase = _validate_finance_storage_health(storage_health)
            finance_facts = reports.locator("body").evaluate(
                """
                () => {
                  const wrap = document.getElementById('wbFinanceReportTableWrap');
                  const table = wrap && wrap.querySelector('table');
                  const rows = table ? [...table.querySelectorAll('tbody tr')] : [];
                  const value = (label) => {
                    const row = rows.find((item) => item.cells[0] && item.cells[0].innerText.trim() === label);
                    return row ? [...row.cells].slice(1).map((cell) => cell.innerText.trim()) : [];
                  };
                  const bodyStyle = getComputedStyle(document.body);
                  return {
                    status: document.getElementById('wbFinanceReportStatus').innerText.trim(),
                    headerCount: table ? table.querySelectorAll('thead th').length : 0,
                    rowCount: rows.length,
                    cogs: value('Себестоимость наша'),
                    profit: value('Прибыль после себестоимости'),
                    margin: value('Итоговая рентабельность, %'),
                    agent: value('Агентское вознаграждение WB'),
                    acquiring: value('Эквайринг'),
                    withMarketing: value('Расходы WB с маркетингом'),
                    noMarketing: value('Расходы WB без маркетинга'),
                    hasTechnicalExpenseRow: value('Расходы периода, учитываемые в прибыли').length > 0,
                    hasDuplicatePercentRow: value('Расходы WB, % от чистой выручки').length > 0,
                    expenseCells: [...table.querySelectorAll('.wb-finance-expense-share')].map((cell) => cell.innerText.trim()),
                    notes: document.getElementById('wbFinanceReportNotes').innerText.trim(),
                    storageHealth: document.getElementById('wbFinanceStorageHealth').innerText.trim(),
                    horizontalOverflow: Boolean(wrap && wrap.scrollWidth > wrap.clientWidth),
                    bodyBackground: bodyStyle.backgroundColor,
                  };
                }
                """
            )
            _assert(
                int(finance_facts["headerCount"]) == len(finance_weeks) + 1,
                "Finance table renders every API week",
            )
            _assert(int(finance_facts["rowCount"]) > 20, "Finance metric matrix is rendered")
            _assert(
                len(finance_facts["cogs"]) == len(finance_weeks)
                and len(finance_facts["profit"]) == len(finance_weeks)
                and len(finance_facts["margin"]) == len(finance_weeks)
                and len(finance_facts["agent"]) == len(finance_weeks)
                and len(finance_facts["acquiring"]) == len(finance_weeks)
                and len(finance_facts["withMarketing"]) == len(finance_weeks)
                and len(finance_facts["noMarketing"]) == len(finance_weeks),
                "Finance canonical metrics align to all weeks",
            )
            _assert(
                not finance_facts["hasTechnicalExpenseRow"]
                and not finance_facts["hasDuplicatePercentRow"],
                "Finance removes technical and duplicate percentage rows",
            )
            _assert(
                all("п.п." not in value for value in finance_facts["expenseCells"]),
                "Finance expense microcells contain no numeric percentage-point delta",
            )
            _assert("недоступен" not in finance_facts["status"].casefold(), "Finance UI is available")
            _assert("стоимость того же nmid на 01.07" in finance_facts["notes"].casefold(), "Finance temporal method is visible")
            _assert("retro-map" in finance_facts["notes"], "Finance rejects an independent retro-cost source")
            _validate_finance_storage_card(
                finance_facts["storageHealth"],
                storage_health=storage_health,
                storage_phase=storage_phase,
            )
            _assert(bool(finance_facts["horizontalOverflow"]), "Finance table scrolls locally")
            finance_screenshot = evidence_dir / "finance_weekly_desktop.png"
            page.screenshot(path=str(finance_screenshot), full_page=True)
            screenshots.append(str(finance_screenshot))

            reports.locator('[data-report-section-button="partner"]').click()
            reports.locator("#partnerReportControls").wait_for(timeout=120_000)
            partner_payload = _protected_json_get(
                context,
                normalized_base_url + PARTNER_OPTIONS_API_PATH,
                label="Partner report options API",
            )
            partner_facts = reports.locator("body").evaluate(
                """
                () => {
                  const controls = document.getElementById('partnerReportControls');
                  const settings = controls.querySelector('.partner-report-settings');
                  const wrap = document.getElementById('partnerReportTableWrap');
                  return {
                    status: document.getElementById('partnerReportStatus').innerText.trim(),
                    cardOptions: document.getElementById('partnerReportNmId').options.length,
                    weekOptions: document.querySelectorAll('#partnerReportWeekList input[type=checkbox]').length,
                    saveVisible: !document.getElementById('partnerReportSaveSettings').hidden,
                    generateVisible: !document.getElementById('partnerReportGenerate').hidden,
                    downloadDisabled: document.getElementById('partnerReportDownload').disabled,
                    hasFinalize: Boolean(document.getElementById('partnerReportFinalize')),
                    hasPackage: Boolean(document.getElementById('partnerReportPackage')),
                    settingsColumns: getComputedStyle(settings).gridTemplateColumns.split(' ').length,
                  };
                }
                """
            )
            _assert(
                int(partner_facts["cardOptions"]) == len(partner_payload.get("cards") or []),
                "Partner card options reconcile to API",
            )
            _assert(
                int(partner_facts["weekOptions"]) == len(partner_payload.get("weeks") or []),
                "Partner week options reconcile to API",
            )
            _assert("недоступен" not in partner_facts["status"].casefold(), "Partner UI is available")
            _assert(
                bool(partner_facts["saveVisible"])
                and bool(partner_facts["generateVisible"])
                and bool(partner_facts["downloadDisabled"])
                and not bool(partner_facts["hasFinalize"])
                and not bool(partner_facts["hasPackage"]),
                "Partner actions expose UI preview and XLSX without finalization/ZIP",
            )
            preview_required_fields = (
                "partner_share_pct",
                "invested_capital_rub",
                "replenishment_reserve_pct",
                "weekly_office_expense_rub",
                "tax_rate_pct",
                "common_expense_rule",
            )
            preview_card = next(
                (
                    card
                    for card in partner_payload.get("cards") or []
                    if _has_complete_partner_settings(card, preview_required_fields)
                ),
                None,
            )
            can_preview = bool(partner_payload.get("weeks")) and preview_card is not None
            preview_evidence["reason"] = (
                "server-owned settings are not complete; production settings were not mutated"
            )
            if can_preview:
                reports.locator("#partnerReportNmId").select_option(
                    str(preview_card.get("nm_id"))
                )
                reports.locator("#partnerReportWeekSummary").click()
                reports.locator("#partnerReportWeeksAll").click()
                selected_weeks = sorted(
                    str(item.get("week_start") or "")
                    for item in partner_payload.get("weeks") or []
                    if str(item.get("week_start") or "")
                )
                selected_nm_id = str(preview_card.get("nm_id") or "")
                reports.locator("#partnerReportGenerate").click()
                expect(reports.locator("#partnerReportStatus")).not_to_contain_text(
                    "Формируем preview",
                    timeout=120_000,
                )
                preview_status = reports.locator("#partnerReportStatus").inner_text().strip()
                table_visible = reports.locator(
                    "#partnerReportTableWrap:not([hidden]) table"
                ).count() == 1
                preview_evidence = {
                    "attempted": True,
                    "ready": False,
                    "status": preview_status,
                    "reason": "Partner preview API request has not completed",
                    "api_status": "",
                    "api_code": "",
                    "api_http_status": None,
                    "blockers": [],
                    "table_visible": table_visible,
                    "download_enabled": False,
                    "nm_id": selected_nm_id,
                    "selected_weeks": selected_weeks,
                    "source_digest": "",
                    "xlsx": dict(preview_evidence["xlsx"]),
                }
                preview_http_status, preview_payload, preview_json_object = _protected_json_post(
                    context,
                    normalized_base_url + PARTNER_PREVIEW_API_PATH,
                    {"nm_id": selected_nm_id, "selected_weeks": selected_weeks},
                    label="Partner preview API",
                )
                response_evidence = _partner_preview_api_evidence(
                    preview_http_status,
                    preview_payload,
                )
                preview_evidence.update(response_evidence)
                _assert(preview_json_object, "Partner preview API: JSON object")
                _assert(preview_http_status == 200, "Partner preview API: HTTP 200")
                blockers = list(response_evidence["blockers"])
                download_enabled = reports.locator(
                    "#partnerReportDownload"
                ).is_enabled()
                preview_ready = (
                    str(preview_payload.get("status") or "") == "ready"
                    and not blockers
                    and table_visible
                    and download_enabled
                )
                ui_table = _partner_ui_table_facts(reports)
                category_definitions = _partner_category_definitions(preview_payload)
                preview_evidence = {
                    "attempted": True,
                    "ready": preview_ready,
                    "status": preview_status,
                    "reason": (
                        "ready preview rendered from existing server-owned settings"
                        if preview_ready
                        else "Partner preview is not ready or its source blockers are non-empty"
                    ),
                    "api_status": str(response_evidence["api_status"]),
                    "api_code": str(response_evidence["api_code"]),
                    "api_http_status": preview_http_status,
                    "blockers": blockers,
                    "table_visible": table_visible,
                    "download_enabled": download_enabled,
                    "notes": reports.locator("#partnerReportNotes").inner_text().strip(),
                    "nm_id": selected_nm_id,
                    "selected_weeks": selected_weeks,
                    "source_digest": str(preview_payload.get("source_digest") or ""),
                    "ui_table": ui_table,
                    "other_expense_category_definitions": [
                        {"key": key, "label": label}
                        for key, label in category_definitions
                    ],
                    "xlsx": dict(preview_evidence["xlsx"]),
                }
                preflight_partner = evidence_dir / "partner_report_preflight.png"
                page.screenshot(path=str(preflight_partner), full_page=True)
                screenshots.append(str(preflight_partner))
                _assert(preview_ready, "Partner preview.ready=true with blockers=[]")
                _assert(
                    reports.locator("#partnerReportTableWrap")
                    .get_by_text("Дивиденды", exact=True)
                    .count()
                    == 1,
                    "Partner preview renders dividend reconciliation",
                )
                _assert(
                    ui_table["main_label"] == OTHER_DIRECT_ALLOCATED_LABEL
                    and ui_table["old_label_count"] == 0
                    and [
                        (str(item.get("key") or ""), str(item.get("label") or ""))
                        for item in ui_table["categories"]
                    ]
                    == category_definitions,
                    "Partner preview renders exactly the non-zero classified expense rows",
                )
                _assert(
                    bool(ui_table["tooltip_keyboard_focus_visible"]),
                    "Partner expense explanation works on keyboard focus",
                )
                _assert(
                    _negative_profit_dividends_valid(preview_payload),
                    "negative profit never accrues negative dividends",
                )
                with page.expect_download(timeout=120_000) as download_info:
                    reports.locator("#partnerReportDownload").click()
                download = download_info.value
                excel_path = evidence_dir / download.suggested_filename
                download.save_as(str(excel_path))
                _assert(
                    excel_path.is_file() and excel_path.stat().st_size > 0,
                    "Partner preview XLSX downloaded",
                )
                workbook_verification = _verify_partner_xlsx(
                    excel_path,
                    preview=preview_payload,
                    ui_table=ui_table,
                )
                preview_evidence["xlsx"] = _xlsx_evidence(
                    excel_path,
                    filename=download.suggested_filename,
                    verification=workbook_verification,
                )
                _assert(
                    bool(workbook_verification.get("passed")),
                    "Partner XLSX semantic verification",
                )
            else:
                _assert(False, "Partner preview requires existing server-owned settings and weeks")
            partner_desktop = evidence_dir / "partner_report_desktop_readonly.png"
            page.screenshot(path=str(partner_desktop), full_page=True)
            screenshots.append(str(partner_desktop))

            page.set_viewport_size({"width": 390, "height": 844})
            narrow_facts = reports.locator("body").evaluate(
                """
                () => {
                  const controls = document.getElementById('partnerReportControls');
                  const settings = controls.querySelector('.partner-report-settings');
                  const wrap = document.getElementById('partnerReportTableWrap');
                  return {
                    pageOverflow: document.documentElement.scrollWidth > document.documentElement.clientWidth,
                    settingsColumns: getComputedStyle(settings).gridTemplateColumns.split(' ').length,
                    controlsWithinViewport: controls.getBoundingClientRect().right <= document.documentElement.clientWidth + 1,
                    localTableScroll: !wrap.hidden && wrap.scrollWidth > wrap.clientWidth,
                  };
                }
                """
            )
            _assert(
                narrow_facts
                == {
                    "pageOverflow": False,
                    "settingsColumns": 1,
                    "controlsWithinViewport": True,
                    "localTableScroll": bool(preview_evidence.get("ready")),
                },
                "Partner narrow layout",
            )
            partner_narrow = evidence_dir / "partner_report_narrow_readonly.png"
            page.screenshot(path=str(partner_narrow), full_page=True)
            screenshots.append(str(partner_narrow))
            final_url = page.url
            browser.close()

        _assert(not page_errors, f"pageerror list is empty: {page_errors}")
        _assert(not console_errors, f"console error list is empty: {console_errors}")
        _assert(not server_errors, f"5xx response list is empty: {server_errors}")
        _assert(not client_errors, f"4xx response list is empty: {client_errors}")
        _assert(not failed_requests, f"failed request list is empty: {failed_requests}")
        _assert(
            _partner_acceptance_passed(preview_evidence),
            "Partner preview/XLSX acceptance is fail-closed",
        )
        unexpected_non_read_requests = [
            item
            for item in non_read_requests
            if urlparse(item["url"]).path
            not in {PARTNER_PREVIEW_API_PATH, PARTNER_PREVIEW_XLSX_API_PATH}
        ]
        _assert(
            not unexpected_non_read_requests,
            "UI flow did not save settings or finalize a report: "
            f"{unexpected_non_read_requests}",
        )
        report = {
            "status": "passed",
            "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "requested_url": requested_url,
            "final_url": final_url,
            "deployed_sha": str(deployed_sha or ""),
            "document_responses": document_responses,
            "redirects": _document_redirects(document_responses),
            "page_errors": page_errors,
            "console_errors": console_errors,
            "server_errors": server_errors,
            "client_errors": client_errors,
            "failed_requests": failed_requests,
            "non_read_requests": non_read_requests,
            "allowed_read_only_post_paths": [
                PARTNER_PREVIEW_API_PATH,
                PARTNER_PREVIEW_XLSX_API_PATH,
            ],
            "finance": {
                "week_count": len(finance_weeks),
                "first_week": finance_weeks[0].get("week_start"),
                "last_week": finance_weeks[-1].get("week_end"),
                "facts": finance_facts,
                "storage_phase": storage_phase,
                "storage_health": storage_health,
            },
            "partner": {
                "card_count": len(partner_payload.get("cards") or []),
                "week_count": len(partner_payload.get("weeks") or []),
                "desktop": partner_facts,
                "narrow": narrow_facts,
                "preview": preview_evidence,
                "finalized_report_created": False,
            },
            "screenshots": screenshots,
        }
        report_path.write_text(
            json.dumps(report, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
            encoding="utf-8",
        )
        return report
    except Exception as exc:
        failed_report = {
            "status": "failed",
            "requested_url": requested_url,
            "final_url": locals().get("final_url", ""),
            "deployed_sha": str(deployed_sha or ""),
            "error_type": type(exc).__name__,
            "error": str(exc),
            "document_responses": document_responses,
            "redirects": _document_redirects(document_responses),
            "page_errors": page_errors,
            "console_errors": console_errors,
            "server_errors": server_errors,
            "client_errors": client_errors,
            "failed_requests": failed_requests,
            "non_read_requests": non_read_requests,
            "partner": {"preview": preview_evidence},
            "screenshots": screenshots,
        }
        report_path.write_text(
            json.dumps(failed_report, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
            encoding="utf-8",
        )
        raise


def _protected_json_get(context: Any, url: str, *, label: str) -> dict[str, Any]:
    response = context.request.get(url, headers={"Accept": "application/json"}, timeout=120_000)
    _assert(response.status == 200, f"{label}: HTTP 200")
    payload = response.json()
    _assert(isinstance(payload, dict), f"{label}: JSON object")
    return payload


def _protected_json_post(
    context: Any,
    url: str,
    payload: dict[str, Any],
    *,
    label: str,
) -> tuple[int, dict[str, Any], bool]:
    response = context.request.post(
        url,
        headers={"Accept": "application/json", "Content-Type": "application/json"},
        data=payload,
        timeout=120_000,
    )
    try:
        result = response.json()
    except Exception:
        result = None
    if not isinstance(result, dict):
        status = int(response.status)
        return (
            status,
            {
                "error": f"{label} returned a non-JSON response",
                "code": "response_not_json",
                "blockers": [
                    {"code": "response_not_json", "http_status": status}
                ],
            },
            False,
        )
    return int(response.status), result, True


def _partner_preview_api_evidence(
    http_status: int,
    payload: dict[str, Any],
) -> dict[str, Any]:
    """Preserve bounded preview failure facts before fail-closed assertions."""

    detail = payload.get("detail")
    container = detail if isinstance(detail, dict) else payload
    raw_blockers = container.get("blockers") or payload.get("blockers") or []
    if isinstance(raw_blockers, list):
        blockers = list(raw_blockers)
    else:
        blockers = [raw_blockers]
    if not blockers and isinstance(detail, list):
        blockers = list(detail)

    api_status = str(container.get("status") or payload.get("status") or "")
    api_code = str(container.get("code") or payload.get("code") or "")
    reason_value = (
        container.get("reason")
        or container.get("message")
        or container.get("error")
        or payload.get("reason")
        or payload.get("message")
        or payload.get("error")
        or (detail if isinstance(detail, str) else "")
    )
    if isinstance(reason_value, (dict, list)):
        reason = json.dumps(reason_value, ensure_ascii=False, sort_keys=True)
    else:
        reason = str(reason_value or "").strip()
    if not reason:
        reason = (
            "Partner preview API returned no human-readable reason"
            if int(http_status) == 200
            else f"Partner preview API rejected the request with HTTP {int(http_status)}"
        )
    return {
        "api_http_status": int(http_status),
        "api_status": api_status,
        "api_code": api_code,
        "blockers": blockers,
        "reason": reason,
        "source_digest": str(container.get("source_digest") or payload.get("source_digest") or ""),
    }


def _partner_ui_table_facts(reports: Any) -> dict[str, Any]:
    anchor = reports.locator(".partner-report-tooltip-anchor")
    _assert(anchor.count() == 1, "Partner expense tooltip anchor exists once")
    anchor.focus()
    facts = reports.locator("body").evaluate(
        """
        () => {
          const table = document.querySelector('#partnerReportTableWrap table');
          const main = table && table.querySelector('[data-partner-row-key="other_direct_and_allocated_expenses"]');
          const categories = table ? [...table.querySelectorAll('[data-partner-expense-category]')] : [];
          const values = (row) => row ? [...row.cells].slice(1).map((cell) => cell.innerText.trim()) : [];
          const tooltip = document.querySelector('.partner-report-tooltip');
          const label = main && main.querySelector('.partner-report-tooltip-wrap > span:first-child');
          return {
            tableVisible: Boolean(table),
            main_label: label ? label.innerText.trim() : '',
            main_values: values(main),
            rows: [...table.querySelectorAll('[data-partner-row-key]')].map((row) => ({
              key: row.dataset.partnerRowKey || '',
              values: values(row),
            })),
            old_label_count: table ? [...table.querySelectorAll('td')].filter((cell) => cell.innerText.trim() === 'Прочие атрибутируемые расходы').length : 0,
            categories: categories.map((row) => ({
              key: row.dataset.partnerExpenseCategory || '',
              label: row.cells[0] ? row.cells[0].innerText.trim() : '',
              values: values(row),
            })),
            tooltip_text: tooltip ? tooltip.innerText.trim() : '',
            tooltip_keyboard_focus_visible: Boolean(tooltip && getComputedStyle(tooltip).display !== 'none'),
          };
        }
        """
    )
    return dict(facts)


def _partner_category_definitions(preview: dict[str, Any]) -> list[tuple[str, str]]:
    """Validate the server-owned ordered visibility contract for Partner subrows."""

    canonical = dict(OTHER_EXPENSE_CATEGORIES)
    raw = preview.get("other_expense_category_definitions") or []
    if not isinstance(raw, list):
        raise AssertionError("Partner expense category definitions must be a list")
    result: list[tuple[str, str]] = []
    for item in raw:
        if not isinstance(item, dict):
            raise AssertionError("Partner expense category definition must be an object")
        key = str(item.get("key") or "")
        label = str(item.get("label") or "")
        if key not in canonical or canonical[key] != label:
            raise AssertionError(f"unknown Partner expense category definition: {key!r}")
        if key in {existing for existing, _label in result}:
            raise AssertionError(f"duplicate Partner expense category definition: {key}")
        result.append((key, label))
    expected_order = [key for key, _label in OTHER_EXPENSE_CATEGORIES if key in dict(result)]
    if [key for key, _label in result] != expected_order:
        raise AssertionError("Partner expense category definitions are out of canonical order")
    return result


def _verify_partner_xlsx(
    path: Path,
    *,
    preview: dict[str, Any],
    ui_table: dict[str, Any],
) -> dict[str, Any]:
    findings: list[dict[str, Any]] = []
    raw = path.read_bytes()
    sheet_names: list[str] = []
    hidden_sheets: list[str] = []
    external_links: list[str] = []
    macro_members: list[str] = []
    try:
        category_definitions = _partner_category_definitions(preview)
        with zipfile.ZipFile(BytesIO(raw)) as archive:
            for member in archive.namelist():
                lowered = member.casefold()
                if lowered.startswith("xl/externallinks/"):
                    external_links.append(member)
                if "vbaproject" in lowered or lowered.endswith(".bin"):
                    macro_members.append(member)
        workbook = load_workbook(BytesIO(raw), read_only=False, data_only=False)
        sheet_names = list(workbook.sheetnames)
        hidden_sheets = [
            sheet.title for sheet in workbook.worksheets if sheet.sheet_state != "visible"
        ]
        if sheet_names != ["Партнёрский отчёт", "Параметры"]:
            findings.append(
                {"code": "unexpected_sheets", "actual": sheet_names}
            )
        if hidden_sheets:
            findings.append({"code": "hidden_sheets", "actual": hidden_sheets})
        if external_links:
            findings.append({"code": "external_links", "actual": external_links})
        if macro_members or getattr(workbook, "vba_archive", None) is not None:
            findings.append({"code": "macros", "actual": macro_members})

        parameters = {
            str(row[0].value): str(row[1].value or "")
            for row in workbook["Параметры"].iter_rows(min_col=1, max_col=2)
            if row[0].value not in (None, "")
        }
        expected_nm_id = str(preview.get("nm_id") or "")
        expected_weeks = [str(item) for item in preview.get("selected_weeks") or []]
        if parameters.get("nmId") != expected_nm_id:
            findings.append(
                {
                    "code": "wrong_nm_id",
                    "expected": expected_nm_id,
                    "actual": parameters.get("nmId"),
                }
            )
        workbook_weeks = [
            item.strip()
            for item in parameters.get("Недели", "").split(",")
            if item.strip()
        ]
        if workbook_weeks != expected_weeks:
            findings.append(
                {
                    "code": "wrong_weeks",
                    "expected": expected_weeks,
                    "actual": workbook_weeks,
                }
            )
        if parameters.get("Source digest") != str(preview.get("source_digest") or ""):
            findings.append({"code": "source_digest_mismatch"})
        if parameters.get("Formula version") != str(preview.get("formula_version") or ""):
            findings.append({"code": "formula_version_mismatch"})

        sheet = workbook["Партнёрский отчёт"]
        rows_by_label = {
            str(row[1].value): row
            for row in sheet.iter_rows(min_row=2)
            if len(row) > 1 and row[1].value not in (None, "")
        }
        missing_metric_rows = [
            key for key, label in REPORT_ROWS if label not in rows_by_label
        ]
        if missing_metric_rows:
            findings.append(
                {"code": "metric_rows_missing", "keys": missing_metric_rows}
            )
        for key, label in REPORT_ROWS:
            row = rows_by_label.get(label)
            if row is None:
                continue
            expected_exact = [
                _metric_exact((week.get("values") or {}).get(key))
                for week in preview.get("weeks") or []
            ] + [_metric_exact((preview.get("totals") or {}).get(key))]
            actual_exact = [
                _xlsx_metric_exact(key, cell.value)
                for cell in row[2 : 3 + len(expected_weeks)]
            ]
            if actual_exact != expected_exact:
                findings.append(
                    {
                        "code": "metric_values_mismatch",
                        "metric": key,
                        "expected": _optional_decimal_list(expected_exact),
                        "actual": _optional_decimal_list(actual_exact),
                    }
                )
        if "Прочие атрибутируемые расходы" in rows_by_label:
            findings.append({"code": "legacy_label_present"})
        main_row = rows_by_label.get(OTHER_DIRECT_ALLOCATED_LABEL)
        if main_row is None:
            findings.append({"code": "main_expense_row_missing"})
        category_rows = {
            key: rows_by_label.get(label) for key, label in category_definitions
        }
        missing_categories = [key for key, row in category_rows.items() if row is None]
        if missing_categories:
            findings.append(
                {"code": "expense_categories_missing", "keys": missing_categories}
            )
        expected_keys = {key for key, _label in category_definitions}
        unexpected_categories = [
            key
            for key, label in OTHER_EXPENSE_CATEGORIES
            if key not in expected_keys and label in rows_by_label
        ]
        if unexpected_categories:
            findings.append(
                {
                    "code": "zero_only_expense_categories_present",
                    "keys": unexpected_categories,
                }
            )

        week_breakdowns = [
            _breakdown_values(week.get("other_expense_breakdown") or [])
            for week in preview.get("weeks") or []
        ]
        week_breakdowns.append(
            _breakdown_values(preview.get("other_expense_breakdown_total") or [])
        )
        expected_main = [
            _money_cent((week.get("values") or {}).get(OTHER_DIRECT_ALLOCATED_KEY))
            for week in preview.get("weeks") or []
        ] + [_money_cent((preview.get("totals") or {}).get(OTHER_DIRECT_ALLOCATED_KEY))]

        if main_row is not None and not missing_categories:
            actual_main = [_money_cent(cell.value) for cell in main_row[2 : 3 + len(expected_weeks)]]
            if actual_main != expected_main:
                findings.append(
                    {
                        "code": "main_expense_values_mismatch",
                        "expected": _decimal_list(expected_main),
                        "actual": _decimal_list(actual_main),
                    }
                )
            for position, breakdown in enumerate(week_breakdowns):
                actual_categories = {
                    key: _money_cent(category_rows[key][position + 2].value)  # type: ignore[index]
                    for key, _label in category_definitions
                }
                expected_categories = {
                    key: _money_cent(breakdown.get(key))
                    for key, _label in category_definitions
                }
                if actual_categories != expected_categories:
                    findings.append(
                        {
                            "code": "category_values_mismatch",
                            "position": position,
                            "expected": _decimal_mapping(expected_categories),
                            "actual": _decimal_mapping(actual_categories),
                        }
                    )
                if sum(actual_categories.values(), Decimal("0")) != actual_main[position]:
                    findings.append(
                        {"code": "displayed_kopeck_conservation_failed", "position": position}
                    )
                if any(category_rows[key][0].value is not None for key, _label in category_definitions):  # type: ignore[index]
                    findings.append({"code": "category_coefficient_exposed"})
                    break

        ui_main = [_parse_ui_money(value) for value in ui_table.get("main_values") or []]
        if ui_main != expected_main:
            findings.append(
                {
                    "code": "ui_main_reconciliation_failed",
                    "expected": _decimal_list(expected_main),
                    "actual": _decimal_list(ui_main),
                }
            )
        ui_categories = {
            str(item.get("key") or ""): [
                _parse_ui_money(value) for value in item.get("values") or []
            ]
            for item in ui_table.get("categories") or []
        }
        if [
            (str(item.get("key") or ""), str(item.get("label") or ""))
            for item in ui_table.get("categories") or []
        ] != category_definitions:
            findings.append({"code": "ui_category_definitions_mismatch"})
        for key, _label in category_definitions:
            expected = [
                _money_cent(item.get(key)) for item in week_breakdowns
            ]
            if ui_categories.get(key) != expected:
                findings.append(
                    {
                        "code": "ui_category_reconciliation_failed",
                        "category": key,
                    }
                )
        if any(
            "%" in value
            for item in ui_table.get("categories") or []
            for value in item.get("values") or []
        ):
            findings.append({"code": "category_percentage_exposed"})

        ui_rows = {
            str(item.get("key") or ""): list(item.get("values") or [])
            for item in ui_table.get("rows") or []
        }
        for key, _label in REPORT_ROWS:
            expected_display = [
                _metric_display((week.get("values") or {}).get(key))
                for week in preview.get("weeks") or []
            ] + [_metric_display((preview.get("totals") or {}).get(key))]
            actual_display = [
                _parse_ui_metric(value) for value in ui_rows.get(key, [])
            ]
            if actual_display != expected_display:
                findings.append(
                    {
                        "code": "ui_metric_reconciliation_failed",
                        "metric": key,
                        "expected": _optional_decimal_list(expected_display),
                        "actual": _optional_decimal_list(actual_display),
                    }
                )

        dividends = [
            _decimal_or_none((week.get("values") or {}).get("dividends"))
            for week in preview.get("weeks") or []
        ]
        capital = _decimal_or_none(
            (preview.get("parameters") or {}).get("invested_capital_rub")
        )
        annualized = _decimal_or_none(
            (preview.get("totals") or {}).get("annualized_return_pct")
        )
        if dividends and all(value is not None for value in dividends) and capital and capital > 0:
            expected_annualized = (
                sum((value for value in dividends if value is not None), Decimal("0"))
                / Decimal(len(dividends))
                * Decimal("52")
                / capital
                * Decimal("100")
            ).quantize(Decimal("0.0001"))
            if annualized is None or annualized.quantize(Decimal("0.0001")) != expected_annualized:
                findings.append({"code": "annualized_return_mismatch"})
        workbook.close()
    except Exception as exc:
        findings.append({"code": "workbook_open_or_verify_failed", "reason": str(exc)})

    return {
        "passed": not findings,
        "findings": findings,
        "sheet_names": sheet_names,
        "hidden_sheets": hidden_sheets,
        "external_links": external_links,
        "macro_members": macro_members,
        "nm_id": str(preview.get("nm_id") or ""),
        "selected_weeks": list(preview.get("selected_weeks") or []),
        "source_digest": str(preview.get("source_digest") or ""),
        "ui_xlsx_reconciled": not findings,
    }


def _breakdown_values(rows: list[dict[str, Any]]) -> dict[str, Decimal | None]:
    return {
        str(item.get("key") or ""): _decimal_or_none(item.get("amount_rub"))
        for item in rows
    }


def _money_cent(value: Any) -> Decimal:
    parsed = _decimal_or_none(value)
    if parsed is None:
        raise ValueError(f"expected a money value, got {value!r}")
    return parsed.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _xlsx_evidence(
    path: Path,
    *,
    filename: str,
    verification: dict[str, Any],
) -> dict[str, Any]:
    return {
        "downloaded": True,
        "path": str(path),
        "filename": str(filename),
        "sha256": "sha256:" + hashlib.sha256(path.read_bytes()).hexdigest(),
        "size_bytes": path.stat().st_size,
        "verification": verification,
    }


def _metric_exact(value: Any) -> Decimal | None:
    parsed = _decimal_or_none(value)
    return None if parsed is None else parsed.quantize(Decimal("0.0001"))


def _xlsx_metric_exact(metric_key: str, value: Any) -> Decimal | None:
    parsed = _decimal_or_none(value)
    if parsed is None:
        return None
    if metric_key == "annualized_return_pct":
        parsed *= Decimal("100")
    return parsed.quantize(Decimal("0.0001"))


def _metric_display(value: Any) -> Decimal | None:
    parsed = _decimal_or_none(value)
    return (
        None
        if parsed is None
        else parsed.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    )


def _decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    return result if result.is_finite() else None


def _parse_ui_money(value: Any) -> Decimal:
    normalized = (
        str(value or "")
        .replace("\u00a0", "")
        .replace("\u202f", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace("−", "-")
        .replace(",", ".")
        .strip()
    )
    if normalized in {"", "—"}:
        raise ValueError(f"expected visible money, got {value!r}")
    return _money_cent(normalized)


def _parse_ui_metric(value: Any) -> Decimal | None:
    normalized = (
        str(value or "")
        .replace("\u00a0", "")
        .replace("\u202f", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace("%", "")
        .replace("−", "-")
        .replace(",", ".")
        .strip()
    )
    if normalized in {"", "—"}:
        return None
    return _metric_display(normalized)


def _decimal_list(values: list[Decimal]) -> list[str]:
    return [format(value, "f") for value in values]


def _optional_decimal_list(values: list[Decimal | None]) -> list[str | None]:
    return [None if value is None else format(value, "f") for value in values]


def _decimal_mapping(values: dict[str, Decimal]) -> dict[str, str]:
    return {key: format(value, "f") for key, value in values.items()}


def _document_redirects(document_responses: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        {"from": str(item.get("redirected_from") or ""), "to": str(item.get("url") or "")}
        for item in document_responses
        if str(item.get("redirected_from") or "")
    ]


def _partner_acceptance_passed(preview: dict[str, Any]) -> bool:
    xlsx = dict(preview.get("xlsx") or {})
    verification = dict(xlsx.get("verification") or {})
    return bool(
        preview.get("attempted")
        and preview.get("ready")
        and str(preview.get("api_status") or "") == "ready"
        and not list(preview.get("blockers") or [])
        and preview.get("table_visible")
        and preview.get("download_enabled")
        and xlsx.get("downloaded")
        and xlsx.get("path")
        and xlsx.get("sha256")
        and int(xlsx.get("size_bytes") or 0) > 0
        and verification.get("passed")
        and verification.get("ui_xlsx_reconciled")
    )


def _negative_profit_dividends_valid(preview: dict[str, Any]) -> bool:
    """Validate present loss weeks without requiring production to have one."""

    negative_weeks = [
        week
        for week in preview.get("weeks") or []
        if _decimal_or_none((week.get("values") or {}).get("net_profit"))
        is not None
        and _decimal_or_none((week.get("values") or {}).get("net_profit")) < 0
    ]
    return all(
        _decimal_or_none((week.get("values") or {}).get("dividends"))
        == Decimal("0")
        for week in negative_weeks
    )


def _has_complete_partner_settings(
    card: dict[str, Any],
    required_fields: tuple[str, ...],
) -> bool:
    parameters = dict(((card or {}).get("settings") or {}).get("parameters") or {})
    return all(parameters.get(key) not in {None, ""} for key in required_fields)


def _assert(condition: bool, label: str) -> None:
    if not condition:
        raise AssertionError(label)
