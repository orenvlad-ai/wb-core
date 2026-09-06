"""Exercise atomic revision, exact links, drift and transaction rollback."""

from copy import deepcopy
from io import BytesIO
import hashlib
import json
from pathlib import Path
import socket
import sqlite3
import sys
from tempfile import TemporaryDirectory
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from apps.production_apply_launcher import execute
from apps.supplier_invoice_parser_smoke import _build_invoice_fixture
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime
from packages.application.supplier_invoice_parser import parse_supplier_invoice_xlsx
from packages.application.supplier_shipment_invoice_revision import SupplierInvoiceRevisionAdapter, snapshot, connect, candidate, digest


def fixture(root):
    original = _build_invoice_fixture()
    parsed = parse_supplier_invoice_xlsx(original, filename="old.xlsx")
    (root / "old.xlsx").write_bytes(original)
    workbook = load_workbook(BytesIO(original))
    # This fixture has three products followed by packaging; adjust a product
    # and its amount, plus any declared total, with the same unit prices.
    sheet = workbook.active
    row = parsed["lines"][0]["raw"]["worksheet_row"]
    sheet.cell(row, 5).value += 10
    sheet.cell(row, 7).value = sheet.cell(row, 5).value * sheet.cell(row, 6).value
    sheet.cell(sheet.max_row, 7).value += 10
    raw = BytesIO(); workbook.save(raw)
    updated = raw.getvalue()
    new_path = root / "new.xlsx"; new_path.write_bytes(updated)
    new_parsed = parse_supplier_invoice_xlsx(updated, filename="new.xlsx")
    assert not new_parsed["errors"], new_parsed["errors"]
    runtime = RegistryUploadDbBackedRuntime(runtime_dir=root)
    now = "2026-07-01T00:00:00Z"
    meta = parsed["metadata"]
    sha = hashlib.sha256(original).hexdigest()
    header = {**meta, **parsed["summary"], "shipment_id": "sup_test", "created_at": now, "updated_at": now, "shipment_date": "2026-09-20", "source_filename": "old.xlsx", "source_file_path": "old.xlsx", "source_file_sha256": sha, "invoice_document_id": "invoice_test", "supplier_name": "Factory", "match_status": "all_matched"}
    lines = [{**line, "line_id": "line_" + str(i), "internal_nm_id": i + 1 if line["line_type"] == "product" else None} for i, line in enumerate(parsed["lines"])]
    runtime.save_supplier_shipment(header=header, lines=lines)
    runtime.save_trade_document({"document_id": "invoice_test", "document_type": "invoice", "number": header["invoice_no"], "document_date": header["invoice_date"], "supplier_name": "Factory", "currency": header["currency"], "amount_total": header["invoice_amount_total"], "file_path": "old.xlsx", "file_original_name": "old.xlsx", "file_sha256": sha, "status": "active", "created_at": now, "updated_at": now})
    runtime.save_trade_document({"document_id": "contract_test", "document_type": "contract", "number": "contract", "status": "active", "created_at": now, "updated_at": now})
    runtime.save_invoice_contract_link(invoice_document_id="invoice_test", contract_document_id="contract_test", created_at=now, updated_at=now, linked_by="operator", source="operator")
    with sqlite3.connect(runtime.db_path) as conn:
        conn.execute("INSERT INTO sheet_vitrina_v1_cny_documents(document_id,document_type,source,source_order_id,context_order_id,natural_key,uploaded_at,created_at,updated_at,status,operation_date,cny_amount) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", ("payment_test", "supplier_cny_payment", "operator", "sup_test", "sup_test", "payment-key", now, now, now, "posted", "2026-07-20", "15.00"))
        conn.execute("INSERT INTO sheet_vitrina_v1_supplier_financial_documents(document_id,supplier_order_id,document_type,original_filename,stored_file_path,file_content_type,file_sha256,uploaded_at,updated_at,parse_status,total_amount_rub) VALUES(?,?,?,?,?,?,?,?,?,?,?)", ("fee_test", "sup_test", "bank_fee_statement", "fee.pdf", "fee.pdf", "application/pdf", "b" * 64, now, now, "confirmed", 100.0))
    marker = root / "sha"; marker.write_text("a" * 40)
    request = {"runtime_dir": str(root), "hostname": socket.gethostname(), "runtime_sha_file": str(marker), "runtime_sha": "a" * 40, "shipment_id": "sup_test", "invoice_no": header["invoice_no"], "source_path": str(new_path), "source_sha256": hashlib.sha256(updated).hexdigest()}
    return request, runtime, parsed, new_parsed


def main():
    adapter = SupplierInvoiceRevisionAdapter()
    adapters = {"test": adapter}
    with TemporaryDirectory() as raw:
        root = Path(raw)
        request, runtime, old, new = fixture(root)
        with connect(runtime.db_path) as conn:
            before = snapshot(conn, "sup_test")
        preview = execute(action="preview", adapter_name="test", operation_id="test-revision", request=request, adapters=adapters)
        assert preview["scope"]["after_quantity"] == preview["scope"]["before_quantity"] + 10
        assert not (root / "supplier_invoice_revisions").exists()
        changed = deepcopy(new); changed["lines"][0]["barcode"] = "9999999999999"
        for bad in (changed,):
            try:
                candidate(before, bad, root / "new.xlsx", request["source_sha256"])
            except ValueError:
                raise
            except Exception as exc:
                assert "assortment" in str(exc)
            else:
                raise AssertionError("changed barcode admitted")
        with patch.object(adapter, "enqueue", return_value={"queue_id": "test-queue", "status": "queued"}):
            result = execute(action="apply", adapter_name="test", operation_id="test-revision", request=request, expected_prestate=preview["prestate_sha256"], expected_candidate=preview["candidate_sha256"], adapters=adapters)
        assert result["state"] == "applied", result
        with connect(runtime.db_path) as conn:
            after = snapshot(conn, "sup_test")
        assert after["protected"] == before["protected"]
        assert len(after["protected"]["cny_documents"]) == 1
        assert len(after["protected"]["supplier_financial_documents"]) == 1
        assert {x["line_id"] for x in before["lines"]} == {x["line_id"] for x in after["lines"]}
        assert (root / after["header"]["source_file_path"]).read_bytes() == (root / "new.xlsx").read_bytes()
        assert json.loads((root / "supplier_invoice_revisions/test-revision/before.json").read_text()) == before
        assert execute(action="readback", adapter_name="test", operation_id="test-revision", request=request, adapters=adapters)["state"] == "applied"
        # A second operation pinned to the old prestate cannot mutate the order.
        try:
            execute(action="apply", adapter_name="test", operation_id="test-drift", request=request, expected_prestate=preview["prestate_sha256"], expected_candidate=preview["candidate_sha256"], adapters=adapters)
        except Exception as exc:
            assert "drift" in str(exc)
        else:
            raise AssertionError("stale preview admitted")
        with sqlite3.connect(runtime.db_path) as conn:
            conn.execute("UPDATE sheet_vitrina_v1_supplier_shipments SET actual_ff_acceptance_date='2026-09-01' WHERE shipment_id='sup_test'")
        try:
            adapter.preview(request, "test-accepted")
        except Exception as exc:
            assert "unshipped" in str(exc)
        else:
            raise AssertionError("received order admitted")
    with TemporaryDirectory() as raw:
        request, runtime, _, _ = fixture(Path(raw))
        preview = adapter.preview(request, "test-rollback")
        with connect(runtime.db_path) as conn:
            before = snapshot(conn, "sup_test")
        # Abort on the document write after the header/lines have been updated.
        with sqlite3.connect(runtime.db_path) as conn:
            conn.execute("CREATE TRIGGER reject_revision BEFORE UPDATE ON sheet_vitrina_v1_trade_documents BEGIN SELECT RAISE(ABORT,'injected failure'); END")
        try:
            adapter.apply(request, "test-rollback", preview)
        except sqlite3.IntegrityError:
            pass
        else:
            raise AssertionError("injected failure did not abort")
        with connect(runtime.db_path) as conn:
            assert digest(snapshot(conn, "sup_test")) == digest(before)
    print("supplier_invoice_revision_smoke: OK")


if __name__ == "__main__":
    main()
