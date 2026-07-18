"""Smoke-check Fulfillment services XLSX/PDF/upload overlay contour."""

from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
import socket
import sys
from tempfile import TemporaryDirectory
import threading
from urllib import error as urllib_error, request as urllib_request

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.adapters.registry_upload_http_entrypoint import (  # noqa: E402
    DEFAULT_FULFILLMENT_SERVICES_TEMPLATE_PATH,
    DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH,
    DEFAULT_SHEET_OPERATOR_UI_PATH,
    DEFAULT_SHEET_PLAN_PATH,
    DEFAULT_SHEET_STATUS_PATH,
    DEFAULT_UPLOAD_PATH,
    DEFAULT_WB_SUPPLIES_PATH,
    build_registry_upload_http_server,
)
from packages.application.fulfillment_services import TEMPLATE_HEADERS, FulfillmentServicesBlock  # noqa: E402
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime  # noqa: E402
from packages.application.registry_upload_http_entrypoint import RegistryUploadHttpEntrypoint  # noqa: E402
from packages.contracts.registry_upload_http_entrypoint import RegistryUploadHttpEntrypointConfig  # noqa: E402


NOW = "2026-07-06T08:00:00Z"


def main() -> None:
    with TemporaryDirectory(prefix="fulfillment-services-app-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        _seed_wb_supplies(runtime)
        block = FulfillmentServicesBlock(runtime=runtime, timestamp_factory=lambda: NOW)

        template_bytes, template_filename, template_content_type = block.build_template()
        if template_filename != "sheet-vitrina-v1-fulfillment-services-template.xlsx":
            raise AssertionError(f"unexpected template filename: {template_filename}")
        if "spreadsheetml" not in template_content_type:
            raise AssertionError(f"unexpected template content type: {template_content_type}")
        template_headers = _read_headers(template_bytes)
        if template_headers != TEMPLATE_HEADERS:
            raise AssertionError(f"template headers must match PNG-derived form, got {template_headers}")

        if template_headers[0] != "Номер поставки" or template_headers[1] != "Склад":
            raise AssertionError(f"template must expose first columns Номер поставки/Склад, got {template_headers[:2]}")

        legacy_payload = block.upload_xlsx(
            _build_workbook([_valid_row("1001")], headers=_legacy_headers()),
            uploaded_filename="legacy-ok.xlsx",
        )
        if legacy_payload.get("validation_status") != "ok":
            raise AssertionError(f"legacy Стоимость услуг header must remain accepted, got {legacy_payload}")
        block.delete_upload(legacy_payload["upload"]["upload_id"])

        ok_payload = block.upload_xlsx(
            _build_workbook([_valid_row("1001"), _valid_row("1002"), _storage_row()]),
            uploaded_filename="ok.xlsx",
        )
        ok_upload = ok_payload["upload"]
        if (
            ok_payload.get("validation_status") != "ok"
            or ok_upload.get("rows_total") != 2
            or ok_upload.get("rows_matched") != 2
            or ok_upload.get("amount_without_vat_total") != 3500.0
            or ok_upload.get("vat_total") != 175.0
            or ok_upload.get("amount_with_vat_total") != 3675.0
            or not ok_upload.get("payment_validation_id")
            or not ok_payload.get("pdf_available")
        ):
            raise AssertionError(f"valid upload must pass and generate PDF, got {ok_payload}")
        detail_lines = ok_payload.get("lines") or []
        if not any(line.get("is_storage_line") and line.get("storage_total_amount_with_vat") == 525.0 for line in detail_lines):
            raise AssertionError(f"STORAGE line must be persisted as storage, got {detail_lines}")
        ordinary_1001 = next(line for line in detail_lines if line.get("supply_id_input") == "1001")
        if (
            ordinary_1001.get("storage_allocated_amount_with_vat") != 210.0
            or ordinary_1001.get("service_amount_with_storage") != 1785.0
        ):
            raise AssertionError(f"storage must allocate by boxes to ordinary rows, got {ordinary_1001}")
        accepted_list = block.list_uploads()
        if [item.get("upload_id") for item in accepted_list.get("uploads") or []] != [ok_upload["upload_id"]]:
            raise AssertionError(f"list must expose accepted OK uploads only, got {accepted_list}")
        pdf_bytes, _, _ = block.download_pdf(ok_upload["upload_id"])
        pdf_text = _pdf_text(pdf_bytes)
        for expected in (
            "Виза на оплату услуг ФФ",
            ok_upload["upload_id"],
            ok_upload["payment_validation_id"],
            ok_upload["short_file_hash"],
            "К оплате = Итого + НДС 5%",
            "Хранение",
            "1001",
            "1002",
        ):
            if expected not in pdf_text:
                raise AssertionError(f"PDF must contain {expected!r}, text={pdf_text!r}")

        unmatched = block.upload_xlsx(_build_workbook([_valid_row("9999")]), uploaded_filename="unmatched.xlsx")
        if unmatched.get("validation_status") != "failed" or unmatched["upload"].get("pdf_available"):
            raise AssertionError(f"unmatched upload must fail without PDF, got {unmatched}")
        if "does not match cached WB supply" not in json.dumps(unmatched.get("row_errors"), ensure_ascii=False):
            raise AssertionError(f"unmatched row error missing, got {unmatched}")
        if [item.get("upload_id") for item in block.list_uploads().get("uploads") or []] != [ok_upload["upload_id"]]:
            raise AssertionError("failed unmatched upload must not enter accepted uploads list")

        duplicate = block.upload_xlsx(_build_workbook([_valid_row("1001"), _valid_row("1001")]), uploaded_filename="duplicate.xlsx")
        if duplicate.get("validation_status") != "failed" or duplicate["upload"].get("pdf_available"):
            raise AssertionError(f"duplicate upload must fail without PDF, got {duplicate}")
        if "Duplicate" not in json.dumps(duplicate.get("row_errors"), ensure_ascii=False):
            raise AssertionError(f"duplicate row error missing, got {duplicate}")
        if [item.get("upload_id") for item in block.list_uploads().get("uploads") or []] != [ok_upload["upload_id"]]:
            raise AssertionError("failed duplicate upload must not enter accepted uploads list")
        only_storage = block.upload_xlsx(_build_workbook([_storage_row()]), uploaded_filename="only-storage.xlsx")
        if only_storage.get("validation_status") != "failed" or only_storage["upload"].get("pdf_available"):
            raise AssertionError(f"only STORAGE upload must fail without PDF, got {only_storage}")
        bad_storage = block.upload_xlsx(_build_workbook([_valid_row("1001"), _storage_row(total="bad")]), uploaded_filename="bad-storage.xlsx")
        if bad_storage.get("validation_status") != "failed" or bad_storage["upload"].get("pdf_available"):
            raise AssertionError(f"bad STORAGE amount must fail without PDF, got {bad_storage}")
        missing_denominator = block.upload_xlsx(
            _build_workbook([_valid_row_without_boxes("1003"), _storage_row()]),
            uploaded_filename="missing-denominator.xlsx",
        )
        if missing_denominator.get("validation_status") != "failed" or missing_denominator["upload"].get("pdf_available"):
            raise AssertionError(f"missing storage allocation denominator must fail without PDF, got {missing_denominator}")
        if "Невозможно распределить хранение" not in json.dumps(missing_denominator.get("row_errors"), ensure_ascii=False):
            raise AssertionError(f"missing denominator error missing, got {missing_denominator}")
        overlay_before_delete = block.approved_overlay_by_supply()
        if (
            overlay_before_delete.get("1001", {}).get("amount_with_vat_total") != 1785.0
            or overlay_before_delete.get("1001", {}).get("storage_allocated_amount_with_vat_total") != 210.0
        ):
            raise AssertionError(f"approved overlay missing before delete, got {overlay_before_delete}")
        deleted = block.delete_upload(ok_upload["upload_id"])
        if not deleted.get("deleted") or not deleted.get("soft_deleted"):
            raise AssertionError(f"delete must soft-delete upload, got {deleted}")
        if block.list_uploads().get("uploads"):
            raise AssertionError("deleted upload must disappear from accepted uploads list")
        try:
            block.download_pdf(ok_upload["upload_id"])
        except ValueError:
            pass
        else:
            raise AssertionError("deleted upload PDF must be unavailable")
        if block.approved_overlay_by_supply():
            raise AssertionError("deleted upload lines must disappear from approved overlay")

    with TemporaryDirectory(prefix="fulfillment-services-http-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        _seed_wb_supplies(runtime)
        port = _reserve_free_port()
        entrypoint = RegistryUploadHttpEntrypoint(
            runtime_dir=runtime_dir,
            runtime=runtime,
            activated_at_factory=lambda: NOW,
        )
        cfg = RegistryUploadHttpEntrypointConfig(
            host="127.0.0.1",
            port=port,
            upload_path=DEFAULT_UPLOAD_PATH,
            sheet_plan_path=DEFAULT_SHEET_PLAN_PATH,
            sheet_refresh_path="/v1/sheet-vitrina-v1/refresh",
            sheet_status_path=DEFAULT_SHEET_STATUS_PATH,
            sheet_operator_ui_path=DEFAULT_SHEET_OPERATOR_UI_PATH,
            runtime_dir=runtime_dir,
        )
        server = build_registry_upload_http_server(cfg, entrypoint=entrypoint)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            base_url = f"http://127.0.0.1:{cfg.port}"
            template_status, template_body, template_headers = _get_bytes(f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_TEMPLATE_PATH}")
            if template_status != 200 or "spreadsheetml" not in template_headers.get("Content-Type", ""):
                raise AssertionError(f"template route must return XLSX, got {template_status} {template_headers}")
            load_workbook(BytesIO(template_body), data_only=True)

            upload_status, upload_payload = _post_multipart(
                f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}",
                _build_workbook([_valid_row("1001"), _valid_row("1002"), _storage_row()]),
                filename="http-ok.xlsx",
            )
            if upload_status != 200 or upload_payload.get("validation_status") != "ok":
                raise AssertionError(f"upload route must accept valid XLSX, got {upload_status} {upload_payload}")
            if (upload_payload.get("warehouse_targeted_recalculation") or {}).get("status") != "queued":
                raise AssertionError(f"confirmed FF upload must enqueue bounded warehouse replay, got {upload_payload}")
            upload_id = upload_payload["upload"]["upload_id"]

            list_status, list_payload = _get_json(f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}")
            if list_status != 200 or not list_payload.get("uploads"):
                raise AssertionError(f"list route must return uploads, got {list_status} {list_payload}")
            detail_status, detail_payload = _get_json(f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}/{upload_id}")
            if detail_status != 200 or len(detail_payload.get("lines") or []) != 3:
                raise AssertionError(f"detail route must return parsed lines, got {detail_status} {detail_payload}")
            if not any(line.get("is_storage_line") for line in detail_payload.get("lines") or []):
                raise AssertionError(f"detail route must expose STORAGE line, got {detail_payload}")
            pdf_status, pdf_body, pdf_headers = _get_bytes(
                f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}/{upload_id}/payment-validation.pdf"
            )
            if pdf_status != 200 or "application/pdf" not in pdf_headers.get("Content-Type", ""):
                raise AssertionError(f"PDF route must return protected PDF, got {pdf_status} {pdf_headers}")
            if upload_id not in _pdf_text(pdf_body):
                raise AssertionError("PDF route must return the generated payment-validation PDF")

            wb_status, wb_payload = _get_json(f"{base_url}{DEFAULT_WB_SUPPLIES_PATH}?size_filter=all&limit=20")
            if wb_status != 200:
                raise AssertionError(f"WB supplies route failed: {wb_status} {wb_payload}")
            labels = [item.get("label") for item in wb_payload.get("schema", {}).get("columns", [])]
            if "Транзит" not in labels or "Услуги ФФ" not in labels or "Стоимость" in labels:
                raise AssertionError(f"WB supplies schema must expose transit/fulfillment labels, got {labels}")
            by_number = {str(row.get("visible_number") or row.get("wb_supply_id")): row for row in wb_payload.get("rows") or []}
            row1001 = by_number.get("1001") or {}
            if (
                row1001.get("fulfillment_amount_with_vat_total") != 1785.0
                or row1001.get("fulfillment_storage_allocated_amount_with_vat_total") != 210.0
                or "₽/шт" not in str(row1001.get("fulfillment_per_unit_display") or "")
                or "₽/шт" not in str(row1001.get("fulfillment_storage_per_unit_display") or "")
                or "₽/шт" not in str(row1001.get("transit_per_unit_display") or "")
            ):
                raise AssertionError(f"WB supplies overlay must include fulfillment/transit per-unit, got {row1001}")

            delete_status, delete_payload = _delete_json(f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}/{upload_id}")
            if delete_status != 200 or not delete_payload.get("deleted"):
                raise AssertionError(f"DELETE route must soft-delete upload, got {delete_status} {delete_payload}")
            if (delete_payload.get("warehouse_targeted_recalculation") or {}).get("status") != "queued":
                raise AssertionError(f"FF upload cancellation must enqueue a new bounded revision, got {delete_payload}")
            pdf_after_delete_status, _, _ = _get_bytes(
                f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}/{upload_id}/payment-validation.pdf"
            )
            if pdf_after_delete_status != 404:
                raise AssertionError(f"deleted upload PDF must be unavailable, got HTTP {pdf_after_delete_status}")
            list_after_delete_status, list_after_delete = _get_json(f"{base_url}{DEFAULT_FULFILLMENT_SERVICES_UPLOADS_PATH}")
            if list_after_delete_status != 200 or any(item.get("upload_id") == upload_id for item in list_after_delete.get("uploads") or []):
                raise AssertionError(f"deleted upload must not be listed, got {list_after_delete_status} {list_after_delete}")
            wb_after_delete_status, wb_after_delete = _get_json(f"{base_url}{DEFAULT_WB_SUPPLIES_PATH}?search=1001&size_filter=all")
            row1001_after_delete = (wb_after_delete.get("rows") or [{}])[0]
            if wb_after_delete_status != 200 or row1001_after_delete.get("fulfillment_amount_with_vat_total") is not None:
                raise AssertionError(f"deleted upload must disappear from WB overlay, got {wb_after_delete_status} {wb_after_delete}")

            operator_status, operator_html = _get_text(f"{base_url}{DEFAULT_SHEET_OPERATOR_UI_PATH}?embedded_tab=factory-order")
            for expected in (
                "Услуги ФФ",
                "Скачать шаблон",
                "Загрузить заполненный файл",
                "Загруженные документы",
                "Дата загрузки",
                "Удалить",
                "Транзит",
                "fulfillment_services_template_path",
                "fulfillment_services_uploads_path",
            ):
                if operator_status != 200 or expected not in operator_html:
                    raise AssertionError(f"operator HTML must expose Fulfillment UI token {expected!r}")
            for forbidden in ("fulfillmentLatestBlock", "fulfillmentLatestSummary", "Услуги fulfillment", "Услуги фулфилмента", ">Fulfillment<", "<th>Стоимость</th>"):
                if forbidden in operator_html:
                    raise AssertionError(f"operator UI must not expose forbidden Fulfillment token {forbidden!r}")
        finally:
            server.shutdown()
            thread.join(timeout=5)

    print("sheet_vitrina_v1_fulfillment_services_smoke: OK")


def _seed_wb_supplies(runtime: RegistryUploadDbBackedRuntime) -> None:
    runtime.save_wb_supply_rows(
        rows=[
            _wb_supply_row("1001", accepted_quantity=10, quantity_added=10, cost_total=200),
            _wb_supply_row("1002", accepted_quantity=20, quantity_added=20, cost_total=400),
            _wb_supply_row("1003", accepted_quantity=0, quantity_added=0, cost_total=600),
        ],
        warehouses=[{"ID": 777, "name": "Электросталь"}],
        synced_at=NOW,
    )


def _wb_supply_row(supply_id: str, *, accepted_quantity: int, quantity_added: int, cost_total: float) -> dict:
    return {
        "supply_id": supply_id,
        "cache_key": f"supply:{supply_id}",
        "wb_supply_id": supply_id,
        "preorder_id": "",
        "visible_number": supply_id,
        "number_label": supply_id,
        "status_id": 5,
        "status_label": "Принято",
        "status_visual_tone": "success",
        "status_class": "success",
        "type_label": "Короб",
        "warehouse_id": "777",
        "warehouse_display": "Электросталь",
        "warehouse_fact_line": "",
        "quantity_added": quantity_added,
        "packed_quantity": quantity_added,
        "accepted_quantity": accepted_quantity,
        "quantity_for_size_filter": accepted_quantity,
        "acceptance_coefficient": 0,
        "cost_total": cost_total,
        "has_transit_cost_marker": True,
        "supply_date": "2026-07-06",
        "fact_date": "2026-07-06",
        "updated_date": "2026-07-06T08:00:00Z",
        "source_created_at": "2026-07-06T07:00:00Z",
        "raw_list_hash": supply_id,
        "raw_list": {"supplyID": supply_id},
        "raw_detail": {"quantity": quantity_added},
        "raw_goods": [{"nmID": 1, "quantity": quantity_added, "acceptedQuantity": accepted_quantity}],
        "raw_package": [],
    }


def _valid_row(supply_id: str) -> list[object]:
    if supply_id == "1002":
        return [supply_id, "Обухово/Краснодар", 3, 450, 0, 850, 1350, 150, 1500, 75]
    return [supply_id, "Сталь", 2, 450, 1, 500, 1400, 100, 1500, 75]


def _valid_row_without_boxes(supply_id: str) -> list[object]:
    return [supply_id, "Сталь", None, None, None, None, None, None, 1500, 75]


def _storage_row(*, total: object = 500, vat: object = 25) -> list[object]:
    return ["STORAGE", "Хранение", None, None, None, None, None, None, total, vat]


def _legacy_headers() -> list[str]:
    headers = list(TEMPLATE_HEADERS)
    headers[1] = "Стоимость услуг"
    return headers


def _build_workbook(rows: list[list[object]], *, headers: list[str] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Услуги ФФ"
    sheet.append(headers or TEMPLATE_HEADERS)
    for row in rows:
        sheet.append(row)
    if rows:
        total = sum(_float_or_zero(row[8]) for row in rows)
        vat = sum(_float_or_zero(row[9]) for row in rows)
        sheet.append(["", "", "", "", "", "", "", "", total, vat])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _float_or_zero(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _read_headers(workbook_bytes: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    return ["" if value is None else str(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def _pdf_text(pdf_bytes: bytes) -> str:
    reader = PdfReader(BytesIO(pdf_bytes))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _reserve_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def _get_json(url: str) -> tuple[int, dict]:
    req = urllib_request.Request(url, method="GET", headers={"Accept": "application/json"})
    return _open_json(req)


def _get_text(url: str) -> tuple[int, str]:
    req = urllib_request.Request(url, method="GET")
    try:
        with urllib_request.urlopen(req, timeout=20) as response:
            return response.status, response.read().decode("utf-8")
    except urllib_error.HTTPError as exc:
        return exc.code, exc.read().decode("utf-8")


def _get_bytes(url: str) -> tuple[int, bytes, dict[str, str]]:
    req = urllib_request.Request(url, method="GET")
    try:
        with urllib_request.urlopen(req, timeout=20) as response:
            return response.status, response.read(), dict(response.headers.items())
    except urllib_error.HTTPError as exc:
        return exc.code, exc.read(), dict(exc.headers.items())


def _delete_json(url: str) -> tuple[int, dict]:
    req = urllib_request.Request(url, method="DELETE", headers={"Accept": "application/json"})
    return _open_json(req)


def _post_multipart(url: str, file_bytes: bytes, *, filename: str) -> tuple[int, dict]:
    boundary = "----wbcorefulfillmentboundary"
    body = b"".join(
        [
            f"--{boundary}\r\n".encode("utf-8"),
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'.encode("utf-8"),
            b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n",
            file_bytes,
            b"\r\n",
            f"--{boundary}--\r\n".encode("utf-8"),
        ]
    )
    req = urllib_request.Request(
        url,
        data=body,
        method="POST",
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Accept": "application/json",
        },
    )
    return _open_json(req)


def _open_json(req: urllib_request.Request) -> tuple[int, dict]:
    try:
        with urllib_request.urlopen(req, timeout=20) as response:
            return response.status, json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        return exc.code, json.loads(exc.read().decode("utf-8"))


if __name__ == "__main__":
    main()
