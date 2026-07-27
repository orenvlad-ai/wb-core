"""Targeted smoke-check for the WB regional supply block."""

from __future__ import annotations

from dataclasses import asdict
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import json
from pathlib import Path
import sys
from tempfile import TemporaryDirectory
from types import SimpleNamespace
import zipfile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.application.factory_order_supply import FactoryOrderSupplyBlock
from packages.application.factory_order_sales_history import persist_sales_history_result_exact_dates
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime
from packages.application.simple_xlsx import build_single_sheet_workbook_bytes, read_first_sheet_rows
from packages.application.wb_supply_overlay import build_warehouse_district_mapping
from packages.application.wb_incident_policy import save_policy_revision
from packages.application.wb_regional_supply import WbRegionalSupplyBlock, _allocate_boxes
from packages.application.wb_regional_supply_export import recommendation_identity, recommendation_prefix
from packages.contracts.factory_order_supply import DATASET_STOCK_FF
from packages.contracts.sales_funnel_history_block import SalesFunnelHistoryItem, SalesFunnelHistorySuccess
from packages.contracts.stocks_block import StocksEnvelope, StocksItem, StocksSuccess
from packages.contracts.wb_regional_supply import (
    DISTRICT_FAR_SIBERIA,
    DISTRICT_NORTHWEST,
    DISTRICT_SOUTH_CAUCASUS,
)
from packages.contracts.wb_supply_planning_zones import (
    PLANNING_ZONE_CENTRAL_EAST,
    PLANNING_ZONE_CENTRAL_NORTH,
    PLANNING_ZONE_CENTRAL_SOUTH,
    SUPPLY_PLANNING_ZONE_KEYS,
)

INPUT_BUNDLE_FIXTURE = (
    ROOT / "artifacts" / "registry_upload_http_entrypoint" / "input" / "registry_upload_bundle__fixture.json"
)
NOW = datetime(2026, 4, 18, 9, 0, tzinfo=timezone.utc)
ACTIVATED_AT = "2026-04-18T09:00:00Z"
MAIN_NM_ID = 210183919


class FakeStocksBlock:
    def __init__(self, nm_ids: list[int]) -> None:
        self.nm_ids = list(nm_ids)

    def execute(self, request_obj: object) -> SimpleNamespace:
        items = []
        warehouse_rows = []
        for nm_id in self.nm_ids:
            central = 0.0
            northwest = 0.0
            if nm_id == MAIN_NM_ID:
                central = 100.0
                northwest = 100.0
            items.append(
                SimpleNamespace(
                    nm_id=nm_id,
                    stock_total=central + northwest,
                    stock_ru_central=central,
                    stock_ru_northwest=northwest,
                    stock_ru_volga=0.0,
                    stock_ru_ural=0.0,
                    stock_ru_south_caucasus=0.0,
                    stock_ru_far_siberia=0.0,
                    stock_ru_central_north=central,
                    stock_ru_central_east=500.0 if nm_id == MAIN_NM_ID else 0.0,
                    stock_ru_central_south=500.0 if nm_id == MAIN_NM_ID else 0.0,
                )
            )
            if nm_id == MAIN_NM_ID:
                warehouse_rows.extend(
                    [
                        SimpleNamespace(
                            nm_id=nm_id,
                            warehouse_id=120762,
                            warehouse_name="Электросталь",
                            quantity=60.0,
                            in_way_to_client=5.0,
                            in_way_from_client=3.0,
                            planning_zone_key=PLANNING_ZONE_CENTRAL_NORTH,
                        ),
                        SimpleNamespace(
                            nm_id=nm_id,
                            warehouse_id=999,
                            warehouse_name="Другой склад",
                            quantity=40.0,
                            in_way_to_client=0.0,
                            in_way_from_client=0.0,
                            planning_zone_key=PLANNING_ZONE_CENTRAL_NORTH,
                        ),
                        SimpleNamespace(
                            nm_id=nm_id,
                            warehouse_id=0,
                            warehouse_name="Остальные",
                            quantity=100.0,
                            in_way_to_client=0.0,
                            in_way_from_client=0.0,
                            planning_zone_key=DISTRICT_NORTHWEST,
                        ),
                    ]
                )
        return SimpleNamespace(
            result=SimpleNamespace(
                kind="success",
                items=items,
                warehouse_rows=warehouse_rows,
                planning_reconciliation={},
                snapshot_date="2026-04-18",
                fetched_at="2026-04-18T09:00:00Z",
                pagination_complete=True,
                raw_rows_digest="sha256:regional-smoke",
            )
        )


class NoopSalesHistoryBlock:
    def execute(self, request_obj: object) -> SimpleNamespace:  # pragma: no cover - should not be called
        raise AssertionError("runtime coverage is fully seeded; live fetch must not be called in smoke")


def main() -> None:
    _check_planning_zone_deficit_allocation()
    bundle = json.loads(INPUT_BUNDLE_FIXTURE.read_text(encoding="utf-8"))
    with TemporaryDirectory(prefix="wb-regional-supply-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        runtime.ingest_bundle(bundle, activated_at=ACTIVATED_AT)
        active_nm_ids = [item.nm_id for item in runtime.load_current_state().config_v2 if item.enabled]
        _seed_runtime_sales_history(runtime, active_nm_ids=active_nm_ids)
        _seed_runtime_stock_history(runtime, active_nm_ids=active_nm_ids)
        _seed_export_nomenclature(runtime, active_nm_ids=active_nm_ids)

        factory_block = FactoryOrderSupplyBlock(
            runtime=runtime,
            now_factory=lambda: NOW,
            timestamp_factory=lambda: ACTIVATED_AT,
        )
        regional_block = WbRegionalSupplyBlock(
            runtime=runtime,
            stocks_block=FakeStocksBlock(active_nm_ids),
            sales_funnel_history_block=NoopSalesHistoryBlock(),
            now_factory=lambda: NOW,
            timestamp_factory=lambda: ACTIVATED_AT,
        )

        stock_template, _ = factory_block.build_template(DATASET_STOCK_FF)
        stock_rows = read_first_sheet_rows(stock_template)
        stock_upload_rows = [list(row) for row in stock_rows]
        for row in stock_upload_rows[1:]:
            row[2] = 0
        for row in stock_upload_rows[1:]:
            if int(row[0]) == MAIN_NM_ID:
                row[2] = 120
                break
        upload_result = factory_block.upload_dataset(
            DATASET_STOCK_FF,
            build_single_sheet_workbook_bytes("Остатки ФФ", stock_upload_rows),
            uploaded_filename="shared-stock-ff.xlsx",
        )
        if upload_result.dataset.uploaded_filename != "shared-stock-ff.xlsx":
            raise AssertionError("shared stock_ff upload must keep the original filename")

        factory_status = factory_block.build_status()
        regional_status = regional_block.build_status()
        if factory_status.datasets["stock_ff"].uploaded_filename != "shared-stock-ff.xlsx":
            raise AssertionError("factory status must expose the shared stock_ff filename")
        if regional_status.shared_datasets["stock_ff"].uploaded_filename != "shared-stock-ff.xlsx":
            raise AssertionError("regional status must reuse the shared stock_ff state")
        if len(regional_status.district_options) != 8:
            raise AssertionError("regional status must expose district options for the operator selector")
        if DISTRICT_FAR_SIBERIA not in regional_status.default_included_district_keys:
            raise AssertionError("regional status default district selection must include far/siberia")

        legacy_payload = {
            "status": "success",
            "calculation_id": "legacy-central-result",
            "calculated_at": "2026-04-17T09:00:00Z",
            "report_date": "2026-04-17",
            "summary": {"total_qty": 50, "estimated_weight": 0.0, "estimated_volume": 0.0},
            "districts": [{"district_key": "central", "total_qty": 50, "rows": []}],
        }
        runtime.save_wb_regional_supply_result_state(
            calculated_at="2026-04-17T09:00:00Z",
            payload=legacy_payload,
        )
        legacy_status = regional_block.build_status()
        legacy_migration = legacy_status.migration_status or {}
        if legacy_status.status != "recalculation_required" or legacy_status.last_result is not None:
            raise AssertionError(f"legacy result must be read fail-closed until recalculation: {legacy_status}")
        if (legacy_migration.get("legacy_snapshot") or {}).get("district_keys") != ["central"]:
            raise AssertionError(f"legacy calculation metadata must remain readable: {legacy_migration}")
        if runtime.load_wb_regional_supply_result_state() != legacy_payload:
            raise AssertionError("lazy migration status must not destroy or rewrite the legacy payload")

        result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
            }
        )
        legacy_alias_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "supply_horizon_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
            }
        )
        if result.summary.total_qty != 100:
            raise AssertionError(f"regional summary total must reflect FF-limited allocation, got {result.summary.total_qty}")
        if result.diagnostics is None or result.diagnostics.get("regional_demand_method") not in {
            "full_clean_days",
            "regional_share_ladder",
        }:
            raise AssertionError(f"result diagnostics must expose share ladder methodology, got {result.diagnostics}")
        if result.diagnostics.get("fallback_sku_count") != 0:
            raise AssertionError(f"share ladder must not use current-stock fallback in this fixture, got {result.diagnostics}")
        if not result.diagnostics.get("share_source_counts"):
            raise AssertionError(f"share source counts must be exposed, got {result.diagnostics}")
        if result.diagnostics.get("requested_valid_day_count") != 14:
            raise AssertionError("result diagnostics must expose requested valid depletion day count")
        if result.settings.included_district_keys != regional_status.default_included_district_keys:
            raise AssertionError("old payload without included_district_keys must default to all districts")
        if result.settings.lead_time_to_region_days_by_district != {key: 2 for key in SUPPLY_PLANNING_ZONE_KEYS}:
            raise AssertionError("legacy scalar lead_time_to_region_days must expand to every district")
        if result.diagnostics.get("district_selection_mode") != "all_districts":
            raise AssertionError("default district selection must be all_districts")
        if legacy_alias_result.summary.total_qty != result.summary.total_qty:
            raise AssertionError("legacy supply_horizon_days alias must keep the same WB regional math")
        districts = {item.district_key: item for item in result.districts}
        if districts[PLANNING_ZONE_CENTRAL_NORTH].total_qty != 50 or districts[PLANNING_ZONE_CENTRAL_NORTH].deficit_qty != 100:
            raise AssertionError("central north zone must keep one box allocated and truthful deficit")
        if districts["northwest"].total_qty != 50 or districts["northwest"].deficit_qty != 100:
            raise AssertionError("northwest district must keep one box allocated and truthful deficit")
        if sum(item.total_qty for item in result.districts) != result.summary.total_qty:
            raise AssertionError("summary total must equal the sum of district totals")
        if sum(item.deficit_qty for item in result.districts) != 200:
            raise AssertionError("deficit totals must equal full recommendation minus allocated supply")
        central_main_row = next(row for row in districts[PLANNING_ZONE_CENTRAL_NORTH].rows if row.nm_id == MAIN_NM_ID)
        if not central_main_row.demand_diagnostics:
            raise AssertionError("district row must carry regional demand diagnostics")
        if central_main_row.demand_diagnostics.get("regional_demand_method") != "full_clean_days":
            raise AssertionError("main SKU must use full-clean-day methodology")
        if central_main_row.demand_diagnostics.get("selected_valid_day_count") != 14:
            raise AssertionError("main SKU must use 14 selected stock-depletion days")
        if abs(central_main_row.daily_demand_total - 60.0) > 1e-9:
            raise AssertionError("total demand must remain based on orderCount, not absolute depletion")
        if central_main_row.lead_time_to_region_days != 2:
            raise AssertionError("legacy scalar lead time must be exposed on district rows")
        if central_main_row.demand_diagnostics.get("lead_time_to_region_days") != 2:
            raise AssertionError("row diagnostics must expose the lead time used by the formula")
        save_policy_revision(
            runtime,
            payload={
                "base_revision": 0,
                "active": True,
                "excluded_wb_warehouse_ids": [120762],
                "reason": "regional smoke incident",
                "effective_from": "2026-04-18",
                "effective_to": "",
                "status": "active",
            },
            actor="regional-smoke",
            warehouse_options=[
                {"warehouse_id": 120762, "warehouse_name": "Электросталь"},
            ],
            timestamp="2026-04-18T09:30:00Z",
        )
        excluded_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "excluded_wb_warehouse_ids": [999],
            }
        )
        excluded_districts = {
            item.district_key: item for item in excluded_result.districts
        }
        excluded_central = next(
            row
            for row in excluded_districts[PLANNING_ZONE_CENTRAL_NORTH].rows
            if row.nm_id == MAIN_NM_ID
        )
        if excluded_central.current_stock != 40:
            raise AssertionError(
                "regional exclusion must happen by warehouseId before planning-zone aggregation"
            )
        if excluded_central.daily_demand_total != central_main_row.daily_demand_total:
            raise AssertionError("warehouse exclusion must not remove demand history")
        exclusion = excluded_result.wb_warehouse_exclusion or {}
        if (
            exclusion.get("actual_stock_total_mp") != 200
            or exclusion.get("excluded_stock_total_mp") != 60
            or exclusion.get("effective_stock_total_mp") != 140
            or exclusion.get("reconciliation_difference") != 0
        ):
            raise AssertionError(f"regional exclusion reconciliation changed: {exclusion}")
        save_policy_revision(
            runtime,
            payload={
                "base_revision": 1,
                "active": False,
                "excluded_wb_warehouse_ids": [120762],
                "reason": "regional smoke incident resolved",
                "effective_from": "2026-04-18",
                "effective_to": "",
                "status": "resolved",
            },
            actor="regional-smoke",
            warehouse_options=[
                {"warehouse_id": 120762, "warehouse_name": "Электросталь"},
            ],
            timestamp="2026-04-18T09:31:00Z",
        )

        legacy_saved_payload = asdict(result)
        legacy_saved_payload["calculation_id"] = "legacy-saved-regional-smoke"
        legacy_saved_payload["settings"].pop("lead_time_to_region_days_by_district", None)
        legacy_saved_payload["diagnostics"].pop("lead_time_to_region_days_by_district", None)
        for district_payload in legacy_saved_payload.get("districts", []):
            district_payload.pop("lead_time_to_region_days", None)
            for row_payload in district_payload.get("rows", []):
                row_payload.pop("lead_time_to_region_days", None)
                diagnostics_payload = row_payload.get("demand_diagnostics")
                if isinstance(diagnostics_payload, dict):
                    diagnostics_payload.pop("lead_time_to_region_days", None)
        runtime.save_wb_regional_supply_result_state(
            calculated_at=result.calculated_at,
            payload=legacy_saved_payload,
        )
        legacy_loaded_status = regional_block.build_status()
        if legacy_loaded_status.last_result is None:
            raise AssertionError("legacy saved regional result must load without falling")
        legacy_loaded_map = legacy_loaded_status.last_result.settings.lead_time_to_region_days_by_district
        if legacy_loaded_map != {key: 2 for key in SUPPLY_PLANNING_ZONE_KEYS}:
            raise AssertionError(f"legacy saved result must restore lead-time map from scalar, got {legacy_loaded_map}")
        legacy_loaded_central = next(
            row
            for district in legacy_loaded_status.last_result.districts
            if district.district_key == PLANNING_ZONE_CENTRAL_NORTH
            for row in district.rows
            if row.nm_id == MAIN_NM_ID
        )
        if legacy_loaded_central.lead_time_to_region_days != 2:
            raise AssertionError("legacy saved row without lead-time must use scalar fallback on status load")

        mixed_lead_times = {key: 2 for key in SUPPLY_PLANNING_ZONE_KEYS}
        mixed_lead_times[DISTRICT_NORTHWEST] = 10
        mixed_lead_time_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days_by_district": mixed_lead_times,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "included_district_keys": [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST],
            }
        )
        mixed_districts = {item.district_key: item for item in mixed_lead_time_result.districts}
        mixed_central_row = next(row for row in mixed_districts[PLANNING_ZONE_CENTRAL_NORTH].rows if row.nm_id == MAIN_NM_ID)
        mixed_northwest_row = next(row for row in mixed_districts[DISTRICT_NORTHWEST].rows if row.nm_id == MAIN_NM_ID)
        expected_central_eta = max(
            mixed_central_row.current_stock
            - mixed_central_row.district_daily_demand * mixed_central_row.lead_time_to_region_days,
            0.0,
        )
        expected_northwest_eta = max(
            mixed_northwest_row.current_stock
            - mixed_northwest_row.district_daily_demand * mixed_northwest_row.lead_time_to_region_days,
            0.0,
        )
        if mixed_central_row.lead_time_to_region_days != 2 or mixed_northwest_row.lead_time_to_region_days != 10:
            raise AssertionError("per-district lead times must be exposed on result rows")
        if abs(mixed_central_row.projected_stock_on_eta - expected_central_eta) > 1e-9:
            raise AssertionError("central projected_stock_on_eta must use central lead time")
        if abs(mixed_northwest_row.projected_stock_on_eta - expected_northwest_eta) > 1e-9:
            raise AssertionError("northwest projected_stock_on_eta must use northwest lead time")
        if mixed_central_row.projected_stock_on_eta <= mixed_northwest_row.projected_stock_on_eta:
            raise AssertionError("longer northwest lead time must lower projected stock in the same-demand fixture")
        if (mixed_lead_time_result.diagnostics or {}).get("lead_time_to_region_days_by_district") != mixed_lead_times:
            raise AssertionError("result diagnostics must expose the applied lead-time map")

        try:
            regional_block.calculate(
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days_by_district": {
                        **mixed_lead_times,
                        "unknown": 3,
                    },
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                }
            )
        except ValueError as exc:
            if "Неизвестный федеральный округ в сроках доставки" not in str(exc):
                raise AssertionError(f"unknown lead-time district key must return clear error, got {exc}") from exc
        else:
            raise AssertionError("unknown lead-time district key must be rejected")

        incomplete_lead_times = dict(mixed_lead_times)
        incomplete_lead_times.pop(DISTRICT_NORTHWEST)
        try:
            regional_block.calculate(
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days_by_district": incomplete_lead_times,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                }
            )
        except ValueError as exc:
            if "Не задан срок доставки для федерального округа" not in str(exc):
                raise AssertionError(f"incomplete lead-time map must return clear error, got {exc}") from exc
        else:
            raise AssertionError("incomplete lead-time map must be rejected")

        _seed_wb_regional_overlay_fixture(
            runtime,
            supply_id="wb-regional-central",
            nm_id=MAIN_NM_ID,
            quantity=50.0,
            supply_date="2026-04-20",
            warehouse_name="Коледино",
            district_key="central",
            warehouse_id=507,
        )
        regional_overlay_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "selected_wb_supply_ids": ["wb-regional-central"],
            }
        )
        overlay_diagnostics = regional_overlay_result.wb_supply_overlay or {}
        overlay_stock = overlay_diagnostics.get("stock_ff", {})
        overlay_regional = overlay_diagnostics.get("wb_regional", {})
        if overlay_stock.get("by_nm_id", {}).get(str(MAIN_NM_ID), {}).get("effective_stock_ff") != 70.0:
            raise AssertionError(f"regional selected supply must reduce available FF pool, got {overlay_stock}")
        added_by_district = overlay_regional.get("added_qty_by_district", {})
        if added_by_district.get(PLANNING_ZONE_CENTRAL_SOUTH) != 50.0:
            raise AssertionError(f"regional selected supply must add qty to mapped district, got {overlay_regional}")
        if any(float(qty or 0) for key, qty in added_by_district.items() if key != PLANNING_ZONE_CENTRAL_SOUTH):
            raise AssertionError(f"selected central supply must not be spread to other districts, got {added_by_district}")
        overlay_districts = {item.district_key: item for item in regional_overlay_result.districts}
        overlay_central_row = next(row for row in overlay_districts[PLANNING_ZONE_CENTRAL_SOUTH].rows if row.nm_id == MAIN_NM_ID)
        overlay_northwest_row = next(row for row in overlay_districts[DISTRICT_NORTHWEST].rows if row.nm_id == MAIN_NM_ID)
        if overlay_central_row.demand_diagnostics.get("selected_wb_supply_qty") != 50.0:
            raise AssertionError("central row diagnostics must expose selected WB qty")
        if overlay_northwest_row.demand_diagnostics.get("selected_wb_supply_qty") != 0.0:
            raise AssertionError("other districts must not receive selected central WB qty")
        if overlay_central_row.projected_stock_on_eta <= central_main_row.projected_stock_on_eta:
            raise AssertionError("selected WB supply must improve projected stock only in the mapped district")
        audit_rows = runtime.list_wb_regional_supply_calculation_audit(limit=5)
        if not audit_rows or audit_rows[0].get("calculation_id") != regional_overlay_result.calculation_id:
            raise AssertionError(f"latest regional audit row must track the last calculation, got {audit_rows}")
        latest_audit = audit_rows[0]
        central_planning_total = sum(
            overlay_districts[key].total_qty
            for key in (
                PLANNING_ZONE_CENTRAL_NORTH,
                PLANNING_ZONE_CENTRAL_EAST,
                PLANNING_ZONE_CENTRAL_SOUTH,
            )
        )
        if latest_audit.get("central_total_qty") != central_planning_total:
            raise AssertionError(f"regional audit must expose central aggregate totals, got {latest_audit}")
        if latest_audit.get("settings", {}).get("selected_wb_supply_ids_count") != 1:
            raise AssertionError(f"regional audit must store selected supply count, got {latest_audit}")
        if "wb-regional-central" in json.dumps(latest_audit, ensure_ascii=False):
            raise AssertionError(f"regional audit must not persist selected WB supply ids, got {latest_audit}")

        _seed_wb_regional_overlay_fixture(
            runtime,
            supply_id="wb-regional-planned",
            nm_id=MAIN_NM_ID,
            quantity=50.0,
            supply_date="2026-04-20",
            warehouse_name="Коледино",
            district_key="central",
            warehouse_id=507,
            status_id=2,
            status_label="Запланировано",
        )
        _seed_wb_regional_overlay_fixture(
            runtime,
            supply_id="wb-regional-accepted",
            nm_id=MAIN_NM_ID,
            quantity=50.0,
            supply_date="2026-04-20",
            warehouse_name="Коледино",
            district_key="central",
            warehouse_id=507,
            status_id=5,
            status_label="Принято",
        )
        skipped_status_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "selected_wb_supply_ids": ["wb-regional-planned", "wb-regional-accepted"],
            }
        )
        skipped_status_overlay = skipped_status_result.wb_supply_overlay or {}
        skipped_status_stock = skipped_status_overlay.get("stock_ff", {})
        skipped_status_regional = skipped_status_overlay.get("wb_regional", {})
        if skipped_status_stock.get("by_nm_id", {}).get(str(MAIN_NM_ID), {}).get("effective_stock_ff") != 120.0:
            raise AssertionError(f"status 2/5 must not reduce regional available FF, got {skipped_status_stock}")
        if skipped_status_regional.get("added_qty_total") != 0.0:
            raise AssertionError(f"status 2/5 must not add regional projected stock, got {skipped_status_regional}")
        if not any("Запланировано" in warning for warning in skipped_status_result.warnings) or not any(
            "Принято" in warning for warning in skipped_status_result.warnings
        ):
            raise AssertionError(f"status 2/5 direct payload skips must be warned, got {skipped_status_result.warnings}")

        _seed_wb_regional_overlay_fixture(
            runtime,
            supply_id="wb-regional-provider-mapped",
            nm_id=MAIN_NM_ID,
            quantity=25.0,
            supply_date="2026-04-20",
            warehouse_name="Электросталь",
            district_key="unmapped",
            warehouse_id=120762,
        )
        regional_block.wb_supply_district_mapping_provider = lambda: build_warehouse_district_mapping(
            warehouse_rows=runtime.list_wb_supplies_warehouses(),
            supply_rows=runtime.list_wb_supplies(),
            office_rows=[{"name": "Электросталь", "federalDistrict": "Центральный федеральный округ"}],
        )
        provider_overlay_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "selected_wb_supply_ids": ["wb-regional-provider-mapped"],
            }
        )
        provider_overlay = provider_overlay_result.wb_supply_overlay or {}
        provider_regional = provider_overlay.get("wb_regional", {})
        provider_added_by_district = provider_regional.get("added_qty_by_district", {})
        if provider_added_by_district.get(PLANNING_ZONE_CENTRAL_EAST) != 25.0:
            raise AssertionError(f"provider mapping must map historical Elektrostal to central east, got {provider_regional}")
        if any(float(qty or 0) for key, qty in provider_added_by_district.items() if key != PLANNING_ZONE_CENTRAL_EAST):
            raise AssertionError(f"provider-mapped central supply must not affect other districts, got {provider_added_by_district}")

        _seed_wb_regional_overlay_fixture(
            runtime,
            supply_id="wb-regional-krasnodar-transit",
            nm_id=MAIN_NM_ID,
            quantity=30.0,
            supply_date="2026-04-20",
            warehouse_name="Краснодар (Тихорецкая)",
            district_key="central",
            actual_warehouse_name="Обухово",
            transit_warehouse_name="Обухово",
        )
        regional_block.wb_supply_district_mapping_provider = lambda: build_warehouse_district_mapping(
            warehouse_rows=runtime.list_wb_supplies_warehouses(),
            supply_rows=runtime.list_wb_supplies(),
            tariff_rows=[{"warehouseName": "Обухово", "geoName": "Центральный федеральный округ"}],
        )
        routed_overlay_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "selected_wb_supply_ids": ["wb-regional-krasnodar-transit"],
            }
        )
        routed_overlay = routed_overlay_result.wb_supply_overlay or {}
        routed_regional = routed_overlay.get("wb_regional", {})
        routed_added_by_district = routed_regional.get("added_qty_by_district", {})
        if routed_added_by_district.get(DISTRICT_SOUTH_CAUCASUS) != 30.0:
            raise AssertionError(f"planned Краснодар warehouse must add qty to south_caucasus, got {routed_regional}")
        if any(
            routed_added_by_district.get(key) != 0.0
            for key in (
                PLANNING_ZONE_CENTRAL_NORTH,
                PLANNING_ZONE_CENTRAL_EAST,
                PLANNING_ZONE_CENTRAL_SOUTH,
            )
        ):
            raise AssertionError(f"actual/transit Обухово must not leak qty into central, got {routed_regional}")
        routed_selected = (routed_overlay.get("selected_supplies") or [{}])[0]
        if (
            routed_selected.get("district_source_warehouse_name") != "Краснодар (Тихорецкая)"
            or routed_selected.get("warehouse_display") != "Краснодар (Тихорецкая) → Обухово"
            or routed_selected.get("district_key") != DISTRICT_SOUTH_CAUCASUS
        ):
            raise AssertionError(f"regional overlay diagnostics must expose planned district source, got {routed_selected}")

        _seed_wb_regional_overlay_fixture(
            runtime,
            supply_id="wb-regional-unmapped",
            nm_id=MAIN_NM_ID,
            quantity=10.0,
            supply_date="2026-04-20",
            warehouse_name="Склад без ФО",
            district_key="unmapped",
        )
        unmapped_overlay_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "selected_wb_supply_ids": ["wb-regional-unmapped"],
            }
        )
        unmapped_overlay = unmapped_overlay_result.wb_supply_overlay or {}
        unmapped_regional = unmapped_overlay.get("wb_regional", {})
        if unmapped_regional.get("added_qty_total") != 0.0 or not unmapped_regional.get("unmapped_events"):
            raise AssertionError(f"unmapped warehouse must not add regional qty and must be diagnosed, got {unmapped_regional}")
        if not any("склад не сопоставлен" in warning for warning in unmapped_overlay_result.warnings):
            raise AssertionError(f"unmapped warehouse must emit regional warning, got {unmapped_overlay_result.warnings}")

        selected_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "included_district_keys": [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST],
            }
        )
        selected_diagnostics = selected_result.diagnostics or {}
        if selected_diagnostics.get("included_district_keys") != [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST]:
            raise AssertionError(f"selected district diagnostics not exposed: {selected_diagnostics}")
        if DISTRICT_FAR_SIBERIA not in selected_diagnostics.get("excluded_district_keys", []):
            raise AssertionError("selected district diagnostics must include excluded far/siberia")
        selected_district_keys = [item.district_key for item in selected_result.districts]
        if selected_district_keys != [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST]:
            raise AssertionError(f"selected result must expose included districts only, got {selected_district_keys}")

        try:
            regional_block.calculate(
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                    "included_district_keys": [],
                }
            )
        except ValueError as exc:
            if "Выберите хотя бы один округ" not in str(exc):
                raise AssertionError(f"empty selected districts must return clear validation error, got {exc}") from exc
        else:
            raise AssertionError("empty selected districts must be rejected")

        central_workbook, central_filename = regional_block.download_district_recommendation(PLANNING_ZONE_CENTRAL_NORTH)
        central_rows = read_first_sheet_rows(central_workbook)
        if central_filename != "wb_regional_central_north_fo.xlsx" or not _is_ascii(central_filename):
            raise AssertionError(f"central district filename must be stable ASCII, got {central_filename!r}")
        if central_rows[0][:2] != ["Федеральный округ", "ЦФО Север"]:
            raise AssertionError("planning-zone workbook must start with direction identification")
        if central_rows[2] != ["nmId", "SKU", "Количество к поставке"]:
            raise AssertionError("district workbook must keep compact Russian headers without deficit")
        if any(cell == "Дефицит" for row in central_rows for cell in row):
            raise AssertionError("district workbook must not export the removed deficit column")
        load_workbook(BytesIO(central_workbook), data_only=True)
        central_allocated_sum = sum(int(row[2]) for row in central_rows[3:] if len(row) >= 3 and str(row[2]).strip())
        if central_allocated_sum != districts[PLANNING_ZONE_CENTRAL_NORTH].total_qty:
            raise AssertionError("district workbook sum must equal district total in summary")

        try:
            regional_block.download_district_recommendation("far_siberia")
        except ValueError as exc:
            if "Округ не участвовал в последнем расчёте: far_siberia" not in str(exc):
                raise AssertionError(f"excluded district download must return clear error, got {exc}") from exc
        else:
            raise AssertionError("excluded district direct download must be blocked")

        archive_bytes, archive_filename = regional_block.download_all_recommendations_archive()
        if (
            not archive_filename.startswith("Рекомендации_поставок_2026-04-18_14-00_")
            or selected_result.calculation_id not in archive_filename
            or not archive_filename.endswith(".zip")
        ):
            raise AssertionError(f"ZIP filename must carry calculation identity and stable timestamp, got {archive_filename!r}")
        expected_archive_names: list[str] = []
        for ordinal, district in enumerate(selected_result.districts, start=1):
            recommendation_id = recommendation_identity(
                report_date=selected_result.report_date,
                calculation_id=selected_result.calculation_id,
                ordinal=ordinal,
            )
            prefix = recommendation_prefix(
                ordinal=ordinal,
                recommendation_id=recommendation_id,
                destination_name=district.planning_zone_label or district.district_name_ru,
            )
            expected_archive_names.extend(
                [
                    f"{prefix}/{prefix}__01_РЕКОМЕНДАЦИЯ.xlsx",
                    f"{prefix}/{prefix}__02_ЗАГРУЗКА_WB.xlsx",
                ]
            )
        with zipfile.ZipFile(BytesIO(archive_bytes)) as archive:
            archive_names = archive.namelist()
            if archive_names != expected_archive_names:
                raise AssertionError(f"ZIP must contain paired district XLSX folders in UI order, got {archive_names}")
            for name in archive_names:
                load_workbook(BytesIO(archive.read(name)), data_only=True)

        _seed_runtime_sales_history(runtime, active_nm_ids=active_nm_ids, all_active_signal=False)
        _seed_runtime_stock_history(
            runtime,
            active_nm_ids=active_nm_ids,
            all_active_signal=False,
            persistent_zero_south_for_main=True,
        )
        seed_stock_upload_rows = [list(row) for row in stock_rows]
        for row in seed_stock_upload_rows[1:]:
            row[2] = 0
        for row in seed_stock_upload_rows[1:]:
            if int(row[0]) == MAIN_NM_ID:
                row[2] = 400
                break
        factory_block.upload_dataset(
            DATASET_STOCK_FF,
            build_single_sheet_workbook_bytes("Остатки ФФ", seed_stock_upload_rows),
            uploaded_filename="shared-stock-ff-seed.xlsx",
        )
        seed_result = regional_block.calculate(
            {
                "sales_avg_period_days": 14,
                "cycle_supply_days": 5,
                "lead_time_to_region_days": 2,
                "safety_days": 1,
                "order_batch_qty": 50,
                "report_date_override": "2026-04-18",
                "included_district_keys": [PLANNING_ZONE_CENTRAL_NORTH, DISTRICT_NORTHWEST, "south_caucasus"],
            }
        )
        seed_diagnostics = seed_result.diagnostics or {}
        if seed_diagnostics.get("fallback_sku_count") != 0:
            raise AssertionError(f"seed-floor fixture must not need current-stock fallback, got {seed_diagnostics}")
        if seed_diagnostics.get("seed_floor_sku_district_count", 0) < 1:
            raise AssertionError(f"seed-floor diagnostics must count affected SKU-districts, got {seed_diagnostics}")
        if seed_diagnostics.get("seed_sku_count") != 1 or seed_diagnostics.get("seed_allocated_qty_total") != 50:
            raise AssertionError(f"seed diagnostics must expose one allocated test box, got {seed_diagnostics}")
        seed_districts = {item.district_key: item for item in seed_result.districts}
        south_seed_row = next(row for row in seed_districts["south_caucasus"].rows if row.nm_id == MAIN_NM_ID)
        if south_seed_row.district_daily_demand != 0:
            raise AssertionError("seed floor must not create demand-based district demand")
        if south_seed_row.seed_qty != 50 or not south_seed_row.seed_floor_applied:
            raise AssertionError(f"south_caucasus row must carry one seed box, got {south_seed_row}")
        if south_seed_row.demand_allocated_qty != 0 or south_seed_row.allocated_qty != 50:
            raise AssertionError("seed row must separate demand allocation from test-box allocation")
        if south_seed_row.share_source != "seed_floor":
            raise AssertionError(f"seed row must expose seed_floor share source, got {south_seed_row}")
        if seed_result.summary.total_qty > 400:
            raise AssertionError("seed allocation must not exceed available stock_ff")
        south_workbook, _ = regional_block.download_district_recommendation("south_caucasus")
        south_rows = read_first_sheet_rows(south_workbook)
        south_allocated_sum = sum(int(row[2]) for row in south_rows[3:] if len(row) >= 3 and str(row[2]).strip())
        if south_allocated_sum != seed_districts["south_caucasus"].total_qty:
            raise AssertionError("district XLSX total must include seed-floor qty")

        latest_before_export_failure = runtime.load_wb_regional_supply_result_state()
        registry_total_before_export_failure = runtime.list_supply_calculation_registry(
            calculation_type="wb_regional",
            limit=1,
        )["pagination"]["total"]
        broken_nomenclature = runtime.list_nomenclature_items(active_only=True)
        for item in broken_nomenclature:
            if int(item.get("nm_id") or 0) == MAIN_NM_ID:
                item["barcode"] = ""
                item["barcodes"] = []
                item["barcode_status"] = "missing"
        runtime.save_nomenclature_items_atomic(broken_nomenclature)
        try:
            regional_block.calculate(
                {
                    "sales_avg_period_days": 14,
                    "cycle_supply_days": 5,
                    "lead_time_to_region_days": 2,
                    "safety_days": 1,
                    "order_batch_qty": 50,
                    "report_date_override": "2026-04-18",
                    "included_district_keys": [
                        PLANNING_ZONE_CENTRAL_NORTH,
                        DISTRICT_NORTHWEST,
                        "south_caucasus",
                    ],
                }
            )
        except ValueError as exc:
            if (
                "отсутствует баркод" not in str(exc)
                and "баркод имеет неподтверждённый статус" not in str(exc)
            ):
                raise AssertionError(
                    f"historical ZIP build failure must stay explicit: {exc}"
                ) from exc
        else:
            raise AssertionError(
                "regional calculate must fail before persistence when exact ZIP cannot be built"
            )
        if (
            runtime.load_wb_regional_supply_result_state()
            != latest_before_export_failure
            or runtime.list_supply_calculation_registry(
                calculation_type="wb_regional",
                limit=1,
            )["pagination"]["total"]
            != registry_total_before_export_failure
        ):
            raise AssertionError(
                "regional export-build failure must leave latest and immutable history unchanged"
            )

        print(f"shared_stock_ff_reuse: ok -> {regional_status.shared_datasets['stock_ff'].uploaded_filename}")
        print(f"regional_total_qty: ok -> {result.summary.total_qty}")
        print(f"central_north_deficit: ok -> {districts[PLANNING_ZONE_CENTRAL_NORTH].deficit_qty}")
        print(f"northwest_deficit: ok -> {districts['northwest'].deficit_qty}")
        print(f"regional_lead_times: ok -> central {mixed_central_row.lead_time_to_region_days}, northwest {mixed_northwest_row.lead_time_to_region_days}")
        print(f"seed_floor: ok -> {seed_diagnostics.get('seed_allocated_qty_total')}")
        print(f"regional_audit: ok -> {latest_audit.get('calculation_id')}")
        print(f"district_xlsx_sum: ok -> {central_allocated_sum}")
        print(f"recommendations_zip: ok -> {archive_names}")


def _check_planning_zone_deficit_allocation() -> None:
    kwargs = {
        "full_recommendation_by_key": {
            PLANNING_ZONE_CENTRAL_NORTH: 300,
            PLANNING_ZONE_CENTRAL_EAST: 100,
            PLANNING_ZONE_CENTRAL_SOUTH: 0,
        },
        "raw_recommendation_by_key": {
            PLANNING_ZONE_CENTRAL_NORTH: 260.0,
            PLANNING_ZONE_CENTRAL_EAST: 90.0,
            PLANNING_ZONE_CENTRAL_SOUTH: 0.0,
        },
        "district_daily_demand_by_key": {
            PLANNING_ZONE_CENTRAL_NORTH: 20.0,
            PLANNING_ZONE_CENTRAL_EAST: 10.0,
            PLANNING_ZONE_CENTRAL_SOUTH: 0.0,
        },
        "projected_stock_by_key": {
            PLANNING_ZONE_CENTRAL_NORTH: 0.0,
            PLANNING_ZONE_CENTRAL_EAST: 0.0,
            PLANNING_ZONE_CENTRAL_SOUTH: 500.0,
        },
        "available_stock_ff": 200.0,
        "order_batch_qty": 50,
    }
    first = _allocate_boxes(**kwargs)
    second = _allocate_boxes(**kwargs)
    if first != {
        PLANNING_ZONE_CENTRAL_NORTH: 150,
        PLANNING_ZONE_CENTRAL_EAST: 50,
        PLANNING_ZONE_CENTRAL_SOUTH: 0,
    }:
        raise AssertionError(f"zones must be allocated by deficits, not equal thirds: {first}")
    if second != first or sum(first.values()) > 200 or any(value % 50 for value in first.values()):
        raise AssertionError(f"planning-zone allocation must be bounded and deterministic: {first}, {second}")
    print(f"planning_zone_deficit_allocation: ok -> {first}")


def _seed_wb_regional_overlay_fixture(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    supply_id: str,
    nm_id: int,
    quantity: float,
    supply_date: str,
    warehouse_name: str,
    district_key: str,
    warehouse_id: int | str | None = None,
    actual_warehouse_name: str = "",
    transit_warehouse_name: str = "",
    status_id: int = 3,
    status_label: str = "Отгрузка разрешена",
) -> None:
    effective_warehouse_id = str(warehouse_id or supply_id)
    warehouse_display = (
        f"{warehouse_name} → {transit_warehouse_name}"
        if transit_warehouse_name and transit_warehouse_name != warehouse_name
        else warehouse_name
    )
    runtime.save_wb_supply_rows(
        rows=[
            {
                "supply_id": supply_id,
                "cache_key": supply_id,
                "wb_supply_id": supply_id,
                "preorder_id": "pre-" + supply_id,
                "number_label": supply_id,
                "status_id": status_id,
                "status_label": status_label,
                "warehouse_id": effective_warehouse_id,
                "warehouse_name": warehouse_name,
                "planned_warehouse_id": effective_warehouse_id,
                "planned_warehouse_name": warehouse_name,
                "target_warehouse_id": effective_warehouse_id,
                "target_warehouse_name": warehouse_name,
                "actual_warehouse_id": ("actual-" + supply_id) if actual_warehouse_name else "",
                "actual_warehouse_name": actual_warehouse_name,
                "transit_warehouse_id": ("transit-" + supply_id) if transit_warehouse_name else "",
                "transit_warehouse_name": transit_warehouse_name,
                "warehouse_from_name": warehouse_name,
                "warehouse_to_name": transit_warehouse_name,
                "warehouse_actual_name": actual_warehouse_name,
                "warehouse_display": warehouse_display,
                "district_source_warehouse_id": effective_warehouse_id,
                "district_source_warehouse_name": warehouse_name,
                "district_source_warehouse_role": "planned",
                "district_source_warehouse_evidence": "fixture.warehouse_name",
                "supply_date": supply_date,
                "district_key": district_key,
                "district_label_ru": "",
                "quantity_for_size_filter": quantity,
                "raw_list": {"supplyID": supply_id, "statusID": status_id, "supplyDate": supply_date},
                "raw_detail": {
                    "warehouseName": warehouse_name,
                    "actualWarehouseName": actual_warehouse_name,
                    "transitWarehouseName": transit_warehouse_name,
                },
                "raw_goods": [{"nmID": int(nm_id), "quantity": float(quantity)}],
                "raw_package": [],
            }
        ],
        warehouses=[{"warehouse_id": effective_warehouse_id, "warehouse_name": warehouse_name}],
        synced_at=ACTIVATED_AT,
    )


def _is_ascii(value: str) -> bool:
    try:
        value.encode("ascii")
    except UnicodeEncodeError:
        return False
    return True


def _seed_export_nomenclature(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    active_nm_ids: list[int],
) -> None:
    runtime.save_nomenclature_items_atomic(
        [
            {
                "item_id": f"regional-export-{int(nm_id)}",
                "is_active": True,
                "nm_id": int(nm_id),
                "barcode": f"46{int(nm_id)}",
                "barcodes": [f"46{int(nm_id)}"],
                "barcode_source": "manual",
                "barcode_status": "manual",
                "nomenclature_name": f"Regional export SKU {int(nm_id)}",
                "product_type": "clear",
                "match_key": f"regional-export-{int(nm_id)}",
                "created_at": ACTIVATED_AT,
                "updated_at": ACTIVATED_AT,
            }
            for nm_id in active_nm_ids
        ]
    )


def _seed_runtime_sales_history(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    active_nm_ids: list[int],
    all_active_signal: bool = False,
) -> None:
    report_date = date(2026, 4, 18)
    items: list[SalesFunnelHistoryItem] = []
    for offset in range(14, 0, -1):
        snapshot_date = (report_date - timedelta(days=offset)).isoformat()
        for nm_id in active_nm_ids:
            value = 60.0 if nm_id == MAIN_NM_ID or all_active_signal else 0.0
            items.append(
                SalesFunnelHistoryItem(
                    date=snapshot_date,
                    nm_id=int(nm_id),
                    metric="orderCount",
                    value=value,
                )
            )
    persist_sales_history_result_exact_dates(
        runtime=runtime,
        payload=SalesFunnelHistorySuccess(
            kind="success",
            date_from="2026-04-04",
            date_to="2026-04-17",
            count=len(items),
            items=items,
        ),
        captured_at=ACTIVATED_AT,
    )


def _seed_runtime_stock_history(
    runtime: RegistryUploadDbBackedRuntime,
    *,
    active_nm_ids: list[int],
    all_active_signal: bool = False,
    persistent_zero_south_for_main: bool = False,
) -> None:
    first_snapshot = date(2026, 4, 3)
    for index in range(15):
        snapshot_date = first_snapshot + timedelta(days=index)
        items: list[StocksItem] = []
        for nm_id in active_nm_ids:
            central = 0.0
            northwest = 0.0
            volga = 0.0
            ural = 0.0
            south = 0.0
            far = 0.0
            if nm_id == MAIN_NM_ID or all_active_signal:
                central = 1000.0 - (10.0 * index)
                northwest = 1000.0 - (10.0 * index)
                volga = 500.0
                ural = 500.0
                south = 0.0 if persistent_zero_south_for_main and nm_id == MAIN_NM_ID else 500.0
                far = 500.0
            items.append(
                StocksItem(
                    nm_id=int(nm_id),
                    stock_total=central + northwest + volga + ural + south + far,
                    stock_ru_central=central,
                    stock_ru_northwest=northwest,
                    stock_ru_volga=volga,
                    stock_ru_ural=ural,
                    stock_ru_south_caucasus=south,
                    stock_ru_far_siberia=far,
                    stock_ru_central_north=central,
                    stock_ru_central_east=500.0 if nm_id == MAIN_NM_ID or all_active_signal else 0.0,
                    stock_ru_central_south=500.0 if nm_id == MAIN_NM_ID or all_active_signal else 0.0,
                )
            )
        runtime.save_temporal_source_snapshot(
            source_key="stocks",
            snapshot_date=snapshot_date.isoformat(),
            captured_at=ACTIVATED_AT,
            payload=StocksEnvelope(
                result=StocksSuccess(
                    kind="success",
                    snapshot_date=snapshot_date.isoformat(),
                    count=len(items),
                    items=items,
                )
            ),
        )


if __name__ == "__main__":
    main()
