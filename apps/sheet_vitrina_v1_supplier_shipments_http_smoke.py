"""HTTP smoke-check for supplier invoice shipment parse/storage/API routes."""

from __future__ import annotations

from io import BytesIO
import hashlib
import json
from pathlib import Path
import socket
import sqlite3
import sys
from tempfile import TemporaryDirectory
import threading
import time
from unittest.mock import patch
from urllib import error as urllib_error, request as urllib_request
from uuid import uuid4

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.adapters.registry_upload_http_entrypoint import (  # noqa: E402
    DEFAULT_NOMENCLATURE_EXPORT_PATH,
    DEFAULT_NOMENCLATURE_IMPORT_PATH,
    DEFAULT_NOMENCLATURE_PATH,
    DEFAULT_SHEET_OPERATOR_UI_PATH,
    DEFAULT_SHEET_PLAN_PATH,
    DEFAULT_SHEET_STATUS_PATH,
    DEFAULT_SUPPLIER_SHIPMENT_REGISTRY_PATH,
    DEFAULT_SUPPLIER_SHIPMENTS_PARSE_PATH,
    DEFAULT_SUPPLIER_SHIPMENTS_PATH,
    DEFAULT_UPLOAD_PATH,
    build_registry_upload_http_server,
)
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime  # noqa: E402
from packages.application.registry_upload_http_entrypoint import RegistryUploadHttpEntrypoint  # noqa: E402
from packages.application.supplier_shipments import SupplierShipmentsBlock  # noqa: E402
from packages.application.ff_pool_dense_fbs import DenseFbsService  # noqa: E402
from packages.contracts.registry_upload_http_entrypoint import RegistryUploadHttpEntrypointConfig  # noqa: E402


TARGET_FACILITY_ID = "fac_supplier_smoke"
INACTIVE_TARGET_FACILITY_ID = "fac_supplier_smoke_inactive"


def _seed_target_facilities(runtime: RegistryUploadDbBackedRuntime) -> None:
    staged_at = "2026-05-30T08:00:00Z"
    with sqlite3.connect(runtime.db_path) as conn:
        latest_epoch = conn.execute(
            """SELECT epoch,writer_enabled
                 FROM sheet_vitrina_v1_ff_pool_feature_epochs
                ORDER BY epoch DESC LIMIT 1"""
        ).fetchone()
        if latest_epoch is None or not bool(latest_epoch[1]):
            conn.execute(
                """INSERT INTO sheet_vitrina_v1_ff_pool_feature_epochs(
                       epoch,writer_enabled,reader_enabled,source_revision,
                       created_at,metadata_json
                   ) VALUES(?,?,?,?,?,?)""",
                (
                    1 if latest_epoch is None else int(latest_epoch[0]) + 1,
                    1,
                    1,
                    "supplier-shipments-http-dense-fixture-v1",
                    staged_at,
                    json.dumps(
                        {
                            "fixture": "supplier_shipments_http",
                            "purpose": "production_shaped_dense_fbs_activation",
                        },
                        sort_keys=True,
                        separators=(",", ":"),
                    ),
                ),
            )
        conn.executemany(
            """
            INSERT INTO sheet_vitrina_v1_ff_facilities(
                facility_id,code,name,active,display_timezone,created_at,updated_at
            ) VALUES(?,?,?,?,?,?,?)
            """,
            (
                (
                    TARGET_FACILITY_ID,
                    "SMOKE",
                    "FF Smoke",
                    0,
                    "Europe/Moscow",
                    staged_at,
                    staged_at,
                ),
                (
                    INACTIVE_TARGET_FACILITY_ID,
                    "SMOKE-OFF",
                    "FF Smoke inactive",
                    0,
                    "Europe/Moscow",
                    staged_at,
                    staged_at,
                ),
            ),
        )
        conn.commit()
    DenseFbsService(
        db_path=runtime.db_path,
        runtime_dir=runtime.runtime_dir,
        timestamp_factory=lambda: "2026-05-30T10:00:00Z",
    ).activate_facility(
        facility_id=TARGET_FACILITY_ID,
        expected_updated_at=staged_at,
        request_id="supplier-shipments-http-facility-activation",
        request_identity="sha256:"
        + hashlib.sha256(b"supplier-shipments-http-facility-activation").hexdigest(),
        actor="supplier-shipments-http-smoke",
    )
    with sqlite3.connect(runtime.db_path) as conn:
        incomplete = conn.execute(
            """SELECT item.nm_id,balance.quantity,balance.capital_rub,balance.wac_rub
                 FROM sheet_vitrina_v1_nomenclature_items item
                 LEFT JOIN sheet_vitrina_v1_ff_pool_balances balance
                   ON balance.facility_id=? AND balance.pool='FBS'
                  AND balance.nm_id=item.nm_id
                WHERE item.is_active=1 AND item.is_hidden=0 AND item.nm_id>0
                  AND (balance.nm_id IS NULL OR balance.quantity<>0
                       OR CAST(balance.capital_rub AS NUMERIC)<>0
                       OR balance.wac_rub IS NOT NULL)
                ORDER BY item.nm_id""",
            (TARGET_FACILITY_ID,),
        ).fetchall()
        if incomplete:
            raise AssertionError(
                "supplier fixture facility must have complete canonical-zero FBS coverage: "
                f"{incomplete}"
            )


def _assert_dense_sku_activation(
    runtime: RegistryUploadDbBackedRuntime, *, nm_id: int
) -> None:
    with sqlite3.connect(runtime.db_path) as conn:
        item = conn.execute(
            """SELECT item_id,is_active,is_hidden
                 FROM sheet_vitrina_v1_nomenclature_items WHERE nm_id=?""",
            (int(nm_id),),
        ).fetchone()
        balance = conn.execute(
            """SELECT quantity,capital_rub,wac_rub
                 FROM sheet_vitrina_v1_ff_pool_balances
                WHERE facility_id=? AND pool='FBS' AND nm_id=?""",
            (TARGET_FACILITY_ID, int(nm_id)),
        ).fetchone()
        intent = conn.execute(
            """SELECT intent.intent_id
                 FROM sheet_vitrina_v1_ff_pool_fbs_dense_intents intent,
                      json_each(intent.plan_json,'$.materialize_nm_ids') materialized_nm
                WHERE intent.subject_kind='sku_activation'
                  AND CAST(materialized_nm.value AS INTEGER)=?
                ORDER BY intent.created_at DESC,intent.intent_id DESC LIMIT 1""",
            (int(nm_id),),
        ).fetchone()
        states = (
            {
                str(row[0])
                for row in conn.execute(
                    """SELECT state
                         FROM sheet_vitrina_v1_ff_pool_fbs_dense_intent_events
                        WHERE intent_id=? ORDER BY event_sequence""",
                    (str(intent[0]),),
                ).fetchall()
            }
            if intent is not None
            else set()
        )
    if item is None or tuple(item[1:]) != (1, 0):
        raise AssertionError(f"dense SKU must publish active only after coverage: {item}")
    if balance is None or tuple(balance) != (0, "0", None):
        raise AssertionError(f"dense SKU must materialize canonical explicit zero: {balance}")
    if not {"staged", "materializing", "materialized", "active"} <= states:
        raise AssertionError(f"dense SKU lifecycle receipt is incomplete: {states}")


def _assert_price_conformity_application_smoke() -> None:
    timestamp_counter = {"value": 0}

    def next_timestamp() -> str:
        timestamp_counter["value"] += 1
        return f"2026-05-30T09:{timestamp_counter['value']:02d}:00Z"

    with TemporaryDirectory(prefix="supplier-price-conformity-") as tmp:
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=Path(tmp) / "runtime")
        block = SupplierShipmentsBlock(runtime=runtime, timestamp_factory=next_timestamp)
        first = block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501001,
                "barcode": "1111111111111",
                "nomenclature_name": "Clear iPhone 14 Pro",
                "product_type": "clean",
                "match_key": "clean|iphone_14_pro",
                "purchase_price_yuan": "1",
            }
        )["item"]
        block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501002,
                "barcode": "2222222222222",
                "nomenclature_name": "Anti-Spy iPhone 14",
                "product_type": "anti_spy",
                "match_key": "anti_spy|iphone_14",
                "purchase_price_yuan": "3",
            }
        )
        _seed_target_facilities(runtime)
        block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501003,
                "barcode": "4444444444444",
                "nomenclature_name": "Clear iPhone 15",
                "product_type": "clean",
                "match_key": "clean|iphone_15",
                "purchase_price_yuan": None,
            }
        )
        _assert_dense_sku_activation(runtime, nm_id=501003)
        block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501004,
                "barcode": "5555555555555",
                "nomenclature_name": "Matte iPhone 16",
                "product_type": "matte",
                "match_key": "matte|iphone_16",
                "purchase_price_yuan": "5",
            }
        )

        parsed = block.parse_upload(_build_price_conformity_invoice_fixture(), uploaded_filename="price-check.xlsx")
        statuses = [line.get("price_conformity_status") for line in parsed.get("lines", []) if line.get("line_type") == "product"]
        expected = ["matched", "mismatched", "sku_not_found", "reference_price_missing", "invoice_price_missing"]
        if statuses != expected:
            raise AssertionError(f"price conformity parse statuses changed: {statuses}")
        block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501005,
                "barcode": "3333333333333",
                "nomenclature_name": "Temporary exact barcode owner",
                "product_type": "other",
                "match_key": "",
                "purchase_price_yuan": "4",
            }
        )
        accepted_parsed = json.loads(json.dumps(parsed))
        accepted_products = [line for line in accepted_parsed.get("lines") or [] if line.get("line_type") == "product"]
        accepted_products[-1]["unit_price"] = 5
        accepted_products[-1]["amount"] = 5
        detail = block.create_shipment(
            {
                "upload_id": parsed["upload_id"],
                "shipment_date": "2026-05-30",
                "target_facility_id": TARGET_FACILITY_ID,
                "payload": accepted_parsed,
            }
        )
        shipment_id = detail["shipment_id"]
        if detail["product_lines"][0].get("price_conformity_check_mode") != "initial_parse":
            raise AssertionError("initial create must persist initial_parse price check mode")
        block.update_nomenclature_item(str(first["item_id"]), {**first, "purchase_price_yuan": "1.25"})
        ordinary_open = block.get_shipment(shipment_id)
        if (
            ordinary_open["product_lines"][0].get("price_conformity_status") != "matched"
            or ordinary_open["product_lines"][0].get("reference_purchase_price_yuan_snapshot") != 1.0
        ):
            raise AssertionError("ordinary detail open must not auto-recalculate price conformity")
        rechecked = block.recheck_shipment_prices(
            shipment_id,
            actor="operator:smoke",
            context={"source": "application_smoke"},
        )
        first_line = rechecked["product_lines"][0]
        if (
            first_line.get("price_conformity_status") != "mismatched"
            or first_line.get("reference_purchase_price_yuan_snapshot") != 1.25
            or first_line.get("price_conformity_check_mode") != "manual_recheck"
            or first_line.get("price_conformity_actor") != "operator:smoke"
            or first_line.get("price_conformity_context", {}).get("source") != "application_smoke"
        ):
            raise AssertionError(f"manual recheck must update persisted price metadata, got {first_line}")

        runtime.save_supplier_shipment(
            header={
                "shipment_id": "sup_legacy_price_check",
                "created_at": "2026-05-30T08:10:00Z",
                "updated_at": "2026-05-30T08:10:00Z",
                "shipment_date": "2026-05-16",
                "invoice_no": "LEGACY-PRICE",
                "invoice_date": "2026-05-15",
                "contract_no": "",
                "contract_date": "",
                "supplier_name": "",
                "customer_name": "",
                "currency": "RMB",
                "product_qty_total": 1,
                "product_amount_total": 1,
                "extras_amount_total": 0,
                "invoice_amount_total": 1,
                "declared_invoice_total": 1,
                "match_status": "all_matched",
                "source_filename": "legacy.xlsx",
                "source_file_sha256": "",
                "source_file_path": "",
                "parser_version": "legacy",
                "warnings": [],
                "errors": [],
            },
            lines=[
                {
                    "line_id": "ln_legacy_price",
                    "line_type": "product",
                    "sort_order": 1,
                    "product_type": "anti_spy",
                    "model_raw": "iPhone 14",
                    "model_normalized": "iphone_14",
                    "match_key": "anti_spy|iphone_14",
                    "internal_nm_id": 501002,
                    "internal_name": "Anti-Spy iPhone 14",
                    "qty": 1,
                    "unit_price": 3,
                    "amount": 3,
                    "currency": "RMB",
                    "match_status": "matched",
                    "manual_override": False,
                    "raw": {"preserve": True},
                }
            ],
        )
        backfill = block.backfill_price_conformity_checks()
        if backfill.get("processed_shipments") != 1 or backfill.get("matched_count") != 1:
            raise AssertionError(f"backfill must fill only missing legacy line, got {backfill}")
        legacy = block.get_shipment("sup_legacy_price_check")
        if (
            legacy["product_lines"][0].get("price_conformity_check_mode") != "migration_backfill"
            or legacy["product_lines"][0].get("barcode") != ""
            or legacy["product_lines"][0].get("raw", {}).get("preserve") is not True
        ):
            raise AssertionError(f"backfill must preserve unrelated line fields, got {legacy['product_lines'][0]}")
        legacy_rematch = block.rematch_shipment("sup_legacy_price_check")
        if (
            legacy_rematch.get("rematch_diagnostics", {}).get("reason") != "legacy_product_barcode_missing"
            or legacy_rematch.get("product_lines", [{}])[0].get("internal_nm_id") != 501002
        ):
            raise AssertionError(f"legacy rematch must skip safely without fuzzy overwrite: {legacy_rematch}")
        second_backfill = block.backfill_price_conformity_checks()
        if second_backfill.get("processed_shipments") != 0 or second_backfill.get("updated_line_count") != 0:
            raise AssertionError(f"backfill must be idempotent, got {second_backfill}")

        supplier_parsed = block.parse_upload_supplier_safe(
            _build_price_conformity_invoice_fixture(),
            uploaded_filename="supplier-unassigned-target.xlsx",
        )
        supplier_line_fields = (
            "line_id",
            "source_row_token",
            "line_type",
            "sort_order",
            "source_no",
            "barcode",
            "model_raw",
            "qty",
            "unit_price",
            "amount",
            "currency",
            "comment",
        )
        supplier_lines = [
            {field: line.get(field) for field in supplier_line_fields}
            for line in supplier_parsed.get("lines") or []
        ]
        supplier_products = [line for line in supplier_lines if line.get("line_type") == "product"]
        supplier_products[-1]["unit_price"] = 5
        supplier_products[-1]["amount"] = 5
        supplier_payload = {
            "upload_id": supplier_parsed["upload_id"],
            "shipment_date": "2026-05-30",
            "payload": {
                "shipment_date": "2026-05-30",
                "metadata": supplier_parsed.get("metadata") or {},
                "lines": supplier_lines,
                "warnings": supplier_parsed.get("warnings") or [],
                "errors": supplier_parsed.get("errors") or [],
            },
        }
        supplier_created = block.create_shipment_supplier_safe(supplier_payload)
        supplier_shipment_id = str(supplier_created.get("shipment_id") or "")
        if (
            not supplier_shipment_id
            or supplier_created.get("target_facility_id")
            or supplier_created.get("target_facility_name")
        ):
            raise AssertionError(f"supplier create must persist an unassigned target, got {supplier_created}")
        if "target_facility_options" in block.list_shipments_supplier_safe():
            raise AssertionError("supplier-safe list must not expose facility assignment options")
        with sqlite3.connect(runtime.db_path) as conn:
            stored_target = conn.execute(
                "SELECT target_facility_id,target_facility_name FROM sheet_vitrina_v1_supplier_shipments WHERE shipment_id=?",
                (supplier_shipment_id,),
            ).fetchone()
        if stored_target != (None, None):
            raise AssertionError(f"unassigned supplier target must persist as SQL NULL, got {stored_target}")
        try:
            block.create_shipment_supplier_safe(
                {**supplier_payload, "target_facility_id": TARGET_FACILITY_ID}
            )
        except ValueError as exc:
            if "unsupported" not in str(exc):
                raise
        else:
            raise AssertionError("supplier create must reject target facility mass assignment")
        try:
            block.update_shipment(
                supplier_shipment_id,
                {"target_facility_id": INACTIVE_TARGET_FACILITY_ID},
            )
        except ValueError as exc:
            if "не существует или не active" not in str(exc):
                raise
        else:
            raise AssertionError("operator assignment must reject inactive target facilities")
        assigned = block.update_shipment(
            supplier_shipment_id,
            {"target_facility_id": TARGET_FACILITY_ID},
        )
        if (
            assigned.get("target_facility_id") != TARGET_FACILITY_ID
            or assigned.get("target_facility_name") != "FF Smoke"
        ):
            raise AssertionError(f"operator assignment readback must preserve exact id/name, got {assigned}")
        try:
            block.update_shipment_supplier_safe(
                supplier_shipment_id,
                {"target_facility_id": INACTIVE_TARGET_FACILITY_ID},
            )
        except ValueError as exc:
            if "unsupported" not in str(exc):
                raise
        else:
            raise AssertionError("supplier update must not assign or change a target facility")


def _assert_authoritative_group_rebinding_smoke() -> None:
    timestamp_counter = {"value": 0}

    def next_timestamp() -> str:
        timestamp_counter["value"] += 1
        minutes = timestamp_counter["value"]
        return f"2026-06-01T{9 + minutes // 60:02d}:{minutes % 60:02d}:00Z"

    with TemporaryDirectory(prefix="supplier-authoritative-groups-") as tmp:
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=Path(tmp) / "runtime")
        block = SupplierShipmentsBlock(runtime=runtime, timestamp_factory=next_timestamp)
        try:
            block.create_nomenclature_item(
                {
                    "is_active": True,
                    "nm_id": 600999,
                    "barcode": "6009990000000",
                    "nomenclature_name": "Invoice-local group must fail",
                    "product_type": "invoice_local_group",
                    "match_key": "invoice_local_group|iphone_14",
                }
            )
        except ValueError as exc:
            if "server-owned SKU groups" not in str(exc):
                raise
        else:
            raise AssertionError("new nomenclature rows must reference the server-owned SKU-group registry")
        owner_specs = [
            ("1111111111111", 601001, "SKU-NFC", "No Frame Clean iPhone 14 Pro", "no_frame_clean", "iphone_14_pro"),
            ("2222222222222", 601002, "SKU-NFAS", "No Frame Anti-spy iPhone 16e", "no_frame_anti_spy", "iphone_16e"),
            ("3333333333333", 601003, "SKU-NFAS-PM", "No Frame Anti-spy iPhone 14 Pro Max", "no_frame_anti_spy", "iphone_14_pro_max"),
        ]
        owners = []
        for barcode, nm_id, sku, name, group_key, model_key in owner_specs:
            owners.append(
                block.create_nomenclature_item(
                    {
                        "is_active": True,
                        "our_sku": sku,
                        "nm_id": nm_id,
                        "barcode": barcode,
                        "nomenclature_name": name,
                        "product_type": group_key,
                        "match_key": f"{group_key}|{model_key}",
                        "compatible_model_keys": [model_key],
                    }
                )["item"]
            )
        _seed_target_facilities(runtime)

        parsed = block.parse_upload(_build_invoice_fixture(), uploaded_filename="authoritative-groups.xlsx")
        products = [line for line in parsed["lines"] if line.get("line_type") == "product"]
        if [line.get("source_product_type") for line in products] != ["clear", "anti_spy", "anti_spy"]:
            raise AssertionError(f"source parser classification must remain diagnostic: {products}")
        if [line.get("product_type") for line in products] != [
            "no_frame_clean",
            "no_frame_anti_spy",
            "no_frame_anti_spy",
        ]:
            raise AssertionError(f"parse must project barcode-owner groups: {products}")
        if [line.get("group_label") for line in products] != [
            "No Frame Clean",
            "No Frame Anti-spy",
            "No Frame Anti-spy",
        ]:
            raise AssertionError(f"parse must resolve labels from SKU groups: {products}")

        barcode_tamper = json.loads(json.dumps(parsed))
        barcode_tamper["lines"][0]["barcode"] = "9999999999999"
        barcode_tamper_result = block.create_shipment(
            {
                "upload_id": parsed["upload_id"],
                "shipment_date": "2026-06-01",
                "target_facility_id": TARGET_FACILITY_ID,
                "payload": barcode_tamper,
            }
        )
        if barcode_tamper_result["product_lines"][0].get("barcode") != "1111111111111":
            raise AssertionError("client-provided barcode must be ignored in favor of source-owned invoice evidence")

        tampered = json.loads(json.dumps(parsed))
        for line in tampered["lines"]:
            if line.get("line_type") != "product":
                continue
            line.update(
                {
                    "nomenclature_item_id": "client-owner",
                    "product_type": "clear",
                    "group_key": "clear",
                    "group_label": "Client group",
                    "match_key": "client|match",
                    "internal_sku": "CLIENT-SKU",
                    "internal_nm_id": 999,
                    "internal_name": "Client name",
                    "match_status": "matched",
                    "match_evidence": {"method": "client"},
                }
            )
        created = block.create_shipment(
            {
                "upload_id": parsed["upload_id"],
                "shipment_date": "2026-06-01",
                "target_facility_id": TARGET_FACILITY_ID,
                "payload": tampered,
            }
        )
        created_products = created["product_lines"]
        if [line.get("internal_nm_id") for line in created_products] != [601001, 601002, 601003]:
            raise AssertionError(f"create must rebind all server-owned owner fields: {created_products}")
        if any(line.get("match_evidence", {}).get("method") != "barcode" for line in created_products):
            raise AssertionError("client match evidence must be ignored")

        update_payload = json.loads(json.dumps(created))
        update_payload["lines"][0].update(
            {
                "product_type": "matte",
                "group_key": "matte",
                "group_label": "Client matte",
                "match_key": "matte|client",
                "internal_nm_id": 1,
                "internal_name": "Client update",
            }
        )
        updated = block.update_shipment(created["shipment_id"], {"payload": update_payload})
        if (
            updated["product_lines"][0].get("product_type") != "no_frame_clean"
            or updated["product_lines"][0].get("internal_nm_id") != 601001
        ):
            raise AssertionError(f"update must repeat server-side barcode rebinding: {updated['product_lines'][0]}")
        certified = block.update_expenses_complete(created["shipment_id"], True)
        if not certified.get("expenses_complete"):
            raise AssertionError("test shipment must enter certified state before a source rematch")
        completed_queue_id = str(
            certified.get("warehouse_targeted_recalculation", {}).get("queue_id") or ""
        )
        if not completed_queue_id:
            raise AssertionError("certification must enqueue its exact source revision")
        with sqlite3.connect(runtime.db_path) as conn:
            conn.execute(
                """UPDATE sheet_vitrina_v1_warehouse_targeted_recalc_queue
                   SET status='complete' WHERE queue_id=?""",
                (completed_queue_id,),
            )
            conn.commit()
        metadata_update = json.loads(json.dumps(certified))
        metadata_update["metadata"]["invoice_no"] = "26GN-METADATA-REVISION"
        metadata_update["metadata"]["invoice_date"] = "2026-05-16"
        metadata_changed = block.update_shipment(
            created["shipment_id"],
            {"payload": metadata_update},
        )
        if metadata_changed.get("expenses_complete"):
            raise AssertionError("invoice identity/date source change must remove certification")
        metadata_queue = metadata_changed.get("warehouse_targeted_recalculation", {})
        if not metadata_queue.get("queue_id"):
            raise AssertionError("invoice identity/date source change must enqueue targeted replay")
        if metadata_queue.get("queue_id") == completed_queue_id or metadata_queue.get("status") != "queued":
            raise AssertionError(
                "invoice identity/date source change must create a new queued revision after a completed one"
            )
        if metadata_queue.get("effective_date") != "2026-05-16":
            raise AssertionError(f"metadata replay must retain the edited invoice date: {metadata_queue}")
        certified = block.update_expenses_complete(created["shipment_id"], True)

        checksum_edit = json.loads(json.dumps(certified))
        checksum_edit["metadata"]["declared_invoice_total"] = (
            float(checksum_edit["summary"]["invoice_amount_total"]) + 1
        )
        checksum_changed = block.update_shipment(
            created["shipment_id"],
            {"payload": checksum_edit},
        )
        if (
            checksum_changed.get("expenses_complete")
            or checksum_changed.get("match_status") != "checksum_error"
            or checksum_changed.get("exact_cost_status") != "unavailable"
            or "контрольной суммы" not in " ".join(checksum_changed.get("exact_cost_blockers") or [])
            or not checksum_changed.get("warehouse_targeted_recalculation", {}).get("queue_id")
        ):
            raise AssertionError(
                "declared invoice checksum mutation must invalidate certification and queue replay: "
                f"{checksum_changed}"
            )
        checksum_fixed = json.loads(json.dumps(checksum_changed))
        checksum_fixed["metadata"]["declared_invoice_total"] = checksum_changed["summary"][
            "invoice_amount_total"
        ]
        block.update_shipment(created["shipment_id"], {"payload": checksum_fixed})
        certified = block.update_expenses_complete(created["shipment_id"], True)

        first_owner = owners[0]
        block.update_nomenclature_item(
            str(first_owner["item_id"]),
            {
                **first_owner,
                "product_type": "no_frame_matte",
                "match_key": "no_frame_matte|iphone_14_pro",
            },
        )
        rematched = block.rematch_shipment(created["shipment_id"])
        first_line = rematched["product_lines"][0]
        if first_line.get("product_type") != "no_frame_matte" or first_line.get("match_key") != "no_frame_matte|iphone_14_pro":
            raise AssertionError(f"explicit rematch must rebind group and match key from the current barcode owner: {first_line}")
        if rematched.get("expenses_complete"):
            raise AssertionError("a changed source rematch must return the shipment to provisional state")
        if not rematched.get("warehouse_targeted_recalculation", {}).get("queue_id"):
            raise AssertionError("a changed source rematch must enqueue bounded warehouse recalculation")

        matte_group = next(group for group in block.list_sku_groups()["groups"] if group["group_key"] == "no_frame_matte")
        block.update_sku_group("no_frame_matte", {**matte_group, "label": "Runtime-renamed group"})
        reopened = block.get_shipment(created["shipment_id"])
        if reopened["product_lines"][0].get("group_label") != "Runtime-renamed group":
            raise AssertionError("saved invoice labels must resolve live through the same SKU-group registry")

        frozen = runtime.load_supplier_shipment(created["shipment_id"])
        frozen_header = dict(frozen["header"])
        frozen_header["actual_ff_acceptance_date"] = "2026-06-30"
        runtime.save_supplier_shipment(header=frozen_header, lines=frozen["lines"])
        matte_group = next(
            group
            for group in block.list_sku_groups()["groups"]
            if group["group_key"] == "no_frame_matte"
        )
        block.update_sku_group(
            "no_frame_matte",
            {**matte_group, "label": "Frozen presentation rename"},
        )
        presentation_rematched = block.rematch_shipment(created["shipment_id"])
        if presentation_rematched.get("warehouse_targeted_recalculation"):
            raise AssertionError(
                "presentation-only rematch must not enqueue a capital recalculation"
            )
        if (
            presentation_rematched["product_lines"][0].get("group_label")
            != "Frozen presentation rename"
        ):
            raise AssertionError(
                "presentation-only rematch must refresh current labels after FF receipt"
            )
        frozen_after_presentation = runtime.load_supplier_shipment(created["shipment_id"])
        if _canonical_supplier_cost_signature(frozen_after_presentation) != _canonical_supplier_cost_signature(frozen):
            raise AssertionError(
                "presentation-only rematch must preserve canonical invoice cost inputs"
            )
        current_owner = next(
            item
            for item in block.list_nomenclature()["items"]
            if str(item.get("item_id") or "") == str(first_owner["item_id"])
        )
        block.update_nomenclature_item(
            str(current_owner["item_id"]),
            {**current_owner, "match_key": "no_frame_matte|frozen-rematch"},
        )
        try:
            block.rematch_shipment(created["shipment_id"])
        except ValueError as exc:
            if "append-only ledger" not in str(exc):
                raise
        else:
            raise AssertionError("post-FF-receipt identity rematch must fail closed")
        frozen_readback = runtime.load_supplier_shipment(created["shipment_id"])
        if frozen_readback["lines"] != frozen_after_presentation["lines"]:
            raise AssertionError("rejected post-receipt rematch must preserve frozen invoice identity")


def _canonical_supplier_cost_signature(payload: dict[str, object]) -> tuple[object, ...]:
    header = dict(payload.get("header") or {})
    return (
        header.get("invoice_no"),
        header.get("invoice_date"),
        header.get("currency"),
        header.get("declared_invoice_total"),
        header.get("invoice_amount_total"),
        header.get("match_status"),
        tuple(
            (
                item.get("line_id"),
                item.get("line_type"),
                item.get("internal_nm_id"),
                item.get("product_type"),
                item.get("match_key"),
                item.get("qty"),
                item.get("unit_price"),
                item.get("amount"),
            )
            for item in payload.get("lines") or []
            if isinstance(item, dict)
        ),
    )


def _assert_blocked_canonical_cost_suppresses_legacy_aggregate() -> None:
    with TemporaryDirectory(prefix="supplier-canonical-cost-guard-") as tmp:
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=Path(tmp) / "runtime")
        block = SupplierShipmentsBlock(runtime=runtime)
        payload = {
            "shipment_id": "blocked-canonical-cost",
            "exact_bank_fees_rub": 12.0,
            "exact_currency_payment_cost_rub": 100.0,
            "exact_landed_cost_total_rub": 112.0,
            "exact_landed_cost_per_unit_rub": 11.2,
            "exact_cost_status": "complete",
            "lines": [{"line_id": "line-1", "line_type": "product"}],
        }
        blocked = {
            "shipment_id": payload["shipment_id"],
            "lines": [{"line_id": "line-1", "unit_cost_rub": None}],
            "blockers": [
                {
                    "code": "supplier_payment_rub_valuation_unavailable",
                    "reason_ru": "Фактическая RUB-стоимость платежа отсутствует.",
                }
            ],
        }
        with patch(
            "packages.application.warehouse_functional.load_supplier_line_cost_breakdown",
            return_value=blocked,
        ):
            result = block._with_supplier_line_costs(payload)
        if (
            result.get("exact_landed_cost_total_rub") is not None
            or result.get("exact_landed_cost_per_unit_rub") is not None
            or result.get("exact_cost_status") != "unavailable"
            or "RUB-стоимость" not in " ".join(result.get("exact_cost_blockers") or [])
        ):
            raise AssertionError(
                "blocked canonical proof must suppress every legacy exact aggregate: "
                f"{result}"
            )
        registry_row = {
            **payload,
            "expenses_complete": True,
            "invoice_no": "BLOCKED-LIST",
        }
        with (
            patch.object(block, "migrate_existing_supplier_shipments_into_trade_documents"),
            patch.object(runtime, "list_supplier_shipments", return_value=[registry_row]),
            patch.object(block, "_with_document_fields", side_effect=lambda value: value),
            patch.object(block, "_with_approx_cost_fields", side_effect=lambda value: value),
            patch(
                "packages.application.warehouse_functional.load_supplier_line_cost_breakdown",
                side_effect=AssertionError("collection must not load a full proof per shipment"),
            ),
            patch(
                "packages.application.warehouse_functional.load_supplier_cost_summary_fields",
                return_value={
                    payload["shipment_id"]: {
                        "exact_bank_fees_rub": None,
                        "exact_currency_payment_cost_rub": None,
                        "exact_landed_cost_total_rub": None,
                        "exact_landed_cost_per_unit_rub": None,
                        "exact_cost_status": "unavailable",
                        "exact_cost_blockers": [
                            "Фактическая RUB-стоимость платежа отсутствует."
                        ],
                        "exact_cost_warnings": [],
                        "expense_allocation": {
                            "status": "none",
                            "label": "Расходы не распределены",
                            "diagnostics": {
                                "eligible_documents": 1,
                                "eligible_components": 1,
                                "allocated_components": 0,
                            },
                            "reasons": ["Фактическая RUB-стоимость платежа отсутствует."],
                        },
                    }
                },
            ),
        ):
            list_row = block.list_shipments()["shipments"][0]
        if (
            list_row.get("exact_landed_cost_per_unit_rub") is not None
            or list_row.get("exact_cost_status") != "unavailable"
            or (list_row.get("expense_allocation") or {}).get("status") != "none"
            or "RUB-стоимость" not in " ".join(list_row.get("exact_cost_blockers") or [])
        ):
            raise AssertionError(
                "collection registry response must apply the same canonical blockers as detail: "
                f"{list_row}"
            )

        provisional = {
            "shipment_id": payload["shipment_id"],
            "capital_rub": "112",
            "average_unit_cost_rub": "11.2",
            "lines": [{"line_id": "line-1", "unit_cost_rub": "11.2"}],
            "blockers": [],
            "component_controls": [],
            "certification": {
                "certified": False,
                "source_fingerprint_matches": False,
                "status_label_ru": "Предварительная себестоимость — источники изменились",
            },
        }
        with patch(
            "packages.application.warehouse_functional.load_supplier_line_cost_breakdown",
            return_value=provisional,
        ):
            result = block._with_supplier_line_costs(registry_row)
        if result.get("exact_cost_status") != "provisional":
            raise AssertionError(
                "expenses_complete alone must not upgrade a fingerprint-mismatched proof: "
                f"{result}"
            )


def main() -> None:
    _assert_barcode_schema_upgrade_smoke()
    _assert_price_conformity_application_smoke()
    _assert_authoritative_group_rebinding_smoke()
    _assert_blocked_canonical_cost_suppresses_legacy_aggregate()
    workbook_bytes = _build_invoice_fixture()
    workbook_sha256 = hashlib.sha256(workbook_bytes).hexdigest()
    with TemporaryDirectory(prefix="supplier-shipments-http-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        config = RegistryUploadHttpEntrypointConfig(
            host="127.0.0.1",
            port=_reserve_free_port(),
            upload_path=DEFAULT_UPLOAD_PATH,
            sheet_plan_path=DEFAULT_SHEET_PLAN_PATH,
            sheet_refresh_path="/v1/sheet-vitrina-v1/refresh",
            sheet_status_path=DEFAULT_SHEET_STATUS_PATH,
            sheet_operator_ui_path=DEFAULT_SHEET_OPERATOR_UI_PATH,
            runtime_dir=runtime_dir,
        )
        entrypoint = RegistryUploadHttpEntrypoint(
            runtime_dir=runtime_dir,
            runtime=runtime,
            activated_at_factory=lambda: "2026-05-30T08:00:00Z",
        )
        _seed_target_facilities(runtime)
        server = build_registry_upload_http_server(config, entrypoint=entrypoint)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            base_url = f"http://127.0.0.1:{config.port}"
            list_status, list_payload = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}")
            if list_status != 200 or list_payload.get("shipments") != []:
                raise AssertionError(f"empty registry must load, got {list_status} {list_payload}")

            unsupported_status, unsupported_payload = _post_multipart(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PARSE_PATH}",
                b"not an invoice",
                filename="invoice.txt",
                content_type="text/plain",
            )
            if unsupported_status != 400 or "xlsx" not in str(unsupported_payload.get("error", "")).lower():
                raise AssertionError(f"unsupported file type must return controlled JSON 400, got {unsupported_status} {unsupported_payload}")

            nomenclature_status, nomenclature_payload = _get_json(f"{base_url}{DEFAULT_NOMENCLATURE_PATH}")
            if nomenclature_status != 200 or nomenclature_payload.get("items") != []:
                raise AssertionError(f"empty nomenclature must load, got {nomenclature_status} {nomenclature_payload}")
            create_nom_status, create_nom_payload = _post_json(
                f"{base_url}{DEFAULT_NOMENCLATURE_PATH}",
                {
                    "is_active": True,
                    "our_sku": "SKU-CLEAR-14P",
                    "nm_id": 210183919,
                    "barcode": "1111111111111",
                    "nomenclature_name": "Clear iPhone 14 Pro",
                    "product_type": "clean",
                    "match_key": "clean|iphone_14_pro",
                    "purchase_price_yuan": "1,0",
                    "aliases": ["iPhone 14 Pro"],
                    "compatible_models_text": "iPhone 14 Pro",
                    "comment": "smoke",
                },
            )
            if create_nom_status != 200 or create_nom_payload.get("item", {}).get("nm_id") != 210183919:
                raise AssertionError(f"nomenclature create must persist item, got {create_nom_status} {create_nom_payload}")
            if create_nom_payload.get("item", {}).get("purchase_price_yuan") != 1.0:
                raise AssertionError("nomenclature create must normalize purchase_price_yuan decimal comma")
            if create_nom_payload.get("item", {}).get("compatible_model_keys") != ["iphone_14_pro"]:
                raise AssertionError("nomenclature create must normalize compatible model keys")
            duplicate_nom_status, duplicate_nom_payload = _post_json(
                f"{base_url}{DEFAULT_NOMENCLATURE_PATH}",
                {
                    "is_active": True,
                    "nomenclature_name": "Duplicate",
                    "product_type": "clean",
                    "match_key": "clean|iphone_14_pro",
                },
            )
            if duplicate_nom_status != 400 or "duplicate" not in str(duplicate_nom_payload.get("error", "")).lower():
                raise AssertionError(f"duplicate active match_key must be rejected, got {duplicate_nom_status} {duplicate_nom_payload}")
            compat_nom_status, compat_nom_payload = _post_json(
                f"{base_url}{DEFAULT_NOMENCLATURE_PATH}",
                {
                    "is_active": True,
                    "our_sku": "SKU-AS-141313P",
                    "nm_id": 391662410,
                    "barcode": "2222222222222",
                    "nomenclature_name": "anti-spy iPhone 14 / 13 / 13Pro",
                    "product_type": "anti_spy",
                    "match_key": "anti_spy|iphone_14_13_13pro",
                    "purchase_price_yuan": "3",
                    "compatible_models_text": "iPhone 14, iPhone 13, iPhone 13 Pro",
                    "comment": "compatibility smoke",
                },
            )
            if compat_nom_status != 200 or compat_nom_payload.get("item", {}).get("compatible_model_keys") != [
                "iphone_14",
                "iphone_13",
                "iphone_13_pro",
            ]:
                raise AssertionError(f"compatible nomenclature item must save normalized keys, got {compat_nom_status} {compat_nom_payload}")
            third_nom_status, third_nom_payload = _post_json(
                f"{base_url}{DEFAULT_NOMENCLATURE_PATH}",
                {
                    "is_active": True,
                    "our_sku": "SKU-AS-14PM",
                    "nm_id": 210184534,
                    "barcode": "3333333333333",
                    "nomenclature_name": "Anti-Spy iPhone 14 Pro Max",
                    "product_type": "anti_spy",
                    "match_key": "anti_spy|iphone_14_pro_max",
                    "purchase_price_yuan": "2",
                    "comment": "atomic matching smoke",
                },
            )
            if third_nom_status != 200 or third_nom_payload.get("item", {}).get("nm_id") != 210184534:
                raise AssertionError("third nomenclature item must exist before atomic invoice acceptance")
            nomenclature_import_bytes = _build_nomenclature_import_fixture(
                first_item_id=str(create_nom_payload["item"]["item_id"]),
                compat_item_id=str(compat_nom_payload["item"]["item_id"]),
            )
            export_status, export_bytes, export_headers = _get_bytes(f"{base_url}{DEFAULT_NOMENCLATURE_EXPORT_PATH}")
            if export_status != 200 or "spreadsheetml.sheet" not in str(export_headers.get("Content-Type", "")):
                raise AssertionError(f"nomenclature export must return XLSX, got {export_status} {export_headers}")
            exported = load_workbook(BytesIO(export_bytes), data_only=True)
            exported_headers = [cell.value for cell in exported.active[1]]
            expected_headers = [
                "ID строки",
                "Включено",
                "Скрыто",
                "nmId",
                "ШК / barcode",
                "Все ШК",
                "Источник ШК",
                "Статус ШК",
                "Артикул продавца WB / vendorCode",
                "Название WB",
                "WB subject",
                "WB updatedAt",
                "Статус WB sync",
                "Номенклатура",
                "Группа",
                "Match key",
                "Цена закупки, ¥",
                "Заводской короб, шт.",
                "Совместимые модели",
                "Ключи совместимости",
                "Обновлено",
            ]
            if exported_headers != expected_headers:
                raise AssertionError(f"nomenclature export headers changed unexpectedly: {exported_headers}")
            if {"Наш SKU", "Aliases", "Комментарий"} & set(exported_headers):
                raise AssertionError("nomenclature export must not expose hidden legacy fields by default")
            dry_run_status, dry_run_payload = _post_multipart(
                f"{base_url}{DEFAULT_NOMENCLATURE_IMPORT_PATH}?dry_run=1",
                nomenclature_import_bytes,
                filename="nomenclature.xlsx",
            )
            if (
                dry_run_status != 200
                or dry_run_payload.get("dry_run") is not True
                or dry_run_payload.get("created_count") != 1
                or dry_run_payload.get("updated_count") != 1
                or dry_run_payload.get("deactivated_count") != 1
            ):
                raise AssertionError(f"nomenclature import dry-run must validate counts, got {dry_run_status} {dry_run_payload}")
            after_dry_run_status, after_dry_run_payload = _get_json(f"{base_url}{DEFAULT_NOMENCLATURE_PATH}")
            after_dry_run_items = {item["item_id"]: item for item in after_dry_run_payload.get("items", [])}
            if (
                after_dry_run_status != 200
                or after_dry_run_items[str(create_nom_payload["item"]["item_id"])].get("purchase_price_yuan") != 1.0
                or after_dry_run_items[str(compat_nom_payload["item"]["item_id"])].get("is_active") is not True
            ):
                raise AssertionError("nomenclature import dry-run must not mutate runtime DB")

            parse_status, parse_payload = _post_multipart(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PARSE_PATH}",
                workbook_bytes,
                filename="PI-test 26GN390 (14.5.2026).xlsx",
            )
            if parse_status != 200 or not parse_payload.get("upload_id"):
                raise AssertionError(f"parse route must stage upload and return editable payload, got {parse_status} {parse_payload}")
            if parse_payload.get("source_file_sha256") != workbook_sha256:
                raise AssertionError("parse route must expose sha256 of original upload")
            if parse_payload.get("metadata", {}).get("supplier_name") != "HanShang Technology":
                raise AssertionError("parse route must default supplier_name to HanShang Technology")
            if parse_payload.get("metadata", {}).get("customer_name") not in {"", None}:
                raise AssertionError("parse route must not require or persist parsed customer_name")
            if (
                parse_payload.get("metadata", {}).get("contract_no") != "CNT-2026-0513"
                or parse_payload.get("metadata", {}).get("contract_date") != "2026-05-13"
            ):
                raise AssertionError(f"parse route must expose contract no/date, got {parse_payload.get('metadata')}")
            product_lines = [item for item in parse_payload.get("lines", []) if item.get("line_type") == "product"]
            if product_lines[0].get("internal_nm_id") != 210183919 or product_lines[0].get("match_status") != "matched_by_barcode":
                raise AssertionError("parse route must resolve active primary barcode into nmId/name")
            if (
                product_lines[1].get("match_status") != "matched_by_barcode"
                or product_lines[1].get("internal_nm_id") != 391662410
            ):
                raise AssertionError(f"parse route must resolve a barcode from all known barcodes, got {product_lines[1]}")
            if product_lines[2].get("match_status") != "matched_by_barcode" or product_lines[2].get("internal_nm_id") != 210184534:
                raise AssertionError("all product lines must match exact active barcodes before acceptance")
            price_statuses = [line.get("price_conformity_status") for line in product_lines]
            if price_statuses != ["matched", "mismatched", "matched"]:
                raise AssertionError(f"parse route must attach price conformity statuses, got {price_statuses}")
            if (
                product_lines[0].get("invoice_price_yuan_snapshot") != 1.0
                or product_lines[0].get("reference_purchase_price_yuan_snapshot") != 1.0
                or product_lines[0].get("price_conformity_check_mode") != "initial_parse"
            ):
                raise AssertionError(f"parse route must expose price snapshots/mode, got {product_lines[0]}")

            missing_date_status, missing_date_payload = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}",
                {"upload_id": parse_payload["upload_id"], "payload": parse_payload},
            )
            if missing_date_status != 400 or "shipment_date" not in str(missing_date_payload.get("error", "")):
                raise AssertionError("create must reject missing shipment_date")

            invalid_actual_status, invalid_actual_payload = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}",
                {
                    "upload_id": parse_payload["upload_id"],
                    "shipment_date": "2026-05-14",
                    "actual_shipment_date": "2026/05/16",
                    "payload": parse_payload,
                },
            )
            if invalid_actual_status != 400 or "actual_shipment_date" not in str(invalid_actual_payload.get("error", "")):
                raise AssertionError(f"create must reject invalid actual_shipment_date, got {invalid_actual_status} {invalid_actual_payload}")

            invalid_rate_status, invalid_rate_payload = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}",
                {
                    "upload_id": parse_payload["upload_id"],
                    "shipment_date": "2026-05-14",
                    "approx_yuan_rate": "-1",
                    "payload": parse_payload,
                },
            )
            if invalid_rate_status != 400 or "approx_yuan_rate" not in str(invalid_rate_payload.get("error", "")):
                raise AssertionError(f"create must reject non-positive approx_yuan_rate, got {invalid_rate_status} {invalid_rate_payload}")

            parse_payload["lines"][0]["barcode"] = "9999999999999"
            parse_payload["lines"][0]["internal_nm_id"] = 999999
            parse_payload["lines"][0]["internal_name"] = "Client supplied identity"
            parse_payload["lines"][0]["match_status"] = "matched"
            parse_payload["lines"][0]["manual_override"] = True
            create_status, detail = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}",
                {
                    "upload_id": parse_payload["upload_id"],
                    "shipment_date": "2026-05-14",
                    "actual_shipment_date": "2026-05-16",
                    "target_facility_id": TARGET_FACILITY_ID,
                    "approx_yuan_rate": "13,2",
                    "payload": parse_payload,
                },
            )
            if create_status != 200 or not detail.get("shipment_id"):
                raise AssertionError(f"create route must persist shipment, got {create_status} {detail}")
            shipment_id = detail["shipment_id"]
            if detail.get("shipment_date") != "2026-05-14" or detail.get("match_status") != "all_matched":
                raise AssertionError("created shipment must keep date and fully matched status")
            if (
                detail.get("planned_shipment_date") != "2026-05-14"
                or detail.get("actual_shipment_date") != "2026-05-16"
                or detail.get("actual_ff_acceptance_date") not in {"", None}
            ):
                raise AssertionError(f"created shipment must expose planned/fact dates, got {detail}")
            if detail.get("order_status") != "in_transit" or detail.get("order_status_display") != "В пути с 16.05.2026":
                raise AssertionError(f"created shipment status must derive from actual date, got {detail.get('order_status_display')}")
            if detail.get("approx_yuan_rate") != 13.2 or detail.get("approx_invoice_cost_rub") != 435.6:
                raise AssertionError(f"created shipment must expose approx yuan rate and invoice RUB cost, got {detail}")
            if detail.get("approx_landed_cost_per_unit_rub") is not None:
                raise AssertionError(f"created shipment without factual expenses must not expose approximate landed cost, got {detail}")
            if detail.get("supplier_name") != "HanShang Technology" or detail.get("metadata", {}).get("supplier_name") != "HanShang Technology":
                raise AssertionError("created shipment must persist fixed supplier_name")
            if detail.get("customer_name") not in {"", None} or detail.get("metadata", {}).get("customer_name") not in {"", None}:
                raise AssertionError("created shipment must not require customer_name")
            if detail.get("contract_no") != "CNT-2026-0513" or detail.get("contract_date") != "2026-05-13":
                raise AssertionError("created shipment must persist contract no/date")
            if len(detail.get("product_lines", [])) != 3 or len(detail.get("extra_lines", [])) != 1:
                raise AssertionError("detail must split product and extra lines")
            if detail["product_lines"][0].get("internal_name") != "Clear iPhone 14 Pro":
                raise AssertionError("created shipment must persist nomenclature auto-match")
            if detail["product_lines"][0].get("barcode") != "1111111111111":
                raise AssertionError("shipment detail must persist and return source-owned barcode")
            if detail["product_lines"][0].get("price_conformity_status") != "matched":
                raise AssertionError("created shipment must persist price conformity status")

            _seed_supplier_factual_expense(runtime, shipment_id, amount_rub=48.0)

            detail_status, loaded_detail = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}")
            if detail_status != 200 or loaded_detail.get("shipment_id") != shipment_id:
                raise AssertionError("detail route must return persisted card payload")
            if (
                loaded_detail.get("approx_yuan_rate") != 13.2
                or loaded_detail.get("approx_invoice_cost_rub") != 435.6
                or loaded_detail.get("approx_landed_cost_per_unit_rub") != 25.45
                or (loaded_detail.get("expense_allocation") or {}).get("status") != "partial"
            ):
                raise AssertionError(f"detail route must expose approximate landed cost with factual expenses, got {loaded_detail}")
            post_expense_list_status, post_expense_list = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}")
            listed_created = _shipment_by_id(post_expense_list, shipment_id)
            if (
                post_expense_list_status != 200
                or listed_created.get("approx_yuan_rate") != 13.2
                or listed_created.get("approx_invoice_cost_rub") != 435.6
                or listed_created.get("approx_landed_cost_per_unit_rub") != 25.45
                or (listed_created.get("expense_allocation") or {}).get("status") != "partial"
            ):
                raise AssertionError(f"list route must expose approximate landed cost with factual expenses, got {post_expense_list_status} {post_expense_list}")
            if loaded_detail.get("order_status") != "in_transit":
                raise AssertionError("detail route must expose derived in-transit status")
            if (
                loaded_detail.get("planned_shipment_date") != "2026-05-14"
                or loaded_detail.get("actual_shipment_date") != "2026-05-16"
                or loaded_detail.get("actual_ff_acceptance_date") not in {"", None}
            ):
                raise AssertionError("detail route must expose planned/fact shipment dates")
            if loaded_detail.get("product_lines", [{}])[0].get("price_conformity_checked_at") != "2026-05-30T08:00:00Z":
                raise AssertionError("detail route must expose persisted price conformity metadata without recalculation")
            future_shipment_status, future_shipment_payload = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/preview",
                {"actual_shipment_date": "2026-05-31"},
            )
            if future_shipment_status != 400 or "business today" not in str(future_shipment_payload.get("error", "")):
                raise AssertionError(f"future shipment date must be rejected by API: {future_shipment_status} {future_shipment_payload}")
            future_acceptance_status, future_acceptance_payload = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/preview",
                {"actual_ff_acceptance_date": "2026-05-31"},
            )
            if future_acceptance_status != 400 or "business today" not in str(future_acceptance_payload.get("error", "")):
                raise AssertionError(f"future FF acceptance must be rejected by API: {future_acceptance_status} {future_acceptance_payload}")
            early_acceptance_status, early_acceptance_payload = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/preview",
                {"actual_ff_acceptance_date": "2026-05-15"},
            )
            if early_acceptance_status != 400 or "раньше" not in str(early_acceptance_payload.get("error", "")):
                raise AssertionError(f"acceptance before shipment must be rejected: {early_acceptance_status} {early_acceptance_payload}")
            unchanged_status, unchanged_detail = _get_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}"
            )
            if (
                unchanged_status != 200
                or unchanged_detail.get("actual_shipment_date") != "2026-05-16"
                or unchanged_detail.get("actual_ff_acceptance_date") not in {"", None}
            ):
                raise AssertionError("rejected factual dates must leave source unchanged")
            price_check_status, price_checked = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/price-check",
                {"context": {"source": "http_smoke"}},
            )
            first_price_checked_line = price_checked.get("product_lines", [{}])[0]
            if (
                price_check_status != 200
                or first_price_checked_line.get("price_conformity_check_mode") != "manual_recheck"
                or not str(first_price_checked_line.get("price_conformity_actor") or "").startswith("webcore_user_")
                or first_price_checked_line.get("price_conformity_context", {}).get("source") != "http_smoke"
            ):
                raise AssertionError(f"manual price-check route must persist actor/context/mode, got {price_check_status} {price_checked}")
            if price_checked.get("approx_yuan_rate") != 13.2 or price_checked.get("approx_landed_cost_per_unit_rub") != 25.45:
                raise AssertionError("manual price-check must not erase approximate cost fields")
            loaded_detail = price_checked

            edited = json.loads(json.dumps(loaded_detail, ensure_ascii=False))
            edited["lines"][0]["internal_sku"] = "SKU-MANUAL"
            edited["lines"][0]["internal_nm_id"] = 123456
            edited["lines"][0]["internal_name"] = "Manual SKU"
            edited["lines"][0]["match_status"] = "matched"
            edited["lines"][0]["manual_override"] = True
            edited["lines"][0]["barcode"] = "9999999999999"
            edited["lines"][0]["amount"] = 12
            edited["metadata"]["declared_invoice_total"] = 35
            patch_status, patched = _patch_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}",
                {
                    "shipment_date": "2026-05-15",
                    "actual_shipment_date": "2026-05-16",
                    "approx_yuan_rate": "14.5",
                    "payload": edited,
                },
            )
            if patch_status != 200 or patched.get("shipment_date") != "2026-05-15":
                raise AssertionError(f"patch route must update shipment date, got {patch_status} {patched}")
            if (
                patched.get("planned_shipment_date") != "2026-05-15"
                or patched.get("actual_shipment_date") != "2026-05-16"
                or patched.get("actual_ff_acceptance_date") not in {"", None}
            ):
                raise AssertionError(f"patch route must update fact dates, got {patched}")
            if patched.get("match_status") != "all_matched" or patched.get("summary", {}).get("product_amount_total") != 30.0:
                raise AssertionError("patch route must reapply barcode identity and recalculate totals server-side")
            first_patched_line = patched.get("product_lines", [{}])[0]
            if (
                first_patched_line.get("barcode") != "1111111111111"
                or first_patched_line.get("internal_nm_id") != 210183919
                or first_patched_line.get("internal_name") != "Clear iPhone 14 Pro"
                or first_patched_line.get("match_status") != "matched_by_barcode"
                or first_patched_line.get("manual_override") is not False
            ):
                raise AssertionError(f"manual override must not bypass source barcode identity: {first_patched_line}")
            if patched.get("order_status") != "in_transit":
                raise AssertionError("full patch must keep status derived from factual date")
            if patched.get("approx_yuan_rate") != 14.5 or patched.get("approx_invoice_cost_rub") != 507.5:
                raise AssertionError(f"patch route must update approx_yuan_rate and derived invoice cost, got {patched}")
            if patched.get("approx_landed_cost_per_unit_rub") != 29.24:
                raise AssertionError(f"patch route must recalculate approximate landed cost, got {patched}")

            correction_started = threading.Event()
            correction_release = threading.Event()

            def hold_correction(phase: str) -> None:
                if phase == "before_transaction":
                    correction_started.set()
                    if not correction_release.wait(timeout=5):
                        raise RuntimeError("correction smoke hold timeout")

            entrypoint.supplier_shipment_factual_correction_block.failure_injector = hold_correction
            correction_preview_status, correction_preview = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/preview",
                {"actual_shipment_date": "2026-05-17"},
            )
            if correction_preview_status != 200:
                raise AssertionError(f"date correction preview failed: {correction_preview_status} {correction_preview}")
            correction_status, correction_accepted = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/confirm",
                {"confirmation_token": correction_preview["confirmation_token"]},
            )
            if correction_status != 202 or correction_accepted.get("status") != "accepted":
                raise AssertionError(f"date correction must start one persisted job, got {correction_status} {correction_accepted}")
            if not correction_started.wait(timeout=5):
                raise AssertionError("correction job did not reach a real running phase")
            running_detail_status, running_detail = _get_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}"
            )
            running_correction = running_detail.get("factual_date_correction") or {}
            if running_detail_status != 200 or not running_correction.get("active"):
                raise AssertionError(f"detail reload must expose active persisted correction: {running_detail}")
            duplicate_preview_status, duplicate_preview = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/preview",
                {"actual_shipment_date": "2026-05-17"},
            )
            if duplicate_preview_status != 200:
                raise AssertionError(f"duplicate date preview failed: {duplicate_preview_status} {duplicate_preview}")
            duplicate_status, duplicate_payload = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/confirm",
                {"confirmation_token": duplicate_preview["confirmation_token"]},
            )
            if (
                duplicate_status != 202
                or not (duplicate_payload.get("correction") or {}).get("deduplicated")
                or (duplicate_payload.get("correction") or {}).get("correction_id")
                != (correction_accepted.get("correction") or {}).get("correction_id")
            ):
                raise AssertionError(f"parallel duplicate save must reuse one job: {duplicate_status} {duplicate_payload}")
            correction_release.set()
            entrypoint.supplier_shipment_factual_correction_block.failure_injector = None
            correction = _wait_for_factual_correction(base_url, shipment_id)
            if correction.get("status") != "success":
                raise AssertionError(f"date correction job did not succeed: {correction}")
            detail_status, patched = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}")
            if (
                detail_status != 200
                or patched.get("actual_shipment_date") != "2026-05-17"
                or patched.get("order_status") != "in_transit"
                or patched.get("order_status_display") != "В пути с 17.05.2026"
            ):
                raise AssertionError(f"date correction readback mismatch: {patched}")
            repeated_status, repeated = _patch_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}",
                {"actual_shipment_date": "2026-05-17"},
            )
            if repeated_status != 200 or repeated.get("actual_shipment_date") != "2026-05-17":
                raise AssertionError(f"repeated correction must be zero-change: {repeated_status} {repeated}")

            status_patch_status, status_patched = _patch_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}",
                {"order_status": "in_transit"},
            )
            if status_patch_status != 400 or "status-only PATCH" not in str(status_patched.get("error", "")):
                raise AssertionError(f"matching status-only patch must be rejected, got {status_patch_status} {status_patched}")
            after_status_reject_status, after_status_reject = _get_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}"
            )
            if (
                after_status_reject_status != 200
                or len(after_status_reject.get("product_lines", [])) != 3
                or after_status_reject.get("source_file_sha256") != workbook_sha256
                or after_status_reject.get("invoice_no") != "26GN390"
                or after_status_reject.get("actual_shipment_date") != "2026-05-17"
                or after_status_reject.get("actual_ff_acceptance_date") not in {"", None}
                or after_status_reject.get("approx_yuan_rate") != 14.5
                or after_status_reject.get("approx_landed_cost_per_unit_rub") != 29.24
            ):
                raise AssertionError("rejected status-only patch must not erase lines, metadata, source file, fact dates, or approx yuan rate")
            invalid_status, invalid_payload = _patch_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}",
                {"order_status": "delivered_to_mars"},
            )
            if invalid_status != 400 or "order_status" not in str(invalid_payload.get("error", "")):
                raise AssertionError(f"invalid order_status must be rejected, got {invalid_status} {invalid_payload}")
            divergent_status, divergent_payload = _patch_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}",
                {"order_status": "production"},
            )
            if divergent_status != 400 or "вычисляется" not in str(divergent_payload.get("error", "")):
                raise AssertionError(f"manual divergent status must be rejected, got {divergent_status} {divergent_payload}")
            accepted_preview_status, accepted_preview = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/preview",
                {"actual_ff_acceptance_date": "2026-05-30"},
            )
            if accepted_preview_status != 200:
                raise AssertionError(f"actual FF acceptance preview failed: {accepted_preview_status} {accepted_preview}")
            accepted_status, accepted_patched = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-dates/confirm",
                {"confirmation_token": accepted_preview["confirmation_token"]},
            )
            if accepted_status != 200 or accepted_patched.get("order_status") != "accepted_ff":
                raise AssertionError(f"actual FF acceptance patch must persist accepted_ff, got {accepted_status} {accepted_patched}")
            if accepted_patched.get("actual_ff_acceptance_date") != "2026-05-30":
                raise AssertionError(f"actual FF acceptance patch must keep acceptance date, got {accepted_patched}")
            ff_stock_keys = [str(item.get("source_key") or "") for item in runtime.list_ff_stock_operations()]
            if ff_stock_keys.count(f"supplier_shipment_acceptance:{shipment_id}") != 1:
                raise AssertionError(f"actual FF acceptance must create one idempotent ФФ stock operation, got {ff_stock_keys}")

            frozen_before_update = runtime.load_supplier_shipment(shipment_id)
            changed_after_receipt = json.loads(json.dumps(accepted_patched))
            changed_after_receipt["lines"][0]["qty"] += 1
            changed_after_receipt["lines"][0]["amount"] += changed_after_receipt["lines"][0]["unit_price"]
            try:
                entrypoint.supplier_shipments_block.update_shipment(
                    shipment_id,
                    {"payload": changed_after_receipt},
                )
            except ValueError as exc:
                if "append-only ledger" not in str(exc):
                    raise
            else:
                raise AssertionError("post-FF invoice quantity mutation must fail before persistence")
            frozen_after_update = runtime.load_supplier_shipment(shipment_id)
            if frozen_after_update != frozen_before_update:
                raise AssertionError("rejected post-FF invoice mutation must preserve shipment and lines atomically")

            rematch_status, rematched = _post_json(
                f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/rematch",
                {"overwrite_manual": False},
            )
            if rematch_status != 200:
                raise AssertionError(f"rematch route must return updated detail, got {rematch_status} {rematched}")
            rematched_products = rematched.get("product_lines", [])
            if rematched_products[0].get("internal_sku") != "SKU-CLEAR-14P":
                raise AssertionError("rematch must resolve source barcode and ignore legacy manual override flags")
            if rematched_products[2].get("internal_nm_id") != 210184534:
                raise AssertionError("rematch must preserve deterministic authoritative match")
            if rematched.get("approx_yuan_rate") != 14.5 or rematched.get("approx_landed_cost_per_unit_rub") != 29.24:
                raise AssertionError("rematch must not erase approximate cost fields")

            import_status, import_payload = _post_multipart(
                f"{base_url}{DEFAULT_NOMENCLATURE_IMPORT_PATH}",
                nomenclature_import_bytes,
                filename="nomenclature.xlsx",
            )
            if (
                import_status != 200
                or import_payload.get("created_count") != 1
                or import_payload.get("updated_count") != 1
                or import_payload.get("deactivated_count") != 1
                or import_payload.get("error_count") != 0
            ):
                raise AssertionError(f"nomenclature import must apply batch changes, got {import_status} {import_payload}")
            imported_list_status, imported_list_payload = _get_json(f"{base_url}{DEFAULT_NOMENCLATURE_PATH}")
            imported_items = {item["item_id"]: item for item in imported_list_payload.get("items", [])}
            if imported_list_status != 200:
                raise AssertionError(f"nomenclature list after import failed: {imported_list_status} {imported_list_payload}")
            first_item = imported_items[str(create_nom_payload["item"]["item_id"])]
            if (
                first_item.get("purchase_price_yuan") != 13.75
                or first_item.get("our_sku") != "SKU-CLEAR-14P"
                or first_item.get("aliases") != ["iPhone 14 Pro"]
                or first_item.get("comment") != "smoke"
            ):
                raise AssertionError(f"nomenclature import must preserve hidden fields while updating visible fields, got {first_item}")
            compat_item = imported_items[str(compat_nom_payload["item"]["item_id"])]
            if compat_item.get("is_active") is not False:
                raise AssertionError("nomenclature import must support soft-disable through Включено=нет")
            created_import_items = [item for item in imported_items.values() if item.get("match_key") == "matte|iphone_15"]
            if len(created_import_items) != 1 or created_import_items[0].get("purchase_price_yuan") != 9.5:
                raise AssertionError(f"nomenclature import must create new active rows, got {created_import_items}")
            invalid_status, invalid_import_payload = _post_multipart(
                f"{base_url}{DEFAULT_NOMENCLATURE_IMPORT_PATH}",
                _build_invalid_nomenclature_import_fixture(),
                filename="nomenclature-invalid.xlsx",
            )
            if (
                invalid_status != 400
                or invalid_import_payload.get("status") != "error"
                or invalid_import_payload.get("error_count", 0) < 2
            ):
                raise AssertionError(f"invalid nomenclature import must return row-level errors, got {invalid_status} {invalid_import_payload}")
            after_invalid_status, after_invalid_payload = _get_json(f"{base_url}{DEFAULT_NOMENCLATURE_PATH}")
            if after_invalid_status != 200 or len(after_invalid_payload.get("items", [])) != len(imported_list_payload.get("items", [])):
                raise AssertionError("invalid nomenclature import must not partially mutate rows")
            runtime.save_nomenclature_item(
                {
                    "item_id": "nom_duplicate_a",
                    "is_active": True,
                    "our_sku": "",
                    "nm_id": 700001,
                    "nomenclature_name": "Duplicate A",
                    "product_type": "clean",
                    "match_key": "clean|duplicate",
                    "purchase_price_yuan": None,
                    "aliases": [],
                    "compatible_models_text": "",
                    "compatible_model_keys": [],
                    "comment": "",
                    "created_at": "2026-05-30T08:20:00Z",
                    "updated_at": "2026-05-30T08:20:00Z",
                }
            )
            runtime.save_nomenclature_item(
                {
                    "item_id": "nom_duplicate_b",
                    "is_active": True,
                    "our_sku": "",
                    "nm_id": 700002,
                    "nomenclature_name": "Duplicate B",
                    "product_type": "clean",
                    "match_key": "clean|duplicate",
                    "purchase_price_yuan": None,
                    "aliases": [],
                    "compatible_models_text": "",
                    "compatible_model_keys": [],
                    "comment": "",
                    "created_at": "2026-05-30T08:20:00Z",
                    "updated_at": "2026-05-30T08:20:00Z",
                }
            )
            ambiguous_status, ambiguous_payload = _post_multipart(
                f"{base_url}{DEFAULT_NOMENCLATURE_IMPORT_PATH}",
                _build_ambiguous_nomenclature_import_fixture(),
                filename="nomenclature-ambiguous.xlsx",
            )
            if ambiguous_status != 400 or "неоднозначен" not in json.dumps(ambiguous_payload, ensure_ascii=False):
                raise AssertionError(f"ambiguous match_key import must return row-level error, got {ambiguous_status} {ambiguous_payload}")

            registry_status, registry_payload = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}")
            if registry_status != 200 or len(registry_payload.get("shipments", [])) != 1:
                raise AssertionError("list route must expose saved shipment")
            if registry_payload["shipments"][0].get("supplier_name") != "HanShang Technology":
                raise AssertionError("list route must expose fixed supplier_name")
            if registry_payload["shipments"][0].get("order_status") != "accepted_ff":
                raise AssertionError("list route must expose persisted order_status")
            if (
                registry_payload["shipments"][0].get("planned_shipment_date") != "2026-05-15"
                or registry_payload["shipments"][0].get("actual_shipment_date") != "2026-05-17"
                or registry_payload["shipments"][0].get("actual_ff_acceptance_date") != "2026-05-30"
            ):
                raise AssertionError(f"list route must expose planned/fact dates, got {registry_payload}")
            shipment_registry_status, shipment_registry = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENT_REGISTRY_PATH}")
            if (
                shipment_registry_status != 200
                or shipment_registry.get("contract_name") != "sheet_vitrina_v1_supplier_shipment_registry"
                or shipment_registry.get("meta", {}).get("shipment_count") != 1
            ):
                raise AssertionError(f"shipment registry matrix route mismatch: {shipment_registry_status} {shipment_registry}")
            shipment_registry_json = json.dumps(shipment_registry, ensure_ascii=False)
            if "NaN" in shipment_registry_json or "Infinity" in shipment_registry_json:
                raise AssertionError(f"shipment registry matrix must not expose invalid numbers: {shipment_registry}")
            section_ids = [section.get("section_id") for section in shipment_registry.get("sections", [])]
            expected_section_ids = [
                "passport",
                "quote_logistics",
                "quote_normalized",
                "lead_times",
                "cargo_physics",
                "cargo_value",
                "fact_expenses",
                "fact_normalized",
                "documents",
            ]
            if section_ids != expected_section_ids:
                raise AssertionError(f"shipment registry matrix missing sections: {section_ids}")
            if _registry_cell_display(shipment_registry, "quote_normalized", "quote_total_rub_per_unit", shipment_id) != "нет КП":
                raise AssertionError(f"shipment registry without financial docs must keep quote ₽/шт unavailable: {shipment_registry}")
            if _registry_cell_display(shipment_registry, "fact_expenses", "fact_total_rub_per_unit", shipment_id) != "—":
                raise AssertionError(f"shipment registry without financial docs must keep fact ₽/шт unavailable: {shipment_registry}")
            if _registry_cell_display(shipment_registry, "lead_times", "shipment_date", shipment_id) != "2026-05-15":
                raise AssertionError(f"shipment registry must expose planned shipment date: {shipment_registry}")
            if _registry_cell_display(shipment_registry, "lead_times", "actual_shipment_date", shipment_id) != "2026-05-17":
                raise AssertionError(f"shipment registry must expose actual shipment date: {shipment_registry}")
            if _registry_cell_display(shipment_registry, "lead_times", "actual_ff_acceptance_date", shipment_id) != "2026-05-30":
                raise AssertionError(f"shipment registry must expose actual FF acceptance date: {shipment_registry}")
            if _registry_cell_display(shipment_registry, "lead_times", "actual_delivery_days", shipment_id) != "13 дн.":
                raise AssertionError(f"shipment registry must calculate actual delivery days: {shipment_registry}")
            invoice_path = registry_payload["shipments"][0].get("invoice_download_path")
            invoice_status, invoice_bytes, invoice_headers = _get_bytes(f"{base_url}{invoice_path}")
            if invoice_status != 200 or hashlib.sha256(invoice_bytes).hexdigest() != workbook_sha256:
                raise AssertionError("invoice download must preserve original XLSX bytes")
            if "attachment" not in str(invoice_headers.get("Content-Disposition", "")):
                raise AssertionError("invoice download must be an attachment")
            delete_status, delete_payload = _delete_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}")
            if (
                delete_status != 200
                or delete_payload.get("deleted") is not True
                or delete_payload.get("archived") is not True
                or not str(delete_payload.get("archive_event_id") or "")
                or not str(delete_payload.get("source_fingerprint") or "").startswith("sha256:")
            ):
                raise AssertionError(f"delete route must archive shipment, got {delete_status} {delete_payload}")
            after_delete_status, after_delete_payload = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}")
            if after_delete_status != 200 or after_delete_payload.get("shipments") != []:
                raise AssertionError("archived supplier order must disappear from active registry")
            archived_detail = runtime.load_supplier_shipment(shipment_id)
            archive_event = runtime.load_supplier_shipment_archive(shipment_id)
            if (
                archived_detail is None
                or not archived_detail.get("lines")
                or archived_detail.get("header", {}).get("persisted_order_status") != "archived"
                or not str(archived_detail.get("header", {}).get("archived_at") or "")
                or archive_event is None
                or archive_event.get("source_snapshot", {}).get("header", {}).get("shipment_id") != shipment_id
                or not archive_event.get("source_snapshot", {}).get("lines")
            ):
                raise AssertionError("supplier archive must retain header, lines and fingerprinted source snapshot")
            deleted_invoice_status, _, _ = _get_bytes(f"{base_url}{invoice_path}")
            if deleted_invoice_status != 404:
                raise AssertionError("archived supplier invoice must not remain downloadable")

            runtime.save_supplier_shipment(
                header={
                    "shipment_id": "sup_legacy_missing_supplier",
                    "created_at": "2026-05-30T08:10:00Z",
                    "updated_at": "2026-05-30T08:10:00Z",
                    "shipment_date": "2026-05-16",
                    "invoice_no": "LEGACY-1",
                    "invoice_date": "2026-05-15",
                    "contract_no": "",
                    "contract_date": "",
                    "supplier_name": "",
                    "customer_name": "",
                    "currency": "RMB",
                    "product_qty_total": 0,
                    "product_amount_total": 0,
                    "extras_amount_total": 0,
                    "invoice_amount_total": 0,
                    "declared_invoice_total": 0,
                    "match_status": "all_matched",
                    "source_filename": "legacy.xlsx",
                    "source_file_sha256": "",
                    "source_file_path": "",
                    "parser_version": "legacy",
                    "warnings": [],
                    "errors": [],
                },
                lines=[],
            )
            legacy_list_status, legacy_list_payload = _get_json(f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}")
            if (
                legacy_list_status != 200
                or legacy_list_payload.get("shipments", [{}])[0].get("supplier_name") != "HanShang Technology"
            ):
                raise AssertionError(f"legacy missing supplier must list with default fallback, got {legacy_list_payload}")
            if legacy_list_payload.get("shipments", [{}])[0].get("order_status") != "production":
                raise AssertionError("legacy missing order_status must list with production fallback")
            if (
                legacy_list_payload.get("shipments", [{}])[0].get("actual_shipment_date") != ""
                or legacy_list_payload.get("shipments", [{}])[0].get("actual_ff_acceptance_date") != ""
            ):
                raise AssertionError("legacy rows without fact dates must expose empty fact dates")
            legacy_no_rate = legacy_list_payload.get("shipments", [{}])[0]
            if (
                legacy_no_rate.get("approx_yuan_rate") is not None
                or legacy_no_rate.get("approx_invoice_cost_rub") is not None
                or legacy_no_rate.get("approx_landed_cost_per_unit_rub") is not None
            ):
                raise AssertionError(f"legacy rows without approx_yuan_rate must expose empty approximate fields, got {legacy_no_rate}")
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)
    print("sheet_vitrina_v1_supplier_shipments_http_smoke: OK")


def _assert_barcode_schema_upgrade_smoke() -> None:
    with TemporaryDirectory(prefix="supplier-barcode-schema-upgrade-") as tmp:
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=Path(tmp) / "runtime")
        runtime.save_supplier_shipment(
            header={
                "shipment_id": "sup_legacy_barcode_schema",
                "created_at": "2026-05-01T00:00:00Z",
                "updated_at": "2026-05-01T00:00:00Z",
                "shipment_date": "2026-05-02",
                "invoice_no": "LEGACY-SCHEMA",
                "currency": "RMB",
                "product_qty_total": 1,
                "product_amount_total": 1,
                "extras_amount_total": 0,
                "invoice_amount_total": 1,
                "declared_invoice_total": 1,
                "match_status": "all_matched",
                "parser_version": "legacy",
                "warnings": [],
                "errors": [],
            },
            lines=[
                {
                    "line_id": "ln_legacy_barcode_schema",
                    "line_type": "product",
                    "sort_order": 1,
                    "internal_nm_id": 1,
                    "qty": 1,
                    "unit_price": 1,
                    "amount": 1,
                    "currency": "RMB",
                    "match_status": "matched",
                    "manual_override": False,
                    "raw": {"legacy": True},
                }
            ],
        )
        with sqlite3.connect(runtime.db_path) as conn:
            conn.execute("ALTER TABLE sheet_vitrina_v1_supplier_shipment_lines DROP COLUMN barcode")
            conn.commit()
        loaded = runtime.load_supplier_shipment("sup_legacy_barcode_schema")
        if loaded is None or len(loaded.get("lines") or []) != 1:
            raise AssertionError("non-destructive barcode schema upgrade lost a legacy shipment line")
        legacy_line = loaded["lines"][0]
        if legacy_line.get("barcode") != "" or legacy_line.get("raw", {}).get("legacy") is not True:
            raise AssertionError(f"legacy shipment line must survive with an empty barcode: {legacy_line}")


def _build_invoice_fixture() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"
    sheet.append(["Invoice No:", "26GN390"])
    sheet.append(["Invoice Date:", "14.5.2026"])
    sheet.append(["Contract No.", "CNT-2026-0513"])
    sheet.append(["Date of Contract", "2026.5.13"])
    sheet.append(["Supplier:", "Zhejiang Supplier", "", "Currency:", "RMB"])
    sheet.append(["Invoice Total:", 33])
    sheet.append(["NO.", "MODELS", "NAME & SPECIFICATION", "Braocde\n(条形码）", "QTY", "U.PRICE", "AMOUNT", "COMMENT"])
    sheet.append([1, "iPhone 14 Pro", "高清膜 smk", "1111111111111", 10, 1, 10, ""])
    sheet.append([2, "iPhone 17e / 16e /14 / 13 / 13Pro", "防窥膜 (Anti-Spy)", "2222222222222", 4, 2, 8, ""])
    sheet.append([3, "iPhone 14 Pro Max", "防窥膜 (Anti-Spy)", "3333333333333", 5, 2, 10, ""])
    sheet.append([4, "OPP bag packets", "OPP bag packets", "", 100, 0.05, 5, "OPP packets"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _shipment_by_id(payload: dict, shipment_id: str) -> dict:
    for shipment in payload.get("shipments", []):
        if shipment.get("shipment_id") == shipment_id:
            return shipment
    return {}


def _seed_supplier_factual_expense(
    runtime: RegistryUploadDbBackedRuntime,
    shipment_id: str,
    *,
    amount_rub: float,
) -> None:
    runtime.save_supplier_financial_document(
        document={
            "document_id": f"fdoc_{shipment_id}_logistics",
            "supplier_order_id": shipment_id,
            "document_type": "logistics_invoice",
            "original_filename": "factual-logistics.pdf",
            "stored_file_path": "",
            "file_content_type": "application/pdf",
            "file_sha256": "",
            "uploaded_at": "2026-05-30T08:00:00Z",
            "updated_at": "2026-05-30T08:00:00Z",
            "parse_status": "parsed",
            "vendor": "Smoke Logistics",
            "document_number": "SMOKE-EXPENSE",
            "document_date": "2026-05-20",
            "currency": "RUB",
            "total_amount": amount_rub,
            "total_amount_rub": amount_rub,
            "normalized_parse": {"document_type": "logistics_invoice", "amount_rub": amount_rub},
            "raw_parse": {},
            "parser_version": "smoke",
            "warnings": [],
            "errors": [],
        },
        expense_lines=[
            {
                "line_id": f"fline_{shipment_id}_logistics",
                "financial_document_id": f"fdoc_{shipment_id}_logistics",
                "supplier_order_id": shipment_id,
                "sort_order": 1,
                "category": "domestic_transport",
                "stage": "fact",
                "description": "Factual smoke logistics expense",
                "amount": amount_rub,
                "currency": "RUB",
                "amount_rub": amount_rub,
                "vat_rate": None,
                "vat_amount_rub": None,
                "included_in_logistics_efficiency": True,
                "included_in_customs_total": False,
                "status": "parsed",
                "confidence": 1.0,
                "raw": {},
            }
        ],
    )


def _build_price_conformity_invoice_fixture() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"
    sheet.append(["Invoice No:", "PRICE-CHECK"])
    sheet.append(["Invoice Date:", "30.5.2026"])
    sheet.append(["Contract No.", "CNT-PRICE"])
    sheet.append(["Date of Contract", "2026.5.30"])
    sheet.append(["Supplier:", "Zhejiang Supplier", "", "Currency:", "RMB"])
    sheet.append(["Invoice Total:", 16])
    sheet.append(["NO.", "MODELS", "NAME & SPECIFICATION", "Barcode", "QTY", "U.PRICE", "AMOUNT", "COMMENT"])
    sheet.append([1, "iPhone 14 Pro", "高清膜 smk", "1111111111111", 1, 1, 1, "matched"])
    sheet.append([2, "iPhone 14", "防窥膜 (Anti-Spy)", "2222222222222", 1, 2, 2, "mismatched"])
    sheet.append([3, "iPhone 99", "高清膜 smk", "3333333333333", 1, 4, 4, "sku missing"])
    sheet.append([4, "iPhone 15", "高清膜 smk", "4444444444444", 1, 4, 4, "reference price missing"])
    sheet.append([5, "iPhone 16", "磨砂膜 Matte", "5555555555555", 1, "", "", "invoice price missing"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_nomenclature_import_fixture(*, first_item_id: str, compat_item_id: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Номенклатура"
    sheet.append(
        [
            "ID строки",
            "Включено",
            "nmId",
            "Номенклатура",
            "Тип",
            "Match key",
            "Цена закупки, ¥",
            "Совместимые модели",
            "Ключи совместимости",
            "Обновлено",
        ]
    )
    sheet.append(
        [
            first_item_id,
            "да",
            210183919,
            "Clear iPhone 14 Pro",
            "Прозрачное",
            "clear|iphone_14_pro",
            "13,75",
            "iPhone 14 Pro",
            "",
            "",
        ]
    )
    sheet.append(
        [
            compat_item_id,
            "нет",
            391662410,
            "anti-spy iPhone 14 / 13 / 13Pro",
            "anti_spy",
            "anti_spy|iphone_14_13_13pro",
            "",
            "iPhone 14, iPhone 13, iPhone 13 Pro",
            "iphone_14; iphone_13; iphone_13_pro",
            "",
        ]
    )
    sheet.append(
        [
            "",
            "yes",
            500001,
            "Matte iPhone 15",
            "Матовое",
            "matte|iphone_15",
            "9,5",
            "iPhone 15",
            "",
            "",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_invalid_nomenclature_import_fixture() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Номенклатура"
    sheet.append(["Включено", "nmId", "Номенклатура", "Тип", "Match key", "Цена закупки, ¥"])
    sheet.append(["да", 900001, "Invalid type", "глянцевое", "clear|invalid_type", 1])
    sheet.append(["да", 900002, "Invalid price", "clear", "clear|invalid_price", "-1"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_ambiguous_nomenclature_import_fixture() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Номенклатура"
    sheet.append(["Включено", "nmId", "Номенклатура", "Тип", "Match key", "Цена закупки, ¥"])
    sheet.append(["да", 700003, "Duplicate import", "clean", "clean|duplicate", 2])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _post_multipart(
    url: str,
    workbook_bytes: bytes,
    *,
    filename: str,
    content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
) -> tuple[int, dict[str, object]]:
    boundary = "----wbcore-supplier" + uuid4().hex
    body = b"".join(
        [
            f"--{boundary}\r\n".encode("utf-8"),
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'.encode("utf-8"),
            f"Content-Type: {content_type}\r\n\r\n".encode("utf-8"),
            workbook_bytes,
            f"\r\n--{boundary}--\r\n".encode("utf-8"),
        ]
    )
    request = urllib_request.Request(
        url,
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}", "Accept": "application/json"},
        method="POST",
    )
    return _open_json(request)


def _post_json(url: str, payload: dict[str, object]) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8", "Accept": "application/json"},
        method="POST",
    )
    return _open_json(request)


def _patch_json(url: str, payload: dict[str, object]) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8", "Accept": "application/json"},
        method="PATCH",
    )
    return _open_json(request)


def _delete_json(url: str) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(url, headers={"Accept": "application/json"}, method="DELETE")
    return _open_json(request)


def _registry_cell_display(registry, section_id: str, row_id: str, shipment_id: str) -> str:
    for section in registry.get("sections", []):
        if section.get("section_id") != section_id:
            continue
        for row in section.get("rows", []):
            if row.get("row_id") == row_id:
                return str((row.get("cells", {}).get(shipment_id) or {}).get("display") or "")
    return ""


def _get_json(url: str) -> tuple[int, dict[str, object]]:
    request = urllib_request.Request(url, headers={"Accept": "application/json"}, method="GET")
    return _open_json(request)


def _wait_for_factual_correction(base_url: str, shipment_id: str) -> dict[str, object]:
    url = f"{base_url}{DEFAULT_SUPPLIER_SHIPMENTS_PATH}/{shipment_id}/factual-date-correction"
    for _ in range(200):
        status, payload = _get_json(url)
        if status == 200 and payload.get("status") in {"success", "error"}:
            return payload
        time.sleep(0.02)
    raise AssertionError("factual date correction status did not become terminal")


def _open_json(request: urllib_request.Request) -> tuple[int, dict[str, object]]:
    try:
        with urllib_request.urlopen(request, timeout=5) as response:
            return response.status, json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        return exc.code, json.loads(exc.read().decode("utf-8"))


def _get_bytes(url: str) -> tuple[int, bytes, dict[str, str]]:
    request = urllib_request.Request(url, method="GET")
    try:
        with urllib_request.urlopen(request, timeout=5) as response:
            return response.status, response.read(), dict(response.headers)
    except urllib_error.HTTPError as exc:
        return exc.code, exc.read(), dict(exc.headers)


def _reserve_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


if __name__ == "__main__":
    main()
