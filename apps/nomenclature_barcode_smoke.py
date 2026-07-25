"""Targeted smoke-check for server-owned nomenclature WB barcodes."""

from __future__ import annotations

from contextlib import contextmanager
from io import BytesIO
import json
import os
from pathlib import Path
import socket
import sqlite3
import sys
from tempfile import TemporaryDirectory
import threading
from urllib import error as urllib_error, request as urllib_request

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from packages.adapters.official_api_runtime import OfficialApiRuntimeError  # noqa: E402
from packages.adapters.wb_content import HttpBackedWbContentSource, WbContentHttpStatusError  # noqa: E402
from packages.adapters.registry_upload_http_entrypoint import (  # noqa: E402
    DEFAULT_NOMENCLATURE_BARCODE_SYNC_PATH,
    DEFAULT_NOMENCLATURE_PATH,
    DEFAULT_SKU_GROUPS_PATH,
    DEFAULT_SHEET_OPERATOR_UI_PATH,
    DEFAULT_SHEET_PLAN_PATH,
    DEFAULT_SHEET_STATUS_PATH,
    DEFAULT_UPLOAD_PATH,
    build_registry_upload_http_server,
)
from packages.application.registry_upload_db_backed_runtime import RegistryUploadDbBackedRuntime  # noqa: E402
from packages.application.registry_upload_http_entrypoint import RegistryUploadHttpEntrypoint  # noqa: E402
from packages.application.supplier_shipments import SupplierShipmentsBlock  # noqa: E402
from packages.contracts.registry_upload_http_entrypoint import RegistryUploadHttpEntrypointConfig  # noqa: E402


class FakeBarcodeSource:
    def __init__(
        self,
        mapping: dict[int, list[str]] | None = None,
        exc: Exception | None = None,
        cards: list[dict[str, object]] | None = None,
    ) -> None:
        self.mapping = mapping or {}
        self.exc = exc
        self.cards = cards or []
        self.calls: list[list[int]] = []
        self.card_calls: list[dict[str, int | None]] = []

    def fetch_barcodes_by_nm_ids(self, nm_ids: list[int]):
        self.calls.append(list(nm_ids))
        if self.exc is not None:
            raise self.exc
        return {
            int(nm_id): {
                "nm_id": int(nm_id),
                "barcodes": list(self.mapping.get(int(nm_id), [])),
                "cards_found": 1 if self.mapping.get(int(nm_id)) else 0,
                "pages_fetched": 1,
                "endpoint": "/content/v2/get/cards/list",
            }
            for nm_id in nm_ids
        }

    def fetch_cards(self, *, limit: int | None = None, max_pages: int | None = None):
        self.card_calls.append({"limit": limit, "max_pages": max_pages})
        if self.exc is not None:
            raise self.exc
        return list(self.cards)


class FakeResponse:
    def __init__(self, payload: bytes, *, status: int = 200, headers: dict[str, str] | None = None) -> None:
        self.payload = payload
        self.status = status
        self.headers = headers or {"Content-Type": "application/json"}

    def __enter__(self) -> "FakeResponse":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        return None

    def read(self) -> bytes:
        return self.payload


class RecordingOpener:
    def __init__(self) -> None:
        self.requests = []
        self.next_payload = b"{}"
        self.next_error = None

    def __call__(self, req, timeout):
        self.requests.append((req, timeout))
        if self.next_error is not None:
            raise self.next_error
        return FakeResponse(self.next_payload)


def main() -> None:
    _adapter_smoke()
    _application_runtime_xlsx_smoke()
    _http_route_smoke()
    print("nomenclature_barcode_smoke: OK")


def _adapter_smoke() -> None:
    with _patched_env({"WB_API_TOKEN": "content-smoke-token", "WB_CONTENT_API_BASE_URL": "https://content-api.example.test"}):
        opener = RecordingOpener()
        opener.next_payload = json.dumps(
            {
                "cards": [
                    {
                        "nmID": 501001,
                        "sizes": [
                            {"skus": ["2000000000001", "2000000000002"]},
                            {"skus": ["2000000000001"]},
                        ],
                    }
                ],
                "cursor": {"total": 1},
            }
        ).encode("utf-8")
        source = HttpBackedWbContentSource(opener=opener)
        result = source.fetch_barcodes_by_nm_ids([501001])
        resolution = result[501001]
        if resolution.barcodes != ["2000000000001", "2000000000002"]:
            raise AssertionError(f"content adapter must extract size skus, got {resolution}")
        req, timeout = opener.requests[-1]
        if req.get_method() != "POST" or req.full_url != "https://content-api.example.test/content/v2/get/cards/list":
            raise AssertionError(f"content adapter endpoint changed: {req.get_method()} {req.full_url}")
        if req.get_header("Authorization") != "content-smoke-token":
            raise AssertionError("content adapter must use canonical WB_API_TOKEN")
        body = json.loads(req.data.decode("utf-8"))
        if body["settings"]["filter"]["textSearch"] != "501001" or body["settings"]["cursor"]["limit"] != 100:
            raise AssertionError(f"content adapter body changed unexpectedly: {body}")
        if timeout <= 0:
            raise AssertionError("content adapter must pass timeout")

        opener.next_payload = json.dumps(
            {
                "cards": [
                    {
                        "nmID": 701001,
                        "vendorCode": "No Frame Anti-Spy iPhone 16 Pro Max",
                        "title": "WB No Frame Anti-Spy",
                        "subjectName": "Защитные стекла",
                        "updatedAt": "2026-07-01T10:00:00Z",
                        "sizes": [{"skus": ["7000000000001", "7000000000002"]}],
                    }
                ],
                "cursor": {"total": 1},
            }
        ).encode("utf-8")
        cards = source.fetch_cards(limit=100, max_pages=1)
        if len(cards) != 1:
            raise AssertionError(f"content adapter must parse one card, got {cards}")
        card = cards[0]
        if (
            card.nm_id != 701001
            or card.vendor_code != "No Frame Anti-Spy iPhone 16 Pro Max"
            or card.title != "WB No Frame Anti-Spy"
            or card.subject_name != "Защитные стекла"
            or card.updated_at != "2026-07-01T10:00:00Z"
            or card.barcodes != ["7000000000001", "7000000000002"]
        ):
            raise AssertionError(f"content adapter card parsing mismatch: {card}")

        opener.next_error = urllib_error.HTTPError(
            "https://content-api.example.test/content/v2/get/cards/list",
            401,
            "Unauthorized",
            hdrs={"Content-Type": "application/json"},
            fp=BytesIO(b'{"error":"bad token"}'),
        )
        try:
            source.fetch_barcodes_by_nm_ids([501001])
        except WbContentHttpStatusError as exc:
            if exc.status_code != 401:
                raise AssertionError(f"content adapter must preserve status, got {exc.status_code}")
            if "content-smoke-token" in str(exc):
                raise AssertionError("content adapter error must not print token")
        else:
            raise AssertionError("content adapter HTTP errors must raise")


def _application_runtime_xlsx_smoke() -> None:
    with TemporaryDirectory(prefix="nomenclature-barcode-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        _seed_old_schema_row(runtime_dir)
        runtime = RegistryUploadDbBackedRuntime(runtime_dir=runtime_dir)
        old_rows = runtime.list_nomenclature_items()
        old_row = next(item for item in old_rows if item["item_id"] == "old_nom")
        if old_row.get("barcode") != "" or old_row.get("barcode_status") != "missing":
            raise AssertionError(f"old schema row must survive barcode migration as missing, got {old_row}")

        source = FakeBarcodeSource({501001: ["3000000000002", "3000000000001"]})
        block = SupplierShipmentsBlock(runtime=runtime, barcode_source=source, timestamp_factory=_clock())
        manual = block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501000,
                "barcode": " 1000000000001 ",
                "nomenclature_name": "Manual Clear",
                "product_type": "clean",
                "match_key": "clean|iphone_14",
                "purchase_price_yuan": "1.2",
                "factory_box_size": "250",
            }
        )["item"]
        if (
            manual["barcode"] != "1000000000001"
            or manual["barcode_source"] != "manual"
            or manual["barcode_status"] != "manual"
            or manual["factory_box_size"] != 250
        ):
            raise AssertionError(f"manual barcode must be authoritative, got {manual}")
        try:
            block.update_nomenclature_item(
                manual["item_id"], {"factory_box_size": "250.5"}
            )
        except ValueError as exc:
            if "positive integer" not in str(exc):
                raise
        else:
            raise AssertionError("fractional factory box size must fail closed")
        block.sync_nomenclature_item_barcode(manual["item_id"])
        preserved = runtime.load_nomenclature_item(manual["item_id"])
        if preserved["barcode"] != "1000000000001" or source.calls:
            raise AssertionError("row sync must not overwrite/call WB for manual barcode")

        synced = block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501001,
                "nomenclature_name": "WB Clear",
                "product_type": "clean",
                "match_key": "clean|iphone_15",
                "purchase_price_yuan": "1.5",
            }
        )["item"]
        if (
            synced["barcode"] != "3000000000001"
            or synced["barcodes"] != ["3000000000001", "3000000000002"]
            or synced["barcode_source"] != "wb_content"
            or synced["barcode_status"] != "multiple"
        ):
            raise AssertionError(f"WB sync must pick deterministic primary and keep all barcodes, got {synced}")

        patched = block.update_nomenclature_item(synced["item_id"], {"barcode": "4000000000001"})["item"]
        if patched["barcode_source"] != "manual" or patched["barcode_status"] != "manual":
            raise AssertionError(f"patching barcode must become manual override, got {patched}")

        missing_source = FakeBarcodeSource(exc=OfficialApiRuntimeError("required env WB_API_TOKEN is not set"))
        token_block = SupplierShipmentsBlock(runtime=runtime, barcode_source=missing_source, timestamp_factory=_clock(start=30))
        token_item = token_block.create_nomenclature_item(
            {
                "is_active": True,
                "nm_id": 501002,
                "nomenclature_name": "Token Missing",
                "product_type": "clean",
                "match_key": "clean|iphone_16",
                "purchase_price_yuan": "1.0",
            }
        )["item"]
        if token_item["barcode_status"] != "token_missing" or token_item["barcode"] != "":
            raise AssertionError(f"token missing must be diagnostic, not save blocker, got {token_item}")

        listed = block.list_nomenclature()
        summary = listed.get("summary") or {}
        if summary.get("active_rows_with_barcode", 0) < 2 or summary.get("manual_barcode_count", 0) < 2:
            raise AssertionError(f"list response must expose barcode readiness summary, got {summary}")

        workbook_bytes, _, _ = block.export_nomenclature_xlsx()
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        headers = [cell.value for cell in next(workbook.active.iter_rows(min_row=1, max_row=1))]
        if "ШК / barcode" not in headers or "Все ШК" not in headers:
            raise AssertionError(f"export must include barcode columns, got {headers}")

        imported = block.import_nomenclature_xlsx(_build_import_workbook(), uploaded_filename="nomenclature.xlsx")
        if imported["created_count"] != 1:
            raise AssertionError(f"barcode import must create one row, got {imported}")
        imported_row = next(item for item in runtime.list_nomenclature_items() if item["match_key"] == "matte|iphone_17")
        if imported_row["barcode"] != "5000000000001" or imported_row["barcode_source"] != "manual":
            raise AssertionError(f"imported barcode must be manual, got {imported_row}")

        old_import = block.import_nomenclature_xlsx(_build_old_import_workbook(), uploaded_filename="old.xlsx")
        if old_import["created_count"] != 1:
            raise AssertionError(f"old import workbook without barcode must remain valid, got {old_import}")
        legacy_row = next(item for item in runtime.list_nomenclature_items() if item["match_key"] == "clear|iphone_18")
        if legacy_row["product_type"] != "clean":
            raise AssertionError(f"legacy Russian label must resolve through the current Clean group, got {legacy_row}")
        legacy_export_bytes, _, _ = block.export_nomenclature_xlsx()
        legacy_workbook = load_workbook(BytesIO(legacy_export_bytes), data_only=True)
        legacy_headers = [cell.value for cell in next(legacy_workbook.active.iter_rows(min_row=1, max_row=1))]
        group_index = legacy_headers.index("Группа")
        match_index = legacy_headers.index("Match key")
        legacy_export_row = next(
            row
            for row in legacy_workbook.active.iter_rows(min_row=2, values_only=True)
            if row[match_index] == "clear|iphone_18"
        )
        if legacy_export_row[group_index] != "Clean":
            raise AssertionError(f"runtime Clean label must drive the export, got {legacy_export_row[group_index]}")

        _wb_card_sync_smoke(runtime)


def _wb_card_sync_smoke(runtime: RegistryUploadDbBackedRuntime) -> None:
    now = "2026-06-27T10:00:00Z"
    runtime.save_nomenclature_items_atomic(
        [
            _nomenclature_item(
                "sync_nm",
                nm_id=810001,
                name="Manual name must survive",
                product_type="clean",
                match_key="clean|manual_nm",
                barcode="",
                barcode_source="missing",
                barcode_status="missing",
                vendor_code="",
                is_hidden=False,
                price=9.0,
                now=now,
            ),
            _nomenclature_item(
                "sync_barcode",
                nm_id=None,
                name="Manual barcode must survive",
                product_type="matte",
                match_key="matte|manual_barcode",
                barcode="8100000000002",
                barcode_source="manual",
                barcode_status="manual",
                vendor_code="",
                is_hidden=False,
                price=8.0,
                now=now,
            ),
            _nomenclature_item(
                "sync_vendor",
                nm_id=None,
                name="Vendor existing name",
                product_type="anti_spy",
                match_key="anti_spy|vendor_existing",
                barcode="",
                barcode_source="missing",
                barcode_status="missing",
                vendor_code="Vendor Match Existing",
                is_hidden=False,
                price=7.0,
                now=now,
            ),
            _nomenclature_item(
                "sync_hidden",
                nm_id=None,
                name="Hidden No Frame Clean",
                product_type="no_frame_clean",
                match_key="no_frame_clean|hidden",
                barcode="",
                barcode_source="missing",
                barcode_status="missing",
                vendor_code="No Frame Clean iPhone 16 Pro Max",
                is_hidden=True,
                price=6.0,
                now=now,
            ),
        ]
    )
    cards = [
        {
            "nm_id": 810001,
            "vendor_code": "WB NM Existing",
            "title": "WB nm title",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:00Z",
            "barcodes": ["8100000000001"],
        },
        {
            "nm_id": 810002,
            "vendor_code": "WB Barcode Existing",
            "title": "WB barcode title",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:01Z",
            "barcodes": ["8100000000002", "8100000000003"],
        },
        {
            "nm_id": 810003,
            "vendor_code": "Vendor Match Existing",
            "title": "WB vendor title",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:02Z",
            "barcodes": ["8100000000004"],
        },
        {
            "nm_id": 810004,
            "vendor_code": "No Frame Clean iPhone 16 Pro Max",
            "title": "Hidden card title",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:03Z",
            "barcodes": ["8100000000005"],
        },
        {
            "nm_id": 810005,
            "vendor_code": "No Frame Clean iPhone 17 Pro",
            "title": "New no frame clean",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:04Z",
            "barcodes": ["8100000000006"],
        },
        {
            "nm_id": 810006,
            "vendor_code": "No Frame Anti-Spy iPhone 16 Pro Max",
            "title": "New no frame anti-spy",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:05Z",
            "barcodes": ["8100000000007"],
        },
        {
            "nm_id": 810007,
            "vendor_code": "No Frame Matte iPhone 18 Pro",
            "title": "New no frame matte",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:06Z",
            "barcodes": ["8100000000008"],
        },
        {
            "nm_id": 810008,
            "vendor_code": "Mystery iPhone 19 Pro",
            "title": "Unknown group",
            "subject_name": "Glass",
            "updated_at": "2026-07-01T00:00:07Z",
            "barcodes": ["8100000000009"],
        },
    ]
    sync_block = SupplierShipmentsBlock(
        runtime=runtime,
        barcode_source=FakeBarcodeSource(cards=cards),
        timestamp_factory=_clock(start=30),
    )
    before_count = len(runtime.list_nomenclature_items())
    result = sync_block.sync_nomenclature_barcodes({"limit": 100, "max_pages": 1})
    if result["cards_processed"] != len(cards) or result["created"] != 4:
        raise AssertionError(f"WB card sync counts mismatch: {result}")
    if result["matched_nm_id"] < 1 or result["matched_barcode"] < 1 or result["matched_vendor_code"] < 2:
        raise AssertionError(f"WB card sync must match by nmID, barcode and vendorCode: {result}")
    nm_row = runtime.load_nomenclature_item("sync_nm")
    if (
        nm_row["nomenclature_name"] != "Manual name must survive"
        or nm_row["product_type"] != "clean"
        or nm_row["purchase_price_yuan"] != 9.0
        or nm_row["match_key"] != "clean|manual_nm"
        or nm_row["barcode"] != "8100000000001"
        or nm_row["wb_title"] != "WB nm title"
    ):
        raise AssertionError(f"existing nmID match must update only WB-owned/reference fields, got {nm_row}")
    barcode_row = runtime.load_nomenclature_item("sync_barcode")
    if (
        barcode_row["nm_id"] != 810002
        or barcode_row["barcode"] != "8100000000002"
        or barcode_row["barcode_source"] != "manual"
        or barcode_row["wb_sync_status"] != "matched_barcode"
    ):
        raise AssertionError(f"barcode match must fill nm_id but preserve manual barcode override, got {barcode_row}")
    vendor_row = runtime.load_nomenclature_item("sync_vendor")
    if vendor_row["nm_id"] != 810003 or vendor_row["barcode"] != "8100000000004":
        raise AssertionError(f"vendorCode match must update existing row, got {vendor_row}")
    hidden_row = runtime.load_nomenclature_item("sync_hidden")
    if not hidden_row["is_hidden"] or hidden_row["nm_id"] != 810004:
        raise AssertionError(f"hidden row must match and remain hidden, got {hidden_row}")
    if any(item["item_id"] == "sync_hidden" for item in sync_block.list_nomenclature()["items"]):
        raise AssertionError("hidden row must not be visible by default")
    hidden_list = sync_block.list_nomenclature(visibility="hidden")["items"]
    if not any(item["item_id"] == "sync_hidden" for item in hidden_list):
        raise AssertionError("hidden selector must show hidden row")
    restored = sync_block.update_nomenclature_item("sync_hidden", {"is_hidden": False})["item"]
    if restored["is_hidden"]:
        raise AssertionError(f"restore hidden SKU failed: {restored}")
    if not any(item["item_id"] == "sync_hidden" for item in sync_block.list_nomenclature()["items"]):
        raise AssertionError("restored SKU must return to default visible list")
    sync_block.update_nomenclature_item("sync_hidden", {"is_hidden": True})
    repeat = sync_block.sync_nomenclature_barcodes({"limit": 100, "max_pages": 1})
    after_count = len(runtime.list_nomenclature_items())
    if after_count != before_count + 4 or repeat["created"] != 0:
        raise AssertionError(f"hidden repeat sync must not create duplicate: before={before_count} after={after_count} repeat={repeat}")
    rows = runtime.list_nomenclature_items()
    by_vendor = {row["vendor_code"]: row for row in rows if row.get("vendor_code")}
    if by_vendor["No Frame Clean iPhone 17 Pro"]["product_type"] != "no_frame_clean":
        raise AssertionError("No Frame Clean vendorCode must auto-detect no_frame_clean")
    if by_vendor["No Frame Anti-Spy iPhone 16 Pro Max"]["product_type"] != "no_frame_anti_spy":
        raise AssertionError("No Frame Anti-Spy vendorCode must auto-detect no_frame_anti_spy")
    if by_vendor["No Frame Matte iPhone 18 Pro"]["product_type"] != "no_frame_matte":
        raise AssertionError("No Frame Matte vendorCode must auto-detect no_frame_matte")
    if by_vendor["Mystery iPhone 19 Pro"]["product_type"] != "other" or by_vendor["Mystery iPhone 19 Pro"]["wb_sync_status"] != "needs_review":
        raise AssertionError("unknown vendorCode must become other/needs_review")


def _nomenclature_item(
    item_id: str,
    *,
    nm_id: int | None,
    name: str,
    product_type: str,
    match_key: str,
    barcode: str,
    barcode_source: str,
    barcode_status: str,
    vendor_code: str,
    is_hidden: bool,
    price: float,
    now: str,
) -> dict[str, object]:
    return {
        "item_id": item_id,
        "is_active": True,
        "is_hidden": is_hidden,
        "hidden_at": now if is_hidden else "",
        "hidden_reason": "smoke" if is_hidden else "",
        "our_sku": "",
        "nm_id": nm_id,
        "barcode": barcode,
        "barcodes": [barcode] if barcode else [],
        "barcode_source": barcode_source,
        "barcode_status": barcode_status,
        "barcode_synced_at": "",
        "barcode_updated_at": now if barcode else "",
        "barcode_evidence": {},
        "vendor_code": vendor_code,
        "wb_title": "",
        "wb_subject_name": "",
        "wb_updated_at": "",
        "wb_synced_at": "",
        "wb_sync_status": "",
        "wb_sync_evidence": {},
        "nomenclature_name": name,
        "product_type": product_type,
        "match_key": match_key,
        "purchase_price_yuan": price,
        "aliases": [],
        "compatible_models_text": "",
        "compatible_model_keys": [],
        "comment": "",
        "created_at": now,
        "updated_at": now,
    }


def _http_route_smoke() -> None:
    with TemporaryDirectory(prefix="nomenclature-barcode-http-") as tmp:
        runtime_dir = Path(tmp) / "runtime"
        config = RegistryUploadHttpEntrypointConfig(
            host="127.0.0.1",
            port=_reserve_free_port(),
            upload_path=DEFAULT_UPLOAD_PATH,
            sheet_plan_path=DEFAULT_SHEET_PLAN_PATH,
            sheet_refresh_path="/v1/sheet-vitrina-v1/refresh",
            sheet_status_path=DEFAULT_SHEET_STATUS_PATH,
            sheet_operator_ui_path=DEFAULT_SHEET_OPERATOR_UI_PATH,
            runtime_dir=runtime_dir,
        )
        entrypoint = RegistryUploadHttpEntrypoint(
            runtime_dir=runtime_dir,
            activated_at_factory=lambda: "2026-06-27T10:00:00Z",
        )
        server = build_registry_upload_http_server(config, entrypoint=entrypoint)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        original_token = os.environ.get("WB_API_TOKEN")
        os.environ.pop("WB_API_TOKEN", None)
        thread.start()
        try:
            base_url = f"http://127.0.0.1:{config.port}"
            manual_status, manual_payload = _post_json(
                f"{base_url}{DEFAULT_NOMENCLATURE_PATH}",
                {
                    "is_active": True,
                    "nm_id": 601001,
                    "barcode": "6000000000001",
                    "nomenclature_name": "HTTP Manual",
                    "product_type": "clean",
                    "match_key": "clean|http_manual",
                    "purchase_price_yuan": "1",
                },
            )
            if manual_status != 200 or manual_payload.get("item", {}).get("barcode_source") != "manual":
                raise AssertionError(f"HTTP manual barcode create failed: {manual_status} {manual_payload}")
            auto_status, auto_payload = _post_json(
                f"{base_url}{DEFAULT_NOMENCLATURE_PATH}",
                {
                    "is_active": True,
                    "nm_id": 601002,
                    "nomenclature_name": "HTTP Token Missing",
                    "product_type": "clean",
                    "match_key": "clean|http_token_missing",
                    "purchase_price_yuan": "1",
                },
            )
            if auto_status != 200 or auto_payload.get("item", {}).get("barcode_status") != "token_missing":
                raise AssertionError(f"HTTP auto-sync token warning must not block save: {auto_status} {auto_payload}")
            sync_status, sync_payload = _post_json(
                f"{base_url}{DEFAULT_NOMENCLATURE_BARCODE_SYNC_PATH}",
                {"limit": 10, "max_pages": 1},
            )
            if sync_status != 200 or sync_payload.get("status") != "token_missing":
                raise AssertionError(f"HTTP batch WB sync must expose token_missing status: {sync_status} {sync_payload}")
            list_status, list_payload = _get_json(f"{base_url}{DEFAULT_NOMENCLATURE_PATH}")
            summary = list_payload.get("summary") or {}
            if list_status != 200 or summary.get("active_rows_missing_barcode", 0) < 1:
                raise AssertionError(f"HTTP list must expose barcode summary: {list_status} {list_payload}")
            groups_status, groups_payload = _get_json(f"{base_url}{DEFAULT_SKU_GROUPS_PATH}")
            group_keys = {str(group.get("group_key") or "") for group in groups_payload.get("groups") or []}
            if groups_status != 200 or "no_frame_clean" not in group_keys:
                raise AssertionError(f"HTTP groups route must expose default SKU groups: {groups_status} {groups_payload}")
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)
            if original_token is None:
                os.environ.pop("WB_API_TOKEN", None)
            else:
                os.environ["WB_API_TOKEN"] = original_token


def _seed_old_schema_row(runtime_dir: Path) -> None:
    runtime_dir.mkdir(parents=True, exist_ok=True)
    db_path = runtime_dir / "registry_upload_runtime.sqlite3"
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE sheet_vitrina_v1_nomenclature_items (
                item_id TEXT PRIMARY KEY,
                is_active INTEGER NOT NULL,
                our_sku TEXT,
                nm_id INTEGER,
                nomenclature_name TEXT NOT NULL,
                product_type TEXT NOT NULL,
                match_key TEXT NOT NULL,
                purchase_price_yuan REAL,
                aliases_json TEXT NOT NULL,
                compatible_models_text TEXT NOT NULL DEFAULT '',
                compatible_model_keys_json TEXT NOT NULL DEFAULT '[]',
                comment TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO sheet_vitrina_v1_nomenclature_items(
                item_id, is_active, our_sku, nm_id, nomenclature_name, product_type, match_key,
                purchase_price_yuan, aliases_json, compatible_models_text, compatible_model_keys_json,
                comment, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "old_nom",
                1,
                "",
                500999,
                "Old Clear",
                "clear",
                "clear|iphone_13",
                1.0,
                "[]",
                "",
                "[]",
                "",
                "2026-06-01T00:00:00Z",
                "2026-06-01T00:00:00Z",
            ),
        )
        conn.commit()


def _build_import_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Включено", "nmId", "ШК / barcode", "Номенклатура", "Тип", "Match key", "Цена закупки, ¥"])
    sheet.append(["да", 501017, "5000000000001", "Matte iPhone 17", "Матовое", "matte|iphone_17", 2])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_old_import_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Включено", "nmId", "Номенклатура", "Тип", "Match key", "Цена закупки, ¥"])
    sheet.append(["да", 501018, "Clear iPhone 18", "Прозрачное", "clear|iphone_18", 2])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _clock(start: int = 0):
    state = {"value": start}

    def tick() -> str:
        state["value"] += 1
        return f"2026-06-27T09:{state['value']:02d}:00Z"

    return tick


def _get_json(url: str) -> tuple[int, dict]:
    req = urllib_request.Request(url, method="GET", headers={"Accept": "application/json"})
    try:
        with urllib_request.urlopen(req, timeout=10) as response:
            return int(response.status), json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        return int(exc.code), json.loads(exc.read().decode("utf-8"))


def _post_json(url: str, payload: dict) -> tuple[int, dict]:
    req = urllib_request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Content-Type": "application/json", "Accept": "application/json"},
    )
    try:
        with urllib_request.urlopen(req, timeout=10) as response:
            return int(response.status), json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        return int(exc.code), json.loads(exc.read().decode("utf-8"))


def _reserve_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


@contextmanager
def _patched_env(values: dict[str, str]):
    previous = {key: os.environ.get(key) for key in values}
    os.environ.update(values)
    try:
        yield
    finally:
        for key, value in previous.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


if __name__ == "__main__":
    main()
